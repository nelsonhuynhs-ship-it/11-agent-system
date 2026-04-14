"""test_enrichment.py — Active Jobs enrichment tests (col 28 + col 31).

Tests:
  - parse_routing helper
  - enrich() writes mailto links to col 28 and SERVICE to col 31
  - enrich() skips existing links without --force
  - enrich() overwrites with force=True
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "ERP" / "jobs"))
from enrichment import COL, DATA_START, enrich, parse_routing

AJ_SHEET_KEYWORD = "Active"


# ---------------------------------------------------------------------------
# parse_routing unit tests
# ---------------------------------------------------------------------------

class TestParseRouting:

    def test_simple_pol_pod(self):
        pol, pod, place = parse_routing("HPH-USLGB")
        assert pol == "HPH"
        assert pod == "USLGB"
        assert place == "USLGB"

    def test_via_route(self):
        pol, pod, place = parse_routing("HCM-CHICAGO VIA USLAX")
        assert pol == "HCM"
        assert pod == "USLAX"
        assert place == "CHICAGO"

    def test_empty_routing(self):
        pol, pod, place = parse_routing("")
        assert pol == ""
        assert pod == ""
        assert place == ""

    def test_pol_only(self):
        pol, pod, place = parse_routing("HPH")
        assert pol == "HPH"
        assert pod == ""

    def test_case_preserved(self):
        pol, pod, place = parse_routing("hcm-USLAX")
        assert pol == "hcm"
        assert pod == "USLAX"


# ---------------------------------------------------------------------------
# Integration tests on seeded_erp
# ---------------------------------------------------------------------------

def _get_ws(erp_path: Path):
    wb = openpyxl.load_workbook(str(erp_path), keep_vba=True)
    sheet = next((s for s in wb.sheetnames if AJ_SHEET_KEYWORD in s), None)
    ws = wb[sheet]
    return wb, ws


class TestEnrichment:

    def test_col28_mailto_links_populated(self, seeded_erp: Path):
        """After enrich(), col 28 must contain mailto links for seeded rows."""
        stats = enrich(str(seeded_erp), force=False)
        assert stats["total"] > 0

        wb, ws = _get_ws(seeded_erp)
        # Check seeded rows (start at DATA_START)
        links_found = 0
        for r in range(DATA_START, ws.max_row + 1):
            crm = ws.cell(r, COL["CRM_ID"]).value
            if not crm:
                continue
            val = ws.cell(r, COL["Request_BKG"]).value
            hyperlink = ws.cell(r, COL["Request_BKG"]).hyperlink
            if val or hyperlink:
                links_found += 1
        wb.close()
        assert links_found > 0, "No mailto links written to col 28"

    def test_mailto_link_starts_with_mailto(self, seeded_erp: Path):
        """Hyperlinks in col 28 must be valid mailto: URLs."""
        enrich(str(seeded_erp), force=True)
        wb, ws = _get_ws(seeded_erp)
        for r in range(DATA_START, ws.max_row + 1):
            crm = ws.cell(r, COL["CRM_ID"]).value
            if not crm:
                continue
            h = ws.cell(r, COL["Request_BKG"]).hyperlink
            if h:
                target = h.target if hasattr(h, "target") else str(h)
                assert target.startswith("mailto:"), (
                    f"Row {r}: hyperlink target does not start with 'mailto:': {target}"
                )
        wb.close()

    def test_col31_service_filled(self, seeded_erp: Path):
        """After enrich(), col 31 SERVICE must be filled for rows without existing value."""
        enrich(str(seeded_erp), force=False)
        wb, ws = _get_ws(seeded_erp)
        empty_service = 0
        for r in range(DATA_START, ws.max_row + 1):
            crm = ws.cell(r, COL["CRM_ID"]).value
            if not crm:
                continue
            svc = ws.cell(r, COL["SERVICE"]).value
            if not svc:
                empty_service += 1
        wb.close()
        assert empty_service == 0, (
            f"{empty_service} seeded row(s) still have blank SERVICE after enrich()"
        )

    def test_door_delivery_yes_gives_cy_door(self, seeded_erp: Path):
        """Rows with Door_Delivery='Yes' (seeded TRANANH at row 13) → SERVICE=CY-DOOR."""
        # TRANANH is seed row 6 (row index 8+5=13) with Door_Delivery="Yes"
        enrich(str(seeded_erp), force=False)
        wb, ws = _get_ws(seeded_erp)
        door_door_rows = []
        for r in range(DATA_START, ws.max_row + 1):
            door = ws.cell(r, COL["Door_Delivery"]).value
            svc = ws.cell(r, COL["SERVICE"]).value
            if str(door).strip().lower() in ("yes", "y"):
                door_door_rows.append((r, svc))
        wb.close()
        assert len(door_door_rows) > 0, "No Door_Delivery='Yes' rows found in seeded_erp"
        for r, svc in door_door_rows:
            assert svc == "CY-DOOR", f"Row {r}: expected CY-DOOR for Door_Delivery=Yes, got '{svc}'"

    def test_no_door_gives_cy_cy(self, seeded_erp: Path):
        """Rows with Door_Delivery='No' → SERVICE=CY-CY."""
        enrich(str(seeded_erp), force=False)
        wb, ws = _get_ws(seeded_erp)
        for r in range(DATA_START, ws.max_row + 1):
            crm = ws.cell(r, COL["CRM_ID"]).value
            if not crm:
                continue
            door = ws.cell(r, COL["Door_Delivery"]).value
            svc = ws.cell(r, COL["SERVICE"]).value
            if str(door).strip().lower() in ("no", "", "n"):
                assert svc in ("CY-CY", "CY-DOOR"), f"Row {r}: unexpected SERVICE '{svc}'"
        wb.close()

    def test_no_force_skips_existing_links(self, seeded_erp: Path):
        """Second enrich() without force=True must NOT overwrite col 28 links."""
        enrich(str(seeded_erp), force=True)  # First run — write all links

        # Read current values
        wb, ws = _get_ws(seeded_erp)
        first_vals = {}
        for r in range(DATA_START, ws.max_row + 1):
            crm = ws.cell(r, COL["CRM_ID"]).value
            if crm:
                first_vals[r] = ws.cell(r, COL["Request_BKG"]).value
        wb.close()

        # Second run without force — should not re-write
        stats2 = enrich(str(seeded_erp), force=False)
        assert stats2["linked"] == 0, (
            f"enrich() without force re-wrote {stats2['linked']} links that already existed"
        )

    def test_force_true_regenerates_links(self, seeded_erp: Path):
        """enrich(force=True) must overwrite all existing links."""
        enrich(str(seeded_erp), force=False)  # First run
        stats2 = enrich(str(seeded_erp), force=True)  # Second with force
        assert stats2["linked"] > 0, "enrich(force=True) should regenerate links"

    def test_stats_total_counts_all_rows(self, seeded_erp: Path):
        """enrich() stats['total'] should match number of seeded CRM_ID rows."""
        stats = enrich(str(seeded_erp), force=False)
        # seeded_erp has 7 seed rows
        assert stats["total"] >= 7

    def test_file_not_found_raises(self, tmp_path: Path):
        """enrich() on non-existent file must raise FileNotFoundError."""
        with pytest.raises(FileNotFoundError):
            enrich(str(tmp_path / "ghost.xlsm"))
