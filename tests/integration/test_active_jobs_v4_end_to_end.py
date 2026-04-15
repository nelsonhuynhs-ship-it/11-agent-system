"""
Active Jobs v4 — end-to-end integration test
============================================
Seeds a COPY of ERP_Master_v14.xlsm with 7 test jobs covering all pipeline stages,
then runs every Phase-1/Phase-2 helper in sequence and verifies:
  - Each helper exits 0 or returns expected result
  - Ribbon (customUI14.xml) is still present after every save
  - Each feature touches the expected columns / sheets
"""
from __future__ import annotations

import os
import shutil
import subprocess
import sys
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT / "ERP" / "core"))
from ribbon_guard import save_preserving_ribbon  # noqa: E402
from active_jobs_cols import COL as _COL  # noqa: E402

PYTHON = r"C:\Users\Nelson\anaconda3\python.exe"
LIVE_ERP = Path(r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm")


def _ribbon_intact(xlsm: Path) -> bool:
    with zipfile.ZipFile(xlsm) as z:
        names = z.namelist()
        if "customUI/customUI14.xml" not in names:
            return False
        rels = z.read("_rels/.rels").decode("utf-8")
        return "customUI14.xml" in rels


def _run(script_rel: str, extra_args: list[str]) -> tuple[int, str]:
    script = str(REPO_ROOT / script_rel)
    cmd = [PYTHON, script] + extra_args
    result = subprocess.run(cmd, cwd=str(REPO_ROOT), capture_output=True, text=True,
                            encoding="utf-8", errors="replace", timeout=180)
    return result.returncode, (result.stdout or "") + (result.stderr or "")


def _seed_active_jobs(xlsm: Path):
    """Seed 7 rows covering each pipeline stage + edge cases."""
    now = datetime(2026, 4, 14, 10, 0)
    wb = openpyxl.load_workbook(xlsm, keep_vba=True)
    ws = wb["Active Jobs"]
    rows = [
        # r=8: NAFOODS — booked only, stage 1
        dict(CRM_ID="NAFOODS", Routing="HPH-LOS ANGELES, CA VIA USLGB",
             Bkg_No="YM-8834501", ETD=now - timedelta(days=2),
             ETA=None, Carrier="YML", Contract_Type="SHA0005N25",
             Container_Type="40HQ", Quantity=2,
             Selling_Rate=3200, Buying_Rate=2650, Profit=1100,
             Status="Booked", Door_Delivery="Yes", Door_Address="LOS ANGELES, CA",
             FAST_JOB_NO="se2604/266", HBL_NO="PELP26040260"),
        # r=9: VIFON — in transit, stage 5
        dict(CRM_ID="VIFON EXPORT", Routing="HCM-USNYC", Bkg_No="CMA-V-2001",
             ETD=now - timedelta(days=5), ETA=None, Carrier="CMA",
             Contract_Type="CMA-2026-FIX", Container_Type="40GP", Quantity=1,
             Selling_Rate=2857, Buying_Rate=2789, Profit=68,
             Status="In Transit", Door_Delivery="No", Door_Address="",
             FAST_JOB_NO="SE2604/0280", HBL_NO="PNYC26040596"),
        # r=10: SIRI reefer — needs insurance + stage ATD
        dict(CRM_ID="SIRI", Routing="HPH-CHICAGO, IL VIA USLAX", Bkg_No="ONE-BKG-301",
             ETD=now - timedelta(days=1), ETA=None, Carrier="ONE",
             Contract_Type="SHA0007N26 SOC", Container_Type="40RF", Quantity=1,
             Selling_Rate=5785, Buying_Rate=5655, Profit=130,
             Status="Booked", Door_Delivery="Yes", Door_Address="CHICAGO, IL",
             FAST_JOB_NO="SE2604/0310", HBL_NO="PCHI26040310"),
        # r=11: TRAN ANH — release email sent 3h ago, ETA within window → URGENT
        dict(CRM_ID="TRAN ANH", Routing="HCM-USLGB", Bkg_No="YM-8840200",
             ETD=now - timedelta(days=22), ETA=now + timedelta(days=1),
             Carrier="YML", Contract_Type="YML-FIX", Container_Type="20GP", Quantity=1,
             Selling_Rate=1800, Buying_Rate=1550, Profit=250,
             Status="In Transit", Door_Delivery="No", Door_Address="",
             FAST_JOB_NO="SE2604/0400", HBL_NO="PLGB26040400",
             RELEASE_EMAIL_SENT_col=33, RELEASE_EMAIL_SENT_val=now - timedelta(hours=3)),
        # r=12: WESTFOOD — delivered, stage 7
        dict(CRM_ID="WESTFOOD", Routing="HCM-USNYC", Bkg_No="WAN-3301",
             ETD=now - timedelta(days=30), ETA=now - timedelta(days=2),
             ATA=now - timedelta(days=1),
             Carrier="WAN HAI", Contract_Type="WHL-FIX", Container_Type="40HQ",
             Quantity=1, Selling_Rate=3320, Buying_Rate=3120, Profit=200,
             Status="Delivered", Door_Delivery="No", Door_Address="",
             FAST_JOB_NO="SE2604/0647", HBL_NO="PNYC26040596"),
        # r=13: PANDA — invalid FAST ID (bad format) — F4 should flag
        dict(CRM_ID="PANDA DAD", Routing="HCM-LAX/LGB", Bkg_No="HPL-2501",
             ETD=now - timedelta(days=3), ETA=None, Carrier="HPL",
             Contract_Type="HPL-FAK", Container_Type="40GP", Quantity=1,
             Selling_Rate=2455, Buying_Rate=2327, Profit=128,
             Status="Booked", Door_Delivery="No", Door_Address="",
             FAST_JOB_NO="BAD/X1", HBL_NO="PLGB26040599"),
        # r=14: HML — 20RF for reefer plug optimization
        dict(CRM_ID="HML", Routing="DAD-LAX/LGB", Bkg_No="ONE-BKG-501",
             ETD=now - timedelta(days=5), ETA=now + timedelta(days=2),
             Carrier="ONE", Contract_Type="ONE-SCFI", Container_Type="20RF", Quantity=1,
             Selling_Rate=2900, Buying_Rate=2760, Profit=140,
             Status="In Transit", Door_Delivery="No", Door_Address="",
             FAST_JOB_NO="SE2604/0650", HBL_NO="PLGB26040650"),
    ]

    # Use canonical source-of-truth COL map (ERP/core/active_jobs_cols.py)
    # Map seed-data keys to COL dict names — FAST_JOB_NO is the v4 FAST_ID col.
    KEY_TO_COL = {
        "CRM_ID": "CRM_ID", "Routing": "Routing", "Bkg_No": "Bkg_No",
        "ETD": "ETD", "ETA": "ETA", "ATA": "ATA", "Carrier": "Carrier",
        "Contract_Type": "Contract_Type", "Container_Type": "Container_Type",
        "Quantity": "Quantity", "Selling_Rate": "Selling_Rate",
        "Buying_Rate": "Buying_Rate", "Profit": "Profit", "Status": "Status",
        "Door_Delivery": "Door_Delivery", "Door_Address": "Door_Address",
        "FAST_JOB_NO": "FAST_ID", "HBL_NO": "HBL_NO",
    }
    # Cols that may carry stale values from live ERP — clear before seeding so
    # downstream helpers (enrichment, etc.) treat rows as fresh.
    CLEAR_COLS = ("SERVICE", "TRACKING", "TRACKING_STAGE", "Request_BKG",
                  "Profit", "Profit_Margin", "RELEASE_EMAIL_SENT",
                  "RELEASE_CONFIRMED", "PRICE_WATCH_STATUS", "PRICE_WATCH_DELTA")

    r = 8
    for row in rows:
        # Wipe stale cells from prior content in the live workbook copy
        for col_name in CLEAR_COLS:
            cell = ws.cell(r, _COL[col_name])
            cell.value = None
            cell.hyperlink = None  # critical — see vba-gotchas #7
        for key, col_name in KEY_TO_COL.items():
            if key in row:
                ws.cell(r, _COL[col_name], row[key])
        # special — RELEASE_EMAIL_SENT uses canonical col
        if "RELEASE_EMAIL_SENT_val" in row:
            ws.cell(r, _COL["RELEASE_EMAIL_SENT"], row["RELEASE_EMAIL_SENT_val"])
        r += 1
    save_preserving_ribbon(wb, str(xlsm))
    wb.close()


@pytest.fixture()
def seeded_live_copy(tmp_path: Path) -> Path:
    """Copy live ERP + seed — auto-cleaned by tmp_path fixture."""
    assert LIVE_ERP.exists(), f"Live ERP missing: {LIVE_ERP}"
    dst = tmp_path / "ERP_Master_v14.xlsm"
    shutil.copy2(LIVE_ERP, dst)
    _seed_active_jobs(dst)
    assert _ribbon_intact(dst), "seed stripped ribbon (should not happen)"
    return dst


def test_e2e_pipeline_preserves_ribbon(seeded_live_copy: Path):
    """Run every pipeline on seeded copy; ribbon must survive each step."""
    erp = str(seeded_live_copy)
    steps = [
        ("schema", "ERP/core/active_jobs_schema.py", ["--file", erp]),
        ("tracking", "ERP/jobs/shipment_tracker.py", ["--erp", erp]),
        ("price_watch", "ERP/intelligence/price_watch.py", ["--erp", erp]),
        ("release_alerts", "ERP/jobs/release_alerts.py",
         ["--erp", erp, "--now", "2026-04-14 13:30"]),
        ("transit_time", "ERP/jobs/transit_time.py", ["--erp", erp]),
        ("enrichment", "ERP/jobs/enrichment.py", ["--erp", erp]),
        ("fast_id_check", "ERP/jobs/fast_id.py", ["--erp", erp, "--fix"]),
        ("reefer_plug", "ERP/jobs/reefer_plug.py", ["--erp", erp, "--write"]),
    ]
    for name, script, args in steps:
        rc, out = _run(script, args)
        # release_alerts returns 1 when URGENT present — that's expected, not a fail
        if name == "release_alerts":
            assert rc in (0, 1), f"[{name}] exit={rc}\n{out[-600:]}"
        else:
            assert rc == 0, f"[{name}] exit={rc}\n{out[-600:]}"
        assert _ribbon_intact(seeded_live_copy), (
            f"[{name}] stripped the ribbon — check save_preserving_ribbon wiring"
        )


def test_e2e_monthly_report_generated(seeded_live_copy: Path, tmp_path: Path):
    out = tmp_path / "monthly.xlsx"
    rc, stdout = _run("ERP/intelligence/monthly_report.py",
                      ["--erp", str(seeded_live_copy), "--month", "2026-04",
                       "--out", str(out)])
    assert rc == 0, stdout
    assert out.exists(), "Monthly report not created"
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    # Verify header structure — 24 cols
    assert ws.cell(3, 1).value == "No"
    assert ws.cell(3, 24).value is not None  # Net Profit header
    # Data rows exist (we seeded 7 jobs in April)
    data_rows = [r for r in range(5, ws.max_row + 1)
                 if ws.cell(r, 1).value not in (None, "TOTAL")]
    # 6 rows seeded in April (WESTFOOD is in March, so filtered out)
    assert len(data_rows) >= 5, f"Expected ≥5 April data rows, got {len(data_rows)}"
    wb.close()


def test_e2e_weekly_report_generated(seeded_live_copy: Path, tmp_path: Path):
    out = tmp_path / "weekly.xlsx"
    rc, stdout = _run("ERP/intelligence/weekly_report.py",
                      ["--erp", str(seeded_live_copy),
                       "--year", "2026", "--week", "16",
                       "--out", str(out)])
    assert rc == 0, stdout
    assert out.exists()


def test_e2e_transit_time_fills_eta(seeded_live_copy: Path):
    # Before: most seeded rows have ETA=None
    wb = openpyxl.load_workbook(seeded_live_copy, read_only=True)
    ws = wb["Active Jobs"]
    missing_before = sum(1 for r in range(8, 15)
                         if ws.cell(r, _COL["ETA"]).value is None)
    wb.close()

    rc, out = _run("ERP/jobs/transit_time.py", ["--erp", str(seeded_live_copy)])
    assert rc == 0, out

    wb = openpyxl.load_workbook(seeded_live_copy, read_only=True)
    ws = wb["Active Jobs"]
    missing_after = sum(1 for r in range(8, 15)
                        if ws.cell(r, _COL["ETA"]).value is None)
    wb.close()
    assert missing_after < missing_before, (
        f"Transit time did not fill any ETA: before={missing_before}, after={missing_after}"
    )


def test_e2e_fast_id_normalizes(seeded_live_copy: Path):
    # Row 8 has FAST_JOB_NO="se2604/266" (lowercase, short seq) → should normalize
    rc, out = _run("ERP/jobs/fast_id.py",
                   ["--erp", str(seeded_live_copy), "--fix"])
    assert rc == 0, out
    wb = openpyxl.load_workbook(seeded_live_copy, read_only=True)
    ws = wb["Active Jobs"]
    val = ws.cell(8, _COL["FAST_ID"]).value
    assert val == "SE2604/0266", f"Expected 'SE2604/0266', got {val!r}"
    wb.close()


def test_e2e_release_alerts_catches_urgent(seeded_live_copy: Path):
    # Row 11 TRAN ANH has RELEASE_EMAIL_SENT=3h ago, no confirm, ETA in 1 day
    # Should produce URGENT alert (exit 1)
    rc, out = _run("ERP/jobs/release_alerts.py",
                   ["--erp", str(seeded_live_copy),
                    "--now", "2026-04-14 13:30",
                    "--hours", "2",
                    "--eta-window-days", "3"])
    # Exit 1 = URGENT present (by design)
    assert rc == 1, f"Expected URGENT (exit 1), got exit={rc}\n{out}"
    assert "URGENT" in out


def test_e2e_enrichment_fills_service_and_mailto(seeded_live_copy: Path):
    rc, out = _run("ERP/jobs/enrichment.py", ["--erp", str(seeded_live_copy)])
    assert rc == 0, out
    wb = openpyxl.load_workbook(seeded_live_copy)
    ws = wb["Active Jobs"]
    # Row 8 NAFOODS Door=Yes → SERVICE=CY-DOOR, Request_BKG has mailto hyperlink
    assert ws.cell(8, _COL["SERVICE"]).value == "CY-DOOR"
    req_cell = ws.cell(8, _COL["Request_BKG"])
    assert req_cell.hyperlink is not None
    assert req_cell.hyperlink.target.startswith("mailto:")
    # Row 9 VIFON Door=No → CY-CY
    assert ws.cell(9, _COL["SERVICE"]).value == "CY-CY"
    wb.close()
