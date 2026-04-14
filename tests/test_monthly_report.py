"""test_monthly_report.py — Monthly report exporter tests.

Verifies:
  - 24-col output with correct header structure (row 3 main, row 4 sub)
  - Data starts at row 5, TOTAL row present
  - Volume cell mapping per container type
  - parse_month edge cases
  - extract_volume column assignment
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "ERP" / "intelligence"))
from monthly_report import (
    CONT_COL_MAP,
    extract_volume,
    filter_by_month,
    parse_month,
    write_report,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_rows(n: int = 3) -> list[dict]:
    """Create n minimal Active Jobs rows matching write_report() expectations."""
    base = datetime(2026, 3, 15)
    carriers = ["ONE", "MSC", "CMA"]
    containers = ["40HQ", "20GP", "40RF"]
    customers = ["NAFOODS", "VIFON", "SIRI"]
    rows = []
    for i in range(n):
        rows.append({
            "CRM_ID": customers[i % len(customers)],
            "Customer_Type": "DIRECT",
            "Routing": f"HPH-USLGB",
            "Bkg_No": f"BKG00{i+1}",
            "ETD": datetime(2026, 3, 10 + i),
            "ETA": datetime(2026, 4, 10 + i),
            "ATA": None,
            "Carrier": carriers[i % len(carriers)],
            "Contract_Type": "SHA0005N25",
            "Container_Type": containers[i % len(containers)],
            "Quantity": i + 1,
            "Selling_Rate": 1200 + i * 100,
            "Buying_Rate": 1000 + i * 100,
            "Profit": 200,
            "Profit_Margin": 0.2,
            "Status": "Confirmed",
            "Door_Address": "LOS ANGELES, CA",
            "Notes": "",
            "Created_Date": datetime(2026, 3, 1),
            "FAST_JOB_NO": f"F202600{i+1}",
            "HBL_NO": f"HBL00{i+1}",
            "SERVICE": "CY-CY",
        })
    return rows


# ---------------------------------------------------------------------------
# parse_month tests
# ---------------------------------------------------------------------------

class TestParseMonth:

    def test_iso_format(self):
        y, m, label = parse_month("2026-04")
        assert y == 2026
        assert m == 4
        assert "APR" in label
        assert "2026" in label

    def test_abbr_dash_year(self):
        y, m, label = parse_month("APR-2026")
        assert y == 2026
        assert m == 4

    def test_abbr_no_separator(self):
        y, m, label = parse_month("Apr2026")
        assert y == 2026
        assert m == 4

    def test_lowercase_month(self):
        y, m, _ = parse_month("mar-2026")
        assert m == 3

    def test_none_returns_current_month(self):
        now = datetime.now()
        y, m, label = parse_month(None)
        assert y == now.year
        assert m == now.month

    def test_invalid_raises(self):
        with pytest.raises(ValueError, match="Can't parse"):
            parse_month("NOTAMONTH")

    def test_single_digit_month(self):
        y, m, _ = parse_month("2026-4")
        assert m == 4


# ---------------------------------------------------------------------------
# extract_volume tests
# ---------------------------------------------------------------------------

class TestExtractVolume:

    def test_40hq_maps_to_col_o(self):
        col, qty = extract_volume("40HQ", 2)
        assert col == "O"
        assert qty == 2

    def test_40rf_maps_to_col_p(self):
        col, qty = extract_volume("40RF", 1)
        assert col == "P"
        assert qty == 1

    def test_20gp_maps_to_col_m(self):
        col, qty = extract_volume("20GP", 3)
        assert col == "M"
        assert qty == 3

    def test_40gp_maps_to_col_n(self):
        col, _ = extract_volume("40GP", 1)
        assert col == "N"

    def test_air_detected_from_status(self):
        col, qty = extract_volume("40HQ", 2, status="AIR SHIPMENT")
        assert col == "J"

    def test_lcl_detected_from_status(self):
        col, qty = extract_volume("20GP", 1, status="LCL")
        assert col == "K"

    def test_unknown_container_defaults_to_n(self):
        col, _ = extract_volume("UNKNOWN", 1)
        assert col == "N"

    def test_case_insensitive(self):
        col, _ = extract_volume("40hq", 1)
        assert col == "O"


# ---------------------------------------------------------------------------
# write_report tests
# ---------------------------------------------------------------------------

class TestWriteReport:

    def test_output_has_24_columns_in_data_rows(self, tmp_path: Path):
        rows = _make_rows(3)
        out = tmp_path / "report.xlsx"
        write_report(rows, "MAR 2026", str(out))

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        # Data rows start at row 5; check last non-empty col in first data row
        data_row = ws[5]
        # Count cols from A to X (24)
        values = [ws.cell(5, c).value for c in range(1, 25)]
        # At least the sequence number (col 1) and net profit (col 24) must be set
        assert ws.cell(5, 1).value == 1, "First data row seq number should be 1"
        assert ws.cell(5, 24).value is not None, "Net profit col 24 should be set"
        wb.close()

    def test_main_headers_row3(self, tmp_path: Path):
        rows = _make_rows(2)
        out = tmp_path / "report.xlsx"
        write_report(rows, "MAR 2026", str(out))

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        # Col A row 3 = "No"
        assert ws.cell(3, 1).value == "No"
        # Volume header spans cols 10-17; check col 10 row 3 has "Volume"
        assert ws.cell(3, 10).value == "Volume"
        wb.close()

    def test_sub_headers_row4(self, tmp_path: Path):
        rows = _make_rows(1)
        out = tmp_path / "report.xlsx"
        write_report(rows, "MAR 2026", str(out))

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        # Sub-headers at row 4
        sub = {10: "AIR", 11: "LCL", 12: "20RF", 13: "20'",
               14: "40'", 15: "HC", 16: "40RF", 17: "45"}
        for col, expected in sub.items():
            assert ws.cell(4, col).value == expected, (
                f"Sub-header col {col}: expected '{expected}', got '{ws.cell(4, col).value}'"
            )
        wb.close()

    def test_total_row_present(self, tmp_path: Path):
        rows = _make_rows(3)
        out = tmp_path / "report.xlsx"
        stats = write_report(rows, "MAR 2026", str(out))

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        total_row = 5 + len(rows)
        val = ws.cell(total_row, 1).value
        assert val == "TOTAL", f"Expected 'TOTAL' at row {total_row} col 1, got '{val}'"
        wb.close()

    def test_total_row_buy_sell_sum(self, tmp_path: Path):
        rows = _make_rows(3)
        out = tmp_path / "report.xlsx"
        stats = write_report(rows, "MAR 2026", str(out))

        # Verify stats dict totals
        expected_buy = sum(float(r["Buying_Rate"]) * r["Quantity"] for r in rows)
        expected_sell = sum(float(r["Selling_Rate"]) * r["Quantity"] for r in rows)
        assert abs(stats["total_buy"] - expected_buy) < 0.01
        assert abs(stats["total_sell"] - expected_sell) < 0.01
        assert abs(stats["total_net"] - (expected_sell - expected_buy)) < 0.01

    def test_40hq_volume_in_col_o(self, tmp_path: Path):
        rows = [{
            "CRM_ID": "TEST", "Customer_Type": "DIRECT", "Routing": "HPH-USLGB",
            "Bkg_No": "B001", "ETD": datetime(2026, 3, 1), "ETA": datetime(2026, 4, 1),
            "ATA": None, "Carrier": "ONE", "Contract_Type": "C001",
            "Container_Type": "40HQ", "Quantity": 2,
            "Selling_Rate": 1200, "Buying_Rate": 1000, "Profit": 200,
            "Profit_Margin": 0.2, "Status": "OK", "Door_Address": "LA",
            "Notes": "", "Created_Date": datetime(2026, 3, 1),
            "FAST_JOB_NO": "F001", "HBL_NO": "H001", "SERVICE": "CY-CY",
        }]
        out = tmp_path / "report_40hq.xlsx"
        write_report(rows, "MAR 2026", str(out))

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        # Col O = column 15 = HC (40HQ maps to O)
        assert ws.cell(5, 15).value == 2, "40HQ qty should be in col O (15)"
        wb.close()

    def test_data_starts_at_row5(self, tmp_path: Path):
        rows = _make_rows(2)
        out = tmp_path / "report.xlsx"
        write_report(rows, "MAR 2026", str(out))

        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        # Row 4 = sub-headers, not data
        assert ws.cell(4, 1).value != 1, "Row 4 should be sub-headers, not seq number 1"
        # Row 5 = first data row
        assert ws.cell(5, 1).value == 1
        wb.close()

    def test_empty_rows_no_crash(self, tmp_path: Path):
        out = tmp_path / "empty.xlsx"
        stats = write_report([], "MAR 2026", str(out))
        assert stats["rows"] == 0
        assert stats["total_net"] == 0.0


# ---------------------------------------------------------------------------
# filter_by_month tests
# ---------------------------------------------------------------------------

class TestFilterByMonth:

    def test_filters_by_etd(self):
        rows = [
            {"ETD": datetime(2026, 3, 10), "Created_Date": None},
            {"ETD": datetime(2026, 4, 5), "Created_Date": None},
            {"ETD": datetime(2026, 3, 31), "Created_Date": None},
        ]
        result = filter_by_month(rows, 2026, 3)
        assert len(result) == 2

    def test_falls_back_to_created_date(self):
        rows = [
            {"ETD": None, "Created_Date": datetime(2026, 3, 5)},
        ]
        result = filter_by_month(rows, 2026, 3)
        assert len(result) == 1

    def test_excludes_non_datetime(self):
        rows = [
            {"ETD": "2026-03-10", "Created_Date": None},
        ]
        result = filter_by_month(rows, 2026, 3)
        assert len(result) == 0
