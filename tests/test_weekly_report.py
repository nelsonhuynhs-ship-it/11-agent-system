"""
tests/test_weekly_report.py — Unit tests for ERP.intelligence.weekly_report

Tests:
  1. iso_week returns correct (year, week) for 2026-04-14 → week 16
  2. build_weekly_summary produces correct rows from synthetic data
  3. write_weekly_report xlsx has 12 columns
"""
from __future__ import annotations

import os
import tempfile
from datetime import date, datetime

import pytest

# Module under test
from ERP.intelligence.weekly_report import (
    build_weekly_summary,
    iso_week,
    write_weekly_report,
)


# ── 1. iso_week ────────────────────────────────────────────────────────────────

def test_iso_week_2026_04_14():
    """2026-04-14 (Tuesday) must be ISO week 16."""
    y, w = iso_week(date(2026, 4, 14))
    assert y == 2026
    assert w == 16


def test_iso_week_start_of_year():
    """2026-01-01 (Thursday) is in ISO week 1 of 2026."""
    y, w = iso_week(date(2026, 1, 1))
    assert y == 2026
    assert w == 1


def test_iso_week_cross_year_boundary():
    """2025-12-29 (Monday) should be week 1 of 2026 by ISO-8601."""
    y, w = iso_week(date(2025, 12, 29))
    assert y == 2026
    assert w == 1


# ── 2. build_weekly_summary ────────────────────────────────────────────────────

def _make_job(crm_id: str, cont: str = "40HC", qty: int = 1, profit: float = 500.0) -> dict:
    return {
        "CRM_ID": crm_id,
        "Container_Type": cont,
        "Quantity": qty,
        "Profit": profit,
        "sales": "Nelson",
    }


def test_build_weekly_summary_basic():
    """Two jobs + one email row → Nelson row has correct counts."""
    jobs = [
        _make_job("CUST_A", cont="40HC", qty=1, profit=1000.0),
        _make_job("CUST_B", cont="20GP", qty=2, profit=500.0),
    ]
    emails = {"nelson@pudongprime.vn": 5}

    # No ERP file — use empty path so CRM lookup gracefully returns {}
    rows = build_weekly_summary(
        jobs, emails, {"nelson@pudongprime.vn": "Nelson"},
        erp_file="/nonexistent/path.xlsm",
        year=2026, week=16,
    )

    nelson_row = next(r for r in rows if r["Ten Sales"] == "Nelson")
    assert nelson_row["So shipment"] == 2
    # CUST_A: 1×40HC = 2 TEU; CUST_B: 2×20GP = 2 TEU → total 4 TEU
    assert nelson_row["VOL (TEU)"] == 4.0
    assert nelson_row["PROFIT ($)"] == 1500.0
    # Both customers are NOT in CRM (no erp_file), so they fall into existing bucket
    assert nelson_row["KH SDDV"] + nelson_row["KH MOI"] == 2
    assert nelson_row["HOAT DONG TUAN"] == 5


def test_build_weekly_summary_zero_jobs():
    """Sales with no jobs should have zeros, not KeyError."""
    rows = build_weekly_summary(
        jobs=[], emails={},
        sales_map={},
        erp_file="/nonexistent/path.xlsm",
        year=2026, week=16,
    )
    for r in rows:
        assert r["So shipment"] == 0
        assert r["VOL (TEU)"] == 0.0
        assert r["PROFIT ($)"] == 0.0


def test_build_weekly_summary_returns_all_sales():
    """One row per sales person in SALES_MAP (7 people)."""
    rows = build_weekly_summary(
        jobs=[], emails={},
        sales_map={},
        erp_file="/nonexistent/path.xlsm",
        year=2026, week=16,
    )
    names = {r["Ten Sales"] for r in rows}
    assert "Nelson" in names
    assert "Johnny" in names
    assert len(names) == 7


def test_build_weekly_summary_teu_factors():
    """TEU factors: 20GP=1, 40HC=2."""
    jobs = [
        _make_job("A", cont="20GP", qty=3, profit=0),   # 3 × 1 = 3 TEU
        _make_job("B", cont="40HC", qty=1, profit=0),   # 1 × 2 = 2 TEU
    ]
    rows = build_weekly_summary(
        jobs, {}, {},
        erp_file="/nonexistent/path.xlsm",
        year=2026, week=16,
    )
    nelson = next(r for r in rows if r["Ten Sales"] == "Nelson")
    assert nelson["VOL (TEU)"] == 5.0


# ── 3. write_weekly_report — 12 columns ───────────────────────────────────────

def test_write_weekly_report_12_columns(tmp_path):
    """Output xlsx must have exactly 12 header columns in row 3."""
    rows = [
        {
            "STT": 1, "Ten Sales": "Nelson", "So shipment": 3,
            "VOL (TEU)": 6.0, "PROFIT ($)": 2500.0,
            "KH SDDV": 2, "KH MOI": 1, "GAP KH": 0, "KH TIEM NANG": 0,
            "HOAT DONG TUAN": 12, "% HOAN THANH": "", "PLAN TUAN NAY": "",
        }
    ]
    out = str(tmp_path / "test_weekly.xlsx")
    stats = write_weekly_report(rows, year=2026, week=16, out_file=out)

    assert os.path.exists(out)
    assert stats["rows"] == 1

    import openpyxl
    wb = openpyxl.load_workbook(out, read_only=True)
    ws = wb.active
    # Row 3 is the header row — count non-None cells
    header_cells = [ws.cell(3, c).value for c in range(1, 20)]
    non_empty = [v for v in header_cells if v]
    wb.close()
    assert len(non_empty) == 12


def test_write_weekly_report_title_contains_week(tmp_path):
    """Title cell (A1) must contain the week number and year."""
    out = str(tmp_path / "title_test.xlsx")
    write_weekly_report([], year=2026, week=16, out_file=out)

    import openpyxl
    wb = openpyxl.load_workbook(out, read_only=True)
    ws = wb.active
    title = ws["A1"].value or ""
    wb.close()
    assert "16" in title
    assert "2026" in title


def test_write_weekly_report_total_row(tmp_path):
    """TOTAL row must appear and accumulate shipments correctly."""
    rows = [
        {
            "STT": 1, "Ten Sales": "Nelson", "So shipment": 4,
            "VOL (TEU)": 8.0, "PROFIT ($)": 3000.0,
            "KH SDDV": 3, "KH MOI": 1, "GAP KH": 0, "KH TIEM NANG": 0,
            "HOAT DONG TUAN": 20, "% HOAN THANH": "", "PLAN TUAN NAY": "",
        },
        {
            "STT": 2, "Ten Sales": "Johnny", "So shipment": 2,
            "VOL (TEU)": 4.0, "PROFIT ($)": 1000.0,
            "KH SDDV": 1, "KH MOI": 1, "GAP KH": 0, "KH TIEM NANG": 0,
            "HOAT DONG TUAN": 10, "% HOAN THANH": "", "PLAN TUAN NAY": "",
        },
    ]
    out = str(tmp_path / "total_test.xlsx")
    stats = write_weekly_report(rows, year=2026, week=16, out_file=out)
    assert stats["total_shipments"] == 6
    assert stats["total_profit"] == 4000.0
