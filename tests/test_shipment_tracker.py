"""test_shipment_tracker.py — Shipment tracker 7-stage pipeline tests.

Ports the compute_stage unit tests from the module + adds integration test.
"""
from __future__ import annotations

import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "ERP" / "jobs"))
from shipment_tracker import (
    STAGE_NAMES,
    _is_set,
    compute_stage,
    update_active_jobs,
)

AJ_SHEET_KEYWORD = "Active"
AJ_DATA_START = 8


# ---------------------------------------------------------------------------
# _is_set helper
# ---------------------------------------------------------------------------

class TestIsSet:
    def test_none_returns_false(self):
        assert _is_set(None) is False

    def test_empty_string_returns_false(self):
        assert _is_set("") is False

    def test_zero_int_returns_true(self):
        assert _is_set(0) is True

    def test_nonempty_string_returns_true(self):
        assert _is_set("BKG001") is True

    def test_datetime_returns_true(self):
        assert _is_set(datetime.now()) is True

    def test_float_nan_returns_false(self):
        import math
        assert _is_set(float("nan")) is False


# ---------------------------------------------------------------------------
# compute_stage unit tests
# ---------------------------------------------------------------------------

class TestComputeStage:
    NOW = datetime(2026, 4, 14, 12, 0, 0)

    def _row(self, **kw) -> dict:
        base = {
            "CRM_ID": "TEST", "Routing": "HPH-USLGB", "Bkg_No": None,
            "ETD": None, "ETA": None, "ATA": None,
            "Carrier": "ONE", "Contract_Type": None, "Container_Type": "40HQ",
            "Quantity": 1, "Status": "", "SI_Received": None,
            "CY_Cutoff": None, "Notes": "", "TRACKING_STAGE": None,
            "RELEASE_CONFIRMED": None,
        }
        base.update(kw)
        return base

    def test_stage7_delivered_status(self):
        row = self._row(Status="Delivered")
        assert compute_stage(row, self.NOW) == 7

    def test_stage7_done_status(self):
        row = self._row(Status="Done")
        assert compute_stage(row, self.NOW) == 7

    def test_stage7_release_confirmed(self):
        row = self._row(RELEASE_CONFIRMED=datetime(2026, 4, 13))
        assert compute_stage(row, self.NOW) == 7

    def test_stage6_ata_set(self):
        row = self._row(ATA=datetime(2026, 4, 10))
        assert compute_stage(row, self.NOW) == 6

    def test_stage6_eta_reached(self):
        row = self._row(ETA=self.NOW - timedelta(hours=1))
        assert compute_stage(row, self.NOW) == 6

    def test_stage5_transit_status(self):
        row = self._row(Status="In Transit")
        assert compute_stage(row, self.NOW) == 5

    def test_stage5_etd_past(self):
        row = self._row(ETD=self.NOW - timedelta(days=2))
        assert compute_stage(row, self.NOW) == 5

    def test_stage4_gate_in_status(self):
        row = self._row(Status="GATE IN")
        assert compute_stage(row, self.NOW) == 4

    def test_stage4_gate_in_notes(self):
        row = self._row(Notes="Container GATED at terminal")
        assert compute_stage(row, self.NOW) == 4

    def test_stage4_cy_cutoff_past(self):
        row = self._row(CY_Cutoff=self.NOW - timedelta(hours=6))
        assert compute_stage(row, self.NOW) == 4

    def test_stage3_si_received(self):
        row = self._row(SI_Received=datetime(2026, 4, 10))
        assert compute_stage(row, self.NOW) == 3

    def test_stage2_bkg_and_contract(self):
        row = self._row(Bkg_No="BKG001", Contract_Type="SHA0005N25")
        assert compute_stage(row, self.NOW) == 2

    def test_stage1_bkg_no_only(self):
        """Bkg_No without Contract_Type → stage 1 (not 2)."""
        row = self._row(Bkg_No="BKG001", Contract_Type=None)
        assert compute_stage(row, self.NOW) == 1

    def test_stage1_booked_status(self):
        row = self._row(Status="Booked")
        assert compute_stage(row, self.NOW) == 1

    def test_stage1_default_empty_row(self):
        row = self._row()
        assert compute_stage(row, self.NOW) == 1

    def test_highest_stage_wins(self):
        """When multiple signals present, highest stage takes precedence."""
        row = self._row(
            Bkg_No="BKG001", Contract_Type="C001",  # stage 2
            SI_Received=datetime(2026, 4, 1),         # stage 3
            ETD=self.NOW - timedelta(days=5),          # stage 5
        )
        assert compute_stage(row, self.NOW) == 5

    def test_stage7_beats_all(self):
        """Delivered status overrides all other signals."""
        row = self._row(
            Status="Delivered",
            ATA=datetime(2026, 4, 10),
            ETD=self.NOW - timedelta(days=30),
        )
        assert compute_stage(row, self.NOW) == 7

    def test_stage_names_mapping(self):
        """All 7 stage names must be in STAGE_NAMES."""
        assert set(STAGE_NAMES.keys()) == {1, 2, 3, 4, 5, 6, 7}


# ---------------------------------------------------------------------------
# Integration: update_active_jobs on seeded_erp
# ---------------------------------------------------------------------------

class TestUpdateActiveJobs:

    def test_update_writes_tracking_stage_col36(self, seeded_erp: Path):
        """update_active_jobs() must write 'N/7 Name' labels to col 36 (TRACKING_STAGE)."""
        stats = update_active_jobs(str(seeded_erp), dry_run=False)
        assert stats["total"] > 0

        wb = openpyxl.load_workbook(str(seeded_erp), keep_vba=True)
        sheet = next(s for s in wb.sheetnames if AJ_SHEET_KEYWORD in s)
        ws = wb[sheet]

        filled = 0
        for r in range(AJ_DATA_START, ws.max_row + 1):
            crm = ws.cell(r, 1).value
            if not crm:
                continue
            stage_val = ws.cell(r, 36).value
            if stage_val:
                filled += 1
                # Must match pattern "N/7 Name"
                assert "/" in str(stage_val), f"Row {r} TRACKING_STAGE '{stage_val}' missing '/'"
                assert "7" in str(stage_val), f"Row {r} TRACKING_STAGE '{stage_val}' missing '7'"
        wb.close()
        assert filled > 0, "No TRACKING_STAGE values written to col 36"

    def test_stage_format_correct(self, seeded_erp: Path):
        """TRACKING_STAGE values must follow 'N/7 StageName' format exactly."""
        update_active_jobs(str(seeded_erp), dry_run=False)

        wb = openpyxl.load_workbook(str(seeded_erp), keep_vba=True)
        sheet = next(s for s in wb.sheetnames if AJ_SHEET_KEYWORD in s)
        ws = wb[sheet]

        valid_labels = {f"{n}/7 {STAGE_NAMES[n]}" for n in range(1, 8)}
        for r in range(AJ_DATA_START, ws.max_row + 1):
            crm = ws.cell(r, 1).value
            if not crm:
                continue
            val = ws.cell(r, 36).value
            if val:
                assert val in valid_labels, (
                    f"Row {r} TRACKING_STAGE '{val}' not in valid labels"
                )
        wb.close()

    def test_stats_by_stage_sums_to_total(self, seeded_erp: Path):
        """sum(stats.by_stage.values()) must equal stats.total."""
        stats = update_active_jobs(str(seeded_erp), dry_run=False)
        assert sum(stats["by_stage"].values()) == stats["total"]

    def test_dry_run_does_not_write(self, seeded_erp: Path):
        """dry_run=True must not modify col 36 (TRACKING_STAGE) on disk."""
        # Snapshot col 36 before dry_run
        wb = openpyxl.load_workbook(str(seeded_erp), keep_vba=True)
        sheet = next(s for s in wb.sheetnames if AJ_SHEET_KEYWORD in s)
        ws = wb[sheet]
        before = {r: ws.cell(r, 36).value
                  for r in range(AJ_DATA_START, ws.max_row + 1)}
        wb.close()

        update_active_jobs(str(seeded_erp), dry_run=True)

        # Col 36 on disk must be identical to before — dry_run must not save
        wb2 = openpyxl.load_workbook(str(seeded_erp), keep_vba=True)
        sheet2 = next(s for s in wb2.sheetnames if AJ_SHEET_KEYWORD in s)
        ws2 = wb2[sheet2]
        for r in range(AJ_DATA_START, ws2.max_row + 1):
            after_val = ws2.cell(r, 36).value
            assert after_val == before[r], (
                f"dry_run modified col 36 at row {r}: "
                f"{before[r]!r} → {after_val!r}"
            )
        wb2.close()

    def test_file_not_found_raises(self, tmp_path: Path):
        with pytest.raises(FileNotFoundError):
            update_active_jobs(str(tmp_path / "ghost.xlsm"))
