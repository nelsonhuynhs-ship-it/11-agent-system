"""Shared pytest fixtures for ERP Phase 1 regression tests.

Provides:
  - erp_copy   : per-test copy of live ERP_Master_v14.xlsm in tmpdir (openpyxl-only)
  - seeded_erp : erp_copy pre-seeded with 5+ Active Jobs rows covering all 7 stages
  - sample_rules: booking_rules.json loaded as dict
  - excel_app  : xlwings session (kept for integration markers, no-op if xlwings absent)
  - erp_workbook: xlwings workbook fixture (skipped if xlwings absent)

CRITICAL: No fixture may write to the live OneDrive ERP_Master_v14.xlsm.
"""
from __future__ import annotations

import faulthandler
import gc
import json
import os
import shutil
import time
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import pytest

faulthandler.disable()

_DEFAULT_MASTER = Path("D:/OneDrive/NelsonData/erp/ERP_Master_v14.xlsm")
MASTER_XLSM = Path(os.environ.get("ERP_MASTER_XLSM", str(_DEFAULT_MASTER)))

RULES_FILE = Path(__file__).parent.parent / "ERP" / "carrier_rules" / "booking_rules.json"

# Active Jobs layout constants (mirror ERP modules)
AJ_SHEET_KEYWORD = "Active"
AJ_HDR_ROW = 7
AJ_DATA_START = 8


# ── openpyxl-based fixtures (no COM / no xlwings) ──────────────────────────

@pytest.fixture(scope="function")
def erp_copy(tmp_path):
    """Copy live ERP_Master_v14.xlsm to a tmpdir. Tests mutate freely."""
    if not MASTER_XLSM.exists():
        pytest.skip(f"Live ERP not found: {MASTER_XLSM}")
    dst = tmp_path / "ERP_Master_v14_test.xlsm"
    shutil.copy2(MASTER_XLSM, dst)
    yield dst
    # cleanup handled by tmp_path fixture


def _seed_active_jobs_rows(erp_path: Path) -> None:
    """Write 7 seed rows into Active Jobs (rows 8-14) covering stages 1-7.

    Uses COL dict from ERP.core.active_jobs_cols (v4 migration source of truth).
    Col positions may change; keys are field names, not integers.
    """
    import sys as _sys
    _sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "ERP" / "core"))
    from active_jobs_cols import COL  # type: ignore[import-untyped]

    wb = openpyxl.load_workbook(str(erp_path), keep_vba=True)
    sheet = next((s for s in wb.sheetnames if AJ_SHEET_KEYWORD in s), None)
    if sheet is None:
        wb.close()
        raise RuntimeError("Active Jobs sheet not found in ERP copy")
    ws = wb[sheet]

    now = datetime.now()

    # Seed rows keyed by field name (resolved to col via COL dict at write time)
    seeds = [
        # Stage 1 BKG — booking pending, no contract
        {"CRM_ID": "NAFOODS", "Routing": "HPH-USLGB", "Bkg_No": "NFBKG001",
         "Carrier": "ONE", "Container_Type": "40HQ", "Quantity": 2,
         "Status": "Booked", "Door_Delivery": "No"},
        # Stage 2 Conf — bkg + contract
        {"CRM_ID": "SIRI", "Routing": "HPH-USLGB", "Bkg_No": "SRBKG002",
         "Carrier": "MSC", "Contract_Type": "SHA0005N25",
         "Container_Type": "40HQ", "Quantity": 1,
         "Status": "Confirmed", "Door_Delivery": "No"},
        # Stage 3 SI Cut
        {"CRM_ID": "VIFON", "Routing": "HCM-USLAX", "Bkg_No": "VFBKG003",
         "Carrier": "CMA", "Contract_Type": "CMA-2026",
         "Container_Type": "40GP", "Quantity": 2,
         "Status": "Confirmed", "SI_Received": now - timedelta(days=2),
         "Door_Delivery": "No"},
        # Stage 5 ATD — ETD in past
        {"CRM_ID": "WESTFOOD", "Routing": "HPH-USNYC", "Bkg_No": "WFBKG004",
         "Carrier": "OOCL", "Contract_Type": "OO2026",
         "Container_Type": "40HQ", "Quantity": 1,
         "ETD": now - timedelta(days=10),
         "Status": "In Transit", "Door_Delivery": "No"},
        # Stage 6 ETA — ATA set
        {"CRM_ID": "PANDA", "Routing": "HCM-USLAX", "Bkg_No": "PDBKG005",
         "Carrier": "COSCO", "Contract_Type": "CS2026",
         "Container_Type": "20GP", "Quantity": 3,
         "ETD": now - timedelta(days=20),
         "ATA": now - timedelta(days=1),
         "Status": "Arrived", "Door_Delivery": "No"},
        # Stage 7 Done — delivered
        {"CRM_ID": "TRANANH", "Routing": "HPH-USLGB", "Bkg_No": "TABKG006",
         "Carrier": "ONE", "Contract_Type": "SHA0005N25",
         "Container_Type": "40RF", "Quantity": 1,
         "ETD": now - timedelta(days=30),
         "ATA": now - timedelta(days=10),
         "Status": "Delivered", "Door_Delivery": "Yes"},
        # Release-email-sent case for release_alerts tests
        {"CRM_ID": "OCEANSEA", "Routing": "HCM-USLAX", "Bkg_No": "OSBKG007",
         "Carrier": "CMA", "Contract_Type": "CMA-2026",
         "Container_Type": "40HQ", "Quantity": 1,
         "ETD": now - timedelta(days=25),
         "ETA": now + timedelta(days=2),
         "Status": "Arrived", "Door_Delivery": "No",
         "RELEASE_EMAIL_SENT": now - timedelta(hours=3)},
    ]

    for i, row_data in enumerate(seeds):
        r = AJ_DATA_START + i
        for field_name, val in row_data.items():
            col_idx = COL.get(field_name)
            if col_idx is None:
                raise KeyError(f"Field {field_name!r} not in active_jobs_cols.COL dict")
            ws.cell(r, col_idx, val)

    wb.save(str(erp_path))
    wb.close()


@pytest.fixture(scope="function")
def seeded_erp(tmp_path):
    """Copy + seed 7 Active Jobs rows covering stages 1-7 plus release-alert case."""
    if not MASTER_XLSM.exists():
        pytest.skip(f"Live ERP not found: {MASTER_XLSM}")
    dst = tmp_path / "ERP_Master_seeded.xlsm"
    shutil.copy2(MASTER_XLSM, dst)
    _seed_active_jobs_rows(dst)
    yield dst


@pytest.fixture(scope="session")
def sample_rules():
    """Load booking_rules.json as dict."""
    with open(str(RULES_FILE), "r", encoding="utf-8") as f:
        return json.load(f)


# ── xlwings fixtures (integration only — skipped when xlwings missing) ──────

@pytest.fixture(scope="session")
def excel_app():
    """Start one headless Excel instance for the whole test session."""
    try:
        import xlwings as xw
    except ImportError:
        pytest.skip("xlwings not installed — skip COM integration tests")
        return
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    try:
        yield app
    finally:
        try:
            for wb in list(app.books):
                try:
                    wb.close()
                except Exception:
                    pass
        except Exception:
            pass
        gc.collect()
        time.sleep(0.2)
        try:
            app.quit()
        except Exception:
            pass
        gc.collect()


@pytest.fixture(scope="function")
def erp_workbook(excel_app, tmp_path):
    """Open a fresh copy of ERP_Master_v14.xlsm per test via xlwings."""
    if not MASTER_XLSM.exists():
        pytest.skip(f"Master xlsm not found: {MASTER_XLSM}")
    dst = tmp_path / "ERP_Master_v14_test.xlsm"
    shutil.copy2(MASTER_XLSM, dst)
    wb = excel_app.books.open(str(dst), update_links=False, read_only=False)
    try:
        yield wb
    finally:
        try:
            wb.close()
        except Exception:
            pass
