"""test_price_watch.py — Price Watch alert engine tests.

Uses in-memory openpyxl workbooks to avoid touching live ERP.
Verifies DROP/RISE alert logic, threshold, and place-matching regression.
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "ERP" / "intelligence"))
from price_watch import (
    Alert,
    Q_COL,
    compute_alerts,
    iter_pending_quotes,
    load_latest_pricing,
)


# ---------------------------------------------------------------------------
# Helpers — build in-memory workbooks
# ---------------------------------------------------------------------------

def _make_quotes_wb(quote_rows: list[dict]) -> openpyxl.Workbook:
    """Build an in-memory wb with a Quotes sheet populated from quote_rows dicts."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotes"

    # Write headers at row 1
    for key, col in Q_COL.items():
        ws.cell(1, col, key)

    for i, row in enumerate(quote_rows, start=2):
        for key, col in Q_COL.items():
            ws.cell(i, col, row.get(key))

    return wb


def _make_pricing_dict(
    pol: str, pod: str, place: str, carrier: str,
    cont: str, buy: float,
    eff: datetime | None = None,
) -> dict:
    """Build a pricing_latest dict with one entry."""
    key = (pol.upper(), pod.upper(), place.upper(), carrier.upper(), cont)
    return {key: (buy, eff or datetime(2026, 3, 1), "test_source")}


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestComputeAlertsDropRise:

    def _base_quote(self, **overrides) -> dict:
        q = {
            "QuoteID": "Q001", "Date": datetime(2026, 3, 1),
            "Customer": "NAFOODS", "Carrier": "ONE",
            "POL": "HPH", "POD": "USLGB", "Place": "LOS ANGELES",
            "Via": "", "Eff": datetime(2026, 3, 1), "Exp": datetime(2026, 9, 1),
            "Source": "email", "Buy_40HC": 1000.0, "Status": "",
            "Remark": "", "StatusDate": None, "Qty": 1, "Volume": "1X40HQ",
            "JobID": None, "ContType": "40HQ",
            "_status": "PENDING", "_row": 2,
        }
        # Blank all other buy cols
        for key in Q_COL:
            if key.startswith("Buy_") and key not in q:
                q[key] = None
        q.update(overrides)
        return q

    def test_drop_alert_fires_when_price_falls_below_threshold(self):
        """DROP alert when current_buy < quoted_buy - 50.

        Note: Buy_40HC in quotes maps to price_cont='40HQ' via CONT_TO_PRICE_COL.
        The pricing_latest key must use '40HQ' to match the lookup.
        """
        quote = self._base_quote(Buy_40HC=1000.0)
        # price_cont for 40HC → '40HQ' per CONT_TO_PRICE_COL
        pricing = _make_pricing_dict("HPH", "USLGB", "LOS ANGELES", "ONE", "40HQ", 900.0)
        # delta = 900 - 1000 = -100 (abs > 50) → DROP
        quotes = [(2, quote)]
        alerts = compute_alerts(quotes, pricing, threshold=50.0)
        drops = [a for a in alerts if a.kind == "DROP"]
        assert len(drops) >= 1
        assert drops[0].delta == pytest.approx(-100.0)
        assert drops[0].priority == "P1"

    def test_rise_alert_fires_when_price_increases(self):
        """RISE alert when current_buy > quoted_buy + 50."""
        quote = self._base_quote(Buy_40HC=1000.0, Status="WIN", **{"_status": "WIN"})
        pricing = _make_pricing_dict("HPH", "USLGB", "LOS ANGELES", "ONE", "40HQ", 1100.0)
        quotes = [(2, quote)]
        alerts = compute_alerts(quotes, pricing, threshold=50.0)
        rises = [a for a in alerts if a.kind == "RISE"]
        assert len(rises) >= 1
        assert rises[0].delta == pytest.approx(100.0)

    def test_no_alert_when_delta_below_threshold(self):
        """No alert when |delta| < threshold."""
        quote = self._base_quote(Buy_40HC=1000.0)
        # current = 1030 → delta = +30, below 50 threshold
        pricing = _make_pricing_dict("HPH", "USLGB", "LOS ANGELES", "ONE", "40HC", 1030.0)
        quotes = [(2, quote)]
        alerts = compute_alerts(quotes, pricing, threshold=50.0)
        assert len(alerts) == 0

    def test_exactly_at_threshold_no_alert(self):
        """Delta exactly == threshold should NOT trigger (uses abs(delta) < threshold check)."""
        quote = self._base_quote(Buy_40HC=1000.0)
        pricing = _make_pricing_dict("HPH", "USLGB", "LOS ANGELES", "ONE", "40HC", 1050.0)
        quotes = [(2, quote)]
        alerts = compute_alerts(quotes, pricing, threshold=50.0)
        assert len(alerts) == 0

    def test_drop_alert_route_populated(self):
        """Alert.route must be 'POL-POD' string."""
        quote = self._base_quote(Buy_40HC=1000.0)
        pricing = _make_pricing_dict("HPH", "USLGB", "LOS ANGELES", "ONE", "40HQ", 800.0)
        alerts = compute_alerts([(2, quote)], pricing, threshold=50.0)
        assert len(alerts) >= 1, f"No alerts produced — check pricing key mapping"
        assert alerts[0].route == "HPH-USLGB"

    def test_customer_in_alert(self):
        quote = self._base_quote(Buy_40HC=1000.0)
        pricing = _make_pricing_dict("HPH", "USLGB", "LOS ANGELES", "ONE", "40HQ", 800.0)
        alerts = compute_alerts([(2, quote)], pricing, threshold=50.0)
        assert len(alerts) >= 1, f"No alerts produced — check pricing key mapping"
        assert alerts[0].customer == "NAFOODS"

    def test_no_buy_rate_skipped(self):
        """Quote with Buy_40HC=None should produce no alert."""
        quote = self._base_quote(Buy_40HC=None)
        pricing = _make_pricing_dict("HPH", "USLGB", "LOS ANGELES", "ONE", "40HC", 800.0)
        alerts = compute_alerts([(2, quote)], pricing, threshold=50.0)
        # No cont has a non-null quoted buy → no match
        assert len(alerts) == 0


class TestPlaceMatchingRegression:
    """Regression: inland pricing row must NOT match when quote.Place == POD."""

    def _base_quote(self, place: str, **overrides) -> dict:
        q = {
            "QuoteID": "Q002", "Date": datetime(2026, 3, 1),
            "Customer": "SIRI", "Carrier": "CMA",
            "POL": "HCM", "POD": "USLAX", "Place": place,
            "Via": "", "Eff": datetime(2026, 3, 1), "Exp": datetime(2026, 9, 1),
            "Source": "email", "Buy_40HC": 1000.0, "Status": "",
            "Remark": "", "StatusDate": None, "Qty": 1, "Volume": "1X40HQ",
            "JobID": None, "ContType": "40HQ",
            "_status": "PENDING", "_row": 3,
        }
        for key in Q_COL:
            if key.startswith("Buy_") and key not in q:
                q[key] = None
        q.update(overrides)
        return q

    def test_inland_pricing_not_picked_when_quote_place_is_pod(self):
        """
        Regression test: if quote.Place='CHICAGO' and pricing row Place='CHICAGO'
        with current_buy 800, alert fires. But if quote.Place=POD='USLAX' and
        pricing Place='CHICAGO' (inland), it must NOT match (wrong place).
        """
        # Quote Place = POD (no inland delivery for this quote)
        quote_pod_place = self._base_quote(place="USLAX")  # Place == POD
        # Pricing has an inland CHICAGO entry for CMA
        inland_pricing = _make_pricing_dict("HCM", "USLAX", "CHICAGO", "CMA", "40HC", 800.0)

        alerts = compute_alerts([(3, quote_pod_place)], inland_pricing, threshold=50.0)
        # Must NOT fire: quote.Place=USLAX but inland pricing key place=CHICAGO
        assert len(alerts) == 0, (
            "Bug: inland pricing matched quote where place == POD. "
            "This is the place-matching regression."
        )

    def test_inland_pricing_matches_inland_quote(self):
        """Positive control: inland pricing DOES match when quote.Place=inland city.

        Buy_40HC in quote → price_cont='40HQ'. Pricing key must use '40HQ'.
        """
        quote_inland = self._base_quote(place="CHICAGO")
        # price_cont for 40HC → '40HQ' per CONT_TO_PRICE_COL
        inland_pricing = _make_pricing_dict("HCM", "USLAX", "CHICAGO", "CMA", "40HQ", 800.0)
        alerts = compute_alerts([(3, quote_inland)], inland_pricing, threshold=50.0)
        drops = [a for a in alerts if a.kind == "DROP"]
        assert len(drops) >= 1, "Expected DROP alert for inland quote with matching inland pricing"


class TestLoadLatestPricing:

    def test_load_latest_pricing_from_wb(self):
        """load_latest_pricing picks up rows from Pricing Dry sheet."""
        wb = openpyxl.Workbook()
        ws_dry = wb.create_sheet("Pricing Dry")
        # Headers at row 1
        ws_dry.cell(1, 1, "POL")
        ws_dry.cell(1, 2, "POD")
        # Data row
        ws_dry.cell(2, 1, "HPH")
        ws_dry.cell(2, 2, "USLGB")
        ws_dry.cell(2, 3, "LOS ANGELES")  # Place
        ws_dry.cell(2, 4, "ONE")  # Carrier
        ws_dry.cell(2, 6, datetime(2026, 3, 1))  # Eff
        ws_dry.cell(2, 9, "test")  # Source
        ws_dry.cell(2, 10, 1000.0)  # 20GP
        ws_dry.cell(2, 11, 1200.0)  # 40GP
        ws_dry.cell(2, 12, 1100.0)  # 40HC

        pricing = load_latest_pricing(wb)
        assert len(pricing) > 0

        key = ("HPH", "USLGB", "LOS ANGELES", "ONE", "40HC")
        assert key in pricing
        assert pricing[key][0] == 1100.0

    def test_missing_pricing_sheet_returns_empty(self):
        wb = openpyxl.Workbook()
        # No "Pricing Dry" or "Pricing Reefer" sheets
        pricing = load_latest_pricing(wb)
        assert len(pricing) == 0
