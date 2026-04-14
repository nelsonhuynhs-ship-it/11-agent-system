"""test_ribbon_guard.py — ribbon preservation regression tests.

Documents the regression: plain openpyxl wb.save() strips customUI from xlsm.
Verifies that save_preserving_ribbon() keeps customUI14.xml + _rels/.rels entry.
"""
from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

CUSTOMUI_INNER = "customUI/customUI14.xml"
RELS_INNER = "_rels/.rels"
REL_TARGET = "customUI/customUI14.xml"

_ONEDRIVE_XML = Path("D:/OneDrive/NelsonData/erp/CustomUI_v14.xml")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _has_ribbon(xlsm_path: Path) -> tuple[bool, bool]:
    """Return (has_customui_xml, has_rels_entry) for an xlsm file."""
    with zipfile.ZipFile(str(xlsm_path), "r") as z:
        names = set(z.namelist())
        has_xml = CUSTOMUI_INNER in names
        has_rel = False
        if RELS_INNER in names:
            rels_text = z.read(RELS_INNER).decode("utf-8")
            has_rel = REL_TARGET in rels_text
    return has_xml, has_rel


def _minimal_xlsm(out_path: Path) -> None:
    """Write a minimal xlsm that openpyxl can open (no ribbon — baseline)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "test"
    # Save as .xlsx first, rename — openpyxl doesn't truly create xlsm with ribbon
    wb.save(str(out_path))
    wb.close()


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestRibbonGuardWithLiveERP:
    """Tests that require the live ERP copy (erp_copy fixture)."""

    def test_plain_save_ribbon_behavior_documented(self, erp_copy: Path):
        """Document the openpyxl ribbon behavior for this environment.

        Historic bug (openpyxl < 3.1): plain wb.save() stripped customUI14.xml.
        openpyxl 3.1.5+ preserves the ribbon on save.
        save_preserving_ribbon() guards against older versions AND re-injects
        the canonical CustomUI_v14.xml content if it diverges.

        This test verifies the file remains valid after plain save, regardless
        of which behavior the installed openpyxl version exhibits.
        """
        has_before, _ = _has_ribbon(erp_copy)
        if not has_before:
            pytest.skip("Live ERP copy has no ribbon")

        wb = openpyxl.load_workbook(str(erp_copy), keep_vba=True)
        wb.save(str(erp_copy))
        wb.close()

        # File must be a valid zip either way
        assert zipfile.is_zipfile(str(erp_copy))
        # Document current behavior (ribbon preserved in openpyxl 3.1.5+)
        has_after, _ = _has_ribbon(erp_copy)
        import openpyxl as _opx
        ver = tuple(int(x) for x in _opx.__version__.split(".")[:2])
        if ver >= (3, 1):
            # 3.1+ preserves ribbon — assert it is still there
            assert has_after, (
                f"openpyxl {_opx.__version__} stripped ribbon on plain save "
                "(was expected to preserve it in 3.1+)"
            )
        # If older version strips: save_preserving_ribbon() is the guard (tested separately)

    def test_save_preserving_ribbon_keeps_customui(self, erp_copy: Path):
        """save_preserving_ribbon() must preserve customUI14.xml after save."""
        if not _ONEDRIVE_XML.exists():
            pytest.skip("CustomUI_v14.xml not found on OneDrive — skip ribbon inject test")

        import sys, os
        sys.path.insert(0, str(Path(__file__).parent.parent / "ERP" / "core"))
        from ribbon_guard import save_preserving_ribbon

        wb = openpyxl.load_workbook(str(erp_copy), keep_vba=True)
        wb["Active Jobs"]["A1"].value = "test_marker"  # dirty write
        result = save_preserving_ribbon(wb, str(erp_copy), xml_path=str(_ONEDRIVE_XML))
        wb.close()

        has_xml, has_rel = _has_ribbon(erp_copy)
        assert has_xml, f"customUI14.xml missing after save_preserving_ribbon. result={result}"
        assert has_rel, f"_rels/.rels missing customUI entry. result={result}"

    def test_idempotent_double_inject(self, erp_copy: Path):
        """Calling save_preserving_ribbon twice must not corrupt the file."""
        if not _ONEDRIVE_XML.exists():
            pytest.skip("CustomUI_v14.xml not found — skip")

        import sys
        sys.path.insert(0, str(Path(__file__).parent.parent / "ERP" / "core"))
        from ribbon_guard import save_preserving_ribbon

        wb = openpyxl.load_workbook(str(erp_copy), keep_vba=True)
        save_preserving_ribbon(wb, str(erp_copy), xml_path=str(_ONEDRIVE_XML))
        wb.close()

        # Second inject
        wb2 = openpyxl.load_workbook(str(erp_copy), keep_vba=True)
        result2 = save_preserving_ribbon(wb2, str(erp_copy), xml_path=str(_ONEDRIVE_XML))
        wb2.close()

        has_xml, has_rel = _has_ribbon(erp_copy)
        assert has_xml
        assert has_rel
        # File must still be a valid zip
        assert zipfile.is_zipfile(str(erp_copy))


class TestRibbonGuardFallback:
    """Tests for graceful fallback when resources are missing."""

    def test_missing_xml_returns_skipped(self, erp_copy: Path, tmp_path: Path):
        """save_preserving_ribbon with non-existent xml_path → skipped, no crash."""
        import sys
        sys.path.insert(0, str(Path(__file__).parent.parent / "ERP" / "core"))
        from ribbon_guard import save_preserving_ribbon

        fake_xml = str(tmp_path / "no_such_file.xml")
        wb = openpyxl.load_workbook(str(erp_copy), keep_vba=True)
        result = save_preserving_ribbon(wb, str(erp_copy), xml_path=fake_xml)
        wb.close()

        # Must not raise; result must indicate skipped/error
        assert isinstance(result, dict)
        assert "skipped" in result or "error" in result, (
            f"Expected skipped/error when XML missing, got {result}"
        )
        # File must still be valid (wb.save succeeded even if ribbon inject failed)
        assert zipfile.is_zipfile(str(erp_copy))

    def test_reinject_without_customui_utils(self, tmp_path: Path):
        """reinject_ribbon when customui_utils.py unreachable → skipped dict, no crash."""
        import sys
        # Evict any stub (e.g. F8's yml_email_scan test stubs ribbon_guard)
        # so we import the real module
        sys.modules.pop("ribbon_guard", None)
        sys.path.insert(0, str(Path(__file__).parent.parent / "ERP" / "core"))
        from ribbon_guard import reinject_ribbon

        # Create a minimal xlsx (no ribbon) and try to reinject with bad path
        fake_xlsm = tmp_path / "fake.xlsm"
        wb = openpyxl.Workbook()
        wb.save(str(fake_xlsm))
        wb.close()

        result = reinject_ribbon(str(fake_xlsm), xml_path=str(tmp_path / "nonexistent.xml"))
        assert isinstance(result, dict)
        assert "skipped" in result or "error" in result
