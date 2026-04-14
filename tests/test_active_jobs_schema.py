"""test_active_jobs_schema.py — Active Jobs v4 schema migration tests.

Verifies:
  - migrate() adds cols 31-36 with correct headers
  - migrate() is idempotent (second run = already up-to-date, no changes)
  - Ribbon preserved after migration
"""
from __future__ import annotations

import sys
import zipfile
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "ERP" / "core"))
from active_jobs_schema import HEADER_ROW, NEW_COLUMNS, _col_letter, migrate

CUSTOMUI_INNER = "customUI/customUI14.xml"
AJ_SHEET_KEYWORD = "Active"


def _get_active_sheet(path: Path):
    wb = openpyxl.load_workbook(str(path), keep_vba=True, read_only=True)
    sheet = next((s for s in wb.sheetnames if AJ_SHEET_KEYWORD in s), None)
    ws = wb[sheet] if sheet else None
    return wb, ws, sheet


class TestMigrateHeaders:

    def test_adds_all_six_columns(self, erp_copy: Path):
        """migrate() must write all 6 new header names at row 7, cols 31-36."""
        ret = migrate(str(erp_copy))
        assert ret == 0, f"migrate() returned non-zero: {ret}"

        wb, ws, sheet = _get_active_sheet(erp_copy)
        assert sheet is not None, "Active Jobs sheet not found after migration"

        expected = {col: name for col, name, _ in NEW_COLUMNS}
        for col_idx, expected_name in expected.items():
            actual = ws.cell(HEADER_ROW, col_idx).value
            assert actual == expected_name, (
                f"Col {_col_letter(col_idx)} ({col_idx}): "
                f"expected '{expected_name}', got '{actual}'"
            )
        wb.close()

    def test_correct_column_letters(self):
        """_col_letter helper maps correctly for cols 31-36."""
        expected = {31: "AE", 32: "AF", 33: "AG", 34: "AH", 35: "AI", 36: "AJ"}
        for n, letter in expected.items():
            assert _col_letter(n) == letter, f"col {n} → expected {letter}, got {_col_letter(n)}"

    def test_column_order(self, erp_copy: Path):
        """Column names must appear in exact order SERVICE→PRICE_WATCH_DELTA."""
        migrate(str(erp_copy))
        wb, ws, _ = _get_active_sheet(erp_copy)
        names_in_order = [ws.cell(HEADER_ROW, c).value for c, _, _ in NEW_COLUMNS]
        expected_names = [name for _, name, _ in NEW_COLUMNS]
        assert names_in_order == expected_names
        wb.close()


class TestMigrateIdempotency:

    def test_second_run_no_changes(self, erp_copy: Path, capsys):
        """Second migrate() call must report 'already up-to-date' and return 0."""
        migrate(str(erp_copy))  # first run
        ret2 = migrate(str(erp_copy))  # second run
        captured = capsys.readouterr()
        assert ret2 == 0
        assert "up-to-date" in captured.out.lower() or "skip" in captured.out.lower(), (
            f"Expected idempotency message. Got:\n{captured.out}"
        )

    def test_second_run_headers_unchanged(self, erp_copy: Path):
        """Headers must be identical after two runs."""
        migrate(str(erp_copy))
        wb1, ws1, _ = _get_active_sheet(erp_copy)
        vals1 = [ws1.cell(HEADER_ROW, c).value for c, _, _ in NEW_COLUMNS]
        wb1.close()

        migrate(str(erp_copy))
        wb2, ws2, _ = _get_active_sheet(erp_copy)
        vals2 = [ws2.cell(HEADER_ROW, c).value for c, _, _ in NEW_COLUMNS]
        wb2.close()

        assert vals1 == vals2


class TestMigrateRibbonPreservation:

    def test_ribbon_present_after_migration(self, erp_copy: Path):
        """If live ERP had ribbon, migrate() must not destroy it."""
        # Check pre-condition
        with zipfile.ZipFile(str(erp_copy), "r") as z:
            had_ribbon = CUSTOMUI_INNER in z.namelist()

        if not had_ribbon:
            pytest.skip("Live ERP copy has no ribbon — nothing to verify")

        migrate(str(erp_copy))

        with zipfile.ZipFile(str(erp_copy), "r") as z:
            has_ribbon_after = CUSTOMUI_INNER in z.namelist()

        assert has_ribbon_after, "Ribbon was stripped by migrate() — critical regression"


class TestMigrateEdgeCases:

    def test_file_not_found_returns_error_code(self, tmp_path: Path):
        """migrate() on non-existent file must return 1."""
        ret = migrate(str(tmp_path / "nonexistent.xlsm"))
        assert ret == 1

    def test_existing_file_keeps_data_rows(self, seeded_erp: Path):
        """migrate() must not wipe existing data rows when adding headers."""
        # Count non-empty rows before
        wb_before, ws_before, _ = _get_active_sheet(seeded_erp)
        rows_before = sum(
            1 for r in range(8, ws_before.max_row + 1)
            if ws_before.cell(r, 1).value
        )
        wb_before.close()

        migrate(str(seeded_erp))

        wb_after, ws_after, _ = _get_active_sheet(seeded_erp)
        rows_after = sum(
            1 for r in range(8, ws_after.max_row + 1)
            if ws_after.cell(r, 1).value
        )
        wb_after.close()

        assert rows_after >= rows_before, (
            f"migrate() lost data rows: before={rows_before} after={rows_after}"
        )
