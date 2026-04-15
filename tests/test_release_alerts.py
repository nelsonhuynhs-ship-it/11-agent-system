"""test_release_alerts.py — Release alert Feature 3 tests.

Ports classify() unit tests + integration tests on seeded_erp.
"""
from __future__ import annotations

import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent / "ERP" / "core"))
from active_jobs_cols import COL as _COL

sys.path.insert(0, str(Path(__file__).parent.parent / "ERP" / "jobs"))
from release_alerts import (
    _format_countdown,
    classify,
    scan_alerts,
)

AJ_SHEET_KEYWORD = "Active"
AJ_DATA_START = 8
COL_RELEASE_EMAIL_SENT = _COL["RELEASE_EMAIL_SENT"]   # 37
COL_RELEASE_CONFIRMED = _COL["RELEASE_CONFIRMED"]      # 38
COL_ETA = _COL["ETA"]                                  # 21


# ---------------------------------------------------------------------------
# classify() unit tests
# ---------------------------------------------------------------------------

class TestClassify:

    def test_urgent_when_elapsed_exceeds_deadline(self):
        priority, remaining = classify(elapsed_hours=3.0, deadline_hours=2.0)
        assert priority == "URGENT"
        assert remaining.total_seconds() < 0  # negative = breached

    def test_warn_at_half_deadline(self):
        priority, _ = classify(elapsed_hours=1.0, deadline_hours=2.0)
        assert priority == "WARN"

    def test_info_below_half_deadline(self):
        priority, _ = classify(elapsed_hours=0.5, deadline_hours=2.0)
        assert priority == "INFO"

    def test_exactly_at_deadline_is_urgent(self):
        priority, remaining = classify(elapsed_hours=2.0, deadline_hours=2.0)
        assert priority == "URGENT"

    def test_countdown_positive_for_info(self):
        _, remaining = classify(elapsed_hours=0.5, deadline_hours=2.0)
        assert remaining.total_seconds() > 0

    def test_countdown_negative_for_urgent(self):
        _, remaining = classify(elapsed_hours=4.0, deadline_hours=2.0)
        assert remaining.total_seconds() < 0

    def test_zero_elapsed_is_info(self):
        priority, _ = classify(elapsed_hours=0.0, deadline_hours=2.0)
        assert priority == "INFO"

    def test_custom_deadline(self):
        priority, _ = classify(elapsed_hours=3.0, deadline_hours=6.0)
        assert priority == "WARN"  # 3/6 = 50% exactly → WARN boundary


# ---------------------------------------------------------------------------
# _format_countdown unit tests
# ---------------------------------------------------------------------------

class TestFormatCountdown:

    def test_positive_countdown(self):
        td = timedelta(hours=1, minutes=30)
        result = _format_countdown(td, "WARN")
        assert result.startswith("+")
        assert "01:30" in result

    def test_negative_countdown_breached(self):
        td = timedelta(hours=-2, minutes=-15)
        result = _format_countdown(td, "URGENT")
        assert "BREACHED" in result

    def test_info_no_breached(self):
        td = timedelta(hours=1)
        result = _format_countdown(td, "INFO")
        assert "BREACHED" not in result


# ---------------------------------------------------------------------------
# Integration: scan_alerts on seeded_erp
# ---------------------------------------------------------------------------

def _seed_release_scenario(erp_path: Path, sent_hours_ago: float,
                            confirmed: bool, eta_days_away: float) -> None:
    """Add/overwrite the OCEANSEA row (seed index 6 → row 14) with release scenario."""
    wb = openpyxl.load_workbook(str(erp_path), keep_vba=True)
    sheet = next(s for s in wb.sheetnames if AJ_SHEET_KEYWORD in s)
    ws = wb[sheet]

    now = datetime(2026, 4, 14, 12, 0, 0)
    eta = now + timedelta(days=eta_days_away)
    sent_at = now - timedelta(hours=sent_hours_ago)

    # Overwrite seed row 7 (row 14) — OCEANSEA
    r = AJ_DATA_START + 6
    ws.cell(r, _COL["CRM_ID"], "OCEANSEA_TEST")
    ws.cell(r, _COL["Routing"], "HCM-USLAX")
    ws.cell(r, COL_ETA, eta)
    ws.cell(r, COL_RELEASE_EMAIL_SENT, sent_at)
    ws.cell(r, COL_RELEASE_CONFIRMED, confirmed if confirmed else None)

    wb.save(str(erp_path))
    wb.close()


class TestScanAlerts:

    def test_urgent_alert_when_past_deadline_within_eta_window(self, seeded_erp: Path):
        """Job with release email 3h ago (> 2h deadline) + ETA in 2 days → URGENT."""
        _seed_release_scenario(seeded_erp, sent_hours_ago=3.0,
                               confirmed=False, eta_days_away=2.0)
        now = datetime(2026, 4, 14, 12, 0, 0)
        alerts = scan_alerts(str(seeded_erp), hours=2.0, eta_window_days=3, now=now)
        urgent = [a for a in alerts if a.priority == "URGENT"
                  and a.crm_id == "OCEANSEA_TEST"]
        assert len(urgent) >= 1, (
            "Expected URGENT alert for OCEANSEA_TEST (3h elapsed, 2h deadline, ETA in 2 days)"
        )

    def test_no_alert_when_release_confirmed(self, seeded_erp: Path):
        """If RELEASE_CONFIRMED is set, scan_alerts must not fire an alert."""
        _seed_release_scenario(seeded_erp, sent_hours_ago=3.0,
                               confirmed=True, eta_days_away=2.0)
        now = datetime(2026, 4, 14, 12, 0, 0)
        alerts = scan_alerts(str(seeded_erp), hours=2.0, eta_window_days=3, now=now)
        confirmed_alerts = [a for a in alerts if a.crm_id == "OCEANSEA_TEST"]
        assert len(confirmed_alerts) == 0, (
            "Alert fired even though RELEASE_CONFIRMED is set"
        )

    def test_no_alert_when_eta_outside_window(self, seeded_erp: Path):
        """If ETA is 30 days away (outside 3-day window), no alert."""
        _seed_release_scenario(seeded_erp, sent_hours_ago=3.0,
                               confirmed=False, eta_days_away=30.0)
        now = datetime(2026, 4, 14, 12, 0, 0)
        alerts = scan_alerts(str(seeded_erp), hours=2.0, eta_window_days=3, now=now)
        outside_window = [a for a in alerts if a.crm_id == "OCEANSEA_TEST"]
        assert len(outside_window) == 0, (
            "Alert fired for ETA 30 days away (outside 3-day window)"
        )

    def test_warn_alert_at_half_deadline(self, seeded_erp: Path):
        """Job with email sent 1h ago (50% of 2h deadline) + ETA in 2 days → WARN."""
        _seed_release_scenario(seeded_erp, sent_hours_ago=1.0,
                               confirmed=False, eta_days_away=2.0)
        now = datetime(2026, 4, 14, 12, 0, 0)
        alerts = scan_alerts(str(seeded_erp), hours=2.0, eta_window_days=3, now=now)
        warn_alerts = [a for a in alerts if a.crm_id == "OCEANSEA_TEST"
                       and a.priority == "WARN"]
        assert len(warn_alerts) >= 1

    def test_no_alert_when_release_email_not_sent(self, seeded_erp: Path):
        """Row with no RELEASE_EMAIL_SENT → never included in alerts."""
        # Ensure a row with no sent timestamp
        wb = openpyxl.load_workbook(str(seeded_erp), keep_vba=True)
        sheet = next(s for s in wb.sheetnames if AJ_SHEET_KEYWORD in s)
        ws = wb[sheet]
        # Row 8 = NAFOODS — no release email sent in seed
        ws.cell(AJ_DATA_START, COL_RELEASE_EMAIL_SENT, None)
        wb.save(str(seeded_erp))
        wb.close()

        now = datetime(2026, 4, 14, 12, 0, 0)
        alerts = scan_alerts(str(seeded_erp), hours=2.0, eta_window_days=3, now=now)
        nafoods_alerts = [a for a in alerts if a.crm_id == "NAFOODS"]
        assert len(nafoods_alerts) == 0, "Alert fired for row with no RELEASE_EMAIL_SENT"

    def test_elapsed_hours_correct(self, seeded_erp: Path):
        """elapsed_hours in alert must match actual time since email sent."""
        _seed_release_scenario(seeded_erp, sent_hours_ago=5.0,
                               confirmed=False, eta_days_away=2.0)
        now = datetime(2026, 4, 14, 12, 0, 0)
        alerts = scan_alerts(str(seeded_erp), hours=2.0, eta_window_days=3, now=now)
        ocean = [a for a in alerts if a.crm_id == "OCEANSEA_TEST"]
        assert len(ocean) >= 1
        assert abs(ocean[0].elapsed_hours - 5.0) < 0.1, (
            f"elapsed_hours should be ~5.0, got {ocean[0].elapsed_hours}"
        )
