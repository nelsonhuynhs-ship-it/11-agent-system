# -*- coding: utf-8 -*-
"""
settings_view.py — Settings & Configuration
=============================================
Edit markup, default POL, Parquet path, email templates, Telegram config.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import TYPE_CHECKING

import customtkinter as ctk

from gui.theme import COLORS, FONTS

if TYPE_CHECKING:
    from gui.app_window import AppWindow


class SettingsView(ctk.CTkFrame):

    def __init__(self, parent, app: AppWindow):
        super().__init__(parent, fg_color="transparent")
        self.app = app
        self._fields: dict[str, ctk.StringVar] = {}
        self._build_ui()
        self._load_settings()

    # ══════════════════════════════════════════════════════════
    # UI
    # ══════════════════════════════════════════════════════════

    def _build_ui(self):
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", pady=(0, 12))
        ctk.CTkLabel(hdr, text="Settings", font=FONTS["heading"],
                     text_color=COLORS["text_primary"], anchor="w").pack(side="left")
        ctk.CTkButton(hdr, text="💾 Save All", font=FONTS["button"], width=130,
                      fg_color=COLORS["accent"], hover_color="#c62828",
                      command=self._save_settings).pack(side="right")

        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True)

        # ── Rate Settings ─────────────────────────────────────
        self._section(scroll, "Rate Settings")
        self._field(scroll, "default_markup",    "Default Markup ($)",   "20")
        self._field(scroll, "default_pol",       "Default POL",          "HPH")
        self._field(scroll, "default_dests",     "Default Destinations (comma-sep)",
                    "USCHI,USLAX,USLGB,USEWR,USSAV,USNYC")

        # ── Parquet Source ────────────────────────────────────
        self._section(scroll, "Parquet Data Source")
        self._field(scroll, "parquet_path", "Parquet File Path",
                    str(self.app.project_root.parent / "Pricing_Engine" / "data" / "Cleaned_Master_History.parquet"))
        self._parquet_status = ctk.CTkLabel(scroll, text="", font=FONTS["small"],
                                             text_color=COLORS["text_muted"], anchor="w")
        self._parquet_status.pack(fill="x", padx=4, pady=(0, 8))

        # ── Email Template ────────────────────────────────────
        self._section(scroll, "Email Template (from config.xlsx)")
        self._field(scroll, "subject_templates",
                    "Subject Templates (pipe-separated)",
                    "What Importers Need to Know|Rate Update from Pudong Prime")
        self._field(scroll, "subject_suffix", "Subject Suffix", "NELSON")
        self._field_multiline(scroll, "intro_text", "Intro Text", height=80)
        self._field_multiline(scroll, "closing_text", "Closing Text", height=60)

        # ── Telegram ──────────────────────────────────────────
        self._section(scroll, "Telegram Alerts")
        self._field(scroll, "telegram_token",   "Bot Token (env: TELEGRAM_TOKEN)", "")
        self._field(scroll, "telegram_chat_id", "Chat ID (env: TELEGRAM_CHAT_ID)", "")

        # ── Paths info ────────────────────────────────────────
        self._section(scroll, "System Paths (read-only)")
        paths = {
            "Project Root":  str(self.app.project_root),
            "Data Dir":      str(self.app.data_dir),
            "Logs Dir":      str(self.app.logs_dir),
            "Config Dir":    str(self.app.config_dir),
        }
        for label, path in paths.items():
            row = ctk.CTkFrame(scroll, fg_color="transparent")
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=f"{label}:", font=FONTS["body_bold"],
                         text_color=COLORS["text_secondary"], width=140, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(row, text=path, font=FONTS["mono"],
                         text_color=COLORS["text_muted"], anchor="w").pack(side="left", padx=4)

    def _section(self, parent, title: str):
        sep = ctk.CTkFrame(parent, height=1, fg_color=COLORS["border"])
        sep.pack(fill="x", pady=(16, 8))
        ctk.CTkLabel(parent, text=title, font=FONTS["heading_small"],
                     text_color=COLORS["accent_blue"], anchor="w").pack(fill="x", padx=4, pady=(0, 6))

    def _field(self, parent, key: str, label: str, default: str = "") -> ctk.StringVar:
        row = ctk.CTkFrame(parent, fg_color="transparent")
        row.pack(fill="x", pady=3)
        ctk.CTkLabel(row, text=label, font=FONTS["body"],
                     text_color=COLORS["text_secondary"], width=240, anchor="w").pack(side="left", padx=4)
        var = ctk.StringVar(value=default)
        ctk.CTkEntry(row, textvariable=var, font=FONTS["body"],
                     fg_color=COLORS["bg_input"], width=400).pack(side="left", padx=4)
        self._fields[key] = var
        return var

    def _field_multiline(self, parent, key: str, label: str, height: int = 80):
        ctk.CTkLabel(parent, text=label, font=FONTS["body"],
                     text_color=COLORS["text_secondary"], anchor="w").pack(fill="x", padx=4, pady=(4, 2))
        tb = ctk.CTkTextbox(parent, font=FONTS["body"], height=height,
                            fg_color=COLORS["bg_input"], text_color=COLORS["text_primary"])
        tb.pack(fill="x", padx=4, pady=(0, 4))
        self._fields[key] = tb  # store textbox directly for multiline

    # ══════════════════════════════════════════════════════════
    # DATA
    # ══════════════════════════════════════════════════════════

    def _load_settings(self):
        """Load settings from config.xlsx and customer_rules.json."""
        # Parquet path status
        parquet_path = self.app.project_root.parent / "Pricing_Engine" / "data" / "Cleaned_Master_History.parquet"
        if parquet_path.exists():
            size_mb = parquet_path.stat().st_size / 1024 / 1024
            self._parquet_status.configure(
                text=f"✓ Found — {size_mb:.1f} MB",
                text_color=COLORS["accent_green"])
        else:
            self._parquet_status.configure(
                text="✗ Parquet file not found at default path — check path above",
                text_color=COLORS["danger"])

        # Load from config.xlsx
        try:
            import openpyxl
            cfg_file = self.app.project_root / "data" / "config.xlsx"
            if cfg_file.exists():
                wb = openpyxl.load_workbook(str(cfg_file), data_only=True)
                cfg = {}
                for r in wb.active.iter_rows(max_col=2, values_only=True):
                    k = str(r[0] or "").strip().upper()
                    v = str(r[1] or "").strip() if r[1] else ""
                    if k and k != "KEY":
                        cfg[k] = v

                key_map = {
                    "SUBJECTTEMPLATES": "subject_templates",
                    "SUBJECTSUFFIX":    "subject_suffix",
                    "INTROTEXT":        "intro_text",
                    "CLOSINGTEXT":      "closing_text",
                }
                for cfg_key, field_key in key_map.items():
                    val = cfg.get(cfg_key, "")
                    field = self._fields.get(field_key)
                    if field is None:
                        continue
                    if isinstance(field, ctk.CTkTextbox):
                        field.delete("1.0", "end")
                        field.insert("1.0", val)
                    elif isinstance(field, ctk.StringVar):
                        field.set(val)
        except Exception:
            pass

        # Telegram from env
        import os
        tok = os.environ.get("TELEGRAM_TOKEN", "")
        cid = os.environ.get("TELEGRAM_CHAT_ID", "")
        if tok and "telegram_token" in self._fields:
            self._fields["telegram_token"].set(tok)
        if cid and "telegram_chat_id" in self._fields:
            self._fields["telegram_chat_id"].set(cid)

    def _save_settings(self):
        """Save editable settings back to config.xlsx."""
        try:
            import openpyxl
            cfg_file = self.app.project_root / "data" / "config.xlsx"
            if not cfg_file.exists():
                self.app.set_status("config.xlsx not found — cannot save")
                return

            wb = openpyxl.load_workbook(str(cfg_file))
            ws = wb.active

            key_map = {
                "SUBJECTTEMPLATES": "subject_templates",
                "SUBJECTSUFFIX":    "subject_suffix",
                "INTROTEXT":        "intro_text",
                "CLOSINGTEXT":      "closing_text",
            }

            # Build lookup of row positions
            row_pos = {}
            for r_idx, row in enumerate(ws.iter_rows(max_col=1, values_only=False), 1):
                k = str(row[0].value or "").strip().upper()
                if k:
                    row_pos[k] = r_idx

            for cfg_key, field_key in key_map.items():
                field = self._fields.get(field_key)
                if field is None:
                    continue
                if isinstance(field, ctk.CTkTextbox):
                    val = field.get("1.0", "end").strip()
                else:
                    val = field.get().strip()

                if cfg_key in row_pos:
                    ws.cell(row=row_pos[cfg_key], column=2, value=val)

            wb.save(str(cfg_file))
            self.app.set_status_done("Settings saved to config.xlsx")

        except Exception as e:
            self.app.set_status(f"Save error: {e}")
