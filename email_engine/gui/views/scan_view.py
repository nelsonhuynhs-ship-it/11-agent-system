# -*- coding: utf-8 -*-
"""
scan_view.py — Scan & Classify View
=====================================
Trigger Outlook scan → detect bounce/auto-reply/human-reply →
clean bad emails → classify reply tiers → show results.
"""
from __future__ import annotations

import sys
import threading
from typing import TYPE_CHECKING

import customtkinter as ctk

from gui.theme import COLORS, FONTS

if TYPE_CHECKING:
    from gui.app_window import AppWindow


class ScanView(ctk.CTkFrame):

    def __init__(self, parent, app: AppWindow):
        super().__init__(parent, fg_color="transparent")
        self.app = app
        self._build_ui()
        self._refresh_stats()

    # ══════════════════════════════════════════════════════════
    # UI
    # ══════════════════════════════════════════════════════════

    def _build_ui(self):
        # Header
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", pady=(0, 12))
        ctk.CTkLabel(hdr, text="Scan & Classify", font=FONTS["heading"],
                     text_color=COLORS["text_primary"], anchor="w").pack(side="left")

        # Action buttons row
        actions = ctk.CTkFrame(self, fg_color=COLORS["bg_sidebar"], corner_radius=10)
        actions.pack(fill="x", pady=(0, 12))
        act_inner = ctk.CTkFrame(actions, fg_color="transparent")
        act_inner.pack(fill="x", padx=16, pady=14)

        self.scan_btn = ctk.CTkButton(
            act_inner, text="🔍 Scan Outlook", font=FONTS["button"],
            width=160, fg_color=COLORS["accent_blue"], hover_color=COLORS["bg_hover"],
            command=self._run_scan,
        )
        self.scan_btn.pack(side="left", padx=(0, 10))

        self.classify_btn = ctk.CTkButton(
            act_inner, text="📊 Classify Tiers", font=FONTS["button"],
            width=160, fg_color=COLORS["bg_card"], hover_color=COLORS["bg_hover"],
            command=self._run_classify,
        )
        self.classify_btn.pack(side="left", padx=(0, 10))

        self.clean_btn = ctk.CTkButton(
            act_inner, text="🧹 Clean Bad Emails", font=FONTS["button"],
            width=170, fg_color=COLORS["bg_card"], hover_color=COLORS["bg_hover"],
            command=self._run_clean,
        )
        self.clean_btn.pack(side="left", padx=(0, 10))

        self.full_btn = ctk.CTkButton(
            act_inner, text="⚡ Full Pipeline", font=FONTS["button"],
            width=150, fg_color=COLORS["accent"], hover_color="#c62828",
            command=self._run_full,
        )
        self.full_btn.pack(side="right")

        # Tier stats cards
        tiers_frame = ctk.CTkFrame(self, fg_color="transparent")
        tiers_frame.pack(fill="x", pady=(0, 12))

        self._tier_cards = {}
        tiers = [
            ("NO_REPLY",  "No Reply",    COLORS["text_muted"]),
            ("REPLY_1",   "Reply 1×",    COLORS["accent_blue"]),
            ("REPLY_2",   "Reply 2×",    COLORS["accent_yellow"]),
            ("REPLY_3",   "HOT Leads",   COLORS["accent"]),
            ("BOUNCED",   "Bounced",     COLORS["warning"]),
            ("AUTO_REPLY","Auto-Reply",  COLORS["text_secondary"]),
        ]
        for key, label, color in tiers:
            card = self._make_tier_card(tiers_frame, label, "—", color)
            card.pack(side="left", fill="both", expand=True, padx=(0, 8))
            self._tier_cards[key] = card

        # Two-column: log + last scan results
        body = ctk.CTkFrame(self, fg_color="transparent")
        body.pack(fill="both", expand=True)
        body.columnconfigure(0, weight=3)
        body.columnconfigure(1, weight=2)

        # Left: Live log
        log_frame = ctk.CTkFrame(body, fg_color=COLORS["bg_sidebar"], corner_radius=10)
        log_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        ctk.CTkLabel(log_frame, text="  Pipeline Log", font=FONTS["heading_small"],
                     text_color=COLORS["text_primary"], anchor="w").pack(fill="x", padx=12, pady=(10, 4))

        self.log_text = ctk.CTkTextbox(
            log_frame, font=FONTS["mono"], fg_color="transparent",
            text_color=COLORS["accent_green"], wrap="word",
        )
        self.log_text.pack(fill="both", expand=True, padx=12, pady=(0, 10))

        # Right: Scan results summary
        res_frame = ctk.CTkFrame(body, fg_color=COLORS["bg_sidebar"], corner_radius=10)
        res_frame.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        ctk.CTkLabel(res_frame, text="  Last Scan Summary", font=FONTS["heading_small"],
                     text_color=COLORS["text_primary"], anchor="w").pack(fill="x", padx=12, pady=(10, 4))

        self.result_text = ctk.CTkTextbox(
            res_frame, font=FONTS["small"], fg_color="transparent",
            text_color=COLORS["text_secondary"], wrap="word",
        )
        self.result_text.pack(fill="both", expand=True, padx=12, pady=(0, 10))
        self._set_result("No scan run yet.\nClick 'Scan Outlook' or 'Full Pipeline' to start.")

    def _make_tier_card(self, parent, label: str, value: str, color: str) -> ctk.CTkFrame:
        card = ctk.CTkFrame(parent, fg_color=COLORS["bg_sidebar"], corner_radius=10)
        val = ctk.CTkLabel(card, text=value, font=("Segoe UI", 26, "bold"), text_color=color)
        val.pack(pady=(14, 2))
        card._val = val
        ctk.CTkLabel(card, text=label, font=FONTS["stat_label"],
                     text_color=COLORS["text_muted"]).pack(pady=(0, 12))
        return card

    def _update_tier(self, key: str, value: str):
        c = self._tier_cards.get(key)
        if c and hasattr(c, "_val"):
            c._val.configure(text=str(value))

    # ══════════════════════════════════════════════════════════
    # LOG HELPERS
    # ══════════════════════════════════════════════════════════

    def _log(self, msg: str):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"{msg}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self.update_idletasks()

    def _clear_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    def _set_result(self, text: str):
        self.result_text.configure(state="normal")
        self.result_text.delete("1.0", "end")
        self.result_text.insert("1.0", text)
        self.result_text.configure(state="disabled")

    # ══════════════════════════════════════════════════════════
    # STATS REFRESH
    # ══════════════════════════════════════════════════════════

    def _refresh_stats(self):
        """Read tier counts from customer_final.xlsx."""
        try:
            import openpyxl
            final = self.app.data_dir / "customer_final.xlsx"
            if not final.exists():
                return
            wb = openpyxl.load_workbook(str(final), read_only=True)
            for key in ["NO_REPLY", "REPLY_1", "REPLY_2", "REPLY_3", "BOUNCED", "AUTO_REPLY"]:
                if key in wb.sheetnames:
                    ws = wb[key]
                    count = max(0, ws.max_row - 1)  # minus header
                    self._update_tier(key, str(count))
            wb.close()
        except Exception:
            pass

    # ══════════════════════════════════════════════════════════
    # PIPELINE RUNNERS
    # ══════════════════════════════════════════════════════════

    def _set_buttons_busy(self, busy: bool):
        state = "disabled" if busy else "normal"
        for btn in [self.scan_btn, self.classify_btn, self.clean_btn, self.full_btn]:
            btn.configure(state=state)

    def _run_scan(self):
        self._clear_log()
        self._set_buttons_busy(True)
        self.app.set_status_busy("Scanning Outlook...")
        threading.Thread(target=self._worker_scan, daemon=True).start()

    def _run_classify(self):
        self._clear_log()
        self._set_buttons_busy(True)
        self.app.set_status_busy("Classifying reply tiers...")
        threading.Thread(target=self._worker_classify, daemon=True).start()

    def _run_clean(self):
        self._clear_log()
        self._set_buttons_busy(True)
        self.app.set_status_busy("Cleaning bad emails...")
        threading.Thread(target=self._worker_clean, daemon=True).start()

    def _run_full(self):
        """Full pipeline: clean → scan → classify."""
        self._clear_log()
        self._set_buttons_busy(True)
        self.app.set_status_busy("Running full pipeline...")
        threading.Thread(target=self._worker_full, daemon=True).start()

    # ── Workers ───────────────────────────────────────────────

    def _run_script(self, script_name: str, extra_args: list = None) -> bool:
        """Run a core or ingest script as subprocess, stream output to log."""
        import subprocess
        # Check core/ first, then ingest/
        core_dir = self.app.core_dir
        script_path = core_dir / script_name
        if not script_path.exists():
            script_path = self.app.project_root / "ingest" / script_name
        if not script_path.exists():
            self.after(0, lambda: self._log(f"[SKIP] {script_name} not found"))
            return False

        cmd = [sys.executable, str(script_path)] + (extra_args or [])
        self.after(0, lambda: self._log(f"▶ {script_name}"))

        proc = subprocess.Popen(
            cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, encoding="utf-8", errors="replace",
            cwd=str(self.app.project_root),
        )
        for line in proc.stdout:
            line = line.rstrip()
            if line:
                self.after(0, lambda l=line: self._log(f"  {l}"))
        proc.wait()

        ok = proc.returncode == 0
        status = "✓ Done" if ok else f"✗ Failed (code {proc.returncode})"
        self.after(0, lambda: self._log(f"  {status}\n"))
        return ok

    def _worker_scan(self):
        self._run_script("read_email1.py")
        self.after(0, self._after_scan)

    def _worker_classify(self):
        self._run_script("process_reply.py")
        self.after(0, self._after_classify)

    def _worker_clean(self):
        self._run_script("clean_data.py")
        self.after(0, lambda: self._finish("Clean complete"))

    def _worker_full(self):
        self.after(0, lambda: self._log("=== FULL PIPELINE ===\n"))
        self._run_script("clean_data.py")
        self._run_script("read_email1.py")
        self._run_script("process_reply.py")
        self._run_script("follow_up_engine.py")
        self.after(0, self._after_classify)

    def _after_scan(self):
        """After scan: show signal summary from email_knowledge.csv."""
        import csv as _csv
        knowledge = self.app.logs_dir / "email_knowledge.csv"
        counts = {}
        if knowledge.exists():
            with open(knowledge, encoding="utf-8", errors="replace") as f:
                for row in _csv.DictReader(f):
                    sig = row.get("signal_type", "unknown")
                    counts[sig] = counts.get(sig, 0) + 1

        lines = ["Scan complete. Signal breakdown:\n"]
        for sig, cnt in sorted(counts.items(), key=lambda x: -x[1]):
            icon = {"human_reply": "💬", "hard_bounce": "❌",
                    "soft_bounce": "⚠", "auto_reply": "🤖",
                    "policy_reject": "🚫"}.get(sig, "•")
            lines.append(f"  {icon} {sig}: {cnt}")
        self._set_result("\n".join(lines))
        self._finish("Scan complete")

    def _after_classify(self):
        """After classify: refresh tier cards."""
        self._refresh_stats()
        self._set_result("Classification complete.\nTier counts updated above.")
        self._finish("Classify complete")

    def _finish(self, msg: str):
        self._set_buttons_busy(False)
        self.app.set_status_done(msg)
