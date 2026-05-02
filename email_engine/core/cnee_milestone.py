"""
cnee_milestone.py — Auto CNEE notification (ATD + ETA-7)

Exported:
    on_atd_detected(mail_item, stages, identifiers, sender_smtp) -> bool
    run_eta_reminder() -> int   # CLI entry for Task Scheduler
    flush_telegram_summary()    # called at end of scanner run

Security controls:
    - Authentication-Results header check (SPF/DKIM/DMARC)
    - Explicit OPS sender allowlist (not domain-wide)
    - Placeholder sanitization (strip + length cap + regex whitelist)
    - Rate limiting (MAX_DRAFTS_PER_RUN, MAX_DRAFTS_PER_DAY)
    - Kill switch file
    - Bulk-Bkg detection (reject >3 Bkg in one mail)
    - Date sanity (ATD within ReceivedTime +/- 30d)
    - ActiveInspector check (don't steal Nelson's Outlook)

Active Jobs schema (post Phase 01 migration):
    Header row: 7   (rows 1-6 are title/visual chrome)
    C4=CUSTOMER, C8=Bkg_No, C9=HBL_NO, C13=ETD, C19=EMAIL,
    C21=ETA, C22=ATA, C41=ATD_DATE, C42=ETA_DATE,
    C43=NOTIFIED_ATD, C44=NOTIFIED_ETA7

CRM schema:
    Header row: 1
    C2=Customer_Name, C8=Contact1_Email, C11=Contact2_Email, C44=AUTO_NOTIFY
"""
from __future__ import annotations

import json
import logging
import os
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

# ── Config ──────────────────────────────────────────────────────────────────

# TODO(Nelson): populate after first run — check Telegram weekly digest for
# actual OPS sender SMTPs and add them here.
# Example: "ops@pudongprime.vn", "pricing@pudongprime.vn"
OPS_ALLOWLIST: set[str] = set()

BLACKLIST_REGEX = re.compile(
    r"VESSEL\s+CHANGE|RVS\s+ETD|REVISED\s+ETD|CHANGE\s+VESSEL|NEW\s+ETD",
    re.I,
)

MAX_DRAFTS_PER_RUN = 5
MAX_DRAFTS_PER_DAY = 20

# Files relative to email_engine/data/
_DATA_DIR = Path(__file__).parent.parent / "data"
KILL_SWITCH = _DATA_DIR / "AUTO_NOTIFY_DISABLED"
STATE_FILE = _DATA_DIR / "milestone_state.jsonl"
DAILY_COUNTER = _DATA_DIR / "milestone_daily.json"

# ERP path — try shared.paths first, fallback to hardcoded OneDrive
try:
    from shared.paths import ERP_DATA
    ERP_PATH = ERP_DATA / "ERP_Master_v14.xlsm"
except ImportError:
    ERP_PATH = Path(r"D:/OneDrive/NelsonData/erp/ERP_Master_v14.xlsm")

# Active Jobs layout (after Phase 01 migration)
AJ_HEADER_ROW = 7
AJ_COL = {
    "CUSTOMER": 4,
    "BKG": 8,
    "HBL": 9,
    "ETD": 13,
    "STATUS": 14,
    "EMAIL": 19,
    "ROUTING": 20,
    "ETA": 21,
    "ATA": 22,
    "ATD_DATE": 41,
    "ETA_DATE": 42,
    "NOTIFIED_ATD": 43,
    "NOTIFIED_ETA7": 44,
}

# CRM layout
CRM_HEADER_ROW = 1
CRM_COL = {
    "CUSTOMER_NAME": 2,
    "CONTACT1_EMAIL": 8,
    "CONTACT2_EMAIL": 11,
    "AUTO_NOTIFY": 44,
}

# Placeholder sanitization rules: field -> (max_length, optional_regex_whitelist)
FIELD_LIMITS: dict[str, tuple[int, Optional[str]]] = {
    "bkg":      (20, r"^[A-Z0-9]{5,20}$"),
    "hbl":      (20, r"^[A-Z0-9]{5,20}$"),
    "vessel":   (40, None),
    "carrier":  (20, None),
    "customer": (80, None),
    "pol":      (20, None),
    "pod":      (30, None),
    "etd":      (20, None),
    "eta":      (20, None),
}

# Date extraction patterns (ATD/Loaded)
# Covers:  "ATD: 20/04/2026"  |  "ATD// 20.04.2026"  |  "ATD__ 20-04-2026"
#          "vessel departed on 20/04/2026"  |  "loaded on board 20.04.2026"
_DATE_PATTERNS = [
    r"ATD\s*[:/=_\-]{0,2}\s*(\d{1,2}[/.\-]\d{1,2}[/.\-]\d{2,4})",
    r"vessel\s+departed\s+(?:on\s+)?(\d{1,2}[/.\-]\d{1,2}[/.\-]\d{2,4})",
    r"loaded\s+on\s+board\s+(\d{1,2}[/.\-]\d{1,2}[/.\-]\d{2,4})",
    r"\batd\s*[:/=_\-]{0,2}\s*(\d{1,2}[/.\-]\d{1,2}[/.\-]\d{2,4})",
]

# Telegram aggregation (module-local list, reset per scan run)
_telegram_lines: list[str] = []

# Logging — reuse root logger (shipment_brain already configures handlers)
log = logging.getLogger(__name__)


# ── Security gates ───────────────────────────────────────────────────────────

def _check_kill_switch() -> bool:
    """Return False (abort) if kill switch file exists."""
    if KILL_SWITCH.exists():
        log.warning("cnee_milestone: KILL SWITCH active at %s — aborting all drafts", KILL_SWITCH)
        return False
    return True


def _daily_count() -> int:
    """Return today's draft count from daily counter file."""
    today = date.today().isoformat()
    if DAILY_COUNTER.exists():
        try:
            data = json.loads(DAILY_COUNTER.read_text(encoding="utf-8"))
            if data.get("date") == today:
                return int(data.get("count", 0))
        except (json.JSONDecodeError, ValueError):
            pass
    return 0


def _increment_daily() -> int:
    """Increment today's draft count. Returns new count."""
    today = date.today().isoformat()
    count = _daily_count() + 1
    DAILY_COUNTER.write_text(
        json.dumps({"date": today, "count": count}),
        encoding="utf-8"
    )
    return count


def _check_auth_results(mail_item) -> bool:
    """
    Parse Authentication-Results header via Outlook PropertyAccessor.
    Require spf=pass + dkim=pass + dmarc=pass.
    Returns False if header missing or any check fails — logs reason.
    """
    PR_HEADERS = "http://schemas.microsoft.com/mapi/proptag/0x007D001F"
    try:
        headers = mail_item.PropertyAccessor.GetProperty(PR_HEADERS)
        if not headers:
            log.warning("cnee_milestone: Auth-Results header absent (empty)")
            return False
        # Find Authentication-Results block (may span multiple lines)
        auth_match = re.search(
            r"Authentication-Results\s*:(.*?)(?:\r?\n[^\s]|\Z)",
            headers, re.S | re.I
        )
        if not auth_match:
            log.warning("cnee_milestone: Authentication-Results header not found in raw headers")
            return False
        auth_text = auth_match.group(1).lower()
        for tag in ("spf", "dkim", "dmarc"):
            if tag + "=pass" not in auth_text:
                log.warning("cnee_milestone: Auth-Results %s≠pass, rejecting", tag)
                return False
        return True
    except Exception as e:
        log.warning("cnee_milestone: Auth-Results check error: %s", e)
        return False


# ── Sanitization ─────────────────────────────────────────────────────────────

def _sanitize(field: str, value: str) -> Optional[str]:
    """
    Sanitize a template placeholder value.
    Returns cleaned string, or None if value fails validation.

    Security: reject values containing control characters (newline/tab/CR)
    before any stripping — prevents BEC injection via compromised OPS email.
    """
    if not value:
        return None
    s = str(value)
    # REJECT if any control chars present (red-team A3: injection prevention)
    # Do NOT strip them — stripping could hide injected content
    if re.search(r"[\r\n\t\x00-\x1f]", s):
        log.debug("cnee_milestone: sanitize REJECT field=%s (control chars)", field)
        return None
    cleaned = s.strip()
    if not cleaned:
        return None
    # Collapse multiple spaces
    cleaned = re.sub(r" {2,}", " ", cleaned)

    limit, pattern = FIELD_LIMITS.get(field, (100, None))
    if len(cleaned) > limit:
        log.debug("cnee_milestone: sanitize FAIL field=%s length=%d > %d", field, len(cleaned), limit)
        return None
    if pattern and not re.match(pattern, cleaned, re.I):
        log.debug("cnee_milestone: sanitize FAIL field=%s value=%r regex=%s", field, cleaned, pattern)
        return None
    return cleaned


# ── Date parsing ─────────────────────────────────────────────────────────────

def _parse_date_flex(s: str) -> Optional[date]:
    """
    Parse dd/mm/yyyy, dd-mm-yyyy, dd.mm.yyyy (Vietnam convention: day first).
    Supports 2-digit and 4-digit years. Returns None on invalid input.
    """
    parts = re.split(r"[/.\-]", s.strip())
    if len(parts) != 3:
        return None
    try:
        d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
        if y < 100:
            y += 2000
        if not (1 <= d <= 31 and 1 <= m <= 12 and 2020 <= y <= 2040):
            return None
        return date(y, m, d)
    except (ValueError, OverflowError):
        return None


def extract_atd_date(text: str, received_time: datetime) -> Optional[date]:
    """
    Extract ATD date from email text.
    Validates result is within ReceivedTime +/- 30 days (red-team B4).
    Returns None if no valid date found.
    """
    candidates: list[date] = []
    for pat in _DATE_PATTERNS:
        for m in re.finditer(pat, text, re.I):
            d = _parse_date_flex(m.group(1))
            if d and d not in candidates:
                candidates.append(d)

    if not candidates:
        return None

    # Date sanity window: ATD must be within 30d before → 1d after receive
    lo = received_time.date() - timedelta(days=30)
    hi = received_time.date() + timedelta(days=1)
    for d in candidates:
        if lo <= d <= hi:
            return d

    log.debug(
        "cnee_milestone: ATD candidates %s all outside window [%s, %s]",
        candidates, lo, hi
    )
    return None


# ── ERP data readers (openpyxl read-only — no VBA risk for reads) ────────────

def _load_active_jobs() -> list[dict]:
    """
    Load Active Jobs sheet rows as list of dicts.
    Uses openpyxl in read-only mode (safe for concurrent reads).
    Returns empty list if ERP file not found.
    """
    if not ERP_PATH.exists():
        log.error("cnee_milestone: ERP file not found: %s", ERP_PATH)
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(ERP_PATH), read_only=True, data_only=True)
        ws = wb["Active Jobs"]

        # Build col index map from header row (AJ_HEADER_ROW = 7)
        rows_iter = ws.iter_rows(values_only=True)
        row_num = 0
        col_map: dict[str, int] = {}

        for row in rows_iter:
            row_num += 1
            if row_num == AJ_HEADER_ROW:
                for i, cell_val in enumerate(row):
                    if cell_val:
                        col_map[str(cell_val).strip().upper()] = i
                break

        if not col_map:
            log.error("cnee_milestone: Active Jobs header row not found at row %d", AJ_HEADER_ROW)
            wb.close()
            return []

        # Read data rows
        jobs = []
        today = date.today()
        cutoff = today - timedelta(days=60)

        for row in rows_iter:
            if not any(row):
                continue
            job = {}
            for col_name_key, col_idx_0based in col_map.items():
                if col_idx_0based < len(row):
                    val = row[col_idx_0based]
                    job[col_name_key] = str(val) if val is not None else ""
                else:
                    job[col_name_key] = ""

            # Recency filter: ETD > today - 60d
            etd_val = job.get("ETD", "")
            if etd_val:
                etd_d = _parse_job_date(etd_val)
                if etd_d and etd_d < cutoff:
                    continue  # Too old
            elif not job.get("BKG_NO", "") and not job.get("Bkg_No".upper(), ""):
                continue  # Empty row

            jobs.append(job)

        wb.close()
        return jobs

    except Exception as e:
        log.error("cnee_milestone: Failed to load Active Jobs: %s", e)
        return []


def _find_active_job(bkg: str, max_age_days: int = 60) -> Optional[dict]:
    """
    Find a job in Active Jobs by Bkg_No within recency window.
    Returns job dict or None.
    """
    bkg_upper = bkg.strip().upper()
    cutoff = date.today() - timedelta(days=max_age_days)
    for job in _load_active_jobs():
        job_bkg = job.get("BKG_NO", "").strip().upper()
        if job_bkg == bkg_upper:
            # Recency check
            etd_d = _parse_job_date(job.get("ETD", ""))
            if etd_d and etd_d < cutoff:
                continue
            return job
    return None


def _crm_auto_notify(customer_name: str) -> bool:
    """
    Check CRM.AUTO_NOTIFY = 'Y' for this customer.
    Returns False if customer not found or not opted in.
    """
    if not customer_name:
        return False
    if not ERP_PATH.exists():
        return False
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(ERP_PATH), read_only=True, data_only=True)
        ws = wb["CRM"]
        name_upper = customer_name.strip().upper()
        for row in ws.iter_rows(min_row=CRM_HEADER_ROW + 1, values_only=True):
            crm_name = str(row[CRM_COL["CUSTOMER_NAME"] - 1] or "").strip().upper()
            if crm_name == name_upper:
                auto_notify_val = str(row[CRM_COL["AUTO_NOTIFY"] - 1] or "").strip().upper()
                wb.close()
                return auto_notify_val == "Y"
        wb.close()
        return False
    except Exception as e:
        log.warning("cnee_milestone: CRM AUTO_NOTIFY check failed: %s", e)
        return False


def _lookup_cnee_emails(job: dict) -> list[str]:
    """
    Look up CNEE email addresses for a job.
    Priority: CRM Contact1_Email + Contact2_Email → Active Jobs EMAIL col.
    Supports semicolon-separated lists.
    Returns deduplicated list of valid email strings.
    """
    emails: list[str] = []
    customer = job.get("CUSTOMER", "").strip()

    # Try CRM first
    if customer and ERP_PATH.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(ERP_PATH), read_only=True, data_only=True)
            ws = wb["CRM"]
            name_upper = customer.upper()
            for row in ws.iter_rows(min_row=CRM_HEADER_ROW + 1, values_only=True):
                crm_name = str(row[CRM_COL["CUSTOMER_NAME"] - 1] or "").strip().upper()
                if crm_name == name_upper:
                    for col_key in ("CONTACT1_EMAIL", "CONTACT2_EMAIL"):
                        raw = str(row[CRM_COL[col_key] - 1] or "").strip()
                        if raw:
                            for e in re.split(r"[;,\s]+", raw):
                                e = e.strip()
                                if "@" in e and e not in emails:
                                    emails.append(e)
                    break
            wb.close()
        except Exception as e:
            log.warning("cnee_milestone: CRM email lookup failed: %s", e)

    # Fallback: Active Jobs EMAIL col
    if not emails:
        raw = job.get("EMAIL", "").strip()
        if raw:
            for e in re.split(r"[;,\s]+", raw):
                e = e.strip()
                if "@" in e and e not in emails:
                    emails.append(e)

    return emails


def _parse_job_date(val: str) -> Optional[date]:
    """
    Parse a date value from xlsm cell (may be string, datetime, or ISO string).
    Returns date or None.
    """
    if not val or val in ("None", ""):
        return None
    # Handle datetime objects serialized as strings
    if "+" in str(val) or "T" in str(val):
        try:
            return datetime.fromisoformat(str(val).split("+")[0]).date()
        except ValueError:
            pass
    # Try flex format
    return _parse_date_flex(str(val))


def _build_context(job: dict, atd: Optional[date]) -> dict:
    """
    Build template context dict from job data.
    Returns raw (unsanitized) values — caller must sanitize before use.
    """
    # Extract pol/pod from POL-POD field (format: "HCM→TACOMA" or "HCM-TACOMA")
    pol_pod = job.get("POL-POD", "")
    pol, pod = "", ""
    if pol_pod:
        for sep in ("→", "->", "-", "/"):
            if sep in pol_pod:
                parts = pol_pod.split(sep, 1)
                pol = parts[0].strip()
                pod = parts[1].strip()
                break
        if not pol:
            pol = pol_pod
    if not pod:
        pod = job.get("FINAL DEST", "")

    # Carrier may include type like "ONE FIX" → use as-is
    carrier = job.get("CARRIER", "")

    # ETD/ETA formatting
    etd_raw = job.get("ETD", "")
    eta_raw = job.get("ETA", "")
    etd_d = _parse_job_date(etd_raw)
    eta_d = _parse_job_date(eta_raw)

    return {
        "bkg":      job.get("BKG_NO", ""),
        "hbl":      job.get("HBL_NO", ""),
        "vessel":   "",   # Not in Active Jobs schema; will be N/A
        "carrier":  carrier,
        "customer": job.get("CUSTOMER", ""),
        "pol":      pol,
        "pod":      pod,
        "etd":      etd_d.strftime("%d/%m/%Y") if etd_d else etd_raw[:20],
        "eta":      eta_d.strftime("%d/%m/%Y") if eta_d else eta_raw[:20],
        "atd":      atd.strftime("%d/%m/%Y") if atd else "N/A",
    }


# ── State management ─────────────────────────────────────────────────────────

def _already_notified(bkg: str, notify_type: str) -> bool:
    """
    Dedup check: return True if (bkg, notify_type) already in sidecar JSONL.
    notify_type: "ATD" or "ETA7"
    """
    bkg_upper = bkg.strip().upper()
    if not STATE_FILE.exists():
        return False
    try:
        with STATE_FILE.open(encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    entry = json.loads(line)
                    if (entry.get("bkg", "").upper() == bkg_upper
                            and entry.get("type", "") == notify_type):
                        return True
                except json.JSONDecodeError:
                    continue
    except OSError:
        pass
    return False


def _write_state(bkg: str, notify_type: str, event_date: str,
                 mail_entry_id: Optional[str]) -> None:
    """
    Append a notification record to milestone_state.jsonl sidecar.
    This is the canonical record; xlsm is updated later via VBA Sync button.
    """
    record = {
        "ts": datetime.now().isoformat(),
        "bkg": bkg.strip().upper(),
        "type": notify_type,
        "date": event_date,
        "mail_entry_id": mail_entry_id or "",
    }
    try:
        STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
        with STATE_FILE.open("a", encoding="utf-8") as f:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    except OSError as e:
        log.error("cnee_milestone: Failed to write state: %s", e)


# ── Telegram ──────────────────────────────────────────────────────────────────

def _queue_telegram(line: str) -> None:
    """Add a line to the per-scan Telegram summary buffer."""
    _telegram_lines.append(line)


def flush_telegram_summary() -> None:
    """
    Send consolidated Telegram summary at end of scan run.
    Resets buffer afterward. Safe to call even if nothing queued.
    """
    global _telegram_lines
    if _telegram_lines:
        msg = "CNEE Milestone Summary\n\n" + "\n".join(_telegram_lines)
        _send_telegram(msg[:4000])
    _telegram_lines = []


def _send_telegram(message: str) -> bool:
    """Send a Telegram message. DISABLED 2026-04-26 — no-op."""
    log.debug("cnee_milestone._send_telegram disabled — message dropped (%d chars)", len(message))
    return True
    token = os.environ.get("TELEGRAM_TOKEN", "")  # noqa: unreachable
    chat_id = os.environ.get("TELEGRAM_CHAT_ID", "")
    if not token or not chat_id:
        log.debug("cnee_milestone: Telegram not configured, skip alert")
        return False
    try:
        import httpx
        resp = httpx.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            json={"chat_id": chat_id, "text": message},
            timeout=10,
        )
        return resp.status_code == 200
    except Exception as e:
        log.warning("cnee_milestone: Telegram send failed: %s", e)
        return False


# ── Logging helper ────────────────────────────────────────────────────────────

def _log(msg: str) -> None:
    """Log info-level message (convenience wrapper)."""
    log.info("cnee_milestone: %s", msg)


# ── Outlook draft composer ────────────────────────────────────────────────────

class _SafeDict(dict):
    """dict subclass: missing keys return 'N/A' instead of raising KeyError."""
    def __missing__(self, key: str) -> str:
        return "N/A"


def _create_outlook_draft_polite(to_list: list[str], template_type: str,
                                  ctx: dict,
                                  attachments: list[dict] | None = None) -> bool:
    """Create email draft in Outlook via COM (save, don't send)."""
    import pythoncom, win32com.client
    pythoncom.CoInitialize()

    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
    except Exception as e:
        log.error("cnee_milestone: COM Outlook unavailable: %s", e)
        return False

    m = outlook.CreateItem(0)  # olMailItem = 0

    # Load template
    tmpl_path = (Path(__file__).parent.parent
                 / "templates" / f"milestone_{template_type.lower()}.txt")
    if not tmpl_path.exists():
        log.error("cnee_milestone: Template not found: %s", tmpl_path)
        return False

    tmpl = tmpl_path.read_text(encoding="utf-8")
    filled = tmpl.format_map(_SafeDict(ctx))

    # Extract subject from first line "Subject: ..."
    lines = filled.split("\n", 2)
    if lines[0].upper().startswith("SUBJECT:"):
        subject = lines[0].split(":", 1)[1].strip()
        body = "\n".join(lines[2:]).strip() if len(lines) > 2 else filled
    else:
        subject = f"Shipment Update — {ctx.get('bkg', 'N/A')}"
        body = filled

    m.Subject = f"[AUTO] {subject}"
    m.HTMLBody = f"<html><body><pre>{body}</pre></body></html>"
    m.To = "; ".join(to_list)

    if attachments:
        for a in attachments:
            m.Attachments.Add(a["data"], 1, 0, a["filename"])

    m.Save()  # Draft only — NOT Send()
    log.info("cnee_milestone: COM draft saved for %s / %s",
             ctx.get("customer"), ctx.get("bkg"))
    return True


# ── Per-Bkg processing ────────────────────────────────────────────────────────

def _process_bkg(bkg: str, atd: date, source_mail, notify_type: str = "ATD") -> bool:
    """
    Process one Bkg number: find job, check opt-in, sanitize, create draft.
    Returns True if draft created.
    """
    # Match Active Jobs (recency filter inside _find_active_job)
    job = _find_active_job(bkg, max_age_days=60)
    if not job:
        log.debug("cnee_milestone: Bkg %s not found in Active Jobs (recent)", bkg)
        return False

    # Dedup check
    if _already_notified(bkg, notify_type):
        log.debug("cnee_milestone: Bkg %s already notified for %s", bkg, notify_type)
        return False

    # CRM opt-in
    customer = job.get("CUSTOMER", "")
    if not _crm_auto_notify(customer):
        log.debug("cnee_milestone: %s not opted in (AUTO_NOTIFY != Y)", customer)
        return False

    # Email lookup
    cnee_emails = _lookup_cnee_emails(job)
    if not cnee_emails:
        _queue_telegram(f"WARN: Missing CNEE email for {customer} / {bkg}")
        log.warning("cnee_milestone: No CNEE email found for %s / %s", customer, bkg)
        return False

    # Build context and sanitize all placeholders
    ctx = _build_context(job, atd)
    sanitized_ctx: dict[str, str] = {}
    for k in FIELD_LIMITS:
        raw_val = ctx.get(k, "")
        result = _sanitize(k, raw_val) if raw_val else None
        if result is None and raw_val:
            _queue_telegram(f"WARN: Sanitize fail for field '{k}' in {bkg}")
            log.warning("cnee_milestone: Sanitize fail field=%s bkg=%s", k, bkg)
            return False
        sanitized_ctx[k] = result if result is not None else "N/A"

    # Compose draft
    template_key = "atd" if notify_type == "ATD" else "eta7"
    if not _create_outlook_draft_polite(cnee_emails, template_key, sanitized_ctx):
        return False

    # Record in sidecar (NOT xlsm — avoid race)
    entry_id = None
    try:
        entry_id = source_mail.EntryID if source_mail else None
    except Exception:
        pass
    _write_state(bkg, notify_type, atd.isoformat() if atd else "", entry_id)
    _queue_telegram(f"OK: Draft created — {customer} / {bkg}")
    _increment_daily()
    return True


# ── Main hook (called from shipment_brain.py) ─────────────────────────────────

def on_atd_detected(mail_item, stages: list, identifiers: dict,
                    sender_smtp: str) -> bool:
    """
    Main entry point called by shipment_brain after detecting ATD/LOADED stages.

    Args:
        mail_item:   Outlook MailItem COM object
        stages:      list of detected stage strings (e.g. ["ATD", "LOADED"])
        identifiers: dict with keys "HBL", "BKG", "CTN" (each a list)
        sender_smtp: SMTP email of sender (from get_sender_smtp)

    Returns:
        True if at least one draft was created.
    """
    # Kill switch check
    if not _check_kill_switch():
        return False

    # Sender allowlist (explicit, not domain-wide)
    if OPS_ALLOWLIST and sender_smtp.lower() not in {e.lower() for e in OPS_ALLOWLIST}:
        log.debug("cnee_milestone: Sender %s not in OPS_ALLOWLIST, skip", sender_smtp)
        return False

    # Auth-Results (SPF/DKIM/DMARC)
    if not _check_auth_results(mail_item):
        _log(f"Rejected: auth-results fail from {sender_smtp}")
        return False

    # Daily rate limit
    if _daily_count() >= MAX_DRAFTS_PER_DAY:
        _queue_telegram(f"WARN: Daily draft cap ({MAX_DRAFTS_PER_DAY}) reached")
        return False

    # Extract text for analysis
    try:
        subject = str(mail_item.Subject or "")
        body = str(mail_item.Body or "")[:2000]   # cap to avoid huge mails
    except Exception:
        subject, body = "", ""
    full_text = f"{subject}\n{body}"

    # Blacklist check (VESSEL CHANGE, RVS ETD, etc.)
    if BLACKLIST_REGEX.search(full_text):
        log.info("cnee_milestone: Blacklisted keyword in mail: %s", subject[:60])
        return False

    # Bulk detect: >3 Bkg = likely bulk rate sheet, not voyage notification
    bkg_list: list[str] = identifiers.get("BKG", [])
    if len(bkg_list) > 3:
        _queue_telegram(f"WARN: Bulk mail detected ({len(bkg_list)} Bkg), skipping")
        log.info("cnee_milestone: Bulk mail (%d Bkg), skip: %s", len(bkg_list), subject[:60])
        return False

    # Extract ATD date with sanity cross-check
    try:
        received_time = mail_item.ReceivedTime
        # Convert COM datetime to Python datetime if needed
        if not isinstance(received_time, datetime):
            received_time = datetime.now()
    except Exception:
        received_time = datetime.now()

    atd = extract_atd_date(full_text, received_time)
    if not atd:
        _queue_telegram(f"WARN: ATD/LOADED detected but no parseable date: {subject[:60]}")
        log.info("cnee_milestone: No parseable ATD date in: %s", subject[:60])
        return False

    # Process each Bkg, respect per-run cap
    made_drafts = 0
    for bkg in bkg_list:
        if made_drafts >= MAX_DRAFTS_PER_RUN:
            log.info("cnee_milestone: Per-run cap (%d) reached", MAX_DRAFTS_PER_RUN)
            break
        if _process_bkg(bkg, atd, mail_item, notify_type="ATD"):
            made_drafts += 1

    return made_drafts > 0


# ── ETA-7 reminder (CLI entry for Task Scheduler) ────────────────────────────

def run_eta_reminder() -> int:
    """
    Check Active Jobs for ETAs in [today+1, today+8] window.
    Create drafts for opted-in customers not yet notified.
    Called by Task Scheduler daily at 08:00.
    Returns count of drafts created.
    """
    if not _check_kill_switch():
        return 0

    today = date.today()
    window_lo = today + timedelta(days=1)
    window_hi = today + timedelta(days=8)
    count = 0

    for job in _load_active_jobs():
        if count >= MAX_DRAFTS_PER_RUN:
            break

        eta = _parse_job_date(job.get("ETA_DATE", "") or job.get("ETA", ""))
        if not eta:
            continue
        if not (window_lo <= eta <= window_hi):
            continue

        bkg = job.get("BKG_NO", "").strip()
        if not bkg:
            continue
        if _already_notified(bkg, "ETA7"):
            continue
        if not _crm_auto_notify(job.get("CUSTOMER", "")):
            continue

        cnee_emails = _lookup_cnee_emails(job)
        if not cnee_emails:
            continue

        ctx = _build_context(job, None)
        # Sanitize
        sanitized_ctx: dict[str, str] = {}
        valid = True
        for k in FIELD_LIMITS:
            raw_val = ctx.get(k, "")
            result = _sanitize(k, raw_val) if raw_val else None
            if result is None and raw_val:
                log.warning("cnee_milestone: ETA7 sanitize fail field=%s bkg=%s", k, bkg)
                valid = False
                break
            sanitized_ctx[k] = result if result is not None else "N/A"
        if not valid:
            continue

        if _create_outlook_draft_polite(cnee_emails, "eta7", sanitized_ctx):
            _write_state(bkg, "ETA7", today.isoformat(), None)
            _queue_telegram(f"ETA-7 draft: {ctx.get('customer')} / {bkg}")
            _increment_daily()
            count += 1

    flush_telegram_summary()
    return count


# ── CLI entry ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    # Configure basic logging for CLI runs
    logging.basicConfig(
        level=logging.INFO,
        format="[%(asctime)s] %(levelname)-8s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    if len(sys.argv) > 1 and sys.argv[1] == "eta-reminder":
        n = run_eta_reminder()
        print(f"ETA-7 drafts created: {n}")
    else:
        print("Usage: python -m email_engine.core.cnee_milestone eta-reminder")
        sys.exit(1)
