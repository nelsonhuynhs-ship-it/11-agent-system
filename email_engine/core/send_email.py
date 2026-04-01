"""
send_email.py -- Knowledge-Aware Email Sender
=============================================
Upgrade v4.0: Adds --source and --sequence modes via sequence_engine.

Modes
-----
  CMD_MODE  (default)  -- interactive, select target by CMD name (data.xlsx)
  TIER_MODE            -- target prospects from customer_final.xlsx by tier
                          Activate: python send_email.py --tier
  SOURCE_MODE          -- target from master files (cnee/contact/shipper)
                          Activate: python send_email.py --source contact
  SEQUENCE_MODE        -- 3-step email sequence auto-progression
                          Activate: python send_email.py --sequence
  DRY_RUN              -- simulate any mode without sending
                          Add: --dry-run
"""

# =========================================================
# EARLY CLI ROUTING: delegate --source / --sequence to sequence_engine
# =========================================================
import sys as _sys
_args = _sys.argv[1:]
if any(a in _args for a in ("--source", "--sequence")):
    from sequence_engine import main as _seq_main
    _seq_main()
    _sys.exit(0)

# =========================================================
# AUTO-RATE MODE: per-customer rate tables from Parquet
# =========================================================
if "--auto-rate" in _args:
    import pandas as _pd
    import win32com.client as _w32
    import csv as _csv
    import logging as _log
    import json as _json
    from pathlib import Path as _Path
    from datetime import datetime as _dt

    _log.basicConfig(level=_log.INFO,
                     format="[%(asctime)s] %(levelname)-8s %(message)s",
                     datefmt="%H:%M:%S")
    _logger = _log.getLogger("auto_rate_send")

    _BASE = _Path(__file__).parent
    _ROOT = _BASE.parent
    _DATA_FILE   = _ROOT / "data.xlsx"
    _CONFIG_FILE = _ROOT / "data" / "config.xlsx"
    _PROFILE_PDF = _ROOT / "assets" / "PUDONG PRIME PROFILE.pdf"
    _LOGO_PNG    = _ROOT / "assets" / "logo.png"
    _LOG_DIR     = _ROOT / "logs"
    _LOG_DIR.mkdir(exist_ok=True)
    _EMAIL_LOG   = _LOG_DIR / "email_log.csv"

    def _log_send(email, subject, campaign):
        exists = _EMAIL_LOG.exists()
        with open(_EMAIL_LOG, "a", newline="", encoding="utf-8") as f:
            w = _csv.writer(f)
            if not exists:
                w.writerow(["timestamp", "email", "subject", "campaign_id", "cycle_id", "status"])
            w.writerow([_dt.now().strftime("%Y-%m-%d %H:%M:%S"), email, subject, campaign, "1", "SENT"])

    # Load config for Intro/Closing/Signature/Subject
    from auto_rate_builder import build_rate_table_for_customer

    def _load_config_plain():
        wb = __import__("openpyxl").load_workbook(str(_CONFIG_FILE), data_only=True)
        ws = wb.active
        cfg = {}
        for row in ws.iter_rows(max_col=2, values_only=True):
            k = str(row[0] or "").strip().upper()
            v = str(row[1] or "").strip() if row[1] else ""
            if k and k != "KEY":
                cfg[k] = v
        return cfg

    import random as _rnd
    from datetime import date as _date

    def _pick(text):
        items = [x.strip() for x in text.split("|") if x.strip()]
        return _rnd.choice(items) if items else ""

    def _gen_subject(cfg):
        base = _pick(cfg.get("SUBJECTTEMPLATES", ""))
        suffix = cfg.get("SUBJECTSUFFIX", "NELSON")
        wk = _date.today().isocalendar()[1]
        return f"{base} // {suffix} WEEK {wk}"

    def _build_preheader(cfg):
        ph = _pick(cfg.get("PREHEADER", ""))
        return (f'<span style="display:none!important;visibility:hidden;opacity:0;'
                f'color:transparent;height:0;width:0;overflow:hidden;mso-hide:all;">'
                f'{ph}</span>')

    def auto_rate_send_mode():
        cfg = _load_config_plain()
        _logger.info("Config loaded: %d keys", len(cfg))

        # 1. Load data
        df = _pd.read_excel(_DATA_FILE)
        df.columns = df.columns.str.strip().str.upper()

        # 2. Select CMD
        cmds = sorted(df["CMD_NAME"].dropna().unique())
        print("\n" + "=" * 60)
        print("  AUTO QUOTE SEND — Per-Customer Rate Tables from Parquet")
        print("=" * 60)
        print()
        for i, c in enumerate(cmds, 1):
            count = len(df[df["CMD_NAME"] == c])
            print(f"  {i:2d}. {c:25s} ({count} contacts)")
        print()
        idxs = input("  Select CMD (comma-sep numbers): ").strip()
        if not idxs:
            _sys.exit("No CMD selected.")
        selected = []
        for x in idxs.split(","):
            try:
                selected.append(cmds[int(x.strip()) - 1])
            except (ValueError, IndexError):
                pass
        if not selected:
            _sys.exit("Invalid CMD.")

        # 3. Markup amount
        markup_input = input("  Markup per container (default $20): ").strip()
        markup = float(markup_input) if markup_input else 20.0

        # Default destinations when customer has no DESTINATION data
        DEFAULT_DESTS = "USCHI,USLAX,USLGB,USEWR,USSAV,USNYC"

        # 4. Filter customers — include ALL with valid email (not just those with DESTINATION)
        subset = df[df["CMD_NAME"].isin(selected)].copy()
        subset = subset[subset["CNEE_EMAIL"].notna()]
        subset["_email_lower"] = subset["CNEE_EMAIL"].astype(str).str.lower().str.strip()
        subset = subset.drop_duplicates(subset="_email_lower")

        has_dest = subset["DESTINATION"].notna().sum()
        no_dest  = len(subset) - has_dest
        _logger.info("Total customers: %d (has DESTINATION: %d, will use default: %d)",
                     len(subset), has_dest, no_dest)

        # 5. Build per-customer rate tables
        print(f"\n  Loading Parquet & building rate tables for {len(subset)} customers...")
        print(f"  ({has_dest} with custom routes, {no_dest} will get top US routes)")
        results = []
        no_rates = 0
        for _, row in subset.iterrows():
            email = str(row.get("CNEE_EMAIL", "")).strip()
            pol   = str(row.get("POL", "HPH")).strip() or "HPH"
            raw_dest = row.get("DESTINATION")
            # Use customer DESTINATION if available, otherwise default top US ports
            if _pd.notna(raw_dest) and str(raw_dest).strip() and str(raw_dest).strip().lower() != "nan":
                dest = str(raw_dest).strip()
            else:
                dest = DEFAULT_DESTS
            pic   = str(row.get("CNEE_PIC", "")).strip()
            company = str(row.get("CNEE_NAME", "")).strip()

            result = build_rate_table_for_customer(
                pol=pol, destinations=dest, markup=markup
            )
            if result["routes_found"] == 0:
                no_rates += 1
                continue

            results.append({
                "email":   email,
                "pic":     pic if pic and pic.lower() not in ("nan", "") else "Team",
                "company": company,
                "pol":     pol,
                "dest":    dest,
                "html":    result["html"],
                "routes":  result["routes_found"],
                "rates":   result["total_rates"],
                "detail":  result["routes_detail"],
                "cmd":     row.get("CMD_NAME", ""),
            })

        # 6. Summary
        print("\n" + "=" * 60)
        print("  AUTO QUOTE BUILD REPORT")
        print("=" * 60)
        print(f"  Customers with rates : {len(results)}")
        print(f"  Customers no rates   : {no_rates}")
        print(f"  Markup               : ${markup:.0f}")
        print()
        for r in results[:5]:
            routes_str = ", ".join(f"{d['port']}({len(d['carriers'])})" for d in r["detail"])
            print(f"    {r['company'][:30]:30s} | {r['email'][:35]:35s} | {routes_str}")
        if len(results) > 5:
            print(f"    ... and {len(results) - 5} more")
        print("=" * 60)

        if not results:
            print("  No customers with valid rates. Nothing to send.")
            _sys.exit(0)

        # 7. Preview first email
        print("\n  Generating preview of first email...")
        r0 = results[0]
        outlook = _w32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = r0["email"]
        mail.Subject = _gen_subject(cfg)

        if _PROFILE_PDF.exists():
            mail.Attachments.Add(str(_PROFILE_PDF))
        if _LOGO_PNG.exists():
            logo = mail.Attachments.Add(str(_LOGO_PNG))
            logo.PropertyAccessor.SetProperty(
                "http://schemas.microsoft.com/mapi/proptag/0x3712001F", "pudonglogo"
            )

        intro = cfg.get("INTROTEXT", "")
        closing = cfg.get("CLOSINGTEXT", "")
        signature = cfg.get("SIGNATURE", "")

        mail.HTMLBody = f"""
        <html><body>
        {_build_preheader(cfg)}
        Dear {r0['pic']},<br><br>
        {intro}<br><br>
        {r0['html']}<br><br>
        {closing}<br><br>
        {signature}
        </body></html>
        """
        mail.Display()

        confirm = input("\n  Preview OK? Send to all customers? (Y/N): ").strip().upper()
        if confirm != "Y":
            _sys.exit("  Cancelled by user.")

        # 8. Send loop
        campaign = f"AUTO_RATE_{_dt.now():%Y%m%d_%H%M}"
        sent = 0
        for r in results:
            m = outlook.CreateItem(0)
            m.To = r["email"]
            m.Subject = _gen_subject(cfg)

            if _PROFILE_PDF.exists():
                m.Attachments.Add(str(_PROFILE_PDF))
            if _LOGO_PNG.exists():
                lg = m.Attachments.Add(str(_LOGO_PNG))
                lg.PropertyAccessor.SetProperty(
                    "http://schemas.microsoft.com/mapi/proptag/0x3712001F", "pudonglogo"
                )

            m.HTMLBody = f"""
            <html><body>
            {_build_preheader(cfg)}
            Dear {r['pic']},<br><br>
            {intro}<br><br>
            {r['html']}<br><br>
            {closing}<br><br>
            {signature}
            </body></html>
            """
            m.Send()
            _log_send(r["email"], m.Subject, campaign)
            _logger.info("SENT -> %s (%d routes)", r["email"], r["routes"])
            sent += 1

        print("\n" + "=" * 60)
        print(f"  AUTO QUOTE SEND COMPLETE: {sent} emails sent")
        print(f"  Campaign: {campaign}")
        print("=" * 60)

    auto_rate_send_mode()
    _sys.exit(0)


import pandas as pd
import win32com.client
import random
import sys
import csv
import shutil
import logging
from datetime import date, datetime
from pathlib import Path

# =========================================================
# LOGGING — ASCII only to avoid cp1258 encoding errors
# =========================================================
logging.basicConfig(
    level   = logging.INFO,
    format  = "[%(asctime)s] %(levelname)-8s %(message)s",
    datefmt = "%H:%M:%S",
    handlers= [logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# =========================================================
# BASIC CONFIG
# =========================================================
BATCH_SIZE = 100

BASE_DIR     = Path(__file__).parent
PROJECT_ROOT = BASE_DIR.parent
DATA_FILE    = PROJECT_ROOT / "data.xlsx"
CONFIG_FILE  = PROJECT_ROOT / "data/config.xlsx"
PROFILE_PDF  = PROJECT_ROOT / "assets/PUDONG PRIME PROFILE.pdf"
LOGO_PNG     = PROJECT_ROOT / "assets/logo.png"
LOG_DIR      = PROJECT_ROOT / "logs"
BACKUP_DIR   = PROJECT_ROOT / "backup"

CMD_HISTORY_FILE = PROJECT_ROOT / "logs" / "cmd_send_history.csv"
EMAIL_LOG_FILE   = PROJECT_ROOT / "logs" / "email_log.csv"
KNOWLEDGE_FILE   = PROJECT_ROOT / "logs" / "email_knowledge.csv"

LOG_DIR.mkdir(exist_ok=True)
BACKUP_DIR.mkdir(exist_ok=True)

# =========================================================
# KNOWLEDGE STATUS CATEGORIES
# =========================================================
# Emails with these statuses are SKIPPED entirely (dead emails)
DEAD_STATUSES = {"hard_bounce", "policy_reject", "spam_block", "invalid"}

# Emails with these statuses get a WARNING but are still sent (user decides)
WARM_STATUSES = {"human_reply", "soft_bounce"}

# Auto-reply: user is away — still send but flag it
AUTO_REPLY_STATUSES = {"auto_reply"}

# =========================================================
# LOAD KNOWLEDGE BASE
# =========================================================
def load_knowledge() -> dict[str, dict]:
    """
    Returns dict: email (lowercase) -> {status, replacement_email, remark, count}
    """
    if not KNOWLEDGE_FILE.exists():
        log.info("No email_knowledge.csv found — all emails treated as unknown.")
        return {}

    df = pd.read_csv(KNOWLEDGE_FILE, encoding="utf-8-sig")
    df.columns = df.columns.str.upper()
    df["EMAIL"] = df["EMAIL"].astype(str).str.lower().str.strip()

    knowledge: dict[str, dict] = {}
    for _, row in df.iterrows():
        email = row["EMAIL"]
        if email and "@" in email:
            knowledge[email] = {
                "status":            str(row.get("STATUS", "")).lower().strip(),
                "replacement_email": str(row.get("REPLACEMENT_EMAIL", "")).lower().strip(),
                "remark":            str(row.get("REMARK", "")).strip(),
                "count":             int(row.get("COUNT", 1)),
            }
    log.info("Knowledge base loaded: %d entries", len(knowledge))
    return knowledge


# How many days must pass before the same email can be sent again to the same CMD.
# Set to 0 to disable cooldown (allow resend anytime).
RESEND_COOLDOWN_DAYS = 7


def load_email_log() -> tuple[pd.DataFrame, dict]:
    """
    Returns (df_log, recent_pairs)

    recent_pairs: dict of (email_lower, campaign_id_upper) -> last_sent_date
        Only contains sends within the last RESEND_COOLDOWN_DAYS.
        If RESEND_COOLDOWN_DAYS == 0, dict is empty (no cooldown).
    """
    if EMAIL_LOG_FILE.exists():
        df_log = pd.read_csv(EMAIL_LOG_FILE)
        df_log.columns = df_log.columns.str.strip().str.upper()
    else:
        df_log = pd.DataFrame()

    required = ["TIMESTAMP", "EMAIL", "CAMPAIGN_ID", "CYCLE_ID", "SUBJECT", "STATUS"]
    for col in required:
        if col not in df_log.columns:
            df_log[col] = ""

    df_log["EMAIL"]       = df_log["EMAIL"].astype(str).str.lower()
    df_log["CAMPAIGN_ID"] = df_log["CAMPAIGN_ID"].astype(str).str.upper()
    df_log["CYCLE_ID"]    = df_log["CYCLE_ID"].astype(str)
    df_log["TIMESTAMP"]   = pd.to_datetime(df_log["TIMESTAMP"], errors="coerce")

    if RESEND_COOLDOWN_DAYS > 0:
        cutoff = pd.Timestamp.now() - pd.Timedelta(days=RESEND_COOLDOWN_DAYS)
        recent = df_log[df_log["TIMESTAMP"] >= cutoff]
    else:
        recent = df_log.iloc[0:0]  # empty — no cooldown

    # Build dict: (email, campaign_id) -> most recent send date
    recent_pairs: dict[tuple, pd.Timestamp] = {}
    for _, row in recent.iterrows():
        key = (row["EMAIL"], row["CAMPAIGN_ID"])
        ts  = row["TIMESTAMP"]
        if key not in recent_pairs or ts > recent_pairs[key]:
            recent_pairs[key] = ts

    log.info("Email log loaded: %d total rows, %d sent within last %d days",
             len(df_log), len(recent_pairs), RESEND_COOLDOWN_DAYS)
    return df_log, recent_pairs


def log_email_send(email: str, subject: str, campaign_id: str, cycle_id: str = "1", status: str = "SENT"):
    log_exists = EMAIL_LOG_FILE.exists()
    with open(EMAIL_LOG_FILE, mode="a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not log_exists:
            writer.writerow(["timestamp", "email", "subject", "campaign_id", "cycle_id", "status"])
        writer.writerow([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            email, subject, campaign_id, cycle_id, status
        ])


# =========================================================
# CMD HISTORY
# =========================================================
def ensure_cmd_history_file():
    if not CMD_HISTORY_FILE.exists() or CMD_HISTORY_FILE.stat().st_size == 0:
        with open(CMD_HISTORY_FILE, "w", newline="", encoding="latin1") as f:
            csv.writer(f).writerow(["CMD_NAME", "SENT_DATE", "SENT_COUNT"])

ensure_cmd_history_file()


# =========================================================
# LOAD CONFIG  —  Rich Text Aware (openpyxl)
# =========================================================
def _rich_cell_to_html(cell_value) -> str:
    """
    Convert an openpyxl cell value to Outlook-safe HTML.

    Supports:
      - Plain str  → newlines become <br>, text unchanged
      - CellRichText (formatted runs) → each run's color / bold / italic
        is converted to HTML tags that Outlook renders correctly.

    In config.xlsx the user can:
      • Type text normally and use Excel Format Cells to set:
          - bold        → Ctrl+B
          - color       → Font Color button
          - italic      → Ctrl+I
          - underline   → Ctrl+U
      • Press Alt+Enter for a new line inside the cell.
    All of these will be preserved in the sent email.
    """
    from openpyxl.cell.rich_text import CellRichText, TextBlock

    if cell_value is None:
        return ""

    def _run_to_html(text: str, font) -> str:
        """Wrap a single text run in appropriate HTML tags."""
        if not text:
            return ""
        # Escape HTML special chars
        text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        # Newlines inside Excel cell (Alt+Enter) → <br>
        text = text.replace("\n", "<br>")

        if font is None:
            return text

        # ── Color ──────────────────────────────────────────────────────────
        # <font color=""> is the only reliable way in Outlook
        # Excel cells use two color systems:
        #   type='rgb'   → direct hex, use as-is
        #   type='theme' → index into Office theme palette; resolve manually
        color_hex = None
        try:
            c = font.color
            if c:
                if c.type == "rgb" and c.rgb:
                    raw = str(c.rgb)
                    # Strip alpha (AARRGGBB → RRGGBB); skip black "00000000"/"FF000000"
                    if raw not in ("00000000", "FF000000", "00000000"):
                        color_hex = "#" + raw[-6:]

                elif c.type == "theme":
                    # Office default theme color palette (index 0-11)
                    # Matches Office 2016/2019/365 default "Office" theme
                    OFFICE_THEME_COLORS = {
                        0:  "#FFFFFF",  # Background 1 (white)
                        1:  "#000000",  # Text 1 (black)
                        2:  "#E7E6E6",  # Background 2
                        3:  "#44546A",  # Text 2 (dark blue-grey)
                        4:  "#4472C4",  # Accent 1 (blue)
                        5:  "#ED7D31",  # Accent 2 (orange)
                        6:  "#A9D18E",  # Accent 3 (green light)
                        7:  "#FFC000",  # Accent 4 (yellow)
                        8:  "#5B9BD5",  # Accent 5 (light blue)
                        9:  "#70AD47",  # Accent 6 (green)
                        10: "#264478",  # Darker Blue
                        11: "#843C0C",  # Darker Orange
                    }
                    theme_idx = int(c.theme) if c.theme is not None else -1
                    base_color = OFFICE_THEME_COLORS.get(theme_idx)

                    if base_color and theme_idx not in (0, 1, 2):
                        # Apply tint (Excel lightens/darkens theme colors with tint)
                        tint = float(c.tint) if c.tint else 0.0
                        if abs(tint) > 0.01:
                            # Decode base color
                            r = int(base_color[1:3], 16)
                            g = int(base_color[3:5], 16)
                            b = int(base_color[5:7], 16)
                            if tint > 0:   # lighten
                                r = int(r + (255 - r) * tint)
                                g = int(g + (255 - g) * tint)
                                b = int(b + (255 - b) * tint)
                            else:          # darken
                                r = int(r * (1 + tint))
                                g = int(g * (1 + tint))
                                b = int(b * (1 + tint))
                            r, g, b = max(0, min(255, r)), max(0, min(255, g)), max(0, min(255, b))
                            color_hex = f"#{r:02X}{g:02X}{b:02X}"
                        else:
                            color_hex = base_color
        except Exception:
            pass

        if color_hex:
            text = f'<font color="{color_hex}">{text}</font>'


        # Bold / italic / underline
        if getattr(font, "underline", None) and font.underline != "none":
            text = f"<u>{text}</u>"
        if getattr(font, "italic", False):
            text = f"<em>{text}</em>"
        if getattr(font, "bold", False):
            text = f"<strong>{text}</strong>"

        return text

    # ── Case 1: CellRichText (multiple formatted runs) ──────────────────────
    if isinstance(cell_value, CellRichText):
        parts = []
        for block in cell_value:
            if isinstance(block, TextBlock):
                parts.append(_run_to_html(block.text, block.font))
            else:
                # Plain string segment within rich text
                raw = str(block).replace("\n", "<br>")
                parts.append(raw)
        return "".join(parts)

    # ── Case 2: Plain string ─────────────────────────────────────────────────
    text = str(cell_value)
    text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    text = text.replace("\n", "<br>")
    return text


def _load_config_rich(config_path: Path) -> dict:
    """
    Load config.xlsx using openpyxl with rich_text=True so that cell
    formatting (colors, bold, etc.) is preserved for HTML conversion.

    Returns dict: UPPER_KEY → HTML string (for rich cells)
                           OR plain string (for simple cells).
    """
    from openpyxl import load_workbook
    from openpyxl.cell.rich_text import CellRichText

    wb = load_workbook(str(config_path), rich_text=True, data_only=True)
    ws = wb.active
    result = {}

    for row in ws.iter_rows():
        if not row:
            continue
        key_cell = row[0]
        val_cell  = row[1] if len(row) > 1 else None

        key = str(key_cell.value or "").strip().upper()
        if not key or key == "KEY":
            continue

        if val_cell is None or val_cell.value is None:
            result[key] = ""
            continue

        v = val_cell.value
        if isinstance(v, CellRichText):
            result[key] = _rich_cell_to_html(v)
        else:
            result[key] = str(v).strip()

    return result


cfg = _load_config_rich(CONFIG_FILE)




def pick_random(text: str) -> str:
    if not isinstance(text, str):
        return ""
    items = [x.strip() for x in text.split("|") if x.strip()]
    return random.choice(items) if items else ""

def get_iso_week() -> int:
    return date.today().isocalendar()[1]

def generate_subject() -> str:
    base   = pick_random(cfg.get("SUBJECTTEMPLATES", ""))
    suffix = cfg.get("SUBJECTSUFFIX", "NELSON")
    return f"{base} // {suffix} WEEK {get_iso_week()}"

def generate_preheader() -> str:
    return pick_random(cfg.get("PREHEADER", ""))

def build_preheader_html() -> str:
    return f"""
    <span style="display:none!important;visibility:hidden;opacity:0;
    color:transparent;height:0;width:0;overflow:hidden;mso-hide:all;">
    {generate_preheader()}
    </span>
    """



# NOTE: cfg['INTROTEXT'] and cfg['CLOSINGTEXT'] are already HTML
# (converted from Excel cell rich-text by _load_config_rich above).
# No further conversion needed.



# =========================================================
# TIER_MODE  —  Semi-Auto Send by Reply Tier
# =========================================================

# Per-tier cooldown days (REPLY_3 = most aggressive follow-up)
TIER_COOLDOWN = {"REPLY_3": 3, "REPLY_2": 7, "REPLY_1": 14}

# Column to use as recipient
FINAL_FILE = PROJECT_ROOT / "data" / "customer_final.xlsx"


def tier_mode() -> None:
    """
    TIER_MODE: load customer_final.xlsx, select tier, send with tier-aware
    cooldown and personalised subject copy.
    """
    print("\n" + "=" * 60)
    print("  TIER SEND MODE")
    print("  Source: customer_final.xlsx (prospect-only)")
    print("=" * 60)

    if not FINAL_FILE.exists():
        print("  customer_final.xlsx not found. Run process_reply.py first.")
        sys.exit(1)

    # --- Select tier ---
    print("\nChoose send tier:")
    print("  1. REPLY_3  — Hot leads (booking/price intent)   [cooldown: 3 days]")
    print("  2. REPLY_2  — Warm leads (gratitude/inquiry)      [cooldown: 7 days]")
    print("  3. REPLY_1  — Cold leads (single reply)           [cooldown: 14 days]")
    tier_choice = input("\n  Your choice (1-3): ").strip()
    tier_map    = {"1": "REPLY_3", "2": "REPLY_2", "3": "REPLY_1"}
    selected_tier = tier_map.get(tier_choice)
    if not selected_tier:
        sys.exit("Invalid tier choice.")

    cooldown_days = TIER_COOLDOWN[selected_tier]

    # --- Load tier sheet ---
    try:
        df_tier = pd.read_excel(FINAL_FILE, sheet_name=selected_tier)
        df_tier.columns = df_tier.columns.str.strip().str.upper()
    except Exception as exc:
        print(f"  Cannot read sheet '{selected_tier}': {exc}")
        sys.exit(1)

    if df_tier.empty:
        print(f"  No prospects in {selected_tier} sheet. Nothing to send.")
        return

    # --- Filter by EMAIL and CNEE_EMAIL columns ---
    knowledge_tm = load_knowledge()
    _, recent_pairs_tm = load_email_log()

    targets: list[dict] = []
    skip_dead = skip_cool = 0

    # Build a campaign label for logging
    campaign_label = f"TIER_{selected_tier}_{datetime.now():%Y%m%d}"

    for _, row in df_tier.iterrows():
        email_col = "CNEE_EMAIL" if "CNEE_EMAIL" in df_tier.columns else None
        if not email_col:
            continue
        email = str(row.get(email_col, "")).strip()
        if not email or "@" not in email:
            continue
        email_l = email.lower()

        # Knowledge dead check
        kb = knowledge_tm.get(email_l, {})
        if kb.get("status", "") in DEAD_STATUSES:
            skip_dead += 1
            continue

        # Cooldown check
        cutoff = pd.Timestamp.now() - pd.Timedelta(days=cooldown_days)
        key    = (email_l, campaign_label)
        if key in recent_pairs_tm:
            skip_cool += 1
            continue

        pic = str(row.get("CNEE_PIC", "")).strip() or "Team"
        targets.append({
            "email":   email,
            "pic":     pic,
            "tier":    selected_tier,
            "intent":  str(row.get("INTENT", "general")),
            "cmd":     campaign_label,
        })

    # --- Pre-send summary ---
    print("\n" + "=" * 60)
    print(f"  TIER SEND REPORT  [{selected_tier}]")
    print("=" * 60)
    print(f"  Ready to send  : {len(targets)}")
    print(f"  Dead skipped   : {skip_dead}")
    print(f"  Cooldown skip  : {skip_cool} (within {cooldown_days} days)")

    if not targets:
        print("  Nothing to send.")
        return

    # --- Show top hot leads first (booking / negotiating) ---
    hot = [t for t in targets if t["intent"] in ("booking_intent", "negotiating")]
    if hot:
        print(f"\n  HOT ({len(hot)} with booking/negotiating intent — send these FIRST):")
        for t in hot[:5]:
            print(f"    {t['intent']:20s} | {t['email']}")

    print("=" * 60)

    if input("\n  Proceed to send? (Y/N): ").upper() != "Y":
        sys.exit("Tier send cancelled.")

    # --- Batch limit ---
    batch_n   = int(input(f"  How many to send (max {len(targets)}): "))
    send_list = targets[:batch_n]

    # --- Tier-specific subject prefix ---
    TIER_PREFIX = {
        "REPLY_3": "[Follow-up] ",
        "REPLY_2": "[Re: Freight Rates] ",
        "REPLY_1": "",
    }
    prefix = TIER_PREFIX.get(selected_tier, "")

    # --- Send loop ---
    outlook_tm = win32com.client.Dispatch("Outlook.Application")
    sent_count = 0

    for t in send_list:
        mail          = outlook_tm.CreateItem(0)
        mail.To       = t["email"]
        mail.Subject  = prefix + generate_subject()

        if PROFILE_PDF.exists():
            mail.Attachments.Add(str(PROFILE_PDF))
        if LOGO_PNG.exists():
            logo = mail.Attachments.Add(str(LOGO_PNG))
            logo.PropertyAccessor.SetProperty(
                "http://schemas.microsoft.com/mapi/proptag/0x3712001F", "pudonglogo"
            )

        # Tier-specific intro override
        if selected_tier == "REPLY_3" and t["intent"] in ("booking_intent", "negotiating"):
            intro = (
                f"Dear {t['pic']},<br><br>"
                "Thank you for your continued interest! We would love to help confirm "
                "your shipment — please let us know if you are ready to proceed or "
                "if you have any final questions.<br><br>"
            )
        elif selected_tier == "REPLY_2":
            intro = (
                f"Dear {t['pic']},<br><br>"
                "We appreciate your response and wanted to follow up with our latest "
                "competitive rates. Please find our updated schedule below.<br><br>"
            )
        else:
            intro = f"Dear {t['pic']},<br><br>{cfg.get('INTROTEXT', '')}<br><br>"

        mail.HTMLBody = f"""
        <html><body>
        {build_preheader_html()}
        {intro}
        {cfg.get('RATETABLEHTML', '')}<br><br>
        {cfg.get('CLOSINGTEXT', '')}<br><br>
        {cfg.get('SIGNATURE', '')}
        </body></html>
        """

        subject = mail.Subject
        mail.Send()
        log_email_send(t["email"], subject, campaign_label)
        log.info("TIER SENT [%s|%s] -> %s", selected_tier, t["intent"], t["email"])
        sent_count += 1

    print("\n" + "=" * 60)
    print(f"  TIER SEND COMPLETE: {sent_count} emails sent [{selected_tier}]")
    print("=" * 60)


# =========================================================
# MODE SELECTOR  —  --tier flag activates tier_mode()
# =========================================================
if "--tier" in sys.argv:
    # Run tier mode and exit; skip all CMD_MODE interactive code below
    tier_mode()
    sys.exit(0)


# =========================================================
# LOAD DATA  (CMD_MODE only below this point)
# =========================================================
shutil.copy(DATA_FILE, BACKUP_DIR / f"data_{datetime.now():%Y%m%d_%H%M}.xlsx")

df = pd.read_excel(DATA_FILE)
df.columns = df.columns.str.strip().str.upper().str.replace(" ", "_")
df["CMD_NAME"] = df["CMD_NAME"].astype(str).str.upper()
df["STATUS"]   = df["STATUS"].astype(str).str.upper()


# =========================================================
# SELECT ROLE
# =========================================================
print("\nSelect email target:")
print("1. CNEE")
print("2. SHIPPER")
print("3. BOTH")

ROLE_MAP   = {"1": ["CNEE"], "2": ["SHIPPER"], "3": ["CNEE", "SHIPPER"]}
SEND_TARGET = ROLE_MAP.get(input("Your choice: ").strip())
if not SEND_TARGET:
    sys.exit("Invalid role")

# =========================================================
# SELECT CMD
# =========================================================
cmds = sorted(df["CMD_NAME"].unique())
for i, c in enumerate(cmds, 1):
    print(f"{i}. {c}")

indexes    = [int(x) for x in input("Select CMD (comma-sep): ").split(",")]
TARGET_CMD = [cmds[i - 1] for i in indexes if 1 <= i <= len(cmds)]
if not TARGET_CMD:
    sys.exit("No CMD selected")


# =========================================================
# CMD GUARD
# =========================================================
hist = pd.read_csv(CMD_HISTORY_FILE, encoding="latin1")
hist["CMD_NAME"]   = hist["CMD_NAME"].astype(str).str.upper()
hist["SENT_DATE"]  = pd.to_datetime(hist["SENT_DATE"], errors="coerce")
hist["SENT_COUNT"] = pd.to_numeric(hist["SENT_COUNT"], errors="coerce").fillna(0)

for cmd in TARGET_CMD:
    total      = df[df["CMD_NAME"] == cmd].shape[0]
    sent_hist  = hist[hist["CMD_NAME"] == cmd]
    sent_before = sent_hist["SENT_COUNT"].sum() if not sent_hist.empty else 0

    print(f"\nCMD: {cmd}")
    print(f"  Total     : {total}")
    print(f"  Sent      : {int(sent_before)}")
    print(f"  Remaining : {max(total - int(sent_before), 0)}")

    if not sent_hist.empty:
        last = sent_hist["SENT_DATE"].max()
        days_ago = (pd.Timestamp.today() - last).days
        print(f"  Last sent : {last.date()} ({days_ago} days ago)")
        if days_ago < 3:
            print("  WARNING: Sent less than 3 days ago (spam risk).")

    if input("  Continue? (Y/N): ").upper() != "Y":
        sys.exit("Cancelled")


# =========================================================
# LOAD KNOWLEDGE & LOG
# =========================================================
knowledge            = load_knowledge()
df_log, recent_pairs = load_email_log()


# =========================================================
# BUILD TARGET LIST — with knowledge-aware filtering
# =========================================================
targets        = []
skip_dead      = []  # hard bounce, policy, spam
skip_duplicate = []  # already sent this campaign
warn_replied   = []  # human replied — still send but flag
warn_ooo       = []  # auto reply — still send but flag
seen_emails_per_cmd: dict[str, set] = {}  # dedup per (email, cmd) pair

for idx, row in df[df["CMD_NAME"].isin(TARGET_CMD)].iterrows():
    for role in SEND_TARGET:
        email_col = f"{role}_EMAIL"
        pic_col   = f"{role}_PIC"

        email = row.get(email_col)
        if not isinstance(email, str) or "@" not in email:
            continue

        email_l = email.strip().lower()
        cmd     = row["CMD_NAME"]

        # --- Knowledge check ---
        kb = knowledge.get(email_l, {})
        kb_status = kb.get("status", "")

        if kb_status in DEAD_STATUSES:
            skip_dead.append((email_l, cmd, kb_status, kb.get("remark", "")))
            continue

        # --- Cooldown check: skip if sent to same CMD within RESEND_COOLDOWN_DAYS ---
        if (email_l, cmd) in recent_pairs:
            last_sent = recent_pairs[(email_l, cmd)]
            days_ago  = (pd.Timestamp.now() - last_sent).days
            skip_duplicate.append((email_l, cmd, days_ago))
            continue

        # --- Dedup: same email appearing in multiple rows of same CMD ---
        seen_key = (email_l, cmd)
        if seen_key in seen_emails_per_cmd.get(cmd, set()):
            log.info("[DEDUP] Skipping duplicate email in same CMD: %s", email_l)
            continue
        seen_emails_per_cmd.setdefault(cmd, set()).add(email_l)

        # --- Warm / replied flag ---
        if kb_status in WARM_STATUSES:
            warn_replied.append((email_l, cmd, kb_status))
        elif kb_status in AUTO_REPLY_STATUSES:
            replacement = kb.get("replacement_email", "")
            warn_ooo.append((email_l, cmd, replacement))

        pic = row.get(pic_col)
        targets.append({
            "row":     idx,
            "email":   email,
            "pic":     pic,
            "cmd":     cmd,
            "kb_status": kb_status,
            "replacement": kb.get("replacement_email", ""),
        })

if not targets:
    print()
    print("=" * 60)
    print("  No email to send — filter summary:")
    print(f"  Dead (bounce/policy/spam) : {len(skip_dead)}")
    print(f"  Cooldown ({RESEND_COOLDOWN_DAYS}d, already sent)  : {len(skip_duplicate)}")
    print(f"  No valid email in row     : sum of skipped rows")
    print()
    print(f"  Tip: RESEND_COOLDOWN_DAYS is currently {RESEND_COOLDOWN_DAYS} days.")
    print(f"  Emails sent more than {RESEND_COOLDOWN_DAYS} days ago CAN be resent.")
    print("  If all emails were sent in the last 7 days, wait or set cooldown lower.")
    print("=" * 60)
    sys.exit(1)

# =========================================================
# PRE-SEND INTELLIGENCE SUMMARY
# =========================================================
print("\n" + "=" * 60)
print("  PRE-SEND INTELLIGENCE REPORT")
print("=" * 60)
print(f"  Ready to send     : {len(targets)}")

if skip_dead:
    print(f"\n  [BLOCKED — dead] : {len(skip_dead)} emails skipped")
    for e, cmd, status, remark in skip_dead[:10]:
        print(f"    {status:15s} | {e:40s} | {remark[:40]}")
    if len(skip_dead) > 10:
        print(f"    ... and {len(skip_dead)-10} more")

if skip_duplicate:
    print(f"\n  [SKIP — sent]    : {len(skip_duplicate)} already sent this campaign")

if warn_replied:
    print(f"\n  [WARN — replied] : {len(warn_replied)} emails — customer has replied before")
    for e, cmd, st in warn_replied[:5]:
        print(f"    {st:15s} | {e}")
    if len(warn_replied) > 5:
        print(f"    ... and {len(warn_replied)-5} more")

if warn_ooo:
    print(f"\n  [INFO — OOO]     : {len(warn_ooo)} auto-reply contacts")
    for e, cmd, rep in warn_ooo[:5]:
        rep_str = f" -> replacement: {rep}" if rep else ""
        print(f"    auto_reply     | {e}{rep_str}")
    if len(warn_ooo) > 5:
        print(f"    ... and {len(warn_ooo)-5} more")

print("=" * 60)

if warn_replied:
    if input("\n  Some customers already replied — skip them? (Y/N): ").upper() == "Y":
        replied_emails = {e for e, _, _ in warn_replied}
        targets = [t for t in targets if t["email"].lower() not in replied_emails]
        print(f"  Removed {len(warn_replied)} replied customers. New total: {len(targets)}")

# =========================================================
# BATCH SIZE
# =========================================================
batch_count  = int(input(f"\nTotal {len(targets)} emails. How many batches? "))
batch_targets = targets[: batch_count * BATCH_SIZE]

if not batch_targets:
    sys.exit("Batch size is 0. Cancelled.")

# =========================================================
# OUTLOOK PREVIEW
# =========================================================
outlook = win32com.client.Dispatch("Outlook.Application")

t0  = batch_targets[0]
pic = t0["pic"] if isinstance(t0["pic"], str) and t0["pic"].strip() else "Team"

mail         = outlook.CreateItem(0)
mail.To      = t0["email"]
mail.Subject = generate_subject()

if PROFILE_PDF.exists():
    mail.Attachments.Add(str(PROFILE_PDF))

if LOGO_PNG.exists():
    logo = mail.Attachments.Add(str(LOGO_PNG))
    logo.PropertyAccessor.SetProperty(
        "http://schemas.microsoft.com/mapi/proptag/0x3712001F", "pudonglogo"
    )

mail.HTMLBody = f"""
<html><body>
{build_preheader_html()}
Dear {pic},<br><br>
{cfg.get("INTROTEXT","")}<br><br>
{cfg.get("RATETABLEHTML","")}<br><br>
{cfg.get("CLOSINGTEXT","")}<br><br>
{cfg.get("SIGNATURE","")}
</body></html>
"""
mail.Display()

if input("\nConfirm send? (Y/N): ").upper() != "Y":
    sys.exit("Cancelled")

# =========================================================
# SEND LOOP
# =========================================================
sent_counter: dict[str, int] = {}

for t in batch_targets:
    mail         = outlook.CreateItem(0)
    mail.To      = t["email"]
    mail.Subject = generate_subject()

    if PROFILE_PDF.exists():
        mail.Attachments.Add(str(PROFILE_PDF))

    if LOGO_PNG.exists():
        logo = mail.Attachments.Add(str(LOGO_PNG))
        logo.PropertyAccessor.SetProperty(
            "http://schemas.microsoft.com/mapi/proptag/0x3712001F", "pudonglogo"
        )

    pic = t["pic"] if isinstance(t["pic"], str) and t["pic"].strip() else "Team"
    mail.HTMLBody = f"""
    <html><body>
    {build_preheader_html()}
    Dear {pic},<br><br>
    {cfg.get("INTROTEXT","")}<br><br>
    {cfg.get("RATETABLEHTML","")}<br><br>
    {cfg.get("CLOSINGTEXT","")}<br><br>
    {cfg.get("SIGNATURE","")}
    </body></html>
    """

    subject = mail.Subject
    mail.Send()

    log_email_send(t["email"], subject, t["cmd"])
    df.loc[t["row"], "STATUS"] = "SENT"
    sent_counter[t["cmd"]] = sent_counter.get(t["cmd"], 0) + 1

    kb_tag = f" [{t['kb_status']}]" if t["kb_status"] else ""
    log.info("SENT%s -> %s", kb_tag, t["email"])

# =========================================================
# LOG CMD HISTORY
# =========================================================
with open(CMD_HISTORY_FILE, "a", newline="", encoding="latin1") as f:
    writer = csv.writer(f)
    for cmd, count in sent_counter.items():
        writer.writerow([cmd, datetime.now().date(), count])

df.to_excel(DATA_FILE, index=False)

print("\n" + "=" * 60)
print("  SEND COMPLETE")
print("=" * 60)
for cmd, cnt in sent_counter.items():
    print(f"  {cmd:20s} : {cnt} emails sent")
total_sent = sum(sent_counter.values())
total_dead = len(skip_dead)
print(f"\n  Total sent    : {total_sent}")
print(f"  Dead skipped  : {total_dead}")
print(f"  Dedup skipped : {len(skip_duplicate)}")
print("=" * 60)
