"""
Nelson Briefing — Daily Excel Dashboard
========================================
Generates logs/nelson_briefing_YYYYMMDD.xlsx with 5 sheets:
  1. TODAY_ALERTS     — Unresolved alerts
  2. SALES_HOT        — Actionable sales replies
  3. ACTIVE_SHIPMENTS — Open shipments
  4. MENTEE_PERFORMANCE — 7-day team scorecard
  5. NEEDS_REVIEW     — Low-confidence parses
"""

from __future__ import annotations

import logging
import logging.handlers
import sqlite3
import sys
from datetime import datetime, timedelta
from pathlib import Path

_repo_root = str(Path(__file__).parent.parent.parent)
if _repo_root not in sys.path:
    sys.path.insert(0, _repo_root)
from shared import paths as sp

log = logging.getLogger(__name__)

DB_PATH    = sp.EMAIL_LOG_DIR / 'shipments.db'
OUTPUT_DIR = sp.EMAIL_LOG_DIR


def generate():
    """Generate the daily briefing Excel file."""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("Required packages missing. Run: pip install pandas openpyxl")
        return None

    if not DB_PATH.exists():
        log.error("Database not found at %s — run data_collector.py first.", DB_PATH)
        return None

    from shared.db_connect import get_db
    conn = get_db(DB_PATH, readonly=True)
    today_str = datetime.now().strftime('%Y-%m-%d')
    week_ago  = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')

    # ------------------------------------------------------------------
    # Styling
    # ------------------------------------------------------------------
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    critical_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
    high_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    hot_fill  = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    wb = Workbook()

    def style_sheet(ws, df, color_rules=None):
        """Apply formatting to a worksheet."""
        # Headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        # Data rows
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Color rules
            if color_rules:
                row_dict = dict(zip(df.columns, row))
                for rule_col, rule_val, fill, font in color_rules:
                    if rule_col in row_dict and row_dict[rule_col] == rule_val:
                        for col_idx in range(1, len(df.columns) + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = fill
                            if font:
                                ws.cell(row=row_idx, column=col_idx).font = font
                        break

        # Auto-fit column widths
        for col_idx in range(1, len(df.columns) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value or ''))
                for r in range(1, min(ws.max_row + 1, 50))
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)

        # Freeze top row
        ws.freeze_panes = 'A2'

    # ------------------------------------------------------------------
    # Sheet 1: TODAY_ALERTS
    # ------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "TODAY_ALERTS"
    df1 = pd.read_sql(f"""
        SELECT
            created_at     AS TIME,
            alert_type     AS TYPE,
            risk_level     AS URGENCY,
            customer_name  AS CUSTOMER,
            member_owner   AS MEMBER,
            subject_raw    AS SUBJECT,
            alert_reason   AS REASON,
            primary_stage  AS STAGE
        FROM nelson_alerts
        WHERE date(created_at) >= date('{today_str}')
          AND is_resolved = 0
        ORDER BY
            CASE risk_level
                WHEN 'CRITICAL' THEN 1
                WHEN 'HIGH' THEN 2
                WHEN 'MEDIUM' THEN 3
                ELSE 4
            END,
            created_at DESC
    """, conn)
    style_sheet(ws1, df1, color_rules=[
        ('URGENCY', 'CRITICAL', critical_fill, white_font),
        ('URGENCY', 'HIGH', high_fill, white_font),
    ])

    # ------------------------------------------------------------------
    # Sheet 2: SALES_HOT
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("SALES_HOT")
    df2 = pd.read_sql(f"""
        SELECT
            customer_name   AS CUSTOMER,
            intent          AS INTENT,
            campaign_week   AS CAMPAIGN,
            campaign_type   AS TYPE,
            next_action     AS NEXT_ACTION,
            urgency         AS URGENCY,
            panjiva_vol_month AS PANJIVA_VOL,
            panjiva_carrier AS PANJIVA_CARRIER,
            panjiva_route   AS PANJIVA_ROUTE,
            subject_raw     AS SUBJECT,
            received_at     AS RECEIVED_AT
        FROM sales_replies
        WHERE intent IN ('HOT','WARM','PRICE_FIGHT')
          AND is_actioned = 0
        ORDER BY
            CASE urgency
                WHEN 'IMMEDIATE' THEN 1
                WHEN 'TODAY' THEN 2
                WHEN 'THIS_WEEK' THEN 3
                ELSE 4
            END,
            received_at DESC
    """, conn)
    style_sheet(ws2, df2, color_rules=[
        ('INTENT', 'HOT', hot_fill, white_font),
    ])

    # ------------------------------------------------------------------
    # Sheet 3: ACTIVE_SHIPMENTS
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("ACTIVE_SHIPMENTS")
    df3 = pd.read_sql("""
        SELECT
            shipment_key    AS SHIPMENT_KEY,
            hbl             AS HBL,
            bkg             AS BKG,
            customer_name   AS CUSTOMER,
            member_owner    AS MEMBER,
            current_stage   AS CURRENT_STAGE,
            missing_stages  AS MISSING_STAGES,
            risk_level      AS RISK_LEVEL,
            days_open       AS DAYS_OPEN,
            etd             AS ETD,
            carrier         AS CARRIER
        FROM shipments
        WHERE is_complete = 0
        ORDER BY
            CASE risk_level
                WHEN 'CRITICAL' THEN 1
                WHEN 'HIGH' THEN 2
                WHEN 'MEDIUM' THEN 3
                ELSE 4
            END,
            days_open DESC
    """, conn)
    style_sheet(ws3, df3, color_rules=[
        ('RISK_LEVEL', 'CRITICAL', critical_fill, white_font),
        ('RISK_LEVEL', 'HIGH', high_fill, white_font),
    ])

    # ------------------------------------------------------------------
    # Sheet 4: MENTEE_PERFORMANCE (last 7 days)
    # ------------------------------------------------------------------
    ws4 = wb.create_sheet("MENTEE_PERFORMANCE")
    df4 = pd.read_sql(f"""
        SELECT
            member_owner AS MEMBER,
            COUNT(*)     AS EMAILS_PROCESSED,
            SUM(CASE WHEN risk_level IN ('CRITICAL','HIGH') THEN 1 ELSE 0 END) AS RISK_EVENTS,
            COUNT(DISTINCT shipment_key) AS SHIPMENTS_ACTIVE
        FROM email_events
        WHERE date(received_at) >= date('{week_ago}')
          AND member_owner IS NOT NULL
          AND member_owner != 'NELSON'
        GROUP BY member_owner
        ORDER BY EMAILS_PROCESSED DESC
    """, conn)
    # Calculate score
    if not df4.empty:
        df4['SCORE'] = 100 - (df4['RISK_EVENTS'] * 5)
        df4['SCORE'] = df4['SCORE'].clip(lower=0)
    style_sheet(ws4, df4)

    # ------------------------------------------------------------------
    # Sheet 5: NEEDS_REVIEW
    # ------------------------------------------------------------------
    ws5 = wb.create_sheet("NEEDS_REVIEW")
    df5 = pd.read_sql(f"""
        SELECT
            msg_filename      AS MSG_FILE,
            subject_raw       AS SUBJECT,
            folder_context    AS FOLDER,
            parse_confidence  AS CONFIDENCE,
            processed_at      AS PROCESSED_AT
        FROM email_events
        WHERE needs_review = 1
          AND date(processed_at) >= date('{week_ago}')
        ORDER BY processed_at DESC
    """, conn)
    style_sheet(ws5, df5)

    conn.close()

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------
    today_file = datetime.now().strftime('%Y%m%d')
    out_path = OUTPUT_DIR / f"nelson_briefing_{today_file}.xlsx"
    wb.save(str(out_path))
    log.info("Briefing saved → %s", out_path)
    print(f"\n  ✅ Briefing generated: {out_path}")
    return out_path


# ======================================================================
# CLI
# ======================================================================
if __name__ == '__main__':
    logging.basicConfig(
        level=logging.INFO,
        format="[%(asctime)s] %(levelname)-8s %(message)s",
        datefmt="%H:%M:%S",
    )
    generate()
