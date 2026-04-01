"""
follow_up_engine.py — Tier Follow-up Alert Engine  v1.0
=========================================================
Reads customer_final.xlsx (produced by process_reply.py) and applies
rules-based follow-up alerting for quote-campaign prospects only.

Rules
-----
  Tier 3, intent=booking_intent  → stale > 2 days  → 🔴 URGENT
  Tier 3, intent=price_inquiry   → stale > 3 days  → 🟠 HIGH
  Tier 3, any intent             → stale > 3 days  → 🟠 HIGH
  Tier 2, any intent             → stale > 5 days  → 🟡 MEDIUM
  Tier 1, any intent             → stale > 10 days → 🔵 LOW

"Stale" is measured from the most recent email sent to this address
in email_log.csv.

Outputs
-------
  1. Sheet "FOLLOW_UP" in customer_final.xlsx (refreshed in-place)
  2. Windows toast notification summarising alert counts
  3. logs/followup_alerts.csv (appended each run)

Usage
-----
  python follow_up_engine.py            # standalone
  # OR called from run_all.py pipeline
"""

from __future__ import annotations

import csv
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

BASE_DIR      = Path(__file__).parent
PROJECT_ROOT = BASE_DIR.parent
LOG_DIR       = PROJECT_ROOT / "logs"
FINAL_FILE    = PROJECT_ROOT / "data" / "customer_final.xlsx"
EMAIL_LOG_FILE = PROJECT_ROOT / "logs" / "email_log.csv"
ALERT_LOG_FILE = LOG_DIR / "followup_alerts.csv"

LOG_DIR.mkdir(exist_ok=True)

# =========================================================
# LOGGING
# =========================================================
logging.basicConfig(
    level   = logging.INFO,
    format  = "[%(asctime)s] %(levelname)-8s %(message)s",
    datefmt = "%H:%M:%S",
    handlers= [logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# =========================================================
# FOLLOW-UP RULES
# priority: lower number = shown higher in the list
# =========================================================
RULES: list[dict] = [
    {"tier": "REPLY_3", "intent": "booking_intent", "stale_days": 2,
     "priority": 1, "label": "🔴 URGENT — Chốt booking"},
    {"tier": "REPLY_3", "intent": "negotiating",    "stale_days": 2,
     "priority": 1, "label": "🔴 URGENT — Đang đàm phán"},
    {"tier": "REPLY_3", "intent": "price_inquiry",  "stale_days": 3,
     "priority": 2, "label": "🟠 HIGH — Hỏi giá chưa trả lời"},
    {"tier": "REPLY_3", "intent": None,             "stale_days": 3,
     "priority": 2, "label": "🟠 HIGH — Tier 3 chưa follow"},
    {"tier": "REPLY_2", "intent": None,             "stale_days": 5,
     "priority": 3, "label": "🟡 MEDIUM — Tier 2 chưa follow"},
    {"tier": "REPLY_1", "intent": None,             "stale_days": 10,
     "priority": 4, "label": "🔵 LOW — Tier 1 chưa follow"},
]


# =========================================================
# LOAD LAST SENT DATE PER EMAIL
# =========================================================
def load_last_sent() -> dict[str, datetime]:
    """Return dict: email_lower → most recent sent datetime."""
    if not EMAIL_LOG_FILE.exists():
        return {}

    df = pd.read_csv(EMAIL_LOG_FILE)
    df.columns = df.columns.str.lower()
    df["email"]     = df["email"].astype(str).str.lower().str.strip()
    df["timestamp"] = pd.to_datetime(df.get("timestamp", pd.Series(dtype="str")),
                                     errors="coerce")
    df = df.dropna(subset=["timestamp"])

    last_sent: dict[str, datetime] = {}
    for _, row in df.iterrows():
        e  = row["email"]
        ts = row["timestamp"].to_pydatetime()
        if e not in last_sent or ts > last_sent[e]:
            last_sent[e] = ts
    return last_sent


# =========================================================
# RULE MATCHER
# =========================================================
def match_rule(tier: str, intent: str) -> dict | None:
    """Return highest-priority rule matching this tier/intent combo."""
    best: dict | None = None
    for rule in RULES:
        if rule["tier"] != tier:
            continue
        if rule["intent"] is not None and rule["intent"] != intent:
            continue
        if best is None or rule["priority"] < best["priority"]:
            best = rule
    return best


# =========================================================
# COMPUTE ALERTS
# =========================================================
def compute_alerts(df_hot: pd.DataFrame, last_sent: dict) -> pd.DataFrame:
    """
    Given the REPLY_2 + REPLY_3 rows from customer_final.xlsx,
    return a DataFrame of rows that need follow-up.
    """
    now     = datetime.now()
    alerts  = []

    for _, row in df_hot.iterrows():
        tier   = str(row.get("REPLY_TIER", "")).strip()
        intent = str(row.get("INTENT", "general")).strip()
        rule   = match_rule(tier, intent)
        if rule is None:
            continue

        # Find last sent date for this row's primary email
        email = str(row.get("CNEE_EMAIL", "")).lower().strip()
        if "@" not in email:
            email = str(row.get("SHIPPER_EMAIL", "")).lower().strip()
        if "@" not in email:
            continue

        last = last_sent.get(email)
        if last is None:
            days_stale = 999  # never sent — always alert
        else:
            days_stale = (now - last).days

        if days_stale < rule["stale_days"]:
            continue  # not yet stale

        alert_row = row.to_dict()
        alert_row["ALERT_LABEL"]     = rule["label"]
        alert_row["ALERT_PRIORITY"]  = rule["priority"]
        alert_row["DAYS_STALE"]      = days_stale
        alert_row["LAST_SENT"]       = last.strftime("%Y-%m-%d") if last else "Never"
        alerts.append(alert_row)

    if not alerts:
        return pd.DataFrame()

    df_alerts = pd.DataFrame(alerts)
    df_alerts  = df_alerts.sort_values(
        ["ALERT_PRIORITY", "DAYS_STALE"], ascending=[True, False]
    )
    return df_alerts


# =========================================================
# WRITE FOLLOW_UP SHEET  (replace only this sheet in-place)
# =========================================================
def write_follow_up_sheet(df_alerts: pd.DataFrame) -> None:
    """
    Re-write only the FOLLOW_UP sheet in customer_final.xlsx,
    preserving all other sheets unchanged.
    Uses openpyxl for surgical sheet replacement.
    """
    if not FINAL_FILE.exists():
        log.warning("customer_final.xlsx not found — skipping sheet write.")
        return

    try:
        wb = load_workbook(FINAL_FILE)
        # Remove old FOLLOW_UP sheet if present
        if "FOLLOW_UP" in wb.sheetnames:
            del wb["FOLLOW_UP"]

        ws = wb.create_sheet("FOLLOW_UP")

        # Front columns to show first
        front_cols = [
            "ALERT_LABEL", "ALERT_PRIORITY", "DAYS_STALE", "LAST_SENT",
            "CNEE_NAME", "CNEE_EMAIL", "CAMPAIGN_ID",
            "REPLY_TIER", "INTENT", "REPLY_COUNT",
        ]
        if df_alerts.empty:
            # Write header-only sheet
            for col_idx, col_name in enumerate(front_cols, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
        else:
            ordered = [c for c in front_cols if c in df_alerts.columns]
            others  = [c for c in df_alerts.columns if c not in ordered]
            final_cols = ordered + others

            # Header
            for col_idx, col_name in enumerate(final_cols, 1):
                ws.cell(row=1, column=col_idx, value=col_name)

            # Data rows
            for row_idx, row in enumerate(df_alerts[final_cols].itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(FINAL_FILE)
        log.info("FOLLOW_UP sheet written: %d alerts", len(df_alerts))
    except Exception as exc:
        log.error("Failed to write FOLLOW_UP sheet: %s", exc)


# =========================================================
# WRITE ALERT LOG
# =========================================================
def append_alert_log(df_alerts: pd.DataFrame) -> None:
    if df_alerts.empty:
        return
    file_exists = ALERT_LOG_FILE.exists()
    with open(ALERT_LOG_FILE, "a", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        if not file_exists:
            writer.writerow([
                "scan_date", "email", "campaign_id",
                "tier", "intent", "alert_label",
                "days_stale", "last_sent",
            ])
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
        for _, row in df_alerts.iterrows():
            writer.writerow([
                now_str,
                row.get("CNEE_EMAIL", ""),
                row.get("CAMPAIGN_ID", ""),
                row.get("REPLY_TIER", ""),
                row.get("INTENT", ""),
                row.get("ALERT_LABEL", ""),
                row.get("DAYS_STALE", ""),
                row.get("LAST_SENT", ""),
            ])


# =========================================================
# TOAST NOTIFICATION SUMMARY
# =========================================================
def send_toast_summary(df_alerts: pd.DataFrame) -> None:
    try:
        from notify import toast
    except ImportError:
        log.debug("notify.py not found — skipping toast.")
        return

    if df_alerts.empty:
        toast("📊 Follow-up Engine", "Không có khách nào cần follow-up hôm nay.", kind="none")
        return

    urgent = len(df_alerts[df_alerts["ALERT_PRIORITY"] == 1])
    high   = len(df_alerts[df_alerts["ALERT_PRIORITY"] == 2])
    medium = len(df_alerts[df_alerts["ALERT_PRIORITY"] == 3])
    low    = len(df_alerts[df_alerts["ALERT_PRIORITY"] == 4])

    parts = []
    if urgent: parts.append(f"🔴 {urgent} URGENT")
    if high:   parts.append(f"🟠 {high} HIGH")
    if medium: parts.append(f"🟡 {medium} MEDIUM")
    if low:    parts.append(f"🔵 {low} LOW")

    kind = "error" if urgent else ("warning" if high else "info")
    toast(
        "📊 Email Engine — Follow-up Alerts",
        " | ".join(parts),
        kind=kind,
        duration_ms=8000,
    )


# =========================================================
# MAIN
# =========================================================
def main() -> None:
    log.info("=" * 60)
    log.info("  FOLLOW-UP ALERT ENGINE  v1.0")
    log.info("=" * 60)

    # 1. Load customer_final.xlsx
    if not FINAL_FILE.exists():
        log.error("customer_final.xlsx not found. Run process_reply.py first.")
        return

    # Try FOLLOW_UP sheet first; fall back to REPLY_2 + REPLY_3 sheets
    xls_sheets = pd.ExcelFile(FINAL_FILE).sheet_names

    if "FOLLOW_UP" in xls_sheets:
        df_hot = pd.read_excel(FINAL_FILE, sheet_name="FOLLOW_UP")
        df_hot.columns = df_hot.columns.str.strip().str.upper()
        log.info("Loaded FOLLOW_UP sheet: %d rows", len(df_hot))
    else:
        # Fallback: merge REPLY_2 + REPLY_3 sheets (pre-upgrade customer_final.xlsx)
        log.warning("FOLLOW_UP sheet not found — falling back to REPLY_2 + REPLY_3 sheets.")
        frames = []
        for sheet in ["REPLY_3", "REPLY_2", "REPLY_1"]:
            if sheet in xls_sheets:
                df_s = pd.read_excel(FINAL_FILE, sheet_name=sheet)
                df_s.columns = df_s.columns.str.strip().str.upper()
                # Synthesise missing columns that process_reply v3 would have added
                if "REPLY_TIER" not in df_s.columns:
                    df_s["REPLY_TIER"] = sheet
                if "INTENT" not in df_s.columns:
                    df_s["INTENT"] = "general"
                if "REPLY_COUNT" not in df_s.columns:
                    df_s["REPLY_COUNT"] = {"REPLY_3": 3, "REPLY_2": 2, "REPLY_1": 1}.get(sheet, 0)
                if "CAMPAIGN_ID" not in df_s.columns:
                    df_s["CAMPAIGN_ID"] = ""
                frames.append(df_s)
        if not frames:
            log.error("No REPLY sheets found in customer_final.xlsx. Run process_reply.py first.")
            return
        df_hot = pd.concat(frames, ignore_index=True)
        log.info("Fallback: loaded %d rows from REPLY sheets", len(df_hot))

    if df_hot.empty:
        log.info("No hot prospects found — nothing to alert.")
        send_toast_summary(pd.DataFrame())
        return

    log.info("Hot prospects loaded: %d rows (Tier 1 + 2 + 3)", len(df_hot))

    # 2. Load last sent dates from email_log
    last_sent = load_last_sent()
    log.info("Email log: last-sent dates for %d unique addresses", len(last_sent))

    # 3. Compute which rows are stale and need alerting
    df_alerts = compute_alerts(df_hot, last_sent)

    # 4. Write updated FOLLOW_UP sheet with stale info
    write_follow_up_sheet(df_alerts if not df_alerts.empty else df_hot)

    # 5. Append to alert log CSV
    append_alert_log(df_alerts)

    # 6. Toast notification
    send_toast_summary(df_alerts)

    # 7. Summary
    log.info("")
    log.info("=" * 60)
    log.info("  ALERT SUMMARY")
    log.info("=" * 60)
    if df_alerts.empty:
        log.info("  No stale prospects — all follow-ups are on track.")
    else:
        for priority, label in [(1, "URGENT"), (2, "HIGH"), (3, "MEDIUM"), (4, "LOW")]:
            count = len(df_alerts[df_alerts["ALERT_PRIORITY"] == priority])
            if count:
                log.info("  %-8s : %d prospects", label, count)
    log.info("  Alert log: %s", ALERT_LOG_FILE)
    log.info("=" * 60)


if __name__ == "__main__":
    main()
