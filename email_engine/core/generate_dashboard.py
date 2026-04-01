"""
generate_dashboard.py — Email Master Dashboard Generator
=========================================================
Reads (never writes) from:
  - data.xlsx           → master contact list
  - logs/email_log.csv  → send history
  - logs/email_knowledge.csv → bounce / reply / OOO knowledge

Produces:
  - email_master.xlsx   → 7-sheet formatted dashboard

Sheets:
  OVERVIEW    — key metrics summary
  SEND_QUEUE  — ready-to-send contacts (not dead, past cooldown)
  SENT        — full send history
  REPLIED     — customers who replied
  BOUNCED     — dead emails (hard_bounce / policy_reject / spam_block)
  AUTO_REPLY  — OOO contacts with replacement info
  MASTER      — full contact list with KB_STATUS annotation

Usage:
  python generate_dashboard.py
"""

import re
import sys
import logging
from datetime import datetime, date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================================================
# CONFIG
# =========================================================
RESEND_COOLDOWN_DAYS = 7   # must match send_email.py setting

BASE_DIR       = Path(__file__).parent
PROJECT_ROOT = BASE_DIR.parent
LOG_DIR        = PROJECT_ROOT / "logs"
DATA_FILE      = PROJECT_ROOT / "data.xlsx"
EMAIL_LOG_FILE = PROJECT_ROOT / "logs"  / "email_log.csv"
KNOWLEDGE_FILE = PROJECT_ROOT / "logs"  / "email_knowledge.csv"
OUTPUT_FILE    = PROJECT_ROOT / "logs" / "email_master.xlsx"

DEAD_STATUSES     = {"hard_bounce", "policy_reject", "spam_block", "invalid"}
AUTO_STATUSES     = {"auto_reply"}
WARM_STATUSES     = {"human_reply", "soft_bounce"}

# =========================================================
# LOGGING
# =========================================================
logging.basicConfig(
    level   = logging.INFO,
    format  = "[%(asctime)s] %(levelname)-8s %(message)s",
    datefmt = "%H:%M:%S",
    handlers= [logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# =========================================================
# COLOR PALETTE
# =========================================================
# Headers
C_HEADER_BG    = "1F3864"   # dark navy
C_HEADER_FG    = "FFFFFF"   # white

# Row alternates
C_ALT_ROW      = "EDF2F7"   # light blue-grey

# Sheet accent colors (row highlight)
C_BOUNCE_ROW   = "FFE0E0"   # soft red
C_REPLY_ROW    = "E0FFE4"   # soft green
C_OOO_ROW      = "FFF9C4"   # soft yellow
C_OVERVIEW_VAL = "E8F0FE"   # light blue for value cells

# Tab colors
TAB_COLORS = {
    "OVERVIEW":   "2196F3",
    "SEND_QUEUE": "4CAF50",
    "SENT":       "607D8B",
    "REPLIED":    "00897B",
    "BOUNCED":    "E53935",
    "AUTO_REPLY": "FB8C00",
    "MASTER":     "5C6BC0",
}

FONT_NAME = "Calibri"

# =========================================================
# STYLE HELPERS
# =========================================================
def hdr_style():
    return {
        "font":      Font(name=FONT_NAME, bold=True, color=C_HEADER_FG, size=10),
        "fill":      PatternFill("solid", fgColor=C_HEADER_BG),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }

def cell_style(bold=False, color=None, bg=None, wrap=False, size=10, align="left"):
    f = Font(name=FONT_NAME, bold=bold, color=color or "000000", size=size)
    fil = PatternFill("solid", fgColor=bg) if bg else PatternFill()
    aln = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return {"font": f, "fill": fil, "alignment": aln}

def thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header(ws, columns: list[str]):
    ws.append(columns)
    for cell in ws[1]:
        for k, v in hdr_style().items():
            setattr(cell, k, v)
        cell.border = thin_border()
    ws.row_dimensions[1].height = 30

def apply_rows(ws, df: pd.DataFrame, accent_col: str | None = None,
               accent_color: str | None = None):
    """Write dataframe rows with alternating row color and optional accent column."""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        ws.append(row)
        bg = C_ALT_ROW if r_idx % 2 == 0 else None
        for c_idx, cell in enumerate(ws[r_idx], start=1):
            col_name = df.columns[c_idx - 1] if c_idx <= len(df.columns) else ""
            cell_bg = bg
            if accent_col and col_name == accent_col and accent_color:
                cell_bg = accent_color
            for k, v in cell_style(bg=cell_bg).items():
                setattr(cell, k, v)
            cell.border = thin_border()

def freeze_and_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

def auto_col_width(ws, min_w=8, max_w=50):
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value or "")) for c in col_cells),
            default=min_w,
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_w), max_w)

# =========================================================
# DATA LOADERS
# =========================================================
def load_data() -> pd.DataFrame:
    df = pd.read_excel(DATA_FILE)
    df.columns = df.columns.str.strip().str.upper().str.replace(" ", "_")
    for col in ["CNEE_EMAIL", "SHIPPER_EMAIL"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.lower().str.strip().replace("nan", "")
    df["CMD_NAME"] = df["CMD_NAME"].astype(str).str.upper().str.strip()
    return df

def load_log() -> pd.DataFrame:
    if not EMAIL_LOG_FILE.exists():
        return pd.DataFrame(columns=["timestamp", "email", "subject",
                                     "campaign_id", "cycle_id", "status"])
    df = pd.read_csv(EMAIL_LOG_FILE)
    df.columns = df.columns.str.lower().str.strip()
    df["email"]       = df["email"].astype(str).str.lower().str.strip()
    df["campaign_id"] = df["campaign_id"].astype(str).str.upper().str.strip()
    df["timestamp"]   = pd.to_datetime(df["timestamp"], errors="coerce")
    return df

def load_knowledge() -> pd.DataFrame:
    if not KNOWLEDGE_FILE.exists():
        return pd.DataFrame()
    df = pd.read_csv(KNOWLEDGE_FILE, encoding="utf-8-sig")
    df.columns = df.columns.str.upper().str.strip()
    df["EMAIL"] = df["EMAIL"].astype(str).str.lower().str.strip()
    return df

def _parse_name_from_remark(remark: str) -> str:
    m = re.search(r"alt contact:\s*([^|\n]+)", str(remark), re.I)
    return m.group(1).strip() if m else ""


# =========================================================
# BUILD DATASET
# =========================================================
def build_datasets() -> dict:
    log.info("Loading source files...")
    df_data = load_data()
    df_log  = load_log()
    df_kb   = load_knowledge()

    # ---- Knowledge lookup dict ----
    kb_dict: dict[str, dict] = {}
    if not df_kb.empty:
        for _, row in df_kb.iterrows():
            email = row["EMAIL"]
            if email and "@" in email:
                kb_dict[email] = {
                    "status":      str(row.get("STATUS", "")).lower().strip(),
                    "replacement": str(row.get("REPLACEMENT_EMAIL", "")).strip(),
                    "role_hint":   str(row.get("ROLE_HINT", "")).strip(),
                    "remark":      str(row.get("REMARK", "")).strip(),
                    "count":       int(row.get("COUNT", 1)),
                }

    # ---- Last sent date per (email, campaign) ----
    last_sent: dict[str, pd.Timestamp] = {}
    if not df_log.empty:
        for _, row in df_log.iterrows():
            e = row["email"]
            t = row["timestamp"]
            if pd.notna(t) and e:
                if e not in last_sent or t > last_sent[e]:
                    last_sent[e] = t

    today     = pd.Timestamp.now()
    cutoff    = today - pd.Timedelta(days=RESEND_COOLDOWN_DAYS)

    # ---- Annotate MASTER ----
    def get_kb_status(row) -> str:
        for col in ["CNEE_EMAIL", "SHIPPER_EMAIL"]:
            e = row.get(col, "")
            if isinstance(e, str) and "@" in e and e in kb_dict:
                return kb_dict[e]["status"]
        return ""

    def get_last_sent(row) -> str:
        for col in ["CNEE_EMAIL", "SHIPPER_EMAIL"]:
            e = row.get(col, "")
            if isinstance(e, str) and "@" in e and e in last_sent:
                return last_sent[e].strftime("%Y-%m-%d")
        return ""

    def get_days_since(row) -> str:
        for col in ["CNEE_EMAIL", "SHIPPER_EMAIL"]:
            e = row.get(col, "")
            if isinstance(e, str) and "@" in e and e in last_sent:
                return str((today - last_sent[e]).days)
        return "Never"

    df_master = df_data.copy()
    df_master["KB_STATUS"]      = df_master.apply(get_kb_status, axis=1)
    df_master["LAST_SENT_DATE"] = df_master.apply(get_last_sent,  axis=1)
    df_master["DAYS_SINCE_SENT"]= df_master.apply(get_days_since, axis=1)

    # ---- SEND_QUEUE ----
    def is_sendable(row) -> bool:
        kb = row["KB_STATUS"].lower()
        if kb in DEAD_STATUSES:
            return False
        # Must have at least one valid email
        has_email = any(
            isinstance(row.get(c, ""), str) and "@" in row.get(c, "")
            for c in ["CNEE_EMAIL", "SHIPPER_EMAIL"]
        )
        if not has_email:
            return False
        # Cooldown check
        for col in ["CNEE_EMAIL", "SHIPPER_EMAIL"]:
            e = row.get(col, "")
            if isinstance(e, str) and "@" in e:
                ls = last_sent.get(e)
                if ls is not None and ls >= cutoff:
                    return False
        return True

    df_queue = df_master[df_master.apply(is_sendable, axis=1)].copy()
    queue_cols = ["CMD_NAME", "CNEE_NAME", "CNEE_EMAIL", "CNEE_PIC",
                  "SHIPPER_NAME", "SHIPPER_EMAIL", "SHIPPER_PIC",
                  "DESTINATION", "CARRIER", "LAST_SENT_DATE", "DAYS_SINCE_SENT", "KB_STATUS"]
    df_queue = df_queue[[c for c in queue_cols if c in df_queue.columns]]

    # ---- SENT (from log) ----
    df_sent = df_log.copy()
    if not df_sent.empty:
        df_sent["timestamp"] = df_sent["timestamp"].dt.strftime("%Y-%m-%d %H:%M").fillna("")
        df_sent = df_sent.sort_values("timestamp", ascending=False)

    # ---- REPLIED ----
    df_replied = pd.DataFrame()
    if not df_kb.empty:
        replied_mask = df_kb["STATUS"].str.lower().isin(WARM_STATUSES)
        df_replied = df_kb[replied_mask][["EMAIL", "DOMAIN", "COMPANY",
                                          "STATUS", "COUNT", "LAST_SEEN", "REMARK"]].copy()
        df_replied = df_replied.sort_values("COUNT", ascending=False)

    # ---- BOUNCED ----
    df_bounced = pd.DataFrame()
    if not df_kb.empty:
        bounce_mask = df_kb["STATUS"].str.lower().isin(DEAD_STATUSES)
        df_bounced = df_kb[bounce_mask][["EMAIL", "DOMAIN", "COMPANY",
                                          "STATUS", "COUNT", "LAST_SEEN", "REMARK"]].copy()
        df_bounced = df_bounced.sort_values("DOMAIN")

    # ---- AUTO_REPLY ----
    df_ooo = pd.DataFrame()
    if not df_kb.empty:
        ooo_mask = df_kb["STATUS"].str.lower().isin(AUTO_STATUSES)
        df_ooo = df_kb[ooo_mask].copy()
        df_ooo["REPLACEMENT_NAME"] = df_ooo["REMARK"].apply(_parse_name_from_remark)
        ooo_cols = ["EMAIL", "DOMAIN", "COMPANY", "REPLACEMENT_EMAIL",
                    "REPLACEMENT_NAME", "ROLE_HINT", "COUNT", "LAST_SEEN", "REMARK"]
        df_ooo = df_ooo[[c for c in ooo_cols if c in df_ooo.columns]]
        df_ooo = df_ooo.sort_values("COMPANY") if "COMPANY" in df_ooo.columns else df_ooo

    return {
        "master":     df_master,
        "queue":      df_queue,
        "sent":       df_sent,
        "replied":    df_replied,
        "bounced":    df_bounced,
        "auto_reply": df_ooo,
        "stats": {
            "total_contacts":   len(df_data),
            "valid_email":      df_data["CNEE_EMAIL"].str.contains("@", na=False).sum(),
            "sent_all_time":    len(df_log),
            "unique_sent":      df_log["email"].nunique() if not df_log.empty else 0,
            "bounced_dead":     len(df_bounced),
            "auto_reply_ooo":   len(df_ooo),
            "human_reply":      len(df_replied),
            "send_queue_today": len(df_queue),
            "generated_at":     datetime.now().strftime("%Y-%m-%d %H:%M"),
            "cooldown_days":    RESEND_COOLDOWN_DAYS,
        }
    }


# =========================================================
# SHEET BUILDERS
# =========================================================
def build_overview(ws, stats: dict):
    ws.tab_color = TAB_COLORS["OVERVIEW"]

    metrics = [
        ("", ""),
        ("  EMAIL ENGINE — MASTER DASHBOARD", ""),
        ("  Generated", stats["generated_at"]),
        ("  Cooldown setting", f"{stats['cooldown_days']} days"),
        ("", ""),
        ("  CONTACTS", ""),
        ("  Total rows in data.xlsx", stats["total_contacts"]),
        ("  Rows with valid email (@)", stats["valid_email"]),
        ("", ""),
        ("  SEND ACTIVITY", ""),
        ("  Total sends (all time)", stats["sent_all_time"]),
        ("  Unique addresses ever sent", stats["unique_sent"]),
        ("", ""),
        ("  EMAIL HEALTH", ""),
        ("  Bounced / Dead emails", stats["bounced_dead"]),
        ("  Auto-Reply / OOO", stats["auto_reply_ooo"]),
        ("  Human Replies received", stats["human_reply"]),
        ("", ""),
        ("  READY TO SEND NOW", ""),
        (f"  Send Queue (cooldown > {stats['cooldown_days']}d)", stats["send_queue_today"]),
    ]

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20

    title_fill = PatternFill("solid", fgColor=C_HEADER_BG)
    title_font = Font(name=FONT_NAME, bold=True, color=C_HEADER_FG, size=14)

    section_font = Font(name=FONT_NAME, bold=True, color="1F3864", size=10)
    val_fill     = PatternFill("solid", fgColor=C_OVERVIEW_VAL)
    normal_font  = Font(name=FONT_NAME, size=10)

    SECTION_LABELS = {
        "  CONTACTS", "  SEND ACTIVITY",
        "  EMAIL HEALTH", "  READY TO SEND NOW",
        "  EMAIL ENGINE — MASTER DASHBOARD",
    }

    for label, value in metrics:
        ws.append([label, value if value != "" else None])
        row = ws.max_row
        a_cell = ws.cell(row, 1)
        b_cell = ws.cell(row, 2)

        if label == "  EMAIL ENGINE — MASTER DASHBOARD":
            a_cell.font  = title_font
            a_cell.fill  = title_fill
            b_cell.fill  = title_fill
            ws.row_dimensions[row].height = 28
        elif label in SECTION_LABELS:
            a_cell.font  = section_font
            ws.row_dimensions[row].height = 20
        elif value != "" and label.strip():
            a_cell.font  = normal_font
            b_cell.font  = Font(name=FONT_NAME, bold=True, size=11)
            b_cell.fill  = val_fill
            b_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 22


def build_data_sheet(ws, df: pd.DataFrame, sheet_key: str,
                     accent_col: str | None = None, accent_color: str | None = None):
    ws.tab_color = TAB_COLORS[sheet_key]
    if df.empty:
        ws.append(["No data available"])
        return
    # Clean display: replace NaN with empty string
    df = df.fillna("").astype(str)
    df = df.replace("nan", "").replace("None", "").replace("NaT", "")
    apply_header(ws, list(df.columns))
    apply_rows(ws, df, accent_col=accent_col, accent_color=accent_color)
    freeze_and_filter(ws)
    auto_col_width(ws)


# =========================================================
# MAIN
# =========================================================
def main():
    log.info("=" * 60)
    log.info("  EMAIL MASTER DASHBOARD GENERATOR")
    log.info("=" * 60)

    data = build_datasets()
    stats = data["stats"]

    log.info("Stats: contacts=%d, valid_email=%d, queue=%d, bounced=%d, ooo=%d, replied=%d",
             stats["total_contacts"], stats["valid_email"], stats["send_queue_today"],
             stats["bounced_dead"], stats["auto_reply_ooo"], stats["human_reply"])

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # 1. OVERVIEW
    log.info("Building OVERVIEW sheet...")
    ws_ov = wb.create_sheet("OVERVIEW")
    build_overview(ws_ov, stats)

    # 2. SEND_QUEUE
    log.info("Building SEND_QUEUE sheet (%d rows)...", len(data["queue"]))
    ws_sq = wb.create_sheet("SEND_QUEUE")
    build_data_sheet(ws_sq, data["queue"], "SEND_QUEUE")

    # 3. SENT
    log.info("Building SENT sheet (%d rows)...", len(data["sent"]))
    ws_sent = wb.create_sheet("SENT")
    build_data_sheet(ws_sent, data["sent"], "SENT")

    # 4. REPLIED
    log.info("Building REPLIED sheet (%d rows)...", len(data["replied"]))
    ws_rep = wb.create_sheet("REPLIED")
    build_data_sheet(ws_rep, data["replied"], "REPLIED",
                     accent_col="EMAIL", accent_color=C_REPLY_ROW)

    # 5. BOUNCED
    log.info("Building BOUNCED sheet (%d rows)...", len(data["bounced"]))
    ws_bn = wb.create_sheet("BOUNCED")
    build_data_sheet(ws_bn, data["bounced"], "BOUNCED",
                     accent_col="EMAIL", accent_color=C_BOUNCE_ROW)

    # 6. AUTO_REPLY
    log.info("Building AUTO_REPLY sheet (%d rows)...", len(data["auto_reply"]))
    ws_ooo = wb.create_sheet("AUTO_REPLY")
    build_data_sheet(ws_ooo, data["auto_reply"], "AUTO_REPLY",
                     accent_col="REPLACEMENT_EMAIL", accent_color=C_OOO_ROW)

    # 7. MASTER
    log.info("Building MASTER sheet (%d rows)...", len(data["master"]))
    ws_m = wb.create_sheet("MASTER")
    build_data_sheet(ws_m, data["master"], "MASTER")

    wb.save(OUTPUT_FILE)

    log.info("")
    log.info("=" * 60)
    log.info("  OUTPUT: %s", OUTPUT_FILE.name)
    log.info("=" * 60)
    log.info("  OVERVIEW    : key metrics")
    log.info("  SEND_QUEUE  : %d contacts ready to send", stats["send_queue_today"])
    log.info("  SENT        : %d total send history rows", stats["sent_all_time"])
    log.info("  REPLIED     : %d human replies", stats["human_reply"])
    log.info("  BOUNCED     : %d dead emails", stats["bounced_dead"])
    log.info("  AUTO_REPLY  : %d OOO contacts", stats["auto_reply_ooo"])
    log.info("  MASTER      : %d total contacts", stats["total_contacts"])
    log.info("=" * 60)


if __name__ == "__main__":
    main()
