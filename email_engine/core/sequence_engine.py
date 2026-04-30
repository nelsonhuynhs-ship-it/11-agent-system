"""
sequence_engine.py -- Master-Aware Multi-Source Email Sequence Engine
====================================================================
Reads from the 3 master files (cnee_master, contact_master, shipper_master)
and manages a 3-step email sequence with automatic progression.

Modes:
  --source [cnee|contact|shipper|all]   Select which master to target
  --sequence                            Run automatic sequence progression
  --dry-run                             Simulate without sending

Usage:
  python sequence_engine.py --source contact --dry-run
  python sequence_engine.py --sequence --dry-run
  python sequence_engine.py --source contact --sequence
"""

from __future__ import annotations

import argparse
import csv
import logging
import sys
from datetime import datetime, date
from pathlib import Path

import pandas as pd

# =========================================================
# CONFIG (paths via shared.paths — OneDrive data, local runtime)
# =========================================================
_repo_root = str(Path(__file__).parent.parent.parent)
if _repo_root not in sys.path:
    sys.path.insert(0, _repo_root)
from shared import paths as sp

LOG_DIR      = sp.EMAIL_LOG_DIR
CONFIG_FILE  = sp.CONFIG_XLSX
PROFILE_PDF  = sp.COMPANY_PDF
LOGO_PNG     = sp.LOGO_FILE

CNEE_MASTER    = sp.CNEE_MASTER
CONTACT_MASTER = sp.CONTACT_MASTER
SHIPPER_MASTER = sp.SHIPPER_MASTER

EMAIL_LOG_FILE   = sp.EMAIL_LOG
KNOWLEDGE_FILE   = sp.EMAIL_LOG_DIR / "email_knowledge.csv"

LOG_DIR.mkdir(parents=True, exist_ok=True)

# Sequence timing
SEQ_STEP2_DELAY = 4   # days after step 1
SEQ_STEP3_DELAY = 5   # days after step 2

BATCH_SIZE = 100

# =========================================================
# LOGGING
# =========================================================
logging.basicConfig(
    level   = logging.INFO,
    format  = "[%(asctime)s] %(levelname)-8s %(message)s",
    datefmt = "%H:%M:%S",
    handlers= [logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


# =========================================================
# EMAIL LOG
# =========================================================
def log_email_send(email, subject, campaign_id, cycle_id="1", status="SENT"):
    log_exists = EMAIL_LOG_FILE.exists()
    with open(EMAIL_LOG_FILE, mode="a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not log_exists:
            writer.writerow(["timestamp", "email", "subject",
                             "campaign_id", "cycle_id", "status"])
        writer.writerow([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            email, subject, campaign_id, cycle_id, status,
        ])


# =========================================================
# LOAD CONFIG (rich text body from config.xlsx)
# =========================================================
def load_config() -> dict:
    """Load email config from config.xlsx (INTROTEXT, RATETABLEHTML, etc)."""
    if not CONFIG_FILE.exists():
        return {}
    try:
        from openpyxl import load_workbook
        from openpyxl.cell.rich_text import CellRichText
        wb = load_workbook(CONFIG_FILE, rich_links=True)
        ws = wb.active
        cfg = {}
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=False):
            key_cell = row[0]
            val_cell = row[1]
            if key_cell.value:
                key = str(key_cell.value).strip().upper()
                val = val_cell.value
                if isinstance(val, CellRichText):
                    cfg[key] = _rich_to_html(val)
                else:
                    cfg[key] = str(val or "").replace("\n", "<br>")
        return cfg
    except Exception as exc:
        log.warning("Could not load config.xlsx: %s", exc)
        return {}


def _rich_to_html(rich_text) -> str:
    """Minimal rich text to HTML converter."""
    from openpyxl.cell.rich_text import TextBlock
    parts = []
    for item in rich_text:
        if isinstance(item, str):
            parts.append(item.replace("\n", "<br>"))
        elif isinstance(item, TextBlock):
            text = str(item.text or "").replace("\n", "<br>")
            font = item.font
            if font:
                if getattr(font, "bold", False):
                    text = f"<strong>{text}</strong>"
                if getattr(font, "italic", False):
                    text = f"<em>{text}</em>"
            parts.append(text)
    return "".join(parts)


# =========================================================
# SEQUENCE TEMPLATES
# =========================================================
def _first_name(full_name: str) -> str:
    """Extract first name from a full name string."""
    if not isinstance(full_name, str) or not full_name.strip():
        return "Team"
    parts = full_name.strip().split()
    return parts[0].capitalize()


def build_step1_email(row: dict, source: str, cfg: dict) -> tuple[str, str]:
    """Build Step 1 email (initial outreach). Returns (subject, html_body)."""
    company = str(row.get("COMPANY", "")).strip()
    dest = str(row.get("DESTINATION", "")).strip()
    campaign = str(row.get("CAMPAIGN_ID", "")).strip()

    if source == "contact":
        name = _first_name(row.get("CONTACT_NAME", ""))
        position = str(row.get("POSITION", "")).strip()
        greeting = f"Dear {name},"
        if position:
            opener = (f"As {position} at {company}, you likely manage "
                      f"import logistics from Vietnam -- I wanted to share "
                      f"something relevant.")
        else:
            opener = (f"I noticed {company} imports from Vietnam and wanted "
                      f"to share a quick update.")
    elif source == "shipper":
        name = _first_name(row.get("SHIPPER_PIC", ""))
        greeting = f"Dear {name},"
        opener = (f"We help Vietnamese exporters like {company} reach "
                  f"US markets with competitive ocean freight rates and "
                  f"reliable schedules.")
    else:  # cnee
        name = _first_name(row.get("CNEE_PIC", ""))
        greeting = f"Dear {name},"
        opener = (f"I noticed {company} imports goods from Vietnam "
                  f"and wanted to introduce our freight services for "
                  f"your trade lane.")

    if dest:
        subject = f"{company} shipment to {dest} -- quick question"
    else:
        subject = f"{company} -- Vietnam freight rates // NELSON"

    intro = cfg.get("INTROTEXT", "")
    rate_table = cfg.get("RATETABLEHTML", "")
    closing = cfg.get("CLOSINGTEXT", "")
    signature = cfg.get("SIGNATURE", "")

    body = f"""<html><body>
    {greeting}<br><br>
    {opener}<br><br>
    {intro}<br><br>
    {rate_table}<br><br>
    {closing}<br><br>
    {signature}
    </body></html>"""

    return subject, body


def build_step2_email(row: dict, source: str, cfg: dict) -> tuple[str, str]:
    """Build Step 2 email (value-add follow-up). Returns (subject, html_body)."""
    company = str(row.get("COMPANY", "")).strip()
    dest = str(row.get("DESTINATION", "")).strip()
    pol = str(row.get("POL", "")).strip()

    if source == "contact":
        name = _first_name(row.get("CONTACT_NAME", ""))
    elif source == "shipper":
        name = _first_name(row.get("SHIPPER_PIC", ""))
    else:
        name = _first_name(row.get("CNEE_PIC", ""))

    route = f"{pol}-{dest}" if pol and dest else "Vietnam-US"
    subject = f"Re: {company} -- {route} rate update"

    rate_table = cfg.get("RATETABLEHTML", "")
    signature = cfg.get("SIGNATURE", "")

    body = f"""<html><body>
    Hi {name},<br><br>
    Following up on my previous email -- I wanted to share the latest
    rates for the <strong>{route}</strong> route that may be relevant
    for {company}.<br><br>
    {rate_table}<br><br>
    We have direct services with competitive transit times on this lane.
    Happy to put together a custom quote if you share your typical
    container volume and commodity.<br><br>
    {signature}
    </body></html>"""

    return subject, body


def build_step3_email(row: dict, source: str, cfg: dict) -> tuple[str, str]:
    """Build Step 3 email (direct CTA). Returns (subject, html_body)."""
    if source == "contact":
        name = _first_name(row.get("CONTACT_NAME", ""))
    elif source == "shipper":
        name = _first_name(row.get("SHIPPER_PIC", ""))
    else:
        name = _first_name(row.get("CNEE_PIC", ""))

    subject = f"{name} -- 10 minutes this week?"
    signature = cfg.get("SIGNATURE", "")

    body = f"""<html><body>
    Hi {name},<br><br>
    I know your inbox is busy, so I will keep this short:<br><br>
    Would you have <strong>10 minutes this week</strong> for a quick call?
    I can walk you through our latest rates and how we have helped similar
    importers save 15-20% on their Vietnam freight.<br><br>
    No pressure at all -- just reply with a time that works, or let me know
    if now is not the right time.<br><br>
    {signature}
    </body></html>"""

    return subject, body


# =========================================================
# OUTLOOK SEND
# =========================================================


# =========================================================
# UPDATE MASTER FILE
# =========================================================
def update_master_seq(filepath: Path, email: str, seq_step: int,
                      completed: bool = False) -> None:
    """Update SEQ_STEP and SEQ_LAST_SENT for a specific email in a master file."""
    df = pd.read_excel(filepath)
    df.columns = df.columns.str.strip().str.upper()

    mask = df["EMAIL"].astype(str).str.lower() == email.lower()
    if mask.any():
        df.loc[mask, "SEQ_STEP"] = seq_step
        df.loc[mask, "SEQ_LAST_SENT"] = datetime.now().strftime("%Y-%m-%d")
        if completed:
            df.loc[mask, "SEQ_STATUS"] = "COMPLETED"
        df.to_excel(filepath, index=False)


def batch_update_master(filepath: Path, updates: list[dict]) -> None:
    """Batch update multiple rows in a master file.
    updates: list of {email, seq_step, completed}
    """
    if not updates:
        return
    df = pd.read_excel(filepath)
    df.columns = df.columns.str.strip().str.upper()
    df["_EMAIL_LOWER"] = df["EMAIL"].astype(str).str.lower()

    today = datetime.now().strftime("%Y-%m-%d")
    for u in updates:
        mask = df["_EMAIL_LOWER"] == u["email"].lower()
        if mask.any():
            df.loc[mask, "SEQ_STEP"] = u["seq_step"]
            df.loc[mask, "SEQ_LAST_SENT"] = today
            if u.get("completed"):
                df.loc[mask, "SEQ_STATUS"] = "COMPLETED"
            if u.get("already_sent"):
                df.loc[mask, "ALREADY_SENT"] = "Y"
                df.loc[mask, "LAST_SENT_DATE"] = today

    df.drop(columns=["_EMAIL_LOWER"], inplace=True)
    df.to_excel(filepath, index=False)


# =========================================================
# LOAD MASTER + FILTER
# =========================================================
def load_master(source: str) -> tuple[pd.DataFrame, Path]:
    """Load the appropriate master file. Returns (df, filepath)."""
    paths = {
        "cnee":    CNEE_MASTER,
        "contact": CONTACT_MASTER,
        "shipper": SHIPPER_MASTER,
    }
    fpath = paths[source]
    if not fpath.exists():
        log.error("%s not found. Run combine_all.py first.", fpath.name)
        return pd.DataFrame(), fpath

    df = pd.read_excel(fpath)
    df.columns = df.columns.str.strip().str.upper()

    # Normalize
    for col in ["EMAIL", "COMPANY", "CAMPAIGN_ID", "SEQ_STATUS",
                "ALREADY_SENT", "KB_STATUS"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()

    # Convert numeric columns
    for col in ["SEQ_STEP", "PRIORITY_SCORE", "TOTAL_SHIPMENT",
                "EMAIL_QUALITY_SCORE"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    return df, fpath


def get_sendable(df: pd.DataFrame, source: str) -> pd.DataFrame:
    """Filter to rows that are eligible for step 1 (new contacts)."""
    mask = (
        (df["SEQ_STATUS"].str.upper() == "ACTIVE") &
        (df["SEQ_STEP"] == 0) &
        (df["EMAIL"].str.contains("@", na=False))
    )
    result = df[mask].copy()

    # Sort by priority
    if source == "contact" and "PRIORITY_SCORE" in result.columns:
        result = result.sort_values("PRIORITY_SCORE", ascending=False)
    elif "TOTAL_SHIPMENT" in result.columns:
        result = result.sort_values("TOTAL_SHIPMENT", ascending=False)

    return result


def get_step2_ready(df: pd.DataFrame) -> pd.DataFrame:
    """Get contacts ready for Step 2 (SEQ_STEP=1 and days >= SEQ_STEP2_DELAY)."""
    mask = (
        (df["SEQ_STATUS"].str.upper() == "ACTIVE") &
        (df["SEQ_STEP"] == 1)
    )
    candidates = df[mask].copy()
    if candidates.empty:
        return candidates

    today = pd.Timestamp.now()
    candidates["SEQ_LAST_SENT"] = pd.to_datetime(
        candidates["SEQ_LAST_SENT"], errors="coerce"
    )
    candidates["_DAYS_SINCE"] = (
        today - candidates["SEQ_LAST_SENT"]
    ).dt.days.fillna(999).astype(int)

    return candidates[candidates["_DAYS_SINCE"] >= SEQ_STEP2_DELAY]


def get_step3_ready(df: pd.DataFrame) -> pd.DataFrame:
    """Get contacts ready for Step 3 (SEQ_STEP=2 and days >= SEQ_STEP3_DELAY)."""
    mask = (
        (df["SEQ_STATUS"].str.upper() == "ACTIVE") &
        (df["SEQ_STEP"] == 2)
    )
    candidates = df[mask].copy()
    if candidates.empty:
        return candidates

    today = pd.Timestamp.now()
    candidates["SEQ_LAST_SENT"] = pd.to_datetime(
        candidates["SEQ_LAST_SENT"], errors="coerce"
    )
    candidates["_DAYS_SINCE"] = (
        today - candidates["SEQ_LAST_SENT"]
    ).dt.days.fillna(999).astype(int)

    return candidates[candidates["_DAYS_SINCE"] >= SEQ_STEP3_DELAY]


# =========================================================
# SOURCE MODE (send step 1 to new contacts from a source)
# =========================================================
def run_source_mode(source: str, dry_run: bool = False,
                    batch: int = BATCH_SIZE) -> dict:
    """Send Step 1 emails from a specific master source."""
    df, fpath = load_master(source)
    if df.empty:
        return {"sent": 0, "source": source}

    sendable = get_sendable(df, source)
    cfg = load_config()

    log.info("Source: %s | Total: %d | Sendable (step 0): %d",
             source, len(df), len(sendable))

    if source == "contact" and "PRIORITY_SCORE" in sendable.columns:
        high_pri = (sendable["PRIORITY_SCORE"] >= 8).sum()
        log.info("  High priority (score >= 8): %d", high_pri)

    if sendable.empty:
        log.info("No new contacts to send.")
        return {"sent": 0, "source": source}

    targets = sendable.head(batch)
    log.info("Batch: %d contacts (limit %d)", len(targets), batch)

    sent_count = 0
    updates = []

    for _, row in targets.iterrows():
        email = row["EMAIL"]
        subject, body = build_step1_email(row.to_dict(), source, cfg)

        if dry_run:
            log.info("  [DRY] Step 1 -> %s (%s)", email,
                     row.get("COMPANY", ""))
        else:
            from email_engine.senders import send_html_via_graph
            ok, _ = send_html_via_graph(to=email, subject=subject, html_body=body)
            if ok:
                log_email_send(email, subject,
                               row.get("CAMPAIGN_ID", "SEQUENCE"))
                log.info("  SENT Step 1 -> %s", email)
            else:
                continue

        sent_count += 1
        updates.append({
            "email": email, "seq_step": 1,
            "completed": False, "already_sent": True,
        })

    if not dry_run and updates:
        batch_update_master(fpath, updates)

    return {"sent": sent_count, "source": source}


# =========================================================
# SEQUENCE MODE (auto-advance steps 1/2/3 across all sources)
# =========================================================
def run_sequence_mode(sources: list[str] | None = None,
                      dry_run: bool = False,
                      batch: int = BATCH_SIZE) -> dict:
    """Run the full 3-step sequence across specified sources."""
    if sources is None:
        sources = ["contact", "cnee", "shipper"]

    cfg = load_config()
    totals = {"step1": 0, "step2": 0, "step3": 0,
              "step1_highpri": 0, "sources_processed": []}

    for source in sources:
        df, fpath = load_master(source)
        if df.empty:
            continue

        log.info("")
        log.info("=" * 50)
        log.info("  SEQUENCE: %s (%d rows)", source.upper(), len(df))
        log.info("=" * 50)

        updates = []

        # --- Step 3 first (most mature contacts) ---
        step3 = get_step3_ready(df)
        if not step3.empty:
            log.info("  Step 3 ready: %d contacts", len(step3))
            for _, row in step3.head(batch).iterrows():
                email = row["EMAIL"]
                subject, body = build_step3_email(row.to_dict(), source, cfg)
                if dry_run:
                    log.info("    [DRY] Step 3 -> %s (%dd since last)",
                             email, row.get("_DAYS_SINCE", "?"))
                else:
                    from email_engine.senders import send_html_via_graph
                    ok, _ = send_html_via_graph(to=email, subject=subject, html_body=body)
                    if ok:
                        log_email_send(email, subject,
                                       row.get("CAMPAIGN_ID", "SEQ-3"))
                        log.info("    SENT Step 3 -> %s", email)
                    else:
                        continue
                totals["step3"] += 1
                updates.append({
                    "email": email, "seq_step": 3, "completed": True,
                })

        # --- Step 2 (follow-up) ---
        step2 = get_step2_ready(df)
        if not step2.empty:
            log.info("  Step 2 ready: %d contacts", len(step2))
            for _, row in step2.head(batch).iterrows():
                email = row["EMAIL"]
                subject, body = build_step2_email(row.to_dict(), source, cfg)
                if dry_run:
                    log.info("    [DRY] Step 2 -> %s (%dd since last)",
                             email, row.get("_DAYS_SINCE", "?"))
                else:
                    from email_engine.senders import send_html_via_graph
                    ok, _ = send_html_via_graph(to=email, subject=subject, html_body=body)
                    if ok:
                        log_email_send(email, subject,
                                       row.get("CAMPAIGN_ID", "SEQ-2"))
                        log.info("    SENT Step 2 -> %s", email)
                    else:
                        continue
                totals["step2"] += 1
                updates.append({
                    "email": email, "seq_step": 2, "completed": False,
                })

        # --- Step 1 (new contacts) ---
        step1 = get_sendable(df, source)
        if not step1.empty:
            step1_batch = step1.head(batch)
            high_pri = 0
            if source == "contact" and "PRIORITY_SCORE" in step1_batch.columns:
                high_pri = (step1_batch["PRIORITY_SCORE"] >= 8).sum()
                totals["step1_highpri"] += high_pri

            log.info("  Step 1 ready: %d contacts (batch %d)",
                     len(step1), len(step1_batch))
            if high_pri:
                log.info("    High-priority: %d", high_pri)

            for _, row in step1_batch.iterrows():
                email = row["EMAIL"]
                subject, body = build_step1_email(
                    row.to_dict(), source, cfg
                )
                if dry_run:
                    pri_tag = ""
                    if source == "contact":
                        pri = row.get("PRIORITY_SCORE", 0)
                        pri_tag = f" [P{pri}]" if pri >= 8 else ""
                    log.info("    [DRY] Step 1 -> %s%s (%s)",
                             email, pri_tag, row.get("COMPANY", ""))
                else:
                    from email_engine.senders import send_html_via_graph
                    ok, _ = send_html_via_graph(to=email, subject=subject, html_body=body)
                    if ok:
                        log_email_send(email, subject,
                                       row.get("CAMPAIGN_ID", "SEQ-1"))
                        log.info("    SENT Step 1 -> %s", email)
                    else:
                        continue
                totals["step1"] += 1
                updates.append({
                    "email": email, "seq_step": 1,
                    "completed": False, "already_sent": True,
                })

        # Batch update master file
        if not dry_run and updates:
            batch_update_master(fpath, updates)
            log.info("  Updated %s: %d rows", fpath.name, len(updates))

        totals["sources_processed"].append(source)

    return totals


# =========================================================
# SUMMARY
# =========================================================
def print_summary(totals: dict, dry_run: bool = False):
    prefix = "[DRY RUN] Would send" if dry_run else "Sent"
    log.info("")
    log.info("=" * 60)
    log.info("  SEQUENCE ENGINE -- SUMMARY")
    log.info("=" * 60)
    log.info("  %s Step 1 (initial):    %d emails", prefix, totals.get("step1", 0))
    if totals.get("step1_highpri", 0):
        log.info("    High-priority:        %d", totals["step1_highpri"])
    log.info("  %s Step 2 (follow-up):  %d emails", prefix, totals.get("step2", 0))
    log.info("  %s Step 3 (CTA):        %d emails", prefix, totals.get("step3", 0))
    total = totals.get("step1", 0) + totals.get("step2", 0) + totals.get("step3", 0)
    log.info("  Total emails %s:   %d",
             "queued" if dry_run else "sent", total)
    log.info("=" * 60)


# =========================================================
# CLI
# =========================================================
def main():
    parser = argparse.ArgumentParser(description="Email Sequence Engine")
    parser.add_argument("--source", choices=["cnee", "contact", "shipper", "all"],
                        default=None,
                        help="Master file to target")
    parser.add_argument("--sequence", action="store_true",
                        help="Run full sequence progression (steps 1/2/3)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Simulate without sending")
    parser.add_argument("--batch", type=int, default=BATCH_SIZE,
                        help="Max emails per batch")
    args = parser.parse_args()

    dry_label = " [DRY RUN]" if args.dry_run else ""
    log.info("=" * 60)
    log.info("  SEQUENCE ENGINE%s", dry_label)
    log.info("=" * 60)

    if args.sequence:
        # Sequence mode: process all (or specified) sources
        sources = None
        if args.source and args.source != "all":
            sources = [args.source]
        elif args.source == "all":
            sources = ["contact", "cnee", "shipper"]

        totals = run_sequence_mode(
            sources=sources, dry_run=args.dry_run, batch=args.batch
        )
        print_summary(totals, dry_run=args.dry_run)

    elif args.source:
        # Source mode only: send step 1 to new contacts
        if args.source == "all":
            all_totals = {"step1": 0, "step2": 0, "step3": 0, "step1_highpri": 0}
            for src in ["contact", "cnee", "shipper"]:
                result = run_source_mode(src, dry_run=args.dry_run,
                                         batch=args.batch)
                all_totals["step1"] += result["sent"]
            print_summary(all_totals, dry_run=args.dry_run)
        else:
            result = run_source_mode(args.source, dry_run=args.dry_run,
                                     batch=args.batch)
            log.info("Sent %d Step 1 emails from %s",
                     result["sent"], result["source"])

    else:
        parser.print_help()
        print("\nExamples:")
        print("  python sequence_engine.py --source contact --dry-run")
        print("  python sequence_engine.py --sequence --dry-run")
        print("  python sequence_engine.py --source all --sequence")


if __name__ == "__main__":
    main()
