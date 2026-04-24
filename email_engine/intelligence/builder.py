# -*- coding: utf-8 -*-
"""
builder.py — Orchestrator for smart emails.

Combines:
    profile (optional, caller-supplied — usually from intel/memory Phase 02)
    + market_engine.analyze_lane() per destination
    + template_selector.match() by dominant state
    + template_renderer.render_email()
    + rate_table HTML (color-coded by state — Jinja-lite partial)

Public API
----------
build_email(cnee_email, pol, destinations, markup=20.0, profile=None) -> dict
    Returns {to, subject, html_body, meta}.
"""
from __future__ import annotations

import logging
from datetime import date
from pathlib import Path
from typing import Any

from .market_engine import analyze_lane
from .template_selector import match as tmpl_match, dominant_state, load_rules
from .template_renderer import render_email, render_text

log = logging.getLogger("intelligence_builder")

# Default destinations + max cap loaded from default_routes.yaml (SOT).
# Keeping the loader here so build_email() never drifts from web_server.py.
_SAFE_FALLBACK_DESTS = ["USLAX", "USSAV", "USNYC"]
_SAFE_FALLBACK_MAX = 10


def _extract_pod_codes(fast_bulk_value) -> list[str]:
    """
    Parse fast_bulk_default from YAML — supports two formats:
      Legacy (list of strings): [USLAX, USSAV, ...]
      New (dict with pod_list):  {pod_list: [{code: USLAX, ...}, ...], ...}
    Returns list of port code strings (upper-stripped).
    """
    if fast_bulk_value is None:
        return []
    if isinstance(fast_bulk_value, list):
        # Legacy format: plain list of strings
        return [str(d).strip().upper() for d in fast_bulk_value if str(d).strip()]
    if isinstance(fast_bulk_value, dict):
        pod_list = fast_bulk_value.get("pod_list") or []
        codes = []
        for entry in pod_list:
            if isinstance(entry, str):
                # String fallback inside dict format
                codes.append(entry.strip().upper())
            elif isinstance(entry, dict):
                code = str(entry.get("code", "")).strip().upper()
                if code:
                    codes.append(code)
        return codes
    return []


def _load_routing_config() -> tuple[list[str], int]:
    """Return (default_destinations, max_destinations_per_email) from YAML."""
    try:
        from shared.paths import DEFAULT_ROUTES_CFG
        _onedrive = Path(DEFAULT_ROUTES_CFG)
    except Exception:
        _onedrive = None
    _local = Path(__file__).resolve().parent.parent / "config" / "default_routes.yaml"
    yaml_path = _onedrive if (_onedrive and _onedrive.exists()) else _local
    try:
        import yaml  # type: ignore
        with open(yaml_path, "r", encoding="utf-8") as fh:
            data = yaml.safe_load(fh) or {}
        fast_bulk = data.get("fast_bulk_default")
        cleaned = _extract_pod_codes(fast_bulk)
        if not cleaned:
            cleaned = _extract_pod_codes(data.get("global_default"))
        # max cap: prefer nested dict value, then top-level key
        cap_raw = _SAFE_FALLBACK_MAX
        if isinstance(fast_bulk, dict):
            cap_raw = fast_bulk.get("max_destinations_per_email", cap_raw)
        cap_raw = data.get("max_destinations_per_email", cap_raw)
        cap = int(cap_raw) if cap_raw else _SAFE_FALLBACK_MAX
        # Enforce cap — truncate with warning if list exceeds limit
        if len(cleaned) > cap:
            log.warning(
                "[builder] pod_list has %d entries > max %d — truncating", len(cleaned), cap
            )
            cleaned = cleaned[:cap]
        if cleaned:
            log.info("[builder] routing config loaded: %d lanes, cap=%d", len(cleaned), cap)
            return cleaned, cap
    except Exception as e:
        log.warning("[builder] default_routes.yaml load failed (%s) — using safe fallback", e)
    return list(_SAFE_FALLBACK_DESTS), _SAFE_FALLBACK_MAX


_DEFAULT_DESTINATIONS, _MAX_DESTINATIONS = _load_routing_config()

# Rate-table partial (simple Jinja-lite template)
_RATE_TABLE_PATH = (
    Path(__file__).resolve().parent.parent / "templates" / "rate_table.html"
)

# State → row background color + badge + icon
_STATE_COLOR = {
    "URGENT":      {"bg": "#fef2f2", "badge": "#b91c1c", "icon": "🚨",
                    "bar": "#dc2626", "text": "#991b1b"},
    "COMPETITIVE": {"bg": "#ecfdf5", "badge": "#047857", "icon": "💰",
                    "bar": "#10b981", "text": "#064e3b"},
    "STABLE":      {"bg": "#f0fdf4", "badge": "#15803d", "icon": "✓",
                    "bar": "#22c55e", "text": "#14532d"},
    "DECLINING":   {"bg": "#eff6ff", "badge": "#1d4ed8", "icon": "🔽",
                    "bar": "#3b82f6", "text": "#1e3a8a"},
}

# Port code → human-readable name (for Pudong Prime style rate table)
_POL_NAMES = {
    "HPH": "Hai Phong",
    "HCM": "Ho Chi Minh",
    "DAD": "Da Nang",
    "SGN": "Ho Chi Minh",
    "HAN": "Hanoi",
    "MYPKG": "Port Klang",
    "SIN": "Singapore",
}

_POD_NAMES = {
    "USLAX": "Los Angeles, CA",
    "USLGB": "Long Beach, CA",
    "USOAK": "Oakland, CA",
    "USSEA": "Seattle, WA",
    "USTIW": "Tacoma, WA",
    "USNYC": "New York, NY",
    "USHOU": "Houston, TX",
    "USSAV": "Savannah, GA",
    "USCHS": "Charleston, SC",
    "USMIA": "Miami, FL",
    "USMEM": "Memphis, TN",
    "USORF": "Norfolk, VA",
    "USBAL": "Baltimore, MD",
    "USILG": "Wilmington, DE",
    "USPDX": "Portland, OR",
    "USJAX": "Jacksonville, FL",
    "USMSY": "New Orleans, LA",
    "USBOS": "Boston, MA",
    "USPHL": "Philadelphia, PA",
    "CAVAN": "Vancouver, BC",
    "CAMTR": "Montreal, QC",
    "CAPRR": "Prince Rupert, BC",
    "CATOR": "Toronto, ON",
    "MXZLO": "Manzanillo",
    "MXVER": "Veracruz",
}


def _load_rate_table_tpl() -> str:
    """Load the rate_table.html partial (returns '' if missing)."""
    try:
        if _RATE_TABLE_PATH.exists():
            return _RATE_TABLE_PATH.read_text(encoding="utf-8")
    except Exception as e:
        log.warning("[builder] could not load rate_table.html: %s", e)
    return ""


def _pol_name(code: str) -> str:
    """POL full name for display. Falls back to code if unknown."""
    return _POL_NAMES.get((code or "").upper(), code.upper() if code else "")


def _pod_name(code: str) -> str:
    """POD full name for display."""
    return _POD_NAMES.get((code or "").upper(), code.upper() if code else "")


def _fmt_rate(v) -> str:
    """Format USD rate as bold number (no prefix — column header says USD)."""
    if v is None or v == "" or v == 0:
        return "—"
    try:
        return f"{int(float(v)):,}"
    except Exception:
        return "—"


def _pick_best_lane(lane_intels: list[dict]) -> dict | None:
    """Return the lane that should get the BEST badge (lowest 40HQ rate wins).
    Falls back to first lane if rates missing.
    """
    if not lane_intels:
        return None
    rated = [ln for ln in lane_intels if ln.get("current_rate_40hq")]
    if not rated:
        return lane_intels[0]
    return min(rated, key=lambda ln: float(ln.get("current_rate_40hq") or 1e9))


# ── Rate Table v2 helpers: POD format + SVC derivation ─────────────────────

# Carriers whose FAK rate is shipper-owned container (SOC) by contract convention
_FAK_SOC_CARRIERS = {"ONE", "CMA", "YML", "HPL"}


def _derive_svc(rate_type: str, carrier: str) -> str:
    """
    Derive SVC column label from Parquet Rate_Type + carrier.

    Rules (per Nelson 2026-04-17, updated 2026-04-19):
      FAK + ONE/CMA/YML/HPL  → SOC
      FAK + others           → COC
      FIX + HPL              → SOC
      FIX + others           → Provide commodity
      SCFI                   → SCFI market
      else                   → direct
    """
    rt = (rate_type or "").upper().strip()
    c = (carrier or "").upper().strip()
    if rt == "FIX":
        return "SOC" if c == "HPL" else "Provide commodity"
    if rt == "FAK":
        return "SOC" if c in _FAK_SOC_CARRIERS else "COC"
    if rt == "SCFI":
        return "SCFI market"
    return "direct"


# Known US/Canada ocean ports (show only port code, no city name)
_US_OCEAN_PORTS = {
    "USLAX", "USLGB", "USOAK", "USSEA", "USTIW", "USPDX",   # West Coast
    "USNYC", "USEWR", "USBAL", "USPHL", "USBOS", "USILG",   # North East
    "USSAV", "USCHS", "USJAX", "USMIA",                      # South East
    "USHOU", "USMSY",                                        # Gulf
    "USORF",                                                 # Mid-Atlantic
    "CAVAN", "CAPRR", "CAMTR", "CAHAL",                      # Canada ocean
}

# Common POD → main ocean port routing hint (for inland "via X" display)
_INLAND_VIA = {
    "USCHI": "LAX/LGB", "USMEM": "LAX/LGB", "USDAL": "LAX/LGB",
    "USATL": "SAV",     "USNSH": "SAV",     "USCLT": "CHS",
    "USSTL": "LAX/LGB", "USKC":  "LAX/LGB", "USMCI": "LAX/LGB",
    "USDEN": "LAX/LGB", "USMSP": "LAX/LGB", "USDTW": "NYC",
    "USCLE": "NYC",     "USCOL": "NYC",     "USIAH": "HOU",
    "USCVG": "NYC",     "CATOR": "VAN",
}


def _is_main_port(pod_code: str) -> bool:
    """True if POD is a direct ocean port (no inland IPI leg)."""
    return (pod_code or "").upper() in _US_OCEAN_PORTS


def _fmt_pod_header(pod_code: str, place_name: str) -> tuple[str, str]:
    """
    Return (main_text, sub_text) for POD chevron header.

    Main port  → ("USLAX", "")                          -- clean code only
    Inland     → ("CHICAGO, IL", "via LAX/LGB")         -- place + routing hint
    """
    code = (pod_code or "").upper()
    place = (place_name or "").strip()

    if _is_main_port(code):
        return code, ""  # Clean: no city name duplication

    # Inland — prefer actual place name, fall back to POD name map
    display_place = place.upper() or _POD_NAMES.get(code, code).upper()
    via_hint = _INLAND_VIA.get(code, "")
    sub = f"via {via_hint}" if via_hint else ""
    return display_place, sub


def _fmt_validity(eff, exp) -> str:
    """Format eff/exp pair as '8 Apr – 30 Apr' or '– 30 Apr' (expiry-only)."""
    from datetime import date as _d, datetime as _dt
    def _to_d(v):
        if v is None or v == "":
            return None
        if isinstance(v, _d) and not isinstance(v, _dt):
            return v
        try:
            import pandas as pd  # local import, pandas already loaded upstream
            if pd.isna(v):
                return None
            return pd.to_datetime(v).date()
        except Exception:
            try:
                return _dt.fromisoformat(str(v)[:10]).date()
            except Exception:
                return None
    d_eff = _to_d(eff)
    d_exp = _to_d(exp)
    if d_eff and d_exp:
        return f"{d_eff.strftime('%d %b')} – {d_exp.strftime('%d %b')}"
    if d_exp:
        return f"– {d_exp.strftime('%d %b')}"
    if d_eff:
        return f"{d_eff.strftime('%d %b')} –"
    return "—"


def _render_rate_table(lane_intels: list[dict], primary_dest: str = "") -> str:
    """
    Render per-lane rate table — Pudong Prime branded quote style (v2).

    Layout:
      [Column header: CARRIER | 20GP | 40HQ | VALID | SVC]
      [POL band — full-width blue bar]
      [POD chevron row per lane]
      [N carrier rows per POD — cheapest first, BEST pill on global cheapest]
      [Footer strip — local charges]

    v3 (2026-04-19): primary_dest — if CNEE has known DESTINATION, that POD
    row is highlighted with "YOUR LANE" pill + amber accent. Email shows all
    9 lanes (breadth) + flags the one we know matters to them. Per Nelson:
    "highlight known, show 9 luôn" — balance personalization + showing coverage.
    """
    if not lane_intels:
        return (
            "<p style='color:#94a3b8;font-style:italic;padding:12px 16px;"
            "border:1px dashed #e2e8f0;border-radius:6px;margin:8px 0;'>"
            "No rate data available for these lanes.</p>"
        )

    pol_code = (lane_intels[0].get("pol") or "HPH").upper()
    pol_full = _pol_name(pol_code)

    # Global BEST = cheapest (carrier, lane) pair across all lanes
    global_best_40 = float("inf")
    global_best_key = None
    for lane in lane_intels:
        for c in lane.get("carriers") or []:
            r40 = c.get("rate_40") or 0
            if r40 and r40 < global_best_40:
                global_best_40 = r40
                global_best_key = (id(lane), c.get("carrier", "").upper())

    rows = []

    # ─── Column header row ────────────────────────────────────────
    _th = ("padding:6px 12px;color:#64748b;font-size:9px;"
           "letter-spacing:1.5px;font-weight:700;")
    rows.append(
        "<tr style='background:#f8fafc;border-bottom:1px solid #e2e8f0;'>"
        f"<td style='{_th}padding-left:16px;text-transform:uppercase;'>CARRIER</td>"
        f"<td style='{_th}text-align:right;'>20GP</td>"
        f"<td style='{_th}text-align:right;'>40HQ</td>"
        f"<td style='{_th}'>VALID</td>"
        f"<td style='{_th}'>SVC</td>"
        "</tr>"
    )

    # ─── POL band (blue full-width) ──────────────────────────────
    rows.append(
        "<tr><td colspan='5' style='padding:0;'>"
        "<div style='background:#2553e2;color:#ffffff;padding:8px 16px;margin:2px 0 0;'>"
        f"<span style='font-size:14px;font-weight:800;letter-spacing:0.5px;'>{pol_code}</span>"
        f"<span style='color:#c7d2fe;margin-left:12px;font-size:12px;'>{pol_full}</span>"
        "</div></td></tr>"
    )

    # ─── Per-POD rows ────────────────────────────────────────────
    # Hard filter: drop any lane whose destination is NaN/empty so the email
    # never contains a "› NAN  No rates available" section.
    lane_intels = [
        ln for ln in lane_intels
        if str(ln.get("destination") or "").strip().upper() not in ("", "NAN", "NONE")
    ]
    if not lane_intels:
        return (
            "<p style='color:#94a3b8;font-style:italic;padding:12px 16px;"
            "border:1px dashed #e2e8f0;border-radius:6px;margin:8px 0;'>"
            "No rate data available for these lanes.</p>"
        )

    for lane in lane_intels:
        dest_code = str(lane.get("destination", "")).upper()
        carriers = lane.get("carriers") or []

        # Determine place from first carrier's parquet_place (reliable), fallback to pod_name
        first_place = ""
        if carriers:
            first_place = carriers[0].get("place_name") or carriers[0].get("parquet_place") or ""
        if not first_place:
            first_place = _POD_NAMES.get(dest_code, "")

        pod_main, pod_sub = _fmt_pod_header(dest_code, first_place)

        # Highlight CNEE's primary lane (if known) — amber "YOUR LANE" pill
        is_primary = bool(primary_dest) and dest_code == (primary_dest or "").upper()
        if is_primary:
            header_bg = "background:#fffbeb;border-left:4px solid #f59e0b;"
            primary_pill = (
                "<span style='background:#f59e0b;color:#ffffff;padding:3px 10px;"
                "border-radius:10px;font-size:10px;font-weight:800;letter-spacing:0.5px;"
                "margin-left:10px;vertical-align:middle;'>YOUR LANE</span>"
            )
        else:
            header_bg = ""
            primary_pill = ""

        rows.append(
            f"<tr><td colspan='5' style='padding:8px 16px 2px;{header_bg}'>"
            "<span style='color:#334155;font-size:13px;font-weight:700;'>"
            f"› {pod_main}"
            f"<span style='color:#64748b;font-weight:400;margin-left:8px;font-size:11px;'>"
            f"{pod_sub}</span>"
            f"{primary_pill}"
            "</span>"
            "</td></tr>"
        )

        if not carriers:
            rows.append(
                "<tr><td colspan='5' style='padding:4px 16px 6px;"
                "color:#94a3b8;font-style:italic;font-size:11px;'>"
                "No rates available for this lane.</td></tr>"
            )
            continue

        # Cap at TOP 3 carriers per POD (Nelson 2026-04-19: only show 3 best
        # options per destination — avoid overwhelming recipient with 4-5 rows).
        # Carriers already sorted cheapest-first in build_email → top 3 = 3 best.
        top_carriers = carriers[:3]

        # Render one row per carrier (cheapest already first — sorted in build_email)
        for c in top_carriers:
            carrier = str(c.get("carrier", "")).upper()
            is_best = (id(lane), carrier) == global_best_key

            r20 = _fmt_rate(c.get("rate_20"))
            r40 = _fmt_rate(c.get("rate_40"))
            valid = _fmt_validity(c.get("eff"), c.get("exp"))
            svc = _derive_svc(c.get("rate_type"), carrier)

            if is_best:
                row_bg = "#ecfdf5"
                accent = "#10b981"
                rate_color = "#064e3b"
                carrier_label = (
                    f"<strong style='color:#064e3b;'>{carrier}</strong> "
                    "<span style='background:#10b981;color:#ffffff;padding:1px 7px;"
                    "border-radius:10px;font-size:9px;font-weight:800;"
                    "letter-spacing:0.5px;margin-left:6px;vertical-align:middle;'>BEST</span>"
                )
            else:
                row_bg = "#ffffff"
                accent = "#e2e8f0"
                rate_color = "#334155"
                carrier_label = f"<strong style='color:#334155;'>{carrier}</strong>"

            rows.append(
                f"<tr style='background:{row_bg};border-left:3px solid {accent};'>"
                f"<td style='padding:5px 16px;font-size:12px;'>{carrier_label}</td>"
                f"<td style='padding:5px 12px;text-align:right;font-weight:700;"
                f"color:{rate_color};font-size:13px;'>{r20}</td>"
                f"<td style='padding:5px 12px;text-align:right;font-weight:700;"
                f"color:{rate_color};font-size:13px;'>{r40}</td>"
                f"<td style='padding:5px 12px;color:#475569;font-size:10px;'>{valid}</td>"
                f"<td style='padding:5px 12px;color:#475569;font-size:10px;'>{svc}</td>"
                "</tr>"
            )

    # ─── Footer info strip ───────────────────────────────────────
    rows.append(
        "<tr><td colspan='5' style='padding:8px 16px;background:#eff6ff;"
        "border-top:1px solid #dbeafe;'>"
        "<span style='color:#1e3a8a;font-size:11px;'>"
        "› <strong>Handling Fee:</strong> "
        "<span style='color:#2553e2;font-weight:700;'>$65</span> / shipment (US) · "
        "<span style='color:#ea580c;'>Canada: $85 / shipment</span>"
        "</span></td></tr>"
    )

    return (
        "<table cellpadding='0' cellspacing='0' width='100%' "
        "style='border-collapse:collapse;margin:10px 0;width:100%;"
        "font-family:Segoe UI,Arial,sans-serif;border:1px solid #e2e8f0;"
        "border-radius:6px;overflow:hidden;'>"
        f"<tbody>{''.join(rows)}</tbody></table>"
    )


def _wrap_branded_shell(
    intro_html: str,
    rate_table_html: str,
    cta_html: str,
    signature_html: str,
    recipient_label: str,
    week: int,
) -> str:
    """
    Wrap email body in Pudong Prime branded shell:
      [Header: logo + title + recipient + date]
      [Intro]
      [Rate table]
      [CTA text]
      [Footer CTA band with button]
      [Signature]
      [Domain bottom]
    """
    from datetime import date
    today = date.today()
    date_str = today.strftime("%d %b %Y")

    return f"""
<div style="margin:0;padding:0;background:#f1f5f9;font-family:Segoe UI,Arial,sans-serif;">
<table cellpadding="0" cellspacing="0" width="100%"
       style="background:#f1f5f9;padding:24px 0;">
  <tr><td align="center">
    <table cellpadding="0" cellspacing="0" width="680"
           style="max-width:680px;background:#ffffff;border-radius:12px;
                  overflow:hidden;box-shadow:0 1px 3px rgba(15,23,42,.06),
                  0 4px 16px rgba(15,23,42,.04);">

      <!-- ══ HEADER ══ -->
      <tr><td style="padding:28px 36px 20px;border-bottom:1px solid #f1f5f9;">
        <table width="100%"><tr>
          <td valign="top">
            <div style="font-size:19px;font-weight:800;color:#0f172a;
                        letter-spacing:-0.3px;line-height:1.2;">
              Pudong Prime Group
            </div>
            <div style="font-size:10px;color:#64748b;letter-spacing:2.5px;
                        margin-top:6px;font-weight:600;">
              OCEAN FREIGHT UPDATE
            </div>
            <div style="font-size:10px;color:#94a3b8;margin-top:8px;">
              A member of VLA · JCTrans · NVOCC
            </div>
          </td>
          <td valign="top" align="right">
            <div style="color:#2553e2;font-weight:700;font-size:15px;">
              {recipient_label}
            </div>
            <div style="color:#64748b;font-size:11px;margin-top:6px;">
              {date_str} · USD · WEEK {week}
            </div>
          </td>
        </tr></table>
      </td></tr>

      <!-- ══ INTRO ══ -->
      <tr><td style="padding:24px 36px 8px;color:#334155;font-size:14px;
                     line-height:1.7;">
        {intro_html}
      </td></tr>

      <!-- ══ RATE TABLE ══ -->
      <tr><td style="padding:0 20px;">
        {rate_table_html}
      </td></tr>

      <!-- ══ CTA TEXT ══ -->
      <tr><td style="padding:8px 36px 20px;color:#334155;font-size:14px;
                     line-height:1.7;">
        {cta_html}
      </td></tr>

      <!-- ══ FOOTER CTA BAND ══ -->
      <tr><td style="padding:22px 36px;background:#2553e2;">
        <table width="100%"><tr>
          <td style="color:#c7d2fe;font-size:14px;vertical-align:middle;">
            Ready to book?
          </td>
          <td align="right">
            <a href="mailto:nelson@pudongprime.vn?subject=Re: Ocean Freight Quote"
               style="background:#ffffff;color:#2553e2;padding:12px 24px;
                      border-radius:6px;text-decoration:none;font-weight:700;
                      font-size:13px;display:inline-block;
                      letter-spacing:0.3px;">
              Confirm Booking →
            </a>
          </td>
        </tr></table>
      </td></tr>

      <!-- ══ SIGNATURE ══ -->
      <tr><td style="padding:24px 36px;color:#475569;font-size:12px;
                     line-height:1.6;border-top:1px solid #f1f5f9;
                     white-space:pre-line;">
        {signature_html or ""}
      </td></tr>

      <!-- ══ DOMAIN FOOTER ══ -->
      <tr><td style="padding:14px 36px;background:#f8fafc;
                     text-align:right;font-size:11px;color:#94a3b8;
                     letter-spacing:0.5px;">
        pudongprime.com
      </td></tr>

    </table>
  </td></tr>
</table>
</div>
""".strip()


def _build_tokens(
    profile: dict | None,
    lane_intels: list[dict],
    pol: str,
    destinations: list[str],
) -> dict:
    """Assemble the tokens dict for the template renderer."""
    profile = profile or {}
    today = date.today()
    iso_week = today.isocalendar()[1]

    # Pick lane with highest-priority state for headline tokens
    from .template_selector import _STATE_PRIORITY
    headline = None
    best_sc = -1
    for ln in lane_intels:
        sc = _STATE_PRIORITY.get(str(ln.get("state", "STABLE")).upper(), 0)
        if sc > best_sc:
            headline = ln
            best_sc = sc

    headline = headline or {}
    mean_90d = headline.get("mean_90d")
    current = headline.get("current_rate_40hq")
    gap_to_mean = None
    if mean_90d and current:
        gap_to_mean = round(float(mean_90d) - float(current), 2)

    tokens = {
        # profile tokens
        "first_name": profile.get("first_name") or profile.get("name") or "",
        "company": profile.get("company") or profile.get("cnee_name") or "",
        "last_rate_quoted": profile.get("last_rate_quoted", ""),
        "days_since_last": profile.get("days_since_last", ""),
        "profile": profile,

        # lane tokens (forecast DISABLED per Nelson's request — "làm hư logic giá")
        "delta": "",  # disabled — market_engine delta unreliable with current data
        "current_rate_40hq": current or "",
        "prev_rate_40hq": "",
        "mean_90d": "",
        "forecast_next_week": "",  # disabled
        "gap_to_mean": "",
        "sample_size": headline.get("sample_size", 0),
        "confidence": headline.get("confidence", 0),
        "typical_pol": pol.upper(),
        "pol": pol.upper(),
        "typical_dest": destinations[0].upper() if destinations else "",

        # meta
        "week": iso_week,
        "year": today.year,
        "date": today.isoformat(),
        "suffix": "NELSON",

        # fallback intro — random pick from IntroTemplates pool (config.xlsx)
        # Each send gets different angle (Observation/Problem/Proof/Question/etc)
        "default_intro": _random_intro(profile),
        "default_closing": _random_closing(),

        # HTML chunks
        "rate_table_html": _render_rate_table(
            lane_intels,
            primary_dest=(profile.get("primary_dest") or "").upper() if profile else "",
        ),
    }

    # Signature priority:
    #   1. config.xlsx SIGNATURE field (HTML — Pudong Prime brand sig with logo cid)
    #   2. YAML defaults.signature (plain text fallback)
    sig_html_from_config = _load_signature_html_from_config()
    if sig_html_from_config:
        tokens["signature_html"] = sig_html_from_config  # raw HTML, bypass escape

    rules = load_rules()
    defaults_cfg = rules.get("defaults") or {}
    sig = defaults_cfg.get("signature")
    if sig and not sig_html_from_config:
        tokens["signature"] = sig  # plain text, will be escaped + <br>-wrapped
    sfx = defaults_cfg.get("subject_suffix")
    if sfx:
        tokens["suffix"] = sfx

    return tokens


# ────────────────────────────────────────────────────────────────────
# Signature loader — config.xlsx SIGNATURE field (HTML with logo cid)
# ────────────────────────────────────────────────────────────────────
_CONFIG_XLSX = (
    Path(__file__).resolve().parent.parent / "data" / "config.xlsx"
)
_SIG_CACHE: dict = {"mtime": 0.0, "html": None}
_TEMPLATE_CACHE: dict = {"mtime": 0.0, "intros": [], "closings": []}


def _load_intro_closing_templates() -> tuple[list[str], list[str]]:
    """Load IntroTemplates + ClosingTemplates from config.xlsx (pipe-split).
    Cached by mtime. Returns ([intros], [closings]) — fallback empty lists."""
    try:
        if not _CONFIG_XLSX.exists():
            return [], []
        mtime = _CONFIG_XLSX.stat().st_mtime
        if _TEMPLATE_CACHE.get("mtime") == mtime and _TEMPLATE_CACHE.get("intros"):
            return _TEMPLATE_CACHE["intros"], _TEMPLATE_CACHE["closings"]

        import openpyxl
        wb = openpyxl.load_workbook(str(_CONFIG_XLSX), data_only=True)
        intros, closings = [], []
        for row in wb.active.iter_rows(max_col=2, values_only=True):
            k = str(row[0] or "").strip().lower()
            v = str(row[1] or "").strip()
            if k == "introtemplates" and v:
                intros = [t.strip() for t in v.split("|") if t.strip()]
            elif k == "closingtemplates" and v:
                closings = [t.strip() for t in v.split("|") if t.strip()]

        _TEMPLATE_CACHE["mtime"] = mtime
        _TEMPLATE_CACHE["intros"] = intros
        _TEMPLATE_CACHE["closings"] = closings
        return intros, closings
    except Exception as e:
        log.warning("[builder] intro/closing templates load failed: %s", e)
        return [], []


def _random_intro(profile: dict) -> str:
    """Random pick from IntroTemplates pool, fallback to legacy default."""
    import random as _r
    intros, _ = _load_intro_closing_templates()
    if intros:
        return _r.choice(intros)
    return (f"Dear {profile.get('first_name', 'Team')},\n"
            f"Weekly Asia-US ocean freight update for {profile.get('company', 'your team')}.")


def _random_closing() -> str:
    """Random pick from ClosingTemplates pool, empty fallback."""
    import random as _r
    _, closings = _load_intro_closing_templates()
    if closings:
        return _r.choice(closings)
    return ""


def _load_signature_html_from_config() -> str:
    """
    Load the HTML signature from email_engine/data/config.xlsx (SIGNATURE row).
    Cached by mtime so anh edits file → next email picks up new sig.
    Returns '' if file missing or no SIGNATURE row.
    """
    try:
        if not _CONFIG_XLSX.exists():
            return ""
        mtime = _CONFIG_XLSX.stat().st_mtime
        if _SIG_CACHE.get("mtime") == mtime and _SIG_CACHE.get("html") is not None:
            return _SIG_CACHE["html"]

        import openpyxl
        wb = openpyxl.load_workbook(str(_CONFIG_XLSX), data_only=True)
        html = ""
        for row in wb.active.iter_rows(max_col=2, values_only=True):
            k = str(row[0] or "").strip().lower()
            if k == "signature":
                html = str(row[1] or "").strip()
                break

        _SIG_CACHE["mtime"] = mtime
        _SIG_CACHE["html"] = html
        return html
    except Exception as e:
        log.warning("[builder] config signature load failed: %s", e)
        return ""


_VN_POLS = frozenset({"HPH", "HCM", "SGN", "HAN", "DAD", "VUT", "CMT", "UIH", "DONG NAI"})


def build_email(
    cnee_email: str,
    pol: str,
    destinations: list[str],
    markup: float = 20.0,
    profile: dict | None = None,
    arb_origin: str | None = None,
) -> dict:
    """
    Build a complete smart email for one CNEE.

    Args:
        cnee_email:  recipient email (required)
        pol:         port of loading (HPH / HCM)
        destinations: list of POD port codes (e.g. ['USLAX','USLGB'])
        markup:      USD markup per container (default 20)
        profile:     optional CNEE profile dict (from Phase 02 intel/memory).
                     Keys: first_name, company, last_rate_quoted, days_since_last, ...

    Returns:
        {
            to: str,
            subject: str,
            html_body: str,
            meta: {
                template_id, dominant_state, lanes_analyzed,
                match_reason, markup, pol, destinations
            }
        }
    """
    pol = (pol or "HPH").strip().upper()
    # Drop NaN/None/empty strings — pandas reads empty cells as "nan" literal
    destinations = [
        d.strip().upper()
        for d in (destinations or [])
        if d and d.strip() and d.strip().lower() not in ("nan", "none")
    ]
    if not destinations:
        destinations = list(_DEFAULT_DESTINATIONS)

    if len(destinations) > _MAX_DESTINATIONS:
        log.info("[builder] truncating destinations %d→%d", len(destinations), _MAX_DESTINATIONS)
        destinations = destinations[:_MAX_DESTINATIONS]

    # 1. Get REAL rates from auto_rate_builder (proven correct).
    # auto_rate_builder uses proper Place/POD mapping + Exp>=today filter.
    # market_engine was returning inflated INLAND rates ($4,282 Nashville
    # matched as "USLAX"). auto_rate_builder returns actual PORT ocean freight.
    # Rate Table v2: keep LIST of carriers per POD (not just cheapest) for multi-row display
    rate_by_dest: dict[str, list[dict]] = {}
    rate_html_from_builder = ""
    try:
        import sys
        if str(Path(__file__).resolve().parent.parent / "core") not in sys.path:
            sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "core"))
        from auto_rate_builder import build_rate_table_for_customer
        # Non-VN POLs (PKG/BKK/SHA/etc) aren't in parquet — lookup via HCM + ARB surcharge layer
        lookup_pol = pol if pol in _VN_POLS else "HCM"
        arb_result = build_rate_table_for_customer(
            pol=lookup_pol, destinations=",".join(destinations), markup=float(markup or 0),
            top_per_route=3, arb_origin=arb_origin,
        )
        rate_html_from_builder = arb_result.get("html", "")
        if rate_html_from_builder:
            idx = rate_html_from_builder.find("<style")
            if idx == -1:
                idx = rate_html_from_builder.find("<table")
            if idx > 0:
                rate_html_from_builder = rate_html_from_builder[idx:]
        for rr in arb_result.get("rates", []):
            pod = str(rr.get("pod_code", "")).upper()
            r40 = rr.get("rate_40")
            r20 = rr.get("rate_20")
            if not (pod and r40):
                continue
            rate_by_dest.setdefault(pod, []).append({
                "rate_40":   int(r40),
                "rate_20":   int(r20) if r20 else 0,
                "carrier":   str(rr.get("carrier", "")),
                "eff":       rr.get("eff"),
                "exp":       rr.get("exp"),
                "rate_type": str(rr.get("rate_type", "")).upper(),
                "place_name": str(rr.get("place_name", "")),
                "parquet_place": str(rr.get("parquet_place", "")),
                "parquet_pod":   str(rr.get("parquet_pod", "")),
            })
        # Sort each POD's carriers by 40HQ rate ASC (cheapest first)
        for pod in rate_by_dest:
            rate_by_dest[pod].sort(key=lambda x: x.get("rate_40", 1e9))
    except Exception as e:
        log.warning("[builder] auto_rate_builder unavailable: %s", e)

    # 2. Per-lane market analysis (for STATE classification: URGENT/STABLE/etc.)
    # market_engine state is used for template selection + subject line only.
    # Dollar amounts come from auto_rate_builder above (NOT market_engine).
    lane_intels = []
    for dest in destinations:
        intel = analyze_lane(pol, dest)
        intel["pol"] = pol

        # Rate Table v2: attach full carrier list (multi-row per POD)
        carriers = rate_by_dest.get(dest, [])
        intel["carriers"] = carriers  # list[dict] — may be empty

        # Override dollar amounts with auto_rate_builder's cheapest (first in sorted list)
        if carriers:
            top = carriers[0]
            intel["current_rate_40hq"] = top["rate_40"]
            intel["current_rate_20gp"] = top.get("rate_20", 0)
            intel["carrier"] = top.get("carrier", "")
        elif intel.get("current_rate_40hq"):
            r40 = intel["current_rate_40hq"]
            intel["current_rate_40hq"] = float(r40) + float(markup or 0)
            intel["current_rate_20gp"] = round(float(r40) * 0.78 + float(markup or 0))

        fcst = intel.get("forecast_next_week")
        if fcst:
            intel["forecast_next_week"] = float(fcst) + float(markup or 0)
        lane_intels.append(intel)

    # 2. Dominant state across all lanes
    dom = dominant_state(lane_intels)

    # 3. Match template
    tmpl = tmpl_match(destinations, [dom])

    # 4. Build tokens
    tokens = _build_tokens(profile, lane_intels, pol, destinations)

    # Rate Table v2 renderer: side-by-side HPH/HCM layout with inland styling.
    # Used when fast_bulk_default pod_list is available (dual-POL workflow).
    # Falls back to legacy auto_rate_builder HTML for single-POL or legacy paths.
    rate_table_v2_html = ""
    try:
        import sys as _sys
        _tpl_dir = str(Path(__file__).resolve().parent.parent / "templates")
        if _tpl_dir not in _sys.path:
            _sys.path.insert(0, _tpl_dir)
        from rate_table_renderer import render_dual_rate_table

        # Load pod_list from YAML (includes type/gateway metadata)
        from shared.paths import DEFAULT_ROUTES_CFG
        import yaml  # type: ignore
        _pod_list_full: list[dict] = []
        try:
            _yml_path = Path(DEFAULT_ROUTES_CFG) if Path(DEFAULT_ROUTES_CFG).exists() else (
                Path(__file__).resolve().parent.parent / "config" / "default_routes.yaml"
            )
            with open(_yml_path, "r", encoding="utf-8") as _fh:
                _yd = yaml.safe_load(_fh) or {}
            _fb = _yd.get("fast_bulk_default", {})
            if isinstance(_fb, dict):
                _pod_list_full = [
                    p for p in (_fb.get("pod_list") or []) if isinstance(p, dict)
                ]
        except Exception as _e:
            log.debug("[builder] pod_list load for renderer: %s", _e)

        # Only use v2 renderer when pol_list has dual POLs AND pod_list is available
        _dual_pol = isinstance(_yd.get("fast_bulk_default"), dict) and (
            len(_yd["fast_bulk_default"].get("pol_list") or []) >= 2
        ) if '_yd' in dir() else False

        if _pod_list_full and _dual_pol:
            # WHY: CNEE-specific destinations + 10 YAML defaults can merge to >10 POD,
            # then truncation drops the tail (typically inland IPI PODs: USCHI/USDAL/USDEN).
            # v2 always renders exactly the 10 YAML pod_list → query rates against THAT list
            # directly so every row has data, independent of per-CNEE destinations.
            _full_codes = ",".join(
                str(p.get("code") or "").upper()
                for p in _pod_list_full if p.get("code")
            )
            _hph_rates: list[dict] = []
            _hcm_rates: list[dict] = []
            try:
                from auto_rate_builder import build_rate_table_for_customer as _brtc
                _hph_result = _brtc(
                    pol="HPH",
                    destinations=_full_codes,
                    markup=float(markup or 0),
                    top_per_route=3,
                    arb_origin=arb_origin if pol.upper() == "HPH" else None,
                )
                _hph_rates = _hph_result.get("rates", []) or []
                _hcm_result = _brtc(
                    pol="HCM",
                    destinations=_full_codes,
                    markup=float(markup or 0),
                    top_per_route=3,
                    arb_origin=arb_origin if pol.upper() == "HCM" else None,
                )
                _hcm_rates = _hcm_result.get("rates", []) or []
            except Exception as _e2:
                log.warning("[builder] v2 dual-POL full rate query failed: %s", _e2)
                # Fallback: use arb_result (may be incomplete if destinations truncated)
                _hph_rates = arb_result.get("rates", []) if pol.upper() == "HPH" else []
                _hcm_rates = arb_result.get("rates", []) if pol.upper() == "HCM" else []

            from datetime import date as _date
            _week = _date.today().isocalendar()[1]
            rate_table_v2_html = render_dual_rate_table(
                hph_rates=_hph_rates,
                hcm_rates=_hcm_rates,
                pod_list=_pod_list_full,
                week=_week,
            )
            log.info(
                "[builder] rate_table_v2: dual HPH+HCM rendered, %d pods, %d+%d rates",
                len(_pod_list_full), len(_hph_rates), len(_hcm_rates),
            )
        elif _pod_list_full:
            # Single POL mode with v2 styling
            _rates = arb_result.get("rates", []) if 'arb_result' in dir() else []
            _hph = _rates if pol.upper() == "HPH" else []
            _hcm = _rates if pol.upper() != "HPH" else []
            from datetime import date as _date
            _week = _date.today().isocalendar()[1]
            rate_table_v2_html = render_dual_rate_table(
                hph_rates=_hph,
                hcm_rates=_hcm,
                pod_list=_pod_list_full,
                week=_week,
            )
    except Exception as _rv2_err:
        log.warning("[builder] rate_table_v2 renderer failed (%s) — using legacy", _rv2_err)

    # Priority: v2 side-by-side renderer > auto_rate_builder HTML > _render_rate_table()
    if rate_table_v2_html:
        tokens["rate_table_html"] = rate_table_v2_html
    elif rate_html_from_builder:
        tokens["rate_table_html"] = rate_html_from_builder

    # 5. Render intro + rate table + cta + signature (standard structure).
    # The rate table itself IS the Pudong Prime visual upgrade — no shell wrap.
    rendered = render_email(tmpl, tokens)

    _rate_table_ver = "v2_dual" if rate_table_v2_html else ("v1_arb" if rate_html_from_builder else "legacy")
    return {
        "to": cnee_email,
        "subject": rendered["subject"],
        "html_body": rendered["html_body"],
        "meta": {
            "template_id": tmpl.get("id", "unknown"),
            "dominant_state": dom,
            "lanes_analyzed": len(lane_intels),
            "match_reason": tmpl.get("match_reason", ""),
            "markup": markup,
            "pol": pol,
            "destinations": destinations,
            "lane_states": [ln.get("state") for ln in lane_intels],
            "rate_table": _rate_table_ver,
        },
    }
