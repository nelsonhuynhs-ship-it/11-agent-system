# -*- coding: utf-8 -*-
"""
refresh-v14.py — Refresh ERP v14 from Parquet
==============================================
Called by VBA Ribbon "Refresh Rates" button.
Updates: Pricing Dashboard, ChargeBreakdown, RateVersions sheets.
Preserves: Quotes, CRM, Active Jobs, Markup_Store (untouched).

Usage:
    python refresh-v14.py [path_to_xlsm]
"""
import os, sys, json
from datetime import datetime, timedelta
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')

# ── Paths ──
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.normpath(os.path.join(SCRIPT_DIR, '..', '..'))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

# Try shared.paths, fallback to hardcoded
try:
    from shared import paths as sp
    PARQUET = str(sp.PARQUET_FILE)
    PRICING_DATA_DIR = str(sp.PRICING_DATA)
except ImportError:
    PARQUET = "D:/OneDrive/NelsonData/pricing/Cleaned_Master_History.parquet"
    PRICING_DATA_DIR = "D:/OneDrive/NelsonData/pricing"


def _resolve_puc_file() -> str:
    """Auto-detect latest PUC file.

    Unified 2026-04-12: drop new PUC file into processed/ alongside FAK/SCFI/FIX
    (one folder for all rate files). Search order:
      1. processed/        (new canonical location)
      2. rate-tables/      (legacy fallback — being phased out)
      3. pricing/ root     (PUC_SOC.xlsx legacy layout)
    Returns path to newest PUC*.xlsx by mtime.
    """
    import glob as _g
    legacy = os.path.join(PRICING_DATA_DIR, "PUC_SOC.xlsx")
    if os.path.exists(legacy):
        return legacy
    search_dirs = [
        os.path.join(PRICING_DATA_DIR, "processed"),
        os.path.join(PRICING_DATA_DIR, "rate-tables"),
        PRICING_DATA_DIR,
    ]
    candidates = []
    for d in search_dirs:
        if os.path.isdir(d):
            candidates.extend(_g.glob(os.path.join(d, "PUC*.xlsx")))
    candidates = [f for f in candidates if "PUC_SOC" not in os.path.basename(f)]
    candidates.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    return candidates[0] if candidates else legacy


PUC_FILE = _resolve_puc_file()

# ERP file — passed as arg or default
if len(sys.argv) > 1:
    ERP_FILE = sys.argv[1]
else:
    ERP_FILE = "D:/OneDrive/NelsonData/erp/ERP_Master_v14.xlsm"

print(f"[refresh-v14] {datetime.now().strftime('%H:%M:%S')}")
print(f"  Parquet: {PARQUET}")
print(f"  ERP: {ERP_FILE}")

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Colors ──
C_SEARCH = PatternFill("solid", fgColor="FFF7ED")
C_DRY    = PatternFill("solid", fgColor="1E40AF")
C_MISC   = PatternFill("solid", fgColor="6B7280")
C_RF     = PatternFill("solid", fgColor="7C3AED")
thin = Side(style='thin', color='CBD5E1')
card = Border(left=thin, right=thin, top=thin, bottom=thin)

DATA_START_ROW = 2
MAX_ROWS = 5000


# ===============================================================
# HELPER: normalize container type variants (P4 2026-04-11)
# ===============================================================
def normalize_container_types(df_in: pd.DataFrame) -> pd.DataFrame:
    """Collapse "45'HQ" into "45HQ" so pivot sees one container column.

    Parquet contains ~105K rows tagged "45'HQ" (with a literal quote character)
    alongside the standard "45HQ". Without this pre-pivot normalization, the
    pivot_table creates two separate columns, and a post-pivot rename causes
    a column-name collision that silently overwrites data. This helper MUST
    run BEFORE any pivot_table on Container_Type.

    Added per plan 260411-2121-erp-workflow-upgrade phase-04.
    """
    if 'Container_Type' in df_in.columns:
        df_in = df_in.copy()
        df_in['Container_Type'] = df_in['Container_Type'].replace({"45'HQ": "45HQ"})
    return df_in


def normalize_commodity_display(df_in: pd.DataFrame) -> pd.DataFrame:
    """Shorten verbose Commodity strings for display in the ERP sheet.

    Nelson flagged 2026-04-12: ONE carrier's REEFER commodity reads
    "REEFER FAK (NOT VALID FOR SEASONAL COMMODITIES PHARMACEUTICALS HIGH VALUE)\\n"
    which bloats row height + hides other columns. Shortening rules (applied
    in order):

    1. Trim whitespace + trailing newlines
    2. If contains "REEFER" → set to "REEFER" (Nelson's explicit preference)
    3. Else if has parenthesis detail "(... )" → keep text before " ("
    4. Else keep as-is

    Pure display layer — does NOT affect dedup or joins (applied AFTER pivot).
    """
    if 'Commodity' not in df_in.columns:
        return df_in
    df_in = df_in.copy()

    def _shorten(raw):
        if raw is None:
            return raw
        try:
            s = str(raw).strip().strip('\r\n').strip()
        except Exception:
            return raw
        if not s:
            return s
        upper = s.upper()
        if 'REEFER' in upper:
            return 'REEFER'
        # Strip parenthetical detail: "FAK (Excluding Garment)" -> "FAK"
        paren_idx = s.find(' (')
        if paren_idx > 0:
            return s[:paren_idx].strip()
        return s

    df_in['Commodity'] = df_in['Commodity'].map(_shorten)
    return df_in


def clean_pod_contamination(df_in: pd.DataFrame) -> pd.DataFrame:
    """Strip surcharge/note fragments that leaked into the POD column.

    Nelson flagged 2026-04-16: HPL SOC FIX rows show POD values like
    'SUBJECT TO EFS', 'included EFS' in the Pricing sheet. Root cause is a
    legacy convert_pricing.py ingest that copied a header/note row's text
    into POD instead of the actual port code.

    Strategy (defensive layer — see docs/vba-gotchas.md):
      1. Detect rows where POD matches surcharge-note patterns
      2. Merge the bad POD text into Note (preserves the info for audit)
      3. Swap POD <- Place if Place is a valid port/city
      4. Drop the row if Place is also empty (unrecoverable)

    Should run BEFORE any pivot/dedup so cleaned POD participates in keys.
    Prints a single-line summary so the refresh log shows the fix.
    """
    if 'POD' not in df_in.columns:
        return df_in

    pod_str = df_in['POD'].astype(str).str.strip()
    bad_pat = r'SUBJECT\s*TO|^incl(?:uded)?\s*EFS|^EFS\b'
    bad = pod_str.str.contains(bad_pat, case=False, na=False, regex=True)
    n_bad = int(bad.sum())
    if n_bad == 0:
        return df_in

    df_in = df_in.copy()

    # 1. Preserve original POD text into Note
    if 'Note' in df_in.columns:
        old_pod = df_in.loc[bad, 'POD'].astype(str).str.strip()
        cur_note = df_in.loc[bad, 'Note'].fillna('').astype(str).str.strip()
        sep = cur_note.where(cur_note == '', ' | ').where(cur_note != '', '')
        df_in.loc[bad, 'Note'] = (cur_note + sep + old_pod).str.strip(' |').str.strip()

    # 2. Swap POD <- Place where Place is valid, else drop
    if 'Place' in df_in.columns:
        place = df_in.loc[bad, 'Place'].astype(str).str.strip()
        place_ok = (place != '') & (place.str.lower() != 'nan')
        bad_idx = df_in.loc[bad].index
        recover_idx = bad_idx[place_ok.values]
        drop_idx = bad_idx[~place_ok.values]
        if len(recover_idx) > 0:
            df_in.loc[recover_idx, 'POD'] = df_in.loc[recover_idx, 'Place'].values
        if len(drop_idx) > 0:
            df_in = df_in.drop(drop_idx).reset_index(drop=True)
        print(f"  POD cleanup: {n_bad:,} contaminated -> {len(recover_idx):,} recovered via Place, {len(drop_idx):,} dropped")
    else:
        df_in = df_in.loc[~bad].reset_index(drop=True)
        print(f"  POD cleanup: {n_bad:,} rows dropped (no Place fallback)")

    return df_in


# ===============================================================
# STEP 1: Load Parquet
# ===============================================================
print("\n[1/4] Loading Parquet...")
df = pd.read_parquet(PARQUET)
print(f"  {len(df):,} rows total")

# Normalize container variants PRE-pivot (fixes 45'HQ column collision)
df = normalize_container_types(df)

# Clean POD contamination (HPL FIX legacy — 'SUBJECT TO EFS' / 'included EFS')
df = clean_pod_contamination(df)

df['Eff'] = pd.to_datetime(df['Eff'], errors='coerce')
df['Exp'] = pd.to_datetime(df['Exp'], errors='coerce')
df['RefreshDate'] = df['Eff'].where(df['Eff'].notna(), df['Exp'])

missing_eff_fix = 0
if 'Rate_Type' in df.columns:
    missing_eff_fix = int(((df['Rate_Type'] == 'FIX') & df['Eff'].isna()).sum())
if missing_eff_fix:
    print(f"  FIX rows missing Eff: {missing_eff_fix:,} (using Exp as fallback)")

# Filter: 15d -> 30d -> 90d fallback
for days in [15, 30, 90]:
    cutoff = datetime.now() - timedelta(days=days)
    df_recent = df[df['RefreshDate'] >= cutoff].copy()
    if len(df_recent) >= 100:
        print(f"  {days}-day filter: {len(df_recent):,} rows")
        break
else:
    df_recent = df.copy()
    print(f"  Using all dates: {len(df_recent):,} rows")

# ===============================================================
# STEP 2: Build Pricing Dashboard data (Total Ocean Freight pivot)
# ===============================================================
print("\n[2/4] Building Pricing Dashboard...")
if 'Charge_Name' in df_recent.columns:
    # Primary: "Total Ocean Freight" (FAK, SCFI, older FIX files)
    df_tof = df_recent[df_recent['Charge_Name'].str.contains('Total Ocean', case=False, na=False)].copy()

    # 2026-04-13 FIX: FIX NO.22+ files only have "Base Ocean Freight" (no Total
    # Ocean row). For FIX Rate_Type rows that have NO "Total Ocean" entry, fall
    # back to "Base Ocean Freight" so they appear in the pricing sheet.
    if 'Rate_Type' in df_recent.columns:
        fix_rows = df_recent[df_recent['Rate_Type'] == 'FIX']
        fix_with_tof = fix_rows[fix_rows['Charge_Name'].str.contains('Total Ocean', case=False, na=False)]
        # Routes that already have Total Ocean → exclude from fallback
        tof_keys = set()
        route_cols = [c for c in ['POL', 'POD', 'Place', 'Carrier', 'Container_Type', 'Note'] if c in fix_with_tof.columns]
        if route_cols and len(fix_with_tof) > 0:
            tof_keys = set(fix_with_tof[route_cols].apply(lambda r: '|'.join(str(v) for v in r), axis=1))
        # FIX rows with Base Ocean Freight that DON'T have a Total Ocean row
        fix_base = fix_rows[fix_rows['Charge_Name'].str.contains('Base Ocean', case=False, na=False)].copy()
        if len(fix_base) > 0 and tof_keys:
            fix_base_keys = fix_base[route_cols].apply(lambda r: '|'.join(str(v) for v in r), axis=1)
            fix_base = fix_base[~fix_base_keys.isin(tof_keys)]
        if len(fix_base) > 0:
            print(f"  FIX fallback (Base Ocean → Total): {len(fix_base):,} rows")
            df_tof = pd.concat([df_tof, fix_base], ignore_index=True)
else:
    df_tof = df_recent.copy()

# 2026-04-12: include Rate_Type in id_cols so it survives the pivot.
# After dedup we rename it to "Source" so col 9 shows FAK/SCFI/FIX.
# 2026-04-21 Phase 3: add Contract, Group Rate, Group_Code so they survive
# the pivot as index columns (metadata — one value per route key).
id_cols = [c for c in ['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note', 'Rate_Type',
                        'Contract', 'Group Rate', 'Group_Code']
           if c in df_tof.columns]
rate_col = 'Amount' if 'Amount' in df_tof.columns else 'Rate'

# 2026-04-21 Phase 3: fill NaN in metadata cols so pivot_table (dropna=True default)
# does not silently drop rows. Non-ONE carriers have no Group_Code; some may have
# no Contract or Group Rate. Empty string keeps rows alive through pivot.
for _meta_col in ['Contract', 'Group Rate', 'Group_Code']:
    if _meta_col in df_tof.columns:
        df_tof[_meta_col] = df_tof[_meta_col].fillna('')

# 2026-04-13: FIX NO.22 has Eff=NaT for ALL rows. pivot_table(dropna=True)
# silently drops them. Instead: fill NaT Eff with Exp before pivoting so
# rows survive the default dropna behaviour.
if 'Eff' in df_tof.columns and 'Exp' in df_tof.columns:
    eff_nat_count = int(df_tof['Eff'].isna().sum())
    if eff_nat_count:
        df_tof['Eff'] = df_tof['Eff'].fillna(df_tof['Exp'])
        print(f"  Filled {eff_nat_count:,} NaT Eff with Exp (FIX rate fallback)")

pivot = df_tof.pivot_table(
    index=id_cols, columns='Container_Type', values=rate_col,
    aggfunc='first'
).reset_index()

# NOTE: post-pivot "45'HQ" → "45HQ" rename removed 2026-04-11 P4.
# Normalization happens pre-pivot via normalize_container_types() — see top of file.

# 2026-04-12: rename Rate_Type -> Source (col 9 in Pricing Dry/Reefer)
if 'Rate_Type' in pivot.columns:
    pivot = pivot.rename(columns={'Rate_Type': 'Source'})

# 2026-04-12: shorten verbose commodity strings for display
pivot = normalize_commodity_display(pivot)

# Dedup: latest per route+carrier+note+source
# 2026-04-13: Added 'Source' so FAK/SCFI/FIX for the same route all survive.
# Without Source, dedup keeps only the newest → FIX rows disappear when FAK is newer.
route_key = [c for c in ['POL', 'POD', 'Place', 'Carrier', 'Note', 'Source'] if c in pivot.columns]
if route_key:
    sort_col = 'RefreshDate' if 'RefreshDate' in pivot.columns else 'Eff'
    pivot = pivot.sort_values(sort_col, ascending=False, na_position='last')
    pivot = pivot.drop_duplicates(subset=route_key, keep='first')

sort_col = 'RefreshDate' if 'RefreshDate' in pivot.columns else 'Eff'
if sort_col in pivot.columns:
    pivot = pivot.sort_values(sort_col, ascending=False, na_position='last')

# Count RF data
rf_cols = [c for c in ['20RF', '40RF'] if c in pivot.columns]
rf_count = pivot[rf_cols].notna().any(axis=1).sum() if rf_cols else 0
print(f"  {len(pivot):,} unique routes ({rf_count} with reefer)")

dry_cols = [c for c in ['20GP', '40GP', '40HQ', '45HQ', '40NOR'] if c in pivot.columns]
reefer_only_notes = pivot['Note'].astype(str).str.contains('REEFER', case=False, na=False) if 'Note' in pivot.columns else pd.Series(False, index=pivot.index)
reefer_only_commodities = pivot['Commodity'].astype(str).str.contains('REEFER', case=False, na=False) if 'Commodity' in pivot.columns else pd.Series(False, index=pivot.index)
dry_mask = pivot[dry_cols].notna().any(axis=1) if dry_cols else pd.Series(False, index=pivot.index)
reefer_rate_mask = pivot[rf_cols].notna().any(axis=1) if rf_cols else pd.Series(False, index=pivot.index)
reefer_mask = reefer_rate_mask | reefer_only_notes | reefer_only_commodities

pivot_dry = pivot[dry_mask & ~reefer_only_notes & ~reefer_only_commodities].copy()
pivot_reefer = pivot[reefer_mask].copy()

# Fix: If a route made it to reefer but has no 20RF or 40RF rate, drop it.
if rf_cols:
    pivot_reefer = pivot_reefer.dropna(how='all', subset=rf_cols)

# ===============================================================
# STEP 3: Build ChargeBreakdown data (per route+carrier+container)
# ===============================================================
print("\n[3/4] Building ChargeBreakdown...")

# Exclude "Total Ocean Freight" — we want individual charges
df_charges = df_recent[~df_recent['Charge_Name'].str.contains('Total Ocean', case=False, na=False)].copy()

# Get latest charges per route+carrier+container+charge
charge_key = [c for c in ['POL', 'POD', 'Place', 'Carrier', 'Container_Type', 'Charge_Name', 'Note']
              if c in df_charges.columns]
if 'RefreshDate' in df_charges.columns:
    df_charges = df_charges.sort_values('RefreshDate', ascending=False, na_position='last')
    df_charges = df_charges.drop_duplicates(subset=charge_key, keep='first')

print(f"  {len(df_charges):,} charge rows (deduped)")

# Build lookup: key -> list of (charge_name, amount)
charge_lookup = defaultdict(list)
for _, row in df_charges.iterrows():
    key = f"{row.get('POL','')}" \
          f"|{row.get('POD','')}" \
          f"|{row.get('Place','')}" \
          f"|{row.get('Carrier','')}" \
          f"|{row.get('Container_Type','')}" \
          f"|{row.get('Note','')}"
    key = key.upper()
    charge_name = str(row.get('Charge_Name', ''))
    amount = float(row.get('Amount', 0)) if pd.notna(row.get('Amount')) else 0
    if amount > 0:
        charge_lookup[key].append((charge_name, amount))

print(f"  {len(charge_lookup):,} unique route+container keys")

# ===============================================================
# STEP 4: Build RateVersions (latest Source_File per Rate_Type)
# 2026-04-12: Sort by FILENAME upload date (Harry's send date) not
# by Eff date. Also produce short display labels for ribbon.
# ===============================================================
import re as _re_rv

def _parse_upload_date(source_file: str) -> str:
    """Extract YYYYMMDD date key from filename for sort order.

    Primary:  'FAK_20260410_...xlsx'     -> '20260410' (rate_importer canonical)
    Fallback: 'Update rate ... 14 APR NO. 1.xlsx' -> '20260414'
              (when rate_importer prefix missing — e.g. file dropped manually
              into incoming/ without the FAK_YYYYMMDD_ prefix).

    Returns '00000000' when no date signal found — sorts to bottom.
    """
    if not source_file:
        return '00000000'
    s = str(source_file)
    # Primary: 8-digit prefix
    m = _re_rv.search(r'(\d{8})', s)
    if m:
        return m.group(1)
    # Fallback: only when filename looks like an email-downloaded rate sheet
    # (keywords present). Ignores legacy archive tags like "FAK 28NOV NO.1"
    # which are just version markers, not dated filenames.
    if _re_rv.search(r'UPDATE\s*RATE|US[\s_]*CANADA|US[\s_]*CAD|RATE\s*SHEET|SCFI\s*CONTRACT', s, _re_rv.IGNORECASE):
        m2 = _re_rv.search(r'(\d{1,2})\s*([A-Z]{3})', s, _re_rv.IGNORECASE)
        if m2:
            _MONTHS = {'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
                       'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12}
            mo = _MONTHS.get(m2.group(2).upper())
            if mo:
                from datetime import datetime
                return f"{datetime.now().year}{mo:02d}{int(m2.group(1)):02d}"
    return '00000000'


def _short_fak_label(source_file: str) -> str:
    """'FAK_20260410_Update rate to US CANADA_ 10 APR NO. 1.xlsx' -> '10APR No1'.
    Ribbon displays as 'FAK: 10APR No1' (VBA prepends 'FAK: ').
    """
    if not source_file:
        return '?'
    s = str(source_file)
    # Match pattern like "08 APR NO. 1" or "10 APR NO.2"
    m = _re_rv.search(r'(\d{1,2})\s*([A-Z]{3})\s*NO\.?\s*(\d+)', s, _re_rv.IGNORECASE)
    if m:
        day, mon, no = m.group(1), m.group(2).upper(), m.group(3)
        return f'{int(day):02d}{mon} No{no}'
    # Fallback: try upload date prefix
    m2 = _re_rv.search(r'(\d{4})(\d{2})(\d{2})', s)
    if m2:
        return f'{int(m2.group(3)):02d}/{int(m2.group(2)):02d}'
    return s[:20]


def _short_scfi_label(source_file: str) -> str:
    """'SCFI_20260403_HPL SCFI CONTRACT 40.xlsx' -> 'No40'.
    Ribbon displays as 'SCFI: No40'.
    """
    if not source_file:
        return '?'
    s = str(source_file)
    m = _re_rv.search(r'(?:CONTRACT|N|NO\.?)\s*(\d+)', s, _re_rv.IGNORECASE)
    if m:
        return f'No{m.group(1)}'
    return '?'


def _short_fix_label(source_file: str) -> str:
    """'Fixed Rate Summary Table NO.21.xlsx' -> 'No21'.
    Ribbon displays as 'FIX: No21'.
    """
    if not source_file:
        return '?'
    s = str(source_file)
    m = _re_rv.search(r'NO\.?\s*(\d+)', s, _re_rv.IGNORECASE)
    if m:
        return f'No{m.group(1)}'
    return '?'


def _short_puc_label(puc_filename: str) -> str:
    """'PUC MAR 2026.xlsx' -> 'PUC MAR 2026'."""
    if not puc_filename:
        return 'PUC ?'
    return str(puc_filename).replace('.xlsx', '').strip()


print("\n  Building RateVersions...")
versions = []

def _latest_file(df_in, rate_type):
    """Pick latest Source_File for given rate_type, sorted by filename upload date."""
    if 'Rate_Type' not in df_in.columns:
        return None
    sub = df_in[df_in['Rate_Type'] == rate_type].copy()
    if len(sub) == 0:
        return None
    sub['_upload_key'] = sub['Source_File'].astype(str).map(_parse_upload_date)
    sub = sub.sort_values(['_upload_key', 'RefreshDate'], ascending=[False, False], na_position='last')
    return str(sub['Source_File'].iloc[0])

# FAK: sort by filename upload date
latest_fak_file = _latest_file(df_recent, 'FAK')
if latest_fak_file:
    versions.append(('FAK', latest_fak_file, _short_fak_label(latest_fak_file)))

# SCFI
latest_scfi_file = _latest_file(df_recent, 'SCFI')
if latest_scfi_file:
    versions.append(('SCFI', latest_scfi_file, _short_scfi_label(latest_scfi_file)))

# FIX
latest_fix_file = _latest_file(df_recent, 'FIX')
if latest_fix_file:
    versions.append(('FIX', latest_fix_file, _short_fix_label(latest_fix_file)))

# PUC — resolved at top via _resolve_puc_file() (processed/ first, rate-tables fallback)
if os.path.exists(PUC_FILE):
    puc_name = os.path.basename(PUC_FILE).replace('.xlsx', '')
    versions.append(('PUC', os.path.basename(PUC_FILE), _short_puc_label(puc_name)))

for vtype, vfull, vshort in versions:
    print(f"  {vtype}: {vshort}  (full: {vfull})")

# ===============================================================
# STEP 5: Write to Excel
# ===============================================================
print(f"\n[4/4] Writing to {ERP_FILE}...")

wb = openpyxl.load_workbook(ERP_FILE, keep_vba=True)

col_map_dry = {
    'POL': 1, 'POD': 2, 'Place': 3, 'Carrier': 4, 'Commodity': 5,
    'Eff': 6, 'Exp': 7, 'Note': 8, 'Source': 9,
    '20GP': 10, '40GP': 11, '40HQ': 12, '45HQ': 13,
    '40NOR': 14,
    # 2026-04-21 Phase 3: hidden cols — Contract #, Group Rate text, ONE group code
    'Contract': 15, 'Group Rate': 16, 'Group_Code': 17,
}

col_map_reefer = {
    'POL': 1, 'POD': 2, 'Place': 3, 'Carrier': 4, 'Commodity': 5,
    'Eff': 6, 'Exp': 7, 'Note': 8, 'Source': 9,
    '20RF': 10, '40RF': 11,
    # 2026-04-21 Phase 3: hidden cols — same 3 as Dry
    'Contract': 15, 'Group Rate': 16, 'Group_Code': 17,
}

def get_or_create_sheet(book, name):
    if name in book.sheetnames:
        return book[name]
    return book.create_sheet(name)

# 2026-04-21 Phase 3: display names for header row (col_name may differ from
# what Nelson sees — e.g. 'Group_Code' stored as underscore, header = 'Group Code')
_HEADER_DISPLAY = {
    'Group_Code': 'Group Code',
}

# Cols that are TEXT metadata — skip int(float()) conversion even though col_idx >= 10
_TEXT_COLS = {'Contract', 'Group Rate', 'Group_Code'}

# Cols 15, 16, 17 — hidden from Nelson's normal view (VBA reads them for tooltip/email)
_HIDDEN_COL_INDICES = {15, 16, 17}

def write_pricing_sheet(ws, data, view_mode, current_col_map):
    ws.delete_rows(1, ws.max_row)  # Clear entire sheet since we build from scratch

    # RESET view state — Excel saves scroll position per sheet. If Nelson
    # scrolled Reefer to row 47 then saved, freeze_panes + topLeftCell
    # persist across refreshes and make row 1 invisible on open.
    ws.sheet_view.topLeftCell = None
    ws.freeze_panes = None  # clear, will re-apply "J2" below

    # Write headers at row 1
    for col_name, col_idx in current_col_map.items():
        cell = ws.cell(1, col_idx)
        cell.value = _HEADER_DISPLAY.get(col_name, col_name)
        cell.font = Font(bold=True, name='Segoe UI', size=10, color='FFFFFF')
        cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # Add Freeze Panes at J2
    ws.freeze_panes = "J2"

    row_count = 0
    for _, row in data.head(MAX_ROWS).iterrows():
        r_idx = DATA_START_ROW + row_count
        for col_name, col_idx in current_col_map.items():
            val = row.get(col_name)
            try:
                if val is None or (isinstance(val, float) and np.isnan(val)):
                    continue
            except (TypeError, ValueError):
                pass
            cell = ws.cell(r_idx, col_idx)
            # 2026-04-21 Phase 3: cols 15/16/17 are text metadata, NOT numeric rates
            if col_idx >= 10 and col_name not in _TEXT_COLS:
                try:
                    cell.value = int(float(val))
                    cell.number_format = '#,##0'
                except (ValueError, TypeError):
                    pass
            else:
                if isinstance(val, pd.Timestamp):
                    cell.value = val.to_pydatetime()
                    if col_name in ("Eff", "Exp"):
                        cell.number_format = "dd-mmm"
                else:
                    cell.value = str(val)
            cell.font = Font(name='Segoe UI', size=10)
            cell.border = card
        row_count += 1

    # 2026-04-12: Reset row heights to fixed 18pt so rows don't overlap/stack.
    # Also disable wrap_text on data rows (was pulling row height huge when
    # ONE's verbose REEFER commodity had embedded newlines).
    for r in range(DATA_START_ROW, DATA_START_ROW + row_count):
        ws.row_dimensions[r].hidden = False
        ws.row_dimensions[r].height = 18
    # Header row slightly taller for readability
    ws.row_dimensions[1].height = 22

    ws.auto_filter.ref = f"A1:Q{max(DATA_START_ROW, DATA_START_ROW + row_count - 1)}"

    # 2026-04-12: Unhide visible cols 1-14 (Dry) / 1-11 (Reefer) + safety buffer to 14.
    # 2026-04-21 Phase 3: Expanded to 17 then hide cols 15/16/17 explicitly below.
    # Previously unhid up to col 16, which is now a hidden metadata col. Unhide 1-14
    # first, then apply hide for 15/16/17 to ensure clean state regardless of prior save.
    for c in range(1, 15):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].hidden = False

    # 2026-04-12: Set sane default widths for commonly-narrow cols
    col_widths = {
        1: 6,    # POL
        2: 13,   # POD
        3: 16,   # Place
        4: 10,   # Carrier
        5: 30,   # Commodity (widest)
        6: 8,    # Eff
        7: 8,    # Exp
        8: 18,   # Note
        9: 8,    # Source
        10: 9, 11: 9, 12: 9, 13: 9, 14: 9,  # price cols (20GP/40GP/40HQ/45HQ/40NOR or 20RF/40RF)
    }
    for c_idx, w in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = w

    # 2026-04-21 Phase 3: Hide cols O/P/Q (15/16/17) — Contract, Group Rate, Group Code.
    # Nelson must NOT see these in normal view; VBA reads them via .Cells(row, 15/16/17).
    for c_idx in sorted(_HIDDEN_COL_INDICES):
        letter = openpyxl.utils.get_column_letter(c_idx)
        ws.column_dimensions[letter].hidden = True
        ws.column_dimensions[letter].width = 9  # reasonable width when unhidden for audit

    return row_count

if "Pricing Dashboard" in wb.sheetnames:
    # We no longer use this sheet, delete it to avoid confusion or keep it empty
    del wb["Pricing Dashboard"]

dry_rows = write_pricing_sheet(get_or_create_sheet(wb, "Pricing Dry"), pivot_dry, "dry", col_map_dry)
reefer_rows = write_pricing_sheet(get_or_create_sheet(wb, "Pricing Reefer"), pivot_reefer, "reefer", col_map_reefer)
print(f"  Pricing Dry: {dry_rows} rows")
print(f"  Pricing Reefer: {reefer_rows} rows")

# --- Write/Update ChargeBreakdown sheet ---
if "ChargeBreakdown" in wb.sheetnames:
    del wb["ChargeBreakdown"]
ws_cb = wb.create_sheet("ChargeBreakdown")
ws_cb.cell(1, 1, value="Key").font = Font(bold=True)
ws_cb.cell(1, 2, value="Charge_Name").font = Font(bold=True)
ws_cb.cell(1, 3, value="Amount").font = Font(bold=True)

cb_row = 2
for key, charges in charge_lookup.items():
    for charge_name, amount in charges:
        ws_cb.cell(cb_row, 1, value=key)
        ws_cb.cell(cb_row, 2, value=charge_name)
        ws_cb.cell(cb_row, 3, value=amount)
        ws_cb.cell(cb_row, 3).number_format = '#,##0'
        cb_row += 1

ws_cb.sheet_state = 'hidden'
print(f"  ChargeBreakdown: {cb_row - 2} rows")

# --- Write/Update RateVersions sheet ---
# Columns: Type | Short (ribbon display) | Full (source filename for audit) | Updated
if "RateVersions" in wb.sheetnames:
    del wb["RateVersions"]
ws_rv = wb.create_sheet("RateVersions")
ws_rv.cell(1, 1, value="Type").font = Font(bold=True)
ws_rv.cell(1, 2, value="Version").font = Font(bold=True)  # short display
ws_rv.cell(1, 3, value="SourceFile").font = Font(bold=True)  # full filename
ws_rv.cell(1, 4, value="Updated").font = Font(bold=True)

for i, (vtype, vfull, vshort) in enumerate(versions, 2):
    ws_rv.cell(i, 1, value=vtype)
    ws_rv.cell(i, 2, value=vshort)
    ws_rv.cell(i, 3, value=vfull)
    ws_rv.cell(i, 4, value=datetime.now().strftime('%Y-%m-%d %H:%M'))

ws_rv.sheet_state = 'hidden'
print(f"  RateVersions: {len(versions)} entries")

# --- Update PUC_Lookup if PUC file exists ---
# PUC_FILE resolved at top by _resolve_puc_file() — already searches
# processed/ first with rate-tables/ fallback.
puc_path = PUC_FILE

if os.path.exists(puc_path):
    if "PUC_Lookup" in wb.sheetnames:
        ws_puc = wb["PUC_Lookup"]
        # Clear old data (keep header)
        for r in range(2, ws_puc.max_row + 1):
            for c in range(1, 5):
                ws_puc.cell(r, c).value = None
        # Write new — PUC file cols: Destination, PUC_20, PUC_40, [Expiration Date]
        # Only write first 3 cols (Place, 20, 40). Col 4 (Expiry) is NOT PUC_40HC.
        # PUC_40HC = same as PUC_40 for most carriers.
        dfp = pd.read_excel(puc_path)
        valid_entries = 0
        for ri, (_, row) in enumerate(dfp.iterrows(), 2):
            place = row.iloc[0]
            if pd.isna(place) or str(place).startswith("*") or str(place).startswith("1/"):
                continue
            valid_entries += 1
            ws_puc.cell(ri, 1, value=str(place).strip())
            # PUC_20 (col 2)
            v20 = row.iloc[1]
            if pd.notna(v20):
                try:
                    ws_puc.cell(ri, 2, value=int(float(v20)))
                except (ValueError, TypeError):
                    ws_puc.cell(ri, 2, value=0)
            # PUC_40 (col 3) — also used as PUC_40HC
            v40 = row.iloc[2]
            if pd.notna(v40):
                try:
                    v40_int = int(float(v40))
                    ws_puc.cell(ri, 3, value=v40_int)
                    ws_puc.cell(ri, 4, value=v40_int)  # HC = same as 40
                except (ValueError, TypeError):
                    ws_puc.cell(ri, 3, value=0)
                    ws_puc.cell(ri, 4, value=0)
        print(f"  PUC_Lookup: {valid_entries} entries from {os.path.basename(puc_path)}")

# Save
wb.save(ERP_FILE)
print(f"\n  Saved! File: {ERP_FILE}")
print(f"  Size: {os.path.getsize(ERP_FILE) / 1024:.0f} KB")

# Re-inject CustomUI14 XML (openpyxl strips it on save)
print("  Re-injecting CustomUI14 XML...")
customui_xml_path = os.path.join(SCRIPT_DIR, '..', 'vba', 'v14-upgrade', 'CustomUI_v14.xml')
# Also check same folder as ERP file
if not os.path.exists(customui_xml_path):
    customui_xml_path = os.path.join(os.path.dirname(ERP_FILE), 'CustomUI_v14.xml')
if os.path.exists(customui_xml_path):
    from customui_utils import ensure_customui
    cui_result = ensure_customui(ERP_FILE, customui_xml_path=customui_xml_path)
    if cui_result.get("injected"):
        print("  CustomUI14 re-injected OK")
    elif cui_result.get("already_ok"):
        print("  CustomUI14 already present")
    else:
        print(f"  CustomUI14 warning: {cui_result}")
else:
    print(f"  WARN: CustomUI_v14.xml not found — ribbon may be missing")

# Write status file for VBA to read
status_file = os.path.join(os.path.dirname(ERP_FILE), "refresh_status.txt")
with open(status_file, 'w') as f:
    f.write(f"OK|{dry_rows}|{cb_row-2}|{len(versions)}|{datetime.now().strftime('%H:%M')}")
print(f"  Status: {status_file}")
print(f"\n[refresh-v14] Done!")
