"""
ERP Integration - Sync Quotes and Jobs to ERP_Master.xlsx sheets
Workflow: Pricing → Quote sheet → Mark WIN → Active_Jobs sheet
"""

import os
import sys
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')

ERP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ERP_FILE = os.path.join(ERP_DIR, "data", "ERP_Master.xlsm")
QUOTE_HISTORY = os.path.join(ERP_DIR, "data", "Quote_History.xlsx")
JOBS_MASTER = os.path.join(ERP_DIR, "data", "Jobs_Master.xlsx")


def ensure_erp_sheets():
    """Ensure ERP_Master has Quotes and Active_Jobs sheets"""
    if not os.path.exists(ERP_FILE):
        print(f"❌ ERP file not found: {ERP_FILE}")
        return False
    
    wb = load_workbook(ERP_FILE)
    sheets_needed = ['Quotes', 'Active_Jobs']
    
    for sheet_name in sheets_needed:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            if sheet_name == 'Quotes':
                headers = ['QuoteID', 'Date', 'Customer', 'POL', 'POD', 'Place', 
                          'Carrier', 'Container', 'BasePrice', 'Markup', 'FinalPrice',
                          'Effective', 'Status', 'StatusDate', 'Notes']
                for col, h in enumerate(headers, 1):
                    ws.cell(1, col).value = h
            elif sheet_name == 'Active_Jobs':
                headers = ['Job_ID', 'Quote_ID', 'Customer', 'Routing', 'Carrier',
                          'ETD', 'Bkg_No', 'Container', 'Rate', 'Status', 
                          'Created_Date', 'Notes']
                for col, h in enumerate(headers, 1):
                    ws.cell(1, col).value = h
    
    wb.save(ERP_FILE)
    print("✅ ERP sheets ready: Quotes, Active_Jobs")
    return True


def sync_quotes_to_erp():
    """Sync quotes from Quote_History to ERP Quotes sheet"""
    if not os.path.exists(QUOTE_HISTORY):
        print("❌ Quote_History not found")
        return 0
    
    # Load quotes
    df_quotes = pd.read_excel(QUOTE_HISTORY)
    
    # Add Status column if not exists
    if 'Status' not in df_quotes.columns:
        df_quotes['Status'] = 'PENDING'
    if 'StatusDate' not in df_quotes.columns:
        df_quotes['StatusDate'] = None
    if 'Notes' not in df_quotes.columns:
        df_quotes['Notes'] = ''
    
    # Load ERP workbook
    wb = load_workbook(ERP_FILE)
    
    if 'Quotes' not in wb.sheetnames:
        ensure_erp_sheets()
        wb = load_workbook(ERP_FILE)
    
    ws = wb['Quotes']
    
    # Clear existing data (keep headers)
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None
    
    # Write quote data
    headers = ['QuoteID', 'Date', 'Customer', 'POL', 'POD', 'Place', 
               'Carrier', 'Container', 'BasePrice', 'Markup', 'FinalPrice',
               'Effective', 'Status', 'StatusDate', 'Notes']
    
    # Map columns
    col_map = {
        'QuoteID': 'QuoteID', 'Date': 'Date', 'Customer': 'Customer',
        'POL': 'POL', 'POD': 'POD', 'Place': 'Details',
        'Carrier': 'Carrier', 'Container': 'Container', 'BasePrice': 'BasePrice',
        'Markup': 'Markup', 'FinalPrice': 'FinalPrice', 'Effective': 'Effective',
        'Status': 'Status', 'StatusDate': 'StatusDate', 'Notes': 'Notes'
    }
    
    for row_idx, row in df_quotes.iterrows():
        for col_idx, header in enumerate(headers, 1):
            src_col = col_map.get(header, header)
            val = row.get(src_col, '')
            ws.cell(row_idx + 2, col_idx).value = val
    
    wb.save(ERP_FILE)
    print(f"✅ Synced {len(df_quotes)} quote lines to ERP Quotes sheet")
    return len(df_quotes)


def mark_quote_win_in_erp(quote_id, notes=''):
    """Mark quote as WIN in ERP and create Active Job"""
    wb = load_workbook(ERP_FILE)
    
    # Update Quotes sheet
    ws_quotes = wb['Quotes']
    found = False
    quote_data = []
    
    for row in range(2, ws_quotes.max_row + 1):
        if ws_quotes.cell(row, 1).value == quote_id:
            ws_quotes.cell(row, 13).value = 'WIN'  # Status
            ws_quotes.cell(row, 14).value = datetime.now().strftime('%Y-%m-%d %H:%M')  # StatusDate
            ws_quotes.cell(row, 15).value = notes
            
            # Collect data for job creation (first line only)
            if not found:
                quote_data = {
                    'quote_id': quote_id,
                    'customer': ws_quotes.cell(row, 3).value,
                    'pol': ws_quotes.cell(row, 4).value,
                    'pod': ws_quotes.cell(row, 5).value,
                    'carrier': ws_quotes.cell(row, 7).value,
                    'container': ws_quotes.cell(row, 8).value,
                    'rate': ws_quotes.cell(row, 11).value,
                }
            found = True
    
    if not found:
        print(f"❌ Quote {quote_id} not found in ERP")
        wb.close()
        return None
    
    # Create Active Job
    ws_jobs = wb['Active_Jobs']
    
    # Generate Job ID
    today = datetime.now().strftime('%y%m%d')
    existing_jobs = []
    for row in range(2, ws_jobs.max_row + 1):
        jid = ws_jobs.cell(row, 1).value
        if jid and str(jid).startswith(f'JOB-{today}'):
            existing_jobs.append(jid)
    
    seq = len(existing_jobs) + 1
    job_id = f"JOB-{today}-{seq:03d}"
    
    # Find next empty row
    next_row = ws_jobs.max_row + 1
    if ws_jobs.cell(2, 1).value is None:
        next_row = 2
    
    # Write job data
    ws_jobs.cell(next_row, 1).value = job_id
    ws_jobs.cell(next_row, 2).value = quote_id
    ws_jobs.cell(next_row, 3).value = quote_data.get('customer', '')
    ws_jobs.cell(next_row, 4).value = f"{quote_data.get('pol', '')}-{quote_data.get('pod', '')}"
    ws_jobs.cell(next_row, 5).value = quote_data.get('carrier', '')
    ws_jobs.cell(next_row, 9).value = quote_data.get('rate', 0)
    ws_jobs.cell(next_row, 10).value = 'BOOKED'
    ws_jobs.cell(next_row, 11).value = datetime.now().strftime('%Y-%m-%d')
    
    wb.save(ERP_FILE)
    print(f"✅ Quote {quote_id} marked as WIN")
    print(f"✅ Job {job_id} created in Active_Jobs sheet")
    
    return job_id


def mark_quote_lost_in_erp(quote_id, reason=''):
    """Mark quote as LOST in ERP"""
    wb = load_workbook(ERP_FILE)
    ws_quotes = wb['Quotes']
    
    found = False
    for row in range(2, ws_quotes.max_row + 1):
        if ws_quotes.cell(row, 1).value == quote_id:
            ws_quotes.cell(row, 13).value = 'LOST'
            ws_quotes.cell(row, 14).value = datetime.now().strftime('%Y-%m-%d %H:%M')
            ws_quotes.cell(row, 15).value = reason
            found = True
    
    if found:
        wb.save(ERP_FILE)
        print(f"✅ Quote {quote_id} marked as LOST")
    else:
        print(f"❌ Quote {quote_id} not found")
    
    wb.close()
    return found


def get_quote_stats_from_erp():
    """Get quote statistics from ERP"""
    if not os.path.exists(ERP_FILE):
        return {}
    
    df = pd.read_excel(ERP_FILE, sheet_name='Quotes')
    
    if df.empty:
        return {}
    
    grouped = df.groupby('QuoteID')['Status'].first()
    total = len(grouped)
    wins = (grouped == 'WIN').sum()
    lost = (grouped == 'LOST').sum()
    pending = (grouped == 'PENDING').sum()
    
    return {
        'total': total,
        'wins': wins,
        'lost': lost,
        'pending': pending,
        'win_rate': (wins / total * 100) if total > 0 else 0
    }


def erp_quote_manager_menu():
    """ERP-based Quote Manager menu"""
    # Ensure sheets exist
    ensure_erp_sheets()
    
    while True:
        print("\n" + "="*60)
        print("📋 ERP QUOTE MANAGER")
        print("="*60)
        print("\n  [1] Sync Quotes từ CRM → ERP")
        print("  [2] Xem Quotes trong ERP (mở Excel)")
        print("  [3] Đánh dấu WIN → Tạo Job")
        print("  [4] Đánh dấu LOST")
        print("  [5] Xem thống kê")
        print("\n  [0] Quay lại")
        
        choice = input("\nChọn: ").strip()
        
        if choice == '1':
            sync_quotes_to_erp()
        
        elif choice == '2':
            os.startfile(ERP_FILE)
            print("✅ Đã mở ERP_Master.xlsx - xem sheet 'Quotes'")
        
        elif choice == '3':
            quote_id = input("\nNhập QuoteID để đánh dấu WIN: ").strip()
            if quote_id:
                notes = input("Ghi chú (Enter để bỏ qua): ").strip()
                mark_quote_win_in_erp(quote_id, notes)
        
        elif choice == '4':
            quote_id = input("\nNhập QuoteID để đánh dấu LOST: ").strip()
            if quote_id:
                reason = input("Lý do (Enter để bỏ qua): ").strip()
                mark_quote_lost_in_erp(quote_id, reason)
        
        elif choice == '5':
            stats = get_quote_stats_from_erp()
            if stats:
                print(f"\n📊 QUOTE STATISTICS:")
                print(f"   Total: {stats['total']}")
                print(f"   🟡 PENDING: {stats['pending']}")
                print(f"   🟢 WIN: {stats['wins']} ({stats['win_rate']:.1f}%)")
                print(f"   🔴 LOST: {stats['lost']}")
            else:
                print("   Chưa có dữ liệu!")
        
        elif choice == '0':
            break
        
        input("\n[Enter để tiếp tục...]")


if __name__ == "__main__":
    erp_quote_manager_menu()
