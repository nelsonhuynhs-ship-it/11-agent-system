"""
Generate Quote Image V8 - Wharfage Integration
- Header: Split layout with Quote Number
- Table: Separate Current Price and Trend columns  
- Wharfage: SUBJECT TO section below table
- Footer: Single signature (left corner)
"""

import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
from matplotlib.offsetbox import OffsetImage, AnnotationBbox
from matplotlib.patches import Rectangle
from datetime import datetime
from math import ceil
import os
import sys
import re

sys.stdout.reconfigure(encoding='utf-8')

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ERP_FILE = os.path.join(BASE_DIR, "ERP", "data", "ERP_Master.xlsm")
WHARFAGE_FILE = os.path.join(BASE_DIR, "ERP", "data", "Wharface.xlsx")
MASTER_PRICING_FILE = os.path.join(BASE_DIR, "Pricing_Engine", "data", "MasterFullPricing.xlsx") 
QUOTE_LOG_FILE = os.path.join(BASE_DIR, "CRM", "data", "Quote_History.xlsx")
LOGO_PATH = os.path.join(BASE_DIR, "logo_pudong.png")

QUOTE_OUTPUT_DIR = os.path.join(BASE_DIR, "Quotes")
if not os.path.exists(QUOTE_OUTPUT_DIR): 
    os.makedirs(QUOTE_OUTPUT_DIR)

# Colors
C_HEADER_BG = '#003366'
C_TABLE_HEAD = '#001F3F'
C_TREND_UP = '#C00000'    # RED for increase
C_TREND_DOWN = '#00B050'  # GREEN for decrease

# ============================================================================
# WHARFAGE LOADER
# ============================================================================
def load_wharfage_data():
    """Load wharfage fees from Wharface.xlsx"""
    wharfage = {}
    
    if not os.path.exists(WHARFAGE_FILE):
        print("⚠️ Wharface.xlsx not found")
        return wharfage
    
    try:
        df = pd.read_excel(WHARFAGE_FILE, sheet_name='Sheet1', header=None)
        
        # CMA: Rows 2-9, Cols 0-8
        wharfage['CMA'] = {}
        for r in range(2, 10):
            pod = str(df.iloc[r, 0]).strip().upper() if pd.notna(df.iloc[r, 0]) else ""
            if not pod: continue
            wharfage['CMA'][pod] = {
                '20GP': df.iloc[r, 2] if pd.notna(df.iloc[r, 2]) else 0,
                '40GP': df.iloc[r, 3] if pd.notna(df.iloc[r, 3]) else 0,
                '40HC': df.iloc[r, 4] if pd.notna(df.iloc[r, 4]) else 0,
                '45HQ': df.iloc[r, 5] if pd.notna(df.iloc[r, 5]) else 0,
            }
        
        # ONE Vancouver special case: Per TEU pricing
        # 20GP = 1 TEU, 40GP/40HC/45HQ = 2 TEU
        wharfage['ONE'] = {}  # Initialize ONE dict
        for r in range(14, 19):
            pod = str(df.iloc[r, 0]).strip().upper() if pd.notna(df.iloc[r, 0]) else ""
            if not pod: continue
            unit = str(df.iloc[r, 1]).strip().lower() if pd.notna(df.iloc[r, 1]) else ""
            base_20 = df.iloc[r, 2] if pd.notna(df.iloc[r, 2]) else 0
            base_40 = df.iloc[r, 3] if pd.notna(df.iloc[r, 3]) else 0
            
            # If unit is "per teu", 40ft containers cost double
            if 'teu' in unit:
                wharfage['ONE'][pod] = {
                    '20GP': base_20,
                    '40GP': base_40,  # Already doubled in source or needs doubling
                    '40HC': base_40,
                    '45HQ': base_40,
                }
            else:
                wharfage['ONE'][pod] = {
                    '20GP': base_20,
                    '40GP': base_40,
                    '40HC': df.iloc[r, 4] if pd.notna(df.iloc[r, 4]) else base_40,
                    '45HQ': df.iloc[r, 5] if pd.notna(df.iloc[r, 5]) else base_40,
                }
        
        # MSC: Rows 2-8, Cols 11-13 (flat rate)
        wharfage['MSC'] = {}
        for r in range(2, 9):
            pod = str(df.iloc[r, 11]).strip().upper() if pd.notna(df.iloc[r, 11]) else ""
            if not pod: continue
            rate = df.iloc[r, 13] if pd.notna(df.iloc[r, 13]) else 0
            wharfage['MSC'][pod] = {'20GP': rate, '40GP': rate, '40HC': rate, '45HQ': rate}
        
        # EMC: Rows 13-18, Cols 11-13
        wharfage['EMC'] = {}
        for r in range(13, 19):
            pod = str(df.iloc[r, 11]).strip().upper() if pd.notna(df.iloc[r, 11]) else ""
            if not pod: continue
            rate = df.iloc[r, 13] if pd.notna(df.iloc[r, 13]) else 0
            wharfage['EMC'][pod] = {'20GP': rate, '40GP': rate, '40HC': rate, '45HQ': rate}
        
        # ZIM: Rows 2-7, Cols 16-19 (WAF + Port Tariff, round up to $10)
        wharfage['ZIM'] = {}
        for r in range(2, 8):
            pod = str(df.iloc[r, 16]).strip().upper() if pd.notna(df.iloc[r, 16]) else ""
            if not pod: continue
            
            # Parse WAF (e.g., "5$/CTNR")
            waf_str = str(df.iloc[r, 17]) if pd.notna(df.iloc[r, 17]) else "0"
            waf = float(re.findall(r'[\d.]+', waf_str)[0]) if re.findall(r'[\d.]+', waf_str) else 0
            
            # Parse Port Tariff (e.g., "63.06$/CTNR")
            port_str = str(df.iloc[r, 18]) if pd.notna(df.iloc[r, 18]) else "0"
            port = float(re.findall(r'[\d.]+', port_str)[0]) if re.findall(r'[\d.]+', port_str) else 0
            
            # Round up to nearest $10
            total = ceil((waf + port) / 10) * 10
            wharfage['ZIM'][pod] = {'20GP': total, '40GP': total, '40HC': total, '45HQ': total}
        
        print(f"✅ Loaded wharfage for: {list(wharfage.keys())}")
        
    except Exception as e:
        print(f"⚠️ Error loading wharfage: {e}")
    
    return wharfage

def get_wharfage_notes(data_rows, selected_cols, wharfage):
    """Get wharfage notes for selected carriers/PODs"""
    notes = {}  # {carrier: fee}
    
    for row in data_rows:
        carrier = str(row.get('Carrier', '')).strip().upper()
        # Try to match POD from Place field (destination)
        place = str(row.get('Place', '')).strip().upper()
        
        if carrier not in wharfage:
            continue
        
        # Try to find POD match
        carrier_pods = wharfage.get(carrier, {})
        matched_pod = None
        
        for pod_key in carrier_pods.keys():
            if pod_key in place or place in pod_key:
                matched_pod = pod_key
                break
        
        if not matched_pod:
            continue
        
        # Get fee for first selected container
        cont_key = selected_cols[0] if selected_cols else '40HC'
        cont_key = cont_key.replace("'", "")  # Clean 45'HQ -> 45HQ
        
        fee = carrier_pods.get(matched_pod, {}).get(cont_key, 0)
        if fee and carrier not in notes:
            notes[carrier] = int(fee)
    
    return notes

# ============================================================================
# QUOTE FUNCTIONS
# ============================================================================
def get_next_quote_id():
    now = datetime.now()
    month_str = now.strftime("%y%b").upper()
    try:
        if not os.path.exists(QUOTE_LOG_FILE): return f"{month_str}-001"
        df = pd.read_excel(QUOTE_LOG_FILE)
        if df.empty: return f"{month_str}-001"
        curr = df[df['QuoteID'].astype(str).str.startswith(month_str)]
        if curr.empty: return f"{month_str}-001"
        max_s = 0
        for q in curr['QuoteID']:
            try: 
                s = int(q.split('-')[1])
                if s > max_s: max_s = s
            except: pass
        return f"{month_str}-{max_s+1:03d}"
    except: return f"{month_str}-001"

def log_quote(qid, cust, rows, cols, m_map):
    recs = []
    now = datetime.now()
    for r in rows:
        for c in cols:
            m = m_map.get(c, 0)
            bp = r.get(c, 0)
            fp = (bp if bp else 0) + m
            recs.append({
                'QuoteID': qid, 'Date': now, 'Customer': cust,
                'POL': r['POL'], 'POD': r['POD'], 'Carrier': r['Carrier'],
                'Container': c, 'Base': bp, 'Markup': m, 'Final': fp, 'Eff': r['Eff']
            })
    df_n = pd.DataFrame(recs)
    if os.path.exists(QUOTE_LOG_FILE):
        try: 
            pd.concat([pd.read_excel(QUOTE_LOG_FILE), df_n]).to_excel(QUOTE_LOG_FILE, index=False)
        except: 
            df_n.to_excel(QUOTE_LOG_FILE, index=False)
    else: 
        df_n.to_excel(QUOTE_LOG_FILE, index=False)

def get_trend(df_h, pol, pod, carr, cont, curr_p, curr_eff):
    """Get trend vs previous validity - RED for up, GREEN for down"""
    if df_h.empty or not curr_p: 
        return "black", "-"
    
    # Map container types
    cont_map = {'20GP': '20GP', '40GP': '40GP', '40HQ': '40HQ', '45HQ': '45HQ', 
                '40NOR': '40NOR', '20RF': '20RF', '40RF': '40RF'}
    price_col = cont_map.get(cont, '40HQ')
    
    # Check if required columns exist
    if 'POL' not in df_h.columns or price_col not in df_h.columns:
        return "black", "-"
    
    # Filter by route and carrier
    mask = (df_h['POL'] == pol) & (df_h['POD'] == pod) & (df_h['Carrier'] == carr)
    df_f = df_h[mask].copy()
    
    if df_f.empty:
        return "black", "-"
    
    # Sort by effective date descending
    if 'EffectiveDate' in df_f.columns:
        df_f = df_f.sort_values('EffectiveDate', ascending=False)
    
    # Find previous price (different from current)
    prev = 0
    for _, r in df_f.iterrows():
        hp = r.get(price_col, 0)
        if pd.notna(hp) and hp > 0:
            hp = float(hp)
            # If price is meaningfully different (at least $50 difference or different value)
            if abs(hp - curr_p) > 50:
                prev = hp
                break
    
    if prev == 0: 
        return "black", "-"
    
    delta = curr_p - prev
    if delta < 0: 
        return C_TREND_DOWN, f"▼ ${abs(int(delta)):,}"  # GREEN - decrease
    elif delta > 0: 
        return C_TREND_UP, f"▲ ${int(delta):,}"         # RED - increase
    return "black", "-"

# ============================================================================
# MAIN GENERATION
# ============================================================================
def generate_image():
    print("🎨 GENERATING QUOTE V8 (WHARFAGE INTEGRATION)...")
    
    # Load Wharfage
    wharfage = load_wharfage_data()
    
    # READ EXCEL
    wb = openpyxl.load_workbook(ERP_FILE, data_only=True)
    ws = wb['📊 Pricing Dashboard']
    customer = ws['B4'].value
    
    markup_map = {
        '20GP': ws['K7'].value or 0, '40GP': ws['L7'].value or 0,
        '40HQ': ws['M7'].value or 0, "45'HQ": ws['N7'].value or 0,
        '45HQ': ws['N7'].value or 0, '40NOR': ws['O7'].value or 0,
        '20RF': ws['P7'].value or 0, '40RF': ws['Q7'].value or 0
    }
    
    # Container Selection
    cont_opts = ['20GP', '40GP', '40HQ', '45HQ', '40NOR', '20RF', '40RF']
    selected_cols = []
    for i, c in enumerate(cont_opts):
        v = ws.cell(6, 11+i).value
        if v and str(v).lower() in ['x', 'true', '1']: 
            selected_cols.append(c)
    if not selected_cols: 
        selected_cols = ['20GP', '40GP', '40HQ']

    # Row Selection
    data = []
    explicit = False
    for r in range(11, ws.max_row+1):
        v = ws.cell(r, 1).value
        if v is True or str(v).upper() in ["TRUE", "X", "1"]: 
            explicit = True
            break
    
    limit_vis = 0
    for r in range(11, ws.max_row+1):
        inc = False
        v = ws.cell(r, 1).value
        chk = (v is True or str(v).upper() in ["TRUE", "X", "1"])
        if explicit: 
            if chk: inc = True
        else:
            if not ws.row_dimensions[r].hidden and ws.cell(r, 2).value: 
                inc = True
                limit_vis += 1
        
        if inc:
            data.append({
                'POL': ws.cell(r, 2).value, 'POD': ws.cell(r, 3).value,
                'Place': ws.cell(r, 4).value, 'Carrier': ws.cell(r, 5).value,
                'Eff': ws.cell(r, 7).value, 'Exp': ws.cell(r, 8).value,
                'Note': ws.cell(r, 9).value,
                '20GP': ws.cell(r, 11).value, '40GP': ws.cell(r, 12).value,
                '40HQ': ws.cell(r, 13).value, '45HQ': ws.cell(r, 14).value,
                '40NOR': ws.cell(r, 15).value, '20RF': ws.cell(r, 16).value,
                '40RF': ws.cell(r, 17).value,
            })
            if not explicit and limit_vis >= 20: break
            
    if not data: 
        print("❌ No rows")
        return

    # Get Wharfage Notes
    wharfage_notes = get_wharfage_notes(data, selected_cols, wharfage)
    
    # Load History from Recent_History sheet
    try: 
        df_h = pd.read_excel(MASTER_PRICING_FILE, sheet_name='Recent_History')
        # Rename columns for consistency
        df_h = df_h.rename(columns={
            # POL and POD already use standardized names
            'Base Ocean Freight_20GP': '20GP',
            'Base Ocean Freight_40GP': '40GP', 
            'Base Ocean Freight_40HQ': '40HQ',
            "Base Ocean Freight_45'HQ": '45HQ',
            'Base Ocean Freight_40NOR': '40NOR',
            'Base Ocean Freight_20RF': '20RF',
            'Base Ocean Freight_40RF': '40RF',
        })
        print(f"✅ Loaded {len(df_h)} history records for trend comparison")
    except Exception as e:
        print(f"⚠️ Could not load history: {e}")
        df_h = pd.DataFrame()

    # Build Table Data
    final_rows = []
    for row in data:
        r_out = [
            row['POL'] or "",
            row['POD'] or "",
            row['Place'] or "",
            row['Carrier'] or "",
            row['Eff'] or "",
            row['Exp'] or "",
            row['Note'] if row['Note'] and 'SOC' in str(row['Note']).upper() else ""
        ]
        
        cont_map = {'20GP':'20GP','40GP':'40GP','40HQ':'40HQ',"45'HQ":'45HQ',
                    '45HQ':'45HQ','40NOR':'40NOR','20RF':'20RF','40RF':'40RF'}
        
        for h in selected_cols:
            k = cont_map.get(h, h)
            p = row.get(k, 0)
            m = markup_map.get(h, 0)
            fp = (p if p else 0) + m
            
            if not p:
                r_out.append("-")
                r_out.append("-")
            else:
                r_out.append(f"${int(fp):,}")
                c, t = get_trend(df_h, row['POL'], row['POD'], row['Carrier'], k, p, row['Eff'])
                r_out.append((t, c))
            
        final_rows.append(r_out)

    # Generate Quote ID and Log
    qid = get_next_quote_id()
    log_quote(qid, customer, data, selected_cols, markup_map)
    
    # PLOT
    num_cols = 7 + (2 * len(selected_cols))
    fig_w = max(16, num_cols * 1.4)
    fig_h = max(8, 4 + len(final_rows) * 0.8)
    
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis('off')
    
    # Header
    rect_header = Rectangle((0, 0.85), 1, 0.15, transform=ax.transAxes, color=C_HEADER_BG, clip_on=False)
    ax.add_patch(rect_header)
    
    # Logo
    if os.path.exists(LOGO_PATH):
        img = mpimg.imread(LOGO_PATH)
        ib = OffsetImage(img, zoom=0.04)
        ab = AnnotationBbox(ib, (0.08, 0.925), frameon=False, box_alignment=(0.5, 0.5))
        ax.add_artist(ab)
        
    # Title & Info
    ax.text(0.92, 0.94, "OCEAN FREIGHT QUOTATION", fontsize=22, weight='bold', color='white', ha='right')
    cust_str = f"CUSTOMER: {str(customer).upper()}" if customer else "CUSTOMER: AD-HOC"
    ax.text(0.92, 0.89, cust_str, fontsize=13, color='white', ha='right')
    ax.text(0.92, 0.86, f"QUOTE NO: {qid}", fontsize=12, color='#FFD700', weight='bold', ha='right')
    
    # Market Update - Clean text without color names
    ax.text(0.02, 0.78, "MARKET UPDATE", fontsize=13, weight='bold', color='black')
    ax.text(0.02, 0.75, "Rates compared to previous validity.", fontsize=10, color='#444')
    # Add colored legend inline
    ax.text(0.30, 0.75, "▲", fontsize=12, weight='bold', color=C_TREND_UP)
    ax.text(0.32, 0.75, "= Increase  ", fontsize=10, color='#444')
    ax.text(0.42, 0.75, "▼", fontsize=12, weight='bold', color=C_TREND_DOWN)
    ax.text(0.44, 0.75, "= Decrease", fontsize=10, color='#444')

    # Table
    base_headers = ['POL', 'POD', 'Destination', 'Carrier', 'Eff', 'Exp', 'Note']
    cont_headers = []
    for c in selected_cols:
        cont_headers.append(f"{c}\n(Current)")
        cont_headers.append("Trend")
    col_labels = base_headers + cont_headers
    
    table = plt.table(
        cellText=[[c[0] if isinstance(c, tuple) else c for c in r] for r in final_rows],
        colLabels=col_labels,
        cellLoc='center', loc='upper center',
        bbox=[0.02, 0.68 - (len(final_rows)+1)*0.075, 0.96, (len(final_rows)+1)*0.075]
    )
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    
    for (row, col), cell in table.get_celld().items():
        cell.set_linewidth(0.5)
        cell.set_edgecolor('black')
        
        if row == 0:
            cell.set_facecolor(C_TABLE_HEAD)
            cell.set_text_props(weight='bold', color='white')
            cell.set_height(0.08)
        else:
            cell.set_height(0.075)
            cell.set_facecolor('white')
            
            if col >= 7:
                col_offset = col - 7
                is_trend_col = (col_offset % 2 == 1)
                
                if is_trend_col:
                    val = final_rows[row-1][col]
                    if isinstance(val, tuple):
                        text, color = val
                        cell.get_text().set_text(text)
                        if text != "-":
                            cell.set_text_props(weight='bold', color=color)

    # SUBJECT TO (Wharfage) - Concise format
    fy = 0.68 - (len(final_rows)+1)*0.075 - 0.06
    
    if wharfage_notes:
        subject_parts = [f"{carr} ${fee}" for carr, fee in wharfage_notes.items()]
        subject_text = "SUBJECT TO: " + " | ".join(subject_parts) + " (Wharfage per cont)"
        ax.text(0.02, fy, subject_text, fontsize=10, style='italic', color='#333')
        fy -= 0.04

    # Footer
    fy -= 0.04
    ax.text(0.02, fy, "Nelson Huynh", fontsize=12, weight='bold')
    ax.text(0.02, fy-0.04, "Sales Team Leader", fontsize=10)
    ax.text(0.02, fy-0.08, "Pudong Prime International Co., Ltd.", fontsize=10)
    ax.text(0.02, fy-0.12, "+84 931 301 014", fontsize=10)
    ax.text(0.02, fy-0.16, "nelson@pudongprime.vn", fontsize=10)

    # Save
    out = os.path.join(QUOTE_OUTPUT_DIR, f"Quote_{qid}.png")
    plt.savefig(out, dpi=200, bbox_inches='tight')
    plt.close()
    
    print(f"✅ Quote saved: {out}")
    os.startfile(out)

if __name__ == "__main__":
    try: 
        generate_image()
    except Exception as e: 
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        input("Press Enter...")
