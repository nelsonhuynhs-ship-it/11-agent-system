"""
monthly_report.py — Auto-generate Nelson's monthly performance report
=====================================================================
Exports Active Jobs data to 24-col format matching Nelson's
"SALES PROFIT - {MON YYYY} - NELSON CHINH.xlsx" template.

Schema (row 1 top headers, row 2 sub-headers for merged groups):
  A  (1)  No
  B  (2)  SHIPPER/CONSIGNEE
  C  (3)  POL/POD
  D  (4)  FINAL DEST
  E  (5)  ETD
  F  (6)  ETA
  G  (7)  CARRIER/COLOADER
  H  (8)  HBL
  I  (9)  JOB NO
  J-Q (10-17)  Volume: AIR | LCL | 20RF | 20' | 40' | HC | 40RF | 45
  R  (18) Buying
  S  (19) Selling
  T  (20) Profit Share
  U-W (21-23) KICK BACK: Client | Carrier | Tax
  X  (24) Net Profit

Data source: ERP_Master_v14.xlsm "Active Jobs" sheet (30+ cols).
Filter: by ETD month (default = current month).

Usage:
    python ERP/intelligence/monthly_report.py
    python ERP/intelligence/monthly_report.py --month 2026-04
    python ERP/intelligence/monthly_report.py --month APR-2026 --out monthly.xlsx
"""
from __future__ import annotations

import argparse
import calendar
import os
import re
import sys
from datetime import datetime
from typing import Final

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "core"))
from active_jobs_cols import COL as AJ_COL, HDR_ROW as AJ_HDR_ROW, DATA_START as AJ_DATA_START  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]

DEFAULT_ERP_FILE: Final = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"
DEFAULT_OUT_DIR: Final = r"D:\OneDrive\NelsonData\erp\monthly_reports"

# Container type → Volume column letter (sub-header row 2)
CONT_COL_MAP: Final = {
    "20GP":  "M",  # 20'
    "20DC":  "M",
    "40GP":  "N",  # 40'
    "40DC":  "N",
    "40HC":  "O",  # HC
    "40HQ":  "O",
    "40":    "N",
    "20RF":  "L",
    "40RF":  "P",
    "45HC":  "Q",
    "45HQ":  "Q",
    # AIR / LCL handled by special Status detection
}
COL_AIR = "J"
COL_LCL = "K"




# ── Helpers ──
def parse_month(s: str | None) -> tuple[int, int, str]:
    """Return (year, month, label_MON_YYYY). Accepts 2026-04, APR-2026, Apr2026, or None."""
    if not s:
        now = datetime.now()
        return now.year, now.month, f"{calendar.month_abbr[now.month].upper()} {now.year}"
    m = re.match(r"^\s*(\d{4})-(\d{1,2})\s*$", s)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        return y, mo, f"{calendar.month_abbr[mo].upper()} {y}"
    m = re.match(r"^\s*([A-Za-z]{3,})\W?(\d{4})\s*$", s)
    if m:
        name, y = m.group(1).lower(), int(m.group(2))
        for i in range(1, 13):
            if calendar.month_abbr[i].lower() == name[:3]:
                return y, i, f"{calendar.month_abbr[i].upper()} {y}"
    raise ValueError(f"Can't parse month: {s}. Use 2026-04 or APR-2026")


def parse_pol_from_routing(routing: str) -> str:
    """Extract POL from Routing like 'HPH-USLGB' or 'HPH-Chicago VIA USLGB' → 'VNHPH' format."""
    if not routing:
        return ""
    pol = routing.split("-")[0].strip()
    # Nelson uses VNHPH/VNCMT format in report but HPH/CMT in ERP routing
    MAP = {"HPH": "VNHPH", "HCM": "VNCMT", "CAT": "VNCMT", "CMT": "VNCMT"}
    return MAP.get(pol.upper(), pol)


def extract_volume(container_type: str, quantity: int, status: str = "") -> tuple[str, int]:
    """Return (column_letter, quantity). AIR/LCL detected from status/notes."""
    status_u = (status or "").upper()
    if "AIR" in status_u:
        return COL_AIR, quantity
    if "LCL" in status_u:
        return COL_LCL, quantity
    col = CONT_COL_MAP.get((container_type or "").upper().strip(), "N")
    return col, quantity


def load_active_jobs(erp_file: str) -> list[dict]:
    """Load Active Jobs rows as list of dicts."""
    if not os.path.exists(erp_file):
        raise FileNotFoundError(erp_file)
    wb = openpyxl.load_workbook(erp_file, read_only=True, data_only=True, keep_vba=True)
    sheet = next((s for s in wb.sheetnames if "Active" in s), None)
    if not sheet:
        wb.close()
        raise RuntimeError("Active Jobs sheet not found")
    ws = wb[sheet]
    rows: list[dict] = []
    for r in range(AJ_DATA_START, ws.max_row + 1):
        crm = ws.cell(r, AJ_COL["CRM_ID"]).value
        if not crm:
            continue
        row = {k: ws.cell(r, c).value for k, c in AJ_COL.items()}
        rows.append(row)
    wb.close()
    return rows


def filter_by_month(rows: list[dict], year: int, month: int) -> list[dict]:
    """Filter rows where ETD falls in year-month. If ETD missing, use Created_Date."""
    out = []
    for r in rows:
        dt = r.get("ETD") or r.get("Created_Date")
        if not isinstance(dt, datetime):
            continue
        if dt.year == year and dt.month == month:
            out.append(r)
    return out


# ── Excel writer ──
THIN = Side(style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _title_style(cell):
    cell.font = Font(bold=True, size=14, color="1F4E79", name="Segoe UI")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _hdr_style(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=10, name="Segoe UI")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def _sub_style(cell):
    cell.font = Font(bold=True, color="1F4E79", size=9, name="Segoe UI")
    cell.fill = PatternFill("solid", fgColor="D9E1F2")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def _body_style(cell, fmt: str | None = None, color_profit=False):
    cell.font = Font(size=10, name="Segoe UI")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if color_profit and isinstance(cell.value, (int, float)):
        if cell.value > 0:
            cell.font = Font(size=10, name="Segoe UI", color="00804A", bold=True)
        elif cell.value < 0:
            cell.font = Font(size=10, name="Segoe UI", color="C00000", bold=True)


def write_report(rows: list[dict], month_label: str, out_file: str) -> dict:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"MONTHLY {month_label}"

    # Title (merged A1:X1)
    ws.merge_cells("A1:X1")
    ws["A1"] = f"MONTHLY PERFORMANCE — {month_label}"
    _title_style(ws["A1"])
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:X2")
    ws["A2"] = "Pudong Prime — Nelson Huynh"
    ws["A2"].font = Font(italic=True, size=9, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # ── Header row 3 (main) + row 4 (sub for merged groups) ──
    MAIN = [
        (1, 1, "No"),
        (2, 2, "SHIPPER/\nCONSIGNEE"),
        (3, 3, "POL/POD"),
        (4, 4, "FINAL DEST"),
        (5, 5, "ETD"),
        (6, 6, "ETA"),
        (7, 7, "CARRIER/\nCOLOADER"),
        (8, 8, "HBL"),
        (9, 9, "JOB NO"),
        (10, 17, "Volume"),
        (18, 18, "Buying"),
        (19, 19, "Selling"),
        (20, 20, "Profit\nShare"),
        (21, 23, "KICK BACK"),
        (24, 24, "Net\nProfit"),
    ]
    for c_start, c_end, label in MAIN:
        if c_start == c_end:
            ws.cell(3, c_start, label)
            ws.merge_cells(start_row=3, start_column=c_start, end_row=4, end_column=c_end)
        else:
            ws.cell(3, c_start, label)
            ws.merge_cells(start_row=3, start_column=c_start, end_row=3, end_column=c_end)
        _hdr_style(ws.cell(3, c_start))

    SUB = {
        10: "AIR", 11: "LCL", 12: "20RF", 13: "20'",
        14: "40'", 15: "HC", 16: "40RF", 17: "45",
        21: "Client", 22: "Carrier", 23: "Tax",
    }
    for c, label in SUB.items():
        ws.cell(4, c, label)
        _sub_style(ws.cell(4, c))

    # style already-rendered single-span headers at row 4 via border
    for c in range(1, 25):
        if c not in SUB:
            cell = ws.cell(4, c)
            if cell.value is None:
                # merged spanning down from row 3: border was applied via merge; ensure style
                pass

    ws.row_dimensions[3].height = 32
    ws.row_dimensions[4].height = 18

    # Column widths
    widths = {
        "A": 5, "B": 22, "C": 10, "D": 22, "E": 11, "F": 11, "G": 18,
        "H": 18, "I": 14,
        "J": 6, "K": 6, "L": 6, "M": 6, "N": 6, "O": 6, "P": 6, "Q": 6,
        "R": 11, "S": 11, "T": 10,
        "U": 8, "V": 8, "W": 8, "X": 11,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Data rows ──
    data_start = 5
    total_buy = total_sell = total_net = 0.0

    for idx, r in enumerate(rows, start=1):
        rr = data_start + idx - 1

        # Routing: "HPH-USLGB" → POL=HPH POD=USLGB. For POL/POD col just show POL per template.
        routing = r.get("Routing") or ""
        pol = parse_pol_from_routing(routing)
        door = r.get("Door_Address") or ""
        if not door and "-" in routing:
            door = routing.split("-", 1)[1].strip()  # fallback = POD

        cont = r.get("Container_Type") or ""
        qty = int(r.get("Quantity") or 1)
        status = r.get("Status") or ""

        buy = float(r.get("Buying_Rate") or 0) * qty
        sell = float(r.get("Selling_Rate") or 0) * qty
        net = sell - buy

        # Volume distribution — one non-zero cell among J-Q
        vol_col_letter, vol_qty = extract_volume(cont, qty, status)

        # Write row
        ws.cell(rr, 1, idx)
        ws.cell(rr, 2, str(r.get("CRM_ID") or "").strip())
        ws.cell(rr, 3, pol)
        ws.cell(rr, 4, door)
        ws.cell(rr, 5, r.get("ETD"))
        ws.cell(rr, 6, r.get("ETA"))
        ws.cell(rr, 7, r.get("Carrier"))
        ws.cell(rr, 8, r.get("HBL_NO"))
        ws.cell(rr, 9, r.get("FAST_ID"))
        ws.cell(rr, ord(vol_col_letter) - ord("A") + 1, vol_qty)
        ws.cell(rr, 18, buy)
        ws.cell(rr, 19, sell)
        # col 20 Profit Share, cols 21/22/23 KB, col 24 Net — blank for manual kickback entry
        ws.cell(rr, 24, net)

        for c in range(1, 25):
            cell = ws.cell(rr, c)
            if c in (5, 6):
                _body_style(cell, fmt="dd/mm/yy")
            elif c in (18, 19):
                _body_style(cell, fmt='"$"#,##0')
            elif c == 24:
                _body_style(cell, fmt='"$"#,##0', color_profit=True)
            elif c == 20 or c in (21, 22, 23):
                _body_style(cell, fmt='"$"#,##0')
            else:
                _body_style(cell)

        total_buy += buy
        total_sell += sell
        total_net += net

    # ── TOTAL row ──
    if rows:
        tr = data_start + len(rows)
        ws.cell(tr, 1, "TOTAL")
        ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=17)
        ws.cell(tr, 18, total_buy)
        ws.cell(tr, 19, total_sell)
        ws.cell(tr, 24, total_net)
        for c in range(1, 25):
            cell = ws.cell(tr, c)
            cell.font = Font(bold=True, size=11, name="Segoe UI", color="1F4E79")
            cell.fill = PatternFill("solid", fgColor="FFF2CC")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
            if c in (18, 19, 24):
                cell.number_format = '"$"#,##0'

    # Freeze panes below headers
    ws.freeze_panes = "A5"

    wb.save(out_file)
    wb.close()
    return {
        "rows": len(rows),
        "total_buy": total_buy,
        "total_sell": total_sell,
        "total_net": total_net,
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--erp", default=DEFAULT_ERP_FILE)
    ap.add_argument("--month", default=None, help="2026-04 or APR-2026; default=current")
    ap.add_argument("--out", default=None, help="output .xlsx path")
    args = ap.parse_args()

    year, month, label = parse_month(args.month)
    print(f"[+] Building monthly report for {label}")

    rows = load_active_jobs(args.erp)
    print(f"    -> loaded {len(rows)} total jobs")
    filtered = filter_by_month(rows, year, month)
    print(f"    -> {len(filtered)} jobs in {label}")

    out = args.out
    if not out:
        os.makedirs(DEFAULT_OUT_DIR, exist_ok=True)
        out = os.path.join(
            DEFAULT_OUT_DIR,
            f"SALES_PROFIT_{label.replace(' ', '_')}.xlsx",
        )

    stats = write_report(filtered, label, out)
    print(f"\n[SUCCESS] {stats['rows']} jobs exported")
    print(f"    buy:  ${stats['total_buy']:>12,.0f}")
    print(f"    sell: ${stats['total_sell']:>12,.0f}")
    print(f"    net:  ${stats['total_net']:>12,.0f}")
    print(f"    -> {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
