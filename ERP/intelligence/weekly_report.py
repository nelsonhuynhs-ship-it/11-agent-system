"""weekly_report.py — Nelson Freight weekly sales KPI report (12 cols).
Replaces previous 4C market-analysis script — see f6-weekly-report.md.
Usage: python ERP/intelligence/weekly_report.py [--year Y --week W] [--out path]
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import date, datetime, timedelta
from typing import Final

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "core"))
from active_jobs_cols import COL as AJ_COL, DATA_START as AJ_DATA_START  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]

DEFAULT_ERP_FILE: Final = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"
DEFAULT_OUT_DIR: Final = r"D:\OneDrive\NelsonData\erp\weekly_reports"
DEFAULT_EMAIL_LOG: Final = r"D:\NELSON\2. Areas\Engine_test\email_engine\logs\email_log.csv"

TEU_FACTOR: Final = {
    "20GP": 1, "20DC": 1, "20RF": 1,
    "40GP": 2, "40DC": 2, "40HC": 2, "40HQ": 2, "40RF": 2, "45HC": 2, "45HQ": 2,
}

SALES_MAP: Final = {
    "Nelson":  ["nelsonhuynhs@gmail.com", "nelson@pudongprime.vn"],
    "Johnny":  ["johnny@pudongprime.vn"],
    "Jennie":  ["jennie@pudongprime.vn"],
    "Blue":    ["blue@pudongprime.vn"],
    "Lina":    ["lina@pudongprime.vn"],
    "Otis":    ["otis@pudongprime.vn"],
    "Jun":     ["jun@pudongprime.vn"],
}

EMAIL_TO_SALES: Final = {
    email: name
    for name, emails in SALES_MAP.items()
    for email in emails
}


def iso_week(d: date) -> tuple[int, int]:
    """Return (iso_year, iso_week_number) for a given date per ISO-8601."""
    iso = d.isocalendar()
    return iso[0], iso[1]


def week_bounds(year: int, week: int) -> tuple[date, date]:
    """Return (monday, sunday) for the given ISO year/week."""
    jan4 = date(year, 1, 4)  # Jan 4 is always in week 1 per ISO-8601
    week1_monday = jan4 - timedelta(days=jan4.weekday())
    monday = week1_monday + timedelta(weeks=week - 1)
    sunday = monday + timedelta(days=6)
    return monday, sunday


def load_active_jobs_for_week(erp_file: str, year: int, week: int) -> list[dict]:
    """Load Active Jobs rows where Created_Date falls in the ISO week.
    All attributed to Nelson (safe default — no Owner col in v14).
    """
    if not os.path.exists(erp_file):
        raise FileNotFoundError(f"ERP file not found: {erp_file}")

    monday, sunday = week_bounds(year, week)
    start = datetime(monday.year, monday.month, monday.day)
    end = datetime(sunday.year, sunday.month, sunday.day, 23, 59, 59)

    wb = openpyxl.load_workbook(erp_file, read_only=True, data_only=True, keep_vba=True)
    sheet_name = next((s for s in wb.sheetnames if "Active" in s), None)
    if not sheet_name:
        wb.close()
        raise RuntimeError("Active Jobs sheet not found in workbook")

    ws = wb[sheet_name]
    rows: list[dict] = []
    for r in range(AJ_DATA_START, ws.max_row + 1):
        crm = ws.cell(r, AJ_COL["CRM_ID"]).value
        if not crm:
            continue
        created = ws.cell(r, AJ_COL["Created_Date"]).value
        if not isinstance(created, datetime):
            continue
        if not (start <= created <= end):
            continue
        rows.append({
            "CRM_ID": str(crm).strip(),
            "Container_Type": ws.cell(r, AJ_COL["Container_Type"]).value or "",
            "Quantity": int(ws.cell(r, AJ_COL["Quantity"]).value or 1),
            "Profit": float(ws.cell(r, AJ_COL["Profit"]).value or 0),
            "sales": "Nelson",
        })
    wb.close()
    return rows


def load_emails_for_week(log_csv: str, year: int, week: int) -> dict[str, int]:
    """Count emails per sender for the ISO week. Returns {sender_email: count}.
    Auto-detects sender_email/sender/from_email column if present;
    otherwise credits all to Nelson (current log has no sender column).
    """
    if not os.path.exists(log_csv):
        return {}

    import pandas as pd

    try:
        df = pd.read_csv(log_csv, dtype=str)
    except Exception:
        return {}

    sender_col = None
    for candidate in ("sender_email", "sender", "from_email"):
        if candidate in df.columns:
            sender_col = candidate
            break

    monday, sunday = week_bounds(year, week)

    def _parse_ts(val: str) -> datetime | None:
        for fmt in ("%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return datetime.strptime(str(val).strip(), fmt)
            except ValueError:
                continue
        return None

    df["_ts"] = df["timestamp"].apply(_parse_ts)
    df = df.dropna(subset=["_ts"])
    df = df[df["_ts"].apply(
        lambda dt: date(dt.year, dt.month, dt.day) >= monday
        and date(dt.year, dt.month, dt.day) <= sunday
    )]

    if sender_col:
        counts = df[sender_col].value_counts().to_dict()
        return {str(k): int(v) for k, v in counts.items()}
    else:
        return {"nelson@pudongprime.vn": int(len(df))}


def _load_crm_first_transactions(erp_file: str) -> dict[str, datetime]:
    """Return {CRM_ID: first_transaction_date} from CRM sheet."""
    if not os.path.exists(erp_file):
        return {}
    try:
        wb = openpyxl.load_workbook(erp_file, read_only=True, data_only=True, keep_vba=True)
        sname = next((s for s in wb.sheetnames if "CRM" in s.upper()), None)
        if not sname:
            wb.close(); return {}
        ws = wb[sname]
        hdr_row = ft_col = None
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, min(40, ws.max_column + 1)):
                v = ws.cell(r, c).value
                if v and "first" in str(v).lower() and "transaction" in str(v).lower():
                    ft_col, hdr_row = c, r; break
            if ft_col:
                break
        if not ft_col:
            wb.close(); return {}
        result = {}
        for r in range(hdr_row + 1, ws.max_row + 1):
            cid, ft = ws.cell(r, 1).value, ws.cell(r, ft_col).value
            if cid and isinstance(ft, datetime):
                result[str(cid).strip()] = ft
        wb.close()
        return result
    except Exception:
        return {}


def build_weekly_summary(
    jobs: list[dict],
    emails: dict[str, int],
    sales_map: dict[str, str],
    erp_file: str = DEFAULT_ERP_FILE,
    year: int | None = None,
    week: int | None = None,
) -> list[dict]:
    """One row per sales person with 12 KPI fields.
    KH MOI = first transaction in this week per CRM; KH SDDV = all others.
    """
    if year is None or week is None:
        y, w = iso_week(date.today())
        year = year or y
        week = week or w

    monday, sunday = week_bounds(year, week)
    week_start = datetime(monday.year, monday.month, monday.day)
    week_end = datetime(sunday.year, sunday.month, sunday.day, 23, 59, 59)

    crm_first = _load_crm_first_transactions(erp_file)

    data: dict[str, dict] = {
        name: {"shipments": 0, "vol_teu": 0.0, "profit": 0.0, "customers": set()}
        for name in SALES_MAP
    }

    for job in jobs:
        sales_name = job.get("sales", "Nelson")
        if sales_name not in data:
            data[sales_name] = {"shipments": 0, "vol_teu": 0.0, "profit": 0.0, "customers": set()}
        d = data[sales_name]
        d["shipments"] += 1
        cont = str(job.get("Container_Type") or "").upper().strip()
        qty = int(job.get("Quantity") or 1)
        teu = TEU_FACTOR.get(cont, 1)
        d["vol_teu"] += qty * teu
        d["profit"] += float(job.get("Profit") or 0)
        cid = job.get("CRM_ID")
        if cid:
            d["customers"].add(cid)

    email_counts: dict[str, int] = {}
    for key, cnt in emails.items():
        name = sales_map.get(key) or EMAIL_TO_SALES.get(key) or key
        email_counts[name] = email_counts.get(name, 0) + cnt

    rows: list[dict] = []
    for stt, sales_name in enumerate(SALES_MAP.keys(), start=1):
        d = data[sales_name]
        customers = d["customers"]

        new_custs = set()
        existing_custs = set()
        for cid in customers:
            ft = crm_first.get(cid)
            if ft and week_start <= ft <= week_end:
                new_custs.add(cid)
            else:
                existing_custs.add(cid)

        rows.append({
            "STT": stt,
            "Ten Sales": sales_name,
            "So shipment": d["shipments"],
            "VOL (TEU)": round(d["vol_teu"], 1),
            "PROFIT ($)": round(d["profit"], 0),
            "KH SDDV": len(existing_custs),
            "KH MOI": len(new_custs),
            "GAP KH": 0,
            "KH TIEM NANG": 0,
            "HOAT DONG TUAN": email_counts.get(sales_name, 0),
            "% HOAN THANH": "",
            "PLAN TUAN NAY": "",
        })

    return rows


_THIN = Side(style="thin", color="AAAAAA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BLUE_HDR = "1F4E79"
_ROW_ODD = "EAF0FB"
_ROW_EVEN = "FFFFFF"
_TOTAL_BG = "FFF2CC"

COLUMNS = [
    "STT", "Ten Sales", "So shipment", "VOL (TEU)", "PROFIT ($)",
    "KH SDDV", "KH MOI", "GAP KH", "KH TIEM NANG",
    "HOAT DONG TUAN", "% HOAN THANH", "PLAN TUAN NAY",
]
COL_WIDTHS = [5, 14, 12, 10, 12, 10, 10, 10, 14, 16, 14, 16]


_CENTER = Alignment(horizontal="center", vertical="center")

def _style_title(cell):
    cell.font = Font(bold=True, size=14, color=_BLUE_HDR, name="Segoe UI")
    cell.alignment = _CENTER

def _style_hdr(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=9, name="Segoe UI")
    cell.fill = PatternFill("solid", fgColor=_BLUE_HDR)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _BORDER

def _style_body(cell, odd: bool = True, bold: bool = False, fmt: str | None = None):
    cell.font = Font(size=10, name="Segoe UI", bold=bold)
    cell.fill = PatternFill("solid", fgColor=_ROW_ODD if odd else _ROW_EVEN)
    cell.alignment = _CENTER
    cell.border = _BORDER
    if fmt:
        cell.number_format = fmt

def _style_total(cell, fmt: str | None = None):
    cell.font = Font(bold=True, size=11, name="Segoe UI", color=_BLUE_HDR)
    cell.fill = PatternFill("solid", fgColor=_TOTAL_BG)
    cell.alignment = _CENTER
    cell.border = _BORDER
    if fmt:
        cell.number_format = fmt


_COL_FMT = {"PROFIT ($)": '"$"#,##0', "VOL (TEU)": "#,##0.0"}

def write_weekly_report(rows: list[dict], year: int, week: int, out_file: str) -> dict:
    """Write formatted xlsx. Returns stats dict."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"W{week:02d}_{year}"
    nc = len(COLUMNS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
    ws["A1"] = f"WEEKLY REPORT — Week {week} / {year}"
    _style_title(ws["A1"]); ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nc)
    ws["A2"] = "Nelson Freight (Pudong Prime) — Sales KPI"
    ws["A2"].font = Font(italic=True, size=9, color="555555")
    ws["A2"].alignment = _CENTER; ws.row_dimensions[2].height = 16
    for c, (label, width) in enumerate(zip(COLUMNS, COL_WIDTHS), start=1):
        _style_hdr(ws.cell(3, c, label))
        ws.column_dimensions[chr(ord("A") + c - 1)].width = width
    ws.row_dimensions[3].height = 32
    totals = {"So shipment": 0, "VOL (TEU)": 0.0, "PROFIT ($)": 0.0,
              "KH SDDV": 0, "KH MOI": 0, "GAP KH": 0, "KH TIEM NANG": 0, "HOAT DONG TUAN": 0}
    for idx, row in enumerate(rows):
        rr, odd = 4 + idx, idx % 2 == 0
        for c, col in enumerate(COLUMNS, start=1):
            _style_body(ws.cell(rr, c, row.get(col, "")), odd=odd, fmt=_COL_FMT.get(col))
        for k in totals:
            v = row.get(k, 0)
            if isinstance(v, (int, float)):
                totals[k] += v
        ws.row_dimensions[rr].height = 20
    tr = 4 + len(rows)
    ws.cell(tr, 1, "TOTAL")
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=2)
    for c, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(tr, c)
        if col in totals:
            cell.value = totals[col]
        _style_total(cell, fmt=_COL_FMT.get(col))
    ws.row_dimensions[tr].height = 22
    ws.freeze_panes = "A4"
    os.makedirs(os.path.dirname(out_file), exist_ok=True)
    wb.save(out_file); wb.close()
    return {"rows": len(rows), "total_shipments": totals["So shipment"],
            "total_vol_teu": totals["VOL (TEU)"], "total_profit": totals["PROFIT ($)"],
            "out_file": out_file}


def main() -> int:
    ap = argparse.ArgumentParser(description="Generate weekly sales KPI report")
    ap.add_argument("--year", type=int, default=None)
    ap.add_argument("--week", type=int, default=None)
    ap.add_argument("--erp", default=DEFAULT_ERP_FILE)
    ap.add_argument("--email-log", default=DEFAULT_EMAIL_LOG)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    if args.year and args.week:
        year, week = args.year, args.week
    else:
        year, week = iso_week(date.today())

    print(f"[+] Weekly KPI report — Year {year} Week {week}")

    jobs = load_active_jobs_for_week(args.erp, year, week)
    print(f"    -> {len(jobs)} jobs in week")

    emails = load_emails_for_week(args.email_log, year, week)
    total_emails = sum(emails.values())
    print(f"    -> {total_emails} emails sent this week")

    rows = build_weekly_summary(
        jobs, emails, EMAIL_TO_SALES,
        erp_file=args.erp, year=year, week=week,
    )

    out = args.out
    if not out:
        out = os.path.join(DEFAULT_OUT_DIR, f"WEEKLY_{year}_W{week:02d}.xlsx")

    stats = write_weekly_report(rows, year, week, out)
    print(f"\n[SUCCESS] Report written: {stats['out_file']}")
    print(f"    sales rows : {stats['rows']}")
    print(f"    shipments  : {stats['total_shipments']}")
    print(f"    volume TEU : {stats['total_vol_teu']:.1f}")
    print(f"    profit     : ${stats['total_profit']:,.0f}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
