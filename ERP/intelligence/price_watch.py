"""
price_watch.py — Price Watch / Re-quote Alert (Active Jobs v4 Feature 1)
=========================================================================
Monitors PENDING quotes in ERP_Master_v14.xlsm against latest Pricing Dry/Reefer.
When a carrier's buy price DROPS below the quoted buy rate by >= threshold,
fires a re-quote alert — Nelson can resend a lower offer to win the deal.

Inputs (ERP_Master_v14.xlsm):
  - Quotes sheet (42 cols, headers row 1): PENDING = Status blank or not WIN/LOST/EXPIRED
  - Pricing Dry (14+ cols) / Pricing Reefer — latest buy rates per POL/POD/Carrier/Cont

Outputs:
  - Fills Quotes Status cell yellow/red based on delta vs current buy
  - Creates/refreshes "Price_Watch" sheet listing alerts sorted by priority
  - Stamps Active Jobs col 35 (PRICE_WATCH_STATUS) + col 36 (PRICE_WATCH_DELTA)
    for WIN-converted quotes whose rate moved post-WIN

Usage:
    python ERP/intelligence/price_watch.py
    python ERP/intelligence/price_watch.py --threshold 50
"""
from __future__ import annotations

import argparse
import os
import sys
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import Final

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_CORE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "core")
_SCRIPTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "scripts")
sys.path.insert(0, os.path.abspath(_CORE_DIR))
sys.path.insert(0, os.path.abspath(_SCRIPTS_DIR))
from ribbon_guard import save_preserving_ribbon  # noqa: E402
from active_jobs_cols import COL as AJ_COL  # noqa: E402

# Phase 02: carrier alias + normalize
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from carrier_alias import normalize_carrier, CARRIER_ALIAS  # noqa: E402

# Telegram notify (optional — fails gracefully if env not set)
try:
    import importlib.util as _ilu
    _tg_spec = _ilu.spec_from_file_location(
        "notify_telegram",
        os.path.join(os.path.abspath(_SCRIPTS_DIR), "notify-telegram.py"),
    )
    _tg_mod = _ilu.module_from_spec(_tg_spec)  # type: ignore[arg-type]
    _tg_spec.loader.exec_module(_tg_mod)  # type: ignore[union-attr]
    _telegram_send = _tg_mod.send  # type: ignore[attr-defined]
except Exception:
    _telegram_send = None

sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]

DEFAULT_ERP_FILE: Final = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"

# Quotes sheet (header row 1)
Q_COL: Final = {
    "QuoteID": 1, "Date": 2, "Customer": 3, "Carrier": 4,
    "POL": 5, "POD": 6, "Place": 7, "Via": 8,
    "Eff": 9, "Exp": 10, "Source": 11,
    "Buy_20GP": 12, "Buy_40GP": 13, "Buy_40HC": 14, "Buy_45HC": 15,
    "Buy_40NOR": 16, "Buy_20RF": 17, "Buy_40RF": 18,
    "Sell_20GP": 29, "Sell_40GP": 30, "Sell_40HC": 31, "Sell_45HC": 32,
    "Sell_40NOR": 33, "Sell_20RF": 34, "Sell_40RF": 35,
    "Status": 36, "Remark": 37, "StatusDate": 38,
    "Qty": 39, "Volume": 40, "JobID": 41, "ContType": 42,
}

# Pricing Dry/Reefer sheet (header row 1)
P_COL: Final = {
    "POL": 1, "POD": 2, "Place": 3, "Carrier": 4, "Commodity": 5,
    "Eff": 6, "Exp": 7, "Note": 8, "Source": 9,
    "20GP": 10, "40GP": 11, "40HQ": 12, "45HQ": 13, "40NOR": 14,
    "20RF": 10, "40RF": 11,  # Reefer sheet uses different positions
}

CONT_TO_BUY_COL: Final = {
    "20GP": "Buy_20GP", "40GP": "Buy_40GP", "40HC": "Buy_40HC",
    "40HQ": "Buy_40HC", "45HC": "Buy_45HC", "45HQ": "Buy_45HC",
    "40NOR": "Buy_40NOR", "20RF": "Buy_20RF", "40RF": "Buy_40RF",
}
CONT_TO_PRICE_COL: Final = {
    "20GP": ("Dry", "20GP"), "40GP": ("Dry", "40GP"),
    "40HC": ("Dry", "40HQ"), "40HQ": ("Dry", "40HQ"),
    "45HC": ("Dry", "45HQ"), "45HQ": ("Dry", "45HQ"),
    "40NOR": ("Dry", "40NOR"),
    "20RF": ("Reefer", "20RF"), "40RF": ("Reefer", "40RF"),
}


@dataclass
class Alert:
    quote_id: str
    row: int
    customer: str
    route: str
    carrier: str
    cont_type: str
    quoted_buy: float
    current_buy: float
    delta: float
    kind: str  # DROP | RISE | NO_MATCH
    priority: str  # P1 | P2 | P3
    action: str
    tier: str = ""       # "ROUTINE" | "LINE" | "" (legacy)
    carrier_new: str = ""  # best alternative carrier (Tier 1 only)


# ── Phase 02 dataclasses ──
@dataclass
class PricingRow:
    """Normalized pricing record used by Phase 02 dual-index."""
    pol: str
    pod: str
    place: str
    carrier_raw: str
    carrier_norm: str
    eff: object       # datetime | None
    exp: object       # datetime | None
    source: str       # "Dry" or "Reefer"
    buy_by_cont: dict = field(default_factory=dict)  # {"20GP": 1200, "40HQ": 2500}


@dataclass
class PWConfig:
    """Price Watch runtime configuration — can be overridden via PW_Config sheet or CLI."""
    threshold_routine: float = 100.0
    threshold_line: float = 50.0
    enabled_tier1: bool = True
    enabled_tier2: bool = True
    ignore_expired: bool = True
    autorun_on_refresh: bool = True


# ── Loaders ──
def load_latest_pricing(wb) -> dict:
    """Return dict: (pol, pod, place, carrier, cont_type) → (buy, eff_date, source)."""
    out: dict[tuple, tuple[float, datetime | None, str]] = {}

    def _ingest(sheet_name: str, cont_cols: dict[str, int]) -> int:
        if sheet_name not in wb.sheetnames:
            return 0
        ws = wb[sheet_name]
        n = 0
        for r in range(2, ws.max_row + 1):
            pol = (ws.cell(r, 1).value or "")
            if not pol:
                continue
            pod = (ws.cell(r, 2).value or "")
            place = (ws.cell(r, 3).value or "")
            carrier = (ws.cell(r, 4).value or "")
            eff = ws.cell(r, 6).value
            source = ws.cell(r, 9).value or ""
            for cont, col in cont_cols.items():
                val = ws.cell(r, col).value
                if not isinstance(val, (int, float)) or val <= 0:
                    continue
                key = (str(pol).upper().strip(),
                       str(pod).upper().strip(),
                       str(place).upper().strip(),
                       str(carrier).upper().strip(),
                       cont)
                prev = out.get(key)
                # keep the most recent eff date
                if prev is None or (isinstance(eff, datetime)
                                    and (not isinstance(prev[1], datetime) or eff > prev[1])):
                    out[key] = (float(val), eff if isinstance(eff, datetime) else None, str(source))
                n += 1
        return n

    n_dry = _ingest("Pricing Dry", {"20GP": 10, "40GP": 11, "40HC": 12, "45HC": 13, "40NOR": 14})
    n_rf = _ingest("Pricing Reefer", {"20RF": 10, "40RF": 11})
    print(f"    -> pricing scanned: dry={n_dry} reefer={n_rf} unique_keys={len(out)}")
    return out


def iter_pending_quotes(wb):
    """Yield (row_idx, quote_dict) for PENDING quotes (Status not WIN/LOST/EXPIRED/blank is PENDING)."""
    ws = wb["Quotes"]
    for r in range(2, ws.max_row + 1):
        qid = ws.cell(r, Q_COL["QuoteID"]).value
        if not qid:
            continue
        status = (ws.cell(r, Q_COL["Status"]).value or "").strip().upper()
        if status in ("LOST", "EXPIRED"):
            continue
        # "" (blank) and "WIN" both matter: blank = pending re-quote; WIN = monitor for post-WIN moves
        q = {k: ws.cell(r, c).value for k, c in Q_COL.items()}
        q["_row"] = r
        q["_status"] = status or "PENDING"
        yield r, q


# ── Comparison ──
def compute_alerts(quotes, pricing_latest: dict, threshold: float) -> list[Alert]:
    alerts: list[Alert] = []
    for row, q in quotes:
        qid = str(q.get("QuoteID") or "").strip()
        carrier = str(q.get("Carrier") or "").upper().strip()
        pol = str(q.get("POL") or "").upper().strip()
        pod = str(q.get("POD") or "").upper().strip()
        place = str(q.get("Place") or "").upper().strip()
        cust = str(q.get("Customer") or "")

        # For each container type the quote has a Buy rate for
        for cont, buy_key in CONT_TO_BUY_COL.items():
            quoted = q.get(buy_key)
            if not isinstance(quoted, (int, float)) or quoted <= 0:
                continue

            _, price_cont = CONT_TO_PRICE_COL[cont]
            # Exact match: POL + POD + Place + Carrier + Cont
            key = (pol, pod, place, carrier, price_cont)
            cur = pricing_latest.get(key)
            if cur is None:
                # Fallback 1: same POL/POD/Carrier/Cont, Place == POD (direct port, no inland)
                key2 = (pol, pod, pod, carrier, price_cont)
                cur = pricing_latest.get(key2)
            if cur is None:
                # Fallback 2: fuzzy carrier — "ONE" ⊂ "Ocean Network Express"
                matches = [v for k, v in pricing_latest.items()
                           if k[0] == pol and k[1] == pod and k[2] == place
                           and k[4] == price_cont
                           and carrier and (carrier in k[3] or k[3] in carrier)]
                if not matches:
                    continue
                cur = matches[0]

            current_buy = cur[0]
            delta = current_buy - float(quoted)  # negative = price dropped

            if abs(delta) < threshold:
                continue

            if delta < 0:
                alerts.append(Alert(
                    quote_id=qid, row=row, customer=cust,
                    route=f"{pol}-{pod}", carrier=carrier, cont_type=cont,
                    quoted_buy=float(quoted), current_buy=current_buy,
                    delta=delta, kind="DROP",
                    priority="P1" if q["_status"] == "PENDING" else "P2",
                    action=f"Re-quote {cust}: buy dropped ${abs(delta):,.0f} ({cont})",
                ))
            else:
                alerts.append(Alert(
                    quote_id=qid, row=row, customer=cust,
                    route=f"{pol}-{pod}", carrier=carrier, cont_type=cont,
                    quoted_buy=float(quoted), current_buy=current_buy,
                    delta=delta, kind="RISE",
                    priority="P2" if q["_status"] == "WIN" else "P3",
                    action=f"⚠ Cost rose ${delta:,.0f} ({cont}) — margin squeeze",
                ))
    return alerts


# ── Phase 02: Dual-Index Loader ──
def load_pricing_v2(wb) -> tuple[dict, dict]:
    """Build two price indices in a single pass over Pricing Dry + Reefer sheets.

    Returns:
        pricing_by_routine: dict[(POL, POD, CONT)] -> list[PricingRow]
            Used by Tier 1 — any carrier, cheapest wins.
        pricing_by_line:    dict[(POL, POD, CARRIER_NORM, CONT)] -> list[PricingRow]
            Used by Tier 2 — same carrier as quote.
    """
    pricing_by_routine: dict[tuple, list] = {}
    pricing_by_line: dict[tuple, list] = {}

    sheet_cont_map = {
        "Pricing Dry":   {"20GP": 10, "40GP": 11, "40HC": 12, "45HC": 13, "40NOR": 14},
        "Pricing Reefer": {"20RF": 10, "40RF": 11},
    }

    total_rows = 0
    for sheet_name, cont_cols in sheet_cont_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        src_label = "Dry" if "Dry" in sheet_name else "Reefer"

        for r in range(2, ws.max_row + 1):
            raw_pol = ws.cell(r, 1).value
            if not raw_pol:
                continue
            pol = str(raw_pol).upper().strip()
            pod = str(ws.cell(r, 2).value or "").upper().strip()
            place = str(ws.cell(r, 3).value or "").upper().strip()
            carrier_raw = str(ws.cell(r, 4).value or "").strip()
            carrier_norm = normalize_carrier(carrier_raw)
            eff = ws.cell(r, 6).value
            exp = ws.cell(r, 7).value
            source = str(ws.cell(r, 9).value or src_label)

            # Collect valid buy rates for this row across cont types
            buy_by_cont: dict[str, float] = {}
            for cont, col in cont_cols.items():
                val = ws.cell(r, col).value
                if isinstance(val, (int, float)) and val > 0:
                    buy_by_cont[cont] = float(val)

            if not buy_by_cont:
                continue

            pr = PricingRow(
                pol=pol, pod=pod, place=place,
                carrier_raw=carrier_raw, carrier_norm=carrier_norm,
                eff=eff, exp=exp, source=source,
                buy_by_cont=buy_by_cont,
            )

            # Build both indices for each cont type in this row
            for cont, buy in buy_by_cont.items():
                # Tier 1 index: (POL, POD, CONT)
                r_key = (pol, pod, cont)
                pricing_by_routine.setdefault(r_key, []).append(pr)

                # Tier 2 index: (POL, POD, CARRIER_NORM, CONT)
                l_key = (pol, pod, carrier_norm, cont)
                pricing_by_line.setdefault(l_key, []).append(pr)

            total_rows += 1

    print(f"    -> v2 pricing rows ingested: {total_rows}  "
          f"routine_keys={len(pricing_by_routine)}  line_keys={len(pricing_by_line)}")
    return pricing_by_routine, pricing_by_line


def load_pw_config(wb) -> PWConfig:
    """Read PWConfig from PW_Config sheet if it exists; return defaults otherwise.
    Creates the sheet with defaults if missing (for first-run setup).
    """
    cfg = PWConfig()

    if "PW_Config" not in wb.sheetnames:
        ws = wb.create_sheet("PW_Config")
        ws["A1"], ws["B1"] = "Key", "Value"
        ws["A2"], ws["B2"] = "threshold_routine", cfg.threshold_routine
        ws["A3"], ws["B3"] = "threshold_line", cfg.threshold_line
        ws["A4"], ws["B4"] = "enabled_tier1", str(cfg.enabled_tier1)
        ws["A5"], ws["B5"] = "enabled_tier2", str(cfg.enabled_tier2)
        ws["A6"], ws["B6"] = "ignore_expired", str(cfg.ignore_expired)
        ws["A7"], ws["B7"] = "autorun_on_refresh", str(cfg.autorun_on_refresh)
        return cfg

    ws = wb["PW_Config"]
    kv: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        k = ws.cell(r, 1).value
        v = ws.cell(r, 2).value
        if k and v is not None:
            kv[str(k).strip()] = str(v).strip()

    def _float(key: str, default: float) -> float:
        try:
            return float(kv.get(key, default))
        except (ValueError, TypeError):
            return default

    def _bool(key: str, default: bool) -> bool:
        val = kv.get(key, "").upper()
        if val in ("TRUE", "1", "YES"):
            return True
        if val in ("FALSE", "0", "NO"):
            return False
        return default

    cfg.threshold_routine = _float("threshold_routine", cfg.threshold_routine)
    cfg.threshold_line = _float("threshold_line", cfg.threshold_line)
    cfg.enabled_tier1 = _bool("enabled_tier1", cfg.enabled_tier1)
    cfg.enabled_tier2 = _bool("enabled_tier2", cfg.enabled_tier2)
    cfg.ignore_expired = _bool("ignore_expired", cfg.ignore_expired)
    cfg.autorun_on_refresh = _bool("autorun_on_refresh", cfg.autorun_on_refresh)
    return cfg


def compute_alerts_v2(
    quotes: list[tuple[int, dict]],
    pricing_by_routine: dict,
    pricing_by_line: dict,
    cfg: PWConfig,
) -> list[Alert]:
    """Two-tier alert detection replacing compute_alerts().

    Tier 2 (LINE): same carrier as quote lowered its price → P2 alert
    Tier 1 (ROUTINE): a *different* carrier is cheaper than threshold → P1 alert

    Both tiers can fire for the same quote+cont — both are emitted.
    Tier 1 sorts above Tier 2 in the output sheet.
    """
    today = datetime.now().date()
    alerts: list[Alert] = []

    # Map quote cont type to pricing sheet cont code
    quote_to_price_cont: dict[str, str] = {
        "20GP": "20GP", "40GP": "40GP",
        "40HC": "40HC", "40HQ": "40HC",
        "45HC": "45HC", "45HQ": "45HC",
        "40NOR": "40NOR",
        "20RF": "20RF", "40RF": "40RF",
    }

    for row, q in quotes:
        qid = str(q.get("QuoteID") or "").strip()
        carrier_raw = str(q.get("Carrier") or "").strip()
        carrier_norm = normalize_carrier(carrier_raw)
        pol = str(q.get("POL") or "").upper().strip()
        pod = str(q.get("POD") or "").upper().strip()
        cust = str(q.get("Customer") or "")
        status = q.get("_status", "PENDING")

        # Skip expired quotes if configured
        if cfg.ignore_expired:
            exp_date = q.get("Exp")
            if isinstance(exp_date, datetime) and exp_date.date() < today:
                continue

        for cont, buy_key in CONT_TO_BUY_COL.items():
            quoted = q.get(buy_key)
            if not isinstance(quoted, (int, float)) or quoted <= 0:
                continue
            quoted_f = float(quoted)

            price_cont = quote_to_price_cont.get(cont, cont)

            # ── Tier 2 (LINE): same carrier, same route ──
            if cfg.enabled_tier2:
                l_key = (pol, pod, carrier_norm, price_cont)
                line_rows = pricing_by_line.get(l_key, [])
                if line_rows:
                    best_buy = min(pr.buy_by_cont.get(price_cont, float("inf"))
                                   for pr in line_rows)
                    delta = best_buy - quoted_f  # negative = dropped
                    if delta < 0 and abs(delta) >= cfg.threshold_line:
                        priority = "P2" if status == "WIN" else "P2"
                        alerts.append(Alert(
                            quote_id=qid, row=row, customer=cust,
                            route=f"{pol}-{pod}",
                            carrier=carrier_raw, cont_type=cont,
                            quoted_buy=quoted_f, current_buy=best_buy,
                            delta=delta, kind="DROP",
                            priority=priority,
                            action=(f"[Tier2 LINE] {carrier_raw} hạ giá "
                                    f"${abs(delta):,.0f} ({cont}) — kiểm tra lại offer"),
                            tier="LINE",
                            carrier_new=carrier_raw,
                        ))

            # ── Tier 1 (ROUTINE): any carrier, exclude same carrier ──
            if cfg.enabled_tier1:
                r_key = (pol, pod, price_cont)
                routine_rows = pricing_by_line  # just using the pool via routine index
                all_rows = pricing_by_routine.get(r_key, [])

                # Filter out same carrier (that's Tier 2)
                alt_rows = [
                    pr for pr in all_rows
                    if pr.carrier_norm != carrier_norm
                ]
                if not alt_rows:
                    # Fallback: try place==pod match
                    r_key2 = (pol, pod, price_cont)
                    alt_rows = [
                        pr for pr in pricing_by_routine.get(r_key2, [])
                        if pr.carrier_norm != carrier_norm
                    ]

                if alt_rows:
                    # Find cheapest alternative carrier
                    best_pr = min(
                        alt_rows,
                        key=lambda pr: pr.buy_by_cont.get(price_cont, float("inf")),
                    )
                    best_buy = best_pr.buy_by_cont.get(price_cont, float("inf"))
                    if best_buy == float("inf"):
                        continue
                    delta = best_buy - quoted_f  # negative = cheaper alternative exists
                    if delta < 0 and abs(delta) >= cfg.threshold_routine:
                        priority = "P1" if status != "WIN" else "P2"
                        alerts.append(Alert(
                            quote_id=qid, row=row, customer=cust,
                            route=f"{pol}-{pod}",
                            carrier=carrier_raw, cont_type=cont,
                            quoted_buy=quoted_f, current_buy=best_buy,
                            delta=delta, kind="DROP",
                            priority=priority,
                            action=(f"[Tier1 ROUTINE] {best_pr.carrier_norm} rẻ hơn "
                                    f"${abs(delta):,.0f} ({cont}) — cân nhắc đổi carrier"),
                            tier="ROUTINE",
                            carrier_new=best_pr.carrier_norm,
                        ))

    # Sort: Tier 1 above Tier 2, then by abs delta desc
    tier_order = {"ROUTINE": 0, "LINE": 1, "": 2}
    priority_order = {"P1": 0, "P2": 1, "P3": 2}
    alerts.sort(key=lambda a: (
        priority_order.get(a.priority, 9),
        tier_order.get(a.tier, 9),
        -abs(a.delta),
    ))
    return alerts


# ── Target Watch ──

# Col indices (1-based) matching docs/s1v2-target-watch-schema.md
TW_COL: Final = {
    "Target_ID": 1, "Created": 2, "QuoteID": 3, "Customer": 4,
    "POL": 5, "POD": 6, "Carrier": 7, "ContType": 8,
    "Target_USD": 9, "CurrentQuote_USD": 10, "Status": 11,
    "LastCheck": 12, "Matched_Rate": 13, "Matched_Carrier": 14,
    "Matched_Date": 15, "Remark": 16,
}
TW_HEADERS: Final = [
    "Target_ID", "Created", "QuoteID", "Customer",
    "POL", "POD", "Carrier", "ContType",
    "Target_USD", "CurrentQuote_USD", "Status",
    "LastCheck", "Matched_Rate", "Matched_Carrier",
    "Matched_Date", "Remark",
]
TW_WATCH_EXPIRE_DAYS: Final = 30


def ensure_target_watch_sheet(wb) -> object:
    """Return Target_Watch worksheet, creating it with headers if missing."""
    if "Target_Watch" in wb.sheetnames:
        return wb["Target_Watch"]

    ws = wb.create_sheet("Target_Watch")
    # Header row
    for i, h in enumerate(TW_HEADERS, 1):
        cell = ws.cell(1, i, h)
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="Segoe UI")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    widths = [18, 16, 16, 20, 8, 8, 12, 8, 12, 16, 12, 16, 14, 16, 16, 24]
    for i, w in enumerate(widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(TW_HEADERS))}1"
    print("    -> created Target_Watch sheet with headers")
    return ws


@dataclass
class TargetMatch:
    """Result of a single Target Watch row evaluation."""
    row: int
    target_id: str
    customer: str
    pol: str
    pod: str
    cont_type: str
    target_usd: float
    buy_rate: float
    markup: float
    sell_rate: float          # buy + markup
    matched_carrier: str
    matched: bool             # True = MATCHED, False = still watching
    expired: bool = False


def _find_min_buy(
    pol: str,
    pod: str,
    cont_type: str,
    carrier_filter: str,    # "" or "ANY" means all carriers
    pricing_by_routine: dict,
    pricing_by_line: dict,
    carrier_norm_filter: str = "",
) -> tuple[float, str]:
    """Return (min_buy, best_carrier_raw) for the given lane+cont.

    If carrier_filter is non-empty and not ANY, only check that carrier (line index).
    Otherwise scan all carriers in routine index.
    """
    # Map cont_type to pricing sheet key
    cont_map = {
        "20GP": "20GP", "40GP": "40GP",
        "40HC": "40HC", "40HQ": "40HC",
        "45HC": "45HC", "45HQ": "45HC",
        "40NOR": "40NOR",
        "20RF": "20RF", "40RF": "40RF",
    }
    price_cont = cont_map.get(cont_type.upper(), cont_type.upper())

    best_buy = float("inf")
    best_carrier = ""

    use_line = carrier_filter and carrier_filter.upper() not in ("ANY", "")
    if use_line:
        l_key = (pol, pod, carrier_norm_filter, price_cont)
        rows = pricing_by_line.get(l_key, [])
        for pr in rows:
            buy = pr.buy_by_cont.get(price_cont, float("inf"))
            if buy < best_buy:
                best_buy = buy
                best_carrier = pr.carrier_raw
    else:
        r_key = (pol, pod, price_cont)
        rows = pricing_by_routine.get(r_key, [])
        for pr in rows:
            buy = pr.buy_by_cont.get(price_cont, float("inf"))
            if buy < best_buy:
                best_buy = buy
                best_carrier = pr.carrier_raw

    return (best_buy if best_buy != float("inf") else -1.0), best_carrier


def scan_target_matches(
    wb,
    pricing_by_routine: dict,
    pricing_by_line: dict,
    default_markup: float = 200.0,
) -> list[TargetMatch]:
    """Scan Target_Watch sheet for WATCHING rows and evaluate against pricing.

    - Updates Status to MATCHED or EXPIRED in-place on the worksheet.
    - Returns list of TargetMatch for new MATCHes (for Telegram + Price_Watch section).
    - MATCHED rows are skipped (idempotent).
    - EXPIRED: Created + 30 days < now → Status = EXPIRED.
    """
    ws_tw = ensure_target_watch_sheet(wb)
    matches: list[TargetMatch] = []
    now = datetime.now()

    watching_count = 0
    for r in range(2, ws_tw.max_row + 1):
        status_cell = ws_tw.cell(r, TW_COL["Status"])
        status = (status_cell.value or "").strip().upper()

        # Only process WATCHING rows
        if status != "WATCHING":
            continue
        watching_count += 1

        target_id = str(ws_tw.cell(r, TW_COL["Target_ID"]).value or "").strip()
        customer = str(ws_tw.cell(r, TW_COL["Customer"]).value or "").strip()
        pol = str(ws_tw.cell(r, TW_COL["POL"]).value or "").upper().strip()
        pod = str(ws_tw.cell(r, TW_COL["POD"]).value or "").upper().strip()
        carrier_raw = str(ws_tw.cell(r, TW_COL["Carrier"]).value or "").strip()
        cont_type = str(ws_tw.cell(r, TW_COL["ContType"]).value or "").strip()
        target_raw = ws_tw.cell(r, TW_COL["Target_USD"]).value
        created_val = ws_tw.cell(r, TW_COL["Created"]).value

        # Parse target USD
        try:
            target_usd = float(target_raw or 0)
        except (ValueError, TypeError):
            target_usd = 0.0

        if target_usd <= 0 or not pol or not pod or not cont_type:
            continue

        # Check expiry
        if isinstance(created_val, datetime):
            age_days = (now - created_val).days
        else:
            age_days = 0

        if age_days > TW_WATCH_EXPIRE_DAYS:
            status_cell.value = "EXPIRED"
            ws_tw.cell(r, TW_COL["LastCheck"]).value = now
            continue

        # Resolve carrier filter
        carrier_norm = normalize_carrier(carrier_raw) if carrier_raw else ""
        use_specific = bool(carrier_raw) and carrier_raw.upper() not in ("ANY", "")

        # Find min buy rate
        min_buy, best_carrier = _find_min_buy(
            pol=pol, pod=pod, cont_type=cont_type,
            carrier_filter=carrier_raw,
            pricing_by_routine=pricing_by_routine,
            pricing_by_line=pricing_by_line,
            carrier_norm_filter=carrier_norm if use_specific else "",
        )

        # Stamp LastCheck
        ws_tw.cell(r, TW_COL["LastCheck"]).value = now

        if min_buy < 0:
            # No pricing data found
            continue

        sell_rate = min_buy + default_markup

        tm = TargetMatch(
            row=r, target_id=target_id, customer=customer,
            pol=pol, pod=pod, cont_type=cont_type,
            target_usd=target_usd,
            buy_rate=min_buy, markup=default_markup, sell_rate=sell_rate,
            matched_carrier=best_carrier,
            matched=(sell_rate <= target_usd),
        )

        if tm.matched:
            # Update worksheet in-place
            status_cell.value = "MATCHED"
            ws_tw.cell(r, TW_COL["Matched_Rate"]).value = round(sell_rate, 2)
            ws_tw.cell(r, TW_COL["Matched_Carrier"]).value = best_carrier
            ws_tw.cell(r, TW_COL["Matched_Date"]).value = now
            # Highlight matched row
            match_fill = PatternFill("solid", fgColor="D1FAE5")
            for c in range(1, len(TW_HEADERS) + 1):
                ws_tw.cell(r, c).fill = match_fill
            matches.append(tm)

    print(f"    -> target_watch: watching={watching_count} new_matches={len(matches)}")
    return matches


def _send_target_match_alerts(matches: list[TargetMatch]) -> None:
    """Fire Telegram notification for each new target match. Fails silently."""
    if not matches or _telegram_send is None:
        return
    for tm in matches:
        msg = (
            f"<b>TARGET MATCHED</b>\n"
            f"Customer: {tm.customer}\n"
            f"Route: {tm.pol}-{tm.pod}  {tm.cont_type}\n"
            f"Carrier: {tm.matched_carrier}\n"
            f"Buy: ${tm.buy_rate:,.0f} + markup ${tm.markup:,.0f} = <b>${tm.sell_rate:,.0f}</b>\n"
            f"Target: ${tm.target_usd:,.0f}  "
            f"(saved ${tm.target_usd - tm.sell_rate:,.0f})"
        )
        try:
            _telegram_send(msg, parse_mode="HTML", silent=False)
        except Exception as exc:
            print(f"    -> telegram send failed: {exc}")


def write_target_matches_section(ws_pw, matches: list[TargetMatch]) -> None:
    """Write/overwrite the TARGET MATCHES block at the bottom of Price_Watch sheet.

    Clears any existing section (rows tagged with TARGET_MATCH sentinel) then
    re-writes fresh, so it never accumulates stale rows.
    """
    # Find last row with content (skip empty rows at bottom)
    last_data_row = ws_pw.max_row
    while last_data_row > 1 and ws_pw.cell(last_data_row, 1).value is None:
        last_data_row -= 1

    # Delete old TARGET MATCHES block: scan upward for sentinel tag
    sentinel = "TARGET_MATCH_SECTION"
    start_row = None
    for r in range(last_data_row, 0, -1):
        val = str(ws_pw.cell(r, 1).value or "")
        if sentinel in val or "TARGET MATCHES" in val.upper():
            start_row = r
            break

    if start_row is not None:
        # Clear from start_row to last_data_row
        for r in range(start_row, last_data_row + 1):
            for c in range(1, 14):
                ws_pw.cell(r, c).value = None
                ws_pw.cell(r, c).fill = PatternFill(fill_type=None)

    # Write new block starting 2 rows below last content
    section_start = (start_row if start_row else last_data_row + 2)

    # Section header row
    n_cols = 10
    ws_pw.merge_cells(f"A{section_start}:J{section_start}")
    title_cell = ws_pw.cell(section_start, 1)
    title_cell.value = f"{sentinel} | TARGET MATCHES — {datetime.now():%d %b %Y %H:%M}"
    title_cell.font = Font(bold=True, size=12, color="1F4E79")
    title_cell.fill = PatternFill("solid", fgColor="EFF6FF")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_pw.row_dimensions[section_start].height = 20

    if not matches:
        ws_pw.cell(section_start + 1, 1).value = "No target matches this run."
        ws_pw.cell(section_start + 1, 1).font = Font(italic=True, color="666666")
        return

    # Column headers
    th_row = section_start + 1
    hdrs = ["Target_ID", "Customer", "POL", "POD", "ContType",
            "Target $", "Sell Rate", "Buy Rate", "Matched Carrier", "Saved $"]
    for i, h in enumerate(hdrs, 1):
        c = ws_pw.cell(th_row, i, h)
        c.font = Font(bold=True, color="FFFFFF", size=9, name="Segoe UI")
        c.fill = PatternFill("solid", fgColor="166534")
        c.alignment = Alignment(horizontal="center")

    fill_match = PatternFill("solid", fgColor="D1FAE5")
    for r_offset, tm in enumerate(matches, start=2):
        r = section_start + r_offset
        saved = tm.target_usd - tm.sell_rate
        row_data = [
            tm.target_id, tm.customer, tm.pol, tm.pod, tm.cont_type,
            tm.target_usd, tm.sell_rate, tm.buy_rate, tm.matched_carrier, saved,
        ]
        for i, v in enumerate(row_data, 1):
            cell = ws_pw.cell(r, i, v)
            cell.font = Font(size=9, name="Segoe UI")
            cell.fill = fill_match
            cell.alignment = Alignment(horizontal="center")
            if i in (6, 7, 8, 10):
                cell.number_format = '"$"#,##0'


# ── Writers ──
FILL_ALERT = PatternFill("solid", fgColor="FEE2E2")
FILL_WARN = PatternFill("solid", fgColor="FEF3C7")
FILL_OK = PatternFill("solid", fgColor="D1FAE5")
THIN = Side(style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def stamp_quotes_sheet(wb, alerts: list[Alert]):
    """Color the Status cell of each alerted quote row."""
    ws = wb["Quotes"]
    touched_rows: set[int] = set()
    by_row: dict[int, list[Alert]] = {}
    for a in alerts:
        by_row.setdefault(a.row, []).append(a)
    for r, row_alerts in by_row.items():
        # strongest signal = max priority drop
        has_drop = any(a.kind == "DROP" for a in row_alerts)
        cell = ws.cell(r, Q_COL["Status"])
        cell.fill = FILL_ALERT if has_drop else FILL_WARN
        # Remark col: prepend alert
        remark_cell = ws.cell(r, Q_COL["Remark"])
        note = "; ".join(f"{a.kind}:{a.cont_type}:${a.delta:+,.0f}" for a in row_alerts)
        existing = (remark_cell.value or "").strip()
        tag = f"[PW {datetime.now():%d%b %H:%M}] {note}"
        if existing and not existing.startswith("[PW "):
            remark_cell.value = f"{tag} | {existing}"
        else:
            remark_cell.value = tag
        touched_rows.add(r)
    return touched_rows


def write_price_watch_sheet(wb, alerts: list[Alert]):
    """Create/refresh Price_Watch summary sheet.
    Phase 02: adds Tier + Alt Carrier columns (12 total).
    Defect 4 fix: always writes header row + summary row even when 0 alerts.
    """
    if "Price_Watch" in wb.sheetnames:
        del wb["Price_Watch"]
    ws = wb.create_sheet("Price_Watch")

    n_cols = 12  # extended in Phase 02

    # Title row (row 1)
    ws.merge_cells(f"A1:{chr(64 + n_cols)}1")
    ws["A1"] = f"PRICE WATCH v2 — {datetime.now():%d %b %Y %H:%M}"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Header row (row 3)
    hdrs = ["Priority", "Tier", "Kind", "QuoteID", "Customer", "Route",
            "Carrier", "Alt Carrier", "Cont", "Quoted Buy", "Current Buy", "Δ Delta"]
    widths = [10, 10, 8, 12, 18, 14, 12, 12, 8, 12, 12, 12]
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(3, i, h)
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Segoe UI")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # No alerts: write a single summary row (Defect 4 fix)
    if not alerts:
        ws.merge_cells(f"A4:{chr(64 + n_cols)}4")
        ws["A4"] = f"No alerts as of {datetime.now():%d %b %Y %H:%M}"
        ws["A4"].font = Font(italic=True, color="666666", size=10, name="Segoe UI")
        ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A4"
        return

    # Sort by priority → tier → abs delta desc (compute_alerts_v2 already sorted,
    # but legacy compute_alerts output may need re-sort)
    priority_order = {"P1": 0, "P2": 1, "P3": 2}
    tier_order = {"ROUTINE": 0, "LINE": 1, "": 2}
    alerts_sorted = sorted(alerts, key=lambda a: (
        priority_order.get(a.priority, 9),
        tier_order.get(getattr(a, "tier", ""), 9),
        -abs(a.delta),
    ))

    for r, a in enumerate(alerts_sorted, start=4):
        tier_label = getattr(a, "tier", "") or "—"
        carrier_new = getattr(a, "carrier_new", "") or "—"
        row_data = [
            a.priority, tier_label, a.kind, a.quote_id, a.customer,
            a.route, a.carrier, carrier_new, a.cont_type,
            a.quoted_buy, a.current_buy, a.delta,
        ]
        for i, v in enumerate(row_data, 1):
            cell = ws.cell(r, i, v)
            cell.font = Font(size=10, name="Segoe UI")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
            if i in (10, 11, 12):
                cell.number_format = '"$"#,##0'
        # Row background by priority
        fill = FILL_ALERT if a.priority == "P1" else FILL_WARN if a.priority == "P2" else None
        if fill:
            for i in range(1, n_cols + 1):
                ws.cell(r, i).fill = fill
        # Delta cell bold + color
        dc = ws.cell(r, 12)
        dc.font = Font(size=10, name="Segoe UI", bold=True,
                       color="00804A" if a.delta < 0 else "C00000")

    ws.freeze_panes = "A4"


def stamp_active_jobs(wb, alerts: list[Alert]):
    """Write PRICE_WATCH_STATUS (col 35) + PRICE_WATCH_DELTA (col 36) for WIN quotes linked to Active Jobs."""
    if "Active Jobs" not in wb.sheetnames:
        return 0
    ws = wb["Active Jobs"]
    # Build map: quote_id → [alerts]
    by_qid: dict[str, list[Alert]] = {}
    for a in alerts:
        by_qid.setdefault(a.quote_id, []).append(a)

    # Active Jobs rows: match via Bkg_No (col 4) OR Notes (col 24) referencing Quote_ID?
    # Actually the quote has JobID (col 41) which is the Active Jobs Job_ID (Nelson uses CRM_ID col 1 as Job_ID in v14).
    # For now match by customer name via CRM_ID.
    stamped = 0
    for r in range(8, ws.max_row + 1):
        crm = ws.cell(r, AJ_COL["CRM_ID"]).value
        if not crm:
            continue
        # Find any alert whose customer name matches CRM_ID (best-effort)
        crm_up = str(crm).upper().strip()
        matched = []
        for a in alerts:
            if a.customer and a.customer.upper().strip() in crm_up:
                matched.append(a)
            elif crm_up in (a.customer or "").upper().strip():
                matched.append(a)
        if not matched:
            continue
        # Strongest signal: largest magnitude drop = most actionable
        drops = [a for a in matched if a.kind == "DROP"]
        pick = max(drops, key=lambda a: abs(a.delta)) if drops else min(matched, key=lambda a: -abs(a.delta))
        ws.cell(r, AJ_COL["PRICE_WATCH_STATUS"], pick.kind)
        ws.cell(r, AJ_COL["PRICE_WATCH_DELTA"], round(pick.delta))
        ws.cell(r, AJ_COL["PRICE_WATCH_STATUS"]).fill = FILL_ALERT if pick.kind == "DROP" else FILL_WARN
        stamped += 1
    return stamped


# ── Main ──
def main() -> int:
    ap = argparse.ArgumentParser(
        description="Price Watch v2 — Two-tier re-quote alert engine"
    )
    ap.add_argument("--erp", default=DEFAULT_ERP_FILE)
    ap.add_argument("--threshold", type=float, default=50.0,
                    help="Legacy single threshold (used by --tier line/routine alias). Default: 50")
    ap.add_argument("--threshold-routine", type=float, default=None,
                    help="Tier 1 ROUTINE alert threshold USD (default: from PW_Config or 100)")
    ap.add_argument("--threshold-line", type=float, default=None,
                    help="Tier 2 LINE alert threshold USD (default: from PW_Config or 50)")
    ap.add_argument("--tier", choices=["all", "routine", "line"], default="all",
                    help="Which tier(s) to run: all | routine | line (default: all)")
    args = ap.parse_args()

    if not os.path.exists(args.erp):
        print(f"[ERROR] ERP file not found: {args.erp}")
        return 1
    try:
        with open(args.erp, "r+b"):
            pass
    except PermissionError:
        print(f"[ERROR] ERP file is open in Excel. Close it first.")
        return 2

    print(f"[+] Price Watch v2 run @ {datetime.now():%Y-%m-%d %H:%M}")
    wb = openpyxl.load_workbook(args.erp, keep_vba=True)

    # Load config from sheet, then apply CLI overrides
    cfg = load_pw_config(wb)
    if args.threshold_routine is not None:
        cfg.threshold_routine = args.threshold_routine
    if args.threshold_line is not None:
        cfg.threshold_line = args.threshold_line
    if args.tier == "routine":
        cfg.enabled_tier2 = False
    elif args.tier == "line":
        cfg.enabled_tier1 = False

    print(f"    -> cfg: routine_thr={cfg.threshold_routine} line_thr={cfg.threshold_line} "
          f"tier1={cfg.enabled_tier1} tier2={cfg.enabled_tier2} ignore_exp={cfg.ignore_expired}")

    # Load v2 dual indices
    pricing_by_routine, pricing_by_line = load_pricing_v2(wb)
    quotes = list(iter_pending_quotes(wb))
    print(f"    -> quotes to inspect: {len(quotes)}")

    # v2 detection
    alerts = compute_alerts_v2(quotes, pricing_by_routine, pricing_by_line, cfg)
    t1 = [a for a in alerts if getattr(a, "tier", "") == "ROUTINE"]
    t2 = [a for a in alerts if getattr(a, "tier", "") == "LINE"]
    drops = [a for a in alerts if a.kind == "DROP"]
    print(f"    -> alerts: {len(alerts)} "
          f"(Tier1-ROUTINE={len(t1)} Tier2-LINE={len(t2)} DROP={len(drops)})")

    # Write Price_Watch sheet (write_price_watch_sheet handles 0-alert case — Defect 4)
    stamp_quotes_sheet(wb, alerts) if alerts else None
    write_price_watch_sheet(wb, alerts)
    if alerts:
        stamped_aj = stamp_active_jobs(wb, alerts)
        print(f"    -> stamped {stamped_aj} Active Jobs row(s)")

    # ── Target Watch scan (S1v2 #4) ──
    print("[+] Target Watch scan...")
    try:
        target_matches = scan_target_matches(
            wb, pricing_by_routine, pricing_by_line, default_markup=200.0
        )
        # Append TARGET MATCHES block to Price_Watch sheet
        if "Price_Watch" in wb.sheetnames:
            write_target_matches_section(wb["Price_Watch"], target_matches)
        # Send Telegram alerts for new matches
        if target_matches:
            print(f"    -> sending {len(target_matches)} Telegram alert(s)...")
            _send_target_match_alerts(target_matches)
    except Exception as exc:
        print(f"    [WARN] Target Watch scan failed (non-fatal): {exc}")

    result = save_preserving_ribbon(wb, args.erp)
    wb.close()
    print(f"[OK] ERP saved: {args.erp}  (ribbon: {result})")

    # Print top 5 alerts
    if alerts:
        print("\nTop alerts:")
        for a in alerts[:5]:
            arrow = "↓" if a.delta < 0 else "↑"
            tier_tag = f"[{a.tier}]" if getattr(a, "tier", "") else ""
            carrier_tag = (f"→{a.carrier_new}" if getattr(a, "carrier_new", "") and a.carrier_new != a.carrier
                           else "")
            print(f"  [{a.priority}]{tier_tag} {a.quote_id} {a.customer[:16]:16s} "
                  f"{a.route:12s} {a.carrier:6s}{carrier_tag} {a.cont_type:5s} "
                  f"{arrow}${abs(a.delta):>6,.0f}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
