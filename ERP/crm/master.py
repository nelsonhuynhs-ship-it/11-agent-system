"""
CRM Master Creator - Customer Segmentation & Relationship Tracking
Creates CRM_Master.xlsx with:
- Customers (Coloader, Shipper, Cnee)
- Shipper_Cnee_Relationships
- Quotes
- Customer_Preferences
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

print("=" * 60)
print("Creating CRM_Master.xlsx...")
print("=" * 60)

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============================================================
# SHEET 1: CUSTOMERS
# ============================================================
print("\n[1/4] Creating Customers sheet...")

ws_customers = wb.create_sheet('Customers', 0)

# Headers
customers_headers = [
    'Customer_ID', 'Customer_Type', 'Company_Name', 'Contact_Person', 
    'Email', 'Phone', 'Country', 'City', 'Payment_Terms', 
    'Credit_Limit', 'Status', 'Notes', 'Created_Date'
]

ws_customers.append(customers_headers)

# Format headers
for col_num, header in enumerate(customers_headers, 1):
    cell = ws_customers.cell(1, col_num)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# Set column widths
widths = [12, 15, 30, 20, 25, 18, 12, 15, 15, 15, 12, 30, 15]
for idx, width in enumerate(widths, 1):
    ws_customers.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

# Data validations
dv_type = DataValidation(type='list', formula1='"Coloader,Shipper,Cnee"', allow_blank=False)
dv_type.error = 'Invalid customer type'
dv_type.errorTitle = 'Invalid Entry'
ws_customers.add_data_validation(dv_type)
dv_type.add('B2:B1000')

dv_status = DataValidation(type='list', formula1='"Active,Inactive"', allow_blank=False)
ws_customers.add_data_validation(dv_status)
dv_status.add('K2:K1000')

dv_payment = DataValidation(type='list', formula1='"7 days,15 days,30 days,45 days,60 days"', allow_blank=False)
ws_customers.add_data_validation(dv_payment)
dv_payment.add('I2:I1000')

# Sample data
sample_customers = [
    ['C001', 'Coloader', 'ABC Logistics', 'John Smith', 'john@abc.com', '+1-555-0101', 
     'USA', 'Los Angeles', '30 days', 50000, 'Active', 'Small forwarder, price sensitive', 
     datetime.now().strftime('%Y-%m-%d')],
    
    ['C002', 'Shipper', 'VN Furniture Co', 'Nguyen Van A', 'a.nguyen@vnfurniture.vn', '+84-28-1234567', 
     'Vietnam', 'Ho Chi Minh', '15 days', 100000, 'Active', 'Furniture manufacturer, FOB terms', 
     datetime.now().strftime('%Y-%m-%d')],
    
    ['C003', 'Cnee', 'USA Import Corp', 'Mike Johnson', 'mike@usaimport.com', '+1-555-0202', 
     'USA', 'New York', '45 days', 200000, 'Active', 'Large importer, buys from multiple VN suppliers', 
     datetime.now().strftime('%Y-%m-%d')],
    
    ['C004', 'Shipper', 'Thai Textile Ltd', 'Somchai Lee', 'somchai@thaitextile.th', '+66-2-9876543', 
     'Thailand', 'Bangkok', '30 days', 80000, 'Active', 'Textile manufacturer', 
     datetime.now().strftime('%Y-%m-%d')],
    
    ['C005', 'Cnee', 'Canada Trading Inc', 'Sarah Brown', 'sarah@canadatrading.ca', '+1-416-5550303', 
     'Canada', 'Toronto', '30 days', 150000, 'Active', 'Imports furniture and textiles', 
     datetime.now().strftime('%Y-%m-%d')],
    
    ['C006', 'Shipper', 'MY Electronics Sdn Bhd', 'Ahmad Hassan', 'ahmad@myelectronics.my', '+60-3-12345678', 
     'Malaysia', 'Kuala Lumpur', '30 days', 120000, 'Active', 'Electronics manufacturer', 
     datetime.now().strftime('%Y-%m-%d')],
    
    ['C007', 'Coloader', 'XYZ Freight Services', 'Emily Chen', 'emily@xyzfreight.com', '+1-555-0303', 
     'USA', 'Seattle', '15 days', 30000, 'Active', 'Small coloader, needs competitive rates', 
     datetime.now().strftime('%Y-%m-%d')],
]

for row in sample_customers:
    ws_customers.append(row)

print(f"   -> Added {len(sample_customers)} sample customers")

# ============================================================
# SHEET 2: SHIPPER_CNEE_RELATIONSHIPS
# ============================================================
print("\n[2/4] Creating Shipper_Cnee_Relationships sheet...")

ws_relationships = wb.create_sheet('Shipper_Cnee_Relationships', 1)

# Headers
rel_headers = [
    'Relationship_ID', 'Shipper_ID', 'Shipper_Name', 'Cnee_ID', 'Cnee_Name', 
    'Commodity', 'Typical_Route', 'Frequency', 'Last_Shipment', 'Status', 'Notes'
]

ws_relationships.append(rel_headers)

# Format headers
for col_num, header in enumerate(rel_headers, 1):
    cell = ws_relationships.cell(1, col_num)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# Set column widths
rel_widths = [15, 12, 30, 12, 30, 20, 20, 12, 15, 12, 30]
for idx, width in enumerate(rel_widths, 1):
    ws_relationships.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

# Sample relationships
sample_relationships = [
    ['R001', 'C002', 'VN Furniture Co', 'C003', 'USA Import Corp', 
     'Furniture', 'HCM-USNYC', 'Monthly', '2026-01-15', 'Active', 
     'Main buyer, 5-10 containers/month'],
    
    ['R002', 'C002', 'VN Furniture Co', 'C005', 'Canada Trading Inc', 
     'Furniture', 'HCM-CATOR', 'Quarterly', '2025-12-20', 'Active', 
     'Seasonal buyer'],
    
    ['R003', 'C004', 'Thai Textile Ltd', 'C003', 'USA Import Corp', 
     'Textile', 'BKK-USNYC', 'Bi-monthly', '2026-01-10', 'Active', 
     'Regular buyer'],
    
    ['R004', 'C004', 'Thai Textile Ltd', 'C005', 'Canada Trading Inc', 
     'Textile', 'BKK-CATOR', 'Monthly', '2026-01-18', 'Active', 
     'Growing relationship'],
    
    ['R005', 'C006', 'MY Electronics Sdn Bhd', 'C003', 'USA Import Corp', 
     'Electronics', 'KUL-USLAX', 'Monthly', '2026-01-12', 'Active', 
     'New relationship, potential to grow'],
]

for row in sample_relationships:
    ws_relationships.append(row)

print(f"   -> Added {len(sample_relationships)} sample relationships")

# ============================================================
# SHEET 3: QUOTES
# ============================================================
print("\n[3/4] Creating Quotes sheet...")

ws_quotes = wb.create_sheet('Quotes', 2)

# Headers
quotes_headers = [
    'Quote_ID', 'Quote_Date', 'Customer_ID', 'Customer_Type', 
    'POL', 'POD', 'Place', 'Carrier', 'Commodity', 'HS_Code', 
    'Container_Type', 'Quantity', 'Ocean_Freight', 'Total_Price', 
    'Valid_Until', 'Pipeline_Status', 'Win_Lose_Reason', 'Notes', 
    'Sales_Person', 'Created_Date'
]

ws_quotes.append(quotes_headers)

# Format headers
for col_num, header in enumerate(quotes_headers, 1):
    cell = ws_quotes.cell(1, col_num)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='D84315', end_color='D84315', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# Set column widths
quote_widths = [12, 12, 12, 15, 10, 10, 20, 10, 20, 10, 15, 10, 15, 15, 12, 15, 20, 30, 15, 15]
for idx, width in enumerate(quote_widths, 1):
    ws_quotes.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

# Data validation for Pipeline Status
dv_pipeline = DataValidation(type='list', formula1='"Inquiry,Quoted,Follow,Won,Lost"', allow_blank=False)
ws_quotes.add_data_validation(dv_pipeline)
dv_pipeline.add('P2:P1000')

# Sample quotes
sample_quotes = [
    ['Q2026001', '2026-01-20', 'C002', 'Shipper', 'HCM', 'USNYC', 'New York', 'ONE', 
     'Furniture', '9403', '40HQ', 2, 1500, 1750, '2026-01-27', 'Quoted', '', 
     'Customer requested quote for Feb shipment', 'Nelson', '2026-01-20'],
    
    ['Q2026002', '2026-01-21', 'C003', 'Cnee', 'HCM', 'USLAX', 'Los Angeles', 'YML', 
     'Furniture', '9403', '40HQ', 5, 1400, 1650, '2026-01-28', 'Follow', '', 
     'Following up, customer comparing with other forwarders', 'Nelson', '2026-01-21'],
    
    ['Q2026003', '2026-01-22', 'C001', 'Coloader', 'HCM', 'USNYC', 'New York', 'HPL', 
     'General Cargo', '', '40GP', 1, 1200, 1400, '2026-01-29', 'Won', 'Best price', 
     'Closed deal, ETD 2026-02-05', 'Nelson', '2026-01-22'],
]

for row in sample_quotes:
    ws_quotes.append(row)

print(f"   -> Added {len(sample_quotes)} sample quotes")

# ============================================================
# SHEET 4: CUSTOMER_PREFERENCES
# ============================================================
print("\n[4/4] Creating Customer_Preferences sheet...")

ws_prefs = wb.create_sheet('Customer_Preferences', 3)

# Headers
prefs_headers = [
    'Customer_ID', 'Customer_Type', 'Preferred_POL', 'Preferred_POD', 
    'Excluded_Carriers', 'Typical_Commodity', 'Avg_Containers_Per_Month', 
    'Typical_ETD_Window', 'Win_Rate', 'Avg_Price_Paid', 
    'Last_Quote_Date', 'Last_Shipment_Date', 'Notes'
]

ws_prefs.append(prefs_headers)

# Format headers
for col_num, header in enumerate(prefs_headers, 1):
    cell = ws_prefs.cell(1, col_num)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='6A1B9A', end_color='6A1B9A', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# Set column widths
for col in range(1, len(prefs_headers) + 1):
    ws_prefs.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

# Sample preferences
sample_prefs = [
    ['C001', 'Coloader', 'HCM', 'USNYC,USLAX', 'None', 'General Cargo', 2, 
     '1-15 of month', '75%', 1250, '2026-01-22', '2026-01-15', 'Price sensitive'],
    
    ['C002', 'Shipper', 'HCM', 'USNYC,USLAX', 'HPL', 'Furniture', 8, 
     '20-30 of month', '60%', 1450, '2026-01-20', '2026-01-15', 'FOB terms, needs reliable service'],
    
    ['C003', 'Cnee', 'HCM,BKK', 'USNYC', 'None', 'Furniture,Textile', 15, 
     'Any', '70%', 1400, '2026-01-21', '2026-01-18', 'Large volume, multiple suppliers'],
]

for row in sample_prefs:
    ws_prefs.append(row)

print(f"   -> Added {len(sample_prefs)} sample preferences")

# ============================================================
# SAVE
# ============================================================
output_file = 'CRM_Master.xlsx'
wb.save(output_file)

print("\n" + "=" * 60)
print(f"SUCCESS! Created {output_file}")
print("=" * 60)
print("\nSheets created:")
print("  1. Customers - Customer database with types (Coloader/Shipper/Cnee)")
print("  2. Shipper_Cnee_Relationships - Track which Shipper sells to which Cnee")
print("  3. Quotes - Quote tracking with pipeline stages")
print("  4. Customer_Preferences - Auto-learned customer preferences")
print("\nKey Features:")
print("  - Customer segmentation (Coloader, Shipper, Cnee)")
print("  - Relationship mapping (Shipper <-> Cnee)")
print("  - Data validation dropdowns")
print("  - Sample data for testing")
print("\nNext Steps:")
print("  1. Open CRM_Master.xlsx")
print("  2. Review sample data")
print("  3. Add your real customers")
print("  4. Link to MasterFullPricing.xlsx for pricing")
