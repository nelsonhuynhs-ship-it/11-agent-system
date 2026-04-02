"""
CRM Visual Dashboard - Single Sheet Layout
Creates a beautiful single-sheet dashboard matching the reference design
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta

print("=" * 70)
print("CRM VISUAL DASHBOARD - Single Sheet Layout")
print("=" * 70)

# Load data
crm_file = 'CRM_Master.xlsx'
print(f"\nLoading {crm_file}...")

df_customers = pd.read_excel(crm_file, sheet_name='Customers')
df_quotes = pd.read_excel(crm_file, sheet_name='Quotes')
df_relationships = pd.read_excel(crm_file, sheet_name='Shipper_Cnee_Relationships')

# Prepare data
df_quotes['Quote_Date'] = pd.to_datetime(df_quotes['Quote_Date'])
df_quotes['Month'] = df_quotes['Quote_Date'].dt.to_period('M').astype(str)
df_relationships['Last_Shipment'] = pd.to_datetime(df_relationships['Last_Shipment'])

today = datetime.now()

print(f"  -> {len(df_customers)} customers")
print(f"  -> {len(df_quotes)} quotes")
print(f"  -> {len(df_relationships)} relationships")

# ============================================================
# CREATE SINGLE SHEET DASHBOARD
# ============================================================
print("\n[1/4] Creating dashboard workbook...")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Dashboard'

# Set column widths
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 8
ws.column_dimensions['G'].width = 8
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 3
ws.column_dimensions['J'].width = 20
ws.column_dimensions['K'].width = 15
ws.column_dimensions['L'].width = 12
ws.column_dimensions['M'].width = 12

# ============================================================
# HEADER
# ============================================================
print("\n[2/4] Creating header...")

ws.merge_cells('A1:M2')
ws['A1'] = 'CRM DASHBOARD - MONTHLY PERFORMANCE & ALERTS'
ws['A1'].font = Font(size=20, bold=True, color='FFFFFF')
ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 25
ws.row_dimensions[2].height = 25

# Subtitle
ws.merge_cells('A3:M3')
ws['A3'] = f'Generated: {today.strftime("%Y-%m-%d %H:%M")} | Use filter to view specific month'
ws['A3'].font = Font(size=10, italic=True, color='666666')
ws['A3'].alignment = Alignment(horizontal='center')
ws.row_dimensions[3].height = 20

# ============================================================
# MONTH FILTER
# ============================================================
print("\n[3/4] Adding month filter...")

ws['B5'] = 'Filter by Month:'
ws['B5'].font = Font(size=11, bold=True)
ws['B5'].alignment = Alignment(horizontal='right')

# Get unique months
unique_months = sorted(df_quotes['Month'].unique(), reverse=True)
month_list = ','.join(['All'] + unique_months)

ws['C5'] = 'All'
ws['C5'].font = Font(size=11)
ws['C5'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
ws['C5'].border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Add dropdown
dv_month = DataValidation(type='list', formula1=f'"{month_list}"', allow_blank=False)
ws.add_data_validation(dv_month)
dv_month.add('C5')

# ============================================================
# LEFT PANEL: MONTHLY PERFORMANCE
# ============================================================
print("\n[4/4] Building performance table...")

# Section header
ws.merge_cells('B7:H7')
ws['B7'] = 'MONTHLY PERFORMANCE'
ws['B7'].font = Font(size=12, bold=True, color='1F4E78')
ws['B7'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
ws['B7'].alignment = Alignment(horizontal='left')
ws.row_dimensions[7].height = 25

# Table headers
headers = ['Month', 'Customer', 'Type', 'Total Quotes', 'Won', 'Lost', 'Win Rate']
header_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H']

for col, header in zip(header_cols, headers):
    cell = ws[f'{col}8']
    cell.value = header
    cell.font = Font(size=10, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

ws.row_dimensions[8].height = 20

# Prepare monthly data
monthly_data = df_quotes.groupby(['Customer_ID', 'Month']).agg({
    'Quote_ID': 'count',
    'Pipeline_Status': lambda x: (x == 'Won').sum()
}).reset_index()

monthly_data.columns = ['Customer_ID', 'Month', 'Total_Quotes', 'Won_Quotes']
monthly_data['Lost_Quotes'] = monthly_data['Total_Quotes'] - monthly_data['Won_Quotes']
monthly_data['Win_Rate'] = (monthly_data['Won_Quotes'] / monthly_data['Total_Quotes'] * 100).round(1)

monthly_data = monthly_data.merge(
    df_customers[['Customer_ID', 'Company_Name', 'Customer_Type']], 
    on='Customer_ID', 
    how='left'
)

# Sort by month descending
monthly_data = monthly_data.sort_values(['Month', 'Company_Name'], ascending=[False, True])

# Add data rows
row_num = 9
for _, row_data in monthly_data.iterrows():
    ws[f'B{row_num}'] = row_data['Month']
    ws[f'C{row_num}'] = row_data['Company_Name']
    ws[f'D{row_num}'] = row_data['Customer_Type']
    ws[f'E{row_num}'] = row_data['Total_Quotes']
    ws[f'F{row_num}'] = row_data['Won_Quotes']
    ws[f'G{row_num}'] = row_data['Lost_Quotes']
    ws[f'H{row_num}'] = f"{row_data['Win_Rate']}%"
    
    # Format cells
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        cell = ws[f'{col}{row_num}']
        cell.border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        cell.alignment = Alignment(horizontal='center' if col in ['B', 'D', 'E', 'F', 'G', 'H'] else 'left')
    
    # Color code win rate
    win_rate = row_data['Win_Rate']
    if win_rate >= 70:
        ws[f'H{row_num}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        ws[f'H{row_num}'].font = Font(bold=True, color='006100')
    elif win_rate >= 50:
        ws[f'H{row_num}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        ws[f'H{row_num}'].font = Font(bold=True, color='9C6500')
    else:
        ws[f'H{row_num}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        ws[f'H{row_num}'].font = Font(bold=True, color='9C0006')
    
    row_num += 1

# ============================================================
# RIGHT PANEL: INACTIVE ALERTS
# ============================================================

# Section header
ws.merge_cells('J7:M7')
ws['J7'] = '⚠ LOSS RISK - No Shipment > 30 Days'
ws['J7'].font = Font(size=11, bold=True, color='FFFFFF')
ws['J7'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
ws['J7'].alignment = Alignment(horizontal='center')

# Table headers
alert_headers = ['Customer', 'Last Shipment', 'Days Since', 'Risk Level']
alert_cols = ['J', 'K', 'L', 'M']

for col, header in zip(alert_cols, alert_headers):
    cell = ws[f'{col}8']
    cell.value = header
    cell.font = Font(size=10, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='D84315', end_color='D84315', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

# Find inactive customers
inactive_alerts = []

for _, customer in df_customers.iterrows():
    customer_id = customer['Customer_ID']
    customer_type = customer['Customer_Type']
    
    if customer_type == 'Shipper':
        rels = df_relationships[df_relationships['Shipper_ID'] == customer_id]
    elif customer_type == 'Cnee':
        rels = df_relationships[df_relationships['Cnee_ID'] == customer_id]
    else:
        continue
    
    if not rels.empty:
        last_shipment = rels['Last_Shipment'].max()
        days_since = (today - last_shipment).days
        
        if days_since > 30:
            inactive_alerts.append({
                'Company_Name': customer['Company_Name'],
                'Last_Shipment': last_shipment.strftime('%Y-%m-%d'),
                'Days_Since': days_since,
                'Risk': 'HIGH' if days_since > 60 else 'MEDIUM'
            })

# Add alert data
alert_row = 9
for alert in inactive_alerts:
    ws[f'J{alert_row}'] = alert['Company_Name']
    ws[f'K{alert_row}'] = alert['Last_Shipment']
    ws[f'L{alert_row}'] = f"{alert['Days_Since']} days"
    ws[f'M{alert_row}'] = alert['Risk']
    
    # Format cells
    for col in ['J', 'K', 'L', 'M']:
        cell = ws[f'{col}{alert_row}']
        cell.border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        cell.alignment = Alignment(horizontal='center' if col in ['K', 'L', 'M'] else 'left')
    
    # Color code risk
    if alert['Risk'] == 'HIGH':
        ws[f'M{alert_row}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        ws[f'M{alert_row}'].font = Font(bold=True, color='9C0006')
    else:
        ws[f'M{alert_row}'].fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        ws[f'M{alert_row}'].font = Font(bold=True, color='9C6500')
    
    alert_row += 1

# ============================================================
# BOTTOM: CUSTOMER SUMMARY CARDS
# ============================================================

summary_row = max(row_num, alert_row) + 2

# Section header
ws.merge_cells(f'B{summary_row}:M{summary_row}')
ws[f'B{summary_row}'] = 'CUSTOMER SUMMARY'
ws[f'B{summary_row}'].font = Font(size=12, bold=True, color='1F4E78')
ws[f'B{summary_row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
ws[f'B{summary_row}'].alignment = Alignment(horizontal='left')
ws.row_dimensions[summary_row].height = 25

card_row = summary_row + 1

# Card 1: Total Customers
ws.merge_cells(f'B{card_row}:D{card_row+2}')
ws[f'B{card_row}'] = f'Total Customers:\n{len(df_customers)}'
ws[f'B{card_row}'].font = Font(size=14, bold=True, color='FFFFFF')
ws[f'B{card_row}'].fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
ws[f'B{card_row}'].alignment = Alignment(horizontal='center', vertical='center')
ws[f'B{card_row}'].border = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)

# Card 2: Active Quotes
ws.merge_cells(f'E{card_row}:G{card_row+2}')
ws[f'E{card_row}'] = f'Active Quotes:\n{len(df_quotes)}'
ws[f'E{card_row}'].font = Font(size=14, bold=True, color='FFFFFF')
ws[f'E{card_row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
ws[f'E{card_row}'].alignment = Alignment(horizontal='center', vertical='center')
ws[f'E{card_row}'].border = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)

# Card 3: Inactive Alerts
ws.merge_cells(f'H{card_row}:J{card_row+2}')
ws[f'H{card_row}'] = f'Inactive Alerts:\n{len(inactive_alerts)}'
ws[f'H{card_row}'].font = Font(size=14, bold=True, color='FFFFFF')
ws[f'H{card_row}'].fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
ws[f'H{card_row}'].alignment = Alignment(horizontal='center', vertical='center')
ws[f'H{card_row}'].border = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)

ws.row_dimensions[card_row].height = 20
ws.row_dimensions[card_row+1].height = 20
ws.row_dimensions[card_row+2].height = 20

# ============================================================
# SAVE
# ============================================================
output_file = 'CRM_Dashboard_Visual.xlsx'
wb.save(output_file)

print("\n" + "=" * 70)
print(f"SUCCESS! Created {output_file}")
print("=" * 70)
print("\nFeatures:")
print("  ✓ Single sheet layout")
print("  ✓ Month filter dropdown (All, 2026-01, 2025-12, 2025-11)")
print("  ✓ Monthly performance table with color-coded win rates")
print("  ✓ Inactive alerts panel")
print("  ✓ Summary cards at bottom")
print("\nColor Coding:")
print("  - Win Rate: Green (≥70%), Yellow (50-69%), Red (<50%)")
print("  - Risk: Red (HIGH >60 days), Yellow (MEDIUM 30-60 days)")
print("\nNext: Open file and use month filter dropdown in cell C5")
