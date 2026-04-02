"""
CRM Dashboard Creator - Monthly Tracking & Loss Alerts
Creates visual dashboard with:
- Monthly quote tracking per customer
- Win rate analysis
- Inactive customer alerts (no shipment in 30 days)
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import calendar

print("=" * 70)
print("CRM DASHBOARD CREATOR - Monthly Tracking & Alerts")
print("=" * 70)

# Load CRM data
crm_file = 'CRM_Master.xlsx'
print(f"\nLoading {crm_file}...")

df_customers = pd.read_excel(crm_file, sheet_name='Customers')
df_quotes = pd.read_excel(crm_file, sheet_name='Quotes')
df_relationships = pd.read_excel(crm_file, sheet_name='Shipper_Cnee_Relationships')

print(f"  -> {len(df_customers)} customers")
print(f"  -> {len(df_quotes)} quotes")
print(f"  -> {len(df_relationships)} relationships")

# ============================================================
# PREPARE DATA
# ============================================================
print("\n[1/4] Preparing monthly data...")

# Convert dates
df_quotes['Quote_Date'] = pd.to_datetime(df_quotes['Quote_Date'])
df_quotes['Month'] = df_quotes['Quote_Date'].dt.to_period('M').astype(str)

df_relationships['Last_Shipment'] = pd.to_datetime(df_relationships['Last_Shipment'])

# Current date
today = datetime.now()
current_month = today.strftime('%Y-%m')
last_30_days = today - timedelta(days=30)

# ============================================================
# MONTHLY QUOTE TRACKING
# ============================================================
print("\n[2/4] Building monthly quote tracking...")

# Group by Customer and Month
monthly_quotes = df_quotes.groupby(['Customer_ID', 'Month']).agg({
    'Quote_ID': 'count',
    'Pipeline_Status': lambda x: (x == 'Won').sum()
}).reset_index()

monthly_quotes.columns = ['Customer_ID', 'Month', 'Total_Quotes', 'Won_Quotes']
monthly_quotes['Lost_Quotes'] = monthly_quotes['Total_Quotes'] - monthly_quotes['Won_Quotes']
monthly_quotes['Win_Rate'] = (monthly_quotes['Won_Quotes'] / monthly_quotes['Total_Quotes'] * 100).round(1)

# Merge with customer info
monthly_quotes = monthly_quotes.merge(
    df_customers[['Customer_ID', 'Company_Name', 'Customer_Type']], 
    on='Customer_ID', 
    how='left'
)

# ============================================================
# INACTIVE CUSTOMER ALERTS
# ============================================================
print("\n[3/4] Identifying inactive customers...")

# Find customers with no recent shipments
inactive_alerts = []

for _, customer in df_customers.iterrows():
    customer_id = customer['Customer_ID']
    customer_type = customer['Customer_Type']
    
    # Check relationships for Shippers
    if customer_type == 'Shipper':
        shipper_rels = df_relationships[df_relationships['Shipper_ID'] == customer_id]
        if not shipper_rels.empty:
            last_shipment = shipper_rels['Last_Shipment'].max()
            days_since = (today - last_shipment).days
            
            if days_since > 30:
                inactive_alerts.append({
                    'Customer_ID': customer_id,
                    'Company_Name': customer['Company_Name'],
                    'Customer_Type': customer_type,
                    'Last_Shipment': last_shipment.strftime('%Y-%m-%d'),
                    'Days_Since': days_since,
                    'Alert': 'NO SHIPMENT > 30 DAYS',
                    'Risk': 'HIGH' if days_since > 60 else 'MEDIUM'
                })
    
    # Check relationships for Cnees
    elif customer_type == 'Cnee':
        cnee_rels = df_relationships[df_relationships['Cnee_ID'] == customer_id]
        if not cnee_rels.empty:
            last_shipment = cnee_rels['Last_Shipment'].max()
            days_since = (today - last_shipment).days
            
            if days_since > 30:
                inactive_alerts.append({
                    'Customer_ID': customer_id,
                    'Company_Name': customer['Company_Name'],
                    'Customer_Type': customer_type,
                    'Last_Shipment': last_shipment.strftime('%Y-%m-%d'),
                    'Days_Since': days_since,
                    'Alert': 'NO SHIPMENT > 30 DAYS',
                    'Risk': 'HIGH' if days_since > 60 else 'MEDIUM'
                })

df_inactive = pd.DataFrame(inactive_alerts)

print(f"  -> Found {len(df_inactive)} inactive customers")

# ============================================================
# CREATE DASHBOARD EXCEL
# ============================================================
print("\n[4/4] Creating dashboard Excel...")

wb = openpyxl.Workbook()
wb.remove(wb.active)

# --- SHEET 1: MONTHLY PERFORMANCE ---
ws_monthly = wb.create_sheet('Monthly_Performance', 0)

# Title
ws_monthly['A1'] = 'MONTHLY CUSTOMER PERFORMANCE DASHBOARD'
ws_monthly['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws_monthly['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
ws_monthly['A1'].alignment = Alignment(horizontal='center')
ws_monthly.merge_cells('A1:H1')
ws_monthly.row_dimensions[1].height = 30

# Subtitle
ws_monthly['A2'] = f'Generated: {today.strftime("%Y-%m-%d %H:%M")}'
ws_monthly['A2'].font = Font(size=10, italic=True)
ws_monthly['A2'].alignment = Alignment(horizontal='center')
ws_monthly.merge_cells('A2:H2')

# Headers
headers = ['Month', 'Customer_ID', 'Company_Name', 'Customer_Type', 
           'Total_Quotes', 'Won', 'Lost', 'Win_Rate (%)']

ws_monthly.append([''])  # Blank row
ws_monthly.append(headers)

# Format headers
for col_num, header in enumerate(headers, 1):
    cell = ws_monthly.cell(4, col_num)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# Add data
for _, row in monthly_quotes.iterrows():
    ws_monthly.append([
        row['Month'],
        row['Customer_ID'],
        row['Company_Name'],
        row['Customer_Type'],
        row['Total_Quotes'],
        row['Won_Quotes'],
        row['Lost_Quotes'],
        row['Win_Rate']
    ])

# Format columns
ws_monthly.column_dimensions['A'].width = 12
ws_monthly.column_dimensions['B'].width = 12
ws_monthly.column_dimensions['C'].width = 30
ws_monthly.column_dimensions['D'].width = 15
ws_monthly.column_dimensions['E'].width = 12
ws_monthly.column_dimensions['F'].width = 8
ws_monthly.column_dimensions['G'].width = 8
ws_monthly.column_dimensions['H'].width = 12

# Conditional formatting for win rate
for row in range(5, ws_monthly.max_row + 1):
    win_rate_cell = ws_monthly.cell(row, 8)
    win_rate = win_rate_cell.value
    
    if win_rate is not None and isinstance(win_rate, (int, float)):
        if win_rate >= 70:
            win_rate_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
        elif win_rate >= 50:
            win_rate_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Yellow
        else:
            win_rate_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red

# --- SHEET 2: INACTIVE ALERTS ---
ws_alerts = wb.create_sheet('Inactive_Alerts', 1)

# Title
ws_alerts['A1'] = 'INACTIVE CUSTOMER ALERTS - LOSS RISK'
ws_alerts['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws_alerts['A1'].fill = PatternFill(start_color='D84315', end_color='D84315', fill_type='solid')
ws_alerts['A1'].alignment = Alignment(horizontal='center')
ws_alerts.merge_cells('A1:G1')
ws_alerts.row_dimensions[1].height = 30

# Subtitle
ws_alerts['A2'] = f'Customers with no shipment in last 30 days | Generated: {today.strftime("%Y-%m-%d")}'
ws_alerts['A2'].font = Font(size=10, italic=True)
ws_alerts['A2'].alignment = Alignment(horizontal='center')
ws_alerts.merge_cells('A2:G2')

# Headers
alert_headers = ['Customer_ID', 'Company_Name', 'Customer_Type', 
                 'Last_Shipment', 'Days_Since', 'Alert', 'Risk_Level']

ws_alerts.append([''])  # Blank row
ws_alerts.append(alert_headers)

# Format headers
for col_num, header in enumerate(alert_headers, 1):
    cell = ws_alerts.cell(4, col_num)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='D84315', end_color='D84315', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# Add data
if not df_inactive.empty:
    for _, row in df_inactive.iterrows():
        ws_alerts.append([
            row['Customer_ID'],
            row['Company_Name'],
            row['Customer_Type'],
            row['Last_Shipment'],
            row['Days_Since'],
            row['Alert'],
            row['Risk']
        ])
else:
    ws_alerts.append(['No inactive customers found', '', '', '', '', '', ''])

# Format columns
ws_alerts.column_dimensions['A'].width = 12
ws_alerts.column_dimensions['B'].width = 30
ws_alerts.column_dimensions['C'].width = 15
ws_alerts.column_dimensions['D'].width = 15
ws_alerts.column_dimensions['E'].width = 12
ws_alerts.column_dimensions['F'].width = 25
ws_alerts.column_dimensions['G'].width = 12

# Conditional formatting for risk level
for row in range(5, ws_alerts.max_row + 1):
    risk_cell = ws_alerts.cell(row, 7)
    risk = risk_cell.value
    
    if risk == 'HIGH':
        risk_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red
        risk_cell.font = Font(bold=True, color='9C0006')
    elif risk == 'MEDIUM':
        risk_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Yellow
        risk_cell.font = Font(bold=True, color='9C6500')

# --- SHEET 3: CUSTOMER SUMMARY ---
ws_summary = wb.create_sheet('Customer_Summary', 2)

# Title
ws_summary['A1'] = 'CUSTOMER SUMMARY - ALL TIME'
ws_summary['A1'].font = Font(size=16, bold=True, color='FFFFFF')
ws_summary['A1'].fill = PatternFill(start_color='6A1B9A', end_color='6A1B9A', fill_type='solid')
ws_summary['A1'].alignment = Alignment(horizontal='center')
ws_summary.merge_cells('A1:I1')
ws_summary.row_dimensions[1].height = 30

# Calculate customer summary
customer_summary = []

for _, customer in df_customers.iterrows():
    customer_id = customer['Customer_ID']
    
    # Quote stats
    customer_quotes = df_quotes[df_quotes['Customer_ID'] == customer_id]
    total_quotes = len(customer_quotes)
    won_quotes = len(customer_quotes[customer_quotes['Pipeline_Status'] == 'Won'])
    win_rate = (won_quotes / total_quotes * 100) if total_quotes > 0 else 0
    
    # Last activity
    last_quote = customer_quotes['Quote_Date'].max() if not customer_quotes.empty else None
    
    # Relationship count
    if customer['Customer_Type'] == 'Shipper':
        rel_count = len(df_relationships[df_relationships['Shipper_ID'] == customer_id])
        rel_type = 'Cnees'
    elif customer['Customer_Type'] == 'Cnee':
        rel_count = len(df_relationships[df_relationships['Cnee_ID'] == customer_id])
        rel_type = 'Shippers'
    else:
        rel_count = 0
        rel_type = 'N/A'
    
    customer_summary.append({
        'Customer_ID': customer_id,
        'Company_Name': customer['Company_Name'],
        'Customer_Type': customer['Customer_Type'],
        'Total_Quotes': total_quotes,
        'Won': won_quotes,
        'Win_Rate': round(win_rate, 1),
        'Last_Quote': last_quote.strftime('%Y-%m-%d') if pd.notna(last_quote) else 'Never',
        'Relationships': rel_count,
        'Rel_Type': rel_type
    })

df_summary = pd.DataFrame(customer_summary)

# Headers
summary_headers = ['Customer_ID', 'Company_Name', 'Type', 'Total_Quotes', 
                   'Won', 'Win_Rate (%)', 'Last_Quote', 'Relationships', 'Rel_Type']

ws_summary.append([''])  # Blank row
ws_summary.append(summary_headers)

# Format headers
for col_num, header in enumerate(summary_headers, 1):
    cell = ws_summary.cell(3, col_num)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='6A1B9A', end_color='6A1B9A', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# Add data
for _, row in df_summary.iterrows():
    ws_summary.append([
        row['Customer_ID'],
        row['Company_Name'],
        row['Customer_Type'],
        row['Total_Quotes'],
        row['Won'],
        row['Win_Rate'],
        row['Last_Quote'],
        row['Relationships'],
        row['Rel_Type']
    ])

# Format columns
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
    ws_summary.column_dimensions[col].width = 15
ws_summary.column_dimensions['B'].width = 30

# ============================================================
# SAVE
# ============================================================
output_file = 'CRM_Dashboard.xlsx'
wb.save(output_file)

print("\n" + "=" * 70)
print(f"SUCCESS! Created {output_file}")
print("=" * 70)
print("\nDashboard sheets:")
print("  1. Monthly_Performance - Quote tracking by month per customer")
print("  2. Inactive_Alerts - Customers with no shipment > 30 days")
print("  3. Customer_Summary - Overall customer performance")
print("\nKey Metrics:")
print(f"  - Total customers tracked: {len(df_customers)}")
print(f"  - Customers with quotes: {df_quotes['Customer_ID'].nunique()}")
print(f"  - Inactive customers (>30 days): {len(df_inactive)}")
print("\nColor Coding:")
print("  - Win Rate: Green (>=70%), Yellow (50-69%), Red (<50%)")
print("  - Risk Level: Red (HIGH), Yellow (MEDIUM)")
