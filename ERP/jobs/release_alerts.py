"""
release_alerts.py — Feature 3 (Active Jobs v4): ETA Release Alert
==================================================================
When docsteam sends a release email (ask customer to confirm pickup near ETA),
Nelson records timestamp in col 33 RELEASE_EMAIL_SENT.
If >= 2 hours pass with RELEASE_CONFIRMED (col 34) still blank → URGENT alert:
Nelson must push the customer (otherwise: demurrage / detention).

Triggers + thresholds (configurable via CLI):
  - email_sent_at NOT NULL  AND  release_confirmed IS NULL
  - elapsed since email_sent_at >= --hours (default 2)
  - AND ETA within --eta-window-days (default 3) — ignore jobs not near ETA

Outputs:
  - Prints alert list to stdout with countdown remaining
  - Creates/refreshes "Release_Alerts" sheet in ERP_Master_v14.xlsm
  - Returns non-zero exit if any P1 (URGENT) alert present — VBA can poll

Usage:
    python ERP/jobs/release_alerts.py
    python ERP/jobs/release_alerts.py --hours 2 --eta-window-days 3
    python ERP/jobs/release_alerts.py --now "2026-04-14 15:00"   # test override
"""
from __future__ import annotations

import argparse
import os
import sys
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Final

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "core"))
from ribbon_guard import save_preserving_ribbon  # noqa: E402
from active_jobs_cols import COL, DATA_START  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]

DEFAULT_ERP_FILE: Final = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"

AJ_DATA_START = DATA_START


@dataclass
class ReleaseAlert:
    row: int
    crm_id: str
    hbl: str
    fast_id: str
    customer: str
    carrier: str
    routing: str
    eta: datetime | None
    email_sent_at: datetime
    elapsed_hours: float
    countdown_to_deadline: timedelta
    priority: str  # URGENT | WARN | INFO


def classify(elapsed_hours: float, deadline_hours: float) -> tuple[str, timedelta]:
    """Return (priority, countdown). Countdown is time remaining until breach."""
    remaining = timedelta(hours=deadline_hours - elapsed_hours)
    if elapsed_hours >= deadline_hours:
        return "URGENT", remaining  # negative remaining
    if elapsed_hours >= deadline_hours * 0.5:
        return "WARN", remaining
    return "INFO", remaining


def scan_alerts(
    erp_file: str,
    hours: float = 2.0,
    eta_window_days: int = 3,
    now: datetime | None = None,
) -> list[ReleaseAlert]:
    now = now or datetime.now()
    window_end = now + timedelta(days=eta_window_days)

    wb = openpyxl.load_workbook(erp_file, read_only=True, data_only=True, keep_vba=True)
    sheet = next((s for s in wb.sheetnames if "Active" in s), None)
    if not sheet:
        wb.close()
        return []
    ws = wb[sheet]

    alerts: list[ReleaseAlert] = []
    for r in range(AJ_DATA_START, ws.max_row + 1):
        crm = ws.cell(r, COL["CRM_ID"]).value
        if not crm:
            continue
        sent = ws.cell(r, COL["RELEASE_EMAIL_SENT"]).value
        confirmed = ws.cell(r, COL["RELEASE_CONFIRMED"]).value
        if not isinstance(sent, datetime):
            continue
        if confirmed is not None and confirmed != "":
            continue  # confirmed → no alert

        eta = ws.cell(r, COL["ETA"]).value
        # Only alert if ETA is within window (near arrival or past)
        if isinstance(eta, datetime) and eta > window_end:
            continue

        elapsed = (now - sent).total_seconds() / 3600.0
        if elapsed < 0:  # email sent in the future? skip
            continue

        priority, remaining = classify(elapsed, hours)

        alerts.append(ReleaseAlert(
            row=r,
            crm_id=str(crm),
            hbl=str(ws.cell(r, COL["HBL_NO"]).value or ""),
            fast_id=str(ws.cell(r, COL["FAST_ID"]).value or ""),
            customer=str(crm),
            carrier=str(ws.cell(r, COL["Carrier"]).value or ""),
            routing=str(ws.cell(r, COL["Routing"]).value or ""),
            eta=eta if isinstance(eta, datetime) else None,
            email_sent_at=sent,
            elapsed_hours=elapsed,
            countdown_to_deadline=remaining,
            priority=priority,
        ))

    wb.close()
    return alerts


THIN = Side(style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL_URGENT = PatternFill("solid", fgColor="FECACA")
FILL_WARN = PatternFill("solid", fgColor="FEF3C7")
FILL_INFO = PatternFill("solid", fgColor="DBEAFE")


def write_alerts_sheet(erp_file: str, alerts: list[ReleaseAlert]) -> None:
    try:
        with open(erp_file, "r+b"):
            pass
    except PermissionError:
        raise RuntimeError(f"ERP file is open in Excel: {erp_file}")

    wb = openpyxl.load_workbook(erp_file, keep_vba=True)
    if "Release_Alerts" in wb.sheetnames:
        del wb["Release_Alerts"]
    ws = wb.create_sheet("Release_Alerts")

    ws.merge_cells("A1:J1")
    ws["A1"] = f"ETA RELEASE ALERTS — {datetime.now():%d %b %Y %H:%M}"
    ws["A1"].font = Font(bold=True, size=14, color="991B1B")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    hdrs = ["Priority", "Customer", "HBL", "FAST ID", "Carrier", "Routing",
            "ETA", "Release Email Sent", "Elapsed (h)", "Countdown"]
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(3, i, h)
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Segoe UI")
        c.fill = PatternFill("solid", fgColor="991B1B")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    widths = [10, 18, 16, 14, 10, 16, 12, 18, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # sort by priority + shortest countdown
    prio = {"URGENT": 0, "WARN": 1, "INFO": 2}
    alerts_sorted = sorted(alerts, key=lambda a: (prio.get(a.priority, 9), a.countdown_to_deadline))

    for rr, a in enumerate(alerts_sorted, start=4):
        values = [
            a.priority, a.customer, a.hbl, a.fast_id, a.carrier, a.routing,
            a.eta, a.email_sent_at,
            round(a.elapsed_hours, 2),
            _format_countdown(a.countdown_to_deadline, a.priority),
        ]
        fill = FILL_URGENT if a.priority == "URGENT" else FILL_WARN if a.priority == "WARN" else FILL_INFO
        for i, v in enumerate(values, 1):
            cell = ws.cell(rr, i, v)
            cell.font = Font(size=10, name="Segoe UI",
                             bold=(a.priority == "URGENT"),
                             color="991B1B" if a.priority == "URGENT" else "000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
            cell.fill = fill
            if i in (7, 8):
                cell.number_format = "dd/mm/yy hh:mm"

    ws.freeze_panes = "A4"
    save_preserving_ribbon(wb, erp_file)
    wb.close()


def _format_countdown(td: timedelta, priority: str) -> str:
    total_sec = int(td.total_seconds())
    if total_sec < 0:
        total_sec = -total_sec
        sign = "-"
    else:
        sign = "+"
    h = total_sec // 3600
    m = (total_sec % 3600) // 60
    suffix = " BREACHED" if priority == "URGENT" else ""
    return f"{sign}{h:02d}:{m:02d}{suffix}"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--erp", default=DEFAULT_ERP_FILE)
    ap.add_argument("--hours", type=float, default=2.0,
                    help="Alert deadline after release email sent (default: 2.0)")
    ap.add_argument("--eta-window-days", type=int, default=3,
                    help="Only alert if ETA is within this many days from now")
    ap.add_argument("--now", type=str, default=None,
                    help="Override current time (ISO: 2026-04-14 15:00) for testing")
    args = ap.parse_args()

    if not os.path.exists(args.erp):
        print(f"[ERROR] ERP not found: {args.erp}")
        return 1

    now = datetime.fromisoformat(args.now) if args.now else datetime.now()
    print(f"[+] Release Alerts @ {now:%Y-%m-%d %H:%M}  (deadline={args.hours}h, eta_window={args.eta_window_days}d)")

    alerts = scan_alerts(args.erp, args.hours, args.eta_window_days, now=now)
    print(f"    -> {len(alerts)} pending release (unconfirmed)")

    urgent = [a for a in alerts if a.priority == "URGENT"]
    warn = [a for a in alerts if a.priority == "WARN"]
    info = [a for a in alerts if a.priority == "INFO"]
    print(f"       URGENT: {len(urgent)}  WARN: {len(warn)}  INFO: {len(info)}")

    if alerts:
        write_alerts_sheet(args.erp, alerts)
        print(f"    -> refreshed Release_Alerts sheet")
        print("\nTop alerts:")
        prio = {"URGENT": 0, "WARN": 1, "INFO": 2}
        for a in sorted(alerts, key=lambda x: (prio.get(x.priority, 9), x.countdown_to_deadline))[:10]:
            cd = _format_countdown(a.countdown_to_deadline, a.priority)
            eta_s = a.eta.strftime('%d/%m') if a.eta else "--"
            print(f"  [{a.priority:6s}] {a.customer[:18]:18s} {a.carrier:5s} {a.routing[:14]:14s} "
                  f"ETA={eta_s}  elapsed={a.elapsed_hours:5.1f}h  countdown={cd}")
    else:
        # clear stale sheet if present
        try:
            with open(args.erp, "r+b"):
                pass
        except PermissionError:
            print("[warn] ERP open in Excel, cannot clear stale sheet")
            return 0
        wb = openpyxl.load_workbook(args.erp, keep_vba=True)
        if "Release_Alerts" in wb.sheetnames:
            del wb["Release_Alerts"]
            ws = wb.create_sheet("Release_Alerts")
            ws["A1"] = f"No pending release alerts as of {now:%d %b %Y %H:%M}"
            ws["A1"].font = Font(italic=True, color="666666")
        save_preserving_ribbon(wb, args.erp)
        wb.close()

    return 1 if urgent else 0


if __name__ == "__main__":
    sys.exit(main())
