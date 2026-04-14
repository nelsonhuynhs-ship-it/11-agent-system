# -*- coding: utf-8 -*-
"""
yml_email_scan.py — Feature 8 (Active Jobs v4): YML Email Event Scanner
Reads YML tracking emails from Outlook inbox, parses container events,
and updates Active Jobs rows in ERP_Master_v14.xlsm.

Usage:
    python ERP/jobs/yml_email_scan.py
    python ERP/jobs/yml_email_scan.py --days 30 --dry-run
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import datetime, timedelta, timezone
from typing import Final

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "core"))
from ribbon_guard import save_preserving_ribbon  # noqa: E402
from active_jobs_cols import COL, DATA_START  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]

# ── Constants ──────────────────────────────────────────────────────────────────
DEFAULT_ERP_FILE: Final = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"
AJ_SHEET: Final = "Active Jobs"
AJ_DATA_START: Final = DATA_START

_RE_CONTAINER = re.compile(r"\b([A-Z]{4}\d{7})\b")
_RE_DATE_ISO  = re.compile(r"\b(\d{4}-\d{2}-\d{2})\b")
_RE_DATE_DMY  = re.compile(r"\b(\d{2}/\d{2}/\d{4})\b")
_RE_BKG       = re.compile(r"\b(YM[-\s]?\d{5,10})\b", re.IGNORECASE)

# Event keyword patterns → DCSA milestone code (most specific first)
_EVENT_KEYWORDS: Final[list[tuple[str, str]]] = [
    ("vessel.*depart|sailed|atd|departed", "VD"),
    (r"vessel.*arriv|arrived.*port|arrival.*dest|container.*arriv|arriv.*at\s+\w", "VA"),
    ("gated.?in|gate.?in|gate in at", "GTIN"),
    ("discharg", "DISC"),
    ("on.?board|loaded|load", "LOAD"),
]

_YML_SENDER_TERMS  = ("yangming", "yml.com", "yang-ming")
_YML_SUBJECT_TERMS = ("yml", "yang ming", "yangming", "ym tracking", "yang ming line")
_NOTE_FMT          = "[YML {stamp}] {msg}"
# Container-specific events require a known container number
_CONT_EVENTS       = {"GTIN", "VA", "LOAD", "DISC"}


# ── Pure parser (no COM — fully testable) ─────────────────────────────────────

def _parse_date(text: str) -> str | None:
    """Return YYYY-MM-DD from first date found in text, or None."""
    m = _RE_DATE_ISO.search(text)
    if m:
        return m.group(1)
    m = _RE_DATE_DMY.search(text)
    if m:
        try:
            return datetime.strptime(m.group(1), "%d/%m/%Y").strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def _classify_event(line: str) -> str | None:
    """Map a text line to a DCSA milestone code, or None."""
    lower = line.lower()
    for pattern, code in _EVENT_KEYWORDS:
        if re.search(pattern, lower):
            return code
    return None


def _extract_location(line: str) -> str:
    """Best-effort extract location phrase after 'at' or 'in'."""
    m = re.search(r"\b(?:at|in)\s+([A-Z][A-Za-z ]{1,30}?)(?:\s+on|\s+\d|$|,)", line)
    return m.group(1).strip() if m else ""


def parse_yml_event(email_body: str) -> list[dict]:
    """
    Parse a YML tracking email body. Returns list of event dicts with keys:
        container, event_type, timestamp, location, vessel, raw_line
    Only lines with a recognizable event are included.
    """
    events: list[dict] = []
    current_container = ""
    current_vessel = ""

    for line in email_body.splitlines():
        line = line.strip()
        if not line:
            continue

        cm = _RE_CONTAINER.search(line)
        if cm:
            current_container = cm.group(1)

        vm = re.search(
            r"(?:vessel|m/v|mv)\s+([A-Z][A-Z ]{2,30}?)(?:\s+depart|\s+arriv|\s+ETA|$|,)",
            line, re.IGNORECASE,
        )
        if vm:
            current_vessel = vm.group(1).strip()

        event_type = _classify_event(line)
        if event_type is None and re.search(r"\beta\b", line, re.IGNORECASE) and _parse_date(line):
            event_type = "ETA_INFO"
        if event_type is None:
            continue

        # Container-specific events require a container number on this line or already tracked
        if event_type in _CONT_EVENTS and not cm and not current_container:
            continue

        events.append({
            "container":  current_container,
            "event_type": event_type,
            "timestamp":  _parse_date(line) or "",
            "location":   _extract_location(line),
            "vessel":     current_vessel,
            "raw_line":   line,
        })

    return events


# ── ERP workbook helpers ───────────────────────────────────────────────────────

def _load_active_jobs(ws) -> list[dict]:
    """Read Active Jobs sheet; return list of row dicts."""
    jobs = []
    for r in range(AJ_DATA_START, ws.max_row + 1):
        bkg = ws.cell(r, COL["Bkg_No"]).value
        hbl = ws.cell(r, COL["HBL_NO"]).value
        if not bkg and not hbl:
            continue
        jobs.append({
            "row_idx": r,
            "Bkg_No":  str(bkg or "").strip().upper(),
            "HBL_NO":  str(hbl or "").strip().upper(),
            "ETD":     ws.cell(r, COL["ETD"]).value,
        })
    return jobs


def _match_job(jobs: list[dict], bkg_ref: str, hbl_ref: str = "") -> dict | None:
    """Exact match by Bkg_No or HBL_NO; on multi-match return latest ETD row."""
    bkg_ref = bkg_ref.strip().upper()
    hbl_ref = hbl_ref.strip().upper()
    candidates = [
        j for j in jobs
        if (bkg_ref and j["Bkg_No"] == bkg_ref)
        or (hbl_ref and j["HBL_NO"] == hbl_ref)
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda j: j["ETD"] if isinstance(j["ETD"], datetime) else datetime.min)


def _note_stamp() -> str:
    fmt = "%#d%b %H:%M" if sys.platform == "win32" else "%-d%b %H:%M"
    return datetime.now().strftime(fmt)


def _append_note(ws, row_idx: int, message: str) -> None:
    cell = ws.cell(row_idx, COL["Notes"])
    existing = str(cell.value or "").strip()
    note = _NOTE_FMT.format(stamp=_note_stamp(), msg=message)
    cell.value = f"{existing}\n{note}".strip() if existing else note


def _apply_event(ws, row_idx: int, event: dict) -> str:
    """Write one parsed event to an Active Jobs row. Returns description string."""
    etype     = event["event_type"]
    loc       = event["location"] or "?"
    ts        = event["timestamp"] or ""
    vessel    = event["vessel"]
    container = event["container"]

    if etype == "GTIN":
        msg = f"Gate-in {container} at {loc}" + (f" on {ts}" if ts else "")
        _append_note(ws, row_idx, msg)
        return f"GTIN: {msg}"

    if etype == "VD":
        ws.cell(row_idx, COL["Status"]).value = "In Transit"
        msg = f"ATD {ts or '?'}" + (f" vessel {vessel}" if vessel else "") + (f" from {loc}" if loc != "?" else "")
        _append_note(ws, row_idx, msg)
        return "VD: status=In Transit"

    if etype == "VA":
        ata_cell = ws.cell(row_idx, COL["ATA"])
        if ts and not ata_cell.value:
            try:
                ata_cell.value = datetime.strptime(ts, "%Y-%m-%d")
            except ValueError:
                pass
        _append_note(ws, row_idx, f"ATA {ts or '?'} at {loc}")
        return f"VA: ATA={ts}"

    if etype == "ETA_INFO":
        msg = f"ETA update: {ts}" + (f" ({vessel})" if vessel else "")
        _append_note(ws, row_idx, msg)
        return f"ETA_INFO: {ts}"

    # LOAD, DISC, other
    msg = f"{etype} {container} at {loc}" + (f" on {ts}" if ts else "")
    _append_note(ws, row_idx, msg)
    return f"{etype}: note appended"


# ── Outlook COM layer ──────────────────────────────────────────────────────────

def _is_yml_message(msg) -> bool:
    try:
        sender  = str(msg.SenderEmailAddress or "").lower()
        subject = str(msg.Subject or "").lower()
    except Exception:
        return False
    return any(t in sender for t in _YML_SENDER_TERMS) \
        or any(t in subject for t in _YML_SUBJECT_TERMS)


def _collect_yml_emails(days: int) -> list[dict]:
    """Scan Outlook inbox (+ YML subfolder) for YML emails in last N days."""
    try:
        import win32com.client  # type: ignore
    except ImportError:
        print("Outlook not available — skip")
        sys.exit(0)

    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        ns      = outlook.GetNamespace("MAPI")
    except Exception as e:
        print(f"Outlook not available — skip ({e})")
        sys.exit(0)

    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    emails: list[dict] = []

    def _scan(folder):
        try:
            items = folder.Items
            items.Sort("[ReceivedTime]", True)
            for msg in items:
                try:
                    recv = msg.ReceivedTime
                    recv_utc = recv if (hasattr(recv, "tzinfo") and recv.tzinfo) \
                               else recv.replace(tzinfo=timezone.utc)
                    if recv_utc < cutoff:
                        break
                    if _is_yml_message(msg):
                        emails.append({"subject": str(msg.Subject or ""),
                                       "body":    str(msg.Body or ""),
                                       "received": recv})
                except Exception:
                    continue
        except Exception:
            pass

    inbox = ns.GetDefaultFolder(6)  # olFolderInbox
    _scan(inbox)
    try:
        for folder in inbox.Folders:
            if "yml" in str(folder.Name).lower():
                _scan(folder)
                break
    except Exception:
        pass

    return emails


# ── Main pipeline ──────────────────────────────────────────────────────────────

def run_scan(erp_file: str = DEFAULT_ERP_FILE, days: int = 7, dry_run: bool = False) -> dict:
    """Collect YML emails, parse events, match jobs, update ERP. Returns summary dict."""
    emails = _collect_yml_emails(days)
    print(f"Found {len(emails)} YML email(s) in last {days} day(s).")
    if not emails:
        return {"emails_scanned": 0, "matched": 0, "events_applied": 0, "dry_run": dry_run}

    parsed = []
    for email in emails:
        events = parse_yml_event(email["body"])
        bkg_ref = ""
        m = _RE_BKG.search(email["subject"] + " " + email["body"])
        if m:
            bkg_ref = m.group(1).replace(" ", "").upper()
        parsed.append((email["subject"], bkg_ref, events))

    if dry_run or not os.path.exists(erp_file):
        if not os.path.exists(erp_file):
            print(f"ERP file not found: {erp_file} — dry-run only.")
        total = sum(len(e) for _, _, e in parsed)
        print(f"[dry-run] Would process {total} event(s) from {len(emails)} email(s).")
        return {"emails_scanned": len(emails), "matched": 0, "events_applied": 0, "dry_run": True}

    wb = openpyxl.load_workbook(erp_file, keep_vba=True)
    ws = wb[AJ_SHEET]
    jobs = _load_active_jobs(ws)
    matched_rows: set[int] = set()
    events_applied = 0

    for subject, bkg_ref, events in parsed:
        if not events:
            continue
        job = _match_job(jobs, bkg_ref)
        if not job:
            print(f"  No match: {subject[:60]}")
            continue
        row_idx = job["row_idx"]
        matched_rows.add(row_idx)
        for event in events:
            print(f"  Row {row_idx}: {_apply_event(ws, row_idx, event)}")
            events_applied += 1

    if events_applied > 0:
        save_preserving_ribbon(wb, erp_file)
        print(f"Saved: {erp_file}")

    return {"emails_scanned": len(emails), "matched": len(matched_rows),
            "events_applied": events_applied, "dry_run": False}


# ── CLI ────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Scan YML tracking emails → update Active Jobs.")
    ap.add_argument("--days",     type=int, default=7,             help="Days back to scan (default 7)")
    ap.add_argument("--erp-file", default=DEFAULT_ERP_FILE,        help="Path to ERP_Master_v14.xlsm")
    ap.add_argument("--dry-run",  action="store_true",             help="Parse only, no ERP write")
    args = ap.parse_args()

    r = run_scan(erp_file=args.erp_file, days=args.days, dry_run=args.dry_run)
    print(f"\nSummary: {r['emails_scanned']} emails scanned | "
          f"{r['matched']} matched to jobs | "
          f"{r['events_applied']} events applied"
          + (" [DRY RUN]" if r["dry_run"] else ""))
