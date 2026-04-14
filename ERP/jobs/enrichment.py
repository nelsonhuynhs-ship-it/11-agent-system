"""
enrichment.py — enrich Active Jobs with booking-email hyperlinks
=================================================================
For every Active Jobs row missing a Request_BKG email link (col 28),
build a mailto: URL from `email_builder.build_mailto_link()` and write
it as an Excel hyperlink. Cost_Breakdown (col 27) is populated by VBA
at Mark-WIN time and is NOT touched here.

Also (best-effort): if SERVICE (col 31) is blank, auto-fill from
Door_Delivery (col 19): "Yes" → CY-DOOR, else → CY-CY.

Usage:
    python ERP/jobs/enrichment.py
    python ERP/jobs/enrichment.py --erp <file> --force   # overwrite existing links
"""
from __future__ import annotations

import argparse
import os
import sys
from typing import Final

import openpyxl
from openpyxl.styles import Alignment, Font

sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)
sys.path.insert(0, os.path.join(SCRIPT_DIR, "..", "core"))
from email_builder import build_mailto_link, load_rules  # noqa: E402
from ribbon_guard import save_preserving_ribbon  # noqa: E402
from active_jobs_cols import COL, HDR_ROW, DATA_START  # noqa: E402

DEFAULT_ERP_FILE: Final = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"


def parse_routing(routing: str) -> tuple[str, str, str]:
    """Parse 'HPH-USLGB' / 'HCM-CHICAGO VIA USLAX' → (pol, pod, place)."""
    if not routing:
        return "", "", ""
    parts = [p.strip() for p in str(routing).split("-", 1)]
    pol = parts[0]
    if len(parts) < 2:
        return pol, "", ""
    tail = parts[1]
    # "CHICAGO VIA USLAX" → place=CHICAGO, pod=USLAX
    if " VIA " in tail.upper():
        idx = tail.upper().rfind(" VIA ")
        place = tail[:idx].strip()
        pod = tail[idx + 5:].strip()
        return pol, pod, place
    # "USLGB"  → pod=USLGB, place=USLGB
    return pol, tail, tail


def enrich(erp_file: str, force: bool = False) -> dict:
    if not os.path.exists(erp_file):
        raise FileNotFoundError(erp_file)
    try:
        with open(erp_file, "r+b"):
            pass
    except PermissionError:
        raise RuntimeError(f"ERP file is open in Excel. Close it first: {erp_file}")

    rules = load_rules()
    wb = openpyxl.load_workbook(erp_file, keep_vba=True)

    sheet = next((s for s in wb.sheetnames if "Active" in s), None)
    if not sheet:
        wb.close()
        raise RuntimeError("Active Jobs sheet not found")
    ws = wb[sheet]

    n_total = n_linked = n_service = 0

    for r in range(DATA_START, ws.max_row + 1):
        crm = ws.cell(r, COL["CRM_ID"]).value
        if not crm:
            continue
        n_total += 1

        # Auto-fill SERVICE col 31 if blank
        service = ws.cell(r, COL["SERVICE"]).value
        if not service:
            door = (ws.cell(r, COL["Door_Delivery"]).value or "")
            svc = "CY-DOOR" if str(door).strip().lower() in ("yes", "y", "true", "1") else "CY-CY"
            c = ws.cell(r, COL["SERVICE"], svc)
            c.font = Font(size=10, name="Segoe UI")
            c.alignment = Alignment(horizontal="center", vertical="center")
            n_service += 1

        # Email link (col 28) — skip if already set unless --force
        existing = ws.cell(r, COL["Request_BKG"]).value
        if existing and not force:
            continue

        routing = ws.cell(r, COL["Routing"]).value or ""
        pol, pod, place = parse_routing(str(routing))
        door_addr = ws.cell(r, COL["Door_Address"]).value or place

        carrier = str(ws.cell(r, COL["Carrier"]).value or "").strip()
        contract = str(ws.cell(r, COL["Contract_Type"]).value or "").strip()
        container = str(ws.cell(r, COL["Container_Type"]).value or "").strip()
        qty = int(ws.cell(r, COL["Quantity"]).value or 1)
        is_soc = "SOC" in contract.upper() or "SOC" in str(ws.cell(r, COL["Cost_Breakdown"]).value or "").upper()

        job_data = {
            "Customer_Name": str(crm).strip(),
            "POL": pol, "POD": pod, "Place": door_addr or place,
            "Carrier": carrier,
            "Container_Type": container or "40HQ",
            "Quantity": qty,
            "Contract_No": contract,
            "is_SOC": is_soc,
        }
        try:
            mailto = build_mailto_link(job_data, rules=rules)
        except Exception as e:
            print(f"    [warn] row {r} mailto failed: {e}")
            continue

        cell = ws.cell(r, COL["Request_BKG"], "📧 Send BKG")
        cell.hyperlink = mailto
        cell.font = Font(color="0563C1", underline="single", size=10, name="Segoe UI")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        n_linked += 1

    save_preserving_ribbon(wb, erp_file)
    wb.close()
    return {"total": n_total, "linked": n_linked, "service_filled": n_service}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--erp", default=DEFAULT_ERP_FILE)
    ap.add_argument("--force", action="store_true",
                    help="Overwrite existing Request_BKG links")
    args = ap.parse_args()

    print(f"[+] Enrich Active Jobs: {args.erp}")
    stats = enrich(args.erp, force=args.force)
    print(f"    -> {stats['total']} jobs scanned")
    print(f"    -> {stats['linked']} email link(s) added")
    print(f"    -> {stats['service_filled']} SERVICE cells filled")
    print(f"[OK] saved")
    return 0


if __name__ == "__main__":
    sys.exit(main())
