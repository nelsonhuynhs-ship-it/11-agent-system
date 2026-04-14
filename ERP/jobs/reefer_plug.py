"""reefer_plug.py — Feature 7: Reefer Power Plug Fee Calculator
Computes plug fee + demurrage for 20RF/40RF containers and finds the
optimal (last cost-free) terminal drop-off date.

CLI: python ERP/jobs/reefer_plug.py [--write] [--file path/to/ERP.xlsm]
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Final

import openpyxl
import yaml

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, REPO_ROOT)
from ERP.core.ribbon_guard import save_preserving_ribbon  # noqa: E402
from ERP.core.active_jobs_cols import COL, DATA_START  # noqa: E402

_DATA_DIR: Final = Path(__file__).resolve().parents[1] / "data"
YAML_PATH: Final = _DATA_DIR / "reefer_freetime.yaml"
DEFAULT_ERP: Final = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"

COL_ETA:   Final = COL["ETA"]           # col U (21) — ETA (hidden)
COL_ATA:   Final = COL["ATA"]           # col V (22) — ATA (hidden)
COL_CTYPE: Final = COL["Container_Type"]  # col J (10) — Container_Type
COL_NOTES: Final = COL["Notes"]         # col AF (32) — Notes (hidden)
COL_POD:   Final = COL["POL_POD"]       # col E (5)  — POL_POD rendered string

RF_TYPES = {"20RF", "40RF"}
sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]

_DEFAULT_YAML = """\
terminals:
  USLGB: {freetime_days: 4, daily_fee_20RF: 150, daily_fee_40RF: 200}
  USLAX: {freetime_days: 4, daily_fee_20RF: 160, daily_fee_40RF: 210}
  USNYC: {freetime_days: 5, daily_fee_20RF: 180, daily_fee_40RF: 230}
  default: {freetime_days: 4, daily_fee_20RF: 150, daily_fee_40RF: 200}
demurrage:
  freetime_days: 7
  daily_fee_20RF: 100
  daily_fee_40RF: 140
"""


def load_reefer_rules() -> dict:
    """Load YAML config; auto-create with defaults if missing."""
    if not YAML_PATH.exists():
        _DATA_DIR.mkdir(parents=True, exist_ok=True)
        YAML_PATH.write_text(_DEFAULT_YAML, encoding="utf-8")
        print(f"[reefer_plug] Created default config: {YAML_PATH}")
    rules = yaml.safe_load(YAML_PATH.read_text(encoding="utf-8"))
    if "terminals" not in rules or "demurrage" not in rules:
        raise ValueError(f"reefer_freetime.yaml missing 'terminals'/'demurrage': {YAML_PATH}")
    return rules


def plug_cost(eta: date, drop_date: date, pod: str, cont_type: str, rules: dict) -> dict:
    """Compute plug fee + demurrage for a given drop date.

    Returns {plug_days, plug_fee, demurrage_days, demurrage_fee, total}.
    drop_date = date customer picks up container from terminal.
    plug_fee  starts after terminal freetime_days.
    demurrage starts after demurrage.freetime_days (usually 7).
    """
    ct = cont_type.strip().upper()
    if ct not in RF_TYPES:
        raise ValueError(f"cont_type must be '20RF' or '40RF', got: {cont_type!r}")
    term = rules["terminals"].get(pod.strip().upper()) or rules["terminals"]["default"]
    dem  = rules["demurrage"]
    fk   = f"daily_fee_{ct}"
    days = max(0, (drop_date - eta).days)
    plug_d = max(0, days - term["freetime_days"])
    dem_d  = max(0, days - dem["freetime_days"])
    plug_f = plug_d * term[fk]
    dem_f  = dem_d  * dem[fk]
    return {"plug_days": plug_d, "plug_fee": plug_f,
            "demurrage_days": dem_d, "demurrage_fee": dem_f, "total": plug_f + dem_f}


def optimal_drop_date(eta: date, pod: str, cont_type: str, rules: dict) -> dict:
    """Return latest drop date with $0 plug fee = eta + freetime_days.

    Returns {optimal_date, freetime_days, cost_at_optimal, daily_plug_fee}.
    """
    term      = rules["terminals"].get(pod.strip().upper()) or rules["terminals"]["default"]
    free_days = term["freetime_days"]
    optimal   = eta + timedelta(days=free_days)
    return {
        "optimal_date":    optimal,
        "freetime_days":   free_days,
        "cost_at_optimal": plug_cost(eta, optimal, pod, cont_type, rules),
        "daily_plug_fee":  term[f"daily_fee_{cont_type.strip().upper()}"],
    }


def compute_for_active_jobs(erp_file: str) -> list[dict]:
    """Scan Active Jobs for 20RF/40RF rows; compute optimal drop per row."""
    results: list[dict] = []
    if not os.path.exists(erp_file):
        print(f"[WARN] ERP file not found: {erp_file}")
        return results
    rules = load_reefer_rules()
    wb = openpyxl.load_workbook(erp_file, keep_vba=True, data_only=True)
    ws = _active_sheet(wb)
    if ws is None:
        wb.close()
        return results
    for row in ws.iter_rows(min_row=DATA_START, values_only=False):
        row_idx = row[0].row
        ctype   = _cs(row, COL_CTYPE)
        if ctype not in RF_TYPES:
            continue
        anchor = _to_date(_cv(row, COL_ATA)) or _to_date(_cv(row, COL_ETA))
        if anchor is None:
            results.append({"row": row_idx, "ctype": ctype, "error": "No ETA/ATA"})
            continue
        pod = _cs(row, COL_POD) or "default"
        opt = optimal_drop_date(anchor, pod, ctype, rules)
        opt_str = opt["optimal_date"].strftime("%d/%b").lstrip("0")
        note = (f"[RF {anchor.strftime('%d%b').upper()}] Optimal drop: {opt_str} | "
                f"Plug: $0 (if on time) | ${opt['daily_plug_fee']}/day after freetime")
        results.append({"row": row_idx, "ctype": ctype, "pod": pod, "anchor": anchor,
                        "optimal": opt["optimal_date"], "free_days": opt["freetime_days"],
                        "daily_fee": opt["daily_plug_fee"], "note": note})
    wb.close()
    return results


def write_notes_to_jobs(erp_file: str, results: list[dict]) -> None:
    """Write RF note into col 24 (Notes) for each result row. Ribbon-safe."""
    if not os.path.exists(erp_file):
        print(f"[ERROR] ERP file not found: {erp_file}")
        return
    wb = openpyxl.load_workbook(erp_file, keep_vba=True)
    ws = _active_sheet(wb)
    if ws is None:
        wb.close()
        return
    written = sum(
        1 for r in results if "note" in r
        and not ws.cell(r["row"], COL_NOTES).__setattr__("value", r["note"])  # type: ignore
        or False
    )
    # Direct loop is clearer — redo properly:
    written = 0
    for r in results:
        if "note" in r:
            ws.cell(r["row"], COL_NOTES).value = r["note"]
            written += 1
    ribbon = save_preserving_ribbon(wb, erp_file)
    wb.close()
    print(f"[write_notes] Updated {written} rows. ribbon={ribbon}")


# --- helpers ----------------------------------------------------------------

def _active_sheet(wb):
    name = next((s for s in wb.sheetnames if "active" in s.lower()), None)
    if name is None:
        print("[WARN] No 'Active*' sheet found")
    return wb[name] if name else None


def _cv(row, col: int):
    i = col - 1
    return row[i].value if 0 <= i < len(row) else None


def _cs(row, col: int) -> str:
    v = _cv(row, col)
    return str(v).strip().upper() if v else ""


def _to_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val).strip()[:10], "%Y-%m-%d").date()
    except Exception:
        return None


# --- CLI --------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description="Reefer plug fee scanner for Active Jobs")
    ap.add_argument("--write", action="store_true", help="Write optimal dates to Notes col")
    ap.add_argument("--file",  default=DEFAULT_ERP, help="Path to ERP_Master_v14.xlsm")
    ap.add_argument("--erp",   dest="file", help="Alias for --file")
    args = ap.parse_args()
    results = compute_for_active_jobs(args.file)
    if not results:
        print("[reefer_plug] No 20RF/40RF rows found or ERP file missing.")
        return 0
    print(f"\n{'Row':>4}  {'Type':6}  {'POD':8}  {'Anchor':10}  {'Optimal':10}  {'$/day':>6}")
    print("-" * 56)
    for r in results:
        if "error" in r:
            print(f"{r['row']:>4}  {r['ctype']:6}  {'—':8}  ERROR: {r['error']}")
        else:
            print(f"{r['row']:>4}  {r['ctype']:6}  {r['pod']:8}  "
                  f"{r['anchor'].isoformat():10}  {r['optimal'].isoformat():10}  ${r['daily_fee']:>5}")
    if args.write:
        write_notes_to_jobs(args.file, results)
    else:
        print("\n[Tip] Run with --write to save optimal dates to Notes column.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
