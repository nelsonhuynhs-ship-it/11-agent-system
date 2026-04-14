"""transit_time.py — F9 Active Jobs v4: Transit Time Auto-Calculator.

Auto-fills ETA (col 6) from ETD + route class. Vietnam→USA/Canada.
Route classes: WC 18-20d | EC/GULF 40-50d | CA_WC 18-22d | CA_EC 35-45d | +INLAND +5d.
Usage: python ERP/jobs/transit_time.py [--dry-run] [--overwrite]
"""
from __future__ import annotations

import argparse
import logging
import os
import sys
import warnings
from datetime import datetime, timedelta
from typing import Final

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "core"))
from ribbon_guard import save_preserving_ribbon  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]

log = logging.getLogger(__name__)

DEFAULT_ERP_FILE: Final = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"

AJ_HDR_ROW = 7
AJ_DATA_START = 8
COL: Final = {
    "CRM_ID": 1, "Routing": 3, "ETD": 5, "ETA": 6, "ATA": 7,
    "Door_Address": 20, "Notes": 24,
}

_WC_KEYWORDS: Final = ("LAX", "LGB", "OAK", "SEA", "TAC", "USLAX", "USLGB", "USOAK", "USSEA")
_EC_KEYWORDS: Final = ("NYC", "SAV", "CHS", "NORFOLK", "BAL", "BOS", "SAVANNAH",
                       "CHARLESTON", "BALTIMORE", "BOSTON", "NEW YORK")
_GULF_KEYWORDS: Final = ("HOUSTON", "HOU", "NEW ORLEANS", "MOB", "MIAMI", "MOBILE")
_CA_WC_KEYWORDS: Final = ("PRINCE RUPERT", "VANCOUVER", "CAPRR", "CAVAN")
_CA_EC_KEYWORDS: Final = ("MONTREAL", "HALIFAX", "CAHAL", "CAMTR")

_WINDOWS: Final[dict[str, tuple[int, int]]] = {
    "WC": (18, 20),
    "EC": (40, 50),
    "GULF": (40, 50),
    "CA_WC": (18, 22),
    "CA_EC": (35, 45),
    "WC+INLAND": (23, 25),
    "EC+INLAND": (45, 55),
    "GULF+INLAND": (45, 55),
    "CA_WC+INLAND": (23, 27),
    "CA_EC+INLAND": (40, 50),
}

_INLAND_DAYS: Final = 5


def classify_route(pol: str, pod: str, place: str = "") -> str:
    """Return route class from POD keywords + optional door address. Unknown POD defaults to EC (warning)."""
    pod_up = (pod or "").upper()
    place_up = (place or "").upper()
    base: str
    if any(k in pod_up for k in _CA_WC_KEYWORDS):
        base = "CA_WC"
    elif any(k in pod_up for k in _CA_EC_KEYWORDS):
        base = "CA_EC"
    elif any(k in pod_up for k in _WC_KEYWORDS):
        base = "WC"
    elif any(k in pod_up for k in _EC_KEYWORDS):
        base = "EC"
    elif any(k in pod_up for k in _GULF_KEYWORDS):
        base = "GULF"
    else:
        warnings.warn(f"classify_route: unrecognised POD '{pod}' — defaulting to EC", stacklevel=2)
        base = "EC"

    inland = _is_inland(place_up, pod_up)
    if inland:
        return f"{base}+INLAND"
    return base


def transit_window(route_class: str) -> tuple[int, int]:
    """Return (min_days, max_days). Raises ValueError for unknown class."""
    if route_class not in _WINDOWS:
        raise ValueError(f"Unknown route class '{route_class}'. Valid: {', '.join(_WINDOWS)}")
    return _WINDOWS[route_class]


def estimate_eta(
    etd_or_atd: datetime,
    route_class: str,
) -> tuple[datetime, datetime]:
    """Return (earliest_eta, latest_eta) from a departure date and route class."""
    min_d, max_d = transit_window(route_class)
    return (
        etd_or_atd + timedelta(days=min_d),
        etd_or_atd + timedelta(days=max_d),
    )


def update_active_jobs(erp_file: str, overwrite: bool = False, dry_run: bool = False) -> dict:
    """Fill missing ETA (or overwrite all) from ETD + route rules. Returns stats dict."""
    if not os.path.exists(erp_file):
        raise FileNotFoundError(erp_file)
    if not dry_run:
        try:
            with open(erp_file, "r+b"):
                pass
        except PermissionError:
            raise RuntimeError(f"ERP file locked — close Excel first: {erp_file}")

    wb = openpyxl.load_workbook(erp_file, keep_vba=True)
    sheet = next((s for s in wb.sheetnames if "Active" in s), None)
    if not sheet:
        wb.close()
        raise RuntimeError("Active Jobs sheet not found in workbook")
    ws = wb[sheet]

    stamp = datetime.now().strftime("%d%b %H:%M")
    counts: dict[str, int] = {}
    total = filled = overwritten = 0

    for r in range(AJ_DATA_START, ws.max_row + 1):
        crm = ws.cell(r, COL["CRM_ID"]).value
        if not crm:
            continue
        total += 1

        routing = str(ws.cell(r, COL["Routing"]).value or "")
        etd = ws.cell(r, COL["ETD"]).value
        eta_existing = ws.cell(r, COL["ETA"]).value
        door = str(ws.cell(r, COL["Door_Address"]).value or "")

        if not isinstance(etd, datetime):
            continue
        has_eta = isinstance(eta_existing, datetime)
        if has_eta and not overwrite:
            continue

        pol, pod = _parse_routing(routing)

        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            route_class = classify_route(pol, pod, door)
        if caught:
            log.warning("Row %d: %s", r, caught[0].message)

        earliest, latest = estimate_eta(etd, route_class)
        median_days = (transit_window(route_class)[0] + transit_window(route_class)[1]) // 2
        median_eta = etd + timedelta(days=median_days)

        counts[route_class] = counts.get(route_class, 0) + 1
        if has_eta:
            overwritten += 1
        else:
            filled += 1

        if dry_run:
            print(f"  [DRY] row {r:>3} {routing:<25} {route_class:<12} "
                  f"ETD {etd:%d-%b} -> ETA {median_eta:%d-%b} "
                  f"(window {earliest:%d-%b}—{latest:%d-%b})")
            continue

        ws.cell(r, COL["ETA"]).value = median_eta
        existing_notes = str(ws.cell(r, COL["Notes"]).value or "")
        # Use strftime %d then lstrip('0') — works on Win + Linux (%-d is POSIX-only)
        e_str = earliest.strftime("%d/%b").lstrip("0")
        l_str = latest.strftime("%d/%b").lstrip("0")
        tag = f"[TT {stamp}] ETA {e_str}—{l_str} ({route_class})"
        ws.cell(r, COL["Notes"]).value = f"{existing_notes} {tag}".strip() if existing_notes else tag

    if not dry_run:
        save_preserving_ribbon(wb, erp_file)
    wb.close()

    return {
        "total": total,
        "filled": filled,
        "overwritten": overwritten,
        "by_route": counts,
    }


def _is_inland(place_up: str, pod_up: str) -> bool:
    """True when door address is present and contains no recognised port keyword."""
    if not place_up:
        return False
    all_port_tokens = _WC_KEYWORDS + _EC_KEYWORDS + _GULF_KEYWORDS + _CA_WC_KEYWORDS + _CA_EC_KEYWORDS
    return not any(token in place_up for token in all_port_tokens)


def _parse_routing(routing: str) -> tuple[str, str]:
    """Extract (pol, pod) from 'HPH-USLGB' or 'HCM-CHICAGO VIA USLAX'."""
    upper = routing.upper()
    if " VIA " in upper:
        parts = upper.split(" VIA ")
        pod = parts[-1].strip().split()[0]
        pol = upper.split("-")[0].strip() if "-" in upper else ""
        return pol, pod
    if "-" in routing:
        parts = routing.split("-", 1)
        return parts[0].strip().upper(), parts[1].strip().upper()
    return "", upper

def main() -> int:
    ap = argparse.ArgumentParser(
        description="Transit Time Auto-Calculator for Active Jobs ETA"
    )
    ap.add_argument("--erp", default=DEFAULT_ERP_FILE, help="Path to ERP_Master_v14.xlsm")
    ap.add_argument("--dry-run", action="store_true", help="Show changes without writing")
    ap.add_argument("--overwrite", action="store_true", help="Replace existing ETA values")
    args = ap.parse_args()

    logging.basicConfig(level=logging.WARNING, format="%(levelname)s %(message)s")
    print(f"[+] Transit Time Calculator @ {datetime.now():%Y-%m-%d %H:%M}")
    print(f"    ERP: {args.erp}")
    if args.dry_run:
        print("    Mode: DRY-RUN")
    elif args.overwrite:
        print("    Mode: OVERWRITE existing ETA")
    else:
        print("    Mode: FILL missing ETA only")

    stats = update_active_jobs(args.erp, overwrite=args.overwrite, dry_run=args.dry_run)

    print(f"\n    Scanned : {stats['total']} jobs")
    print(f"    Filled  : {stats['filled']} ETA")
    print(f"    Updated : {stats['overwritten']} ETA (overwrite)")
    if stats["by_route"]:
        print("    By route:")
        for rc, n in sorted(stats["by_route"].items()):
            print(f"      {rc:<16} {n}")
    if args.dry_run:
        print("\n    [DRY-RUN] no changes written")
    else:
        print(f"\n[OK] ERP saved: {args.erp}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
