"""fast_id.py — Feature 4: FAST ID Validator + Normalizer
FAST ID format: {PREFIX}{YY}{MM}/{SEQ}
  PREFIX=2-4 uppercase letters, YY/MM=2-digit, SEQ>=4 digits zero-padded.
  e.g. SE2603/0266  NF2604/1200  ARB2512/9999
CLI: python ERP/jobs/fast_id.py --check|--fix [--file path/to/ERP.xlsm]
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from typing import Final

import openpyxl
from openpyxl.styles import PatternFill

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, REPO_ROOT)
from ERP.core.ribbon_guard import save_preserving_ribbon  # noqa: E402
from ERP.core.active_jobs_cols import COL, DATA_START  # noqa: E402

DEFAULT_ERP: Final = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"
COL_FAST_ID: Final = COL["FAST_ID"]    # col B (2) — FAST_ID
COL_STATUS:  Final = COL["Status"]     # col N (14) — Status
COL_CRM:     Final = COL["CRM_ID"]     # col D (4) — CRM_ID

FILL_YELLOW = PatternFill("solid", fgColor="FFFF00")  # invalid format
FILL_RED    = PatternFill("solid", fgColor="FF0000")  # duplicate

sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]


def normalize_fast_id(raw: str) -> str:
    """Return normalized FAST ID or raise ValueError with a helpful message.

    Rules: strip whitespace, uppercase, enforce PREFIX(2-4 letters)+YYMM/SEQ(4+digits).
    """
    if not raw or not isinstance(raw, str):
        raise ValueError(f"FAST ID must be a non-empty string, got: {raw!r}")
    cleaned = raw.strip().upper()
    if "/" not in cleaned:
        raise ValueError(f"FAST ID '{raw}' missing '/'. Expected: PREFIX YYMM/SEQ  e.g. SE2603/0266")
    left, right = cleaned.split("/", 1)
    if not right.isdigit():
        raise ValueError(f"FAST ID '{raw}': sequence '{right}' must be digits only")
    seq = right.zfill(4)
    m = re.match(r"^([A-Z]{2,4})(\d{4})$", left)
    if not m:
        raise ValueError(
            f"FAST ID '{raw}': left part '{left}' must be 2-4 letters + 4 digits (YYMM). e.g. SE2603"
        )
    yymm = m.group(2)
    if not (1 <= int(yymm[2:]) <= 12):
        raise ValueError(f"FAST ID '{raw}': month '{yymm[2:]}' out of range 01-12")
    return f"{m.group(1)}{yymm}/{seq}"


def validate_active_jobs(erp_file: str) -> dict:
    """Scan Active Jobs, return {invalid, duplicates, missing_delivered}."""
    report: dict = {"invalid": [], "duplicates": [], "missing_delivered": []}
    if not os.path.exists(erp_file):
        print(f"[WARN] ERP file not found: {erp_file}")
        return report
    wb = openpyxl.load_workbook(erp_file, keep_vba=True, data_only=True)
    ws = _active_sheet(wb)
    if ws is None:
        wb.close()
        return report
    seen: dict[str, int] = {}
    for row in ws.iter_rows(min_row=DATA_START, values_only=False):
        row_idx  = row[0].row
        raw_fast = _cv(row, COL_FAST_ID)
        raw_stat = _cv(row, COL_STATUS)
        raw_crm  = _cv(row, COL_CRM)
        if not raw_crm and not raw_fast:
            continue
        if str(raw_stat).strip().lower() == "delivered" and not raw_fast:
            report["missing_delivered"].append((row_idx, raw_crm))
            continue
        if not raw_fast:
            continue
        try:
            norm = normalize_fast_id(str(raw_fast))
        except ValueError as exc:
            report["invalid"].append((row_idx, raw_fast, str(exc)))
            continue
        if norm in seen:
            report["duplicates"].append((seen[norm], row_idx, norm))
        else:
            seen[norm] = row_idx
    wb.close()
    return report


def stamp_warnings(erp_file: str, report: dict) -> None:
    """Color col 29: yellow=invalid, red=duplicate. Saves with ribbon guard."""
    if not os.path.exists(erp_file):
        print(f"[ERROR] ERP file not found: {erp_file}")
        return
    wb = openpyxl.load_workbook(erp_file, keep_vba=True)
    ws = _active_sheet(wb)
    if ws is None:
        wb.close()
        return
    for r, _v, _reason in report["invalid"]:
        ws.cell(r, COL_FAST_ID).fill = FILL_YELLOW
    for r1, r2, _v in report["duplicates"]:
        ws.cell(r1, COL_FAST_ID).fill = FILL_RED
        ws.cell(r2, COL_FAST_ID).fill = FILL_RED
    result = save_preserving_ribbon(wb, erp_file)
    wb.close()
    n = len(report["invalid"]) + len(report["duplicates"]) * 2
    print(f"[stamp_warnings] Colored {n} cells. ribbon={result}")


# --- helpers ----------------------------------------------------------------

def _active_sheet(wb):
    name = next((s for s in wb.sheetnames if "active" in s.lower()), None)
    if name is None:
        print("[WARN] No 'Active*' sheet found")
    return wb[name] if name else None


def _cv(row, col: int):
    """Cell value by 1-based column index."""
    i = col - 1
    return row[i].value if 0 <= i < len(row) else None


def _print_report(report: dict) -> None:
    print("\n" + "=" * 60 + "\nFAST ID VALIDATION REPORT\n" + "=" * 60)
    if report["invalid"]:
        print(f"\n[INVALID FORMAT] {len(report['invalid'])} rows:")
        for row, val, reason in report["invalid"]:
            print(f"  row {row:>4} | {val!r:20} | {reason}")
    if report["duplicates"]:
        print(f"\n[DUPLICATES] {len(report['duplicates'])} pairs:")
        for r1, r2, val in report["duplicates"]:
            print(f"  rows {r1} & {r2} share: {val}")
    if report["missing_delivered"]:
        print(f"\n[MISSING — Delivered] {len(report['missing_delivered'])} rows:")
        for row, crm in report["missing_delivered"]:
            print(f"  row {row:>4} | CRM: {crm}")
    if not any(report.values()):
        print("\n[OK] All FAST IDs valid and unique.")
    print("=" * 60)


def _fix_in_place(erp_file: str) -> None:
    """Normalize valid-but-unstandardized FAST IDs in-place (ribbon-safe)."""
    if not os.path.exists(erp_file):
        print(f"[ERROR] ERP file not found: {erp_file}")
        return
    wb = openpyxl.load_workbook(erp_file, keep_vba=True)
    ws = _active_sheet(wb)
    if ws is None:
        wb.close()
        return
    fixed = 0
    for row in ws.iter_rows(min_row=DATA_START):
        cell = row[COL_FAST_ID - 1]
        raw  = cell.value
        if not raw:
            continue
        try:
            norm = normalize_fast_id(str(raw))
            if norm != str(raw).strip().upper():
                cell.value = norm
                fixed += 1
        except ValueError:
            pass
    if fixed:
        result = save_preserving_ribbon(wb, erp_file)
        print(f"[fix] Normalized {fixed} FAST ID(s). ribbon={result}")
    else:
        print("[fix] Nothing to normalize.")
    wb.close()


# --- CLI --------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description="FAST ID validator for Active Jobs")
    ap.add_argument("--check", action="store_true", help="Report only — no writes")
    ap.add_argument("--fix",   action="store_true", help="Normalize + stamp colors")
    ap.add_argument("--file",  default=DEFAULT_ERP, help="Path to ERP_Master_v14.xlsm")
    ap.add_argument("--erp",   dest="file", help="Alias for --file (consistent with other v4 scripts)")
    args = ap.parse_args()
    # default to --check if neither flag given (safer than printing help + exit 1)
    if not args.check and not args.fix:
        args.check = True
    report = validate_active_jobs(args.file)
    _print_report(report)
    if args.fix:
        _fix_in_place(args.file)
        stamp_warnings(args.file, report)
    return 0


if __name__ == "__main__":
    sys.exit(main())
