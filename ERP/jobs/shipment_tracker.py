"""
shipment_tracker.py — Feature 2 (Active Jobs v4): 7-stage pipeline
===================================================================
Computes current stage for each Active Job based on existing ERP fields,
writes a compact stage code into col 32 TRACKING_STAGE.

Pipeline stages (matching Nelson's workflow, HTML plan):
  1 BKG       — booking request sent
  2 Conf      — carrier confirmed (Bkg_No issued)
  3 SI Cut    — shipping instruction submitted / SI received
  4 Gate-in   — container gated into POL yard
  5 ATD       — vessel departed POL
  6 ETA       — vessel arrived POD (ATA set)
  7 Done      — released & delivered to customer

Inference rules (best-effort from existing cols — all optional, graceful):
  Stage 1 (BKG)    : Status ∈ {Booked, Pending} or Bkg_No present
  Stage 2 (Conf)   : Bkg_No matches pattern + Contract_Type present
  Stage 3 (SI Cut) : SI_Received (col 17) not null
  Stage 4 (Gate-in): Notes or Status contains "GATE", OR today >= CY_Cutoff (col 18)
  Stage 5 (ATD)    : Status contains "TRANSIT"/"DEPART" OR today >= ETD
  Stage 6 (ETA)    : ATA set OR today >= ETA
  Stage 7 (Done)   : Status ∈ {Delivered, Done, Released}
                    OR RELEASE_CONFIRMED (col 34) set

Output:
  col 32 (AF) TRACKING_STAGE = "N/7 <name>" e.g. "4/7 Gate-in"
  cell fill colored by stage bucket (green=done, blue=transit, amber=pre-ATD)

Usage:
    python ERP/jobs/shipment_tracker.py
    python ERP/jobs/shipment_tracker.py --dry-run
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime
from typing import Final

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "core"))
from ribbon_guard import save_preserving_ribbon  # noqa: E402
from active_jobs_cols import COL, HDR_ROW, DATA_START, STAGE_NAMES  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]

DEFAULT_ERP_FILE: Final = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"

AJ_HDR_ROW = HDR_ROW
AJ_DATA_START = DATA_START

# Color per stage bucket
STAGE_FILLS: Final = {
    1: PatternFill("solid", fgColor="FEF3C7"),  # amber — pending
    2: PatternFill("solid", fgColor="FEF3C7"),
    3: PatternFill("solid", fgColor="FEF3C7"),
    4: PatternFill("solid", fgColor="DBEAFE"),  # light blue — prep
    5: PatternFill("solid", fgColor="BFDBFE"),  # blue — in transit
    6: PatternFill("solid", fgColor="BFDBFE"),
    7: PatternFill("solid", fgColor="D1FAE5"),  # green — done
}


def _is_set(v) -> bool:
    if v is None or v == "":
        return False
    try:
        if isinstance(v, float) and v != v:  # NaN
            return False
    except TypeError:
        pass
    return True


def compute_stage(row: dict, now: datetime | None = None) -> int:
    """Return integer stage 1..7 based on row fields. Highest achieved stage wins."""
    now = now or datetime.now()
    status = str(row.get("Status") or "").upper()
    notes = str(row.get("Notes") or "").upper()

    # Stage 7: Done
    if any(k in status for k in ("DELIVERED", "DONE", "RELEASED", "COMPLETED")):
        return 7
    if _is_set(row.get("RELEASE_CONFIRMED")):
        return 7

    # Stage 6: ETA reached
    if _is_set(row.get("ATA")):
        return 6
    eta = row.get("ETA")
    if isinstance(eta, datetime) and now >= eta:
        return 6

    # Stage 5: Vessel departed
    if any(k in status for k in ("TRANSIT", "DEPART", "ATD", "SAILED")):
        return 5
    etd = row.get("ETD")
    if isinstance(etd, datetime) and now >= etd:
        return 5

    # Stage 4: Gate-in
    if "GATE" in status or "GATE" in notes or "GATED" in status:
        return 4
    cy_cut = row.get("CY_Cutoff")
    if isinstance(cy_cut, datetime) and now >= cy_cut:
        return 4

    # Stage 3: SI cut received
    if _is_set(row.get("SI_Received")):
        return 3

    # Stage 2: Confirmed — Bkg_No + Contract_Type
    if _is_set(row.get("Bkg_No")) and _is_set(row.get("Contract_Type")):
        return 2

    # Stage 1: Booked (default if any booking signal)
    if _is_set(row.get("Bkg_No")) or any(k in status for k in ("BOOKED", "PENDING")):
        return 1

    return 1  # safe default


def update_active_jobs(erp_file: str, dry_run: bool = False) -> dict:
    if not os.path.exists(erp_file):
        raise FileNotFoundError(erp_file)
    if not dry_run:
        try:
            with open(erp_file, "r+b"):
                pass
        except PermissionError:
            raise RuntimeError(f"ERP file locked — close Excel: {erp_file}")

    wb = openpyxl.load_workbook(erp_file, keep_vba=True)
    sheet = next((s for s in wb.sheetnames if "Active" in s), None)
    if not sheet:
        wb.close()
        raise RuntimeError("Active Jobs sheet not found")
    ws = wb[sheet]

    counts = {s: 0 for s in range(1, 8)}
    total = 0

    for r in range(AJ_DATA_START, ws.max_row + 1):
        crm = ws.cell(r, COL["CRM_ID"]).value
        if not crm:
            continue
        row = {k: ws.cell(r, c).value for k, c in COL.items()}
        stage = compute_stage(row)
        counts[stage] += 1
        total += 1

        label = f"{stage}/7 {STAGE_NAMES[stage]}"
        cell = ws.cell(r, COL["TRACKING_STAGE"], label)
        cell.font = Font(size=10, name="Segoe UI", bold=stage == 7)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = STAGE_FILLS[stage]

    if not dry_run:
        save_preserving_ribbon(wb, erp_file)
    wb.close()

    return {"total": total, "by_stage": counts}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--erp", default=DEFAULT_ERP_FILE)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    print(f"[+] Shipment Tracker @ {datetime.now():%Y-%m-%d %H:%M}")
    stats = update_active_jobs(args.erp, dry_run=args.dry_run)
    print(f"    -> {stats['total']} jobs tracked")
    for s in range(1, 8):
        n = stats["by_stage"][s]
        if n:
            print(f"       stage {s}/7 {STAGE_NAMES[s]:<8s}: {n}")
    if args.dry_run:
        print("    [DRY-RUN] no changes written")
    else:
        print(f"[OK] ERP saved: {args.erp}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
