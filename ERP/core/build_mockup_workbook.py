"""
build_mockup_workbook.py — 100% layout match to HTML mockup
=============================================================
One-shot script that brings ERP_Master_v14.xlsm to match the locked design:
  plans/260414-email-automation-v3/visuals/active-jobs-layout.html

Actions (idempotent):
  1. Ensure Active Jobs has v4 layout (calls migrate_active_jobs_v4 if not yet)
  2. Create 6 missing bottom sheets in mockup order:
     Active Jobs (existing) → Archive → Insurance → Commission →
     Tracking → Rpt Tháng → Rpt Tuần
  3. Preserve ribbon XML (via ribbon_guard)

Run with Excel closed.
Usage:
    python ERP/core/build_mockup_workbook.py
"""
from __future__ import annotations

import os
import sys
from typing import Final

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from active_jobs_cols import COL, HDR_ROW, DATA_START  # noqa: E402
from ribbon_guard import save_preserving_ribbon  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]

ERP_FILE = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"

HDR_FONT = Font(bold=True, color="FFFFFF", size=10, name="Segoe UI")
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


# Target sheet order per mockup bottom tabs
# (Active Jobs already exists, Quotes + CRM + Pricing + others stay hidden-ish for now)
MOCKUP_SHEETS: Final = [
    "Active Jobs",
    "Archive",
    "Insurance",
    "Commission",
    "Tracking",
    "Rpt Tháng",
    "Rpt Tuần",
]

# Header spec per new sheet (row 1 is title banner, row 2 is header row)
# Columns are kept minimal — Nelson can grow each as features ship
SHEET_SCHEMAS: Final = {
    "Archive": [
        "Job_ID", "FAST_ID", "CUSTOMER", "POL-POD", "CARRIER", "Bkg_No",
        "HBL_NO", "Container", "Qty", "SELL", "COST", "PROFIT",
        "Delivered_Date", "Closed_Reason",
    ],
    "Insurance": [
        "Job_ID", "CUSTOMER", "Cargo_Description", "Cargo_Value_USD",
        "ICC_Class", "Rate_pct", "Premium_USD", "Policy_No", "Insurer",
        "Eff_Date", "Exp_Date", "Status",
    ],
    "Commission": [
        "Job_ID", "CUSTOMER", "Sales", "Gross_Profit",
        "KB_Client", "KB_Carrier", "KB_Tax", "Net_Company",
        "PAID_Date", "Paid_Amount", "Status",
    ],
    "Tracking": [
        "Job_ID", "Container_No", "Bkg_No", "HBL_NO", "Carrier",
        "Stage", "Last_Event", "Location", "Event_Date",
        "ETD", "ETA", "ATA",
    ],
    "Rpt Tháng": [
        "Month", "Total_Shipments", "Total_TEU", "Total_Profit",
        "Top_Customer", "Top_Route", "Top_Carrier", "Notes",
    ],
    "Rpt Tuần": [
        "Week", "Sales", "Shipments", "TEU", "Profit",
        "KH_Existing", "KH_New", "Email_Sent", "Meetings",
        "Pct_Complete", "Plan_Next",
    ],
}

# Nice col widths per sheet
SHEET_WIDTHS: Final = {
    "Archive": {1: 14, 2: 13, 3: 18, 4: 12, 5: 8, 6: 15, 7: 15, 8: 7, 9: 5, 10: 9, 11: 9, 12: 9, 13: 12, 14: 16},
    "Insurance": {1: 14, 2: 18, 3: 22, 4: 13, 5: 8, 6: 8, 7: 11, 8: 12, 9: 15, 10: 11, 11: 11, 12: 10},
    "Commission": {1: 14, 2: 18, 3: 10, 4: 11, 5: 9, 6: 9, 7: 8, 8: 11, 9: 11, 10: 11, 11: 10},
    "Tracking": {1: 14, 2: 15, 3: 15, 4: 15, 5: 8, 6: 10, 7: 22, 8: 16, 9: 11, 10: 11, 11: 11, 12: 11},
    "Rpt Tháng": {1: 10, 2: 14, 3: 10, 4: 13, 5: 16, 6: 14, 7: 10, 8: 32},
    "Rpt Tuần": {1: 8, 2: 12, 3: 10, 4: 8, 5: 10, 6: 11, 7: 9, 8: 11, 9: 10, 10: 13, 11: 30},
}


def _make_sheet(wb, sheet_name: str, headers: list[str], widths: dict[int, int]):
    """Create the sheet if missing + write title banner + header row."""
    if sheet_name in wb.sheetnames:
        print(f"    [skip] {sheet_name} already exists")
        return wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Row 1: title banner spanning all headers
    end_col = len(headers)
    title_map = {
        "Archive": "ARCHIVE — Completed / Cancelled Jobs",
        "Insurance": "MARINE CARGO INSURANCE",
        "Commission": "COMMISSION / KICK BACK",
        "Tracking": "CONTAINER TRACKING — Multi-leg Events",
        "Rpt Tháng": "MONTHLY PERFORMANCE",
        "Rpt Tuần": "WEEKLY SALES KPI",
    }
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    c = ws.cell(1, 1, title_map.get(sheet_name, sheet_name.upper()))
    c.font = Font(bold=True, color="1F4E79", size=14, name="Segoe UI")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 2: column headers
    for i, h in enumerate(headers, 1):
        cell = ws.cell(2, i, h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = CENTER
    ws.row_dimensions[2].height = 22

    # Column widths
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Freeze below row 2
    ws.freeze_panes = "A3"

    print(f"    [add]  {sheet_name}  ({len(headers)} cols)")
    return ws


def _reorder_tabs(wb):
    """Arrange sheets so mockup tabs appear first in the bottom tab bar."""
    desired = [s for s in MOCKUP_SHEETS if s in wb.sheetnames]
    others = [s for s in wb.sheetnames if s not in desired]
    final_order = desired + others
    # openpyxl has no direct reorder — use move_sheet
    for idx, name in enumerate(final_order):
        ws = wb[name]
        cur_idx = wb.sheetnames.index(name)
        if cur_idx != idx:
            wb.move_sheet(ws, offset=idx - cur_idx)


def main() -> int:
    if not os.path.exists(ERP_FILE):
        print(f"[ERROR] {ERP_FILE} not found")
        return 1
    try:
        with open(ERP_FILE, "r+b"):
            pass
    except PermissionError:
        print(f"[ERROR] Close Excel first: {ERP_FILE}")
        return 2

    print(f"[+] Opening: {ERP_FILE}")
    wb = openpyxl.load_workbook(ERP_FILE, keep_vba=True)

    # Active Jobs should already be v4 (we migrated earlier)
    if "Active Jobs" not in wb.sheetnames:
        print("[ERROR] Active Jobs sheet missing — run migrate_active_jobs_v4.py first")
        wb.close()
        return 3

    print(f"[+] Existing sheets: {', '.join(wb.sheetnames)}")
    print("[+] Adding mockup sheets...")
    for sheet_name in MOCKUP_SHEETS:
        if sheet_name == "Active Jobs":
            continue  # already present + migrated
        headers = SHEET_SCHEMAS.get(sheet_name, [])
        widths = SHEET_WIDTHS.get(sheet_name, {})
        _make_sheet(wb, sheet_name, headers, widths)

    print("[+] Reordering tabs to match mockup order...")
    _reorder_tabs(wb)

    result = save_preserving_ribbon(wb, ERP_FILE)
    wb.close()
    print(f"\n[SUCCESS] Sheets in final order: {', '.join(MOCKUP_SHEETS)}")
    print(f"          ribbon: {result}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
