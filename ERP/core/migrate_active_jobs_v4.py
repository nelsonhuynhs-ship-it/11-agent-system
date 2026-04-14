"""
migrate_active_jobs_v4.py — one-shot Active Jobs v4 layout migration
=====================================================================
Reorders Active Jobs sheet columns to match the mockup
(plans/260414-email-automation-v3/visuals/active-jobs-layout.html):

  A MONTH | B FAST_ID | C Job_ID | D CUSTOMER | E POL-POD | F FINAL DEST |
  G CARRIER | H Bkg_No | I HBL_NO | J CONT | K QTY | L SERVICE | M ETD |
  N STATUS | O TRACKING | P SELL | Q COST | R PROFIT | S EMAIL
  T..AN = hidden preserved data + v4 extras

Also applies visual styling: frozen pane, conditional formatting on Status
(IN TRANSIT blue, DELIVERED green, PENDING yellow, ETA ALERT red),
Profit red/green, tracking dots rendered from TRACKING_STAGE raw text.

Idempotent: skips reorder if row 7 header already matches new layout.
Run with Excel closed.

Usage:
    python ERP/core/migrate_active_jobs_v4.py
    python ERP/core/migrate_active_jobs_v4.py --dry-run   # preview, no write
"""
from __future__ import annotations

import argparse
import os
import sys
from typing import Final

import openpyxl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from active_jobs_cols import (COL, DATA_START, HDR_ROW, LAST_VISIBLE_COL,
                              TOTAL_COLS, col_letter, derive_month,
                              derive_pol_pod, render_tracking_dots)
from ribbon_guard import save_preserving_ribbon  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]

ERP_FILE = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"


# Map from OLD col name → OLD col index (before migration)
# This is the v14 36-col layout present RIGHT NOW in the workbook
OLD_COL: Final = {
    "CRM_ID": 1, "Customer_Type": 2, "Routing": 3, "Bkg_No": 4,
    "ETD": 5, "ETA": 6, "ATA": 7, "Carrier": 8, "Contract_Type": 9,
    "Container_Type": 10, "Quantity": 11, "Selling_Rate": 12,
    "Buying_Rate": 13, "Profit": 14, "Profit_Margin": 15, "Status": 16,
    "SI_Received": 17, "CY_Cutoff": 18, "Door_Delivery": 19,
    "Door_Address": 20, "Door_Status": 21, "Delay_Count": 22,
    "Delay_Log": 23, "Notes": 24, "Created_Date": 25, "Last_Updated": 26,
    "Cost_Breakdown": 27, "Request_BKG": 28, "FAST_JOB_NO": 29, "HBL_NO": 30,
    "SERVICE": 31, "TRACKING_STAGE": 32, "RELEASE_EMAIL_SENT": 33,
    "RELEASE_CONFIRMED": 34, "PRICE_WATCH_STATUS": 35, "PRICE_WATCH_DELTA": 36,
}

# Mapping: NEW col index → (header, old_field_or_None, derive_fn_or_None)
MIGRATION_MAP: Final = [
    (1,  "MONTH",        None,                derive_month),        # derived from ETD
    (2,  "FAST_ID",      "FAST_JOB_NO",       None),
    (3,  "Job_ID",       None,                None),                # NEW — blank
    (4,  "CUSTOMER",     "CRM_ID",            None),
    (5,  "POL-POD",      None,                derive_pol_pod),      # derived from Routing
    (6,  "FINAL DEST",   "Door_Address",      None),
    (7,  "CARRIER",      "Carrier",           None),
    (8,  "Bkg_No",       "Bkg_No",            None),
    (9,  "HBL_NO",       "HBL_NO",            None),
    (10, "CONT",         "Container_Type",    None),
    (11, "QTY",          "Quantity",          None),
    (12, "SERVICE",      "SERVICE",           None),
    (13, "ETD",          "ETD",               None),
    (14, "STATUS",       "Status",            None),
    (15, "TRACKING",     None,                None),                # derived from TRACKING_STAGE
    (16, "SELL",         "Selling_Rate",      None),
    (17, "COST",         "Buying_Rate",       None),
    (18, "PROFIT",       "Profit",            None),
    (19, "EMAIL",        "Request_BKG",       None),                # hyperlink preserved separately
    (20, "Routing",          "Routing",           None),
    (21, "ETA",              "ETA",               None),
    (22, "ATA",              "ATA",               None),
    (23, "Contract_Type",    "Contract_Type",     None),
    (24, "Profit_Margin",    "Profit_Margin",     None),
    (25, "Customer_Type",    "Customer_Type",     None),
    (26, "SI_Received",      "SI_Received",       None),
    (27, "CY_Cutoff",        "CY_Cutoff",         None),
    (28, "Door_Delivery",    "Door_Delivery",     None),
    (29, "Door_Status",      "Door_Status",       None),
    (30, "Delay_Count",      "Delay_Count",       None),
    (31, "Delay_Log",        "Delay_Log",         None),
    (32, "Notes",            "Notes",             None),
    (33, "Created_Date",     "Created_Date",      None),
    (34, "Last_Updated",     "Last_Updated",      None),
    (35, "Cost_Breakdown",   "Cost_Breakdown",    None),
    (36, "TRACKING_STAGE",   "TRACKING_STAGE",    None),
    (37, "RELEASE_EMAIL_SENT",  "RELEASE_EMAIL_SENT",  None),
    (38, "RELEASE_CONFIRMED",   "RELEASE_CONFIRMED",   None),
    (39, "PRICE_WATCH_STATUS",  "PRICE_WATCH_STATUS",  None),
    (40, "PRICE_WATCH_DELTA",   "PRICE_WATCH_DELTA",   None),
]

# Target widths per visible col (pixels / Excel units)
COL_WIDTHS: Final = {
    1: 9,   2: 13,  3: 13,  4: 18,  5: 12,  6: 22,  7: 8,   8: 15,
    9: 15,  10: 7,  11: 5,  12: 10, 13: 11, 14: 13, 15: 14, 16: 9,
    17: 9,  18: 9,  19: 9,
}


def already_migrated(ws) -> bool:
    h1 = ws.cell(HDR_ROW, 1).value
    h2 = ws.cell(HDR_ROW, 2).value
    return (h1 == "MONTH") and (h2 == "FAST_ID")


def _capture_old_rows(ws) -> list[dict]:
    """Read all data rows into dicts keyed by OLD col name."""
    rows = []
    for r in range(DATA_START, ws.max_row + 1):
        if not ws.cell(r, OLD_COL["CRM_ID"]).value:
            continue
        row = {"_hyperlink_request_bkg": None}
        # capture hyperlink on Request_BKG cell (col 28)
        bkg_cell = ws.cell(r, OLD_COL["Request_BKG"])
        if bkg_cell.hyperlink:
            row["_hyperlink_request_bkg"] = bkg_cell.hyperlink.target
        for name, idx in OLD_COL.items():
            row[name] = ws.cell(r, idx).value
        rows.append(row)
    return rows


def _write_new_layout(ws, old_rows: list[dict]):
    # Clear everything below title row
    max_col_to_clear = max(ws.max_column, TOTAL_COLS)
    for r in range(HDR_ROW, ws.max_row + 1):
        for c in range(1, max_col_to_clear + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.hyperlink = None

    # Write new header row 7
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Segoe UI")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for new_idx, header, _, _ in MIGRATION_MAP:
        c = ws.cell(HDR_ROW, new_idx, header)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
    ws.row_dimensions[HDR_ROW].height = 26

    # Set visible col widths + hide hidden cols
    for idx in range(1, TOTAL_COLS + 1):
        letter = get_column_letter(idx)
        if idx <= LAST_VISIBLE_COL:
            ws.column_dimensions[letter].width = COL_WIDTHS.get(idx, 12)
            ws.column_dimensions[letter].hidden = False
        else:
            ws.column_dimensions[letter].hidden = True

    # Freeze pane: above row 8, left of col 4 (keep MONTH/FAST/JobID/CUSTOMER visible on scroll)
    ws.freeze_panes = "E8"

    # Write data rows
    for i, old in enumerate(old_rows):
        r = DATA_START + i
        for new_idx, header, old_field, derive in MIGRATION_MAP:
            value = None
            if derive is derive_month:
                value = derive_month(old.get("ETD"))
            elif derive is derive_pol_pod:
                value = derive_pol_pod(old.get("Routing"))
            elif old_field:
                value = old.get(old_field)

            # Special cases
            if new_idx == COL["TRACKING"]:
                value = render_tracking_dots(old.get("TRACKING_STAGE"))

            cell = ws.cell(r, new_idx, value)
            cell.font = Font(size=10, name="Segoe UI")
            cell.alignment = Alignment(vertical="center")

            # Formatting tweaks
            if new_idx in (COL["ETD"], COL["ETA"], COL["ATA"],
                           COL["RELEASE_EMAIL_SENT"], COL["RELEASE_CONFIRMED"]):
                cell.number_format = "dd/mm/yy"
            elif new_idx in (COL["Selling_Rate"], COL["Buying_Rate"], COL["Profit"]):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif new_idx == COL["Profit_Margin"]:
                cell.number_format = '0.0%'
            elif new_idx in (COL["Quantity"], COL["MONTH"]):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif new_idx == COL["TRACKING"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=11, name="Segoe UI", color="1F4E79")
            elif new_idx == COL["FAST_ID"]:
                cell.font = Font(size=10, name="Consolas", color="0563C1", underline="single")

            # Profit coloring
            if new_idx == COL["Profit"] and isinstance(value, (int, float)):
                if value > 0:
                    cell.font = Font(size=10, name="Segoe UI", bold=True, color="00804A")
                elif value < 0:
                    cell.font = Font(size=10, name="Segoe UI", bold=True, color="C00000")

            # Email hyperlink
            if new_idx == COL["Request_BKG"] and old.get("_hyperlink_request_bkg"):
                cell.value = "📧 Send"
                cell.hyperlink = old["_hyperlink_request_bkg"]
                cell.font = Font(color="0563C1", underline="single", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_conditional_formatting(ws, n_data_rows: int):
    """Status col (N=14) colored by value; Profit col (R=18) red/green."""
    if n_data_rows == 0:
        return
    last_row = DATA_START + n_data_rows - 1

    status_col = col_letter(COL["Status"])
    status_range = f"{status_col}{DATA_START}:{status_col}{last_row}"

    STATUS_RULES = [
        ("IN TRANSIT", "DBEAFE", "1E3A8A"),
        ("DELIVERED",  "D1FAE5", "064E3B"),
        ("PENDING",    "FEF3C7", "78350F"),
        ("BOOKED",     "FEF3C7", "78350F"),
        ("ETA ALERT",  "FECACA", "7F1D1D"),
        ("CANCELLED",  "E5E7EB", "374151"),
        ("DELAYED",    "FECACA", "7F1D1D"),
    ]
    for text, fill_hex, font_hex in STATUS_RULES:
        rule = FormulaRule(
            formula=[f'ISNUMBER(SEARCH("{text}",${status_col}{DATA_START}))'],
            fill=PatternFill("solid", fgColor=fill_hex),
            font=Font(bold=True, color=font_hex),
        )
        ws.conditional_formatting.add(status_range, rule)


def migrate(erp_file: str, dry_run: bool = False) -> dict:
    if not os.path.exists(erp_file):
        raise FileNotFoundError(erp_file)
    if not dry_run:
        try:
            with open(erp_file, "r+b"):
                pass
        except PermissionError:
            raise RuntimeError(f"ERP open in Excel — close first: {erp_file}")

    print(f"[+] Opening: {erp_file}")
    wb = openpyxl.load_workbook(erp_file, keep_vba=True)
    sheet = next((s for s in wb.sheetnames if "Active" in s and "Archive" not in s), None)
    if not sheet:
        wb.close()
        raise RuntimeError("Active Jobs sheet not found")
    ws = wb[sheet]
    print(f"    -> sheet: '{sheet}'  current max_row={ws.max_row}  max_col={ws.max_column}")

    if already_migrated(ws):
        print("[OK] Already migrated (row 7 header = MONTH | FAST_ID | ...). No changes.")
        wb.close()
        return {"skipped": True}

    old_rows = _capture_old_rows(ws)
    print(f"    -> captured {len(old_rows)} data rows from old layout")

    if dry_run:
        print("    [DRY-RUN] would reorder, apply formatting, hide cols 20+")
        for i, row in enumerate(old_rows[:3]):
            print(f"      sample old row {i+1}: CRM={row.get('CRM_ID')} "
                  f"Routing={row.get('Routing')} stage={row.get('TRACKING_STAGE')}")
        wb.close()
        return {"dry_run": True, "captured_rows": len(old_rows)}

    _write_new_layout(ws, old_rows)
    _apply_conditional_formatting(ws, n_data_rows=len(old_rows))

    save_preserving_ribbon(wb, erp_file)
    wb.close()
    print(f"\n[SUCCESS] migrated {len(old_rows)} rows to v4 layout + conditional formatting applied")
    print(f"          Visible cols A-S (19), hidden cols T-AN (21)")
    return {"migrated": True, "rows": len(old_rows)}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--erp", default=ERP_FILE)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    try:
        migrate(args.erp, dry_run=args.dry_run)
    except Exception as e:
        print(f"[ERROR] {e}")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
