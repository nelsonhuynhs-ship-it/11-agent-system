# ============================================================
#  BUILD ERP V13 - Quick Search + Ribbon Layout B
#  Row 1 A-I = dual header/search (placeholder labels)
#  Row 1 J-P = container type headers (fixed)
#  Row 2+ = Data
# ============================================================

import os, time, shutil, zipfile

# Script is in ERP/core/ — go up to Engine_test/ as project root
BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
VBA_DIR = os.path.join(BASE, "ERP", "vba")
PARQUET = os.path.join(BASE, "Pricing_Engine", "data", "Cleaned_Master_History.parquet")
PUC_FILE = os.path.join(BASE, "Pricing_Engine", "PUC_SOC.xlsx")

DATA_START_ROW = 2

# ===============================================================
# DATA LOADING - PIVOT from long format to wide
# ===============================================================
data_rows = []
MAX_ROWS = 2000

try:
    import pandas as pd
    from datetime import datetime, timedelta
    print(f"Loading data from {PARQUET}...")
    df = pd.read_parquet(PARQUET)
    print(f"  Raw Parquet: {len(df):,} rows")

    # Filter Total Ocean Freight only
    if 'Charge_Name' in df.columns:
        df_tof = df[df['Charge_Name'].str.contains('Total Ocean', case=False, na=False)].copy()
    else:
        df_tof = df.copy()
    print(f"  Total Ocean Freight: {len(df_tof):,} rows")

    # 90-day window filter
    if 'Eff' in df_tof.columns:
        df_tof['Eff'] = pd.to_datetime(df_tof['Eff'], errors='coerce')
        cutoff = datetime.now() - timedelta(days=90)
        df_recent = df_tof[df_tof['Eff'] >= cutoff].copy()
        if len(df_recent) < 100:
            df_recent = df_tof.copy()
            print(f"  90-day filter yielded <100 rows, using all dates")
        else:
            print(f"  After 90-day filter: {len(df_recent):,} rows")
    else:
        df_recent = df_tof.copy()

    # PIVOT: Container_Type x Amount -> columns
    id_cols = [c for c in ['POL','POD','Place','Carrier','Commodity','Eff','Exp','Note'] if c in df_recent.columns]
    rate_col = 'Amount' if 'Amount' in df_recent.columns else 'Rate'

    if 'Container_Type' in df_recent.columns and rate_col in df_recent.columns:
        pivot = df_recent.pivot_table(
            index=id_cols, columns='Container_Type', values=rate_col,
            aggfunc='first'
        ).reset_index()
        
        # Rename 45'HQ -> 45HQ
        if "45'HQ" in pivot.columns:
            pivot = pivot.rename(columns={"45'HQ": "45HQ"})

        # SMART DEDUP: keep only the latest rate per POL+POD+Place+Carrier
        route_key = [c for c in ['POL','POD','Place','Carrier'] if c in pivot.columns]
        if 'Eff' in pivot.columns and route_key:
            pivot = pivot.sort_values('Eff', ascending=False)
            pivot = pivot.drop_duplicates(subset=route_key, keep='first')
            print(f"  After dedup (latest per route+carrier): {len(pivot):,} unique routes")

        # Sort by Eff desc
        if 'Eff' in pivot.columns:
            pivot = pivot.sort_values('Eff', ascending=False)

        # Build data rows (capped at MAX_ROWS)
        col_map = {
            'POL': 1, 'POD': 2, 'Place': 3, 'Carrier': 4, 'Commodity': 5,
            'Eff': 6, 'Exp': 7, 'Note': 8,
            '20GP': 10, '40GP': 11, '40HQ': 12, '45HQ': 13,
            '40NOR': 14, '20RF': 15, '40RF': 16,
        }

        for _, row in pivot.head(MAX_ROWS).iterrows():
            r = [None] * 16
            for col_name, col_idx in col_map.items():
                val = row.get(col_name)
                if pd.notna(val):
                    if col_idx >= 10:
                        try:
                            r[col_idx - 1] = int(float(val))
                        except:
                            pass
                    else:
                        r[col_idx - 1] = str(val)
            data_rows.append(r)

        # Report carrier coverage
        if 'Carrier' in pivot.columns:
            loaded_carriers = pivot.head(MAX_ROWS)['Carrier'].unique()
            print(f"  Carriers loaded: {sorted(loaded_carriers)} ({len(loaded_carriers)} carriers)")
        print(f"  Pivoted: {len(pivot):,} routes -> loaded {len(data_rows)} rows (max {MAX_ROWS})")

except ImportError:
    print("  pandas/pyarrow not installed - using sample data")
except Exception as e:
    print(f"  Data load error: {e}")

# ===============================================================
# STEP 1: Create xlsx layout with openpyxl
# ===============================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors - match V10
C_SEARCH = PatternFill("solid", fgColor="FFF7ED")    # orange tint for A-I header/search
C_DRY    = PatternFill("solid", fgColor="1E40AF")     # blue for 20GP/40GP/40HQ
C_MISC   = PatternFill("solid", fgColor="6B7280")     # gray for 45HQ/40NOR
C_RF     = PatternFill("solid", fgColor="7C3AED")     # purple for 20RF/40RF

thin = Side(style='thin', color='CBD5E1')
card = Border(left=thin, right=thin, top=thin, bottom=thin)

# --- Sheet 1: Pricing Dashboard ---
ws1 = wb.active
ws1.title = "Pricing Dashboard"

# Row 1: Dual header/search cells (A-I) + Container headers (J-P)
ws1.row_dimensions[1].height = 30

# A1-I1: Search/header cells (placeholder labels, dimmed gray text, orange bg)
search_labels = ['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note', 'Source']
for i, lbl in enumerate(search_labels, 1):
    cell = ws1.cell(1, i, value=lbl)
    cell.font = Font(name='Segoe UI', size=9, color="B0B0B0", italic=True)  # dimmed placeholder style
    cell.fill = C_SEARCH
    cell.border = card
    cell.alignment = Alignment(horizontal='center', vertical='center')

# J1-P1: Container type headers (bold white, colored bg)
cont_colors = {
    '20GP': C_DRY, '40GP': C_DRY, '40HQ': C_DRY,
    '45HQ': C_MISC, '40NOR': C_MISC,
    '20RF': C_RF, '40RF': C_RF,
}
cont_names = ['20GP', '40GP', '40HQ', '45HQ', '40NOR', '20RF', '40RF']
for i, cont in enumerate(cont_names):
    cell = ws1.cell(1, 10 + i, value=cont)
    cell.font = Font(name='Segoe UI', size=9, bold=True, color="FFFFFF")
    cell.fill = cont_colors.get(cont, C_DRY)
    cell.border = card
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Column widths (match V10)
widths = {'A': 8, 'B': 12, 'C': 20, 'D': 10, 'E': 22, 'F': 7, 'G': 7, 'H': 10, 'I': 10}
for col_letter, w in widths.items():
    ws1.column_dimensions[col_letter].width = w
for col in range(10, 17):
    ws1.column_dimensions[get_column_letter(col)].width = 10

# Data rows (Row 2+)
if data_rows:
    for ri, row_data in enumerate(data_rows):
        for ci, val in enumerate(row_data):
            if val is not None:
                cell = ws1.cell(DATA_START_ROW + ri, ci + 1)
                cell.value = val
                cell.font = Font(name='Segoe UI', size=10)
                cell.border = card
                if ci >= 9:  # Price columns
                    cell.number_format = '#,##0'
else:
    # Sample data if no Parquet
    sample = [
        ["DAD", "BALTIMORE, MD", "BALTIMORE, MD", "CMA", "", "2024-02", "2024-02", "via YANTIAN", None, 6192, 7652, 7652, None, None, None, None],
        ["DAD", "BALTIMORE, MD", "BALTIMORE, MD", "ONE", "", "2024-03", "2024-03", "SOC", None, 4672, 5752, 5752, None, None, None, None],
        ["HPH", "LOS ANGELES, CA", "LOS ANGELES, CA", "YML", "", "2026-03", "2026-03", "SOC DIRECT", None, 1800, 5200, 5200, None, None, None, None],
        ["HPH", "TACOMA, WA", "DENVER, CO", "CMA", "Stone", "2026-03", "2026-03", "DIRECT", None, 2100, 5600, 5700, 6800, None, None, None],
        ["HCM", "LOS ANGELES, CA", "LOS ANGELES, CA", "MSK", "Frozen", "2026-02", "2026-03", "REEFER", None, None, None, None, None, None, 3500, 5500],
    ]
    for ri, row in enumerate(sample):
        for ci, val in enumerate(row):
            if val is not None:
                cell = ws1.cell(DATA_START_ROW + ri, ci + 1)
                cell.value = val
                cell.font = Font(name='Segoe UI', size=10)
                cell.border = card
                if ci >= 9:
                    cell.number_format = '#,##0'
    print("  Added 5 sample rows for testing")

ws1.freeze_panes = "A2"  # Freeze Row 1 only

# --- Sheet 2: Quotes (42 columns — V13 layout) ---
ws2 = wb.create_sheet("Quotes")
q_hdr = ["QuoteID", "Date", "Customer", "Carrier", "POL", "POD",
         "Place", "Via", "Eff", "Exp", "Source",
         "Buy_20GP", "Buy_40GP", "Buy_40HC", "Buy_45HC", "Buy_40NOR", "Buy_20RF", "Buy_40RF",
         "Mar_20GP", "Mar_40GP", "Mar_40HC", "Mar_45HC", "Mar_40NOR", "Mar_20RF", "Mar_40RF",
         "PUC_20", "PUC_40", "PUC_40HC",
         "Sell_20GP", "Sell_40GP", "Sell_40HC", "Sell_45HC", "Sell_40NOR", "Sell_20RF", "Sell_40RF",
         "Status", "Remark", "StatusDate", "Qty", "Volume", "JobID", "ContType"]
for i, h in enumerate(q_hdr, 1):
    ws2.cell(1, i, value=h).font = Font(bold=True, size=9)
ws2.freeze_panes = "A2"

# --- Sheet 3: CRM (Customer SOPs — 43 columns, row 1 = header) ---
from openpyxl.worksheet.datavalidation import DataValidation

ws_crm = wb.create_sheet("CRM")
crm_headers = [
    'CRM_ID', 'Customer_Name', 'Customer_Type', 'Sales_Owner', 'FAS_ID', 'Status',
    'Contact1_Name', 'Contact1_Email', 'Contact1_Phone',
    'Contact2_Name', 'Contact2_Email', 'Contact2_Phone',
    'Preferred_Carriers', 'POL_Options', 'POD_Options', 'Container_Types',
    'Commodity', 'HS_Code', 'GW_Per_Container', 'Stuffing_Place',
    'MT_Pickup_ICD', 'Full_Return_Port',
    'Is_Reefer', 'Reefer_Temp', 'Reefer_Ventilation', 'Reefer_Humidity',
    'Switch_Bill', 'ISF_Filer', 'Shipper_Fix_or_Byshipment', 'Consignee_Fix_or_Byshipment',
    'Payer', 'Credit_Term', 'Invoice_Trigger', 'Debit_Note_Trigger',
    'BL_Fee', 'THC_20', 'THC_40', 'ENS_AMS', 'Seal_Fee', 'Telex_Release',
    'HDL_Fee_Note', 'Carrier_KB_Pct', 'Special_Notes',
]

crm_hdr_font = Font(name='Segoe UI', size=10, bold=True)
crm_hdr_fill = PatternFill('solid', fgColor='DAEEF3')
crm_hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
for i, h in enumerate(crm_headers, 1):
    cell = ws_crm.cell(1, i)
    cell.value = h
    cell.font = crm_hdr_font
    cell.fill = crm_hdr_fill
    cell.alignment = crm_hdr_align
    cell.border = card

# CRM Data Validation
dv_custtype = DataValidation(type='list', formula1='"BCO,Forwarder,Retailer,Enterprise"', allow_blank=True)
dv_custtype.error = 'Choose: BCO, Forwarder, Retailer, Enterprise'
ws_crm.add_data_validation(dv_custtype)
dv_custtype.add('C2:C1000')

dv_status = DataValidation(type='list', formula1='"Active,Inactive"', allow_blank=True)
ws_crm.add_data_validation(dv_status)
dv_status.add('F2:F1000')

dv_reefer = DataValidation(type='list', formula1='"Yes,No"', allow_blank=True)
ws_crm.add_data_validation(dv_reefer)
dv_reefer.add('W2:W1000')

# Pre-fill Row 2: NAFOODS GROUP
nafoods_row2 = {
    1: 'CS001289', 2: 'CÔNG TY CP NAFOODS GROUP', 3: 'BCO', 4: 'NELSON', 6: 'Active',
    7: 'Ms. Thơm', 8: 'thomvt@nafoods.com', 9: '0974 58 7022',
    10: 'Ms. Thúy', 11: 'thuypnn@nafoods.com', 12: '0987729960',
    13: 'COSCO/YML/ONE', 14: 'HCM/HPH',
    15: 'TCM/LAX/OAK/NYC/SAV/HOU/BOSTON/VAN/TORONTO',
    16: '40RF/20RF', 17: 'FROZEN PASSIONFRUIT IQF', 18: '20098999',
    19: '25TONS', 20: 'WAREHOUSE', 21: 'AS RO', 22: 'CAI MEP',
    23: 'Yes', 24: '-23°C', 25: 'CLOSED', 26: 'NO',
    27: 'No request', 28: 'Broker Cnee file',
    31: 'Shipper pay LCC, O/F', 32: 'Net 25 days from invoice date',
    33: 'Sailing date, that day exchange rate',
    34: 'After customer confirms HBL',
    35: 40, 36: 200, 37: 260, 38: 40, 39: 10, 40: 35,
}
for col, val in nafoods_row2.items():
    ws_crm.cell(2, col, value=val).font = Font(name='Segoe UI', size=10)

# Pre-fill Row 3: NAFOODS MIEN NAM (same as row 2 except CRM_ID and Customer_Name)
for col, val in nafoods_row2.items():
    ws_crm.cell(3, col, value=val).font = Font(name='Segoe UI', size=10)
ws_crm.cell(3, 1, value='CS001156')
ws_crm.cell(3, 2, value='CÔNG TY CỔ PHẦN NAFOODS MIỀN NAM')
ws_crm.cell(3, 5, value='CS001156')  # FAS_ID

# CRM column widths: auto-fit approximation
crm_col_widths = {
    'A': 12, 'B': 35, 'C': 14, 'D': 12, 'E': 12, 'F': 10,
    'G': 15, 'H': 25, 'I': 16,
    'J': 15, 'K': 25, 'L': 16,
    'M': 20, 'N': 12, 'O': 40, 'P': 14,
    'Q': 30, 'R': 12, 'S': 18, 'T': 14,
    'U': 22, 'V': 16,
    'W': 10, 'X': 12, 'Y': 16, 'Z': 14,
    'AA': 12, 'AB': 22, 'AC': 24, 'AD': 24,
    'AE': 24, 'AF': 30, 'AG': 34, 'AH': 28,
    'AI': 8, 'AJ': 8, 'AK': 8, 'AL': 10, 'AM': 10, 'AN': 14,
    'AO': 22, 'AP': 14, 'AQ': 30,
}
for col_letter, w in crm_col_widths.items():
    ws_crm.column_dimensions[col_letter].width = w

ws_crm.freeze_panes = 'A2'
ws_crm.sheet_properties.tabColor = '00B050'  # Green tab
print("  CRM sheet: 43 columns, 2 NAFOODS rows pre-filled, green tab")

# --- Sheet 4: Active Jobs (31 headers in row 7, data row 8+) ---
ws3 = wb.create_sheet("Active Jobs")
ws3.merge_cells('A1:AE1')
title = ws3.cell(1, 1)
title.value = 'ACTIVE JOBS'
title.font = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
title.fill = PatternFill('solid', fgColor='294B93')
title.alignment = Alignment(horizontal='center', vertical='center')

aj_headers = [
    'CRM_ID', 'Customer_Type', 'Routing', 'Bkg_No', 'ETD', 'ETA', 'ATA',
    'Carrier', 'Contract_Type', 'Container_Type', 'Quantity',
    'Selling_Rate', 'Buying_Rate', 'Profit', 'Profit_Margin',
    'Status', 'SI_Received', 'CY_Cutoff',
    'Door_Delivery', 'Door_Address', 'Door_Status',
    'Delay_Count', 'Delay_Log', 'Notes',
    'Created_Date', 'Last_Updated',
    'Cost_Breakdown', 'Request_BKG',
    'FAST_JOB_NO', 'FAST_REF', 'HBL_NO',
]
hdr_font_aj = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
hdr_fill_aj = PatternFill('solid', fgColor='1F4E79')
hdr_align_aj = Alignment(horizontal='center', vertical='center', wrap_text=True)
for i, h in enumerate(aj_headers, 1):
    cell = ws3.cell(7, i)
    cell.value = h
    cell.font = hdr_font_aj
    cell.fill = hdr_fill_aj
    cell.alignment = hdr_align_aj
    cell.border = card

# Active Jobs data validations
dv_aj_custtype = DataValidation(type='list', formula1='"BCO,Forwarder,Retailer,Enterprise"', allow_blank=True)
ws3.add_data_validation(dv_aj_custtype)
dv_aj_custtype.add('B8:B5000')

dv_aj_status = DataValidation(type='list', formula1='"Booked,Confirmed,Sailing,Arrived,Done"', allow_blank=True)
ws3.add_data_validation(dv_aj_status)
dv_aj_status.add('P8:P5000')

dv_aj_door = DataValidation(type='list', formula1='"Yes,No"', allow_blank=True)
ws3.add_data_validation(dv_aj_door)
dv_aj_door.add('S8:S5000')

# Active Jobs column widths
aj_col_widths = {
    'A': 12, 'B': 14, 'C': 28, 'D': 16, 'E': 12, 'F': 12, 'G': 12,
    'H': 10, 'I': 14, 'J': 14, 'K': 10,
    'L': 12, 'M': 12, 'N': 12, 'O': 14,
    'P': 12, 'Q': 12, 'R': 12,
    'S': 14, 'T': 20, 'U': 14,
    'V': 12, 'W': 20, 'X': 20,
    'Y': 16, 'Z': 16,
    'AA': 30, 'AB': 14, 'AC': 14, 'AD': 14, 'AE': 14,
}
for col_letter, w in aj_col_widths.items():
    ws3.column_dimensions[col_letter].width = w

ws3.freeze_panes = 'A8'

# --- Sheet 4: Markup_Store (hidden) ---
ws4 = wb.create_sheet("Markup_Store")
mk_headers = ["Carrier", "Mar_20GP", "Mar_40GP", "Mar_40HC", "Mar_45HC", "Mar_40NOR", "Mar_20RF", "Mar_40RF"]
for ci, h in enumerate(mk_headers, 1):
    ws4.cell(1, ci, value=h).font = Font(bold=True)
carriers = ["CMA", "COSCO", "EMC", "HMM", "HPL", "KMTC", "MSC", "MSK", "ONE",
            "OOCL", "PIL", "SEALEAD", "TSL", "UWL", "WHL", "YML", "ZIM",
            "ESL", "MCK", "APL"]
for ri, c in enumerate(carriers, 2):
    ws4.cell(ri, 1, value=c)

# --- Sheet 5: PUC_Lookup (hidden) ---
ws5 = wb.create_sheet("PUC_Lookup")
for ci, h in enumerate(["Place", "PUC_20", "PUC_40", "PUC_40HC"], 1):
    ws5.cell(1, ci, value=h).font = Font(bold=True)
try:
    import pandas as pd
    if os.path.exists(PUC_FILE):
        dfp = pd.read_excel(PUC_FILE)
        for ri, (_, row) in enumerate(dfp.iterrows(), 2):
            for ci, col in enumerate(dfp.columns[:4], 1):
                val = row[col]
                if pd.notna(val):
                    ws5.cell(ri, ci, value=int(val) if isinstance(val, (int, float)) else str(val))
        print(f"  Loaded {len(dfp)} PUC entries")
except:
    pass

# --- Sheet 6: InvoiceLog (hidden) ---
wb.create_sheet("InvoiceLog")

# ===============================================================
# SAVE XLSX
# ===============================================================
tmp_xlsx = os.path.join(BASE, "ERP_v13_temp.xlsx")
final_xlsm = os.path.join(BASE, "ERP", "data", "ERP_V13_STAGING.xlsm")
wb.save(tmp_xlsx)
print(f"\nStep 1: Layout saved ({len(data_rows)} data rows, Row 1 = search/headers, Row 2+ = data)")

# ===============================================================
# STEP 2: Convert .xlsx to .xlsm via COM
# ===============================================================
print("Step 2: Converting to .xlsm via COM...")
try:
    import win32com.client
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False; xl.DisplayAlerts = False
    wbCom = xl.Workbooks.Open(os.path.abspath(tmp_xlsx))
    if os.path.exists(final_xlsm):
        os.remove(final_xlsm)
    wbCom.SaveAs(os.path.abspath(final_xlsm), FileFormat=52)
    for sn in ["Markup_Store", "PUC_Lookup", "InvoiceLog"]:
        try:
            wbCom.Sheets(sn).Visible = 0
        except:
            pass
    wbCom.Save()
    wbCom.Close()
    xl.Quit(); xl = None
    time.sleep(1)
    print(f"  Converted to {final_xlsm}")
except Exception as e:
    print(f"  COM conversion failed: {e}")
    shutil.copy2(tmp_xlsx, final_xlsm)

# ===============================================================
# STEP 3: Import VBA modules via COM + Save
# ===============================================================
print("Step 3: Injecting VBA modules...")
try:
    import win32com.client
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False; xl.DisplayAlerts = False
    wbCom = xl.Workbooks.Open(os.path.abspath(final_xlsm))

    # 3a. Import VBA modules
    vba_modules = [
        "QuoteBuilder_ERP.bas",
        "CostBreakdown.bas",
        "BookingEmail.bas",
        "MonthlyReport.bas",
        "CRM_Sheet.bas",
    ]
    for vba_file in vba_modules:
        vba_path = os.path.join(VBA_DIR, vba_file)
        if os.path.exists(vba_path):
            wbCom.VBProject.VBComponents.Import(vba_path)
            print(f"  Imported {vba_file}")
        else:
            print(f"  WARN: {vba_file} not found, skipped")

    # 3b. Sheet1 event handler
    #   - Worksheet_Change: Quick Search (Row 1, col A-I)
    #   - Worksheet_SelectionChange: Ribbon row load (Row 2+)
    sheet_handler = (
        "Private Sub Worksheet_Change(ByVal Target As Range)\n"
        "    ' Quick Search: Row 1, columns A-I\n"
        "    If Target.Row = 1 And Target.Column >= 1 And Target.Column <= 9 Then\n"
        "        On Error Resume Next\n"
        "        Call QuoteBuilder.HandleSearchChange(Target)\n"
        "        On Error GoTo 0\n"
        "    End If\n"
        "End Sub\n"
        "\n"
        "Private Sub Worksheet_SelectionChange(ByVal Target As Range)\n"
        "    ' Ribbon: load clicked row (Row 2+)\n"
        "    If Target.Cells.Count > 1 Then Exit Sub\n"
        "    If Target.Row < 2 Then Exit Sub\n"
        "    On Error Resume Next\n"
        "    Call QuoteBuilder.LoadRowToRibbon(Target.Row)\n"
        "    If Not QuoteBuilder.ribbonUI Is Nothing Then\n"
        "        QuoteBuilder.ribbonUI.Invalidate\n"
        "    End If\n"
        "    On Error GoTo 0\n"
        "End Sub\n"
    )

    handler_added = False
    try:
        sh1 = wbCom.VBProject.VBComponents("Sheet1")
        cm = sh1.CodeModule
        if cm.CountOfLines > 0:
            cm.DeleteLines(1, cm.CountOfLines)
        cm.AddFromString(sheet_handler)
        print(f"  Sheet1 handler: {cm.CountOfLines} lines (Change for search + SelectionChange for ribbon)")
        handler_added = True
    except Exception as e:
        print(f"  Sheet1 direct: {e}")

    if not handler_added:
        for i in range(1, wbCom.VBProject.VBComponents.Count + 1):
            comp = wbCom.VBProject.VBComponents.Item(i)
            if comp.Type == 100 and comp.Name.startswith("Sheet"):
                try:
                    cm = comp.CodeModule
                    if cm.CountOfLines == 0:
                        cm.AddFromString(sheet_handler)
                        print(f"  Handler added to {comp.Name}")
                        handler_added = True
                        break
                except:
                    pass

    # 3c. ThisWorkbook — auto-expire check on open
    wb_code = (
        "Private Sub Workbook_Open()\n"
        "    On Error Resume Next\n"
        "    Application.Wait Now + TimeSerial(0, 0, 1)\n"
        "    ' Auto-expire quotes past Exp date\n"
        "    Call QuoteBuilder.CheckAutoExpired\n"
        "    Debug.Print \"[Startup \" & Now() & \"] Auto-expired check complete\"\n"
        "    On Error GoTo 0\n"
        "End Sub\n"
    )
    try:
        twb = wbCom.VBProject.VBComponents("ThisWorkbook")
        cm2 = twb.CodeModule
        if cm2.CountOfLines > 0:
            cm2.DeleteLines(1, cm2.CountOfLines)
        cm2.AddFromString(wb_code)
        print("  ThisWorkbook_Open added")
    except Exception as e:
        print(f"  ThisWorkbook: {e}")

    wbCom.Save()
    wbCom.Close()
    xl.Quit(); xl = None
    time.sleep(2)
    print("  VBA injection complete!")
except Exception as e:
    print(f"  VBA error: {e}")

# ===============================================================
# STEP 4: Inject CustomUI14 XML (LAST STEP!)
# ===============================================================
print("Step 4: Injecting CustomUI14 XML (LAST step)...")

customui_path = os.path.join(VBA_DIR, "CustomUI_ERP.xml")

# Use shared utility for reliable 3-part injection
from customui_utils import ensure_customui
result = ensure_customui(final_xlsm, customui_xml_path=customui_path)
if result.get("injected"):
    print("  CustomUI14 injected (M365 compatible)")
elif result.get("already_ok"):
    print("  CustomUI14 already present")
else:
    print(f"  CustomUI14 error: {result}")

# ===============================================================
# CLEANUP + VERIFY
# ===============================================================
if os.path.exists(tmp_xlsx):
    os.remove(tmp_xlsx)

print("\nVerifying...")
with zipfile.ZipFile(final_xlsm, 'r') as z:
    checks = {
        'customUI14.xml exists': 'customUI/customUI14.xml' in z.namelist(),
        'Content_Types entry': 'customUI14' in z.read('[Content_Types].xml').decode(),
        'Rels entry': 'customUI14' in z.read('_rels/.rels').decode(),
        'onLoad callback': 'onLoad' in z.read('customUI/customUI14.xml').decode(),
    }
    for k, v in checks.items():
        print(f"  {k}: {'OK' if v else 'FAIL'}")

print(f"\n{'='*60}")
print(f"  ERP V13 - QUICK SEARCH + RIBBON - BUILD COMPLETE!")
print(f"  File: {final_xlsm}")
print(f"  Data rows: {len(data_rows)} | Row 1 = search | Row 2+ = data")
print(f"{'='*60}")
print(f"\nTest:")
print(f"  1. Open > Enable Macros > Operations tab visible")
print(f"  2. Type 'DAD' in A1 (replaces 'POL' placeholder) -> data filters")
print(f"  3. Clear A1 -> 'POL' placeholder restored, all rows visible")
print(f"  4. Click data row > Operations tab > Buy Rate shows")
print(f"  5. Enter Margin > Sell Rate calculates")
print(f"  6. Customer > QUOTE > Quotes sheet")
