"""
seed_test_jobs.py — insert/remove 5 test Active Jobs for button demo
=====================================================================
Populates ERP_Master_v14.xlsm "Active Jobs" with realistic test rows that
cover every pipeline stage (1/7 → 7/7) so Nelson can click each ribbon
button and see real output.

Test rows inserted starting at row 8:
  r=8  NAFOODS          stage 1 (Booked only)
  r=9  VIFON EXPORT     stage 5 (In Transit, ATD passed)
  r=10 SIRI (40RF)      stage 2 (Confirmed) — for Reefer Plug
  r=11 TRAN ANH         URGENT release alert (email sent 3h ago, ETA tomorrow)
  r=12 WESTFOOD         stage 7 (Delivered) — for Monthly Report (March)

Usage:
    python ERP/core/seed_test_jobs.py          # insert 5 rows
    python ERP/core/seed_test_jobs.py --clear  # remove rows 8-12

Safe: only touches rows 8-12. Preserves ribbon via ribbon_guard.
"""
from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime, timedelta

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ribbon_guard import save_preserving_ribbon  # noqa: E402
from active_jobs_cols import COL  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]

ERP_FILE = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"

TEST_ROW_START = 8
TEST_ROW_END = 12  # inclusive


def _build_seed_rows(now: datetime | None = None) -> list[dict]:
    now = now or datetime.now()
    return [
        # r=8: NAFOODS stage 1 — just booked, ETD future, ETA empty (Transit Time will fill)
        dict(CRM_ID="NAFOODS", Customer_Type="Direct",
             Routing="HPH-LOS ANGELES, CA VIA USLGB", Bkg_No="YM-8834501",
             ETD=now + timedelta(days=3), ETA=None, ATA=None,
             Carrier="YML", Contract_Type="SHA0005N25",
             Container_Type="40HQ", Quantity=2,
             Selling_Rate=3200, Buying_Rate=2650, Profit=1100, Profit_Margin=0.172,
             Status="Booked",
             Door_Delivery="Yes", Door_Address="LOS ANGELES, CA",
             Notes="", Created_Date=now, Last_Updated=now,
             FAST_ID="se2604/266",  # lowercase + short seq — F4 FAST ID will fix
             HBL_NO="PELP26040260"),

        # r=9: VIFON stage 5 — in transit, ETD passed
        dict(CRM_ID="VIFON EXPORT", Customer_Type="Direct",
             Routing="HCM-USNYC", Bkg_No="CMA-V-2001",
             ETD=now - timedelta(days=20), ETA=None, ATA=None,
             Carrier="CMA", Contract_Type="CMA-2026-FIX",
             Container_Type="40GP", Quantity=1,
             Selling_Rate=2857, Buying_Rate=2789, Profit=68, Profit_Margin=0.024,
             Status="In Transit",
             Door_Delivery="No", Door_Address="",
             Notes="", Created_Date=now - timedelta(days=22), Last_Updated=now,
             FAST_ID="SE2604/0280",
             HBL_NO="PNYC26040596"),

        # r=10: SIRI 40RF stage 2 — confirmed, for Reefer Plug demo
        dict(CRM_ID="SIRI", Customer_Type="Direct",
             Routing="HPH-CHICAGO, IL VIA USLAX", Bkg_No="ONE-BKG-301",
             ETD=now + timedelta(days=5), ETA=now + timedelta(days=25), ATA=None,
             Carrier="ONE", Contract_Type="SHA0007N26 SOC",
             Container_Type="40RF", Quantity=1,
             Selling_Rate=5785, Buying_Rate=5655, Profit=130, Profit_Margin=0.022,
             Status="Booked",
             Door_Delivery="Yes", Door_Address="CHICAGO, IL",
             Notes="", Created_Date=now - timedelta(days=1), Last_Updated=now,
             FAST_ID="SE2604/0310",
             HBL_NO="PCHI26040310"),

        # r=11: TRAN ANH URGENT — release email 3h ago, ETA tomorrow, no confirm
        dict(CRM_ID="TRAN ANH", Customer_Type="Direct",
             Routing="HCM-USLGB", Bkg_No="YM-8840200",
             ETD=now - timedelta(days=22), ETA=now + timedelta(days=1), ATA=None,
             Carrier="YML", Contract_Type="YML-FIX",
             Container_Type="20GP", Quantity=1,
             Selling_Rate=1800, Buying_Rate=1550, Profit=250, Profit_Margin=0.139,
             Status="In Transit",
             Door_Delivery="No", Door_Address="",
             Notes="", Created_Date=now - timedelta(days=25), Last_Updated=now,
             FAST_ID="SE2604/0400",
             HBL_NO="PLGB26040400",
             RELEASE_EMAIL_SENT=now - timedelta(hours=3)),

        # r=12: WESTFOOD stage 7 — delivered (for Monthly/Weekly report)
        dict(CRM_ID="WESTFOOD", Customer_Type="Direct",
             Routing="HCM-USNYC", Bkg_No="WAN-3301",
             ETD=now - timedelta(days=30), ETA=now - timedelta(days=2),
             ATA=now - timedelta(days=1),
             Carrier="WAN HAI", Contract_Type="WHL-FIX",
             Container_Type="40HQ", Quantity=1,
             Selling_Rate=3320, Buying_Rate=3120, Profit=200, Profit_Margin=0.060,
             Status="Delivered",
             Door_Delivery="No", Door_Address="",
             Notes="", Created_Date=now - timedelta(days=33), Last_Updated=now,
             FAST_ID="SE2604/0647",
             HBL_NO="PNYC26040596"),
    ]


def _write_rows(ws, rows: list[dict]):
    r = TEST_ROW_START
    for row in rows:
        for key, col in COL.items():
            if key in row and row[key] is not None:
                ws.cell(r, col, row[key])
        r += 1


def _clear_rows(ws):
    for r in range(TEST_ROW_START, TEST_ROW_END + 1):
        for col in range(1, 40):
            ws.cell(r, col).value = None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--clear", action="store_true", help="Remove test rows 8-12")
    ap.add_argument("--erp", default=ERP_FILE)
    args = ap.parse_args()

    if not os.path.exists(args.erp):
        print(f"[ERROR] ERP not found: {args.erp}")
        return 1
    try:
        with open(args.erp, "r+b"):
            pass
    except PermissionError:
        print(f"[ERROR] ERP file is open in Excel. Close Excel and retry.")
        return 2

    wb = openpyxl.load_workbook(args.erp, keep_vba=True)
    ws = wb["Active Jobs"]

    if args.clear:
        _clear_rows(ws)
        action = "cleared"
    else:
        rows = _build_seed_rows()
        _write_rows(ws, rows)
        action = "seeded"

    save_preserving_ribbon(wb, args.erp)
    wb.close()
    print(f"[OK] {action} rows {TEST_ROW_START}-{TEST_ROW_END} in Active Jobs")

    if not args.clear:
        print("\nTest rows inserted:")
        print("  r=8  NAFOODS      stage 1 (Booked)       — Transit Time fills ETA, FAST ID fixes 'se2604/266'")
        print("  r=9  VIFON EXPORT stage 5 (In Transit)   — Price Watch, Tracking")
        print("  r=10 SIRI 40RF    stage 2 (Confirmed)    — Reefer Plug optimal drop")
        print("  r=11 TRAN ANH     URGENT release alert   — Release Alerts will flag")
        print("  r=12 WESTFOOD     stage 7 (Delivered)    — Monthly Report picks up")
        print("\nNow click any button in Operations tab → Active Jobs v4.")
        print("To remove when done: python ERP/core/seed_test_jobs.py --clear")
    return 0


if __name__ == "__main__":
    sys.exit(main())
