# -*- coding: utf-8 -*-
"""
refp/refresh.py — Unified ERP Refresh (merged from 3 scripts)
=================================================================
Merged from:
  - refresh_erp_from_parquet.py (base — Parquet pipeline + normalization)
  - refresh_erp_data.py (TotalCharge in BasicCost_Lookup)
  - refresh_erp_parquet.py (built-in normalization, config-driven)

Reads Parquet directly, normalizes, writes all data into ERP_Master.xlsm.
Backward compatible: refresh_data() function preserved for control.py.
"""
import os, sys, re, json
import datetime as _dt
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Paths ─────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ERP_BASE_DIR = os.path.dirname(SCRIPT_DIR)  # ERP/
ENGINE_DIR = os.path.normpath(os.path.join(ERP_BASE_DIR, '..'))  # Engine_test/
PE_DIR = os.path.join(ENGINE_DIR, 'Pricing_Engine')  # Pricing_Engine/
DATA_DIR = os.path.join(PE_DIR, "data")
ERP_DIR = os.path.join(ERP_BASE_DIR, "data")
PARQUET_FILE = os.path.join(DATA_DIR, "Cleaned_Master_History.parquet")
ERP_FILE = os.path.join(ERP_DIR, "ERP_Master.xlsm")
PORT_MAP_FILE = os.path.join(DATA_DIR, "Port_Code_Mapping_Final.xlsx")
PUC_SOC_FILE = os.path.join(DATA_DIR, "PUC_SOC.xlsx")
CONFIG_FILE = os.path.join(PE_DIR, 'config', 'pipeline_rules.json')

SHEET_NAME = "📊 Pricing Dashboard"
MARKUP_STORE = "Markup_Store"

# ── ERP Layout Constants ──────────────────────────────────────────────
CONT_NAMES = ['20GP', '40GP', '40HQ', '45HQ', '40NOR', '20RF', '40RF']
PRICE_SOURCE = CONT_NAMES
CARRIERS = ["CMA", "COSCO", "EMC", "HPL", "MSC", "ONE", "WHL", "YML"]
CARRIER_DEFAULTS = {"HPL": 75, "WHL": 50}
# ── Design System Colors (SaaS-style) ──
TITLE_BAR_COLOR = "294B93"     # Deep navy title bar
PAGE_BG = "F5F6F8"             # Light gray page background
DATA_HEADER_COLOR = "475569"   # Data table header
DATA_TITLE_BG = "64748B"       # "MASTER PRICING DATA" bar
SPACER_COLOR = "E2E8F0"        # Thin spacer rows
QUOTE_COLOR = "F8FAFC"         # Quote area bg
INPUT_COLOR = "EFF6FF"         # PUC + preset inputs (blue tint)
MARKUP_BG = "F0FDF4"           # Markup inputs (green tint)
SEARCH_BG = "FFF7ED"           # Search inputs (orange tint)
ACTIVE_BG = "ECFDF5"           # Active carriers badge
QUOTE_BTN_COLOR = "166534"     # Generate Quote button
BORDER_COLOR = "CBD5E1"        # Card borders
# ── Right Sidebar Layout (Option A) ──
# Data: Row 1 = headers, Row 2+ = data, Columns A-P
# Sidebar: Columns R-W (18-23), controls stacked vertically
# Column Q (17) = narrow separator
DATA_START_ROW = 2
DATA_HEADER_ROW = 1
HIDDEN_COL_START = 24
# Sidebar column positions
SB_COL_START = 18       # R = sidebar start
SB_COL_END = 23         # W = sidebar end
SB_COL_VAL = 19         # S = value input column
SB_COL_VAL2 = 20        # T = second value column
# Sidebar row positions (stacked vertically in R-W)
SB_TITLE_ROW = 1        # R1:W1 = "PRICING ENGINE" title
SB_SEARCH_LABEL = 2     # R2 = "🔍 SEARCH"
SB_SEARCH_POL = 3       # R3=POL label, S3=dropdown
SB_SEARCH_POD = 4       # R4=POD label, S4=dropdown
SB_SEARCH_PLACE = 5     # R5=Place label, S5=dropdown
SB_SEARCH_BTN = 6       # R6=Apply, S6=Clear, T6=Quick preset
SB_MARKUP_LABEL = 8     # R8 = "💰 MARKUP"
SB_CARRIER_ROW = 9      # R9=Carrier label, S9=dropdown
SB_MARKUP_HDR = 10      # R10:X10 = container type labels
SB_MARKUP_VAL = 11      # R11:X11 = carrier markup values
SB_GLOBAL_LABEL = 12    # R12 = "ALL (Global)"
SB_GLOBAL_VAL = 13      # R13:X13 = global markup values
SB_PUC_LABEL = 15       # R15 = "📦 PUC"
SB_PUC_ROUTE = 16       # R16=Route label, S16=dropdown
SB_PUC_HDR = 17         # R17:U17 = PUC headers
SB_PUC_VAL = 18         # R18:U18 = PUC values
SB_QUOTE_LABEL = 20     # R20 = "📋 QUOTE"
SB_CUSTOMER_ROW = 21    # R21=Customer label, S21:W21=input
SB_GENERATE_ROW = 22    # R22:W22 = Generate Quote button
SB_ACTIVE_ROW = 24      # R24 = Active carriers summary
# Legacy references (for formula generation)
SEARCH_POL_COL, SEARCH_POD_COL, SEARCH_PLACE_COL = SB_COL_VAL, SB_COL_VAL, SB_COL_VAL
PUC_COL, PUC_ROW = SB_COL_VAL, SB_PUC_ROUTE
MARKUP_CARRIER_COL = SB_COL_VAL       # S9 = carrier dropdown
MARKUP_FIRST_COL = SB_COL_START       # R11 = first markup value
MARKUP_ACTIVE_COL = SB_COL_END        # W = active carriers
# Data columns (UNCHANGED — A through P)
COL_POL, COL_POD, COL_PLACE, COL_CARRIER = 1, 2, 3, 4
COL_COMMODITY, COL_EFF, COL_EXP, COL_NOTE, COL_SOURCE = 5, 6, 7, 8, 9
COL_FIRST_PRICE = 10

# ══════════════════════════════════════════════════════════════════════
# STEP 1: DATA PROCESSING (from create_master_dashboard.py)
# ══════════════════════════════════════════════════════════════════════

# Import normalization functions from create_master_dashboard (in PE/scripts)
PE_SCRIPTS_DIR = os.path.join(PE_DIR, 'scripts')
if PE_SCRIPTS_DIR not in sys.path:
    sys.path.insert(0, PE_SCRIPTS_DIR)
from create_master_dashboard import (
    normalize_notes, normalize_text_data, load_mapping_files,
    shorten_source_file, apply_one_group_codes, create_puc_lookup_table
)


def load_and_process_parquet():
    """Load parquet, normalize, pivot to Master wide format + BasicCost."""
    print(f"\n{'='*60}")
    print(f"  DIRECT PARQUET → ERP REFRESH")
    print(f"{'='*60}")

    if not os.path.exists(PARQUET_FILE):
        print(f"❌ Parquet not found: {PARQUET_FILE}")
        sys.exit(1)

    df_all = pd.read_parquet(PARQUET_FILE)
    print(f"[1/6] Parquet loaded: {len(df_all):,} rows")

    # Normalize
    port_map = load_mapping_files()
    df_all = normalize_text_data(df_all, port_map)

    # Re-identify Rate_Type
    df_all.loc[df_all['Source'].str.upper().str.contains("SCFI", na=False), 'Rate_Type'] = 'SCFI'
    df_all.loc[df_all['Source'].str.upper().str.contains("FIX", na=False), 'Rate_Type'] = 'FIX'
    df_all.loc[df_all['Source'].str.upper().str.contains("FAK", na=False), 'Rate_Type'] = 'FAK'
    df_all.loc[df_all['Source'].str.upper().str.contains("OCR", na=False), 'Rate_Type'] = 'OCR'

    # Normalize charges
    mask_scfi_fix = df_all['Rate_Type'].isin(['SCFI', 'FIX', 'OCR'])
    df_all.loc[mask_scfi_fix, 'Charge_Name'] = "Base Ocean Freight"
    mask_fak = df_all['Rate_Type'] == 'FAK'
    mask_base_fak = df_all['Charge_Name'].str.contains('ALL IN|Base|Basic|O/F|Ocean', case=False, na=False)
    df_all.loc[mask_fak & mask_base_fak, 'Charge_Name'] = "Base Ocean Freight"

    df_all = normalize_notes(df_all)
    df_all['Source'] = df_all['Source'].apply(shorten_source_file)

    # ── Master Sheet (Base Ocean Freight, 30-day filter) ──
    print("[2/6] Creating Master pivot...")
    cutoff_date = pd.Timestamp(_dt.date.today()) - pd.Timedelta(days=30)
    df_current = df_all[df_all['Rate_Type'].isin(['FAK', 'SCFI', 'FIX', 'SPECIAL', 'OCR'])].copy()
    df_current['Exp'] = pd.to_datetime(df_current['Exp'], errors='coerce')
    df_current = df_current[df_current['Exp'] >= cutoff_date]

    df_master_raw = df_current[df_current['Charge_Name'] == "Base Ocean Freight"].copy()
    df_master_raw = df_master_raw.sort_values('Amount', ascending=False)
    df_master_raw = df_master_raw.drop_duplicates(
        subset=['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note', 'Source', 'Container_Type'],
        keep='first'
    )

    df_master = df_master_raw.pivot_table(
        index=['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note', 'Source'],
        columns='Container_Type', values='Amount', aggfunc='max'
    ).reset_index()

    if "45'HQ" in df_master.columns:
        df_master = df_master.rename(columns={"45'HQ": "45HQ"})

    desired_order = ['20GP', '40GP', '40HQ', '45HQ', '40NOR', '20RF', '40RF']
    existing_containers = [c for c in desired_order if c in df_master.columns]
    df_master = df_master[['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note', 'Source'] + existing_containers]
    print(f"    → Master: {len(df_master):,} rows")

    # ── BasicCost (full charge breakdown) ──
    print("[3/6] Creating BasicCost breakdown...")
    df_recent = df_current.copy()
    df_recent = apply_one_group_codes(df_recent)

    charge_mapping = {
        'EMF': 'EIC/GFS/BAF/FDI', 'DLF': 'PCS', 'ISPS': 'ISPS/LSF/CMC',
        'BASIC O/F': 'BASIC O/F', 'ALL IN COST': 'BASIC O/F',
        'Base Ocean Freight': 'BASIC O/F', 'COMMISSION': 'HANDLING FEE FOR CARRIER'
    }
    df_recent['Charge_Name_Mapped'] = df_recent['Charge_Name'].replace(charge_mapping)

    df_history_wide = pd.DataFrame()
    if not df_recent.empty:
        df_recent['Charge_Container'] = df_recent['Charge_Name_Mapped'] + '_' + df_recent['Container_Type']
        index_cols = ['Rate_Type', 'Source', 'POL', 'POD', 'Place', 'Carrier', 'Commodity',
                      'Group Rate', 'Eff', 'Exp', 'Contract', 'Note']
        valid_idx = [c for c in index_cols if c in df_recent.columns]
        df_history_wide = df_recent.pivot_table(
            index=valid_idx, columns='Charge_Container', values='Amount', aggfunc='max'
        ).reset_index()
    print(f"    → BasicCost: {len(df_history_wide):,} rows")

    # ── PUC ──
    print("[4/6] Creating PUC lookup...")
    df_puc = create_puc_lookup_table()
    print(f"    → PUC: {len(df_puc):,} rows")

    # ── Reference sheets from FAK ──
    print("[5/6] Extracting FAK reference sheets...")
    REF_SHEETS = {
        'HANDLING FEE CARRIER': 'Handling Fee',
        'ONE GROUP CODE': 'ONE Group Code',
        'WHARFAGE': 'Wharfage', 'GOH': 'GOH', 'HAWAII': 'Hawaii',
    }
    ref_sheet_data = {}
    fak_file = None
    for f in os.listdir(DATA_DIR):
        if f.endswith('.xlsx') and 'FAK' in f.upper() and not f.startswith('~$'):
            fak_file = os.path.join(DATA_DIR, f)
            break
    if fak_file:
        try:
            fak_xls = pd.ExcelFile(fak_file)
            for src_name, dst_name in REF_SHEETS.items():
                if src_name in fak_xls.sheet_names:
                    ref_sheet_data[dst_name] = pd.read_excel(fak_xls, sheet_name=src_name, header=None)
                    print(f"    → {dst_name}: {ref_sheet_data[dst_name].shape[0]} rows")
        except Exception as e:
            print(f"    ⚠️ Error: {e}")

    # ── Version info ──
    EXCLUDE = ['PUC_SOC', 'Port_Code', 'Schedule', 'Master']
    raw_files = sorted([f for f in os.listdir(DATA_DIR)
                        if f.endswith('.xlsx') and not f.startswith('~$')
                        and not any(e in f for e in EXCLUDE)])
    fak_version = 'N/A'
    for rf in raw_files:
        if 'FAK' in rf.upper():
            m = re.search(r'(\d{1,2})\s+([A-Z]{3})\s+NO\.?\s*(\d+)', rf.upper())
            if m:
                fak_version = f"{m.group(1)}{m.group(2)}NO.{m.group(3)}"
            break

    version_rows = [
        ('Ấn bản bảng giá', fak_version),
        ('Tổng số dòng (Master)', len(df_master)),
        ('Ngày normalize', _dt.date.today().strftime('%d-%b-%Y')),
        ('Pipeline', 'Direct Parquet → ERP (no MasterFullPricing.xlsx)'),
    ]
    for rf in raw_files:
        fu = rf.upper()
        label = 'RAW: FAK' if 'FAK' in fu else 'RAW: SCFI' if 'SCFI' in fu else 'RAW: SPECIAL RATE' if 'SPECIAL' in fu or 'FIX' in fu else 'RAW'
        version_rows.append((label, re.sub(r'\.xlsx$', '', rf, flags=re.I)))

    return df_master, df_history_wide, df_puc, ref_sheet_data, version_rows


# ══════════════════════════════════════════════════════════════════════
# STEP 2: WRITE TO ERP (from refresh_erp_data.py)
# ══════════════════════════════════════════════════════════════════════

def _short_source(filename):
    if not filename or pd.isna(filename): return ""
    fname = str(filename).upper()
    if "FAK" in fname: return "FAK"
    elif "SCFI" in fname: return "SCFI"
    elif "FIX" in fname: return "FIX"
    elif "SPECIAL" in fname: return "SPECIAL"
    return str(filename)[:12]

def _short_date(date_val):
    if pd.isna(date_val): return ""
    try:
        dt = pd.to_datetime(date_val) if isinstance(date_val, str) else date_val
        return f"{dt.day}{dt.strftime('%b').upper()}"
    except: return str(date_val)

def read_markup_store(wb):
    store = {}
    if MARKUP_STORE in wb.sheetnames:
        ws = wb[MARKUP_STORE]
        for r in range(2, ws.max_row + 1):
            carrier = ws.cell(r, 1).value
            if carrier:
                vals = [ws.cell(r, c).value or 0 for c in range(2, 2 + len(CONT_NAMES))]
                store[str(carrier).strip()] = vals
    return store

def write_markup_store(wb, store):
    if MARKUP_STORE in wb.sheetnames:
        wb.remove(wb[MARKUP_STORE])
    ws = wb.create_sheet(MARKUP_STORE)
    ws.sheet_state = 'hidden'
    ws.cell(1, 1).value = "Carrier"
    for i, c in enumerate(CONT_NAMES):
        ws.cell(1, 2 + i).value = c
    ws.cell(2, 1).value = "ALL"
    for i, v in enumerate(store.get("ALL", [0]*len(CONT_NAMES))):
        ws.cell(2, 2 + i).value = int(v) if v else 0
    for ci, carrier in enumerate(CARRIERS):
        r = 3 + ci
        ws.cell(r, 1).value = carrier
        vals = store.get(carrier, [CARRIER_DEFAULTS.get(carrier, 0)] * len(CONT_NAMES))
        for i, v in enumerate(vals):
            ws.cell(r, 2 + i).value = int(v) if v else 0


def write_to_erp(df_master, df_history_wide, df_puc, ref_sheet_data, version_rows):
    """Write all data directly into ERP_Master.xlsm."""
    print(f"\n[6/6] Writing to ERP...")

    if not os.path.exists(ERP_FILE):
        print(f"❌ ERP not found: {ERP_FILE}")
        sys.exit(1)

    # Prep master data
    df_master['Source_Short'] = df_master['Source'].apply(_short_source)
    df_master['Eff_Short'] = df_master['Eff'].apply(_short_date)
    df_master['Exp_Short'] = df_master['Exp'].apply(_short_date)

    wb = openpyxl.load_workbook(ERP_FILE, keep_vba=True)
    print(f"   ✓ ERP opened ({len(wb.sheetnames)} sheets)")

    # ── Read existing markups ──
    existing_markups = read_markup_store(wb)
    if existing_markups:
        print(f"   ✓ Found {len(existing_markups)} saved markups")

    # ── Rebuild PUC_Lookup ──
    if not df_puc.empty:
        if "PUC_Lookup" in wb.sheetnames:
            wb.remove(wb["PUC_Lookup"])
        ws_puc = wb.create_sheet("PUC_Lookup")
        ws_puc.sheet_state = 'hidden'
        for ci, h in enumerate(['Place', '20GP', '40GP', '40HQ', '45HQ'], 1):
            ws_puc.cell(1, ci).value = h
        for ri, (_, row) in enumerate(df_puc.iterrows(), 2):
            ws_puc.cell(ri, 1).value = row['Place']
            ws_puc.cell(ri, 2).value = row.get('20GP', 0)
            ws_puc.cell(ri, 3).value = row.get('40GP', 0)
            ws_puc.cell(ri, 4).value = row.get('40HQ', 0)
            ws_puc.cell(ri, 5).value = row.get('45HQ', 0)
        puc_last_row = len(df_puc) + 1
    else:
        puc_last_row = 2

    # ── Build BasicCost_Lookup (from refresh_erp_data.py logic) ──
    print("   📋 Building BasicCost_Lookup...")
    if "BasicCost_Lookup" in wb.sheetnames:
        wb.remove(wb["BasicCost_Lookup"])
    ws_bc = wb.create_sheet("BasicCost_Lookup")
    ws_bc.sheet_state = 'hidden'
    # 5 columns: Key, Contract, GroupRate, CostBreakdown, TotalCharge (merged from refresh_erp_data.py)
    for ci, h in enumerate(['Key', 'Contract', 'GroupRate', 'CostBreakdown', 'TotalCharge'], 1):
        ws_bc.cell(1, ci).value = h

    charge_groups = [
        ("O/F", "BASIC O/F"), ("ARB", "ARB/OLF"), ("ISPS", "ISPS/LSF/CMC"),
        ("PSS/PUC", "PSS/PUC"), ("OCS/LSS", "OCS/LSS/EFF/ITC/GFS/ SOC COST HDL FEE"),
        ("PCS/ACS", "PCS/ACS/AGS"), ("GRI", "GRI"), ("EIC/BAF", "EIC/GFS/BAF/FDI"),
        ("WHA/BCO", "WHA/BCO/BCD/CFC/EIC"), ("GARMENT", "GARMENT ADD ON"),
        ("PREMIUM", "PREMIUM ADD ON/HDL FEE US FOR SOC"),
    ]

    # PUC override dict
    puc_override = {}
    if not df_puc.empty:
        puc_cont_map = {'20GP': '20GP', '40GP': '40GP', '40HQ': '40HQ', '45HQ': '45HQ'}
        for _, puc_row in df_puc.iterrows():
            route = str(puc_row.get('Place', '')).strip()
            if route:
                puc_override[route] = {k: float(puc_row.get(v, 0)) if pd.notna(puc_row.get(v)) else 0
                                        for k, v in puc_cont_map.items()}

    bc_row_idx = 2
    seen_keys = set()
    if not df_history_wide.empty:
        df_bc = df_history_wide.copy()
        if 'Exp' in df_bc.columns:
            df_bc = df_bc.sort_values('Exp', ascending=False, na_position='last')
        for _, row in df_bc.iterrows():
            pol = str(row.get('POL', '')).strip()
            pod = str(row.get('POD', '')).strip()
            bc_place = str(row.get('Place', '')).strip() if pd.notna(row.get('Place', '')) else ''
            carrier = str(row.get('Carrier', '')).strip()
            contract = row.get('Contract', '')
            group_rate = row.get('Group Rate', '')
            if not pol or not carrier: continue
            for cont in CONT_NAMES:
                note_val = str(row.get('Note', '')).strip().upper()
                key = f"{pol}|{pod}|{bc_place}|{carrier}|{cont}|{note_val}"
                if key in seen_keys: continue
                of_col = f"BASIC O/F_{cont}"
                of_val = row.get(of_col)
                if pd.isna(of_val) or of_val == 0: continue
                seen_keys.add(key)
                cost_parts = []
                total_charge = 0.0  # TotalCharge tracking (from refresh_erp_data.py)
                for short_name, charge_prefix in charge_groups:
                    col_name = f"{charge_prefix}_{cont}"
                    val = row.get(col_name)
                    if short_name == "PSS/PUC" and note_val == "SOC" and puc_override:
                        for rn, rv in puc_override.items():
                            if rn.upper() in bc_place.upper() or bc_place.upper() in rn.upper():
                                pv = rv.get(cont, 0)
                                if pv != 0: val = pv
                                break
                    if pd.notna(val) and val != 0:
                        try:
                            fval = float(val)
                            cost_parts.append(f"{short_name} ${fval:,.0f}")
                            total_charge += fval
                        except: pass
                hdl_col = f"HANDLING FEE FOR CARRIER_{cont}"
                hdl_val = row.get(hdl_col)
                hdl_str = ""
                if pd.notna(hdl_val) and hdl_val != 0:
                    try:
                        hdl_float = float(hdl_val)
                        hdl_str = f"\nHDL FEE: ${hdl_float:,.0f}"
                        total_charge += hdl_float
                    except: pass
                contract_str = str(contract) if pd.notna(contract) else ""
                group_str = str(group_rate) if pd.notna(group_rate) and str(group_rate).strip() else "N/A"
                breakdown = f"BKG: S/C={contract_str} | Group={group_str}\n"
                if cost_parts: breakdown += f"COST: {' + '.join(cost_parts)}"
                if hdl_str: breakdown += hdl_str
                ws_bc.cell(bc_row_idx, 1).value = key
                ws_bc.cell(bc_row_idx, 2).value = contract_str
                ws_bc.cell(bc_row_idx, 3).value = group_str
                ws_bc.cell(bc_row_idx, 4).value = breakdown
                ws_bc.cell(bc_row_idx, 5).value = round(total_charge, 0)  # TotalCharge col E
                bc_row_idx += 1
    print(f"   ✓ {bc_row_idx - 2} BasicCost lookup entries")

    # ── Build Pricing_Data (Phase 1 — clean data backend) ──
    print("   📦 Building Pricing_Data (flat table)...")
    PRICING_DATA_SHEET = "Pricing_Data"
    if PRICING_DATA_SHEET in wb.sheetnames:
        wb.remove(wb[PRICING_DATA_SHEET])
    ws_pd = wb.create_sheet(PRICING_DATA_SHEET)
    ws_pd.sheet_state = 'hidden'

    # Headers: 17 columns — RateID + 9 info cols + 7 base price cols
    pd_headers = ['RateID', 'POL', 'POD', 'Place', 'Carrier', 'Commodity',
                  'Eff', 'Exp', 'Note', 'Source',
                  'Base_20GP', 'Base_40GP', 'Base_40HQ', 'Base_45HQ',
                  'Base_40NOR', 'Base_20RF', 'Base_40RF']
    pd_header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    pd_header_fill = PatternFill('solid', fgColor='1F4E79')
    for ci, h in enumerate(pd_headers, 1):
        cell = ws_pd.cell(1, ci)
        cell.value = h
        cell.font = pd_header_font
        cell.fill = pd_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows: pure values, no formulas
    for ri, (_, row) in enumerate(df_master.iterrows(), 2):
        pol_v = str(row.get('POL', '')).strip()
        pod_v = str(row.get('POD', '')).strip()
        place_v = str(row.get('Place', '')).strip()
        carrier_v = str(row.get('Carrier', '')).strip()
        # RateID = POL|POD|Place|Carrier (4-part composite key)
        rate_id = f"{pol_v}|{pod_v}|{place_v}|{carrier_v}"
        ws_pd.cell(ri, 1).value = rate_id
        ws_pd.cell(ri, 2).value = pol_v
        ws_pd.cell(ri, 3).value = pod_v
        ws_pd.cell(ri, 4).value = place_v
        ws_pd.cell(ri, 5).value = carrier_v
        ws_pd.cell(ri, 6).value = row.get('Commodity', '')
        ws_pd.cell(ri, 7).value = row.get('Eff_Short', '')
        ws_pd.cell(ri, 8).value = row.get('Exp_Short', '')
        ws_pd.cell(ri, 9).value = row.get('Note', '')
        ws_pd.cell(ri, 10).value = row.get('Source_Short', '')
        for ci_cont, cont_name in enumerate(CONT_NAMES):
            base_val = row.get(cont_name)
            ws_pd.cell(ri, 11 + ci_cont).value = float(base_val) if pd.notna(base_val) and base_val != 0 else 0

    # Column widths for readability (even though hidden)
    pd_widths = {'A': 30, 'B': 8, 'C': 10, 'D': 18, 'E': 10, 'F': 14,
                 'G': 8, 'H': 8, 'I': 10, 'J': 10}
    for col_letter, width in pd_widths.items():
        ws_pd.column_dimensions[col_letter].width = width
    for ci in range(11, 18):
        ws_pd.column_dimensions[get_column_letter(ci)].width = 10

    # Add Excel Table for structure
    pd_end_row = len(df_master) + 1
    if pd_end_row > 1:
        pd_table_ref = f"A1:{get_column_letter(len(pd_headers))}{pd_end_row}"
        pd_tab = Table(displayName="PricingData", ref=pd_table_ref)
        pd_tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
        ws_pd.add_table(pd_tab)
    ws_pd.freeze_panes = 'A2'
    print(f"   ✓ Pricing_Data: {len(df_master):,} rows, {len(pd_headers)} columns")

    # ── Build Search_Lists ──
    unique_pols = sorted(df_master['POL'].dropna().unique().tolist())
    unique_pods = sorted(df_master['POD'].dropna().unique().tolist())
    unique_places = sorted(df_master['Place'].dropna().unique().tolist())
    if "Search_Lists" in wb.sheetnames:
        wb.remove(wb["Search_Lists"])
    ws_search = wb.create_sheet("Search_Lists")
    ws_search.sheet_state = 'hidden'
    ws_search.cell(1, 1).value = "POL"
    ws_search.cell(1, 2).value = "POD"
    ws_search.cell(1, 3).value = "Place"
    for i, v in enumerate(unique_pols, 2): ws_search.cell(i, 1).value = v
    for i, v in enumerate(unique_pods, 2): ws_search.cell(i, 2).value = v
    for i, v in enumerate(unique_places, 2): ws_search.cell(i, 3).value = v
    pol_last = len(unique_pols) + 1
    pod_last = len(unique_pods) + 1
    place_last = len(unique_places) + 1
    print(f"   ✓ Search lists: POL({len(unique_pols)}) POD({len(unique_pods)}) Place({len(unique_places)})")

    # ── Write Reference Sheets (hidden) ──
    print("   📋 Writing reference sheets...")
    HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
    DATA_FONT = Font(name='Arial', size=10)
    for ref_name, ref_df in ref_sheet_data.items():
        if ref_name in wb.sheetnames:
            wb.remove(wb[ref_name])
        ws_ref = wb.create_sheet(ref_name)
        ws_ref.sheet_state = 'hidden'
        for ri, (_, row) in enumerate(ref_df.iterrows(), 1):
            for ci, value in enumerate(row, 1):
                ws_ref.cell(ri, ci, value)
        for cell in ws_ref[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        ws_ref.freeze_panes = 'A2'
        print(f"      → {ref_name} ✓")

    # ── Write Version Sheet (hidden) ──
    if "Version" in wb.sheetnames:
        wb.remove(wb["Version"])
    ws_ver = wb.create_sheet("Version")
    ws_ver.sheet_state = 'hidden'
    ws_ver.column_dimensions['A'].width = 30
    ws_ver.column_dimensions['B'].width = 55
    for ri, (label, value) in enumerate(version_rows, 1):
        ws_ver.cell(ri, 1, label)
        ws_ver.cell(ri, 2, value)

    # ══════════════════════════════════════════════════════════════════
    # PRICING DASHBOARD (exact same layout as refresh_erp_data.py)
    # ══════════════════════════════════════════════════════════════════
    print("   📊 Writing Pricing Dashboard...")
    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ Sheet '{SHEET_NAME}' not found!")
        sys.exit(1)
    ws = wb[SHEET_NAME]

    for merge_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge_range))

    # Clear all existing Data Validations to prevent corruption
    # (openpyxl strips DV extensions on load, causing save errors in Excel)
    from openpyxl.worksheet.datavalidation import DataValidationList
    ws.data_validations = DataValidationList()

    max_clear_row = max(ws.max_row, DATA_START_ROW + 10)
    for r in range(1, max_clear_row + 1):
        for c in range(1, 35):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font()
            cell.border = Border()
            cell.alignment = Alignment()

    # Reset old row dimensions (heights) from previous layout
    for r in range(1, 15):
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None

    if ws.tables:
        for t_name in list(ws.tables.keys()):
            del ws.tables[t_name]

    # ── Shared styles ──
    thin = Side(style='thin', color=BORDER_COLOR)
    card_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill(start_color=TITLE_BAR_COLOR, end_color=TITLE_BAR_COLOR, fill_type="solid")
    page_bg_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Darker gray contrast
    data_header_fill = PatternFill(start_color=DATA_HEADER_COLOR, end_color=DATA_HEADER_COLOR, fill_type="solid")
    spacer_fill = PatternFill(start_color=SPACER_COLOR, end_color=SPACER_COLOR, fill_type="solid")
    input_fill = PatternFill(start_color=INPUT_COLOR, end_color=INPUT_COLOR, fill_type="solid")
    markup_fill = PatternFill(start_color=MARKUP_BG, end_color=MARKUP_BG, fill_type="solid")
    search_fill = PatternFill(start_color=SEARCH_BG, end_color=SEARCH_BG, fill_type="solid")
    active_fill = PatternFill(start_color=ACTIVE_BG, end_color=ACTIVE_BG, fill_type="solid")
    global_fill = PatternFill(start_color=INPUT_COLOR, end_color=INPUT_COLOR, fill_type="solid")
    quote_btn_fill = PatternFill(start_color=QUOTE_BTN_COLOR, end_color=QUOTE_BTN_COLOR, fill_type="solid")
    section_font = Font(name='Segoe UI', size=9, bold=True, color="64748B")
    label_font = Font(name='Segoe UI', size=9, color="475569")

    # ══════════════════════════════════════════════════════════════════
    # ROW 1: DATA HEADERS (data starts immediately!)
    # ══════════════════════════════════════════════════════════════════
    ws.row_dimensions[DATA_HEADER_ROW].height = 28
    headers = ['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note', 'Source',
               '20GP', '40GP', '40HQ', '45HQ', '40NOR', '20RF', '40RF']
    for i, h in enumerate(headers):
        cell = ws.cell(DATA_HEADER_ROW, i + 1)
        cell.value = h
        cell.font = Font(name='Segoe UI', bold=True, color="FFFFFF", size=10)
        cell.fill = data_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = card_border

    # ══════════════════════════════════════════════════════════════════
    # SIDEBAR: Columns R-W (18-23) — Control Panel
    # ══════════════════════════════════════════════════════════════════
    # Column Q = narrow separator
    ws.column_dimensions[get_column_letter(17)].width = 2
    # Sidebar column widths
    sb_widths = {SB_COL_START: 12, SB_COL_VAL: 10, SB_COL_VAL2: 10, 21: 10, 22: 10, SB_COL_END: 10}
    for col, w in sb_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Fill sidebar background
    for r in range(1, 26):
        for c in range(SB_COL_START, SB_COL_END + 1):
            ws.cell(r, c).fill = page_bg_fill

    # ── SIDEBAR TITLE ──
    ws.merge_cells(start_row=SB_TITLE_ROW, start_column=SB_COL_START,
                   end_row=SB_TITLE_ROW, end_column=SB_COL_END)
    cell = ws.cell(SB_TITLE_ROW, SB_COL_START)
    cell.value = "PRICING ENGINE"
    cell.font = Font(name='Segoe UI', size=12, bold=True, color="FFFFFF")
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for c in range(SB_COL_START, SB_COL_END + 1):
        ws.cell(SB_TITLE_ROW, c).fill = title_fill

    # ── SEARCH SECTION ──
    ws.cell(SB_SEARCH_LABEL, SB_COL_START).value = u"\U0001F50D SEARCH"
    ws.cell(SB_SEARCH_LABEL, SB_COL_START).font = section_font
    search_items = [
        (SB_SEARCH_POL, "POL:", f"=Search_Lists!$A$2:$A${pol_last}"),
        (SB_SEARCH_POD, "POD:", f"=Search_Lists!$B$2:$B${pod_last}"),
        (SB_SEARCH_PLACE, "Place:", f"=Search_Lists!$C$2:$C${place_last}"),
    ]
    for sb_row, lbl, formula in search_items:
        ws.cell(sb_row, SB_COL_START).value = lbl
        ws.cell(sb_row, SB_COL_START).font = label_font
        ws.cell(sb_row, SB_COL_START).alignment = Alignment(horizontal='right', vertical='center')
        dd_cell = ws.cell(sb_row, SB_COL_VAL)
        dd_cell.fill = search_fill
        dd_cell.font = Font(name='Segoe UI', bold=True, size=10, color="9A3412")
        dd_cell.alignment = Alignment(horizontal='center', vertical='center')
        dd_cell.border = card_border
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(dd_cell)

    # Quick preset in search section
    ws.cell(SB_SEARCH_BTN, SB_COL_START).value = "Quick:"
    ws.cell(SB_SEARCH_BTN, SB_COL_START).font = label_font
    ws.cell(SB_SEARCH_BTN, SB_COL_START).alignment = Alignment(horizontal='right', vertical='center')
    preset_cell = ws.cell(SB_SEARCH_BTN, SB_COL_VAL)
    preset_cell.value = "DRY"
    preset_cell.fill = input_fill
    preset_cell.font = Font(name='Segoe UI', bold=True, size=10, color="1E40AF")
    preset_cell.alignment = Alignment(horizontal='center', vertical='center')
    preset_cell.border = card_border
    dv_preset = DataValidation(type="list", formula1='"DRY,REEFER,FULL,ALL"')
    ws.add_data_validation(dv_preset)
    dv_preset.add(preset_cell)

    # Divider
    for c in range(SB_COL_START, SB_COL_END + 1):
        ws.cell(7, c).fill = spacer_fill
    ws.row_dimensions[7].height = 4

    # ── MARKUP SECTION ──
    ws.cell(SB_MARKUP_LABEL, SB_COL_START).value = u"\U0001F4B0 MARKUP"
    ws.cell(SB_MARKUP_LABEL, SB_COL_START).font = section_font
    # Carrier dropdown
    ws.cell(SB_CARRIER_ROW, SB_COL_START).value = "Carrier:"
    ws.cell(SB_CARRIER_ROW, SB_COL_START).font = label_font
    ws.cell(SB_CARRIER_ROW, SB_COL_START).alignment = Alignment(horizontal='right', vertical='center')
    carrier_cell = ws.cell(SB_CARRIER_ROW, SB_COL_VAL)
    carrier_cell.value = CARRIERS[0]
    carrier_cell.font = Font(name='Segoe UI', bold=True, size=11, color="166534")
    carrier_cell.fill = markup_fill
    carrier_cell.alignment = Alignment(horizontal='center', vertical='center')
    carrier_cell.border = card_border
    dv_carrier = DataValidation(type="list", formula1=f'"{",".join(CARRIERS)}"')
    ws.add_data_validation(dv_carrier)
    dv_carrier.add(carrier_cell)

    # Container type headers (R10:X10)
    for i, cont in enumerate(CONT_NAMES):
        c = ws.cell(SB_MARKUP_HDR, SB_COL_START + i)
        c.value = cont
        c.font = Font(name='Segoe UI', size=8, color="94A3B8")
        c.alignment = Alignment(horizontal='center', vertical='center')

    # Carrier markup values (R11:X11)
    first_carrier = CARRIERS[0]
    first_vals = existing_markups.get(first_carrier, [CARRIER_DEFAULTS.get(first_carrier, 0)] * len(CONT_NAMES))
    for i in range(len(CONT_NAMES)):
        cell = ws.cell(SB_MARKUP_VAL, SB_COL_START + i)
        cell.value = first_vals[i] if i < len(first_vals) else 0
        cell.number_format = '#,##0'
        cell.font = Font(name='Segoe UI', bold=True, size=11, color="166534")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = markup_fill
        cell.border = card_border
        cell.protection = Protection(locked=False)

    # Global ALL markup
    ws.cell(SB_GLOBAL_LABEL, SB_COL_START).value = "ALL (Global):"
    ws.cell(SB_GLOBAL_LABEL, SB_COL_START).font = Font(name='Segoe UI', size=9, bold=True, color="1E40AF")
    ws.merge_cells(start_row=SB_GLOBAL_LABEL, start_column=SB_COL_START,
                   end_row=SB_GLOBAL_LABEL, end_column=SB_COL_START + 1)
    all_vals = existing_markups.get("ALL", [0] * len(CONT_NAMES))
    for i in range(len(CONT_NAMES)):
        cell = ws.cell(SB_GLOBAL_VAL, SB_COL_START + i)
        cell.value = all_vals[i] if i < len(all_vals) else 0
        cell.number_format = '#,##0'
        cell.font = Font(name='Segoe UI', bold=True, size=10, color="1E40AF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = global_fill
        cell.border = card_border
        cell.protection = Protection(locked=False)

    # Divider
    for c in range(SB_COL_START, SB_COL_END + 1):
        ws.cell(14, c).fill = spacer_fill
    ws.row_dimensions[14].height = 4

    # ── PUC SECTION ──
    ws.cell(SB_PUC_LABEL, SB_COL_START).value = u"\U0001F4E6 PUC"
    ws.cell(SB_PUC_LABEL, SB_COL_START).font = section_font
    ws.cell(SB_PUC_ROUTE, SB_COL_START).value = "Route:"
    ws.cell(SB_PUC_ROUTE, SB_COL_START).font = label_font
    ws.cell(SB_PUC_ROUTE, SB_COL_START).alignment = Alignment(horizontal='right', vertical='center')
    route_cell = ws.cell(SB_PUC_ROUTE, SB_COL_VAL)
    route_cell.font = Font(name='Segoe UI', bold=True, color="1E40AF", size=10)
    route_cell.fill = input_fill
    route_cell.alignment = Alignment(horizontal='center', vertical='center')
    route_cell.border = card_border
    if not df_puc.empty:
        dv_route = DataValidation(type="list", formula1=f"=PUC_Lookup!$A$2:$A${puc_last_row}", allow_blank=True)
        ws.add_data_validation(dv_route)
        dv_route.add(route_cell)
        puc_places = df_puc['Place'].dropna().unique().tolist()
        route_cell.value = puc_places[0] if puc_places else ""

    # PUC headers
    for i, cont in enumerate(['20GP', '40GP', '40HQ', '45HQ']):
        c = ws.cell(SB_PUC_HDR, SB_COL_START + i)
        c.value = cont
        c.font = Font(name='Segoe UI', size=8, color="94A3B8")
        c.alignment = Alignment(horizontal='center', vertical='center')

    # PUC values
    route_col_let = get_column_letter(SB_COL_VAL)  # S column
    puc_cell_refs = []
    for i, col_idx in enumerate([2, 3, 4, 5]):
        cell = ws.cell(SB_PUC_VAL, SB_COL_START + i)
        cell.value = f'=IFERROR(VLOOKUP(${route_col_let}${SB_PUC_ROUTE},PUC_Lookup!$A:$E,{col_idx},0),0)'
        cell.number_format = '#,##0'
        cell.fill = input_fill
        cell.font = Font(name='Segoe UI', bold=True, size=11, color="1E40AF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = card_border
        puc_cell_refs.append(f'${get_column_letter(SB_COL_START + i)}${SB_PUC_VAL}')
    puc_cell_refs.extend([None, None, None])

    # Divider
    for c in range(SB_COL_START, SB_COL_END + 1):
        ws.cell(19, c).fill = spacer_fill
    ws.row_dimensions[19].height = 4

    # ── QUOTE SECTION ──
    ws.cell(SB_QUOTE_LABEL, SB_COL_START).value = u"\U0001F4CB QUOTE"
    ws.cell(SB_QUOTE_LABEL, SB_COL_START).font = section_font
    ws.cell(SB_CUSTOMER_ROW, SB_COL_START).value = "Customer:"
    ws.cell(SB_CUSTOMER_ROW, SB_COL_START).font = label_font
    ws.cell(SB_CUSTOMER_ROW, SB_COL_START).alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells(start_row=SB_CUSTOMER_ROW, start_column=SB_COL_VAL,
                   end_row=SB_CUSTOMER_ROW, end_column=SB_COL_END)
    cust_cell = ws.cell(SB_CUSTOMER_ROW, SB_COL_VAL)
    cust_cell.font = Font(name='Segoe UI', bold=True, size=11, color="1E293B")
    cust_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    bottom_accent = Side(style='medium', color='294B93')
    cust_cell.border = Border(bottom=bottom_accent)

    # Generate Quote button
    ws.merge_cells(start_row=SB_GENERATE_ROW, start_column=SB_COL_START,
                   end_row=SB_GENERATE_ROW, end_column=SB_COL_END)
    gen_cell = ws.cell(SB_GENERATE_ROW, SB_COL_START)
    gen_cell.value = u"\u25B6 Generate Quote"
    gen_cell.font = Font(name='Segoe UI', bold=True, color="FFFFFF", size=11)
    gen_cell.fill = quote_btn_fill
    gen_cell.alignment = Alignment(horizontal='center', vertical='center')
    gen_cell.border = card_border
    for c in range(SB_COL_START, SB_COL_END + 1):
        ws.cell(SB_GENERATE_ROW, c).fill = quote_btn_fill

    # Divider
    for c in range(SB_COL_START, SB_COL_END + 1):
        ws.cell(23, c).fill = spacer_fill
    ws.row_dimensions[23].height = 4

    # Active carriers summary
    active_cell = ws.cell(SB_ACTIVE_ROW, SB_COL_START)
    active_cell.value = f'=SUMPRODUCT((MMULT(IF({MARKUP_STORE}!B3:H{2 + len(CARRIERS)}<>0,1,0),ROW(INDIRECT("1:{len(CONT_NAMES)}"))^0)>0)*1)&" active carriers"'
    active_cell.font = Font(name='Segoe UI', size=9, italic=True, color="166534")
    active_cell.alignment = Alignment(horizontal='left', vertical='center')
    active_cell.fill = active_fill
    ws.merge_cells(start_row=SB_ACTIVE_ROW, start_column=SB_COL_START,
                   end_row=SB_ACTIVE_ROW, end_column=SB_COL_END)

    # Timestamp
    ts = datetime.now().strftime('%d%b %H:%M').upper()
    ws.cell(25, SB_COL_START).value = f"Updated: {ts}"
    ws.cell(25, SB_COL_START).font = Font(name='Segoe UI', size=8, italic=True, color="94A3B8")

    # ══════════════════════════════════════════════════════════════════
    # DATA ROWS WITH FORMULAS (Row 2+, same columns A-P)
    # ══════════════════════════════════════════════════════════════════
    print(f"   Writing {len(df_master):,} rows with formulas...")
    ms_carrier_range = f"{MARKUP_STORE}!$A$3:$A${2 + len(CARRIERS)}"

    # Cell references for formulas — now pointing to sidebar positions
    carrier_dd_ref = f"${get_column_letter(SB_COL_VAL)}${SB_CARRIER_ROW}"  # $S$9

    for r_idx, (_, row) in enumerate(df_master.iterrows(), start=DATA_START_ROW):
        ws.cell(r_idx, COL_POL).value = row['POL']
        ws.cell(r_idx, COL_POD).value = row['POD']
        ws.cell(r_idx, COL_PLACE).value = row['Place']
        ws.cell(r_idx, COL_CARRIER).value = row['Carrier']
        ws.cell(r_idx, COL_COMMODITY).value = row['Commodity']
        ws.cell(r_idx, COL_EFF).value = row['Eff_Short']
        ws.cell(r_idx, COL_EXP).value = row['Exp_Short']
        ws.cell(r_idx, COL_NOTE).value = row['Note']
        ws.cell(r_idx, COL_SOURCE).value = row['Source_Short']

        for col_idx_src, price_col in enumerate(PRICE_SOURCE):
            base_val = row.get(price_col)
            ws.cell(r_idx, HIDDEN_COL_START + col_idx_src).value = float(base_val) if pd.notna(base_val) and base_val != 0 else 0

        place_ref = f'$C{r_idx}'
        carrier_ref = f'$D{r_idx}'
        for col_idx, (cont_name, puc_ref) in enumerate(zip(CONT_NAMES, puc_cell_refs)):
            hidden_letter = get_column_letter(HIDDEN_COL_START + col_idx)
            vis_col = COL_FIRST_PRICE + col_idx
            store_data_col = col_idx + 2
            # Markup references → sidebar positions
            markup_col_letter = get_column_letter(SB_COL_START + col_idx)  # R+offset
            global_ref = f"${markup_col_letter}${SB_GLOBAL_VAL}"   # e.g. $R$13
            carrier_cell_ref = f"${markup_col_letter}${SB_MARKUP_VAL}"  # e.g. $R$11
            ms_data_range = f"{MARKUP_STORE}!${get_column_letter(store_data_col)}$3:${get_column_letter(store_data_col)}${2 + len(CARRIERS)}"
            carrier_lookup = (f'IF({carrier_ref}={carrier_dd_ref},{carrier_cell_ref},'
                              f'IFERROR(INDEX({ms_data_range},MATCH({carrier_ref},{ms_carrier_range},0)),0))')
            note_ref = f'${get_column_letter(COL_NOTE)}{r_idx}'
            puc_formula = ''
            if puc_ref:
                puc_lookup_col_map = {0: 'B', 1: 'C', 2: 'D', 3: 'E'}
                if col_idx in puc_lookup_col_map:
                    puc_lk_col = puc_lookup_col_map[col_idx]
                    stored_puc = f'IFERROR(SUMPRODUCT((ISNUMBER(SEARCH(PUC_Lookup!$A$2:$A${puc_last_row},{place_ref})))*PUC_Lookup!${puc_lk_col}$2:${puc_lk_col}${puc_last_row}),0)'
                    puc_formula = (f'+IF({note_ref}="SOC",'
                                   f'IF(ISNUMBER(SEARCH(${route_col_let}${SB_PUC_ROUTE},{place_ref})),{puc_ref},{stored_puc})'
                                   f',0)')

            formula = (f'=IF({hidden_letter}{r_idx}>0,'
                       f'{hidden_letter}{r_idx}+{global_ref}+{carrier_lookup}{puc_formula}'
                       f',"")')
            cell = ws.cell(r_idx, vis_col)
            cell.value = formula
            cell.number_format = '#,##0'

    # Hide helper columns
    for col_idx in range(HIDDEN_COL_START, HIDDEN_COL_START + 7):
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    # Excel Table
    end_row = DATA_START_ROW + len(df_master) - 1
    last_data_col = COL_FIRST_PRICE + len(CONT_NAMES) - 1
    table_ref = f"A{DATA_HEADER_ROW}:{get_column_letter(last_data_col)}{end_row}"
    tab = Table(displayName="MasterPricing", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    ws.add_table(tab)
    ws.auto_filter.ref = table_ref

    # Column widths (data area)
    widths = {'A': 10, 'B': 10, 'C': 20, 'D': 12, 'E': 22, 'F': 8, 'G': 8, 'H': 10, 'I': 12}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    for col in range(COL_FIRST_PRICE, COL_FIRST_PRICE + len(CONT_NAMES)):
        ws.column_dimensions[get_column_letter(col)].width = 10

    # Freeze: Row 1 frozen (headers), sidebar always visible
    ws.freeze_panes = f"R{DATA_START_ROW}"

    # ── Write Markup Store ──
    if not existing_markups:
        existing_markups["ALL"] = [0] * len(CONT_NAMES)
        for carrier in CARRIERS:
            existing_markups[carrier] = [CARRIER_DEFAULTS.get(carrier, 0)] * len(CONT_NAMES)
    write_markup_store(wb, existing_markups)

    # ── Delete excess rows ──
    if ws.max_row > end_row:
        ws.delete_rows(end_row + 1, ws.max_row - end_row)

    # ── Save ──
    print("   💾 Saving...")
    wb.save(ERP_FILE)

    # ── Re-inject customUI (openpyxl strips it on save) ──
    try:
        from customui_utils import ensure_customui
        customui_xml = os.path.join(ERP_BASE_DIR, "vba", "CustomUI_ERP.xml")
        result = ensure_customui(ERP_FILE, customui_xml_path=customui_xml)
        if result.get("already_ok"):
            print("   🎗️  CustomUI: already intact")
        elif result.get("injected"):
            print("   🎗️  CustomUI: re-injected after save")
        elif result.get("error"):
            print(f"   ⚠️  CustomUI: {result['error']}")
    except Exception as e:
        print(f"   ⚠️  CustomUI re-inject skipped: {e}")

    print(f"\n{'='*60}")
    print(f"✅ DIRECT PARQUET → ERP COMPLETE!")
    print(f"   📊 {len(df_master):,} pricing rows")
    print(f"   📋 BasicCost: {bc_row_idx - 2:,} entries")
    print(f"   🔍 Search: POL({len(unique_pols)}) POD({len(unique_pods)}) Place({len(unique_places)})")
    print(f"   📁 Ref sheets: {', '.join(ref_sheet_data.keys()) or 'none'}")
    print(f"   📁 Hidden: PUC_Lookup, BasicCost_Lookup, Search_Lists, Markup_Store, Pricing_Data, Version")
    print(f"   ⚠️  Quotes/Jobs sheets: UNTOUCHED")
    print(f"{'='*60}")


# ══════════════════════════════════════════════════════════════════════
# BACKWARD COMPATIBILITY — refresh_data() for control.py / ERP_Control.py
# ══════════════════════════════════════════════════════════════════════

def refresh_data():
    """Backward-compatible entry point. Called by control.py and ERP_Control.py."""
    df_master, df_history_wide, df_puc, ref_sheet_data, version_rows = load_and_process_parquet()
    write_to_erp(df_master, df_history_wide, df_puc, ref_sheet_data, version_rows)


# ══════════════════════════════════════════════════════════════════════
# AUTO CLOSE/REOPEN — Graceful Excel handling for scheduled runs
# ══════════════════════════════════════════════════════════════════════

def refresh_erp():
    """
    Full refresh with automatic Excel close/reopen.
    Returns dict with refresh status for callers (rate_importer.py etc).
    """
    was_open = False
    xl = None
    refresh_ok = False

    # Step 1 — Detect if Excel has ERP_Master open, save + close
    try:
        import win32com.client
        xl = win32com.client.GetActiveObject("Excel.Application")
        for wb in xl.Workbooks:
            if "ERP_Master" in wb.Name or "ERP_V13" in wb.Name:
                was_open = True
                wb.Save()
                wb.Close(SaveChanges=True)
                print("[INFO] ERP saved and closed for refresh")
                break
    except Exception:
        pass  # Excel not running or ERP not open — fine

    import time
    if was_open:
        time.sleep(2)  # Let Excel fully release the file

    # Step 2 — Run existing refresh logic (unchanged)
    try:
        refresh_data()
        refresh_ok = True
    except PermissionError:
        print("[ERROR] PermissionError — file still locked")
        refresh_ok = False
    except Exception as e:
        print(f"[ERROR] Refresh failed: {e}")
        import traceback
        traceback.print_exc()
        refresh_ok = False

    # Step 3 — Always reopen if it was open before
    reopened = False
    if was_open:
        try:
            import win32com.client
            if xl is None:
                xl = win32com.client.Dispatch("Excel.Application")
            xl.Visible = True
            xl.Workbooks.Open(os.path.abspath(ERP_FILE))
            reopened = True
            print("[INFO] ERP reopened automatically")
        except Exception as e:
            print(f"[WARN] Could not reopen ERP: {e}")

    return {
        "erp_refreshed": refresh_ok,
        "was_open": was_open,
        "reopened": reopened,
    }


if __name__ == "__main__":
    result = refresh_erp()
    if not result["erp_refreshed"]:
        sys.exit(1)
