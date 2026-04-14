"""
active-jobs-schema.py — Active Jobs v4 schema migration
=======================================================
Adds cols 31-36 to Active Jobs sheet in ERP_Master_v14.xlsm:
  31 (AE): SERVICE              — CY-CY / CY-DOOR
  32 (AF): TRACKING_STAGE       — stage code (1/7..7/7) from shipment_tracker
  33 (AG): RELEASE_EMAIL_SENT   — timestamp docsteam sent release email to customer
  34 (AH): RELEASE_CONFIRMED    — timestamp customer confirmed pickup
  35 (AI): PRICE_WATCH_STATUS   — OK / DROP / RISE (from price_alerts)
  36 (AJ): PRICE_WATCH_DELTA    — $ delta (negative = price dropped, good for re-quote)

Run ONCE before features 1/2/3 use those columns.
Idempotent: skips if headers already present.

Usage:
    python ERP/core/active_jobs_schema.py
    # optional: --file path/to/ERP_Master_v14.xlsm
"""
from __future__ import annotations

import argparse
import os
import sys
from typing import Final

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# Ensure ribbon XML survives openpyxl save
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from ribbon_guard import save_preserving_ribbon  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]

DEFAULT_ERP_FILE: Final = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"

HEADER_ROW: Final = 7
NEW_COLUMNS: Final = [
    (31, "SERVICE",            14),
    (32, "TRACKING_STAGE",     16),
    (33, "RELEASE_EMAIL_SENT", 18),
    (34, "RELEASE_CONFIRMED",  18),
    (35, "PRICE_WATCH_STATUS", 16),
    (36, "PRICE_WATCH_DELTA",  14),
]


def _col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def migrate(erp_file: str) -> int:
    if not os.path.exists(erp_file):
        print(f"[ERROR] ERP file not found: {erp_file}")
        return 1
    try:
        with open(erp_file, "r+b"):
            pass
    except PermissionError:
        print(f"[ERROR] ERP file is open in Excel. Close it first: {erp_file}")
        return 2

    print(f"[+] Opening: {erp_file}")
    wb = openpyxl.load_workbook(erp_file, keep_vba=True)

    sheet_name = next((s for s in wb.sheetnames if "Active" in s), None)
    if not sheet_name:
        print("[ERROR] 'Active Jobs' sheet not found")
        wb.close()
        return 3
    ws = wb[sheet_name]
    print(f"    -> sheet: '{sheet_name}'  max_row={ws.max_row}")

    header_font = Font(bold=True, color="FFFFFF", size=10, name="Segoe UI")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    added = 0
    for col_idx, name, width in NEW_COLUMNS:
        cur = ws.cell(HEADER_ROW, col_idx).value
        if cur == name:
            print(f"    [skip] col {_col_letter(col_idx)} ({col_idx}) already = {name}")
            continue
        if cur not in (None, ""):
            print(f"    [WARN] col {_col_letter(col_idx)} has '{cur}', overwriting with '{name}'")
        c = ws.cell(HEADER_ROW, col_idx, name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        ws.column_dimensions[_col_letter(col_idx)].width = width
        added += 1
        print(f"    [add]  col {_col_letter(col_idx)} ({col_idx}) = {name}")

    if added == 0:
        print("[OK] Schema already up-to-date — no changes.")
        wb.close()
        return 0

    result = save_preserving_ribbon(wb, erp_file)
    wb.close()
    print(f"\n[SUCCESS] Added {added} columns. Active Jobs now has 36 cols.")
    print(f"    ribbon: {result}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", default=DEFAULT_ERP_FILE)
    args = ap.parse_args()
    return migrate(args.file)


if __name__ == "__main__":
    sys.exit(main())
