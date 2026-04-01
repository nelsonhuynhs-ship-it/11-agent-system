# -*- coding: utf-8 -*-
"""
auto_quote_router.py — Auto Quote Email Endpoints
====================================================
FastAPI endpoints for auto-generating and sending rate quote emails
to customers using Parquet data + Outlook COM.

Endpoints:
  POST /api/auto-quote/preview     — Preview rate table for 1 customer
  POST /api/auto-quote/bulk        — Build rate tables for CMD group
  POST /api/auto-quote/send        — Send emails via Outlook
  GET  /api/auto-quote/customers   — List customers + their routes
  GET  /api/auto-quote/config      — Get email template config
"""
from __future__ import annotations

import sys
import os
import logging
from typing import Optional, List
from pathlib import Path

from fastapi import APIRouter, Query
from pydantic import BaseModel

log = logging.getLogger("nelson.auto_quote")

# ── Paths ─────────────────────────────────────────────────────────────────────
_ENGINE_TEST = Path(__file__).parent.parent.parent  # Engine_test/
_EMAIL_ENGINE = _ENGINE_TEST / "email_engine"
_CORE_DIR = _EMAIL_ENGINE / "core"
_DATA_DIR = _EMAIL_ENGINE / "data"

# Add email_engine/core to path for auto_rate_builder import
sys.path.insert(0, str(_CORE_DIR))
sys.path.insert(0, str(_EMAIL_ENGINE))

router = APIRouter(prefix="/api/auto-quote", tags=["Auto Quote Email"])


# ── Lazy Imports ──────────────────────────────────────────────────────────────
_rate_builder = None
_customer_rules = None


def _get_rate_builder():
    global _rate_builder
    if _rate_builder is None:
        try:
            from auto_rate_builder import (
                build_rate_table_for_customer,
                build_bulk_preview,
            )
            _rate_builder = {
                "single": build_rate_table_for_customer,
                "bulk": build_bulk_preview,
            }
        except ImportError as e:
            log.error("Cannot import auto_rate_builder: %s", e)
            return None
    return _rate_builder


def _get_customer_rules() -> dict:
    global _customer_rules
    if _customer_rules is None:
        import json
        rules_path = _DATA_DIR / "customer_rules.json"
        if rules_path.exists():
            with open(rules_path, "r", encoding="utf-8") as f:
                _customer_rules = json.load(f)
        else:
            _customer_rules = {}
    return _customer_rules


def _load_data_xlsx() -> list[dict]:
    """Load email_engine/data.xlsx — customer contact list with DESTINATION."""
    import pandas as pd
    data_file = _EMAIL_ENGINE / "data.xlsx"
    if not data_file.exists():
        return []
    df = pd.read_excel(data_file)
    df.columns = df.columns.str.strip().str.upper()
    return df.to_dict("records")


# ── Request Models ────────────────────────────────────────────────────────────

class PreviewRequest(BaseModel):
    """Preview rate table for a single customer."""
    pol: str = "HPH"
    destinations: str  # Comma-separated port codes: "USCHI,USLAX,USSAV"
    markup: float = 20.0
    top_per_route: int = 2
    customer_name: Optional[str] = None


class BulkRequest(BaseModel):
    """Build rate tables for a CMD group."""
    cmd_name: Optional[str] = None  # CMD group from data.xlsx
    markup: float = 20.0


class SendRequest(BaseModel):
    """Send auto-quote emails."""
    cmd_names: List[str]  # CMD groups to send to
    markup: float = 20.0
    dry_run: bool = True  # Preview only by default — safety first
    subject_override: Optional[str] = None
    batch_limit: Optional[int] = None  # Max emails to send


# ══════════════════════════════════════════════════════════════════════════════
# 1. PREVIEW — POST /api/auto-quote/preview
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/preview")
def preview_rate_table(req: PreviewRequest):
    """
    Generate an HTML rate table for a single customer/route.

    Input: POL + destination codes + markup
    Output: Outlook-ready HTML rate table with carrier comparison

    Usage: Call this before sending to verify rates are correct.
    """
    rb = _get_rate_builder()
    if not rb:
        return {"error": "auto_rate_builder not available", "success": False}

    try:
        result = rb["single"](
            pol=req.pol,
            destinations=req.destinations,
            markup=req.markup,
            top_per_route=req.top_per_route,
        )

        return {
            "success": True,
            "customer": req.customer_name or "Preview",
            "pol": req.pol,
            "destinations": req.destinations,
            "markup": req.markup,
            "html": result["html"],
            "routes_found": result["routes_found"],
            "total_rates": result["total_rates"],
            "routes_detail": result["routes_detail"],
        }

    except Exception as e:
        log.error("Preview error: %s", e)
        return {"error": str(e), "success": False}


# ══════════════════════════════════════════════════════════════════════════════
# 2. BULK — POST /api/auto-quote/bulk
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/bulk")
def bulk_preview(req: BulkRequest):
    """
    Build rate tables for all customers in a CMD group.

    Returns summary: how many customers have rates, which routes found.
    Does NOT send — use /send endpoint after reviewing.
    """
    rb = _get_rate_builder()
    if not rb:
        return {"error": "auto_rate_builder not available", "success": False}

    try:
        results = rb["bulk"](
            cmd_filter=req.cmd_name,
            markup=req.markup,
        )

        # Summary stats
        with_rates = [r for r in results if r.get("routes_found", 0) > 0]
        no_rates = [r for r in results if r.get("routes_found", 0) == 0]

        # Summarize without full HTML (too large for API response)
        summaries = []
        for r in results:
            summaries.append({
                "email": r.get("email", ""),
                "company": r.get("company", ""),
                "pol": r.get("pol", ""),
                "destinations": r.get("destinations", ""),
                "routes_found": r.get("routes_found", 0),
                "total_rates": r.get("total_rates", 0),
                "routes_detail": r.get("routes_detail", []),
            })

        return {
            "success": True,
            "cmd": req.cmd_name or "ALL",
            "markup": req.markup,
            "total_customers": len(results),
            "with_rates": len(with_rates),
            "no_rates": len(no_rates),
            "customers": summaries,
        }

    except Exception as e:
        log.error("Bulk preview error: %s", e)
        return {"error": str(e), "success": False}


# ══════════════════════════════════════════════════════════════════════════════
# 3. SEND — POST /api/auto-quote/send
# ══════════════════════════════════════════════════════════════════════════════

@router.post("/send")
def send_auto_quotes(req: SendRequest):
    """
    Send auto-quote emails via Outlook COM.

    REQUIRES: Outlook desktop running on this machine.
    Default: dry_run=True (preview only). Set dry_run=False to actually send.

    Flow:
      1. Load data.xlsx → filter by CMD names
      2. Query Parquet → build rate tables per customer
      3. Apply markup → generate HTML email
      4. Send via Outlook COM (or dry-run preview)
    """
    import pandas as pd
    from datetime import datetime

    rb = _get_rate_builder()
    if not rb:
        return {"error": "auto_rate_builder not available", "success": False}

    # 1. Load customer data
    data_file = _EMAIL_ENGINE / "data.xlsx"
    if not data_file.exists():
        return {"error": "data.xlsx not found", "success": False}

    df = pd.read_excel(data_file)
    df.columns = df.columns.str.strip().str.upper()

    # Filter by CMD names
    subset = df[df["CMD_NAME"].isin(req.cmd_names)].copy()
    subset = subset[subset["CNEE_EMAIL"].notna()]
    subset["_email_lower"] = subset["CNEE_EMAIL"].astype(str).str.lower().str.strip()
    subset = subset.drop_duplicates(subset="_email_lower")

    if subset.empty:
        return {
            "error": f"No customers found for CMD: {req.cmd_names}",
            "success": False,
        }

    # 2. Build rate tables
    DEFAULT_DESTS = "USCHI,USLAX,USLGB,USEWR,USSAV,USNYC"
    results = []
    no_rates = 0

    for _, row in subset.iterrows():
        email = str(row.get("CNEE_EMAIL", "")).strip()
        pol = str(row.get("POL", "HPH")).strip() or "HPH"
        raw_dest = row.get("DESTINATION")

        if pd.notna(raw_dest) and str(raw_dest).strip().lower() not in ("", "nan"):
            dest = str(raw_dest).strip()
        else:
            dest = DEFAULT_DESTS

        pic = str(row.get("CNEE_PIC", "")).strip()
        company = str(row.get("CNEE_NAME", "")).strip()

        result = rb["single"](pol=pol, destinations=dest, markup=req.markup)

        if result["routes_found"] == 0:
            no_rates += 1
            continue

        results.append({
            "email": email,
            "pic": pic if pic and pic.lower() not in ("nan", "") else "Team",
            "company": company,
            "pol": pol,
            "dest": dest,
            "html": result["html"],
            "routes": result["routes_found"],
            "rates": result["total_rates"],
            "routes_detail": result["routes_detail"],
        })

    if not results:
        return {
            "success": True,
            "message": "No customers with valid rates",
            "total_customers": len(subset),
            "no_rates": no_rates,
            "sent": 0,
        }

    # Apply batch limit
    if req.batch_limit:
        results = results[:req.batch_limit]

    # 3. DRY RUN — just return what would be sent
    if req.dry_run:
        return {
            "success": True,
            "dry_run": True,
            "message": f"Would send {len(results)} emails (set dry_run=false to send)",
            "total_customers": len(subset),
            "with_rates": len(results),
            "no_rates": no_rates,
            "preview": [
                {
                    "email": r["email"],
                    "company": r["company"],
                    "routes": r["routes"],
                    "rates": r["rates"],
                    "routes_detail": r["routes_detail"],
                }
                for r in results
            ],
        }

    # 4. ACTUAL SEND via Outlook COM
    try:
        import win32com.client
    except ImportError:
        return {
            "error": "win32com not available — Outlook COM requires Windows + pywin32",
            "success": False,
        }

    try:
        # Load email config
        config_file = _DATA_DIR / "config.xlsx"
        cfg_email = {}
        if config_file.exists():
            import openpyxl
            wb = openpyxl.load_workbook(str(config_file), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(max_col=2, values_only=True):
                k = str(row[0] or "").strip().upper()
                v = str(row[1] or "").strip() if row[1] else ""
                if k and k != "KEY":
                    cfg_email[k] = v

        outlook = win32com.client.Dispatch("Outlook.Application")
        campaign_id = f"AUTO_RATE_{datetime.now():%Y%m%d_%H%M}"

        intro = cfg_email.get("INTROTEXT", "")
        closing = cfg_email.get("CLOSINGTEXT", "")
        signature = cfg_email.get("SIGNATURE", "")

        # Subject
        import random
        from datetime import date
        templates = cfg_email.get("SUBJECTTEMPLATES", "Freight Rate Update")
        suffix = cfg_email.get("SUBJECTSUFFIX", "NELSON")
        week = date.today().isocalendar()[1]
        items = [x.strip() for x in templates.split("|") if x.strip()]
        base_subject = random.choice(items) if items else "Freight Rate Update"
        subject = req.subject_override or f"{base_subject} // {suffix} WEEK {week}"

        # Profile PDF attachment
        profile_pdf = _EMAIL_ENGINE / "assets" / "PUDONG PRIME PROFILE.pdf"
        logo_png = _EMAIL_ENGINE / "assets" / "logo.png"

        sent = 0
        failed = []

        for r in results:
            try:
                mail = outlook.CreateItem(0)
                mail.To = r["email"]
                mail.Subject = subject

                if profile_pdf.exists():
                    mail.Attachments.Add(str(profile_pdf))
                if logo_png.exists():
                    logo = mail.Attachments.Add(str(logo_png))
                    logo.PropertyAccessor.SetProperty(
                        "http://schemas.microsoft.com/mapi/proptag/0x3712001F",
                        "pudonglogo",
                    )

                mail.HTMLBody = f"""
                <html><body>
                Dear {r['pic']},<br><br>
                {intro}<br><br>
                {r['html']}<br><br>
                {closing}<br><br>
                {signature}
                </body></html>
                """
                mail.Send()
                sent += 1
                log.info("SENT -> %s (%d routes)", r["email"], r["routes"])

            except Exception as e:
                log.error("FAILED -> %s: %s", r["email"], e)
                failed.append({"email": r["email"], "error": str(e)})

        # Log to CSV
        import csv
        log_file = _EMAIL_ENGINE / "logs" / "email_log.csv"
        log_file.parent.mkdir(exist_ok=True)
        log_exists = log_file.exists()
        with open(log_file, "a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if not log_exists:
                w.writerow(["timestamp", "email", "subject", "campaign_id", "cycle_id", "status"])
            for r in results:
                status = "SENT" if r["email"] not in [f["email"] for f in failed] else "FAILED"
                w.writerow([
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    r["email"], subject, campaign_id, "1", status,
                ])

        return {
            "success": True,
            "dry_run": False,
            "campaign_id": campaign_id,
            "sent": sent,
            "failed_count": len(failed),
            "failed": failed,
            "total_with_rates": len(results),
            "no_rates": no_rates,
            "subject": subject,
        }

    except Exception as e:
        log.error("Send error: %s", e)
        return {"error": str(e), "success": False}


# ══════════════════════════════════════════════════════════════════════════════
# 4. CUSTOMERS — GET /api/auto-quote/customers
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/customers")
def get_customers(cmd: Optional[str] = Query(None)):
    """
    List customers from data.xlsx with their routes.

    Used by WebApp to populate customer picker before sending quotes.
    """
    import pandas as pd
    data_file = _EMAIL_ENGINE / "data.xlsx"
    if not data_file.exists():
        return {"customers": [], "error": "data.xlsx not found"}

    df = pd.read_excel(data_file)
    df.columns = df.columns.str.strip().str.upper()

    if cmd:
        df = df[df["CMD_NAME"].str.upper() == cmd.upper()]

    # Group by CMD
    cmds = {}
    for _, row in df.iterrows():
        cmd_name = str(row.get("CMD_NAME", "")).strip()
        if cmd_name not in cmds:
            cmds[cmd_name] = {"cmd": cmd_name, "customers": [], "count": 0}

        email = str(row.get("CNEE_EMAIL", "")).strip()
        if email and "@" in email:
            cmds[cmd_name]["customers"].append({
                "email": email,
                "company": str(row.get("CNEE_NAME", "")).strip(),
                "pic": str(row.get("CNEE_PIC", "")).strip(),
                "pol": str(row.get("POL", "HPH")).strip(),
                "destination": str(row.get("DESTINATION", "")).strip(),
            })
            cmds[cmd_name]["count"] += 1

    return {
        "cmd_groups": list(cmds.values()),
        "total_cmds": len(cmds),
        "total_customers": sum(c["count"] for c in cmds.values()),
    }


# ══════════════════════════════════════════════════════════════════════════════
# 5. CONFIG — GET /api/auto-quote/config
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/config")
def get_email_config():
    """Get email template configuration (subject, intro, closing, signature)."""
    config_file = _DATA_DIR / "config.xlsx"
    if not config_file.exists():
        return {"error": "config.xlsx not found", "config": {}}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(config_file), data_only=True)
        ws = wb.active
        cfg = {}
        for row in ws.iter_rows(max_col=2, values_only=True):
            k = str(row[0] or "").strip().upper()
            v = str(row[1] or "").strip() if row[1] else ""
            if k and k != "KEY":
                cfg[k] = v

        return {
            "config": {
                "subject_templates": cfg.get("SUBJECTTEMPLATES", ""),
                "subject_suffix": cfg.get("SUBJECTSUFFIX", "NELSON"),
                "intro_text": cfg.get("INTROTEXT", ""),
                "closing_text": cfg.get("CLOSINGTEXT", ""),
                "preheader": cfg.get("PREHEADER", ""),
                "has_signature": bool(cfg.get("SIGNATURE", "")),
            },
            "total_keys": len(cfg),
        }
    except Exception as e:
        return {"error": str(e), "config": {}}
