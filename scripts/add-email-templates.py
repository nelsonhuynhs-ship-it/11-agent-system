"""
add-email-templates.py
=======================
Add IntroTemplates (10 variants) + ClosingTemplates (5 variants) to config.xlsx.
Pipe-separated like existing SubjectTemplates.

Designed per cold-email skill principles (coreyhaines31/marketingskills@cold-email):
- Peer-to-peer tone (not "I hope this finds you well")
- Observation → Problem → Proof → Ask
- "You/your" dominates over "I/we"
- One low-friction CTA
- Contractions, conversational

Placeholder tokens available (per template_renderer.py):
  {{first_name}}   - PIC (defaults to "Team")
  {{company}}      - CNEE company name
  {{typical_pol}}  - HPH or HCM
  {{typical_dest}} - USLAX, USNYC, etc
  {{week}}         - ISO week number
"""
from __future__ import annotations
import openpyxl
from pathlib import Path

CONFIG = Path("D:/OneDrive/NelsonData/email/config.xlsx")
LOCAL_CONFIG = Path("D:/NELSON/2. Areas/Engine_test/email_engine/data/config.xlsx")

# ──────────────────────────────────────────────────────────────────
# 10 INTRO templates — unique angles, Nelson freight context
# ──────────────────────────────────────────────────────────────────
INTRO_TEMPLATES = [
    # 1. Peer observation — for importers with Panjiva history
    "Hi {{first_name}},\n"
    "Saw {{company}} moving containers out of {{typical_pol}} recently — "
    "wanted to put current {{typical_dest}} levels in front of you for Week {{week}}.",

    # 2. Problem-first — rate volatility
    "Hi {{first_name}},\n"
    "Asia-US rates moved this week — if {{company}} is planning bookings out of "
    "{{typical_pol}}, here's what's live for Week {{week}}:",

    # 3. Proof-first — VIP tier, authority
    "Hi {{first_name}},\n"
    "Moved 400+ TEU on {{typical_pol}}→{{typical_dest}} last quarter with 99% on-time. "
    "Sharing this week's levels since {{company}} ships similar lanes:",

    # 4. Ultra-short — for C-suite, high-tier
    "Hi {{first_name}},\n"
    "Quick one — current {{typical_pol}}→{{typical_dest}} rates below, valid Week {{week}}:",

    # 5. Question-first — engagement
    "Hi {{first_name}},\n"
    "Curious how {{company}} is handling Q2 capacity on Asia-US lanes. Attached this "
    "week's rate + space view for {{typical_pol}} corridor in case useful:",

    # 6. Urgency — time-sensitive, booking push
    "Hi {{first_name}},\n"
    "Space on {{typical_pol}}→{{typical_dest}} tightening end of Week {{week}}. "
    "If {{company}} has containers to move, these are current options:",

    # 7. Lane-specific movement
    "Hi {{first_name}},\n"
    "{{typical_dest}} rates shifted $200+/40HQ this week. Sending to {{company}} since "
    "you've been shipping this corridor:",

    # 8. Follow-up style — re-engagement, busy inbox
    "Hi {{first_name}},\n"
    "Know Week {{week}} is busy — 1-line version: {{typical_pol}}→{{typical_dest}} "
    "is bookable this week at levels below:",

    # 9. Seasonal — peak planning
    "Hi {{first_name}},\n"
    "Peak season planning hits most importers late April. Thought this week's "
    "{{typical_pol}} rate reference might help {{company}}'s team get ahead:",

    # 10. Value-add — certification + fee waive
    "Hi {{first_name}},\n"
    "Pudong Prime is CTPAT-certified + $65 POD fee waived through end of April for "
    "new lanes. Here's {{company}}'s rate context on {{typical_pol}}→{{typical_dest}}:",
]

# ──────────────────────────────────────────────────────────────────
# 5 CLOSING templates — varied CTAs, low friction
# ──────────────────────────────────────────────────────────────────
CLOSING_TEMPLATES = [
    # 1. Soft CTA — reference only
    "Happy to run custom lanes if {{company}} has anything specific. Otherwise, "
    "just a heads up for Week {{week}}.\n"
    "Reply with the lane + POD and I'll send a formal quote.",

    # 2. Direct ask — space hold
    "Want me to hold space at these levels? One-line reply works — I'll "
    "send the booking form + SOW back within 2 hours.",

    # 3. No-ask reference
    "No reply needed if everything's covered for Week {{week}} — filing this "
    "away as your weekly reference.\n"
    "If you want formal pricing with 2-4 week validity, just say so.",

    # 4. Open question — diagnostic
    "What's {{company}}'s biggest pain on Asia-US lanes right now — rate, space, "
    "or transit? One line back and I'll match what we have against it.",

    # 5. Next-step hint — SOW-focused
    "If {{company}} needs a formal SOW with 2-4 week validity for Q2 planning, "
    "1-line reply and I'll send the booking sheet with validated space.",
]


def add_templates():
    # Main config at OneDrive
    wb = openpyxl.load_workbook(str(CONFIG))
    ws = wb.active

    intro_str = " | ".join(INTRO_TEMPLATES)
    closing_str = " | ".join(CLOSING_TEMPLATES)

    # Check existing keys
    existing = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=False), 1):
        key = str(row[0].value or "").strip()
        if key:
            existing[key] = i

    # Upsert IntroTemplates
    if "IntroTemplates" in existing:
        ws.cell(row=existing["IntroTemplates"], column=2).value = intro_str
        print(f"Updated existing IntroTemplates row {existing['IntroTemplates']}")
    else:
        ws.cell(row=ws.max_row + 1, column=1).value = "IntroTemplates"
        ws.cell(row=ws.max_row, column=2).value = intro_str
        print(f"Added IntroTemplates at row {ws.max_row}")

    # Upsert ClosingTemplates
    if "ClosingTemplates" in existing:
        ws.cell(row=existing["ClosingTemplates"], column=2).value = closing_str
        print(f"Updated existing ClosingTemplates row {existing['ClosingTemplates']}")
    else:
        ws.cell(row=ws.max_row + 1, column=1).value = "ClosingTemplates"
        ws.cell(row=ws.max_row, column=2).value = closing_str
        print(f"Added ClosingTemplates at row {ws.max_row}")

    wb.save(str(CONFIG))
    # Also copy to local (gitignored, code reads from local)
    import shutil
    if LOCAL_CONFIG.exists():
        shutil.copy2(str(CONFIG), str(LOCAL_CONFIG))
        print(f"Synced to {LOCAL_CONFIG}")

    print()
    print(f"✅ Saved {len(INTRO_TEMPLATES)} intro + {len(CLOSING_TEMPLATES)} closing templates.")
    print(f"   OneDrive: {CONFIG}")
    print(f"   Local:    {LOCAL_CONFIG}")


if __name__ == "__main__":
    add_templates()
