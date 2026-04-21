# -*- coding: utf-8 -*-
"""
setup-booking-pool-sheet.py — One-shot script to add "Booking Pool" sheet to ERP_Master_v14.xlsm.

Run once:
    python scripts/setup-booking-pool-sheet.py

Steps:
  1. Check if "Booking Pool" sheet already exists → skip if yes
  2. Load xlsm with keep_vba=True
  3. Create "Booking Pool" sheet
  4. Write header row (row 1) using POOL_COLS keys
  5. Style header: bold, blue fill, white text, center-aligned, border
  6. Freeze row 1 (pane split below header)
  7. Auto-filter A1:T1
  8. Set column widths from POOL_COL_WIDTHS
  9. Save
 10. Re-inject customUI XML (openpyxl strips it on save)
 11. Print result
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)

# ── Paths ──────────────────────────────────────────────────────────────────
_SCRIPT_DIR = Path(__file__).resolve().parent
_REPO_ROOT = _SCRIPT_DIR.parent

# Add repo root so we can import Pricing_Engine modules
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from Pricing_Engine.booking_pool_schema import (  # noqa: E402
    POOL_COL_WIDTHS,
    POOL_COLS,
    POOL_SHEET_NAME,
)

ERP = Path("D:/OneDrive/NelsonData/erp/ERP_Master_v14.xlsm")
CUSTOMUI_XML = Path("D:/OneDrive/NelsonData/erp/CustomUI_v14.xml")
CUSTOMUI_UTILS = Path("D:/OneDrive/NelsonData/erp")

# ── Style constants ─────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="1F6AB0")   # Nelson blue
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
_BORDER_SIDE = Side(style="thin", color="D0D7E5")
_HEADER_BORDER = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE, bottom=_BORDER_SIDE,
)


def _apply_header_style(cell) -> None:
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.alignment = _HEADER_ALIGN
    cell.border = _HEADER_BORDER


def _col_letter(col_idx: int) -> str:
    """Convert 1-based column index to Excel letter (A, B, … T)."""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def main() -> int:
    """Return 0 on success, 1 on error."""
    if not ERP.exists():
        print(f"[ERROR] ERP file not found: {ERP}")
        return 1

    # ── Step 1: Check if sheet already exists ──────────────────────────────
    wb = openpyxl.load_workbook(str(ERP), keep_vba=True)
    if POOL_SHEET_NAME in wb.sheetnames:
        print(f"Sheet '{POOL_SHEET_NAME}' already exists, skipped.")
        wb.close()
        return 0

    print(f"[setup] Creating '{POOL_SHEET_NAME}' sheet in {ERP.name} ...")

    # ── Step 2: Create sheet ────────────────────────────────────────────────
    ws = wb.create_sheet(title=POOL_SHEET_NAME)

    # ── Step 3: Write header row ────────────────────────────────────────────
    # POOL_COLS is ordered by column index value; sort for safety
    headers = sorted(POOL_COLS.items(), key=lambda kv: kv[1])  # [(name, idx), ...]
    for col_name, col_idx in headers:
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        _apply_header_style(cell)

    # ── Step 4: Set row height for header ──────────────────────────────────
    ws.row_dimensions[1].height = 22

    # ── Step 5: Freeze pane below header ───────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Step 6: Auto-filter on header row ──────────────────────────────────
    last_col_letter = _col_letter(len(POOL_COLS))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

    # ── Step 7: Column widths ───────────────────────────────────────────────
    for col_name, width in POOL_COL_WIDTHS.items():
        col_idx = POOL_COLS.get(col_name)
        if col_idx is None:
            continue
        letter = _col_letter(col_idx)
        ws.column_dimensions[letter].width = width

    # ── Step 8: Tab color (blue-ish to match Active Jobs family) ───────────
    ws.sheet_properties.tabColor = "1F6AB0"

    # ── Step 9: Save xlsm ──────────────────────────────────────────────────
    wb.save(str(ERP))
    wb.close()
    print(f"[setup] Sheet '{POOL_SHEET_NAME}' created and saved.")

    # ── Step 10: Re-inject customUI XML (openpyxl strips it on save) ───────
    if str(CUSTOMUI_UTILS) not in sys.path:
        sys.path.insert(0, str(CUSTOMUI_UTILS))

    try:
        from customui_utils import ensure_customui  # type: ignore[import-untyped]
        result = ensure_customui(str(ERP), str(CUSTOMUI_XML))
        if result.get("injected"):
            print("[setup] CustomUI ribbon re-injected successfully.")
        elif result.get("already_ok"):
            print("[setup] CustomUI ribbon already intact — no action needed.")
        elif result.get("error"):
            print(f"[WARNING] CustomUI inject failed: {result['error']}")
            print("  Ribbon may be missing tabs. Run customui_utils.py manually.")
    except ImportError:
        print("[WARNING] customui_utils not found — ribbon tabs may be missing.")
        print(f"  Expected: {CUSTOMUI_UTILS / 'customui_utils.py'}")

    print("[setup] Done. Open ERP_Master_v14.xlsm and verify 'Booking Pool' tab.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
