"""
erp-s1v2-column-polish.py — Column cleanup + grouping for Quotes sheet (S1 v2)
===============================================================================
Idempotent script. Safe to re-run anytime.

Actions on Quotes sheet:
  1. Hide column B  (Date)
  2. Hide column StatusDate (lookup by header name, currently col 38)
  3. Hide column JobID      (lookup by header name, currently col 41)
  4. Set outline groups (collapse/expand):
       Buy  group  cols 12-18  (Buy_20GP .. Buy_40RF)   outline_level=1
       Mar  group  cols 19-25  (Mar_20GP .. Mar_40RF)   outline_level=1
       PUC  group  cols 26-28                           outline_level=1
       Sell group  cols 29-35  (Sell_20GP .. Sell_40RF) outline_level=1
  5. summaryRight = False (summary/total col on LEFT side per Nelson's layout)

Saves via save_preserving_ribbon (gotcha #6 — never wb.save() directly).
Creates backup ERP_Master_v14.backup_YYYYMMDD_HHMMSS.xlsm before writing.

Usage:
    python scripts/erp-s1v2-column-polish.py
    python scripts/erp-s1v2-column-polish.py --dry-run
    python scripts/erp-s1v2-column-polish.py --erp "D:/OneDrive/NelsonData/erp/ERP_Master_v14.xlsm"
"""
from __future__ import annotations

import argparse
import os
import shutil
import sys
from datetime import datetime

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]

# ── repo root on sys.path so ERP.core imports work ──
_REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, _REPO_ROOT)

from ERP.core.ribbon_guard import save_preserving_ribbon  # noqa: E402

DEFAULT_ERP_FILE = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"

# Outline groups: (first_col_idx, last_col_idx, group_label)
# openpyxl col indices are 1-based
OUTLINE_GROUPS = [
    (12, 18, "Buy"),   # Buy_20GP..Buy_40RF
    (19, 25, "Mar"),   # Mar_20GP..Mar_40RF
    (26, 28, "PUC"),   # PUC group
    (29, 35, "Sell"),  # Sell_20GP..Sell_40RF
]


def col_letter(n: int) -> str:
    """Convert 1-based column index to Excel letter (A-ZZ)."""
    result = ""
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


QUOTES_HEADER_ROW = 4  # Row 1-3 are KPI dashboard rows; actual headers at row 4


def find_col_by_header(ws, header_name: str) -> int | None:
    """Find 1-based column index by header value in the Quotes header row.

    Scans rows 1-6 to be resilient to minor layout changes.
    """
    for row_idx in range(1, 7):
        for cell in ws[row_idx]:
            if cell.value and str(cell.value).strip().upper() == header_name.upper():
                return cell.column
    return None


def check_file_lock(erp_path: str) -> bool:
    """Return True if file is locked by another process (Excel has it open)."""
    try:
        with open(erp_path, "r+b"):
            return False  # opened fine → not locked
    except PermissionError:
        return True  # locked


def backup_xlsm(erp_path: str) -> str:
    """Create timestamped backup alongside the original. Returns backup path."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(erp_path)
    backup_path = f"{base}.backup_{ts}{ext}"
    shutil.copy2(erp_path, backup_path)
    return backup_path


def apply_column_polish(ws, dry_run: bool = False) -> list[str]:
    """Apply all column hide + outline changes. Returns list of change descriptions."""
    changes: list[str] = []

    # ── 1. Hide column B (Date) ──
    if not ws.column_dimensions["B"].hidden:
        if not dry_run:
            ws.column_dimensions["B"].hidden = True
        changes.append("HIDE col B (Date)")
    else:
        changes.append("SKIP col B already hidden")

    # ── 2. Hide StatusDate (lookup by header) ──
    sd_col = find_col_by_header(ws, "StatusDate")
    if sd_col is not None:
        letter = col_letter(sd_col)
        if not ws.column_dimensions[letter].hidden:
            if not dry_run:
                ws.column_dimensions[letter].hidden = True
            changes.append(f"HIDE col {letter} (StatusDate, col {sd_col})")
        else:
            changes.append(f"SKIP StatusDate col {letter} already hidden")
    else:
        changes.append("WARN: StatusDate header not found in row 1 — skipped")

    # ── 3. Hide JobID (lookup by header) ──
    jid_col = find_col_by_header(ws, "JobID")
    if jid_col is not None:
        letter = col_letter(jid_col)
        if not ws.column_dimensions[letter].hidden:
            if not dry_run:
                ws.column_dimensions[letter].hidden = True
            changes.append(f"HIDE col {letter} (JobID, col {jid_col})")
        else:
            changes.append(f"SKIP JobID col {letter} already hidden")
    else:
        changes.append("WARN: JobID header not found in row 1 — skipped")

    # ── 4. Column grouping (outline levels) ──
    # Use openpyxl range-based .group() API which writes proper
    # <col min=X max=Y outlineLevel=1/> range entries — iterating
    # per-col + setting .outline_level only persisted the first col
    # (serialization collapses runs into ranges).
    for start, end, label in OUTLINE_GROUPS:
        start_letter = col_letter(start)
        end_letter = col_letter(end)
        # Idempotency: if start col already has outline_level=1, skip
        cd = ws.column_dimensions.get(start_letter)
        if cd and cd.outline_level == 1:
            changes.append(f"SKIP {label} group {start_letter}-{end_letter} already grouped")
            continue
        if not dry_run:
            ws.column_dimensions.group(start_letter, end_letter, outline_level=1)
        changes.append(
            f"GROUP {label}: cols {start_letter}{start}-{end_letter}{end} outline_level=1"
        )

    # ── 5. summaryRight = False ──
    try:
        current_sr = ws.sheet_properties.outlinePr.summaryRight
        if current_sr is not False:
            if not dry_run:
                ws.sheet_properties.outlinePr.summaryRight = False
            changes.append("SET outlinePr.summaryRight = False (summary on left)")
        else:
            changes.append("SKIP summaryRight already False")
    except AttributeError:
        # openpyxl may not expose outlinePr on all versions
        if not dry_run:
            try:
                ws.sheet_properties.outlinePr.summaryRight = False
                changes.append("SET outlinePr.summaryRight = False")
            except Exception as exc:
                changes.append(f"WARN: could not set summaryRight: {exc}")
        else:
            changes.append("DRY-RUN: would set summaryRight = False")

    return changes


def main() -> int:
    ap = argparse.ArgumentParser(
        description="ERP S1v2 — Column polish for Quotes sheet (hide + outline groups)"
    )
    ap.add_argument("--erp", default=DEFAULT_ERP_FILE,
                    help=f"Path to ERP xlsm (default: {DEFAULT_ERP_FILE})")
    ap.add_argument("--dry-run", action="store_true",
                    help="Print planned changes without writing to file")
    args = ap.parse_args()

    erp_path = os.path.normpath(args.erp)
    print(f"[+] ERP Column Polish (S1v2)")
    print(f"    erp  : {erp_path}")
    print(f"    mode : {'DRY-RUN' if args.dry_run else 'LIVE'}")

    # ── File exists check ──
    if not os.path.exists(erp_path):
        print(f"[ERROR] ERP file not found: {erp_path}")
        return 1

    # ── File-lock check ──
    if check_file_lock(erp_path):
        print("[ERROR] ERP file is locked (Excel has it open). Close Excel first.")
        return 2

    # ── Backup (skip in dry-run) ──
    if not args.dry_run:
        backup = backup_xlsm(erp_path)
        print(f"    backup: {backup}")
    else:
        print("    backup: (skipped — dry-run)")

    # ── Load workbook ──
    print("    loading workbook (keep_vba=True)...")
    wb = openpyxl.load_workbook(erp_path, keep_vba=True)

    if "Quotes" not in wb.sheetnames:
        print("[ERROR] 'Quotes' sheet not found in workbook.")
        wb.close()
        return 3

    ws = wb["Quotes"]
    print(f"    Quotes sheet: max_col={ws.max_column}  max_row={ws.max_row}")

    # ── Apply changes ──
    changes = apply_column_polish(ws, dry_run=args.dry_run)

    print("\n    Changes:")
    for c in changes:
        prefix = "  [DRY]" if args.dry_run else "  [OK] "
        print(f"{prefix} {c}")

    if args.dry_run:
        print("\n[DRY-RUN] No file written.")
        wb.close()
        return 0

    # ── Save via ribbon guard (gotcha #6) ──
    print("\n    saving via save_preserving_ribbon...")
    result = save_preserving_ribbon(wb, erp_path)
    wb.close()
    print(f"[OK] Saved: {erp_path}  ribbon={result}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
