# -*- coding: utf-8 -*-
"""
si-48h-alert.py — Daily SI cutoff reminder for Nelson.

Scans Active Jobs. For each row where SI_CutOff is within next 48h AND
Docs stage not yet reached (tracking dots < stage 4), fire Telegram alert.

Deduplication: don't re-alert same BKG within 24h.

Runs daily at 08:05 via Windows Task Scheduler (see register-si-alert-task.ps1).

Env vars (same as notify-telegram.py):
    BOT_TOKEN       — Telegram bot token
    ADMIN_CHAT_ID   — Nelson's Telegram user ID

Usage:
    python scripts/si-48h-alert.py        # run once
    python scripts/si-48h-alert.py --test  # dry-run, no Telegram send
"""
from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
import tempfile
import urllib.parse
import urllib.request
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError as e:
    print(f"Missing dep: {e}  ->  pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ERP_PATH = Path("D:/OneDrive/NelsonData/erp/ERP_Master_v14.xlsm")
DEDUP_FILE = (
    Path(__file__).parent.parent / "email_engine" / "data" / "si_alert_dedup.json"
)

# ---------------------------------------------------------------------------
# Telegram (reuse same env vars as notify-telegram.py)
# ---------------------------------------------------------------------------
BOT_TOKEN = os.environ.get("BOT_TOKEN", "").strip()
CHAT_ID = os.environ.get("ADMIN_CHAT_ID", "").strip()


def send_telegram(text: str) -> bool:
    """Send message to Nelson's Telegram. DISABLED 2026-04-26 — no-op."""
    return True
    if not BOT_TOKEN or not CHAT_ID:  # noqa: unreachable
        print(f"[no-telegram — BOT_TOKEN/ADMIN_CHAT_ID not set] {text}")
        return False
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    data = urllib.parse.urlencode(
        {
            "chat_id": CHAT_ID,
            "text": text[:4000],
            "parse_mode": "Markdown",
        }
    ).encode()
    req = urllib.request.Request(url, data=data)
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            resp = json.loads(r.read())
            return bool(resp.get("ok"))
    except Exception as e:
        print(f"Telegram fail: {e}", file=sys.stderr)
        return False


# ---------------------------------------------------------------------------
# Dedup
# ---------------------------------------------------------------------------
def load_dedup() -> dict:
    if not DEDUP_FILE.exists():
        return {}
    try:
        return json.loads(DEDUP_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_dedup(d: dict) -> None:
    DEDUP_FILE.parent.mkdir(parents=True, exist_ok=True)
    DEDUP_FILE.write_text(
        json.dumps(d, indent=2, ensure_ascii=False), encoding="utf-8"
    )


# ---------------------------------------------------------------------------
# Date parsing
# ---------------------------------------------------------------------------
_DT_FORMATS = (
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%d/%m/%Y %H:%M",
    "%d/%m/%Y",
    "%d/%m %H:%M",
)


def parse_dt(val) -> Optional[datetime]:
    """Parse datetime from cell — accepts datetime obj, string ISO, or DD/MM HH:MM."""
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in _DT_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------
def find_header_row(ws) -> int:
    """Find the row containing ERP column headers. Returns 0 if not found."""
    keywords = {"CUSTOMER", "HBL_NO", "BKG_NO", "BOOKING"}
    for r in range(1, 8):
        for c in range(1, 60):
            v = ws.cell(row=r, column=c).value
            if v and str(v).strip().upper() in keywords:
                return r
    return 0


def find_col_by_name(ws, header_row: int, target: str) -> int:
    """Find 1-based column index by header name. Returns 0 if not found."""
    t = target.upper().strip()
    for c in range(1, 60):
        v = ws.cell(row=header_row, column=c).value
        if v and str(v).strip().upper() == t:
            return c
    return 0


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def run_alert(dry_run: bool = False) -> dict:
    """Scan Active Jobs, alert SI cutoff < 48h via Telegram.

    Returns dict with stats:
        {status, alerts_sent, rows_checked, errors}

    Callable from outlook_scanner.py as sub-job, or from main() CLI wrapper.
    """
    if not ERP_PATH.exists():
        return {"status": "error", "error": f"ERP not found: {ERP_PATH}"}

    # Copy xlsm to tmp — file may be open in Excel (read-only lock)
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsm")
    os.close(fd)
    tmp = Path(tmp_path)
    try:
        shutil.copy2(ERP_PATH, tmp)
        wb = openpyxl.load_workbook(
            tmp, read_only=True, data_only=True, keep_vba=False
        )
    except Exception as e:
        tmp.unlink(missing_ok=True)
        return {"status": "error", "error": f"Cannot read xlsm: {e}"}

    if "Active Jobs" not in wb.sheetnames:
        wb.close()
        tmp.unlink(missing_ok=True)
        return {"status": "error", "error": "Active Jobs sheet missing"}

    ws = wb["Active Jobs"]

    header_row = find_header_row(ws)
    if not header_row:
        wb.close()
        tmp.unlink(missing_ok=True)
        return {"status": "error", "error": "Cannot find header row in Active Jobs"}

    # Detect column positions by header name
    col_bkg = find_col_by_name(ws, header_row, "BKG_NO")
    col_si = find_col_by_name(ws, header_row, "SI_CUTOFF")
    col_customer = find_col_by_name(ws, header_row, "CUSTOMER")
    col_tracking = find_col_by_name(ws, header_row, "TRACKING")

    if not col_bkg:
        col_bkg = find_col_by_name(ws, header_row, "BOOKING")
    if not (col_bkg and col_si and col_customer):
        wb.close()
        tmp.unlink(missing_ok=True)
        return {
            "status": "error",
            "error": f"Missing cols bkg={col_bkg} si={col_si} cust={col_customer}",
        }

    now = datetime.now()
    threshold = now + timedelta(hours=48)
    dedup = load_dedup()
    alerts_sent = 0
    rows_checked = 0

    for r in range(header_row + 1, ws.max_row + 1):
        bkg_raw = ws.cell(row=r, column=col_bkg).value
        if not bkg_raw:
            continue
        bkg = str(bkg_raw).strip()
        if not bkg:
            continue

        rows_checked += 1

        si_raw = ws.cell(row=r, column=col_si).value
        si_dt = parse_dt(si_raw)
        if si_dt is None:
            continue

        # Only alert when SI cutoff is within [now, now+48h]
        if not (now <= si_dt <= threshold):
            continue

        customer = str(
            ws.cell(row=r, column=col_customer).value or "?"
        ).strip()

        # Check tracking stage — if >= 4 green dots (Docs done), skip
        if col_tracking:
            tracking_val = ws.cell(row=r, column=col_tracking).value or ""
            done_stages = str(tracking_val).count("\u25cf")  # ● filled circle
        else:
            done_stages = 0

        if done_stages >= 4:
            continue  # Docs stage already reached

        # Dedup: one alert per BKG per calendar day
        dedup_key = f"{bkg}_{now.strftime('%Y%m%d')}"
        if dedup_key in dedup:
            continue

        hours_left = int((si_dt - now).total_seconds() / 3600)
        msg = (
            f"\u26a0 *SI Alert \u2014 {customer}*\n"
            f"BKG: `{bkg}`\n"
            f"SI cutoff: {si_dt.strftime('%d/%m %H:%M')}\n"
            f"C\u00f2n: *{hours_left}h*\n"
            f"\u2192 Check plan kh\u00e1ch tr\u00e1nh cancel c\u1eadn gi\u1edd"
        )

        if dry_run:
            print(f"[dry-run] Would alert: {bkg} ({customer}) SI in {hours_left}h")
            alerts_sent += 1
        else:
            if send_telegram(msg):
                dedup[dedup_key] = now.isoformat()
                alerts_sent += 1

    wb.close()
    tmp.unlink(missing_ok=True)

    if not dry_run:
        save_dedup(dedup)

    print(
        f"SI 48h alert: {alerts_sent} sent, {rows_checked} rows checked"
        f" @ {now.strftime('%Y-%m-%d %H:%M')}"
    )
    return {
        "status": "ok",
        "alerts_sent": alerts_sent,
        "rows_checked": rows_checked,
    }


def main() -> int:
    """CLI wrapper for standalone / manual debug."""
    ap = argparse.ArgumentParser(description="48h SI cutoff alert")
    ap.add_argument("--test", action="store_true", help="Dry-run, no Telegram send")
    args = ap.parse_args()
    result = run_alert(dry_run=args.test)
    if result.get("status") == "error":
        print(f"Error: {result.get('error')}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
