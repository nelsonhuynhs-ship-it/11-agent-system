"""Migrate Markup_Store sheet from v1 (per-carrier) to v2 (per-carrier-lane).

Phase 5 of plans/260411-2121-erp-workflow-upgrade.

Current schema:
    Carrier | Mar20GP | Mar40GP | Mar40HC | Mar45HC | Mar40NOR | Mar20RF | Mar40RF

New schema:
    Carrier | Lane | Mar20GP | Mar40GP | Mar40HC | Mar45HC | Mar40NOR | Mar20RF | Mar40RF

Where Lane is one of {WC, EC, GULF, *}. All pre-existing rows get Lane = "*"
(default), so the updated VBA LoadMarkupForCarrier will still find them as
fallback when no exact lane match exists.

Usage:
    python scripts/migrate-markup-store.py                  # live migrate
    python scripts/migrate-markup-store.py --dry-run        # report only
    python scripts/migrate-markup-store.py --xlsm /tmp/x.xlsm

The script is idempotent: if col 2 header is already "Lane" it exits with
no changes. Always backs up the target xlsm to
    D:/OneDrive/NelsonData/pricing/_backup/pre-p5-markup-migration/<ts>/
before writing, unless --dry-run.
"""
from __future__ import annotations

import argparse
import shutil
import sys
import time
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl not installed — pip install openpyxl\n")
    sys.exit(2)


DEFAULT_XLSM = Path("D:/OneDrive/NelsonData/erp/ERP_Master_v14.xlsm")
BACKUP_ROOT = Path(
    "D:/OneDrive/NelsonData/pricing/_backup/pre-p5-markup-migration"
)
SHEET_NAME = "Markup_Store"
LANE_HEADER = "Lane"
DEFAULT_LANE = "*"


def backup_xlsm(xlsm: Path) -> Path:
    ts = time.strftime("%Y%m%d-%H%M%S")
    dst_dir = BACKUP_ROOT / ts
    dst_dir.mkdir(parents=True, exist_ok=True)
    dst = dst_dir / xlsm.name
    shutil.copy2(xlsm, dst)
    return dst


def already_migrated(ws) -> bool:
    hdr = ws.cell(1, 2).value
    return isinstance(hdr, str) and hdr.strip() == LANE_HEADER


def migrate(xlsm: Path, dry_run: bool) -> int:
    if not xlsm.exists():
        sys.stderr.write(f"ERROR: xlsm not found: {xlsm}\n")
        return 2

    print(f"Target xlsm: {xlsm}")
    wb = openpyxl.load_workbook(xlsm, keep_vba=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.stderr.write(
            f"ERROR: sheet '{SHEET_NAME}' not found in {xlsm}\n"
            f"Available sheets: {wb.sheetnames}\n"
        )
        return 2

    ws = wb[SHEET_NAME]

    if already_migrated(ws):
        print(
            f"Already migrated — col 2 header is '{LANE_HEADER}'. Nothing to do."
        )
        return 0

    rows_before = ws.max_row
    cols_before = ws.max_column
    print(f"Before: {rows_before} rows, {cols_before} cols")
    print(f"Col 1 header: {ws.cell(1, 1).value!r}")
    print(f"Col 2 header (old): {ws.cell(1, 2).value!r}")

    # Count data rows that will be tagged with default lane
    data_rows = 0
    for r in range(2, rows_before + 1):
        if ws.cell(r, 1).value not in (None, ""):
            data_rows += 1
    print(f"Data rows to tag with Lane='{DEFAULT_LANE}': {data_rows}")

    if dry_run:
        print("DRY-RUN — no changes written, no backup created.")
        print(
            "Would: insert col 2 'Lane', set header to 'Lane', "
            f"fill {data_rows} existing rows with '{DEFAULT_LANE}', save xlsm."
        )
        return 0

    # Live path — back up first
    bkp = backup_xlsm(xlsm)
    print(f"Backup: {bkp}")

    ws.insert_cols(2)
    ws.cell(1, 2).value = LANE_HEADER
    tagged = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value not in (None, ""):
            ws.cell(r, 2).value = DEFAULT_LANE
            tagged += 1

    wb.save(xlsm)
    print(
        f"After: {ws.max_row} rows, {ws.max_column} cols — tagged {tagged} rows."
    )
    print("Migration done. Re-import updated VBA modules via "
          "`python scripts/reimport-erp-vba.py`.")
    return 0


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        description="Migrate Markup_Store from per-carrier to per-(carrier,lane)."
    )
    p.add_argument(
        "--xlsm", type=Path, default=DEFAULT_XLSM,
        help=f"xlsm path (default: {DEFAULT_XLSM})",
    )
    p.add_argument(
        "--dry-run", action="store_true",
        help="Report what would change without saving.",
    )
    args = p.parse_args(argv)
    return migrate(args.xlsm, args.dry_run)


if __name__ == "__main__":
    raise SystemExit(main())
