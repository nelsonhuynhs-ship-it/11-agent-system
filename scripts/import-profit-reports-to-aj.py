"""
Import historical Sales Profit reports -> ERP Archive + Active Jobs.

Parses 11 sheets (MAY 2025 -> MAR 2026) from:
  D:/OneDrive/NelsonData/erp/monthly_reports/SALES PROFIT - MAR 2026 - NELSON CHINH.xlsx

Split logic:
  - ETA > today  -> Active Jobs (In Transit)
  - ETA <= today -> Archive (Delivered)

Writes to existing columns only. No new columns added.

Usage:
  python scripts/import-profit-reports-to-aj.py           # dry-run
  python scripts/import-profit-reports-to-aj.py --live    # commit
"""
import sys
import io
import argparse
import shutil
import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl

SRC = Path(r"D:/OneDrive/NelsonData/erp/monthly_reports/SALES PROFIT - MAR 2026 - NELSON CHINH.xlsx")
ERP = Path(r"D:/OneDrive/NelsonData/erp/ERP_Master_v14.xlsm")

TARGET_SHEETS = [
    "MAY 2025", "JUN 2025", "JUL 2025", "AUG 2025", "SEP 2025",
    "OCT 2025", "NOV 2025", "DEC 2025",
    "JAN 2026", "FEB 2026", "MAR 2026",
]

CONT_MAP = {
    10: "AIR", 11: "LCL", 12: "20RF", 13: "20GP",
    14: "40GP", 15: "40HC", 16: "40RF", 17: "45HC",
}

TODAY = datetime.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
TRACKING_IN_TRANSIT = "●●●●●●○○○○"


def clean_str(v):
    return "" if v is None else str(v).strip()


def to_num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def detect_container(row):
    for col_idx, label in CONT_MAP.items():
        val = to_num(row[col_idx - 1])
        if val and val > 0:
            return label, int(val) if val == int(val) else val
    return None, None


def format_month(sheet_name):
    parts = sheet_name.strip().split()
    return f"{parts[0].upper()}-{parts[1][-2:]}"


def format_pol_pod(pol_pod, final_dest):
    pol_pod = clean_str(pol_pod)
    final_dest = clean_str(final_dest)
    if pol_pod and final_dest:
        return f"{pol_pod}->{final_dest}"
    return pol_pod or final_dest


def parse_sheet(ws, sheet_name):
    rows = []
    month_tag = format_month(sheet_name)

    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or len(r) < 24:
            continue

        no_val = r[0]
        customer = clean_str(r[1])
        job_no = clean_str(r[8])

        if not no_val or "TOTAL" in customer.upper():
            continue
        if not job_no and not customer:
            continue

        cont_type, qty = detect_container(r)

        rows.append({
            "FAST_ID": job_no,
            "CUSTOMER": customer,
            "POL_POD": clean_str(r[2]),
            "FINAL_DEST": clean_str(r[3]),
            "POL_POD_COMBO": format_pol_pod(r[2], r[3]),
            "CARRIER": clean_str(r[6]),
            "HBL_NO": clean_str(r[7]),
            "CONT": cont_type or "",
            "QTY": qty,
            "SELL": to_num(r[18]),
            "COST": to_num(r[17]),
            "PROFIT": to_num(r[23]),
            "ETD": r[4] if isinstance(r[4], datetime.datetime) else None,
            "ETA": r[5] if isinstance(r[5], datetime.datetime) else None,
            "MONTH": month_tag,
        })
    return rows


def is_in_transit(row):
    eta = row["ETA"]
    return eta is not None and eta > TODAY


def parse_all():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    all_rows = []
    per_sheet = {}
    for sn in TARGET_SHEETS:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        rs = parse_sheet(ws, sn)
        per_sheet[sn] = len(rs)
        all_rows.extend(rs)
    wb.close()
    return all_rows, per_sheet


def read_existing_ids(erp_path):
    """Return (archive_ids, aj_ids) — existing IDs (FAST_ID/Job_ID) to avoid duplicates."""
    wb = openpyxl.load_workbook(erp_path, read_only=True, data_only=True, keep_vba=False)
    archive_ids = set()
    aj_ids = set()

    if "Archive" in wb.sheetnames:
        aws = wb["Archive"]
        # Header at row 2; Job_ID in col 1
        for r in aws.iter_rows(min_row=3, values_only=True):
            if r and r[0]:
                archive_ids.add(clean_str(r[0]))

    if "Active Jobs" in wb.sheetnames:
        ajs = wb["Active Jobs"]
        # Header at row 7; FAST_ID in col 2, Job_ID in col 3
        for r in ajs.iter_rows(min_row=8, values_only=True):
            if not r:
                continue
            if len(r) >= 2 and r[1]:
                aj_ids.add(clean_str(r[1]))
            if len(r) >= 3 and r[2]:
                aj_ids.add(clean_str(r[2]))

    wb.close()
    return archive_ids, aj_ids


def build_archive_row(r):
    """Archive has 15 cols (header at row 2). Map FAST_ID as Job_ID."""
    return [
        r["FAST_ID"],            # 1 Job_ID
        "",                      # 2 FAST_ID (empty - profit report has no FAST_ID separate)
        r["CUSTOMER"],           # 3 CUSTOMER
        r["POL_POD_COMBO"],      # 4 POL-POD
        r["CARRIER"],            # 5 CARRIER
        "",                      # 6 Bkg_No (empty, scanner fills via HBL)
        r["HBL_NO"],             # 7 HBL_NO
        r["CONT"],               # 8 Container
        r["QTY"],                # 9 Qty
        r["SELL"],               # 10 SELL
        r["COST"],               # 11 COST
        r["PROFIT"],             # 12 PROFIT
        r["ETA"],                # 13 Delivered_Date
        "Delivered",             # 14 Closed_Reason
        r["MONTH"],              # 15 MONTH
    ]


def build_aj_row(r):
    """Active Jobs has 48 cols (header at row 7). Only fill cols we have data for."""
    row = [None] * 48
    row[0]  = r["MONTH"]            # 1 MONTH
    row[1]  = r["FAST_ID"]          # 2 FAST_ID
    # row[2] Job_ID empty - no NF code yet
    row[3]  = r["CUSTOMER"]         # 4 CUSTOMER
    row[4]  = r["POL_POD_COMBO"]    # 5 POL-POD
    row[5]  = r["FINAL_DEST"]       # 6 FINAL DEST
    row[6]  = r["CARRIER"]          # 7 CARRIER
    # row[7] Bkg_No empty
    row[8]  = r["HBL_NO"]           # 9 HBL_NO
    row[9]  = r["CONT"]             # 10 CONT
    row[10] = r["QTY"]              # 11 QTY
    row[11] = "CY-CY"               # 12 SERVICE (default)
    row[12] = r["ETD"]              # 13 ETD
    row[13] = "IN TRANSIT"          # 14 STATUS
    row[14] = TRACKING_IN_TRANSIT   # 15 TRACKING (6/10 dots)
    row[15] = r["SELL"]             # 16 SELL
    row[16] = r["COST"]             # 17 COST
    row[17] = r["PROFIT"]           # 18 PROFIT
    row[20] = r["ETA"]              # 21 ETA
    return row


def run_dry():
    print(f"Source: {SRC.name}")
    print(f"Today:  {TODAY.strftime('%Y-%m-%d')}")
    print("=" * 70)

    all_rows, per_sheet = parse_all()
    for sn, n in per_sheet.items():
        print(f"  {sn:12} {n:3} jobs")
    print(f"TOTAL parsed: {len(all_rows)}")

    in_transit = [r for r in all_rows if is_in_transit(r)]
    delivered = [r for r in all_rows if not is_in_transit(r)]
    print(f"  -> In Transit (ETA > today): {len(in_transit)} -> Active Jobs")
    print(f"  -> Delivered  (ETA <= today): {len(delivered)} -> Archive")

    if not ERP.exists():
        print(f"\nERP file missing: {ERP}")
        return all_rows, [], []

    try:
        archive_ids, aj_ids = read_existing_ids(ERP)
    except PermissionError:
        print(f"\n[ERROR] ERP xlsm locked. Close Excel then retry.")
        return all_rows, [], []

    print(f"\nExisting Archive IDs: {len(archive_ids)}")
    print(f"Existing AJ IDs:      {len(aj_ids)}")

    # Split
    archive_new = [r for r in delivered
                   if r["FAST_ID"] not in archive_ids and r["FAST_ID"] not in aj_ids]
    archive_dup = [r for r in delivered
                   if r["FAST_ID"] in archive_ids or r["FAST_ID"] in aj_ids]

    aj_new = [r for r in in_transit
              if r["FAST_ID"] not in archive_ids and r["FAST_ID"] not in aj_ids]
    aj_dup = [r for r in in_transit
              if r["FAST_ID"] in archive_ids or r["FAST_ID"] in aj_ids]

    print(f"\nArchive : {len(archive_new)} new, {len(archive_dup)} skipped (dup)")
    print(f"AJ      : {len(aj_new)} new, {len(aj_dup)} skipped (dup)")

    if archive_dup:
        print(f"\n  Archive dup FAST_IDs: {[r['FAST_ID'] for r in archive_dup[:5]]}")
    if aj_dup:
        print(f"  AJ dup FAST_IDs:      {[r['FAST_ID'] for r in aj_dup[:5]]}")

    print("\n" + "=" * 70)
    print("Active Jobs preview (In Transit):")
    for r in aj_new:
        eta = r["ETA"].strftime("%Y-%m-%d") if r["ETA"] else "-"
        print(f"  {r['FAST_ID']:14} | {r['CUSTOMER'][:16]:16} | {r['POL_POD_COMBO'][:28]:28} | "
              f"{r['CONT']:5}x{r['QTY']} | ETA={eta}")

    print("\nArchive preview (first 5 + last 5):")
    sample = archive_new[:5] + [{"FAST_ID": "...", "CUSTOMER": "...", "POL_POD_COMBO": "...",
                                  "CONT": "...", "QTY": 0, "ETA": None, "MONTH": "..."}] + archive_new[-5:]
    for r in sample:
        eta = r["ETA"].strftime("%Y-%m-%d") if r.get("ETA") else "-"
        qty = r.get("QTY") or 0
        print(f"  {r['FAST_ID']:14} | {r['CUSTOMER'][:16]:16} | {r['POL_POD_COMBO'][:28]:28} | "
              f"{r['CONT']:5}x{qty} | ETA={eta} | {r['MONTH']}")

    print("\n" + "=" * 70)
    print(f"READY: archive +{len(archive_new)}, AJ +{len(aj_new)}")
    return all_rows, archive_new, aj_new


def run_live(archive_new, aj_new):
    if not archive_new and not aj_new:
        print("Nothing to insert.")
        return

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = ERP.with_name(f"{ERP.stem}_backup_{ts}.xlsm")
    shutil.copy2(ERP, backup)
    print(f"[backup] {backup.name}")

    wb = openpyxl.load_workbook(ERP, keep_vba=True)

    if archive_new:
        aws = wb["Archive"]
        start = aws.max_row + 1
        for r in archive_new:
            aws.append(build_archive_row(r))
        print(f"[Archive] appended rows {start} -> {aws.max_row}  (+{len(archive_new)})")

    if aj_new:
        ajs = wb["Active Jobs"]
        start = ajs.max_row + 1
        for r in aj_new:
            ajs.append(build_aj_row(r))
        print(f"[AJ]      appended rows {start} -> {ajs.max_row}  (+{len(aj_new)})")

    wb.save(ERP)
    wb.close()
    print(f"[save] {ERP.name}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--live", action="store_true")
    args = p.parse_args()

    _, archive_new, aj_new = run_dry()

    if args.live:
        print("\n[LIVE] Committing...")
        run_live(archive_new, aj_new)
        print("DONE.")
    else:
        print("\nDRY-RUN only. Add --live to commit.")


if __name__ == "__main__":
    main()
