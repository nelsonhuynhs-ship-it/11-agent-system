"""
audit-xlsm-vs-parquet.py — verify final Pricing Dry/Reefer matches parquet expectations.
"""
import openpyxl
from pathlib import Path
import shutil
import tempfile

ERP = Path("D:/OneDrive/NelsonData/erp/ERP_Master_v14.xlsm")

fd, tmp_path = tempfile.mkstemp(suffix=".xlsm")
import os; os.close(fd)
shutil.copy2(ERP, tmp_path)

wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)

def audit_sheet(sheet_name: str, carrier_filter: str):
    ws = wb[sheet_name]
    places = {}
    for r in range(2, min(ws.max_row + 1, 5000)):
        carrier = str(ws.cell(row=r, column=4).value or '').strip().upper()
        if carrier != carrier_filter: continue
        pol = str(ws.cell(row=r, column=1).value or '').strip()
        place = str(ws.cell(row=r, column=3).value or '').strip()
        note = str(ws.cell(row=r, column=8).value or '').strip()
        source = str(ws.cell(row=r, column=9).value or '').strip()
        key = (pol, place, source, note)
        places.setdefault(key, []).append(r)
    return places

for carrier in ["ONE", "COSCO"]:
    print(f"\n{'='*70}\n{carrier}\n{'='*70}")
    dry = audit_sheet("Pricing Dry", carrier)
    reefer = audit_sheet("Pricing Reefer", carrier)

    dry_places = set(k[1] for k in dry.keys() if k[0] == "HCM")
    reefer_places = set(k[1] for k in reefer.keys() if k[0] == "HCM")

    print(f"Pricing Dry HCM places:    {len(dry_places)}")
    print(f"Pricing Reefer HCM places: {len(reefer_places)}")
    print(f"\nDRY places:    {sorted(dry_places)}")
    print(f"REEFER places: {sorted(reefer_places)}")

# Specific check: ONE HCM-TACOMA all rows
print(f"\n{'='*70}\nONE HCM-TACOMA all rows (Dry + Reefer)\n{'='*70}")
for sheet in ["Pricing Dry", "Pricing Reefer"]:
    ws = wb[sheet]
    print(f"\n[{sheet}]")
    for r in range(2, min(ws.max_row + 1, 5000)):
        carrier = str(ws.cell(row=r, column=4).value or '').strip().upper()
        pol = str(ws.cell(row=r, column=1).value or '').strip().upper()
        pod = str(ws.cell(row=r, column=2).value or '').strip().upper()
        if carrier == "ONE" and pol == "HCM" and "TACOMA" in pod:
            place = ws.cell(row=r, column=3).value
            source = ws.cell(row=r, column=9).value
            note = ws.cell(row=r, column=8).value
            cmd = str(ws.cell(row=r, column=5).value or '')[:50]
            print(f"  row {r}: Place={place} Source={source} Note={note!r} Cmd={cmd!r}")

wb.close()
os.remove(tmp_path)
