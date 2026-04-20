"""
erp-import-shipments.py — Import Shipments.xlsx → Active Jobs + Archive (ERP v14)
==================================================================================
Reads 12 monthly sheets from Shipments.xlsx and writes rows to ERP_Master_v14.xlsm:

  Active Jobs  — rows where ETD.month == current month (Apr 2026)
  Archive      — rows where ETD.month < current month

Filter rule (Nelson confirmed):
  ETD col is the departure proxy. ETD month == current → Active; < current → Archive.
  Rows with both Bkg_No AND Hbl_No blank are skipped (insufficient ID).

Idempotency:
  Match by Bkg_No in target sheet.  If found → UPDATE.  If not found → INSERT.

Usage:
    python scripts/erp-import-shipments.py [--dry-run] [--source PATH] [--target PATH]

Constraints:
  - openpyxl.load_workbook(keep_vba=True)
  - save via ERP.core.ribbon_guard.save_preserving_ribbon (NEVER wb.save)
  - Backup before write: ERP_Master_v14.backup_YYYYMMDD_HHMMSS.xlsm
  - Excel must be closed before run (file-lock check)
  - Exit 0 success / non-zero failure

Gotcha #6: NEVER call wb.save() — always save_preserving_ribbon().
"""
from __future__ import annotations

import argparse
import os
import re
import shutil
import sys
from datetime import datetime
from typing import Any, Optional

import openpyxl

# ---------------------------------------------------------------------------
# Repo path so we can import ERP.core modules
# ---------------------------------------------------------------------------
_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from ERP.core.active_jobs_cols import (  # noqa: E402
    COL,
    DATA_START,
    HDR_ROW,
    TOTAL_COLS,
    derive_month,
    derive_pol_pod,
)
from ERP.core.ribbon_guard import save_preserving_ribbon  # noqa: E402

# ---------------------------------------------------------------------------
# Defaults
# ---------------------------------------------------------------------------
DEFAULT_SOURCE = r"C:\Users\Nelson\OneDrive\Desktop\Shipments.xlsx"
DEFAULT_TARGET = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"

# Current month for filter rule (Apr 2026)
_NOW = datetime.now()
CURRENT_YEAR_MONTH = (_NOW.year, _NOW.month)

# Monthly sheets to iterate (skip Sheet1/2/3)
MONTHLY_SHEETS = [
    "May 2025", "Jun 2025", "Jul 2025", "Aug 2025", "Sep 2025", "Oct 2025",
    "Nov 2025", "Dec 2025", "Jan 2026", "Feb 2026", "Mar 2026", "Apr 2026",
]

# Archive sheet constants
ARCHIVE_SHEET = "Archive"
# Row 1 = title, Row 2 = actual header, Data from row 3
ARCHIVE_TITLE_ROW = 1
ARCHIVE_HDR_ROW = 2
ARCHIVE_DATA_START = 3

# Archive columns (14 cols, matched to existing sheet)
ARCH_COL = {
    "Job_ID":           1,
    "FAST_ID":          2,
    "CUSTOMER":         3,
    "POL_POD":          4,
    "CARRIER":          5,
    "Bkg_No":           6,
    "HBL_NO":           7,
    "Container":        8,
    "Qty":              9,
    "SELL":             10,
    "COST":             11,
    "PROFIT":           12,
    "Delivered_Date":   13,
    "Closed_Reason":    14,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _log(msg: str) -> None:
    # encode-safe output: replace chars not supported by console codepage
    safe = msg.encode(sys.stdout.encoding or "utf-8", errors="replace").decode(
        sys.stdout.encoding or "utf-8", errors="replace"
    )
    print(safe, flush=True)


def _check_excel_closed(path: str) -> None:
    """Raise SystemExit if file is locked by Excel."""
    try:
        with open(path, "a"):
            pass
    except PermissionError:
        _log(f"[ERROR] File is locked — close Excel first: {path}")
        sys.exit(1)


def _backup(path: str) -> str:
    """Create timestamped backup. Returns backup path."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(path)
    backup_path = f"{base}.backup_{ts}{ext}"
    shutil.copy2(path, backup_path)
    _log(f"[backup] {backup_path}")
    return backup_path


def _parse_float(value: Any) -> Optional[float]:
    """Parse numeric or string like '3571 + 75' into float."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    # Handle expressions like "3571 + 75"
    if re.search(r"[+\-*/]", s):
        try:
            return float(eval(s, {"__builtins__": {}}))  # safe: no builtins
        except Exception:
            pass
    # Try direct parse
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def _parse_date(value: Any) -> Optional[datetime]:
    """Normalize date value from openpyxl (already datetime or string)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def _parse_bkg_no(value: Any) -> Optional[str]:
    """Normalize booking number to string."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _derive_service(routing: str) -> str:
    """Derive SERVICE: CY-DOOR if VIA in routing, else CY-CY."""
    if not routing:
        return "CY-CY"
    if "VIA" in str(routing).upper():
        return "CY-DOOR"
    return "CY-CY"


def _status_map(status: Any) -> str:
    """Map Shipments Status to Active Jobs Status vocabulary."""
    if status is None:
        return ""
    s = str(status).strip().upper()
    mapping = {
        "CONFIRMED": "BOOKED",
        "BOOKED": "BOOKED",
        "IN TRANSIT": "IN TRANSIT",
        "INTRANSIT": "IN TRANSIT",
        "ARRIVED": "ARRIVED",
        "DELIVERED": "DELIVERED",
        "CANCELLED": "CANCELLED",
        "CANCELED": "CANCELLED",
    }
    return mapping.get(s, str(status).strip())


def _build_notes(volume: Any, hdl_fee: Any, status_calc: Any, progress_eta: Any) -> str:
    """Assemble Notes string from auxiliary Shipments cols."""
    parts = []
    if volume is not None:
        parts.append(f"Vol={volume}")
    if hdl_fee is not None and str(hdl_fee).strip():
        parts.append(f"HdlFee={hdl_fee}")
    if status_calc is not None and str(status_calc).strip():
        parts.append(f"Calc={status_calc}")
    if progress_eta is not None and str(progress_eta).strip():
        parts.append(f"Progress={progress_eta}")
    return " | ".join(parts)


def _build_delay_log(etd_original: Any, etd: Any, existing_delay: Any) -> str:
    """Build delay log from ETD_Original vs ETD, merged with existing Delay_Log."""
    parts = []
    etd_orig_d = _parse_date(etd_original)
    etd_d = _parse_date(etd)
    if etd_orig_d and etd_d and etd_orig_d.date() != etd_d.date():
        parts.append(f"Re-sched from ETD_Orig {etd_orig_d.strftime('%Y-%m-%d')}")
    if existing_delay is not None and str(existing_delay).strip():
        existing = str(existing_delay).strip()
        if existing not in " | ".join(parts):
            parts.append(existing)
    return " | ".join(parts)


# ---------------------------------------------------------------------------
# Source schema detection
# ---------------------------------------------------------------------------

def _detect_schema(header_row: tuple) -> dict[str, int]:
    """
    Map header names → 0-based column indices for a given sheet.
    Handles both old schema (col 1 = Customer) and new schema (col 1 = Stt, col 2 = Customer).
    """
    col_map: dict[str, int] = {}
    for i, h in enumerate(header_row):
        if h is None:
            continue
        normalized = str(h).strip()
        col_map[normalized] = i
    return col_map


def _get_val(row: tuple, col_map: dict[str, int], key: str) -> Any:
    """Safe lookup by header name."""
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


# ---------------------------------------------------------------------------
# Job ID counter (sequential per run, 3 digits)
# ---------------------------------------------------------------------------

class _JobIDCounter:
    def __init__(self, existing_ids: set[str], etd: Optional[datetime]):
        self._prefix = f"NF-{etd.strftime('%m%d')}-" if etd else "NF-0000-"
        # Find max seq already used for this prefix
        max_seq = 0
        pattern = re.compile(r"NF-\d{4}-(\d{3})")
        for jid in existing_ids:
            m = pattern.match(str(jid))
            if m:
                max_seq = max(max_seq, int(m.group(1)))
        self._seq = max_seq

    def next(self) -> str:
        self._seq += 1
        return f"{self._prefix}{self._seq:03d}"


# ---------------------------------------------------------------------------
# Read Shipments source
# ---------------------------------------------------------------------------

def read_shipments(source_path: str) -> tuple[list[dict], list[dict]]:
    """
    Read all 12 monthly sheets from Shipments.xlsx.

    Returns (active_rows, archive_rows) as list of dicts keyed by Active Jobs
    or Archive field names.
    """
    wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    active_raw: list[dict] = []
    archive_raw: list[dict] = []
    skipped = 0

    for sheet_name in MONTHLY_SHEETS:
        if sheet_name not in wb.sheetnames:
            _log(f"[WARN] Sheet not found: {sheet_name} — skipping")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Detect schema from header row (row 0)
        header = rows[0]
        col_map = _detect_schema(header)

        for row_data in rows[1:]:
            # Skip fully empty rows
            if all(v is None for v in row_data):
                continue

            # Extract key fields
            bkg_no = _parse_bkg_no(_get_val(row_data, col_map, "Bkg No"))
            hbl_no = _parse_bkg_no(_get_val(row_data, col_map, "Hbl No"))

            # Skip rows with both IDs blank
            if not bkg_no and not hbl_no:
                skipped += 1
                continue

            etd = _parse_date(_get_val(row_data, col_map, "ETD"))
            eta = _parse_date(_get_val(row_data, col_map, "ETA"))
            ata = _parse_date(_get_val(row_data, col_map, "ATA"))
            etd_original = _get_val(row_data, col_map, "ETD_Original") or \
                           _get_val(row_data, col_map, "Etd_Original")

            customer = _get_val(row_data, col_map, "Customer")
            customer_type = _get_val(row_data, col_map, "Customer Type")
            routing = _get_val(row_data, col_map, "Routing")
            carrier = _get_val(row_data, col_map, "Carrier")
            container_type = _get_val(row_data, col_map, "Container Type")
            quantity = _get_val(row_data, col_map, "Quantity")
            volume = _get_val(row_data, col_map, "Volume")
            status = _get_val(row_data, col_map, "Status")
            selling_rate = _parse_float(_get_val(row_data, col_map, "Selling Rate"))
            buying_rate_raw = _get_val(row_data, col_map, "Buying Rate")
            buying_rate = _parse_float(buying_rate_raw)
            profit = _parse_float(_get_val(row_data, col_map, "Profit"))
            si = _get_val(row_data, col_map, "Si")
            cy = _get_val(row_data, col_map, "Cy")
            hdl_fee = _get_val(row_data, col_map, "Hdl Fee Carrier")
            status_calc = _get_val(row_data, col_map, "Status_Calc")
            delay_log_src = _get_val(row_data, col_map, "Delay_Log")
            progress_eta = _get_val(row_data, col_map, "Progress ETA")

            # Derived fields
            pol_pod = derive_pol_pod(str(routing) if routing else "")
            month_label = derive_month(etd) if etd else ""
            service = _derive_service(str(routing) if routing else "")
            notes = _build_notes(volume, hdl_fee, status_calc, progress_eta)
            delay_log = _build_delay_log(etd_original, etd, delay_log_src)
            profit_margin: Optional[float] = None
            if selling_rate and profit is not None and selling_rate != 0:
                profit_margin = round(profit / selling_rate * 100, 2)
            mapped_status = _status_map(status)

            record = {
                # Source identifiers
                "_sheet": sheet_name,
                "_bkg_no": bkg_no,
                "_hbl_no": hbl_no,
                "_etd": etd,
                # Active Jobs mapping
                "MONTH": month_label,
                "FAST_ID": None,
                "Job_ID": None,  # assigned later per sheet
                "CRM_ID": customer,
                "POL_POD": pol_pod,
                "Door_Address": None,
                "Carrier": carrier,
                "Bkg_No": bkg_no,
                "HBL_NO": hbl_no,
                "Container_Type": container_type,
                "Quantity": quantity,
                "SERVICE": service,
                "ETD": etd,
                "Status": mapped_status,
                "TRACKING": None,
                "Selling_Rate": selling_rate,
                "Buying_Rate": buying_rate,
                "Profit": profit,
                "Request_BKG": None,
                "Routing": routing,
                "ETA": eta,
                "ATA": ata,
                "Contract_Type": None,
                "Profit_Margin": profit_margin,
                "Customer_Type": customer_type,
                "SI_Received": si,
                "CY_Cutoff": cy,
                "Door_Delivery": None,
                "Door_Status": None,
                "Delay_Count": None,
                "Delay_Log": delay_log,
                "Notes": notes,
                "Created_Date": datetime.now(),
                "Last_Updated": datetime.now(),
                "Cost_Breakdown": None,
                "TRACKING_STAGE": None,
                "RELEASE_EMAIL_SENT": None,
                "RELEASE_CONFIRMED": None,
                "PRICE_WATCH_STATUS": None,
                "PRICE_WATCH_DELTA": None,
            }

            # Filter rule: ETD month == current → Active; else → Archive
            if etd and (etd.year, etd.month) == CURRENT_YEAR_MONTH:
                active_raw.append(record)
            elif etd:
                archive_raw.append(record)
            else:
                # No ETD → cannot classify reliably, skip with warning
                _log(f"[SKIP] No ETD — Bkg={bkg_no} HBL={hbl_no} in {sheet_name}")
                skipped += 1

    wb.close()
    _log(f"[source] Active={len(active_raw)} Archive={len(archive_raw)} Skipped={skipped}")
    return active_raw, archive_raw


# ---------------------------------------------------------------------------
# Read existing Bkg_No index from a sheet
# ---------------------------------------------------------------------------

def _build_bkg_index(ws, bkg_col_1based: int, data_start: int) -> dict[str, int]:
    """
    Returns {bkg_no_str: row_number} for all existing rows.
    bkg_col_1based: 1-based column number for Bkg_No field.
    """
    index: dict[str, int] = {}
    for row_num in range(data_start, (ws.max_row or data_start) + 1):
        cell = ws.cell(row=row_num, column=bkg_col_1based)
        val = cell.value
        if val is not None:
            key = str(val).strip()
            if key:
                index[key] = row_num
    return index


def _next_empty_row(ws, data_start: int) -> int:
    """Find first truly empty row starting from data_start."""
    row = data_start
    while True:
        # Check if any cell in this row has a value
        has_value = any(
            ws.cell(row=row, column=c).value is not None
            for c in range(1, min(10, TOTAL_COLS + 1))
        )
        if not has_value:
            return row
        row += 1


# ---------------------------------------------------------------------------
# Write to Active Jobs
# ---------------------------------------------------------------------------

def write_active_jobs(
    ws,
    records: list[dict],
    dry_run: bool,
) -> tuple[int, int, int]:
    """
    Upsert records into Active Jobs sheet.
    Returns (inserted, updated, skipped).
    """
    inserted = updated = job_seq = 0

    # Build index of existing Bkg_No → row
    bkg_index = _build_bkg_index(ws, COL["Bkg_No"], DATA_START)
    existing_job_ids: set[str] = set()
    for r in range(DATA_START, (ws.max_row or DATA_START) + 1):
        v = ws.cell(row=r, column=COL["Job_ID"]).value
        if v:
            existing_job_ids.add(str(v).strip())

    for rec in records:
        bkg_no = rec.get("_bkg_no") or ""
        etd = rec.get("_etd")

        # Determine target row
        if bkg_no and bkg_no in bkg_index:
            target_row = bkg_index[bkg_no]
            action = "UPDATE"
        else:
            target_row = _next_empty_row(ws, DATA_START)
            action = "INSERT"
            # Assign Job_ID for new row
            counter = _JobIDCounter(existing_job_ids, etd)
            job_id = counter.next()
            rec["Job_ID"] = job_id
            existing_job_ids.add(job_id)
            if bkg_no:
                bkg_index[bkg_no] = target_row

        label = f"NF {rec.get('CRM_ID','')} {rec.get('POL_POD','')} ETD={etd.strftime('%Y-%m-%d') if etd else 'N/A'}"
        _log(f"  [{action}] row={target_row} Bkg={bkg_no} {label}")

        if not dry_run:
            _write_active_row(ws, target_row, rec)

        if action == "INSERT":
            inserted += 1
        else:
            updated += 1

    return inserted, updated, 0


def _write_active_row(ws, row: int, rec: dict) -> None:
    """Write a single record dict into the Active Jobs worksheet at given row."""
    field_to_col = {
        "MONTH":            COL["MONTH"],
        "FAST_ID":          COL["FAST_ID"],
        "Job_ID":           COL["Job_ID"],
        "CRM_ID":           COL["CRM_ID"],
        "POL_POD":          COL["POL_POD"],
        "Door_Address":     COL["Door_Address"],
        "Carrier":          COL["Carrier"],
        "Bkg_No":           COL["Bkg_No"],
        "HBL_NO":           COL["HBL_NO"],
        "Container_Type":   COL["Container_Type"],
        "Quantity":         COL["Quantity"],
        "SERVICE":          COL["SERVICE"],
        "ETD":              COL["ETD"],
        "Status":           COL["Status"],
        "TRACKING":         COL["TRACKING"],
        "Selling_Rate":     COL["Selling_Rate"],
        "Buying_Rate":      COL["Buying_Rate"],
        "Profit":           COL["Profit"],
        "Request_BKG":      COL["Request_BKG"],
        "Routing":          COL["Routing"],
        "ETA":              COL["ETA"],
        "ATA":              COL["ATA"],
        "Contract_Type":    COL["Contract_Type"],
        "Profit_Margin":    COL["Profit_Margin"],
        "Customer_Type":    COL["Customer_Type"],
        "SI_Received":      COL["SI_Received"],
        "CY_Cutoff":        COL["CY_Cutoff"],
        "Door_Delivery":    COL["Door_Delivery"],
        "Door_Status":      COL["Door_Status"],
        "Delay_Count":      COL["Delay_Count"],
        "Delay_Log":        COL["Delay_Log"],
        "Notes":            COL["Notes"],
        "Created_Date":     COL["Created_Date"],
        "Last_Updated":     COL["Last_Updated"],
        "Cost_Breakdown":   COL["Cost_Breakdown"],
        "TRACKING_STAGE":   COL["TRACKING_STAGE"],
        "RELEASE_EMAIL_SENT":  COL["RELEASE_EMAIL_SENT"],
        "RELEASE_CONFIRMED":   COL["RELEASE_CONFIRMED"],
        "PRICE_WATCH_STATUS":  COL["PRICE_WATCH_STATUS"],
        "PRICE_WATCH_DELTA":   COL["PRICE_WATCH_DELTA"],
    }
    for field, col_num in field_to_col.items():
        value = rec.get(field)
        cell = ws.cell(row=row, column=col_num)
        # Always update Last_Updated on writes
        if field == "Last_Updated":
            cell.value = datetime.now()
        elif field == "Created_Date":
            # Only set Created_Date on new rows (don't overwrite)
            if cell.value is None:
                cell.value = value
        else:
            if value is not None:
                cell.value = value


# ---------------------------------------------------------------------------
# Write to Archive
# ---------------------------------------------------------------------------

def _ensure_archive_header(ws) -> None:
    """
    Ensure Archive sheet has correct structure:
      Row 1: title (keep existing)
      Row 2: 14-col header
      Data from row 3
    """
    # Check if row 2 already has header
    existing_hdr = ws.cell(row=ARCHIVE_HDR_ROW, column=1).value
    if existing_hdr == "Job_ID":
        return  # Already set up correctly

    # Write header row
    arch_headers = [
        "Job_ID", "FAST_ID", "CUSTOMER", "POL-POD", "CARRIER",
        "Bkg_No", "HBL_NO", "Container", "Qty",
        "SELL", "COST", "PROFIT", "Delivered_Date", "Closed_Reason",
    ]
    for i, h in enumerate(arch_headers, 1):
        ws.cell(row=ARCHIVE_HDR_ROW, column=i).value = h


def write_archive(
    ws,
    records: list[dict],
    dry_run: bool,
) -> tuple[int, int]:
    """
    Upsert records into Archive sheet.
    Returns (inserted, updated).
    """
    if not dry_run:
        _ensure_archive_header(ws)

    inserted = updated = 0

    # Build index: Bkg_No → row
    bkg_index = _build_bkg_index(ws, ARCH_COL["Bkg_No"], ARCHIVE_DATA_START)
    existing_job_ids: set[str] = set()
    for r in range(ARCHIVE_DATA_START, (ws.max_row or ARCHIVE_DATA_START) + 1):
        v = ws.cell(row=r, column=ARCH_COL["Job_ID"]).value
        if v:
            existing_job_ids.add(str(v).strip())

    for rec in records:
        bkg_no = rec.get("_bkg_no") or ""
        etd = rec.get("_etd")

        if bkg_no and bkg_no in bkg_index:
            target_row = bkg_index[bkg_no]
            action = "UPDATE"
            job_id = ws.cell(row=target_row, column=ARCH_COL["Job_ID"]).value or ""
        else:
            target_row = _next_empty_arch_row(ws)
            action = "INSERT"
            counter = _JobIDCounter(existing_job_ids, etd)
            job_id = counter.next()
            existing_job_ids.add(job_id)
            if bkg_no:
                bkg_index[bkg_no] = target_row

        label = f"{rec.get('CRM_ID','')} {rec.get('POL_POD','')} ETD={etd.strftime('%Y-%m-%d') if etd else 'N/A'}"
        _log(f"  [Archive {action}] row={target_row} Bkg={bkg_no} {label}")

        if not dry_run:
            _write_archive_row(ws, target_row, rec, job_id)

        if action == "INSERT":
            inserted += 1
        else:
            updated += 1

    return inserted, updated


def _next_empty_arch_row(ws) -> int:
    """Find first empty row in Archive starting from data start."""
    row = ARCHIVE_DATA_START
    while True:
        if ws.cell(row=row, column=ARCH_COL["Bkg_No"]).value is None:
            return row
        row += 1


def _write_archive_row(ws, row: int, rec: dict, job_id: str) -> None:
    """Write a single record into Archive sheet."""
    ws.cell(row=row, column=ARCH_COL["Job_ID"]).value = job_id
    ws.cell(row=row, column=ARCH_COL["FAST_ID"]).value = rec.get("FAST_ID")
    ws.cell(row=row, column=ARCH_COL["CUSTOMER"]).value = rec.get("CRM_ID")
    ws.cell(row=row, column=ARCH_COL["POL_POD"]).value = rec.get("POL_POD")
    ws.cell(row=row, column=ARCH_COL["CARRIER"]).value = rec.get("Carrier")
    ws.cell(row=row, column=ARCH_COL["Bkg_No"]).value = rec.get("Bkg_No")
    ws.cell(row=row, column=ARCH_COL["HBL_NO"]).value = rec.get("HBL_NO")
    ws.cell(row=row, column=ARCH_COL["Container"]).value = rec.get("Container_Type")
    ws.cell(row=row, column=ARCH_COL["Qty"]).value = rec.get("Quantity")
    ws.cell(row=row, column=ARCH_COL["SELL"]).value = rec.get("Selling_Rate")
    ws.cell(row=row, column=ARCH_COL["COST"]).value = rec.get("Buying_Rate")
    ws.cell(row=row, column=ARCH_COL["PROFIT"]).value = rec.get("Profit")
    # Delivered_Date: use ATA if available, else ETA
    delivered = rec.get("ATA") or rec.get("ETA")
    ws.cell(row=row, column=ARCH_COL["Delivered_Date"]).value = delivered
    ws.cell(row=row, column=ARCH_COL["Closed_Reason"]).value = "Delivered"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import Shipments.xlsx → Active Jobs + Archive (ERP v14)"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Preview changes without saving to ERP xlsm",
    )
    parser.add_argument(
        "--source", default=DEFAULT_SOURCE,
        help=f"Path to Shipments.xlsx (default: {DEFAULT_SOURCE})",
    )
    parser.add_argument(
        "--target", default=DEFAULT_TARGET,
        help=f"Path to ERP_Master_v14.xlsm (default: {DEFAULT_TARGET})",
    )
    args = parser.parse_args()

    source_path: str = args.source
    target_path: str = args.target
    dry_run: bool = args.dry_run

    _log("=" * 60)
    _log(f"[erp-import-shipments] {'DRY RUN — ' if dry_run else ''}Starting")
    _log(f"  Source : {source_path}")
    _log(f"  Target : {target_path}")
    _log(f"  Current month filter: {CURRENT_YEAR_MONTH[1]:02d}/{CURRENT_YEAR_MONTH[0]}")
    _log("=" * 60)

    # Validate source exists
    if not os.path.exists(source_path):
        _log(f"[ERROR] Source not found: {source_path}")
        sys.exit(1)

    if not dry_run:
        # Validate target exists
        if not os.path.exists(target_path):
            _log(f"[ERROR] Target not found: {target_path}")
            sys.exit(1)
        # Check file lock
        _check_excel_closed(target_path)
        # Backup
        _backup(target_path)

    # ── Step 1: Read source ──
    _log("\n[Step 1] Reading Shipments.xlsx ...")
    active_rows, archive_rows = read_shipments(source_path)

    if dry_run:
        _log("\n[DRY RUN] Would write:")
        _log(f"  Active Jobs : {len(active_rows)} rows (ETD month={CURRENT_YEAR_MONTH[1]:02d}/{CURRENT_YEAR_MONTH[0]})")
        _log(f"  Archive     : {len(archive_rows)} rows (ETD < current month)")
        _log("\n  Active rows preview:")
        for r in active_rows:
            etd = r.get("_etd")
            _log(f"    [Active] Bkg={r.get('_bkg_no','?')} | {r.get('CRM_ID','?')} | {r.get('POL_POD','?')} | ETD={etd.strftime('%Y-%m-%d') if etd else 'N/A'}")
        _log("\n  Archive rows preview (first 10):")
        for r in archive_rows[:10]:
            etd = r.get("_etd")
            _log(f"    [Archive] Bkg={r.get('_bkg_no','?')} | {r.get('CRM_ID','?')} | {r.get('POL_POD','?')} | ETD={etd.strftime('%Y-%m-%d') if etd else 'N/A'}")
        if len(archive_rows) > 10:
            _log(f"    ... and {len(archive_rows) - 10} more")
        _log("\n[DRY RUN] No changes written. Exiting.")
        sys.exit(0)

    # ── Step 2: Open ERP workbook ──
    _log("\n[Step 2] Opening ERP workbook ...")
    wb = openpyxl.load_workbook(target_path, keep_vba=True)

    ws_aj = wb["Active Jobs"]
    ws_arch = wb[ARCHIVE_SHEET]

    # ── Step 3: Write Active Jobs ──
    _log("\n[Step 3] Writing Active Jobs ...")
    aj_ins, aj_upd, aj_skip = write_active_jobs(ws_aj, active_rows, dry_run=False)

    # ── Step 4: Write Archive ──
    _log("\n[Step 4] Writing Archive ...")
    arch_ins, arch_upd = write_archive(ws_arch, archive_rows, dry_run=False)

    # ── Step 5: Save (MUST use save_preserving_ribbon — gotcha #6) ──
    _log("\n[Step 5] Saving (preserving ribbon) ...")
    result = save_preserving_ribbon(wb, target_path)
    _log(f"  Ribbon guard result: {result}")

    # ── Summary ──
    _log("\n" + "=" * 60)
    _log("[DONE] Import complete:")
    _log(f"  Active Jobs : +{aj_ins} inserted / {aj_upd} updated / {aj_skip} skipped")
    _log(f"  Archive     : +{arch_ins} inserted / {arch_upd} updated")
    _log("=" * 60)


if __name__ == "__main__":
    main()
