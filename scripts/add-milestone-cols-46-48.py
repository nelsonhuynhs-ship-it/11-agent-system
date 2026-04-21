"""
One-shot: add cols 46-48 (ATD_Date, Notified_ATD, Notified_ETA7) to Active Jobs.

Moves CNEE Milestone SyncMilestones cols from 41/43/44 (collision with Phase 3
Booking Pool cols) to new cols 46/47/48.

Re-injects CustomUI after save (openpyxl strips it).
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path("D:/OneDrive/NelsonData/erp")))

import openpyxl
from customui_utils import ensure_customui

ERP = Path("D:/OneDrive/NelsonData/erp/ERP_Master_v14.xlsm")
CUSTOMUI_XML = Path("D:/OneDrive/NelsonData/erp/CustomUI_v14.xml")

HEADERS = {
    46: "ATD_Date",
    47: "Notified_ATD",
    48: "Notified_ETA7",
}

def main() -> int:
    if not ERP.exists():
        print(f"ERP not found: {ERP}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(ERP, keep_vba=True)
    if "Active Jobs" not in wb.sheetnames:
        print("Active Jobs sheet missing", file=sys.stderr)
        return 1

    ws = wb["Active Jobs"]

    # Header row 7 (confirmed earlier)
    HEADER_ROW = 7
    added = []
    for col, name in HEADERS.items():
        existing = ws.cell(row=HEADER_ROW, column=col).value
        if existing in (None, "", name):
            ws.cell(row=HEADER_ROW, column=col).value = name
            # Hide the column
            letter = openpyxl.utils.get_column_letter(col)
            if letter not in ws.column_dimensions:
                ws.column_dimensions[letter].hidden = True
            else:
                ws.column_dimensions[letter].hidden = True
            added.append(f"{letter}={name}")
        else:
            print(f"  SKIP col {col}: already has {existing!r}")

    wb.save(ERP)
    print(f"Saved. Added: {added}")

    # Re-inject customUI
    res = ensure_customui(str(ERP), str(CUSTOMUI_XML))
    print(f"CustomUI: {res}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
