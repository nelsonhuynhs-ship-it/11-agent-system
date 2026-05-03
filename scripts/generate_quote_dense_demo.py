"""
Generate QUOTE_DENSE_DEMO.xlsx — Dense quote dashboard demo using Excel-native features.
Pure Python + openpyxl (no Excel COM required).
Nelson Freight context: carriers ONE/HPL/CMA, routes HCM/HPH → US PODs.
"""
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Constants ────────────────────────────────────────────────────────────────
OUTPUT = "docs/visual-tour/quote-mockups/QUOTE_DENSE_DEMO.xlsx"
SHEET  = "Quotes_DENSE"
N_ROWS  = 20   # sample data rows

# Colors
C_HEADER_BG   = "1E3A5F"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_BANNER_BG   = "F0F4FF"
C_ALT_ROW     = "F8FAFC"
C_WIN         = "4ADE80"    # green fill
C_LOST        = "F87171"    # red fill
C_PENDING     = "FCD34D"    # amber fill
C_EXPIRED     = "9CA3AF"   # gray fill
C_WIN_TXT     = "065F46"
C_LOST_TXT    = "7F1D1D"
C_PENDING_TXT = "713F12"
C_EXPIRED_TXT = "6B7280"
C_MARGIN_HOT  = "FCA5A5"   # red   (margin < 50)
C_MARGIN_WARM = "FDE68A"   # yellow (50 ≤ margin < 150)
C_MARGIN_COOL = "6EE7B7"   # green (margin ≥ 150)
C_MARGIN_HOT_TXT  = "7F1D1D"
C_MARGIN_WARM_TXT = "713F12"
C_MARGIN_COOL_TXT = "064E3B"
C_VIP_TXT     = "8B5CF6"
C_NEW_TXT     = "06B6D4"
C_REG_TXT     = "1E293B"

# ── Sample data ─────────────────────────────────────────────────────────────
# Columns: #, Customer, Carrier, POL, POD, 20GP, 40GP, 40HC, 45HC, 40NOR, Margin, Status, Date, Trend(ref)
CARRIERS = ["ONE", "HPL", "CMA", "YML", "OOL", "EMC", "CSL"]
ROUTES   = [
    ("HCM", "LAX/LGB"), ("HPH", "LAX/LGB"), ("HCM", "USNYC"),
    ("HPH", "USNYC"),   ("HCM", "USEC"),    ("HPH", "SAV"),
    ("SGN", "LAX/LGB"), ("HCM", "ATL"),     ("HPH", "CHI"),
]
CUSTOMERS = [
    ("🌟 GOWIN INTERNATIONAL",         "vip"),
    ("🌟 SUNRISE FURNITURE CORP",       "vip"),
    ("🆕 GREENLINE TRADING",            "new"),
    ("👤 ATLAS WOOD PRODUCTS",           "reg"),
    ("👤 PREMIUM FLOORING INC",         "reg"),
    ("🌟 NOVA COMMODITIES",             "vip"),
    ("👤 ASIA PACIFIC CARGO",           "reg"),
    ("🆕 SWIFT LOGISTICS VN",            "new"),
    ("👤 EVEREST TEXTILE",              "reg"),
    ("👤 DELTA SUPPLY CHAIN",           "reg"),
    ("🌟 PRIME SHIPPING CO",            "vip"),
    ("👤 GOLDEN GATE FREIGHT",          "reg"),
    ("🆕 APEX CARGO SERVICES",          "new"),
    ("👤 PACIFIC TRADE VN",             "reg"),
    ("🌟 CLEARWATER TRADING",            "vip"),
    ("👤 HORIZON LOGISTICS",            "reg"),
    ("👤 TOPAZ SHIPPING",               "reg"),
    ("🆕 ZENITH GLOBAL",                "new"),
    ("👤 AURORA FREIGHT",               "reg"),
    ("👤 MERIDIAN CARGO",               "reg"),
]
STATUSES  = ["WIN", "WIN", "WIN", "PENDING", "PENDING", "LOST", "EXPIRED", "WIN", "PENDING", "WIN"]

def make_row(i):
    c_name, c_tier = CUSTOMERS[i % len(CUSTOMERS)]
    carrier  = CARRIERS[i % len(CARRIERS)]
    pol, pod = ROUTES[i % len(ROUTES)]
    base     = random.randint(1800, 3200)
    margin   = random.randint(30, 320)
    status   = STATUSES[i % len(STATUSES)]
    day_off  = random.randint(0, 14)
    return [str(i+1), c_name, carrier, pol, pod,
            random.randint(1700, 3100) if random.random() > 0.05 else None,
            random.randint(2700, 4700) if random.random() > 0.05 else None,
            random.randint(2800, 4900) if random.random() > 0.05 else None,
            random.randint(3000, 5300) if random.random() > 0.40 else None,
            random.randint(2500, 4300) if random.random() > 0.50 else None,
            margin, status,
            f"2026-04-{25-day_off:02d}"]

# ── Workbook setup ──────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = SHEET

# Helper styles
def hdr_font(bold=True, color=C_HEADER_FG, size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def cell_font(color="000000", bold=False, size=10):
    return Font(color=color, bold=bold, size=size, name="Calibri")

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="DDE2EC")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

# ── TOP BANNER (rows 1-3) ───────────────────────────────────────────────────
ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 20
ws.row_dimensions[3].height = 18

# Stat 1 — HÔM NAY (col A)
ws["A1"] = "📊 HÔM NAY"
ws["A1"].font  = Font(bold=True, size=9, color="6B7280", name="Calibri")
ws["A1"].fill  = fill(C_BANNER_BG)
ws["A1"].alignment = center()
ws["B1"] = 12
ws["B1"].font  = Font(bold=True, size=18, color="1E3A5F", name="Calibri")
ws["B1"].alignment = center()
ws["C1"] = "quotes"
ws["C1"].font  = Font(size=9, color="6B7280", name="Calibri")
ws["C1"].alignment = center()
ws["D1"] = "3 WIN · $4,200"
ws["D1"].font  = Font(size=9, color="10B981", bold=True, name="Calibri")
ws["D1"].alignment = center()

# Stat 2 — TUẦN NÀY (col F)
ws["F1"] = "📅 TUẦN NÀY"
ws["F1"].font  = Font(bold=True, size=9, color="6B7280", name="Calibri")
ws["F1"].fill  = fill(C_BANNER_BG)
ws["F1"].alignment = center()
ws["G1"] = 67
ws["G1"].font  = Font(bold=True, size=18, color="1E3A5F", name="Calibri")
ws["G1"].alignment = center()
ws["H1"] = "quotes"
ws["H1"].font  = Font(size=9, color="6B7280", name="Calibri")
ws["H1"].alignment = center()
ws["I1"] = "19 WIN · ~$28K"
ws["I1"].font  = Font(size=9, color="10B981", bold=True, name="Calibri")
ws["I1"].alignment = center()

# Stat 3 — TOP CUSTOMER (col K)
ws["K1"] = "🌟 TOP CUSTOMER"
ws["K1"].font  = Font(bold=True, size=9, color="6B7280", name="Calibri")
ws["K1"].fill  = fill(C_BANNER_BG)
ws["K1"].alignment = center()
ws["L1"] = "GOWIN"
ws["L1"].font  = Font(bold=True, size=18, color="8B5CF6", name="Calibri")
ws["L1"].alignment = center()
ws["M1"] = "$8.5K"
ws["M1"].font  = Font(size=9, color="6B7280", name="Calibri")
ws["M1"].alignment = center()

# Stat 4 — NEW (col N)
ws["N1"] = "✨ QUOTE MỚI"
ws["N1"].font  = Font(bold=True, size=9, color="6B7280", name="Calibri")
ws["N1"].fill  = fill(C_BANNER_BG)
ws["N1"].alignment = center()
ws["N2"] = "+3"
ws["N2"].font  = Font(bold=True, size=18, color="06B6D4", name="Calibri")
ws["N2"].alignment = center()
ws["N3"] = "chờ xử lý"
ws["N3"].font  = Font(size=9, color="6B7280", name="Calibri")
ws["N3"].alignment = center()

# ── HEADER ROW (row 4) ───────────────────────────────────────────────────────
HEADERS = ["#", "Customer", "Carrier", "POL", "POD",
           "20GP", "40GP", "40HC", "45HC", "40NOR",
           "Margin", "Status", "Date", "Trend"]
ws.row_dimensions[4].height = 20

for col_idx, header in enumerate(HEADERS, 1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font   = hdr_font()
    cell.fill   = fill(C_HEADER_BG)
    cell.border = thin_border()
    cell.alignment = center()

# ── DATA ROWS (rows 5 to 4+N_ROWS) ─────────────────────────────────────────
random.seed(42)
for r in range(N_ROWS):
    row_data = make_row(r)
    excel_row = 5 + r
    ws.row_dimensions[excel_row].height = 18

    # alternating row fill
    row_bg = C_ALT_ROW if r % 2 == 1 else "FFFFFF"

    for c_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=excel_row, column=c_idx, value=value)
        cell.border    = thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # col 2 = Customer name
        if c_idx == 2:
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.font = Font(size=10, name="Calibri")
            if "🌟" in str(value):
                cell.font = Font(size=10, bold=True, color=C_VIP_TXT, name="Calibri")
            elif "🆕" in str(value):
                cell.font = Font(size=10, bold=True, color=C_NEW_TXT, name="Calibri")
            else:
                cell.font = Font(size=10, color=C_REG_TXT, name="Calibri")
            if row_bg != "FFFFFF":
                cell.fill = fill(row_bg)
        # col 3 = Carrier
        elif c_idx == 3:
            cell.font = Font(size=10, bold=True, color="1E3A5F", name="Calibri")
            if row_bg != "FFFFFF":
                cell.fill = fill(row_bg)
        # rate cols (6-10): numeric
        elif c_idx in (6, 7, 8, 9, 10):
            if value is not None:
                cell.number_format = '#,##0'
                cell.font = Font(size=10, color="6B7280", name="Calibri")
            else:
                cell.value = "—"
                cell.font = Font(size=10, color="D1D5DB", name="Calibri")
            if row_bg != "FFFFFF":
                cell.fill = fill(row_bg)
        # col 11 = Margin
        elif c_idx == 11:
            cell.number_format = '$#,##0'
            cell.font = Font(size=10, bold=True, name="Calibri")
            if row_bg != "FFFFFF":
                cell.fill = fill(row_bg)
        # col 12 = Status
        elif c_idx == 12:
            s = str(value)
            if s == "WIN":
                cell.fill = fill(C_WIN)
                cell.font = Font(size=10, bold=True, color=C_WIN_TXT, name="Calibri")
            elif s == "LOST":
                cell.fill = fill(C_LOST)
                cell.font = Font(size=10, bold=True, color=C_LOST_TXT, name="Calibri")
            elif s == "PENDING":
                cell.fill = fill(C_PENDING)
                cell.font = Font(size=10, bold=True, color=C_PENDING_TXT, name="Calibri")
            else:  # EXPIRED
                cell.fill = fill(C_EXPIRED)
                cell.font = Font(size=10, bold=True, color=C_EXPIRED_TXT, name="Calibri")
        else:
            cell.font = Font(size=10, name="Calibri")
            if row_bg != "FFFFFF":
                cell.fill = fill(row_bg)

# ── CONDITIONAL FORMATTING ──────────────────────────────────────────────────
data_last_row = 4 + N_ROWS  # row 24 if N_ROWS=20

# Margin heatmap (col 11 = K) — red <50, yellow 50-150, green >150
margin_range = f"K5:K{data_last_row}"

# Three discrete color rules for margin
margin_hot = CellIsRule(operator="lessThan", formula=["50"],
    fill=PatternFill("solid", fgColor=C_MARGIN_HOT))
margin_warm = CellIsRule(operator="between", formula=["50", "149"],
    fill=PatternFill("solid", fgColor=C_MARGIN_WARM))
margin_cool = CellIsRule(operator="greaterThanOrEqual", formula=["150"],
    fill=PatternFill("solid", fgColor=C_MARGIN_COOL))

ws.conditional_formatting.add(margin_range, margin_hot)
ws.conditional_formatting.add(margin_range, margin_warm)
ws.conditional_formatting.add(margin_range, margin_cool)

# Status colors via discrete CF (col 12 = L)
status_range = f"L5:L{data_last_row}"
for status, bg, fg in [
    ("WIN",     C_WIN,     C_WIN_TXT),
    ("LOST",    C_LOST,    C_LOST_TXT),
    ("PENDING", C_PENDING, C_PENDING_TXT),
    ("EXPIRED", C_EXPIRED, C_EXPIRED_TXT),
]:
    rule = CellIsRule(operator="equal", formula=[f'"{status}"'],
                      fill=PatternFill("solid", fgColor=bg),
                      font=Font(bold=True, color=fg, name="Calibri", size=10))
    ws.conditional_formatting.add(status_range, rule)

# ── SPARKLINE DATA (cols O-Z = 15-26, hidden) ──────────────────────────────
# Generate 12-week history per row (cols O..Z = 15..26)
random.seed(42)
for r in range(N_ROWS):
    excel_row = 5 + r
    # Use 40HC col (index 8) as base for sparkline, fallback to 20GP if missing
    base_val = CUSTOMERS[r % len(CUSTOMERS)]
    base = random.randint(2800, 5000)
    trend = 1 if r % 3 == 0 else (-1 if r % 3 == 1 else 0)
    for c in range(15, 27):  # cols O(15) to Z(26)
        val = base * (0.88 + 0.24 * random.random()) + trend * 30 * (c - 15)
        ws.cell(row=excel_row, column=c, value=round(val, 2))

# Hide sparkline data cols O-Z
for col_letter in [get_column_letter(c) for c in range(15, 27)]:
    ws.column_dimensions[col_letter].hidden = True

# ── TREND COLUMN (col N = 14) — emoji fallback ────────────────────────────
# openpyxl sparkline API not available in this version; use emoji indicators
trends = ["📈", "📉", "➡️"]
for r in range(N_ROWS):
    excel_row = 5 + r
    ws.cell(row=excel_row, column=14, value=trends[r % 3])
    ws.cell(row=excel_row, column=14).alignment = center()
    ws.cell(row=excel_row, column=14).font = Font(size=11, name="Calibri")

# ── EXCEL TABLE (for slicer support) ────────────────────────────────────────
table_ref = f"A4:N{data_last_row}"
tbl = Table(displayName="QuotesTable", ref=table_ref)
tbl.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False)
ws.add_table(tbl)

# ── FROZEN PANES ─────────────────────────────────────────────────────────────
ws.freeze_panes = "A5"   # freeze rows 1-4 + col A

# ── COLUMN WIDTHS ────────────────────────────────────────────────────────────
widths = {
    "A": 4,
    "B": 26,
    "C": 8,
    "D": 6,
    "E": 12,
    "F": 8,
    "G": 8,
    "H": 8,
    "I": 8,
    "J": 8,
    "K": 9,
    "L": 10,
    "M": 12,
    "N": 10,
}
for col_letter, w in widths.items():
    ws.column_dimensions[col_letter].width = w

# ── SAVE ─────────────────────────────────────────────────────────────────────
import os
os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
wb.save(OUTPUT)
print(f"[OK] Saved: {OUTPUT}")

# ── VERIFY ───────────────────────────────────────────────────────────────────
from openpyxl import load_workbook
vb = load_workbook(OUTPUT)
vw = vb.active
print(f"   Sheet:      {vw.title}")
print(f"   Rows:       {vw.max_row}")
print(f"   Columns:    {vw.max_column}")
print(f"   Frozen:     {vw.freeze_panes}")
print(f"   Tables:     {[t.name for t in vw.tables.values()]}")
print(f"   CF rules:   {len(vw.conditional_formatting._cf_rules)}")
