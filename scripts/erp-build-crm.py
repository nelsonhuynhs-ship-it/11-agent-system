"""
erp-build-crm.py — Build CRM sheet from customer_rules.json + Shipments.xlsx
=============================================================================
Merges two sources into the CRM sheet of ERP_Master_v14.xlsm:
  1. D:/OneDrive/NelsonData/email/customer_rules.json  (59 customers)
  2. C:/Users/Nelson/OneDrive/Desktop/Shipments.xlsx   (48 shipment records)

Idempotent: re-run detects existing rows by normalized Customer_Name → UPDATE, not duplicate.
Preserves row 2 (NAFOODS GROUP) untouched.
Uses save_preserving_ribbon (gotcha #6) — NEVER wb.save().

Usage:
    python scripts/erp-build-crm.py
    python scripts/erp-build-crm.py --dry-run
    python scripts/erp-build-crm.py --erp-file "D:/path/to/ERP_Master_v14.xlsm"

Exit: 0 = success | 1 = error
"""
from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Final

import openpyxl

# ---------------------------------------------------------------------------
# Sys path — allow importing ribbon_guard from repo root
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(SCRIPT_DIR)
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

from ERP.core.ribbon_guard import save_preserving_ribbon  # noqa: E402

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ERP_FILE_DEFAULT: Final = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"
CUSTOMER_RULES_PATH: Final = r"D:\OneDrive\NelsonData\email\customer_rules.json"
SHIPMENTS_PATH: Final = r"C:\Users\Nelson\OneDrive\Desktop\Shipments.xlsx"

CRM_SHEET: Final = "CRM"
NAFOODS_ROW: Final = 2          # row index (1-based) that holds NAFOODS GROUP
NAFOODS_CRM_ID: Final = "CS001289"
CRM_ID_PREFIX: Final = "CS"
CRM_ID_DIGITS: Final = 6        # zero-padded: CS001290
SALES_OWNER_DEFAULT: Final = "NELSON"

# Columns (1-based) — must match 43-col schema
COL = {
    "CRM_ID": 1,
    "Customer_Name": 2,
    "Customer_Type": 3,
    "Sales_Owner": 4,
    "FAS_ID": 5,
    "Status": 6,
    "Contact1_Name": 7,
    "Contact1_Email": 8,
    "Contact1_Phone": 9,
    "Contact2_Name": 10,
    "Contact2_Email": 11,
    "Contact2_Phone": 12,
    "Preferred_Carriers": 13,
    "POL_Options": 14,
    "POD_Options": 15,
    "Container_Types": 16,
    "Commodity": 17,
    "HS_Code": 18,
    "GW_Per_Container": 19,
    "Stuffing_Place": 20,
    "MT_Pickup_ICD": 21,
    "Full_Return_Port": 22,
    "Is_Reefer": 23,
    "Reefer_Temp": 24,
    "Reefer_Ventilation": 25,
    "Reefer_Humidity": 26,
    "Switch_Bill": 27,
    "ISF_Filer": 28,
    "Shipper_Fix_or_Byshipment": 29,
    "Consignee_Fix_or_Byshipment": 30,
    "Payer": 31,
    "Credit_Term": 32,
    "Invoice_Trigger": 33,
    "Debit_Note_Trigger": 34,
    "BL_Fee": 35,
    "THC_20": 36,
    "THC_40": 37,
    "ENS_AMS": 38,
    "Seal_Fee": 39,
    "Telex_Release": 40,
    "HDL_Fee_Note": 41,
    "Carrier_KB_Pct": 42,
    "Special_Notes": 43,
}
TOTAL_COLS: Final = 43

# Monthly sheets to parse — exclude helper sheets
SKIP_SHEETS: Final = {"Sheet1", "Sheet2", "Sheet3"}

# Shipments.xlsx columns (from header inspection)
SHIP_COL_CUSTOMER: Final = "Customer"
SHIP_COL_ROUTING: Final = "Routing"
SHIP_COL_CONTAINER_TYPE: Final = "Container Type"
SHIP_COL_CARRIER: Final = "Carrier"
SHIP_COL_ETD: Final = "ETD"

# How far back to consider "recent" for Active status
RECENT_MONTHS: Final = 6

# ---------------------------------------------------------------------------
# Helper: normalize customer name for dedup
# ---------------------------------------------------------------------------
_PUNCT_RE = re.compile(r"[^\w\s]")


def normalize(name: str) -> str:
    """Uppercase, strip punctuation, collapse whitespace."""
    n = _PUNCT_RE.sub(" ", name.upper())
    return " ".join(n.split())


def title_case(name: str) -> str:
    """Title-case for display, handle edge cases."""
    return name.title()


# ---------------------------------------------------------------------------
# Helper: parse routes from customer_rules format
# "HPH-US"   → POL=HPH, POD=US
# "HCM-LAX"  → POL=HCM, POD=LAX
# "HPH-VAN"  → POL=HPH, POD=VAN
# ---------------------------------------------------------------------------
def parse_routes(routes: list[str]) -> tuple[set[str], set[str]]:
    pols: set[str] = set()
    pods: set[str] = set()
    for r in routes:
        if "-" in r:
            parts = r.split("-", 1)
            pol = parts[0].strip().upper()
            pod = parts[1].strip().upper()
            if pol:
                pols.add(pol)
            if pod:
                pods.add(pod)
    return pols, pods


# ---------------------------------------------------------------------------
# Helper: extract POL/POD from Shipments routing string
# "HPH-LAX"           → HPH, LAX
# "HCM-DENVER VIA OAK"→ HCM, DENVER VIA OAK (keep full pod as-is, normalised)
# ---------------------------------------------------------------------------
def parse_routing(routing: str) -> tuple[str | None, str | None]:
    if not routing or not isinstance(routing, str):
        return None, None
    routing = routing.strip()
    if "-" not in routing:
        return None, None
    parts = routing.split("-", 1)
    pol = parts[0].strip().upper()
    pod = parts[1].strip().upper()
    return (pol or None, pod or None)


# ---------------------------------------------------------------------------
# Load source 1: customer_rules.json
# ---------------------------------------------------------------------------
def load_customer_rules(path: str) -> dict:
    """
    Returns dict[normalized_name] -> record with fields needed for CRM.
    """
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    customers_raw = data.get("customers", {})
    result: dict[str, dict] = {}

    for name, rec in customers_raw.items():
        key = normalize(name)

        # Map type
        raw_type = rec.get("type", "DIRECT").upper()
        if raw_type == "DIRECT":
            crm_type = "BCO"
        elif raw_type == "FWD":
            crm_type = "FORWARDER"
        elif raw_type == "CNEE":
            crm_type = "CONSIGNEE"
        else:
            crm_type = raw_type

        # Carriers
        carrier_list = rec.get("carrier_affinity", [])
        # Normalise each carrier: "ONE SOC" → "ONE", "CMA" → "CMA"
        carriers_clean: list[str] = []
        for c in carrier_list:
            c_clean = str(c).split()[0].strip().upper()  # first word only
            if c_clean and c_clean not in carriers_clean:
                carriers_clean.append(c_clean)

        # Routes
        pols, pods = parse_routes(rec.get("routes", []))

        # Contact email
        seen = rec.get("seen_senders", [])
        contact1_email = seen[0] if seen else None

        result[key] = {
            "display_name": title_case(name),
            "crm_type": crm_type,
            "carriers": carriers_clean,
            "pols": pols,
            "pods": pods,
            "contact1_email": contact1_email,
            "notes": rec.get("notes", ""),
            "shipment_count": rec.get("shipment_count", 0),
        }

    return result


# ---------------------------------------------------------------------------
# Load source 2: Shipments.xlsx — aggregate per customer
# ---------------------------------------------------------------------------
def load_shipments(path: str) -> dict:
    """
    Returns dict[normalized_name] -> aggregated fields:
        carriers: list (most-used first)
        pols: set
        pods: set
        container_types: set
        shipment_count: int
        last_etd: datetime | None
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    carrier_counter: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    pols_map: dict[str, set[str]] = defaultdict(set)
    pods_map: dict[str, set[str]] = defaultdict(set)
    ct_map: dict[str, set[str]] = defaultdict(set)
    count_map: dict[str, int] = defaultdict(int)
    last_etd_map: dict[str, datetime | None] = {}

    for sname in wb.sheetnames:
        if sname in SKIP_SHEETS:
            continue
        ws = wb[sname]
        headers: list[str] | None = None

        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(h).strip() if h else "" for h in row]
                continue
            if not row or not row[0]:
                continue
            row_dict = dict(zip(headers, row))

            raw_cust = row_dict.get(SHIP_COL_CUSTOMER, "")
            if not raw_cust or str(raw_cust).strip().upper() in ("NONE", ""):
                continue

            key = normalize(str(raw_cust).strip())

            # Carrier
            carrier_raw = row_dict.get(SHIP_COL_CARRIER, "")
            if carrier_raw and str(carrier_raw).strip().upper() not in ("NONE", ""):
                c_main = str(carrier_raw).split()[0].strip().upper()
                if c_main:
                    carrier_counter[key][c_main] += 1

            # Routing
            routing = row_dict.get(SHIP_COL_ROUTING, "")
            pol, pod = parse_routing(str(routing) if routing else "")
            if pol:
                pols_map[key].add(pol)
            if pod:
                pods_map[key].add(pod)

            # Container type
            ct = row_dict.get(SHIP_COL_CONTAINER_TYPE, "")
            if ct and str(ct).strip().upper() not in ("NONE", ""):
                ct_norm = str(ct).strip().upper()
                ct_map[key].add(ct_norm)

            # Count + ETD
            count_map[key] += 1
            etd = row_dict.get(SHIP_COL_ETD)
            if isinstance(etd, datetime):
                prev = last_etd_map.get(key)
                if prev is None or etd > prev:
                    last_etd_map[key] = etd

    wb.close()

    result: dict[str, dict] = {}
    for key in count_map:
        # Sort carriers by frequency
        c_sorted = sorted(carrier_counter[key], key=lambda x: carrier_counter[key][x], reverse=True)
        result[key] = {
            "carriers": c_sorted,
            "pols": pols_map[key],
            "pods": pods_map[key],
            "container_types": ct_map[key],
            "shipment_count": count_map[key],
            "last_etd": last_etd_map.get(key),
        }

    return result


# ---------------------------------------------------------------------------
# Merge: combine rules + shipments into unified customer records
# ---------------------------------------------------------------------------
def merge_sources(rules: dict, shipments: dict) -> list[dict]:
    """
    Merge both sources. Returns list of dicts ready for CRM rows.
    Sorted by shipment_count DESC.
    """
    all_keys: set[str] = set(rules.keys()) | set(shipments.keys())
    merged: list[dict] = []

    six_months_ago = datetime.now() - timedelta(days=RECENT_MONTHS * 30)

    for key in all_keys:
        r = rules.get(key, {})
        s = shipments.get(key, {})

        # Display name: prefer rules (usually cleaner), else reconstruct from key
        display_name = r.get("display_name") or title_case(key)

        # Customer type: prefer rules
        crm_type = r.get("crm_type", "BCO")

        # Carriers: start with rules list, add shipments top carriers if new
        carriers = list(r.get("carriers", []))
        for c in s.get("carriers", []):
            if c not in carriers:
                carriers.append(c)

        # POL/POD: union of both sources
        pols = r.get("pols", set()) | s.get("pols", set())
        pods = r.get("pods", set()) | s.get("pods", set())

        # Container types: only from shipments (rules don't have this)
        container_types = s.get("container_types", set())

        # Is_Reefer: Yes if any RF container type
        is_reefer = "Yes" if any("RF" in ct for ct in container_types) else None

        # Shipment count: max of both sources
        ship_count_rules = r.get("shipment_count", 0)
        ship_count_ship = s.get("shipment_count", 0)
        total_shipment_count = max(ship_count_rules, ship_count_ship)

        # Status: Active if has recent shipment
        last_etd = s.get("last_etd")
        if last_etd and last_etd >= six_months_ago:
            status = "Active"
        elif total_shipment_count > 0:
            status = "Active"
        else:
            status = "Prospect"

        # Contact
        contact1_email = r.get("contact1_email")

        # Notes
        notes = r.get("notes", "")

        merged.append({
            "norm_key": key,
            "display_name": display_name,
            "crm_type": crm_type,
            "status": status,
            "contact1_email": contact1_email,
            "carriers": carriers,
            "pols": sorted(pols),
            "pods": sorted(pods),
            "container_types": sorted(container_types),
            "is_reefer": is_reefer,
            "notes": notes,
            "shipment_count": total_shipment_count,
        })

    # Sort by shipment_count DESC, then name ASC
    merged.sort(key=lambda x: (-x["shipment_count"], x["display_name"]))
    return merged


# ---------------------------------------------------------------------------
# Check file lock (Excel open check via lock file)
# ---------------------------------------------------------------------------
def check_file_lock(path: str) -> bool:
    """Returns True if file appears to be open/locked by Excel."""
    dir_path = os.path.dirname(path)
    fname = os.path.basename(path)
    lock_file = os.path.join(dir_path, f"~${fname}")
    return os.path.exists(lock_file)


# ---------------------------------------------------------------------------
# Read existing CRM rows to build dedup index
# Returns: (max_crm_id_int, dict[norm_key -> row_number (1-based)])
# ---------------------------------------------------------------------------
def read_existing_crm(ws) -> tuple[int, dict[str, int]]:
    max_id = int(NAFOODS_CRM_ID.replace(CRM_ID_PREFIX, ""))
    existing: dict[str, int] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cust_name = row[COL["Customer_Name"] - 1]
        crm_id_val = row[COL["CRM_ID"] - 1]
        if not cust_name:
            continue
        norm_key = normalize(str(cust_name))
        existing[norm_key] = row_idx

        # Track max CRM_ID
        if crm_id_val and str(crm_id_val).startswith(CRM_ID_PREFIX):
            try:
                id_int = int(str(crm_id_val)[len(CRM_ID_PREFIX):])
                if id_int > max_id:
                    max_id = id_int
            except ValueError:
                pass

    return max_id, existing


# ---------------------------------------------------------------------------
# Build a single row dict (43 values) from merged record
# ---------------------------------------------------------------------------
def build_row_values(rec: dict, crm_id: str) -> list:
    carriers_str = "/".join(rec["carriers"]) if rec["carriers"] else None
    pol_str = "/".join(rec["pols"]) if rec["pols"] else None
    pod_str = "/".join(rec["pods"]) if rec["pods"] else None
    ct_str = "/".join(rec["container_types"]) if rec["container_types"] else None

    row = [None] * TOTAL_COLS
    row[COL["CRM_ID"] - 1] = crm_id
    row[COL["Customer_Name"] - 1] = rec["display_name"]
    row[COL["Customer_Type"] - 1] = rec["crm_type"]
    row[COL["Sales_Owner"] - 1] = SALES_OWNER_DEFAULT
    row[COL["Status"] - 1] = rec["status"]
    row[COL["Contact1_Email"] - 1] = rec["contact1_email"]
    row[COL["Preferred_Carriers"] - 1] = carriers_str
    row[COL["POL_Options"] - 1] = pol_str
    row[COL["POD_Options"] - 1] = pod_str
    row[COL["Container_Types"] - 1] = ct_str
    row[COL["Is_Reefer"] - 1] = rec["is_reefer"]
    row[COL["Special_Notes"] - 1] = rec["notes"] or None
    return row


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Build/update CRM sheet in ERP_Master_v14.xlsm from customer_rules.json + Shipments.xlsx"
    )
    parser.add_argument("--dry-run", action="store_true", help="Print planned rows without saving")
    parser.add_argument("--erp-file", default=ERP_FILE_DEFAULT, help="Path to ERP_Master_v14.xlsm")
    parser.add_argument("--customer-rules", default=CUSTOMER_RULES_PATH)
    parser.add_argument("--shipments", default=SHIPMENTS_PATH)
    args = parser.parse_args(argv)

    erp_file = args.erp_file
    dry_run = args.dry_run

    print(f"[*] erp-build-crm.py {'(DRY-RUN)' if dry_run else ''}")
    print(f"    ERP file     : {erp_file}")
    print(f"    Rules source : {args.customer_rules}")
    print(f"    Shipments    : {args.shipments}")

    # --- Validate source files ---
    for path, label in [(args.customer_rules, "customer_rules.json"), (args.shipments, "Shipments.xlsx")]:
        if not os.path.exists(path):
            print(f"[ERROR] Source not found: {label} @ {path}")
            return 1

    if not os.path.exists(erp_file):
        print(f"[ERROR] ERP file not found: {erp_file}")
        return 1

    # --- File lock check ---
    if check_file_lock(erp_file):
        print(f"[ERROR] ERP file appears to be open in Excel. Close it first.")
        return 1

    # --- Load sources ---
    print("[*] Loading customer_rules.json ...")
    rules = load_customer_rules(args.customer_rules)
    print(f"    {len(rules)} customers loaded from rules")

    print("[*] Loading Shipments.xlsx ...")
    shipments = load_shipments(args.shipments)
    print(f"    {len(shipments)} unique customers aggregated from shipments")

    # --- Merge ---
    print("[*] Merging sources ...")
    merged = merge_sources(rules, shipments)
    print(f"    {len(merged)} total unique customers after merge")

    # --- Filter NAFOODS (row 2 kept as-is) ---
    # Skip any customer whose normalized name matches NAFOODS GROUP variants
    nafoods_variants = {normalize("NAFOODS GROUP"), normalize("NAFOODS"), normalize("NAFOOD GROUP")}

    merged_filtered = [r for r in merged if r["norm_key"] not in nafoods_variants]
    skipped_nafoods = len(merged) - len(merged_filtered)
    if skipped_nafoods:
        print(f"    Skipped {skipped_nafoods} NAFOODS variant(s) — already in row 2")

    # --- Open workbook ---
    print("[*] Opening ERP workbook (keep_vba=True) ...")
    wb = openpyxl.load_workbook(erp_file, keep_vba=True)
    ws = wb[CRM_SHEET]

    # --- Read existing rows ---
    max_id, existing_index = read_existing_crm(ws)
    print(f"    Existing CRM rows: {len(existing_index)} (max ID: {CRM_ID_PREFIX}{max_id:0{CRM_ID_DIGITS}d})")

    # --- Plan updates ---
    next_id = max_id + 1
    new_rows: list[dict] = []
    update_rows: list[dict] = []

    for rec in merged_filtered:
        key = rec["norm_key"]
        if key in existing_index:
            # UPDATE — will overwrite that row
            row_num = existing_index[key]
            crm_id = ws.cell(row=row_num, column=COL["CRM_ID"]).value or f"{CRM_ID_PREFIX}{next_id:0{CRM_ID_DIGITS}d}"
            rec["_row_num"] = row_num
            rec["_crm_id"] = crm_id
            update_rows.append(rec)
        else:
            # NEW
            crm_id = f"{CRM_ID_PREFIX}{next_id:0{CRM_ID_DIGITS}d}"
            next_id += 1
            rec["_crm_id"] = crm_id
            new_rows.append(rec)

    print(f"\n[*] Plan:")
    print(f"    NEW     : {len(new_rows)}")
    print(f"    UPDATE  : {len(update_rows)}")
    print(f"    SKIPPED : {skipped_nafoods + (len(merged) - len(merged_filtered))}")

    # --- Dry-run: print and exit ---
    if dry_run:
        print("\n--- DRY-RUN: Planned rows ---")
        print(f"{'CRM_ID':<12} {'Customer_Name':<35} {'Type':<12} {'Status':<10} {'Carriers':<20} {'POL':<12} {'POD':<30} {'CT':<20} {'Action'}")
        print("-" * 165)
        for rec in update_rows:
            carriers = "/".join(rec["carriers"][:3]) if rec["carriers"] else "-"
            pols = "/".join(rec["pols"][:3]) if rec["pols"] else "-"
            pods = "/".join(rec["pods"][:4]) if rec["pods"] else "-"
            cts = "/".join(rec["container_types"][:3]) if rec["container_types"] else "-"
            print(f"{rec['_crm_id']:<12} {rec['display_name']:<35} {rec['crm_type']:<12} {rec['status']:<10} {carriers:<20} {pols:<12} {pods:<30} {cts:<20} UPDATE row {rec['_row_num']}")
        for rec in new_rows:
            carriers = "/".join(rec["carriers"][:3]) if rec["carriers"] else "-"
            pols = "/".join(rec["pols"][:3]) if rec["pols"] else "-"
            pods = "/".join(rec["pods"][:4]) if rec["pods"] else "-"
            cts = "/".join(rec["container_types"][:3]) if rec["container_types"] else "-"
            print(f"{rec['_crm_id']:<12} {rec['display_name']:<35} {rec['crm_type']:<12} {rec['status']:<10} {carriers:<20} {pols:<12} {pods:<30} {cts:<20} NEW")
        print(f"\n[DRY-RUN] Summary: +{len(new_rows)} new / {len(update_rows)} updated / {skipped_nafoods} skipped (NAFOODS)")
        wb.close()
        return 0

    # --- Backup ---
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = erp_file.replace(".xlsm", f".backup_{ts}.xlsm")
    print(f"\n[*] Backing up to: {backup_path}")
    shutil.copy2(erp_file, backup_path)

    # --- Apply UPDATEs ---
    for rec in update_rows:
        row_num = rec["_row_num"]
        crm_id = rec["_crm_id"]
        row_values = build_row_values(rec, crm_id)
        for col_idx, value in enumerate(row_values, start=1):
            if value is not None:  # Only overwrite non-None fields
                ws.cell(row=row_num, column=col_idx).value = value
        carriers = "/".join(rec["carriers"][:3]) if rec["carriers"] else "-"
        pols = "/".join(rec["pols"][:2]) if rec["pols"] else "-"
        pods = "/".join(rec["pods"][:3]) if rec["pods"] else "-"
        cts = "/".join(rec["container_types"][:2]) if rec["container_types"] else "-"
        print(f"[~] UPD  {crm_id}  {rec['display_name']:<35}  {rec['crm_type']:<12}  {carriers}  {pols}  {pods}  {cts}")

    # --- Append NEWs ---
    # Find first empty row after last data row
    last_data_row = 1
    for row in ws.iter_rows(min_row=2, values_only=False):
        if any(cell.value for cell in row):
            if row[0].row > last_data_row:
                last_data_row = row[0].row

    append_row = last_data_row + 1

    for rec in new_rows:
        crm_id = rec["_crm_id"]
        row_values = build_row_values(rec, crm_id)
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=append_row, column=col_idx).value = value
        carriers = "/".join(rec["carriers"][:3]) if rec["carriers"] else "-"
        pols = "/".join(rec["pols"][:2]) if rec["pols"] else "-"
        pods = "/".join(rec["pods"][:3]) if rec["pods"] else "-"
        cts = "/".join(rec["container_types"][:2]) if rec["container_types"] else "-"
        print(f"[+] NEW  {crm_id}  {rec['display_name']:<35}  {rec['crm_type']:<12}  {carriers}  {pols}  {pods}  {cts}")
        append_row += 1

    # --- Save (preserve ribbon — gotcha #6) ---
    print(f"\n[*] Saving via save_preserving_ribbon ...")
    result = save_preserving_ribbon(wb, erp_file)
    print(f"    Ribbon result: {result}")

    print(f"\n[OK] Done: +{len(new_rows)} new / {len(update_rows)} updated / {skipped_nafoods} skipped (NAFOODS)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
