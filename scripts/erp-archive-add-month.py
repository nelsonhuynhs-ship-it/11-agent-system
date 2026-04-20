"""
erp-archive-add-month.py — Backfill MONTH col (col 15) in Archive sheet
========================================================================
Archive sheet previously had 14 cols. This one-shot script:
  1. Ensures MONTH header at Archive row 2, col 15
  2. Derives MONTH from Delivered_Date (col 13) for each existing data row
     - Format: "APR-26" (matches Active Jobs MONTH format)
  3. Saves via save_preserving_ribbon (gotcha #6)

Idempotent: rows already having a MONTH value are skipped.

Usage:
    python scripts/erp-archive-add-month.py [--target PATH] [--dry-run] [--overwrite]

--overwrite: re-derive MONTH for ALL rows (even if already populated)
"""
from __future__ import annotations

import argparse
import os
import shutil
import sys
from datetime import datetime

import openpyxl

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from ERP.core.ribbon_guard import save_preserving_ribbon  # noqa: E402

DEFAULT_TARGET = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"

ARCHIVE_SHEET = "Archive"
ARCHIVE_HDR_ROW = 2
ARCHIVE_DATA_START = 3

# Fixed col positions in Archive sheet
ARCH_DELIVERED_DATE_COL = 13   # Delivered_Date
ARCH_MONTH_COL = 15            # MONTH (new col, appended)


def _log(msg: str) -> None:
    safe = msg.encode(sys.stdout.encoding or "utf-8", errors="replace").decode(
        sys.stdout.encoding or "utf-8", errors="replace"
    )
    print(safe, flush=True)


def _check_excel_closed(path: str) -> None:
    try:
        with open(path, "a"):
            pass
    except PermissionError:
        _log(f"[ERROR] File locked — close Excel first: {path}")
        sys.exit(1)


def _backup(path: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(path)
    backup_path = f"{base}.backup_{ts}{ext}"
    shutil.copy2(path, backup_path)
    _log(f"[backup] {backup_path}")
    return backup_path


def _derive_month_from_date(val) -> str:
    """Derive 'APR-26' from a date cell value (datetime or string)."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%b-%y").upper()
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return ""
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).strftime("%b-%y").upper()
            except ValueError:
                continue
    # openpyxl may return date as serial (int/float) — skip
    return ""


def backfill_archive_month(target_path: str, dry_run: bool, overwrite: bool) -> None:
    _log("=" * 60)
    _log(f"[erp-archive-add-month] {'DRY RUN — ' if dry_run else ''}Starting")
    _log(f"  Target   : {target_path}")
    _log(f"  Overwrite: {overwrite}")
    _log("=" * 60)

    if not os.path.exists(target_path):
        _log(f"[ERROR] Target not found: {target_path}")
        sys.exit(1)

    if dry_run:
        _log("[DRY RUN] Would add MONTH col 15 to Archive sheet.")
        _log("  Derives from: Delivered_Date (col 13)")
        _log("  Format: APR-26")
        _log("[DRY RUN] No changes written.")
        return

    _check_excel_closed(target_path)
    _backup(target_path)

    _log("\n[Step 1] Opening workbook ...")
    wb = openpyxl.load_workbook(target_path, keep_vba=True)

    if ARCHIVE_SHEET not in wb.sheetnames:
        _log(f"[ERROR] Sheet '{ARCHIVE_SHEET}' not found in workbook.")
        sys.exit(1)

    ws = wb[ARCHIVE_SHEET]

    # Ensure header
    ws.cell(row=ARCHIVE_HDR_ROW, column=ARCH_MONTH_COL).value = "MONTH"
    _log(f"  [OK] Archive row {ARCHIVE_HDR_ROW}, col {ARCH_MONTH_COL} = 'MONTH' (header set)")

    # Backfill data rows
    filled = skipped = already_set = 0
    max_row = ws.max_row or ARCHIVE_DATA_START

    for row_num in range(ARCHIVE_DATA_START, max_row + 1):
        # Skip empty rows (no Bkg_No)
        bkg_cell = ws.cell(row=row_num, column=6)  # Bkg_No = col 6
        if bkg_cell.value is None or str(bkg_cell.value).strip() == "":
            skipped += 1
            continue

        month_cell = ws.cell(row=row_num, column=ARCH_MONTH_COL)
        existing = month_cell.value

        if existing and str(existing).strip() and not overwrite:
            already_set += 1
            continue

        # Derive from Delivered_Date
        delivered_val = ws.cell(row=row_num, column=ARCH_DELIVERED_DATE_COL).value
        month_label = _derive_month_from_date(delivered_val)

        if month_label:
            month_cell.value = month_label
            filled += 1
        else:
            # No Delivered_Date — leave blank (VBA combo will treat as unknown)
            skipped += 1

    _log(f"\n  Results: filled={filled} already_set={already_set} skipped={skipped}")

    # Save via ribbon guard (gotcha #6)
    _log("\n[Step 2] Saving (preserving ribbon) ...")
    result = save_preserving_ribbon(wb, target_path)
    _log(f"  Ribbon guard result: {result}")

    _log("\n[DONE] Archive MONTH col backfill complete.")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Backfill MONTH col (col 15) in Archive sheet of ERP_Master_v14.xlsm"
    )
    parser.add_argument(
        "--target", default=DEFAULT_TARGET,
        help=f"Path to ERP_Master_v14.xlsm (default: {DEFAULT_TARGET})",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Preview — do not write to file",
    )
    parser.add_argument(
        "--overwrite", action="store_true",
        help="Re-derive MONTH for ALL rows (even if already set)",
    )
    args = parser.parse_args()
    backfill_archive_month(args.target, args.dry_run, args.overwrite)


if __name__ == "__main__":
    main()
