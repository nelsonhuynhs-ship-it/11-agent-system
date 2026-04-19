"""
erp-quote-polish.py — Quote sheet polish: KPI rows + conditional format + autofilter
=====================================================================================
Applies 3 fixes to the Quotes sheet in ERP_Master_v14.xlsm **idempotently**:

  Fix A — Conditional formatting on Status column (col 36)
           PENDING=yellow | WIN=green | LOST=red | EXPIRED=gray
           Stale row fade (Date < TODAY-7 AND Status=PENDING) → #EEEEEE

  Fix C — KPI rows at top (rows 1-3 dashboard, row 4 = original header)
           Data rows start at row 5. Idempotent: if A1=="📊 QUOTES TODAY"
           skip insert but refresh formulas + conditional format.

  Fix D — AutoFilter on row 4, Freeze pane at A5

Usage:
    python scripts/erp-quote-polish.py
    python scripts/erp-quote-polish.py --dry-run
    python scripts/erp-quote-polish.py --erp-file "D:/path/to/ERP_Master_v14.xlsm"

Constraints:
    - openpyxl.load_workbook(keep_vba=True)
    - save via ERP.core.ribbon_guard.save_preserving_ribbon (NEVER wb.save)
    - Backup created before write: ERP_Master_v14.backup_YYYYMMDD_HHMMSS.xlsm
    - Excel must be closed before run (file-lock check)
    - Exit 0 success / non-zero failure
"""
from __future__ import annotations

import argparse
import os
import shutil
import sys
from datetime import datetime
from typing import Final

import openpyxl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths & constants
# ---------------------------------------------------------------------------
DEFAULT_ERP_FILE: Final = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"

QUOTES_SHEET: Final = "Quotes"

# Column positions in Quotes sheet (1-based, from price_watch.py Q_COL mapping)
Q_DATE_COL: Final = 2       # Date
Q_STATUS_COL: Final = 36    # Status
Q_SELL_40GP_COL: Final = 30  # Sell_40GP
Q_BUY_40GP_COL: Final = 13   # Buy_40GP

# Excel column letters derived from 1-based positions
DATE_LETTER: Final = get_column_letter(Q_DATE_COL)       # B
STATUS_LETTER: Final = get_column_letter(Q_STATUS_COL)   # AJ
SELL_LETTER: Final = get_column_letter(Q_SELL_40GP_COL)  # AD
BUY_LETTER: Final = get_column_letter(Q_BUY_40GP_COL)    # M

# Last column in Quotes sheet (col AQ = 43 per Q_COL max ContType=42 + 1 buffer)
LAST_COL_LETTER: Final = "AQ"
HEADER_ROW: Final = 4       # After KPI insertion, original header is at row 4
DATA_START_ROW: Final = 5   # First data row after header
DATA_END_ROW: Final = 1000  # Max rows for formatting / filter range
KPI_TITLE: Final = "\U0001f4ca QUOTES TODAY"  # 📊 QUOTES TODAY

# Fill colors for conditional formatting
FILL_PENDING: Final = PatternFill(start_color="FFF4A3", end_color="FFF4A3", fill_type="solid")
FILL_WIN: Final = PatternFill(start_color="B6EEC8", end_color="B6EEC8", fill_type="solid")
FILL_LOST: Final = PatternFill(start_color="FFB8B8", end_color="FFB8B8", fill_type="solid")
FILL_EXPIRED: Final = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
FILL_STALE: Final = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

# KPI row styles
FILL_TITLE: Final = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FILL_LABELS: Final = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
FILL_VALUES: Final = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FONT_TITLE: Final = Font(bold=True, color="FFFFFF", size=13)
FONT_LABEL: Final = Font(bold=True, color="FFFFFF", size=10)
FONT_VALUE: Final = Font(bold=True, color="1F4E79", size=10)

# Thin border helper
_thin = Side(style="thin", color="BFBFBF")
BORDER_THIN: Final = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _log(msg: str) -> None:
    print(f"[erp-quote-polish] {msg}", flush=True)


def _check_excel_closed(erp_file: str) -> None:
    """Raise SystemExit if xlsm is currently open in Excel (file-lock check)."""
    try:
        with open(erp_file, "a"):
            pass
    except PermissionError:
        _log("ERROR: File is locked — close Excel first, then re-run.")
        sys.exit(1)


def _backup(erp_file: str) -> str:
    """Create a timestamped backup next to the live file. Returns backup path."""
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(erp_file)
    backup_path = f"{base}.backup_{stamp}{ext}"
    shutil.copy2(erp_file, backup_path)
    _log(f"Backup created: {backup_path}")
    return backup_path


def _kpi_already_inserted(ws) -> bool:
    """Return True if KPI rows are already present (A1 == KPI_TITLE)."""
    v = ws["A1"].value
    if v is None:
        return False
    return str(v).strip() == KPI_TITLE


# ---------------------------------------------------------------------------
# Fix C — KPI rows
# ---------------------------------------------------------------------------

def _insert_kpi_rows(ws) -> str:
    """Insert 3 KPI rows above the original header. Returns 'INSERTED'."""
    _log("  Inserting 3 KPI rows above header...")
    ws.insert_rows(1, amount=3)

    # --- Row 1: Dashboard title ---
    title_cell = ws["A1"]
    title_cell.value = KPI_TITLE
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_TITLE
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    # Merge A1:F1
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 22

    # --- Row 2: KPI labels ---
    labels = [
        "Today Count", "WIN", "LOST", "PENDING",
        f"Avg {SELL_LETTER}{DATA_START_ROW} (Sell_40GP)", "Avg Profit %",
    ]
    for col_idx, label in enumerate(labels, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = label
        cell.font = FONT_LABEL
        cell.fill = FILL_LABELS
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_THIN
    ws.row_dimensions[2].height = 18

    # --- Row 3: KPI formulas ---
    _write_kpi_formulas(ws)
    ws.row_dimensions[3].height = 18

    # Row 4 is original header (pushed down) — will be styled by autofilter step
    ws.row_dimensions[4].height = 20

    return "INSERTED"


def _write_kpi_formulas(ws) -> None:
    """Write/refresh COUNTIFS/AVERAGEIFS formulas into row 3."""
    data_range = f"{DATA_START_ROW}:{DATA_END_ROW}"
    date_range = f"{DATE_LETTER}{DATA_START_ROW}:{DATE_LETTER}{DATA_END_ROW}"
    status_range = f"{STATUS_LETTER}{DATA_START_ROW}:{STATUS_LETTER}{DATA_END_ROW}"
    sell_range = f"{SELL_LETTER}{DATA_START_ROW}:{SELL_LETTER}{DATA_END_ROW}"
    buy_range = f"{BUY_LETTER}{DATA_START_ROW}:{BUY_LETTER}{DATA_END_ROW}"

    # Col A3: Today count = rows where Date >= TODAY()
    ws["A3"].value = f'=COUNTIFS({date_range},">="&TODAY())'

    # Col B3: WIN today
    ws["B3"].value = (
        f'=COUNTIFS({date_range},">="&TODAY(),'
        f'{status_range},"WIN")'
    )

    # Col C3: LOST today
    ws["C3"].value = (
        f'=COUNTIFS({date_range},">="&TODAY(),'
        f'{status_range},"LOST")'
    )

    # Col D3: PENDING today
    ws["D3"].value = (
        f'=COUNTIFS({date_range},">="&TODAY(),'
        f'{status_range},"PENDING")'
    )

    # Col E3: Avg Sell_40GP today
    ws["E3"].value = (
        f'=AVERAGEIFS({sell_range},'
        f'{date_range},">="&TODAY())'
    )

    # Col F3: Avg Profit % today = AVERAGEIFS((Sell-Buy)/Buy, Date>=TODAY(), Buy>0)
    # Using AVERAGEIF on a helper expression is not native Excel; use SUMPRODUCT instead
    ws["F3"].value = (
        f'=IFERROR('
        f'SUMPRODUCT(({date_range}>=(TODAY()))*({buy_range}>0)'
        f'*(({sell_range}-{buy_range})/{buy_range}))'
        f'/SUMPRODUCT(({date_range}>=(TODAY()))*({buy_range}>0)*1),'
        f'0)'
    )

    # Style row 3
    for col_idx in range(1, 7):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = FONT_VALUE
        cell.fill = FILL_VALUES
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_THIN


def apply_kpi_rows(ws) -> str:
    """
    Idempotent KPI row insertion.
    Returns 'INSERTED' or 'REFRESHED'.
    """
    if _kpi_already_inserted(ws):
        _log("  KPI rows already present — refreshing formulas only.")
        _write_kpi_formulas(ws)
        return "REFRESHED"
    return _insert_kpi_rows(ws)


# ---------------------------------------------------------------------------
# Fix A — Conditional formatting
# ---------------------------------------------------------------------------

def apply_conditional_formatting(ws) -> None:
    """
    Apply status-based conditional formatting to rows DATA_START_ROW:DATA_END_ROW.

    Rules (applied to full row range A{r}:{LAST}:{r}, triggered by Status col):
      1. Status == "WIN"      → green fill
      2. Status == "LOST"     → red fill
      3. Status == "EXPIRED"  → gray fill
      4. Status == "PENDING"  → yellow fill
      5. Stale: Date < TODAY()-7 AND Status=="PENDING" → light gray fade

    Note: openpyxl conditional formatting uses FormulaRule for row-spanning.
    CellIsRule applies only to the target column; FormulaRule can span the row.
    We use FormulaRule for row-wide fill + a simple CellIsRule on status column
    for cell-only highlight as fallback.
    """
    _log("  Applying conditional formatting to Status column and rows...")

    # Clear existing CF rules on Quotes sheet to avoid duplicates on re-run
    ws.conditional_formatting = openpyxl.formatting.formatting.ConditionalFormattingList()

    row_range = f"A{DATA_START_ROW}:{LAST_COL_LETTER}{DATA_END_ROW}"
    # $AJ5 anchors the Status column reference per row in the formula
    status_ref = f"${STATUS_LETTER}{DATA_START_ROW}"
    date_ref = f"${DATE_LETTER}{DATA_START_ROW}"

    # Rule priority: lower number = higher priority in openpyxl CF list
    # Stale (gray fade) checked first so it overrides PENDING yellow when stale
    # Excel applies the FIRST matching rule and stops (stopIfTrue behaviour)

    # 5. Stale row: Date older than 7 days AND Status=PENDING
    stale_formula = (
        f'AND({status_ref}="PENDING",'
        f'{date_ref}<TODAY()-7,'
        f'{date_ref}<>"")'
    )
    ws.conditional_formatting.add(
        row_range,
        FormulaRule(formula=[stale_formula], fill=FILL_STALE, stopIfTrue=True),
    )

    # 1. WIN → green
    ws.conditional_formatting.add(
        row_range,
        FormulaRule(formula=[f'{status_ref}="WIN"'], fill=FILL_WIN, stopIfTrue=True),
    )

    # 2. LOST → red
    ws.conditional_formatting.add(
        row_range,
        FormulaRule(formula=[f'{status_ref}="LOST"'], fill=FILL_LOST, stopIfTrue=True),
    )

    # 3. EXPIRED → gray
    ws.conditional_formatting.add(
        row_range,
        FormulaRule(formula=[f'{status_ref}="EXPIRED"'], fill=FILL_EXPIRED, stopIfTrue=True),
    )

    # 4. PENDING → yellow
    ws.conditional_formatting.add(
        row_range,
        FormulaRule(formula=[f'{status_ref}="PENDING"'], fill=FILL_PENDING, stopIfTrue=True),
    )

    _log(f"  Conditional formatting applied to range {row_range}")


# ---------------------------------------------------------------------------
# Fix D — AutoFilter + Freeze pane
# ---------------------------------------------------------------------------

def apply_autofilter_and_freeze(ws) -> None:
    """Set AutoFilter on header row 4 and freeze rows 1-4."""
    _log("  Setting AutoFilter on A4:AQ1000 and freeze at A5...")
    ws.auto_filter.ref = f"A{HEADER_ROW}:{LAST_COL_LETTER}{DATA_END_ROW}"
    ws.freeze_panes = f"A{DATA_START_ROW}"  # freeze rows 1-4, no col freeze
    _log("  AutoFilter + Freeze: SET")


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def main(erp_file: str, dry_run: bool = False) -> None:
    if not os.path.exists(erp_file):
        _log(f"ERROR: ERP file not found: {erp_file}")
        sys.exit(2)

    # --- Pre-flight: file lock check ---
    _check_excel_closed(erp_file)
    _log(f"Target: {erp_file}")

    if dry_run:
        _log("DRY RUN mode — workbook will be loaded and inspected but NOT saved.")

    # --- Backup ---
    if not dry_run:
        _backup(erp_file)

    # --- Load workbook (keep_vba=True to preserve macros) ---
    _log("Loading workbook (keep_vba=True)...")
    wb = openpyxl.load_workbook(erp_file, keep_vba=True)

    if QUOTES_SHEET not in wb.sheetnames:
        _log(f"ERROR: Sheet '{QUOTES_SHEET}' not found. Sheets: {wb.sheetnames}")
        sys.exit(3)

    ws = wb[QUOTES_SHEET]
    _log(f"Sheet '{QUOTES_SHEET}' loaded. Current max_row={ws.max_row}")

    # --- Apply fixes ---
    _log("--- Fix C: KPI rows ---")
    kpi_status = apply_kpi_rows(ws)

    _log("--- Fix A: Conditional formatting ---")
    apply_conditional_formatting(ws)
    cf_status = "APPLIED"

    _log("--- Fix D: AutoFilter + Freeze ---")
    apply_autofilter_and_freeze(ws)
    ad_status = "SET"

    # --- Save or dry-run report ---
    if dry_run:
        _log("DRY RUN: skipping save.")
    else:
        _log("Saving workbook via save_preserving_ribbon...")
        # Import at runtime to allow testing without OneDrive mounted
        try:
            # Insert repo root into sys.path so ERP.core is importable
            _repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            if _repo_root not in sys.path:
                sys.path.insert(0, _repo_root)
            from ERP.core.ribbon_guard import save_preserving_ribbon  # noqa: PLC0415
        except ImportError as exc:
            _log(f"ERROR: Cannot import ribbon_guard: {exc}")
            _log("Ensure script is run from repo root: python scripts/erp-quote-polish.py")
            sys.exit(4)

        result = save_preserving_ribbon(wb, erp_file)
        _log(f"Save result: {result}")

    # --- Summary diff ---
    print()
    print("=" * 50)
    print("  erp-quote-polish SUMMARY")
    print("=" * 50)
    print(f"  KPI rows          : {kpi_status}")
    print(f"  Conditional format: {cf_status}")
    print(f"  AutoFilter+Freeze : {ad_status}")
    if dry_run:
        print("  [DRY RUN — no file written]")
    print("=" * 50)
    _log("Done. Exit 0.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="ERP Quote sheet polish: KPI rows + conditional format + autofilter",
    )
    parser.add_argument(
        "--erp-file",
        default=DEFAULT_ERP_FILE,
        help="Path to ERP_Master_v14.xlsm (default: OneDrive canonical path)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Load and inspect workbook but do NOT write/save",
    )
    args = parser.parse_args()
    main(erp_file=args.erp_file, dry_run=args.dry_run)
