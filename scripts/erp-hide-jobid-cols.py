"""
erp-hide-jobid-cols.py — One-shot: hide Job_ID columns in ERP_Master_v14.xlsm
==============================================================================
Nelson decision (2026-04-21): Job_ID is no longer the primary key.
  - Active Jobs col C (Job_ID, col 3)  → hidden
  - Archive     col A (Job_ID, col 1)  → hidden

Primary key going forward:
  - Active Jobs: Bkg_No (col H / col 8)
  - Archive:     Bkg_No (col F / col 6)

Existing values are PRESERVED for history/rollback — only visibility changes.

Usage:
    python scripts/erp-hide-jobid-cols.py [--target PATH] [--dry-run]

Gotcha #6: NEVER call wb.save() — always save_preserving_ribbon().
Idempotent: re-running sets hidden=True again (no-op if already hidden).
"""
from __future__ import annotations

import argparse
import os
import shutil
import sys
from datetime import datetime

import openpyxl

_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from ERP.core.ribbon_guard import save_preserving_ribbon  # noqa: E402

DEFAULT_TARGET = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"

ACTIVE_JOBS_SHEET = "Active Jobs"
ARCHIVE_SHEET = "Archive"

# Col letters to hide
ACTIVE_HIDE_COL = "C"   # Job_ID (col 3)
ARCHIVE_HIDE_COL = "A"  # Job_ID (col 1)


def _log(msg: str) -> None:
    safe = msg.encode(sys.stdout.encoding or "utf-8", errors="replace").decode(
        sys.stdout.encoding or "utf-8", errors="replace"
    )
    print(safe, flush=True)


def _check_excel_closed(path: str) -> None:
    try:
        with open(path, "a"):
            pass
    except PermissionError:
        _log(f"[ERROR] File locked — close Excel first: {path}")
        sys.exit(1)


def _backup(path: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(path)
    backup_path = f"{base}.backup_{ts}{ext}"
    shutil.copy2(path, backup_path)
    _log(f"[backup] {backup_path}")
    return backup_path


def hide_jobid_cols(target_path: str, dry_run: bool) -> None:
    _log("=" * 60)
    _log(f"[erp-hide-jobid-cols] {'DRY RUN — ' if dry_run else ''}Starting")
    _log(f"  Target: {target_path}")
    _log("=" * 60)

    if not os.path.exists(target_path):
        _log(f"[ERROR] Target not found: {target_path}")
        sys.exit(1)

    if dry_run:
        _log("[DRY RUN] Would hide:")
        _log(f"  Active Jobs sheet — col {ACTIVE_HIDE_COL} (Job_ID)")
        _log(f"  Archive sheet     — col {ARCHIVE_HIDE_COL} (Job_ID)")
        _log("[DRY RUN] No changes written.")
        return

    _check_excel_closed(target_path)
    _backup(target_path)

    _log("\n[Step 1] Opening workbook ...")
    wb = openpyxl.load_workbook(target_path, keep_vba=True)

    # Active Jobs — hide col C
    if ACTIVE_JOBS_SHEET not in wb.sheetnames:
        _log(f"[WARN] Sheet not found: {ACTIVE_JOBS_SHEET} — skipping")
    else:
        ws_aj = wb[ACTIVE_JOBS_SHEET]
        ws_aj.column_dimensions[ACTIVE_HIDE_COL].hidden = True
        _log(f"  [OK] Active Jobs col {ACTIVE_HIDE_COL} (Job_ID) → hidden")

    # Archive — hide col A
    if ARCHIVE_SHEET not in wb.sheetnames:
        _log(f"[WARN] Sheet not found: {ARCHIVE_SHEET} — skipping")
    else:
        ws_arc = wb[ARCHIVE_SHEET]
        ws_arc.column_dimensions[ARCHIVE_HIDE_COL].hidden = True
        _log(f"  [OK] Archive col {ARCHIVE_HIDE_COL} (Job_ID) → hidden")

    # Save via ribbon guard (gotcha #6)
    _log("\n[Step 2] Saving (preserving ribbon) ...")
    result = save_preserving_ribbon(wb, target_path)
    _log(f"  Ribbon guard result: {result}")

    _log("\n[DONE] Job_ID columns hidden. Values preserved for history/rollback.")
    _log("  Primary key going forward: Bkg_No")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Hide Job_ID columns in ERP_Master_v14.xlsm (Active Jobs col C + Archive col A)"
    )
    parser.add_argument(
        "--target", default=DEFAULT_TARGET,
        help=f"Path to ERP_Master_v14.xlsm (default: {DEFAULT_TARGET})",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Preview — do not write to file",
    )
    args = parser.parse_args()
    hide_jobid_cols(args.target, args.dry_run)


if __name__ == "__main__":
    main()
