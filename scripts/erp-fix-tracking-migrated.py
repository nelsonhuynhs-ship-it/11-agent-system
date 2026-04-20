"""
erp-fix-tracking-migrated.py — Backfill TRACKING + TRACKING_STAGE for migrated rows
======================================================================================
One-shot fix for rows that were imported via erp-import-shipments.py before
derive_tracking_stage() was wired in.  Those rows have:
  - col 15 (TRACKING)       empty
  - col 36 (TRACKING_STAGE) empty
  - col 8  (Bkg_No)         populated  ← "migrated row" signal

Algorithm:
  1. Scan Active Jobs sheet from DATA_START (row 8)
  2. For each row where TRACKING is empty AND Bkg_No is present:
     a. Read ATA (col 22), ETD_Original/Delay_Log (no direct col), ETD (col 13),
        Status (col 14), Bkg_No (col 8) from the sheet
     b. Infer stage via the same derive_tracking_stage() logic from erp-import-shipments.py
     c. Write TRACKING (col 15) and TRACKING_STAGE (col 36)
  3. Save via save_preserving_ribbon (gotcha #6)

Usage:
    python scripts/erp-fix-tracking-migrated.py [--dry-run] [--target PATH]

Constraints:
  - openpyxl.load_workbook(keep_vba=True)
  - save via ERP.core.ribbon_guard.save_preserving_ribbon (NEVER wb.save)
  - Backup before write
  - Excel must be closed before run
  - Idempotent: skips rows that already have TRACKING filled
"""
from __future__ import annotations

import argparse
import os
import shutil
import sys
from datetime import datetime
from typing import Any, Optional

import openpyxl
from openpyxl.comments import Comment

# ---------------------------------------------------------------------------
# Repo path so we can import ERP.core modules
# ---------------------------------------------------------------------------
_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

from ERP.core.active_jobs_cols import COL, DATA_START  # noqa: E402
from ERP.core.ribbon_guard import save_preserving_ribbon  # noqa: E402

# Re-use helpers from erp-import-shipments to stay DRY
# (both live in scripts/ so direct import by path injection works)
_SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
if _SCRIPTS_DIR not in sys.path:
    sys.path.insert(0, _SCRIPTS_DIR)

# Import the shared helpers we need — avoid re-implementing
import importlib.util as _ilu

_spec = _ilu.spec_from_file_location(
    "erp_import_shipments",
    os.path.join(_SCRIPTS_DIR, "erp-import-shipments.py"),
)
_imp_mod = _ilu.module_from_spec(_spec)  # type: ignore[arg-type]
_spec.loader.exec_module(_imp_mod)  # type: ignore[union-attr]

_parse_date = _imp_mod._parse_date
_parse_bkg_no = _imp_mod._parse_bkg_no
_status_map = _imp_mod._status_map
derive_tracking_stage = _imp_mod.derive_tracking_stage

# ---------------------------------------------------------------------------
# Defaults
# ---------------------------------------------------------------------------
DEFAULT_TARGET = r"D:\OneDrive\NelsonData\erp\ERP_Master_v14.xlsm"

# Unicode dots (same as erp-import-shipments.py)
_DOT_FULL = "\u25cf"   # ●
_DOT_EMPTY = "\u25cb"  # ○
_CHECK = "\u2713"      # ✓
_PENDING = "\u29f3"   # (placeholder, uses ○ for consistency)

# 7-stage tracking list — matches VBA ApplyTrackingDots in erp-v14-jobs-automation.bas
_STAGES_7 = ["BKG", "Confirmed", "SI Cut", "Gate-in", "ATD", "ETA", "Delivered"]


def _build_tooltip(stage_label: str) -> str:
    """Build cell comment text showing 7 stages with ✓/○ per status.

    stage_label maps to 'done' count:
      PENDING/0    → 0 done
      BOOKED/1     → 1 done (BKG ✓)
      CONFIRMED/2  → 2 done
      SI_CUT/3     → 3 done
      GATE_IN/4    → 4 done
      ATD/5        → 5 done
      ETA/6        → 6 done
      DELIVERED/7  → 7 done (ARRIVED also maps here for simplicity)
    """
    # Map label to done count — align with derive_tracking_stage outputs
    done_map = {
        "PENDING":   0,
        "BOOKED":    1,
        "CONFIRMED": 2,
        "SI_CUT":    3,
        "GATE_IN":   4,
        "ATD":       5,
        "ETA":       6,
        "ARRIVED":   7,   # treat ARRIVED as Delivered-equivalent
        "DELIVERED": 7,
    }
    done = done_map.get(str(stage_label).upper().strip(), 1)
    lines = []
    for i, stage in enumerate(_STAGES_7):
        marker = _CHECK if i < done else _DOT_EMPTY
        lines.append(f"{marker} {stage}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _log(msg: str) -> None:
    safe = msg.encode(sys.stdout.encoding or "utf-8", errors="replace").decode(
        sys.stdout.encoding or "utf-8", errors="replace"
    )
    print(safe, flush=True)


def _check_excel_closed(path: str) -> None:
    try:
        with open(path, "a"):
            pass
    except PermissionError:
        _log(f"[ERROR] File locked — close Excel first: {path}")
        sys.exit(1)


def _backup(path: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(path)
    backup_path = f"{base}.backup_{ts}{ext}"
    shutil.copy2(path, backup_path)
    _log(f"[backup] {backup_path}")
    return backup_path


def _cell_val(ws, row: int, col: int) -> Any:
    return ws.cell(row=row, column=col).value


# ---------------------------------------------------------------------------
# Core fix logic
# ---------------------------------------------------------------------------

def fix_tracking(ws, dry_run: bool) -> tuple[int, int]:
    """
    Scan Active Jobs ws and backfill TRACKING + TRACKING_STAGE for migrated rows.

    A row is eligible if:
      - Bkg_No (col 8) is present  (migrated row signal)
      - TRACKING (col 15) is empty

    Returns (fixed, skipped_already_filled).
    """
    fixed = 0
    skipped = 0

    max_row = ws.max_row or DATA_START
    for row in range(DATA_START, max_row + 1):
        bkg_no_raw = _cell_val(ws, row, COL["Bkg_No"])
        bkg_no = _parse_bkg_no(bkg_no_raw)

        # Skip rows without Bkg_No (not migrated rows we care about)
        if not bkg_no:
            continue

        tracking_current = _cell_val(ws, row, COL["TRACKING"])
        cell_existing = ws.cell(row=row, column=COL["TRACKING"])
        has_dots = tracking_current is not None and str(tracking_current).strip()
        has_comment = cell_existing.comment is not None

        # Idempotent: skip only if BOTH dots + comment present.
        # If dots exist but comment missing (migrated pre-comment upgrade), refresh.
        if has_dots and has_comment:
            skipped += 1
            continue

        # Read signals from sheet
        ata_raw = _cell_val(ws, row, COL["ATA"])
        etd_raw = _cell_val(ws, row, COL["ETD"])
        status_raw = _cell_val(ws, row, COL["Status"])
        # ETD_Original not a dedicated column in Active Jobs layout —
        # use Delay_Log (col 31) as proxy: if it has "Re-sched" the ETD was changed
        delay_log_raw = _cell_val(ws, row, COL["Delay_Log"])

        # Reconstruct etd_original proxy: if Delay_Log mentions a reschedule,
        # pass a sentinel (non-None) so derive_tracking_stage sees a change.
        # Exact date not needed — presence is enough for the ATD branch.
        etd_original_proxy: Optional[datetime] = None
        if delay_log_raw and "Re-sched" in str(delay_log_raw):
            # Any date != ETD will trigger the rescheduled branch; use a past date
            etd_original_proxy = datetime(2000, 1, 1)

        stage_label, dots = derive_tracking_stage(
            ata=ata_raw,
            etd_original=etd_original_proxy,
            etd=etd_raw,
            status=status_raw,
            bkg_no=bkg_no,
        )

        crm = _cell_val(ws, row, COL["CRM_ID"]) or ""
        etd_str = etd_raw.strftime("%Y-%m-%d") if isinstance(etd_raw, datetime) else str(etd_raw or "")
        _log(f"  [FIX] row={row} Bkg={bkg_no} {crm} ETD={etd_str} -> {stage_label} | {dots}")

        if not dry_run:
            cell = ws.cell(row=row, column=COL["TRACKING"])
            cell.value = dots
            # Attach tooltip comment with 7-stage checklist (✓ done · ○ pending)
            tooltip = _build_tooltip(stage_label)
            cell.comment = Comment(tooltip, "ERP", height=140, width=200)
            ws.cell(row=row, column=COL["TRACKING_STAGE"]).value = stage_label

        fixed += 1

    return fixed, skipped


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Backfill TRACKING + TRACKING_STAGE for migrated Active Jobs rows"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Preview changes without writing to ERP xlsm",
    )
    parser.add_argument(
        "--target", default=DEFAULT_TARGET,
        help=f"Path to ERP_Master_v14.xlsm (default: {DEFAULT_TARGET})",
    )
    args = parser.parse_args()

    target_path: str = args.target
    dry_run: bool = args.dry_run

    _log("=" * 60)
    _log(f"[erp-fix-tracking-migrated] {'DRY RUN — ' if dry_run else ''}Starting")
    _log(f"  Target : {target_path}")
    _log("=" * 60)

    if not os.path.exists(target_path):
        _log(f"[ERROR] Target not found: {target_path}")
        sys.exit(1)

    if not dry_run:
        _check_excel_closed(target_path)
        _backup(target_path)

    _log("\n[Step 1] Opening ERP workbook ...")
    wb = openpyxl.load_workbook(target_path, keep_vba=True)
    ws_aj = wb["Active Jobs"]

    _log("\n[Step 2] Scanning Active Jobs for empty TRACKING rows ...")
    fixed, skipped = fix_tracking(ws_aj, dry_run=dry_run)

    if dry_run:
        _log(f"\n[DRY RUN] Would fix: {fixed} rows  |  Already filled: {skipped} rows")
        _log("[DRY RUN] No changes written.")
        sys.exit(0)

    _log(f"\n  Fixed: {fixed}  |  Already filled (skipped): {skipped}")

    if fixed == 0:
        _log("\n[INFO] Nothing to fix — all rows already have TRACKING. Exiting.")
        sys.exit(0)

    _log("\n[Step 3] Saving (preserving ribbon — gotcha #6) ...")
    result = save_preserving_ribbon(wb, target_path)
    _log(f"  Ribbon guard result: {result}")

    _log("\n" + "=" * 60)
    _log(f"[DONE] Backfilled TRACKING for {fixed} migrated row(s).")
    _log("=" * 60)


if __name__ == "__main__":
    main()
