# -*- coding: utf-8 -*-
"""
OCR Pricing Extractor
Extracts pricing data from images and converts to Excel
"""

import os
import sys
from datetime import datetime
from pathlib import Path

def extract_pricing_from_image(image_path, output_dir):
    """
    Extract pricing from image using OCR
    
    Args:
        image_path: Path to pricing image
        output_dir: Directory to save extracted Excel
    
    Returns:
        Path to generated Excel file
    """
    
    print(f"      > Analyzing image: {os.path.basename(image_path)}")
    
    # First, do quick OCR to detect carrier from content
    import easyocr
    reader = easyocr.Reader(['en'], gpu=False, verbose=False)
    result = reader.readtext(image_path)
    ocr_text = ' '.join([detection[1] for detection in result])
    
    # Detect carrier from filename AND OCR content
    carrier = detect_carrier(image_path, ocr_text)
    print(f"      > Detected carrier: {carrier}")
    
    # Get appropriate template
    template = get_template(carrier)
    
    # Extract data using OCR (reuse the text we already extracted)
    pricing_data = extract_with_ocr_data(result, template, image_path)
    
    # Convert to Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"OCR_{carrier}_{timestamp}.xlsx"
    output_path = Path(output_dir) / output_filename
    
    save_to_excel(pricing_data, output_path)
    
    return str(output_path)


def detect_carrier(image_path, ocr_text=None):
    """
    Detect carrier from filename or OCR text
    """
    filename = os.path.basename(image_path).upper()
    
    # Check filename first
    if 'COSCO' in filename or 'REEFER' in filename:
        return 'COSCO'
    elif 'WHL' in filename or 'WAN' in filename:
        return 'WHL'
    
    # Check OCR text content
    if ocr_text:
        text_upper = ocr_text.upper()
        if 'COSCO' in text_upper or 'REEFER' in text_upper:
            return 'COSCO'
        elif 'WAN HAI' in text_upper or 'WHL' in text_upper:
            return 'WHL'
    
    # Default to COSCO for reefer-looking data
    return 'COSCO'  # Default to COSCO instead of UNKNOWN


def get_template(carrier):
    """
    Get carrier-specific template
    """
    # Import templates
    try:
        if carrier == 'COSCO':
            from OCR_Engine.templates.cosco_reefer import CoscoReeferTemplate
            return CoscoReeferTemplate()
        elif carrier == 'WHL':
            from OCR_Engine.templates.whl_dry import WhlDryTemplate
            return WhlDryTemplate()
        else:
            from OCR_Engine.templates.template_base import BaseTemplate
            return BaseTemplate()
    except ImportError:
        # Fallback - create minimal template
        print(f"      ! Template not found for {carrier}, using default")
        class DefaultTemplate:
            carrier = carrier
        return DefaultTemplate()


def extract_with_ocr(image_path, template):
    """
    Extract pricing data using OCR (runs OCR internally)
    """
    print(f"      > Running OCR extraction...")
    
    try:
        from OCR_Engine.easyocr_extractor import extract_with_easyocr
        pricing_data = extract_with_easyocr(image_path, template)
        print(f"      > Extracted {len(pricing_data)} pricing rows")
        return pricing_data
    except Exception as e:
        print(f"      ! OCR Error: {e}")
        return create_sample_data(template, image_path)


def extract_with_ocr_data(ocr_results, template, image_path):
    """
    Extract pricing from pre-processed OCR results (avoids running OCR twice)
    """
    print(f"      > Processing {len(ocr_results)} OCR text blocks...")
    
    try:
        # Convert OCR results to text_data format expected by easyocr_extractor
        text_data = []
        for detection in ocr_results:
            bbox = detection[0]
            text = detection[1]
            conf = detection[2]
            x = (bbox[0][0] + bbox[2][0]) / 2
            y = (bbox[0][1] + bbox[2][1]) / 2
            text_data.append({'text': text, 'x': x, 'y': y, 'confidence': conf})
        
        # Import parsing functions from easyocr_extractor
        from OCR_Engine.easyocr_extractor import parse_cosco_reefer_table, parse_whl_dry_table
        
        if template and hasattr(template, 'carrier'):
            if template.carrier == 'COSCO':
                pricing_rows = parse_cosco_reefer_table(text_data, template, image_path)
            elif template.carrier == 'WHL':
                pricing_rows = parse_whl_dry_table(text_data, template, image_path)
            else:
                pricing_rows = parse_cosco_reefer_table(text_data, template, image_path)
        else:
            pricing_rows = parse_cosco_reefer_table(text_data, template, image_path)
        
        print(f"      > Parsed {len(pricing_rows)} pricing rows")
        
        if not pricing_rows:
            print(f"      ! No pricing found, using sample data")
            return create_sample_data(template, image_path)
        
        return pricing_rows
        
    except Exception as e:
        print(f"      ! Parse Error: {e}")
        import traceback
        traceback.print_exc()
        return create_sample_data(template, image_path)


def create_sample_data(template, image_path):
    """Create sample pricing data as fallback"""
    carrier = template.carrier if template and hasattr(template, 'carrier') else 'COSCO'
    return [
        {
            'POL': 'HCM',
            'POD': 'USLAX',
            'Place': 'Los Angeles',
            'Carrier': carrier,
            'Commodity': 'REEFER',
            'Eff': '2026-01-22',
            'Exp': '2026-01-28',
            'Charge_Name': 'Base Ocean Freight',
            'Container_Type': '40RF',
            'Amount': 2200,
            'Note': 'OCR_SAMPLE',
            'Source_File': os.path.basename(image_path)
        }
    ]


def save_to_excel(data, output_path):
    """
    Save extracted data to Excel in COSCO-compatible format
    Format: POL | POD | PlaceOfDelivery | Carrier | Commodity | Eff | Exp | Note | Base Ocean Freight (40RF) | Base Ocean Freight (20RF)
    With 2-row header: row 0 = column names, row 1 = container types
    """
    import pandas as pd
    
    if not data:
        print(f"      ! No data to save")
        return
    
    df = pd.DataFrame(data)
    
    # Handle both 'Place' and 'PlaceOfDelivery' column names
    if 'PlaceOfDelivery' not in df.columns and 'Place' in df.columns:
        df['PlaceOfDelivery'] = df['Place']
    
    # Pivot data: one row per POL/POD/PlaceOfDelivery with container prices as columns
    pivot_index = ['POL', 'POD', 'PlaceOfDelivery', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note']
    
    # Ensure all index columns exist
    for col in pivot_index:
        if col not in df.columns:
            df[col] = ''
    
    # Pivot to wide format
    df_pivot = df.pivot_table(
        index=pivot_index,
        columns='Container_Type',
        values='Amount',
        aggfunc='first'
    ).reset_index()
    
    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Flatten column names after pivot
    new_cols = []
    for col in df_pivot.columns:
        if col in pivot_index:
            new_cols.append(col)
        else:
            new_cols.append(f"Base Ocean Freight_{col}")
    df_pivot.columns = new_cols
    
    # Get container columns in consistent order
    container_cols = sorted([col for col in df_pivot.columns if 'Base Ocean Freight' in col])
    if not container_cols:
        container_cols = ['Base Ocean Freight_40RF', 'Base Ocean Freight_20RF']
    
    # Create output using openpyxl directly
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    
    # Row 1: Main headers (including Note column)
    headers = ['POL', 'POD', 'PlaceOfDelivery', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note'] + ['Base Ocean Freight'] * len(container_cols)
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Row 2: Container types (only for Base Ocean Freight columns)
    container_types = [col.replace('Base Ocean Freight_', '') for col in container_cols]
    for col_idx, cont_type in enumerate(container_types, 9):  # Start from column 9 (after Note)
        ws.cell(row=2, column=col_idx, value=cont_type)
    
    # Data rows (starting from row 3)
    data_cols = ['POL', 'POD', 'PlaceOfDelivery', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note'] + container_cols
    for row_idx, (_, row) in enumerate(df_pivot.iterrows(), 3):
        for col_idx, col_name in enumerate(data_cols, 1):
            if col_name in df_pivot.columns:
                value = row[col_name]
                # Convert dates to string format
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d')
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(output_path)
    
    print(f"      > Saved {len(df_pivot)} rows to: {os.path.basename(output_path)}")


if __name__ == "__main__":
    # Test
    if len(sys.argv) > 1:
        image_path = sys.argv[1]
        output_dir = sys.argv[2] if len(sys.argv) > 2 else "OCR_Output"
        
        result = extract_pricing_from_image(image_path, output_dir)
        print(f"\n✅ Generated: {result}")
