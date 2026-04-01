# -*- coding: utf-8 -*-
"""
rate_monitor.py — NELSON Market Intelligence (NMI) Module 1
============================================================
Rate Anomaly Detector: compares current market rates (Parquet) against
open quotes/jobs in ERP, and alerts Nelson via Telegram when significant
rate changes are detected.

Part of the NMI system — Sprint 12.

Usage:
    # Full check with Telegram alerts:
    python rate_monitor.py

    # Dry-run (print alerts, no Telegram):
    python rate_monitor.py --dry-run

    # Show current market rates for a route:
    python rate_monitor.py --route HCM-LAX

    # Custom threshold:
    python rate_monitor.py --threshold 100
"""
import json
import logging
import os
import sys
import argparse
from datetime import datetime, timedelta
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional

import pandas as pd

# ── Logging ────────────────────────────────────────────────────────────────────
log = logging.getLogger("nelson.nmi.rate_monitor")
if not log.handlers:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)-5s | %(name)s | %(message)s",
        datefmt="%H:%M:%S",
    )

# ── Paths ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
CONFIG_FILE = SCRIPT_DIR / "config" / "nmi_config.json"
DATA_DIR = SCRIPT_DIR / "data"
PARQUET_FILE = DATA_DIR / "Cleaned_Master_History.parquet"
ERP_FILE = SCRIPT_DIR.parent / "ERP" / "data" / "ERP_Master.xlsm"


# ==============================================================================
# DATA CLASSES
# ==============================================================================

@dataclass
class NMIConfig:
    """Configurable thresholds and settings."""
    drop_alert_usd: float = 150.0
    drop_alert_pct: float = 3.0
    spike_alert_usd: float = 200.0
    spike_alert_pct: float = 5.0
    stale_quote_days: int = 3
    auto_expire_days: int = 7
    min_days_before_etd: int = 7
    charge_filter: str = "Total Ocean Freight"
    container_priority: list = field(default_factory=lambda: ["20GP", "40GP", "40HQ"])

    @classmethod
    def from_file(cls, path: Path = CONFIG_FILE) -> "NMIConfig":
        """Load config from JSON file."""
        if not path.exists():
            log.warning("Config file not found: %s — using defaults", path)
            return cls()
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        thresholds = data.get("thresholds", {})
        return cls(
            drop_alert_usd=thresholds.get("drop_alert_usd", 150),
            drop_alert_pct=thresholds.get("drop_alert_pct", 3.0),
            spike_alert_usd=thresholds.get("spike_alert_usd", 200),
            spike_alert_pct=thresholds.get("spike_alert_pct", 5.0),
            stale_quote_days=thresholds.get("stale_quote_days", 3),
            auto_expire_days=thresholds.get("auto_expire_days", 7),
            min_days_before_etd=thresholds.get("min_days_before_etd", 7),
            charge_filter=data.get("charge_filter", "Total Ocean Freight"),
            container_priority=data.get("container_priority", ["20GP", "40GP", "40HQ"]),
        )


@dataclass
class RateAlert:
    """A single rate anomaly alert."""
    alert_type: str          # "DROP" or "SPIKE"
    customer: str
    routing: str             # e.g. "HCM → LOS ANGELES, CA"
    carrier: str
    container: str
    quoted_price: float
    market_price: float
    delta: float
    delta_pct: float
    job_id: str
    quote_id: str
    etd: Optional[datetime] = None
    days_to_etd: Optional[int] = None
    status: str = ""


@dataclass
class FreshnessAlert:
    """A quote freshness warning."""
    job_id: str
    customer: str
    routing: str
    carrier: str
    quote_age_days: int
    quoted_price: float
    market_price: float
    rate_changed: float      # absolute delta since quote
    etd: Optional[datetime] = None


# ==============================================================================
# RATE MONITOR CORE
# ==============================================================================

class RateMonitor:
    """
    Compares market rates from Parquet against open quotes/jobs in ERP.
    Generates alerts when rates move beyond configurable thresholds.
    """

    def __init__(self, config: NMIConfig = None,
                 parquet_path: Path = PARQUET_FILE,
                 erp_path: Path = ERP_FILE):
        self.config = config or NMIConfig.from_file()
        self.parquet_path = parquet_path
        self.erp_path = erp_path
        self._market_rates: Optional[pd.DataFrame] = None
        self._open_quotes: Optional[pd.DataFrame] = None

    # ── Load Market Rates ──────────────────────────────────────────────────

    def load_latest_rates(self) -> pd.DataFrame:
        """
        Load the latest market rates from Parquet.
        Filters to 'Total Ocean Freight' charge and gets the most recent
        rate per lane (POL/POD/Carrier/Container_Type).
        """
        log.info("Loading market rates from Parquet...")
        if not self.parquet_path.exists():
            log.error("Parquet file not found: %s", self.parquet_path)
            return pd.DataFrame()

        df = pd.read_parquet(self.parquet_path)

        # Filter to Total Ocean Freight only
        charge = self.config.charge_filter
        df = df[df["Charge_Name"].str.contains(charge, case=False, na=False)]

        # Filter to standard containers
        df = df[df["Container_Type"].isin(self.config.container_priority)]

        # Parse dates
        df["Eff"] = pd.to_datetime(df["Eff"], errors="coerce")
        df["Exp"] = pd.to_datetime(df["Exp"], errors="coerce")

        # Get the most recent rate per lane
        # Sort by Eff descending, keep first (most recent)
        df = df.sort_values("Eff", ascending=False)
        df = df.drop_duplicates(
            subset=["POL", "POD", "Carrier", "Container_Type"],
            keep="first",
        )

        # Clean POL/POD
        df["POL"] = df["POL"].str.strip().str.upper()
        df["POD"] = df["POD"].str.strip().str.upper()
        df["Carrier"] = df["Carrier"].str.strip().str.upper()
        df["Place"] = df["Place"].str.strip().str.upper()

        log.info(
            "  Loaded %d unique rates | %d carriers | Eff range: %s to %s",
            len(df),
            df["Carrier"].nunique(),
            df["Eff"].min().strftime("%Y-%m-%d") if not df.empty else "N/A",
            df["Eff"].max().strftime("%Y-%m-%d") if not df.empty else "N/A",
        )

        self._market_rates = df
        return df

    # ── Load Open Quotes from ERP ──────────────────────────────────────────

    def load_open_quotes(self) -> pd.DataFrame:
        """
        Read Active Jobs from ERP_Master.xlsm.
        Filters to open/quoted status jobs only.
        Returns DataFrame with: job_id, customer, routing, carrier,
                                container, selling_rate, etd, status
        """
        log.info("Loading open quotes from ERP...")

        if not self.erp_path.exists():
            log.warning("ERP file not found: %s", self.erp_path)
            return pd.DataFrame()

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(self.erp_path), keep_vba=True, read_only=True)

            # Find Active Jobs sheet
            target = None
            for s in wb.sheetnames:
                if "active" in s.lower():
                    target = wb[s]
                    break

            if target is None:
                log.warning("No 'Active Jobs' sheet found in ERP")
                wb.close()
                return pd.DataFrame()

            rows = list(target.iter_rows(values_only=True))
            wb.close()

            if len(rows) < 8:
                log.warning("Active Jobs sheet has fewer than 8 rows")
                return pd.DataFrame()

            # Row 7 (index 6) = headers, row 8+ = data
            headers = [str(h).strip() if h else f"Col{i}" for i, h in enumerate(rows[6], 1)]
            data_rows = rows[7:]  # row 8 onward
            df = pd.DataFrame(data_rows, columns=headers)
            df = df.dropna(how="all")

            # Filter to valid jobs
            if "Job_ID" in df.columns:
                # Keep rows with proper Job_ID format OR Quote_ID (some may be pre-job quotes)
                has_job = df["Job_ID"].astype(str).str.match(r"J\d{4,}", na=False)
                has_quote = (
                    df["Quote_ID"].astype(str).str.match(r"SE\d{4}", na=False)
                    if "Quote_ID" in df.columns
                    else pd.Series(False, index=df.index)
                )
                df = df[has_job | has_quote]

            if df.empty:
                log.info("  No active jobs/quotes found in ERP")
                return pd.DataFrame()

            # Filter to open/active status only
            if "Status" in df.columns:
                open_statuses = ["open", "quoted", "pending", "booked", "in_transit", "active"]
                status_mask = df["Status"].astype(str).str.strip().str.lower().isin(open_statuses)
                # Also include blank status (could be pending)
                blank_status = df["Status"].isna() | (df["Status"].astype(str).str.strip() == "")
                df = df[status_mask | blank_status]

            # Parse key columns
            if "Selling_Rate" in df.columns:
                df["Selling_Rate"] = pd.to_numeric(df["Selling_Rate"], errors="coerce")
            if "ETD" in df.columns:
                df["ETD"] = pd.to_datetime(df["ETD"], errors="coerce")

            # Extract POL/POD from Routing (format: "HCM → LOS ANGELES, CA" or "HCM-LAX")
            if "Routing" in df.columns:
                routing = df["Routing"].astype(str).str.strip()
                # Try → separator first, then -
                split = routing.str.split(r"\s*[→\-]\s*", n=1, expand=True)
                if split.shape[1] >= 2:
                    df["_pol"] = split[0].str.strip().str.upper()
                    df["_pod"] = split[1].str.strip().str.upper()
                else:
                    df["_pol"] = ""
                    df["_pod"] = ""
            else:
                df["_pol"] = ""
                df["_pod"] = ""

            if "Carrier" in df.columns:
                df["Carrier"] = df["Carrier"].astype(str).str.strip().str.upper()
            if "Container_Type" in df.columns:
                df["Container_Type"] = df["Container_Type"].astype(str).str.strip().str.upper()

            log.info("  Loaded %d open jobs/quotes from ERP", len(df))
            self._open_quotes = df
            return df

        except Exception as e:
            log.error("Failed to read ERP file: %s", e)
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    # ── Detect Anomalies ───────────────────────────────────────────────────

    def detect_anomalies(self) -> list[RateAlert]:
        """
        Compare each open quote against current market rates.
        Returns list of RateAlert objects for quotes with significant rate changes.
        """
        if self._market_rates is None:
            self.load_latest_rates()
        if self._open_quotes is None:
            self.load_open_quotes()

        rates = self._market_rates
        quotes = self._open_quotes

        if rates is None or rates.empty:
            log.warning("No market rates available — skipping anomaly detection")
            return []
        if quotes is None or quotes.empty:
            log.info("No open quotes — nothing to compare")
            return []

        alerts = []
        now = datetime.now()
        cfg = self.config

        for _, q in quotes.iterrows():
            try:
                quoted_price = q.get("Selling_Rate")
                if pd.isna(quoted_price) or quoted_price <= 0:
                    continue

                carrier = str(q.get("Carrier", "")).upper()
                container = str(q.get("Container_Type", "")).upper()
                pol = str(q.get("_pol", "")).upper()
                pod = str(q.get("_pod", "")).upper()
                customer = str(q.get("Customer_Name", ""))
                routing = str(q.get("Routing", ""))
                job_id = str(q.get("Job_ID", ""))
                quote_id = str(q.get("Quote_ID", ""))
                etd = q.get("ETD")

                if not carrier or not container or not pol:
                    continue

                # Check days to ETD — skip if too close to departure
                days_to_etd = None
                if pd.notna(etd):
                    days_to_etd = (etd - now).days
                    if days_to_etd < cfg.min_days_before_etd:
                        continue  # Too close to ETD, can't reprice

                # Find matching market rate
                # Match on POL + Carrier + Container_Type
                # For POD, try exact match first, then contains (fuzzy)
                mask = (
                    (rates["POL"] == pol)
                    & (rates["Carrier"] == carrier)
                    & (rates["Container_Type"] == container)
                )

                # POD matching: try exact, then contains, then Place
                if pod:
                    pod_exact = mask & (rates["POD"] == pod)
                    if pod_exact.any():
                        mask = pod_exact
                    else:
                        # Fuzzy: POD contains the first word of our POD
                        pod_first = pod.split(",")[0].split()[0] if pod else ""
                        if pod_first:
                            pod_fuzzy = mask & (
                                rates["POD"].str.contains(pod_first, case=False, na=False)
                                | rates["Place"].str.contains(pod_first, case=False, na=False)
                            )
                            if pod_fuzzy.any():
                                mask = pod_fuzzy

                matched = rates[mask]
                if matched.empty:
                    continue

                # Take the most recent rate
                market_price = matched.iloc[0]["Amount"]
                if pd.isna(market_price) or market_price <= 0:
                    continue

                # Calculate delta
                delta = quoted_price - market_price
                delta_pct = (delta / market_price) * 100 if market_price > 0 else 0

                # Check thresholds
                alert_type = None
                if delta > 0 and market_price < quoted_price:
                    # Market dropped below our quote
                    if abs(delta) >= cfg.drop_alert_usd or abs(delta_pct) >= cfg.drop_alert_pct:
                        alert_type = "DROP"
                elif delta < 0 and market_price > quoted_price:
                    # Market spiked above our quote
                    if abs(delta) >= cfg.spike_alert_usd or abs(delta_pct) >= cfg.spike_alert_pct:
                        alert_type = "SPIKE"

                if alert_type:
                    alerts.append(RateAlert(
                        alert_type=alert_type,
                        customer=customer,
                        routing=routing,
                        carrier=carrier,
                        container=container,
                        quoted_price=float(quoted_price),
                        market_price=float(market_price),
                        delta=float(delta),
                        delta_pct=round(float(delta_pct), 1),
                        job_id=job_id,
                        quote_id=quote_id,
                        etd=etd if pd.notna(etd) else None,
                        days_to_etd=days_to_etd,
                        status=str(q.get("Status", "")),
                    ))

            except Exception as e:
                log.debug("Error processing quote row: %s", e)
                continue

        log.info("Detected %d anomalies (%d DROP, %d SPIKE)",
                 len(alerts),
                 sum(1 for a in alerts if a.alert_type == "DROP"),
                 sum(1 for a in alerts if a.alert_type == "SPIKE"))

        return alerts

    # ── Detect Stale Quotes ────────────────────────────────────────────────

    def detect_stale_quotes(self) -> list[FreshnessAlert]:
        """
        Find open quotes that are older than stale_quote_days
        and have had rate changes since quoting.
        """
        if self._open_quotes is None:
            self.load_open_quotes()
        if self._market_rates is None:
            self.load_latest_rates()

        quotes = self._open_quotes
        rates = self._market_rates

        if quotes is None or quotes.empty or rates is None or rates.empty:
            return []

        alerts = []
        now = datetime.now()
        cfg = self.config

        for _, q in quotes.iterrows():
            try:
                # Check Created_Date for age
                created = q.get("Created_Date")
                if pd.isna(created):
                    continue
                created = pd.to_datetime(created, errors="coerce")
                if pd.isna(created):
                    continue

                age_days = (now - created).days
                if age_days < cfg.stale_quote_days:
                    continue

                quoted_price = q.get("Selling_Rate")
                if pd.isna(quoted_price) or quoted_price <= 0:
                    continue

                carrier = str(q.get("Carrier", "")).upper()
                container = str(q.get("Container_Type", "")).upper()
                pol = str(q.get("_pol", "")).upper()

                # Find current market rate
                mask = (
                    (rates["POL"] == pol)
                    & (rates["Carrier"] == carrier)
                    & (rates["Container_Type"] == container)
                )
                matched = rates[mask]
                if matched.empty:
                    continue

                market_price = matched.iloc[0]["Amount"]
                rate_changed = abs(float(quoted_price) - float(market_price))

                # Only alert if rate actually changed
                if rate_changed < 50:  # minimal threshold
                    continue

                alerts.append(FreshnessAlert(
                    job_id=str(q.get("Job_ID", "")),
                    customer=str(q.get("Customer_Name", "")),
                    routing=str(q.get("Routing", "")),
                    carrier=carrier,
                    quote_age_days=age_days,
                    quoted_price=float(quoted_price),
                    market_price=float(market_price),
                    rate_changed=rate_changed,
                    etd=q.get("ETD") if pd.notna(q.get("ETD")) else None,
                ))

            except Exception:
                continue

        log.info("Found %d stale quotes", len(alerts))
        return alerts

    # ── Market Summary ─────────────────────────────────────────────────────

    def get_market_summary(self, pol: str = "HCM") -> str:
        """
        Quick market summary for a given POL.
        Shows current rates per carrier per primary POD.
        """
        if self._market_rates is None:
            self.load_latest_rates()

        rates = self._market_rates
        if rates is None or rates.empty:
            return "No market data available."

        pol = pol.strip().upper()
        pol_rates = rates[rates["POL"] == pol]

        if pol_rates.empty:
            return f"No rates found for POL={pol}"

        # Pivot: Carrier × Container_Type for top PODs
        top_pods = pol_rates["POD"].value_counts().head(10).index.tolist()
        summary_lines = [f"📊 Market Rates — {pol} (as of {datetime.now().strftime('%d-%b-%Y')})"]
        summary_lines.append("━" * 50)

        for pod in top_pods:
            pod_rates = pol_rates[pol_rates["POD"] == pod]
            carriers = pod_rates.groupby("Carrier").apply(
                lambda g: {ct: g[g["Container_Type"] == ct]["Amount"].values[0]
                           for ct in ["20GP", "40HQ"]
                           if ct in g["Container_Type"].values},
                include_groups=False,
            )
            if carriers.empty:
                continue

            summary_lines.append(f"\n🚢 {pol} → {pod}")
            for carrier, prices in carriers.items():
                p20 = prices.get("20GP", "-")
                p40 = prices.get("40HQ", "-")
                p20_str = f"${p20:,.0f}" if isinstance(p20, (int, float)) else p20
                p40_str = f"${p40:,.0f}" if isinstance(p40, (int, float)) else p40
                eff_date = pod_rates[pod_rates["Carrier"] == carrier]["Eff"].iloc[0]
                eff_str = eff_date.strftime("%d-%b") if pd.notna(eff_date) else ""
                summary_lines.append(f"  {carrier:6s} | 20GP: {p20_str:>8s} | 40HQ: {p40_str:>8s} | Eff: {eff_str}")

        return "\n".join(summary_lines)


# ==============================================================================
# TELEGRAM ALERTING
# ==============================================================================

def format_rate_alert(alert: RateAlert) -> str:
    """Format a single rate alert for Telegram."""
    emoji = "📉" if alert.alert_type == "DROP" else "📈"
    direction = "dropped" if alert.alert_type == "DROP" else "spiked"
    recommend = (
        "Reprice before customer confirms"
        if alert.alert_type == "DROP"
        else "Opportunity — consider locking in"
    )

    etd_str = alert.etd.strftime("%d-%b") if alert.etd else "N/A"
    days_str = f" ({alert.days_to_etd}d away)" if alert.days_to_etd else ""

    lines = [
        f"{emoji} RATE {alert.alert_type} ALERT",
        f"━━━━━━━━━━━━━━━━━━━━",
        f"👤 {alert.customer}",
        f"🚢 {alert.routing} | {alert.carrier} {alert.container}",
        f"💰 Quote: ${alert.quoted_price:,.0f} → Market: ${alert.market_price:,.0f}",
        f"📊 Δ ${abs(alert.delta):,.0f} ({abs(alert.delta_pct):.1f}%) {direction}",
        f"📋 Job: {alert.job_id} | Quote: {alert.quote_id}",
        f"📅 ETD: {etd_str}{days_str}",
        f"",
        f"💡 {recommend}",
    ]
    return "\n".join(lines)


def format_freshness_alert(alert: FreshnessAlert) -> str:
    """Format a freshness alert for Telegram."""
    etd_str = alert.etd.strftime("%d-%b") if alert.etd else "N/A"
    return (
        f"⏰ STALE QUOTE WARNING\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 {alert.customer}\n"
        f"🚢 {alert.routing} | {alert.carrier}\n"
        f"📋 Job: {alert.job_id} — {alert.quote_age_days} days old\n"
        f"💰 Quote: ${alert.quoted_price:,.0f} | Market: ${alert.market_price:,.0f}\n"
        f"📊 Rate Δ: ${alert.rate_changed:,.0f} since quoted\n"
        f"📅 ETD: {etd_str}\n"
        f"\n"
        f"💡 Recommend: refresh quote or follow up with customer"
    )


def format_daily_summary(rate_alerts: list[RateAlert],
                         freshness_alerts: list[FreshnessAlert],
                         market_stats: dict) -> str:
    """Format the complete daily summary for Telegram."""
    now = datetime.now().strftime("%d-%b-%Y %H:%M")

    lines = [
        f"📊 NMI DAILY CHECK — {now}",
        f"{'━' * 40}",
    ]

    # Market stats
    if market_stats:
        lines.append(f"📈 Market: {market_stats.get('total_rates', 0)} rates tracked")
        lines.append(f"   Carriers: {market_stats.get('carriers', 0)} | Routes: {market_stats.get('routes', 0)}")

    # Rate alerts summary
    drops = [a for a in rate_alerts if a.alert_type == "DROP"]
    spikes = [a for a in rate_alerts if a.alert_type == "SPIKE"]
    lines.append(f"\n⚠️ Anomalies: {len(drops)} drops | {len(spikes)} spikes")

    if rate_alerts:
        lines.append("")
        for alert in rate_alerts[:5]:  # Top 5
            emoji = "📉" if alert.alert_type == "DROP" else "📈"
            lines.append(
                f"  {emoji} {alert.customer} | {alert.carrier} {alert.container} | "
                f"Δ${abs(alert.delta):,.0f} ({abs(alert.delta_pct):.1f}%)"
            )
        if len(rate_alerts) > 5:
            lines.append(f"  ... +{len(rate_alerts) - 5} more")

    # Freshness alerts
    if freshness_alerts:
        lines.append(f"\n⏰ Stale Quotes: {len(freshness_alerts)}")
        for fa in freshness_alerts[:3]:
            lines.append(
                f"  {fa.customer} | {fa.carrier} | {fa.quote_age_days}d old | Δ${fa.rate_changed:,.0f}"
            )

    if not rate_alerts and not freshness_alerts:
        lines.append("\n✅ All clear — no anomalies detected")

    return "\n".join(lines)


def send_telegram(message: str, dry_run: bool = False) -> bool:
    """Send alert to Telegram. Returns True if sent."""
    if dry_run:
        print("\n" + "=" * 60)
        print("[DRY RUN] Would send to Telegram:")
        print("=" * 60)
        print(message)
        print("=" * 60)
        return True

    token = os.environ.get("TELEGRAM_BOT_TOKEN", "")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID", "")

    if not token or not chat_id:
        log.warning("Telegram not configured (TELEGRAM_BOT_TOKEN / TELEGRAM_CHAT_ID missing)")
        log.info("Alert message:\n%s", message)
        return False

    try:
        import requests
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        resp = requests.post(url, json={
            "chat_id": chat_id,
            "text": message,
            "parse_mode": "HTML",
        }, timeout=10)
        if resp.status_code == 200:
            log.info("Telegram alert sent successfully")
            return True
        else:
            log.error("Telegram send failed: %s", resp.text)
            return False
    except Exception as e:
        log.error("Telegram error: %s", e)
        return False


# ==============================================================================
# MAIN ENTRY POINT
# ==============================================================================

def run_daily_check(dry_run: bool = False, threshold_override: float = None) -> dict:
    """
    Main daily check routine.
    1. Load latest rates from Parquet
    2. Load open quotes from ERP
    3. Detect anomalies
    4. Detect stale quotes
    5. Format and send alerts

    Returns summary dict.
    """
    log.info("=" * 60)
    log.info("NMI RATE MONITOR — Daily Check")
    log.info("=" * 60)

    config = NMIConfig.from_file()
    if threshold_override:
        config.drop_alert_usd = threshold_override
        config.spike_alert_usd = threshold_override

    monitor = RateMonitor(config)

    # 1. Load data
    rates = monitor.load_latest_rates()
    quotes = monitor.load_open_quotes()

    market_stats = {
        "total_rates": len(rates),
        "carriers": rates["Carrier"].nunique() if not rates.empty else 0,
        "routes": rates.groupby(["POL", "POD"]).ngroups if not rates.empty else 0,
    }

    # 2. Detect anomalies
    rate_alerts = monitor.detect_anomalies()

    # 3. Detect stale quotes
    freshness_alerts = monitor.detect_stale_quotes()

    # 4. Format messages
    # Send individual alerts for critical ones
    for alert in rate_alerts:
        msg = format_rate_alert(alert)
        send_telegram(msg, dry_run=dry_run)

    for fa in freshness_alerts:
        if fa.quote_age_days >= config.auto_expire_days:
            msg = format_freshness_alert(fa)
            send_telegram(msg, dry_run=dry_run)

    # 5. Send daily summary
    summary_msg = format_daily_summary(rate_alerts, freshness_alerts, market_stats)
    send_telegram(summary_msg, dry_run=dry_run)

    # 6. Log results
    result = {
        "timestamp": datetime.now().isoformat(),
        "market_rates_loaded": len(rates),
        "open_quotes_loaded": len(quotes),
        "rate_alerts": len(rate_alerts),
        "freshness_alerts": len(freshness_alerts),
        "drops": sum(1 for a in rate_alerts if a.alert_type == "DROP"),
        "spikes": sum(1 for a in rate_alerts if a.alert_type == "SPIKE"),
    }

    # Append to log file
    log_file = DATA_DIR / "nmi_alerts.log"
    try:
        with open(log_file, "a", encoding="utf-8") as f:
            f.write(json.dumps(result) + "\n")
    except Exception as e:
        log.warning("Could not write log: %s", e)

    log.info("\n" + "=" * 60)
    log.info("NMI CHECK COMPLETE")
    log.info("  Market rates: %d | Open quotes: %d", len(rates), len(quotes))
    log.info("  Alerts: %d rate anomalies | %d stale quotes",
             len(rate_alerts), len(freshness_alerts))
    log.info("=" * 60)

    return result


def show_route_rates(route: str):
    """Show current market rates for a route (e.g. HCM-LAX)."""
    parts = route.upper().replace("→", "-").replace(" ", "").split("-")
    pol = parts[0] if parts else "HCM"

    monitor = RateMonitor()
    monitor.load_latest_rates()
    summary = monitor.get_market_summary(pol)
    print(summary)


# ==============================================================================
# CLI
# ==============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="NMI Rate Monitor — Detect rate anomalies vs open quotes"
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Print alerts without sending to Telegram")
    parser.add_argument("--route", type=str,
                        help="Show market rates for a route (e.g. HCM-LAX)")
    parser.add_argument("--threshold", type=float,
                        help="Override alert threshold in USD")
    parser.add_argument("--summary", action="store_true",
                        help="Show market summary only")

    args = parser.parse_args()

    if args.route:
        show_route_rates(args.route)
    elif args.summary:
        monitor = RateMonitor()
        monitor.load_latest_rates()
        print(monitor.get_market_summary())
    else:
        result = run_daily_check(
            dry_run=args.dry_run,
            threshold_override=args.threshold,
        )
        print(f"\n✅ Check complete: {result['rate_alerts']} alerts, "
              f"{result['freshness_alerts']} stale quotes")


if __name__ == "__main__":
    main()
