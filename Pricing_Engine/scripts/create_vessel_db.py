# -*- coding: utf-8 -*-
"""
Vessel Capacity Database — Generator
Creates vessel_database.xlsx with initial data for major carriers.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

VESSELS = [
    # ONE Fleet (major vessels on US trades)
    ("ONE APUS",      "ONE",   "9806079", 14000, "PS3",       2020, "JP"),
    ("ONE AQUILA",    "ONE",   "9806081", 14000, "PS3",       2020, "JP"),
    ("ONE ATLAS",     "ONE",   "9795762", 14000, "PS3,EC4",   2020, "JP"),
    ("ONE OWL",       "ONE",   "9351971",  8600, "EC3",       2008, "PA"),
    ("ONE MINATO",    "ONE",   "9302140",  8600, "EC4",       2006, "PA"),
    ("ONE STORK",     "ONE",   "9312996",  4600, "PS4",       2006, "PA"),
    ("ONE COLUMBA",   "ONE",   "9784816", 14000, "PS3",       2019, "JP"),
    ("ONE CRANE",     "ONE",   "9324898",  4600, "PS4",       2006, "PA"),
    
    # ZIM Fleet
    ("ZIM MOUNT BLANC",  "ZIM", "9811000", 15000, "ZXB",     2021, "IL"),
    ("ZIM MOUNT DENALI", "ZIM", "9811012", 15000, "ZXB",     2021, "IL"),
    ("ZIM MOUNT FUJI",   "ZIM", "9811024", 15000, "Z7S",     2021, "IL"),
    ("ZIM CHARLESTON",   "ZIM", "9367903",  4250, "ZEX",     2008, "IL"),
    ("ZIM SAVANNAH",     "ZIM", "9367915",  4250, "ZEX",     2008, "IL"),
    
    # CMA CGM Fleet
    ("CMA CGM LOIRE",       "CMA", "9839200", 15000, "SAX",  2021, "FR"),
    ("CMA CGM NILE",        "CMA", "9481784",  6500, "OPNW", 2011, "FR"),
    ("CMA CGM BENJAMIN FRANKLIN", "CMA", "9706891", 18000, "SAX", 2015, "FR"),
    
    # COSCO Fleet
    ("COSCO SHIPPING ARIES",   "COSCO", "9783491", 20000, "AWE5",  2018, "CN"),
    ("COSCO SHIPPING TAURUS",  "COSCO", "9783506", 20000, "AWE5",  2018, "CN"),
    
    # MSC Fleet  
    ("MSC ANNA",      "MSC", "9695160", 19000, "Sentosa",    2016, "PA"),
    ("MSC GULSUN",    "MSC", "9839430", 23756, "Liberty",    2019, "PA"),
    ("MSC AMBRA",     "MSC", "9647461", 16600, "Chinook",    2014, "PA"),
    
    # EMC Fleet
    ("EVER ACE",      "EMC", "9893890", 23992, "TP22",       2021, "PA"),
    ("EVER GIVEN",    "EMC", "9811000", 20124, "TP22",       2018, "PA"),
    
    # YML Fleet
    ("YM TRUTH",      "YML", "9684665", 14220, "PS4",        2015, "PA"),
    ("YM TRIUMPH",    "YML", "9684677", 14220, "PS4",        2015, "PA"),
    
    # HPL Fleet
    ("HYUNDAI SPEED",     "HPL", "9700332", 13100, "FP2",    2016, "PA"),
    ("HYUNDAI TOGETHER",  "HPL", "9475686",  8600, "EC2",    2013, "PA"),
    
    # HMM Fleet
    ("HMM ALGECIRAS",  "HMM", "9863297", 23964, "THE Alliance", 2020, "KR"),
    ("HMM OSLO",       "HMM", "9863302", 23964, "THE Alliance", 2020, "KR"),
    
    # WHL Fleet
    ("WAN HAI 613",   "WHL", "9889912",  3055, "PS4",        2021, "TW"),
    ("WAN HAI 615",   "WHL", "9889924",  3055, "PS4",        2021, "TW"),
]

HEADERS = ["Vessel_Name", "Carrier", "IMO", "TEU_Capacity", "Services", "Built_Year", "Flag"]

def create_vessel_db():
    wb = Workbook()
    ws = wb.active
    ws.title = "Vessels"
    
    FONT = 'Arial'
    hdr_font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    data_font = Font(name=FONT, size=10)
    
    for col_idx, header in enumerate(HEADERS, 1):
        c = ws.cell(1, col_idx, header)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_idx, vessel in enumerate(VESSELS, 2):
        for col_idx, val in enumerate(vessel, 1):
            c = ws.cell(row_idx, col_idx, val)
            c.font = data_font
    
    # Auto-fit
    widths = [22, 10, 12, 14, 18, 12, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    # Table
    ref = f"A1:{get_column_letter(len(HEADERS))}{len(VESSELS)+1}"
    tab = Table(displayName="VesselDB", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2', showRowStripes=True,
        showFirstColumn=False, showLastColumn=False, showColumnStripes=False
    )
    ws.add_table(tab)
    ws.freeze_panes = 'A2'
    
    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(1, 1, "Carrier").font = hdr_font
    ws2.cell(1, 1).fill = hdr_fill
    ws2.cell(1, 2, "Vessels").font = hdr_font
    ws2.cell(1, 2).fill = hdr_fill
    ws2.cell(1, 3, "Total TEU").font = hdr_font
    ws2.cell(1, 3).fill = hdr_fill
    ws2.cell(1, 4, "Avg TEU").font = hdr_font
    ws2.cell(1, 4).fill = hdr_fill
    
    from collections import Counter, defaultdict
    carrier_stats = defaultdict(lambda: {"count": 0, "teu": 0})
    for v in VESSELS:
        carrier_stats[v[1]]["count"] += 1
        carrier_stats[v[1]]["teu"] += v[3]
    
    for r, (carrier, stats) in enumerate(sorted(carrier_stats.items()), 2):
        ws2.cell(r, 1, carrier).font = data_font
        ws2.cell(r, 2, stats["count"]).font = data_font
        ws2.cell(r, 3, stats["teu"]).font = data_font
        ws2.cell(r, 4, round(stats["teu"] / stats["count"])).font = data_font
    
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 10
    ws2.freeze_panes = 'A2'
    
    out = r'd:\NELSON\2. Areas\PricingSystem\Engine_test\Pricing_Engine\data\vessel_database.xlsx'
    wb.save(out)
    print(f"✅ Created vessel_database.xlsx — {len(VESSELS)} vessels, {len(carrier_stats)} carriers")
    return out

if __name__ == '__main__':
    create_vessel_db()
