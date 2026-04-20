# -*- coding: utf-8 -*-
"""
MasterFullPricing Generator - Simplified Version
Creates Master + Recent_History + PUC sheets (No Dashboard)
"""
import pandas as pd
import os
import sys

# Fix encoding for Windows console (guard for pythonw.exe where stdout=None)
if sys.platform == 'win32':
    import io
    if sys.stdout and hasattr(sys.stdout, 'buffer'):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    elif sys.stdout is None:
        sys.stdout = open(os.devnull, 'w', encoding='utf-8')

# --- CẤU HÌNH ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)  # Pricing_Engine folder
DATA_DIR = os.path.join(BASE_DIR, "data")  # Pricing_Engine/data
HISTORY_FILE = os.path.join(DATA_DIR, "Cleaned_Master_History.parquet")
OUTPUT_MASTER = os.path.join(DATA_DIR, "MasterFullPricing.xlsx")
PORT_MAP_FILE = os.path.join(DATA_DIR, "Port_Code_Mapping_Final.xlsx")
PUC_SOC_FILE  = os.path.join(DATA_DIR, "PUC_SOC.xlsx")
CONFIG_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'pipeline_rules.json')

# Import normalize functions from standalone module (DRY — single source).
# Fallback: functions defined inline below (legacy path, kept for safety).
_repo_root_cm = os.path.dirname(BASE_DIR)  # Engine_test/
if _repo_root_cm not in sys.path:
    sys.path.insert(0, _repo_root_cm)
try:
    from Pricing_Engine.normalization.text_normalize import (
        normalize_notes as _normalize_notes_ext,
        normalize_text_data as _normalize_text_data_ext,
        normalize_commodity_display as _normalize_commodity_display_ext,
        normalize_container_types as _normalize_container_types_ext,
    )
    _USE_TEXT_NORMALIZE_MODULE = True
    print("  [create_master] Using Pricing_Engine.normalization.text_normalize module")
except ImportError as _tn_err:
    _USE_TEXT_NORMALIZE_MODULE = False
    print(f"  [WARN] text_normalize module unavailable ({_tn_err}); using inline functions")


def _dispatch_normalize_notes(df):
    """Delegate to external module if available, else inline."""
    if _USE_TEXT_NORMALIZE_MODULE:
        return _normalize_notes_ext(df)
    return normalize_notes(df)  # inline definition below


def _dispatch_normalize_text_data(df, port_map):
    """Delegate to external module if available, else inline."""
    if _USE_TEXT_NORMALIZE_MODULE:
        return _normalize_text_data_ext(df, port_map)
    return normalize_text_data(df, port_map)  # inline definition below


def load_mapping_files():
    """Load port code mapping"""
    port_map = {}
    if os.path.exists(PORT_MAP_FILE):
        try:
            df_pm = pd.read_excel(PORT_MAP_FILE)
            df_pm['PortName'] = df_pm['PortName'].astype(str).str.upper().str.strip()
            df_pm['PortCode'] = df_pm['PortCode'].astype(str).str.upper().str.strip()
            port_map = dict(zip(df_pm['PortName'], df_pm['PortCode']))
        except:
            pass
    return port_map


def create_puc_lookup_table():
    """Create PUC lookup table from latest PUC*.xlsx file in data/"""
    import os
    import pandas as pd
    import re
    
    # 1. Find latest PUC file
    latest_puc = None
    files = sorted([f for f in os.listdir(DATA_DIR) if f.startswith('PUC') and f.endswith('.xlsx')], reverse=True)
    if files:
        latest_puc = os.path.join(DATA_DIR, files[0])
        print(f"    → Found PUC file: {files[0]}")
    
    if not latest_puc or not os.path.exists(latest_puc):
        print(f"    ⚠️ No PUC file found in {DATA_DIR}")
        return pd.DataFrame()
    
    try:
        # 2. Read and handle Row 0 sub-headers
        df_puc = pd.read_excel(latest_puc)
        if df_puc.empty: return pd.DataFrame()
        
        # Standardize columns
        # Based on audit: Destination, TYPE, Unnamed: 2, Expiration Date
        # Row 0 often has: NaN, 20DC, 40HC, NaT
        puc_data = []
        for ri, row in df_puc.iterrows():
            dest = str(row.get('Destination', '')).strip().upper()
            if not dest or dest == 'NAN' or dest == 'DESTINATION': continue
            
            # 3. Smart Mapping: Los Angeles = LAX/LGB
            if 'LOS ANGELES' in dest:
                dest = 'LOS ANGELES'
                
            # Get values - handle based on position/name
            # TYPE is usually 20ft, Unnamed: 2 is 40ft
            val_20 = row.get('TYPE', row.get('20GP', row.get('20DC', 0)))
            val_40 = row.get('Unnamed: 2', row.get(40, row.get('40HC', 0)))
            
            # Map 40ft to both 40GP and 40HQ
            puc_data.append({
                'Place': dest,
                '20GP': float(val_20) if pd.notna(val_20) and str(val_20).replace('.', '', 1).replace('-', '', 1).isdigit() else 0,
                '40GP': float(val_40) if pd.notna(val_40) and str(val_40).replace('.', '', 1).replace('-', '', 1).isdigit() else 0,
                '40HQ': float(val_40) if pd.notna(val_40) and str(val_40).replace('.', '', 1).replace('-', '', 1).isdigit() else 0,
                '45HQ': float(val_40) if pd.notna(val_40) and str(val_40).replace('.', '', 1).replace('-', '', 1).isdigit() else 0
            })
        
        # Add special mapping for LAX/LGB if LOS ANGELES is present
        ext_puc = []
        for p in puc_data:
            if p['Place'] == 'LOS ANGELES':
                # Create copies for common port codes
                lax = p.copy(); lax['Place'] = 'LAX'
                lgb = p.copy(); lgb['Place'] = 'LGB'
                ext_puc.extend([lax, lgb])
        
        puc_data.extend(ext_puc)
        return pd.DataFrame(puc_data)
    except Exception as e:
        print(f"    ⚠️ Error parsing PUC file: {e}")
        return pd.DataFrame()


def shorten_source_file(filename):
    """Shorten source file names"""
    if not filename or pd.isna(filename):
        return ""
    
    fname = str(filename).upper()
    if "FAK" in fname:
        import re as _re
        m = _re.search(r'(\d{1,2})\s*([A-Z]{3})\s*NO\.?\s*(\d+)', fname)
        if m:
            return f"FAK {m.group(1)}{m.group(2)} NO.{m.group(3)}"
        return "FAK"
    elif "SCFI" in fname:
        return "SCFI"
    elif "OCR" in fname:
        return "OCR"
    elif "SPECIAL" in fname:
        return "SPECIAL RATE"
    elif "FIX" in fname:
        return "FIX"
    else:
        return filename[:15]


def apply_one_group_codes(df, config_file=None):
    """Auto-populate 'Group Rate' for ONE carrier rows using pipeline_rules.json rules."""
    if config_file is None:
        config_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'pipeline_rules.json')
    
    if not os.path.exists(config_file):
        print(f"  [ONE GROUP] pipeline_rules.json not found - skipping")
        return df
    
    try:
        import json
        with open(config_file, 'r', encoding='utf-8') as f:
            config = json.load(f)
        rules = config.get('stage_2_one_group_codes', {}).get('rules', [])
        if not rules:
            print(f"  [ONE GROUP] No rules found in pipeline_rules.json - skipping")
            return df
        rules = sorted(rules, key=lambda r: r.get('priority', 99))
    except Exception as e:
        print(f"  [ONE GROUP] Error loading rules: {e} - skipping")
        return df

    mask_one = df['Carrier'].str.upper().str.contains('ONE', na=False)
    if not mask_one.any():
        return df

    def _pod_region(pod_code):
        """Return 'CANADA' or 'USA' based on POD prefix."""
        pc = str(pod_code or '').upper().strip()
        if pc[:2] == 'CA' and not pc.startswith('CAI'):
            return 'CANADA'
        return 'USA'

    def _match(value, pattern):
        """True if value contains any keyword in comma-separated pattern, or pattern is *."""
        p = str(pattern).strip()
        if p in ('*', '', 'nan'):
            return True
        val_up = str(value or '').upper()
        return any(kw.strip().upper() in val_up for kw in p.split(',') if kw.strip())

    applied = 0
    for idx in df[mask_one].index:
        row       = df.loc[idx]
        commodity = str(row.get('Commodity', '') or '')
        note      = str(row.get('Note', '') or '')
        pod_code  = str(row.get('POD', '') or '')
        source    = str(row.get('Source', '') or '').upper()
        rate_type = str(row.get('Rate_Type', '') or '').upper()
        contract  = 'FIX' if ('FIX' in rate_type or 'FIX' in source) else 'FAK'
        region    = _pod_region(pod_code)

        for rule in rules:
            r_contract = str(rule.get('contract_type', '*')).strip().upper()
            r_pod      = str(rule.get('pod_region', '*')).strip().upper()
            r_comm     = rule.get('commodity_keywords', '*')
            r_note     = rule.get('note_keywords', '*')
            r_code     = str(rule.get('group_code', '')).strip()
            r_label    = str(rule.get('group_label', '')).strip()

            if r_contract not in ('*', '', 'NAN') and r_contract != contract:
                continue
            if r_pod not in ('*', '', 'NAN') and r_pod != region:
                continue
            if not _match(commodity, r_comm):
                continue
            if not _match(note, r_note):
                continue

            # Format output
            if '|' in r_code:
                codes  = [c.strip() for c in r_code.split('|')]
                labels = [l.strip() for l in r_label.split('|')]
                group_val = ' / '.join(
                    f"{c} | {l}" for c, l in zip(codes, labels) if c
                )
            else:
                group_val = f"{r_code} | {r_label}"

            df.at[idx, 'Group Rate'] = group_val
            applied += 1
            break  # first match wins

    print(f"  [ONE GROUP] Applied group codes: {applied}/{mask_one.sum()} ONE rows")
    return df


def normalize_notes(df):
    """Normalize note variations into standardized short labels.

    Priority order inside _normalize:
      1. ZIM-specific (Z7S / ZXB / ZEX / generic ZIM OWS)
      2. EMC-specific (PCTF/STF/PCS/SUEZ)
      3. COSCO vessel names
      4. MSC service names (shortened with [ref])
      5. CMA service names
      6. SOC variants (TRANSIT / Cai Mep / DIRECT / plain SOC)
      7. Non-SOC routing (TRANSIT / Cai Mep / DIRECT)
      8. Unrecognized -> kept as-is
    """
    if 'Note' not in df.columns:
        return df

    TRANSIT_KEYWORDS = ['YANTIAN', 'KAOHSIUNG', 'HONG KONG', 'SINGAPORE', 'SHANGHAI']
    CAI_MEP_KEYWORDS = ['CAI MEP']
    DIRECT_KEYWORDS  = ['DIRECT', 'HPH']
    HAIPHONG_WORDS   = ['HAIPHONG', 'HAI PHONG']

    import re as _re
    _TONNAGE_RE = _re.compile(r'(?:UP\s*TO|UPTO)\s*([\d]+(?:\.[\d]+)?)', _re.IGNORECASE)

    def _tonnage_warn(nu):
        """Return warning string if OWS tonnage limit < 22 tons detected."""
        m = _TONNAGE_RE.search(nu)
        if m:
            try:
                tons = float(m.group(1))
                if tons < 22:
                    return f' [!OWS<22T:{tons}t]'
            except ValueError:
                pass
        return ''

    def _normalize(note, carrier=''):
        if not note or str(note).strip() in ('', 'nan'):
            return ''
        n = str(note).strip()
        nu = n.upper()
        cu = str(carrier).upper()

        # =========================================================
        # 1. ZIM SERVICE BLOCK
        # =========================================================
        if 'ZIM' in nu or 'Z7S' in nu or 'ZXB' in nu or 'ZEX' in nu:
            if 'OWS' in nu:
                if 'SUBJECT TO OWS' in nu or 'SUBJECT TO  OWS' in nu:
                    ows_tag = ' OWS EXTRA'
                else:
                    ows_tag = ' OWS INCL'
                warn = _tonnage_warn(nu)
            else:
                ows_tag = ''
                warn = ''
            if 'Z7S' in nu:
                return f'Z7S{ows_tag}{warn}'
            if 'ZXB' in nu:
                return f'ZXB{ows_tag}{warn}'
            if 'ZEX' in nu:
                return 'ZEX'
            if ows_tag:
                return f'ZIM{ows_tag}{warn}'
            return n

        # 1b. OWS CATCH-ALL
        if 'OWS' in nu:
            if 'SUBJECT TO OWS' in nu or 'SUBJECT TO  OWS' in nu:
                return f'ZIM OWS EXTRA{_tonnage_warn(nu)}'
            return f'ZIM OWS INCL{_tonnage_warn(nu)}'

        # =========================================================
        # 2. EMC-SPECIFIC: PCTF / STF / PCS / SUEZ
        # =========================================================
        if 'EMC' in cu:
            has_cmep = 'VIA CMEP' in nu or 'CMEP' in nu
            if 'PCTF' in nu or 'PANAMA CANAL TRANSIT' in nu:
                if 'STF' in nu or 'SUEZ TRANSIT' in nu:
                    return 'via CMEP PCTF/STF' if has_cmep else 'PCTF/STF SURCHG'
                return 'via CMEP PCTF' if has_cmep else 'PCTF SURCHG'
            if 'PCS' in nu or 'SUEZ' in nu:
                return 'via CMEP PCS/SUEZ' if has_cmep else 'PCS/SUEZ'
            if has_cmep:
                return 'via CMEP'

        # =========================================================
        # 3. COSCO-SPECIFIC: Vessel / service names
        # =========================================================
        if 'COSCO' in cu:
            if 'CMA CGM NILE' in nu or 'POINTE-NOIRE' in nu or 'POINTE NOIRE' in nu:
                return 'NILE/P-NOIRE'
            if 'OPNW' in nu:
                return 'via OPNW'

        # =========================================================
        # 4. MSC-SPECIFIC: Service name shortening [ref: full name]
        # =========================================================
        if 'MSC' in cu:
            if 'AMERICA' in nu and 'EMPIRE' in nu:
                return 'AMR/EMP/AMB/EMR/ELE/SAN/LION [ref:SvcGroup1]'
            if 'LIBERTY' in nu and ('NOT' in nu or 'OTHER' in nu):
                return 'non-Liberty'
            if 'SENTOSA' in nu and 'PEARL' in nu:
                return 'Sentosa/Pearl'
            if 'LONE STAR' in nu or 'PELICAN' in nu:
                return 'LONE STAR/PELICAN'
            if 'CHINOOK' in nu:
                if 'ON CHINOOK' in nu:
                    return 'on Chinook'
                return 'Chinook'

        # =========================================================
        # 5. CMA-SPECIFIC: service names
        # =========================================================
        if 'CMA' in cu:
            if 'SERVICE' in nu and ':' in n:
                # "service : SAX CS DEVELOPMENT" → "SAX CS"
                svc_name = n.split(':', 1)[1].strip()
                # Remove trailing long descriptions
                svc_short = svc_name.split(' ')[0:2]  # Keep first 2 words
                return ' '.join(svc_short)

        # =========================================================
        # 6. SOC / ROUTING BLOCK
        # =========================================================
        has_soc = 'SOC' in nu

        if any(kw in nu for kw in TRANSIT_KEYWORDS):
            return 'SOC TRANSIT' if has_soc else 'TRANSIT'

        if any(kw in nu for kw in CAI_MEP_KEYWORDS):
            return 'SOC Cai Mep (EC3)' if has_soc else 'Cai Mep (EC3)'

        if any(kw in nu for kw in DIRECT_KEYWORDS) or any(kw in nu for kw in HAIPHONG_WORDS):
            return 'SOC DIRECT' if has_soc else 'DIRECT'

        # Unrecognized -> keep trimmed as-is
        return n

    # Apply with carrier context
    df['Note'] = df.apply(lambda r: _normalize(r.get('Note', ''), r.get('Carrier', '')), axis=1)
    return df


def normalize_text_data(df, port_map):
    """Normalize commodity names and port codes"""
    if 'Commodity' not in df.columns:
        df['Commodity'] = ''

    comm = df['Commodity'].astype(str)
    carrier = df['Carrier'].astype(str).str.upper()

    # === UNIVERSAL: FAK INCLUDING/EXCLUDING GARMENT (all carriers) ===
    # Keep distinction: INCL vs EXCL
    mask_fak_incl = comm.str.contains('FAK', case=False, na=False) & \
                    comm.str.contains('INCLUDING|INCL', case=False, na=False) & \
                    comm.str.contains('GARMENT', case=False, na=False)
    df.loc[mask_fak_incl, 'Commodity'] = 'FAK INCL GARMENT'

    mask_fak_excl = comm.str.contains('FAK', case=False, na=False) & \
                    comm.str.contains('EXCLUDING|EXCL', case=False, na=False) & \
                    comm.str.contains('GARMENT', case=False, na=False)
    df.loc[mask_fak_excl, 'Commodity'] = 'FAK EXCL GARMENT'

    # Refresh after universal rules
    comm = df['Commodity'].astype(str)

    # === COSCO ===
    mask_cosco = carrier.str.contains('COSCO', na=False)
    df.loc[mask_cosco & comm.str.contains('GARMENT|TEXTILE|CONSOL', case=False, na=False) &
           ~comm.str.contains('FAK', case=False, na=False), 'Commodity'] = 'GARMENT'

    # === ONE ===
    mask_one = carrier.str.contains('ONE', na=False)
    df.loc[mask_one & comm.str.contains('REEFER', case=False, na=False), 'Commodity'] = 'REEFER FAK'
    # FAK: TPE1/TPF1/TPE variants → standardize to FAK: TPE1
    df.loc[mask_one & comm.str.contains(r'FAK[:\s]*TP[EF]\d', case=False, na=False, regex=True), 'Commodity'] = 'FAK: TPE1'
    # GDSM Straight (Excluding...) → SHORT TERM GDSM
    df.loc[mask_one & comm.str.contains('GDSM', case=False, na=False), 'Commodity'] = 'SHORT TERM GDSM'
    df.loc[mask_one & comm.str.upper().str.strip().str.startswith('GARMENT'), 'Commodity'] = 'GARMENT'
    # ONE SOC Group: "S1– TPE9 – Group SOC: xxxx" or "S1– TPE10 – Group SOC"
    comm = df['Commodity'].astype(str)  # refresh
    mask_soc_group = comm.str.contains('Group SOC', case=False, na=False)
    if mask_soc_group.any():
        # Strip anything after "Group SOC" (colon or not)
        import re as _re_commodity
        df.loc[mask_one & mask_soc_group, 'Commodity'] = df.loc[mask_one & mask_soc_group, 'Commodity'].apply(
            lambda x: _re_commodity.split(r'Group SOC\b', str(x), maxsplit=1)[0].strip() + 'Group SOC'
            if 'Group SOC' in str(x).upper() or 'group soc' in str(x).lower() else x
        )
    # ONE FURNITURE/FLOORING/QUARTZ → FURNITURE
    df.loc[mask_one & comm.str.contains('FURNITURE', case=False, na=False), 'Commodity'] = 'FURNITURE'
    # ONE HS codes (numeric patterns like 0901.00.0000) → HS CODE
    df.loc[mask_one & comm.str.contains(r'^\d{4}\.\d{2}\.\d{4}', na=False, regex=True), 'Commodity'] = 'HS CODE'

    # === YML ===
    mask_yml = carrier.str.contains('YML', na=False)
    comm = df['Commodity'].astype(str)  # refresh
    # "GROUP A : FAK (NON-HAZ, EXCLUDING REEFER/SHIPS/...)" → "GROUP A : FAK"
    df.loc[mask_yml & comm.str.contains('GROUP A', case=False, na=False) &
           comm.str.contains('FAK', case=False, na=False), 'Commodity'] = 'GROUP A : FAK'
    # Plain FAK with exclusions → "FAK"
    df.loc[mask_yml & comm.str.contains('FAK', case=False, na=False) &
           comm.str.contains('NON-HAZ|EXCLUDING', case=False, na=False) &
           ~comm.str.contains('GROUP', case=False, na=False), 'Commodity'] = 'FAK'
    # "SHIPS/ BOATS/ VEHICLES/ CARS" → "VEHICLES/CARS"
    df.loc[mask_yml & comm.str.contains('SHIPS|BOATS|VEHICLES|CARS', case=False, na=False) &
           ~comm.str.contains('FAK', case=False, na=False), 'Commodity'] = 'VEHICLES/CARS'

    # === CMA ===
    mask_cma = carrier.str.contains('CMA', na=False)
    df.loc[mask_cma & comm.str.contains('PANAMA', case=False, na=False), 'Commodity'] = 'PANAMA SURCHG'
    df.loc[mask_cma & comm.str.contains('DIRECT SERVICE', case=False, na=False), 'Commodity'] = 'DIRECT SVC'

    # === EMC ===
    mask_emc = carrier.str.contains('EMC', na=False)
    comm = df['Commodity'].astype(str)  # refresh
    df.loc[mask_emc & comm.str.contains('RATE 1.*GENERAL CARGO', case=False, na=False), 'Commodity'] = 'RATE 1'

    # === ZIM ===
    mask_zim = carrier.str.contains('ZIM', na=False)
    comm = df['Commodity'].astype(str)  # refresh
    df.loc[mask_zim & comm.str.contains('SUBJECT TO OWS|OWS.*20|20.*OWS|include OWS|OWS include', case=False, na=False), 'Commodity'] = 'OWS 20GP'
    # ZIM HS codes → HS CODE
    df.loc[mask_zim & comm.str.contains(r'^\d{4}\.\d{2}\.\d{4}', na=False, regex=True), 'Commodity'] = 'HS CODE'
    # ZIM INCLUDE/FREETIME/DET notes in commodity → FAK
    df.loc[mask_zim & comm.str.contains('INCLUDE.*FREETIME|FREETIME.*COMBINE|DET.*include', case=False, na=False), 'Commodity'] = 'FAK'
    # ZIM FAK HAWAII
    df.loc[mask_zim & comm.str.contains('HAWAII', case=False, na=False), 'Commodity'] = 'FAK HAWAII'

    # === MSC ===
    mask_msc = carrier.str.contains('MSC', na=False)
    comm = df['Commodity'].astype(str)  # refresh
    # MSC vessel/service names in commodity → move to note-like short label
    df.loc[mask_msc & comm.str.contains('America:|Elephant:|sentosa:|Pearl:', case=False, na=False), 'Commodity'] = 'FAK'

    # === WHL ===
    mask_whl = carrier.str.contains('WHL', na=False)
    df.loc[mask_whl & comm.str.contains('FOODSTUFF|FROZEN|SEAFOOD', case=False, na=False), 'Commodity'] = 'FROZEN FOOD'

    # === HMM / COSCO booking notes in commodity ===
    comm = df['Commodity'].astype(str)  # refresh
    df.loc[comm.str.contains('APPLY.*FILE.*COMMODITY|apply.*commodity.*booking', case=False, na=False), 'Commodity'] = 'FAK'

    # === CATCH-ALL: Truncate any remaining commodity > 25 chars ===
    comm = df['Commodity'].astype(str)
    long_mask = comm.str.len() > 25
    if long_mask.any():
        # Keep first 2 meaningful "words" (split on space/slash/comma), max 25 chars
        df.loc[long_mask, 'Commodity'] = comm[long_mask].apply(
            lambda x: ' '.join(str(x).split()[:3])[:25].strip()
        )

    # Map port codes (overwrite POL/POD directly with codes)
    if port_map:
        df['POL'] = df['POL'].astype(str).str.upper().str.strip().map(port_map).fillna(df['POL'])
        df['POD'] = df['POD'].astype(str).str.upper().str.strip().map(port_map).fillna(df['POD'])

    # Rename Source_File → Source
    if 'Source_File' in df.columns:
        df = df.rename(columns={'Source_File': 'Source'})

    return df


def create_master_v13():
    """Create MasterFullPricing with Master + History + PUC sheets"""
    print(f"\n=== CREATING MASTER FULL PRICING ===")

    # 1. Load Data
    if not os.path.exists(HISTORY_FILE):
        print(f"[!] Không tìm thấy file: {HISTORY_FILE}")
        return

    try:
        df_all = pd.read_parquet(HISTORY_FILE)
    except Exception as e:
        print(f"[!] Lỗi đọc parquet: {e}")
        return
    
    print(f"[1/4] Dữ liệu gốc: {len(df_all):,} dòng")

    # 2. Normalize
    port_map = load_mapping_files()
    df_all = _dispatch_normalize_text_data(df_all, port_map)

    # 3. Re-identify Rate_Type from filename
    df_all.loc[df_all['Source'].str.upper().str.contains("SCFI", na=False), 'Rate_Type'] = 'SCFI'
    df_all.loc[df_all['Source'].str.upper().str.contains("FIX", na=False), 'Rate_Type'] = 'FIX'
    df_all.loc[df_all['Source'].str.upper().str.contains("FAK", na=False), 'Rate_Type'] = 'FAK'
    df_all.loc[df_all['Source'].str.upper().str.contains("OCR", na=False), 'Rate_Type'] = 'OCR'

    # Normalize charge names
    mask_scfi_fix = df_all['Rate_Type'].isin(['SCFI', 'FIX', 'OCR'])
    df_all.loc[mask_scfi_fix, 'Charge_Name'] = "Base Ocean Freight"

    mask_fak = df_all['Rate_Type'] == 'FAK'
    mask_base_fak = df_all['Charge_Name'].str.contains('ALL IN|Base|Basic|O/F|Ocean', case=False, na=False)
    df_all.loc[mask_fak & mask_base_fak, 'Charge_Name'] = "Base Ocean Freight"

    # Normalize all note variations (SOC + non-SOC)
    df_all = _dispatch_normalize_notes(df_all)

    # Shorten source file names
    df_all['Source'] = df_all['Source'].apply(shorten_source_file)

    # --- SHEET MASTER (Base Ocean Freight only, last 30 days) ---
    print("[2/4] Tạo Sheet Master...")
    
    import datetime as _dt
    cutoff_date = pd.Timestamp(_dt.date.today()) - pd.Timedelta(days=30)
    
    current_rate_types = ['FAK', 'SCFI', 'FIX', 'SPECIAL', 'OCR']
    df_current = df_all[df_all['Rate_Type'].isin(current_rate_types)].copy()
    
    # Filter: only rates with Exp >= 30 days ago (active/recent rates)
    df_current['Exp'] = pd.to_datetime(df_current['Exp'], errors='coerce')
    df_current = df_current[df_current['Exp'] >= cutoff_date]
    print(f"    → After 30-day filter: {len(df_current):,} records (cutoff: {cutoff_date.strftime('%d-%b-%Y')})")
    
    df_master_raw = df_current[df_current['Charge_Name'] == "Base Ocean Freight"].copy()
    
    # Deduplicate
    df_master_raw = df_master_raw.sort_values('Amount', ascending=False)
    df_master_raw = df_master_raw.drop_duplicates(
        subset=['POL', 'POD', 'Place', 'Carrier', 'Commodity', 
                'Eff', 'Exp', 'Note', 'Source', 'Container_Type'],
        keep='first'
    )

    # Pivot to wide format
    df_master = df_master_raw.pivot_table(
        index=['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note', 'Source'],
        columns='Container_Type',
        values='Amount',
        aggfunc='max'
    ).reset_index()
    
    # Rename 45'HQ -> 45HQ (raw data uses 45'HQ as Container_Type)
    if "45'HQ" in df_master.columns:
        df_master = df_master.rename(columns={"45'HQ": "45HQ"})
    
    # Reorder columns
    desired_order = ['20GP', '40GP', '40HQ', '45HQ', '40NOR', '20RF', '40RF']
    existing_containers = [col for col in desired_order if col in df_master.columns]
    other_cols = [col for col in df_master.columns if col not in existing_containers 
                  and col not in ['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note', 'Source']]
    df_master = df_master[['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note', 'Source'] + existing_containers + other_cols]

    print(f"    → Master: {len(df_master):,} rows")

    # --- SHEET BASIC COST (same scope as Master, full charge breakdown) ---
    print("[3/4] Tạo Sheet Basic Cost...")

    # Re-read from Parquet WITHOUT the line-434 merge that collapsed
    # ALL IN COST / BASIC O/F / Base Ocean Freight into one charge name.
    # This keeps the raw charge names so BASIC O/F = true base ($1700)
    # and ALL IN COST = total ($1743) remain separate.
    df_bc_raw = pd.read_parquet(HISTORY_FILE)
    df_bc_raw = _dispatch_normalize_text_data(df_bc_raw, port_map)
    df_bc_raw['Exp'] = pd.to_datetime(df_bc_raw['Exp'], errors='coerce')
    df_recent = df_bc_raw[
        df_bc_raw['Rate_Type'].isin(current_rate_types) &
        (df_bc_raw['Exp'] >= cutoff_date)
    ].copy()
    
    # Exclude ALL IN / Total rows — they are redundant totals that would
    # overwrite the true BASIC O/F via aggfunc='max' in the pivot.
    exclude_charges = ['Total Ocean Freight', 'Base Ocean Freight', 'ALL IN COST']
    df_recent = df_recent[~df_recent['Charge_Name'].isin(exclude_charges)].copy()
    
    # Apply normalize_notes again (was applied to df_all but df_bc_raw is fresh)
    df_recent = _dispatch_normalize_notes(df_recent)
    df_recent['Source'] = df_recent['Source'].apply(shorten_source_file) if 'Source' in df_recent.columns else df_recent.get('Source_File', '').apply(shorten_source_file)

    # Apply ONE carrier Group Code rules from pipeline_rules.json
    df_recent = apply_one_group_codes(df_recent)

    # Map charge names for SCFI to FAK equivalents
    charge_mapping = {
        'EMF': 'EIC/GFS/BAF/FDI',
        'DLF': 'PCS',
        'ISPS': 'ISPS/LSF/CMC',
        'BASIC O/F': 'BASIC O/F',
        'COMMISSION': 'HANDLING FEE FOR CARRIER'  # SCFI Commission = HDL FEE
    }
    
    # Apply charge mapping
    df_recent['Charge_Name_Mapped'] = df_recent['Charge_Name'].replace(charge_mapping)
    
    # Pivot: charges as columns using FAK structure
    if not df_recent.empty:
        df_recent['Charge_Container'] = df_recent['Charge_Name_Mapped'] + '_' + df_recent['Container_Type']
        
        index_cols = ['Rate_Type', 'Source', 'POL', 'POD', 'Place', 
                     'Carrier', 'Commodity', 'Group Rate',
                     'Eff', 'Exp', 'Contract', 'Note']
        valid_idx = [c for c in index_cols if c in df_recent.columns]

        df_history_wide = df_recent.pivot_table(
            index=valid_idx, 
            columns='Charge_Container', 
            values='Amount', 
            aggfunc='max'
        ).reset_index()

        # Define FAK-based column order
        container_types = ['20GP', '40GP', '40HQ', '45HQ', '40NOR', '20RF', '40RF']
        charge_groups = [
            'BASIC O/F',           # Basic ocean freight (first - replaces Total Ocean Freight)
            'ARB/OLF',
            'PREMIUM ADD ON/HDL FEE US FOR SOC',
            'ISPS/LSF/CMC',
            'OCS/LSS/EFF/ITC/GFS/ SOC COST HDL FEE',
            'PSS/PUC',
            'GRI',
            'EIC/GFS/BAF/FDI',     # EMF maps here
            'PCS',                  # DLF maps here
            'PCS/ACS/AGS',
            'GARMENT ADD ON',
            'WHA/BCO/BCD/CFC/EIC',
            'HANDLING FEE FOR CARRIER',
            'COMMISSION'
        ]
        
        # Build ordered column list
        ordered_charge_cols = []
        for charge in charge_groups:
            for cont in container_types:
                col_name = f"{charge}_{cont}"
                if col_name in df_history_wide.columns:
                    ordered_charge_cols.append(col_name)
        
        # Add any remaining charge columns not in the predefined order
        remaining_cols = [c for c in df_history_wide.columns 
                         if c not in valid_idx and c not in ordered_charge_cols]
        
        # Final column order: index cols + ordered charge cols + remaining
        final_cols = valid_idx + ordered_charge_cols + sorted(remaining_cols)
        df_history_wide = df_history_wide[[c for c in final_cols if c in df_history_wide.columns]]
    else:
        df_history_wide = pd.DataFrame()

    print(f"    → Basic Cost: {len(df_history_wide):,} rows")

    # --- SHEET VERSION ---
    print("[4a] Tạo Sheet Version...")

    # Extract FAK version from source files in DATA_DIR
    import re, datetime
    today_str = datetime.date.today().strftime('%d-%b-%Y')

    # Collect all raw source files (same filter as master_loader_v2)
    EXCLUDE = ['PUC_SOC', 'Port_Code', 'Schedule', 'Master']
    raw_files = sorted([
        f for f in os.listdir(DATA_DIR)
        if f.endswith('.xlsx') and not f.startswith('~$')
        and not any(e in f for e in EXCLUDE)
    ])

    # Build FAK version label: e.g. "FAK_US CANADA_ 26 FEB NO. 2.xlsx" -> "26FEBNO.2"
    fak_version = 'N/A'
    for rf in raw_files:
        if 'FAK' in rf.upper():
            m = re.search(r'(\d{1,2})\s+([A-Z]{3})\s+NO\.?\s*(\d+)', rf.upper())
            if m:
                fak_version = f"{m.group(1)}{m.group(2)}NO.{m.group(3)}"
            break

    total_master_rows = len(df_master)

    version_rows = [
        ('Ấn bản bảng giá', fak_version),
        ('Tổng số dòng (Master)', total_master_rows),
        ('Ngày normalize', today_str),
        ('', ''),
    ]
    for rf in raw_files:
        fname_up = rf.upper()
        if 'FAK' in fname_up:
            label = 'RAW: FAK'
        elif 'SCFI' in fname_up:
            label = 'RAW: SCFI'
        elif 'SPECIAL' in fname_up or 'FIX' in fname_up:
            label = 'RAW: SPECIAL RATE'
        else:
            label = 'RAW'
        # Shorten display name
        display = re.sub(r'\.xlsx$', '', rf, flags=re.I)
        version_rows.append((label, display))

    # --- SHEET PUC ---
    print("[4/4] Tạo Sheet PUC...")
    df_puc = create_puc_lookup_table()
    print(f"    → PUC: {len(df_puc):,} rows")

    # --- REFERENCE SHEETS FROM FAK ---
    print("[5/4] Extracting reference sheets from FAK...")
    REF_SHEETS = {
        'HANDLING FEE CARRIER': 'Handling Fee',
        'ONE GROUP CODE': 'ONE Group Code',
        'WHARFAGE': 'Wharfage',
        'GOH': 'GOH',
        'HAWAII': 'Hawaii',
    }
    ref_sheet_data = {}
    
    # Find FAK file in DATA_DIR
    fak_file = None
    for f in os.listdir(DATA_DIR):
        if f.endswith('.xlsx') and 'FAK' in f.upper() and not f.startswith('~$'):
            fak_file = os.path.join(DATA_DIR, f)
            break
    
    if fak_file:
        try:
            fak_xls = pd.ExcelFile(fak_file)
            for src_name, dst_name in REF_SHEETS.items():
                if src_name in fak_xls.sheet_names:
                    ref_sheet_data[dst_name] = pd.read_excel(fak_xls, sheet_name=src_name, header=None)
                    print(f"    → {dst_name}: {ref_sheet_data[dst_name].shape[0]} rows")
                else:
                    print(f"    ⚠️ Sheet '{src_name}' not found in FAK")
        except Exception as e:
            print(f"    ⚠️ Error reading FAK reference sheets: {e}")
    else:
        print("    ⚠️ No FAK file found in data/ - skipping reference sheets")

    # --- SAVE TO EXCEL ---
    print(f"\n[SAVING] {OUTPUT_MASTER}")
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.worksheet.table import Table, TableStyleInfo
    import openpyxl

    FONT_NAME = 'Arial'
    FONT_SIZE = 10
    HEADER_FONT = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
    HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    DATA_FONT = Font(name=FONT_NAME, size=FONT_SIZE)
    THIN_BORDER = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    def _autofit_columns(ws, min_width=8, max_width=30):
        """Auto-fit column widths based on actual cell content."""
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = min_width
            for cell in col_cells[:min(500, len(col_cells))]:
                val = str(cell.value) if cell.value is not None else ''
                # Account for formula cells showing shorter
                if val.startswith('='):
                    val = val[:15]
                max_len = max(max_len, min(len(val) + 2, max_width))
            ws.column_dimensions[col_letter].width = max_len

    def _apply_table(ws, table_name, style='TableStyleMedium2'):
        """Convert data range to an Excel Table with auto-filter."""
        if ws.max_row < 2 or ws.max_column < 1:
            return
        max_col = get_column_letter(ws.max_column)
        ref = f'A1:{max_col}{ws.max_row}'
        # Sanitize table name (no spaces, special chars)
        safe_name = ''.join(c if c.isalnum() or c == '_' else '_' for c in table_name)
        tab = Table(displayName=safe_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name=style, showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(tab)

    def _style_header(ws):
        """Style header row: bold white on dark blue, centered."""
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER

    def _style_data(ws):
        """Apply data font to all non-header cells."""
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = DATA_FONT
                cell.border = THIN_BORDER

    def _format_dates(ws, col_names=('Eff', 'Exp')):
        """Format date columns as DD-MMM."""
        header_row = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        for name in col_names:
            if name in header_row:
                cidx = header_row[name]
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(r, cidx)
                    if cell.value:
                        cell.number_format = 'DD-MMM'

    wb = Workbook()
    wb.remove(wb.active)

    # ═══════════════════════════════════════════════
    # SHEET: Master (with PUC formulas)
    # ═══════════════════════════════════════════════
    ws_master = wb.create_sheet('Master', 0)

    puc_lookup = {}
    if not df_puc.empty:
        for _, r in df_puc.iterrows():
            p_name = str(r['Place']).strip().upper()
            if ',' in p_name: p_name = p_name.split(',')[0].strip()
            puc_lookup[p_name] = {
                '20GP': r.get('20GP', 0), '40GP': r.get('40GP', 0),
                '40HQ': r.get('40HQ', 0), '45HQ': r.get('45HQ', 0)
            }

    headers = list(df_master.columns)
    for col_idx, header in enumerate(headers, 1):
        ws_master.cell(1, col_idx, header)

    container_cols = [col for col in headers if any(x in str(col) for x in ['20', '40', '45'])]
    container_col_indices = {col: headers.index(col) + 1 for col in container_cols}
    place_col_idx = headers.index('Place') + 1

    for row_idx, (_, row) in enumerate(df_master.iterrows(), 2):
        is_soc = 'SOC' in str(row.get('Note', '')).upper()
        for col_idx, (col_name, value) in enumerate(row.items(), 1):
            if col_name not in container_cols:
                ws_master.cell(row_idx, col_idx, value)
        for cont_col, col_idx in container_col_indices.items():
            base_value = row.get(cont_col, 0)
            if is_soc and pd.notna(base_value) and base_value != 0:
                clean_base = base_value
                puc_col_map = {'20GP': 2, '40GP': 3, '40HQ': 4, '45HQ': 5}
                puc_col_num = puc_col_map.get(cont_col, 2)
                place_ref = f'${get_column_letter(place_col_idx)}{row_idx}'
                normalized_place = f'IFERROR(TRIM(LEFT({place_ref},FIND(",",{place_ref})-1)),{place_ref})'
                formula = f'={clean_base}+IFERROR(VLOOKUP({normalized_place},PUC!$A:$E,{puc_col_num},FALSE),0)'
                ws_master.cell(row_idx, col_idx, formula)
            else:
                ws_master.cell(row_idx, col_idx, base_value)

    _style_header(ws_master)
    _style_data(ws_master)
    _format_dates(ws_master)
    _autofit_columns(ws_master)
    _apply_table(ws_master, 'Master', 'TableStyleMedium2')
    ws_master.freeze_panes = 'A2'

    # ═══════════════════════════════════════════════
    # SHEET: Basic Cost
    # ═══════════════════════════════════════════════
    if not df_history_wide.empty:
        ws_bc = wb.create_sheet('Basic Cost')
        for col_idx, header in enumerate(df_history_wide.columns, 1):
            ws_bc.cell(1, col_idx, header)
        for row_idx, (_, row) in enumerate(df_history_wide.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                ws_bc.cell(row_idx, col_idx, value)
        _style_header(ws_bc)
        _style_data(ws_bc)
        _format_dates(ws_bc)
        _autofit_columns(ws_bc)
        _apply_table(ws_bc, 'BasicCost', 'TableStyleMedium2')
        ws_bc.freeze_panes = 'A2'

    # ═══════════════════════════════════════════════
    # SHEET: PUC
    # ═══════════════════════════════════════════════
    if not df_puc.empty:
        ws_puc = wb.create_sheet('PUC')
        for col_idx, header in enumerate(df_puc.columns, 1):
            ws_puc.cell(1, col_idx, header)
        for row_idx, (_, row) in enumerate(df_puc.iterrows(), 2):
            for col_idx, value in enumerate(row, 1):
                ws_puc.cell(row_idx, col_idx, value)
        _style_header(ws_puc)
        _style_data(ws_puc)
        _autofit_columns(ws_puc)
        _apply_table(ws_puc, 'PUC', 'TableStyleMedium9')
        ws_puc.freeze_panes = 'A2'

    # ═══════════════════════════════════════════════
    # REFERENCE SHEETS (from FAK)
    # ═══════════════════════════════════════════════
    # These have raw headers in row 1 from the FAK source
    # Detect header row and apply table from there
    REF_STYLES = {
        'Handling Fee': 'TableStyleMedium3',
        'ONE Group Code': 'TableStyleMedium4',
        'Wharfage': 'TableStyleMedium5',
        'GOH': 'TableStyleMedium6',
        'Hawaii': 'TableStyleMedium7',
    }
    for ref_name, ref_df in ref_sheet_data.items():
        ws_ref = wb.create_sheet(ref_name)
        # Write raw data (header=None, so row 0 of df = row 1 in Excel)
        for row_idx, (_, row) in enumerate(ref_df.iterrows(), 1):
            for col_idx, value in enumerate(row, 1):
                ws_ref.cell(row_idx, col_idx, value)
        # Style first row as header
        for cell in ws_ref[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
        # Style data rows
        for row in ws_ref.iter_rows(min_row=2):
            for cell in row:
                cell.font = DATA_FONT
        _autofit_columns(ws_ref)
        ws_ref.freeze_panes = 'A2'

    # ═══════════════════════════════════════════════
    # SHEET: Version
    # ═══════════════════════════════════════════════
    ws_ver = wb.create_sheet('Version')
    ws_ver.column_dimensions['A'].width = 30
    ws_ver.column_dimensions['B'].width = 55
    for r_idx, (label, value) in enumerate(version_rows, 1):
        cell_a = ws_ver.cell(r_idx, 1, label)
        cell_b = ws_ver.cell(r_idx, 2, value)
        if r_idx <= 3:
            cell_a.font = HEADER_FONT
            cell_a.fill = HEADER_FILL
            cell_a.alignment = Alignment(horizontal='left', vertical='center')
            cell_b.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color='1F4E79')
        elif label.startswith('RAW'):
            cell_a.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
            cell_b.font = DATA_FONT
        else:
            cell_a.font = DATA_FONT
            cell_b.font = DATA_FONT

    # ═══════════════════════════════════════════════
    # SAVE
    # ═══════════════════════════════════════════════
    wb.save(OUTPUT_MASTER)

    print(f"\n[SUCCESS] MasterFullPricing created!")
    print(f"  → Master: {len(df_master):,} routes (Table + Filter)")
    print(f"  → Basic Cost: {len(df_history_wide):,} records (Table + Filter)")
    print(f"  → PUC: {len(df_puc):,} entries (Table + Filter)")
    if ref_sheet_data:
        print(f"  → Reference sheets: {', '.join(ref_sheet_data.keys())}")
    print(f"  → All sheets: freeze panes, auto-fit columns, Arial {FONT_SIZE}pt")


if __name__ == "__main__":
    create_master_v13()