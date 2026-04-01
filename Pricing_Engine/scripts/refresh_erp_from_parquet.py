# -*- coding: utf-8 -*-
"""
refresh_erp_from_parquet.py — Direct Parquet → ERP Pipeline
Replaces: create_master_dashboard.py + refresh_erp_data.py (sync_erp.py)
Reads Parquet directly, writes all data into ERP_Master.xlsm.
"""
import os, sys, re, json
import datetime as _dt
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)  # Pricing_Engine
DATA_DIR = os.path.join(BASE_DIR, "data")
ERP_DIR = os.path.join(BASE_DIR, "..", "ERP", "data")
PARQUET_FILE = os.path.join(DATA_DIR, "Cleaned_Master_History.parquet")
ERP_FILE = os.path.join(ERP_DIR, "ERP_Master.xlsm")
PORT_MAP_FILE = os.path.join(DATA_DIR, "Port_Code_Mapping_Final.xlsx")
PUC_SOC_FILE = os.path.join(DATA_DIR, "PUC_SOC.xlsx")
CONFIG_FILE = os.path.join(BASE_DIR, 'config', 'pipeline_rules.json')

SHEET_NAME = "Pricing Dashboard"
MARKUP_STORE = "Markup_Store"

# ── ERP Layout Constants (V13 Ribbon Layout) ─────────────────────────
CONT_NAMES = ['20GP', '40GP', '40HQ', '45HQ', '40NOR', '20RF', '40RF']
PRICE_SOURCE = CONT_NAMES
CARRIERS = ["CMA", "COSCO", "EMC", "HPL", "MSC", "ONE", "WHL", "YML"]
CARRIER_DEFAULTS = {"HPL": 75, "WHL": 50}
DATA_START_ROW = 2   # V13: data starts row 2 (row 1 = headers)
DATA_HEADER_ROW = 1  # V13: row 1 = headers
COL_FIRST_PRICE = 10 # J = first price column

# ══════════════════════════════════════════════════════════════════════
# STEP 1: DATA PROCESSING
# ══════════════════════════════════════════════════════════════════════

sys.path.insert(0, SCRIPT_DIR)
from create_master_dashboard import (
    normalize_notes, normalize_text_data, load_mapping_files,
    shorten_source_file, apply_one_group_codes, create_puc_lookup_table
)

def _safe_val(val, default=0):
    if isinstance(val, pd.Series):
        return val.iloc[0] if not val.empty else default
    return val if pd.notna(val) else default

def load_and_process_parquet():
    """Load parquet, normalize, pivot to Master wide format + BasicCost."""
    print(f"\n{'='*60}")
    print(f"  DIRECT PARQUET → ERP REFRESH")
    print(f"{'='*60}")

    if not os.path.exists(PARQUET_FILE):
        print(f"❌ Parquet not found: {PARQUET_FILE}")
        sys.exit(1)

    df_all = pd.read_parquet(PARQUET_FILE)
    print(f"[1/6] Parquet loaded: {len(df_all):,} rows")

    port_map = load_mapping_files()
    df_all = normalize_text_data(df_all, port_map)

    df_all.loc[df_all['Source'].str.upper().str.contains("SCFI", na=False), 'Rate_Type'] = 'SCFI'
    df_all.loc[df_all['Source'].str.upper().str.contains("FIX", na=False), 'Rate_Type'] = 'FIX'
    df_all.loc[df_all['Source'].str.upper().str.contains("FAK", na=False), 'Rate_Type'] = 'FAK'
    df_all.loc[df_all['Source'].str.upper().str.contains("OCR", na=False), 'Rate_Type'] = 'OCR'

    mask_scfi_fix = df_all['Rate_Type'].isin(['SCFI', 'FIX', 'OCR'])
    df_all.loc[mask_scfi_fix, 'Charge_Name'] = "Base Ocean Freight"
    mask_fak = df_all['Rate_Type'] == 'FAK'
    df_all = normalize_notes(df_all)
    df_all['Source'] = df_all['Source'].apply(shorten_source_file)

    # ── Master Sheet (Base Ocean Freight, 30-day filter) ──
    print("[2/6] Creating Master pivot (Optimized - Active Only)...")
    today = pd.Timestamp(_dt.date.today())
    grace_date = today - pd.Timedelta(days=7)
    
    df_current = df_all[df_all['Rate_Type'].isin(['FAK', 'SCFI', 'FIX', 'SPECIAL', 'OCR'])].copy()
    df_current['Exp'] = pd.to_datetime(df_current['Exp'], errors='coerce')
    df_current['Eff'] = pd.to_datetime(df_current['Eff'], errors='coerce')
    
    # Apply grace period for SCFI/FIX/SPECIAL, strict today for others
    is_flex = df_current['Rate_Type'].isin(['SCFI', 'FIX', 'SPECIAL'])
    
    # [YEAR CORRECTION] If it's SCFI and year is 2025, it's likely a template error, fix to 2026
    mask_2025 = (df_current['Rate_Type'] == 'SCFI') & (df_current['Exp'].dt.year == 2025)
    if mask_2025.any():
        print(f"    → [YEAR CORRECTION] Incrementing year for {mask_2025.sum()} SCFI rows (2025 -> 2026)")
        df_current.loc[mask_2025, 'Exp'] = df_current.loc[mask_2025, 'Exp'] + pd.offsets.DateOffset(years=1)
        df_current.loc[mask_2025, 'Eff'] = df_current.loc[mask_2025, 'Eff'] + pd.offsets.DateOffset(years=1)

    df_current = df_current[
        (is_flex & (df_current['Exp'] >= grace_date)) |
        (~is_flex & (df_current['Exp'] >= today))
    ]
    
    # Preserve "All-In" flag BEFORE mapping Charge_Name
    df_current['Is_All_In'] = df_current['Charge_Name'].str.contains('ALL IN|Total', case=False, na=False)
    
    mask_scfi_fix = df_current['Rate_Type'].isin(['SCFI', 'FIX', 'OCR'])
    df_current.loc[mask_scfi_fix, 'Charge_Name'] = "Base Ocean Freight"
    mask_fak = df_current['Rate_Type'] == 'FAK'
    mask_base_fak = df_current['Charge_Name'].str.contains('ALL IN|Base|Basic|O/F|Ocean', case=False, na=False)
    df_current.loc[mask_fak & mask_base_fak, 'Charge_Name'] = "Base Ocean Freight"
    
    df_master_raw = df_current[df_current['Charge_Name'] == "Base Ocean Freight"].copy()
    sort_cols = ['Exp', 'Source', 'Amount']
    asc_flags = [False, False, True]
    df_master_raw = df_master_raw.sort_values(sort_cols, ascending=asc_flags)
    dedup_subset = ['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Note', 'Container_Type']
    df_master_raw = df_master_raw.drop_duplicates(subset=dedup_subset, keep='first')

    df_master = df_master_raw.pivot_table(
        index=['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note', 'Source', 'Is_All_In'],
        columns='Container_Type', values='Amount', aggfunc='max'
    ).reset_index()

    if "45'HQ" in df_master.columns: df_master = df_master.rename(columns={"45'HQ": "45HQ"})
    desired_order = ['20GP', '40GP', '40HQ', '45HQ', '40NOR', '20RF', '40RF']
    existing_containers = [c for c in desired_order if c in df_master.columns]
    df_master = df_master[['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note', 'Source', 'Is_All_In'] + existing_containers]
    print(f"    → Master: {len(df_master):,} rows")

    print("[3/6] Creating BasicCost breakdown...")
    df_recent = df_current.copy()
    df_recent = apply_one_group_codes(df_recent)
    charge_mapping = {
        'EMF': 'EIC/GFS/BAF/FDI', 'DLF': 'PCS', 'ISPS': 'ISPS/LSF/CMC',
        'BASIC O/F': 'BASIC O/F', 'ALL IN COST': 'BASIC O/F',
        'Base Ocean Freight': 'BASIC O/F', 'COMMISSION': 'HANDLING FEE FOR CARRIER'
    }
    df_recent['Charge_Name_Mapped'] = df_recent['Charge_Name'].replace(charge_mapping)
    df_history_wide = pd.DataFrame()
    if not df_recent.empty:
        df_recent['Charge_Container'] = df_recent['Charge_Name_Mapped'] + '_' + df_recent['Container_Type']
        idx_cols = ['Rate_Type', 'Source', 'POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Group Rate', 'Eff', 'Exp', 'Contract', 'Note']
        v_idx = [c for c in idx_cols if c in df_recent.columns]
        df_history_wide = df_recent.pivot_table(index=v_idx, columns='Charge_Container', values='Amount', aggfunc='max').reset_index()
    print(f"    → BasicCost: {len(df_history_wide):,} rows")

    print("[4/6] Creating PUC lookup...")
    df_puc = create_puc_lookup_table()
    print(f"    → PUC: {len(df_puc):,} rows")

    print("[5/6] Extracting FAK reference sheets...")
    REF_SHEETS = {'HANDLING FEE CARRIER': 'Handling Fee', 'ONE GROUP CODE': 'ONE Group Code', 'WHARFAGE': 'Wharfage', 'GOH': 'GOH', 'HAWAII': 'Hawaii'}
    ref_sheet_data = {}
    fak_file = None
    for f in os.listdir(DATA_DIR):
        if f.endswith('.xlsx') and 'FAK' in f.upper() and not f.startswith('~$'):
            fak_file = os.path.join(DATA_DIR, f)
            break
    if fak_file:
        try:
            fak_xls = pd.ExcelFile(fak_file)
            for src, dst in REF_SHEETS.items():
                if src in fak_xls.sheet_names:
                    ref_sheet_data[dst] = pd.read_excel(fak_xls, sheet_name=src, header=None)
                    print(f"    → {dst}: {ref_sheet_data[dst].shape[0]} rows")
        except Exception as e: print(f"    ⚠️ Error: {e}")

    version_rows = [('Tổng số dòng (Master)', len(df_master)), ('Ngày normalize', _dt.date.today().strftime('%d-%b-%Y')), ('Pipeline', 'Direct Parquet → ERP (Optimized)')]
    return df_master, df_history_wide, df_puc, ref_sheet_data, version_rows, df_current

# ══════════════════════════════════════════════════════════════════════
# STEP 2: WRITE TO ERP
# ══════════════════════════════════════════════════════════════════════

def read_markup_store(wb):
    store = {}
    if MARKUP_STORE in wb.sheetnames:
        ws = wb[MARKUP_STORE]
        for r in range(2, ws.max_row + 1):
            c = ws.cell(r, 1).value
            if c: store[str(c).strip()] = [ws.cell(r, i).value or 0 for i in range(2, 2 + len(CONT_NAMES))]
    return store

def write_markup_store(wb, store):
    if MARKUP_STORE in wb.sheetnames: wb.remove(wb[MARKUP_STORE])
    ws = wb.create_sheet(MARKUP_STORE)
    ws.sheet_state = 'hidden'
    ws.cell(1,1).value = "Carrier"
    for i, c in enumerate(CONT_NAMES): ws.cell(1, 2 + i).value = c
    ws.cell(2,1).value = "ALL"
    for i, v in enumerate(store.get("ALL", [0]*7)): ws.cell(2, 2+i).value = int(v)
    for ci, carrier in enumerate(CARRIERS):
        r = 3 + ci
        ws.cell(r, 1).value = carrier
        vals = store.get(carrier, [CARRIER_DEFAULTS.get(carrier, 0)] * 7)
        for i, v in enumerate(vals): ws.cell(r, 2+i).value = int(v)

def write_to_erp(df_master, df_history_wide, df_puc, ref_sheet_data, version_rows, df_all_charges):
    """Write all data to ERP_Master.xlsm with V13 layout and Ribbon restoration."""
    print(f"\n{'='*60}\n   STEP 2: WRITING TO ERP\n{'='*60}")
    if not os.path.exists(ERP_FILE): sys.exit(1)

    wb = openpyxl.load_workbook(ERP_FILE, keep_vba=True)
    existing_markups = read_markup_store(wb)

    if "PUC_Lookup" in wb.sheetnames: wb.remove(wb["PUC_Lookup"])
    ws_puc = wb.create_sheet("PUC_Lookup")
    ws_puc.sheet_state = 'hidden'
    for ci, h in enumerate(['Place', '20GP', '40GP', '40HQ', '45HQ'], 1): ws_puc.cell(1, ci).value = h
    for ri, (_, r) in enumerate(df_puc.iterrows(), 2):
        ws_puc.cell(ri, 1).value = r['Place']
        for ci, col in enumerate(['20GP', '40GP', '40HQ', '45HQ'], 2): ws_puc.cell(ri, ci).value = r.get(col, 0)

    puc_override = {}
    if not df_puc.empty:
        # Filter for active PUC rates only and sort by latest expiration
        today = datetime.now()
        df_puc_active = df_puc.copy()
        df_puc_active['Expiration Date'] = pd.to_datetime(df_puc_active.get('Expiration Date'), errors='coerce')
        df_puc_active = df_puc_active[df_puc_active['Expiration Date'] >= today].sort_values('Expiration Date', ascending=False)
        
        # Determine the place/destination column name (Destination in the file)
        place_col = 'Destination' if 'Destination' in df_puc_active.columns else 'Place'
        
        for _, pr in df_puc_active.iterrows():
            rt = str(pr.get(place_col, '')).strip().upper()
            if rt and rt not in puc_override:
                puc_override[rt] = {k: float(pr.get(k, 0)) for k in ['20GP', '40GP', '40HQ', '45HQ']}

    print("   📋 Building BasicCost_Lookup...")
    if "BasicCost_Lookup" in wb.sheetnames: wb.remove(wb["BasicCost_Lookup"])
    ws_bc = wb.create_sheet("BasicCost_Lookup")
    ws_bc.sheet_state = 'hidden'
    for ci, h in enumerate(['Key', 'Contract', 'GroupRate', 'CostBreakdown', 'TotalCharge'], 1): ws_bc.cell(1, ci).value = h

    charge_groups = [("O/F", "BASIC O/F"), ("O/F", "Total Ocean Freight"), ("O/F", "Base Ocean Freight"), ("ARB", "ARB/OLF"), ("ISPS", "ISPS/LSF/CMC"), ("PSS/PUC", "PSS/PUC"), ("OCS/LSS", "OCS/LSS/EFF/ITC/GFS/ SOC COST HDL FEE"), ("PCS/ACS", "PCS/ACS/AGS"), ("GRI", "GRI"), ("EIC/BAF", "EIC/GFS/BAF/FDI"), ("WHA/BCO", "WHA/BCO/BCD/CFC/EIC"), ("GARMENT", "GARMENT ADD ON"), ("PREMIUM", "PREMIUM ADD ON/HDL FEE US FOR SOC")]
    bc_row_idx = 2
    seen_keys = set()
    if not df_history_wide.empty:
        for _, row in df_history_wide.iterrows():
            pol, pod, bc_place, carrier = str(row.get('POL','')), str(row.get('POD','')), str(row.get('Place','')), str(row.get('Carrier',''))
            contract, group_rate = str(row.get('Contract','')), str(row.get('Group Rate',''))
            note_u = str(row.get('Note', '')).strip().upper()
            if not pol or not carrier: continue

            for cont in CONT_NAMES:
                key = f"{pol}|{pod}|{bc_place}|{carrier}|{cont}|{note_u}"
                if key in seen_keys: continue
                has_base = any(pd.notna(row.get(f"{b}_{cont}")) and row.get(f"{b}_{cont}") != 0 for b in ["BASIC O/F", "Total Ocean Freight", "Base Ocean Freight"])
                if not has_base: continue
                seen_keys.add(key)
                
                comps = {}
                for sn, pref in charge_groups:
                    v = _safe_val(row.get(f"{pref}_{cont}"), 0)
                    if v != 0: comps[sn] = max(comps.get(sn, 0), float(v))

                if ("SOC" in note_u) and puc_override:
                    for rn, rv in puc_override.items():
                        if rn in bc_place.upper() or bc_place.upper() in rn:
                            ps = rv.get(cont, 0)
                            if ps != 0: comps["PSS/PUC"] = ps
                            break

                t_val = _safe_val(row.get(f"Total Ocean Freight_{cont}", 0))
                cost_parts = []
                if t_val > 0:
                    basic_of = t_val - sum(v for k, v in comps.items() if k != "O/F")
                    if basic_of > 0: cost_parts.append(f"O/F ${basic_of:,.0f}")
                    final_buy = t_val
                else:
                    basic_of = comps.get("O/F", 0)
                    if basic_of > 0: cost_parts.append(f"O/F ${basic_of:,.0f}")
                    final_buy = basic_of + sum(v for k, v in comps.items() if k != "O/F")

                for k, v in comps.items():
                    if k != "O/F" and v != 0: cost_parts.append(f"{k} ${v:,.0f}")

                h_val = _safe_val(row.get(f"HANDLING FEE FOR CARRIER_{cont}"), 0)
                h_str = f"\nHDL FEE: ${float(h_val):,.0f}" if h_val != 0 else ""
                breakdown = f"BKG: S/C={contract} | Group={group_rate}\nCOST: {' + '.join(cost_parts)}{h_str}"

                ws_bc.cell(bc_row_idx, 1).value, ws_bc.cell(bc_row_idx, 2).value, ws_bc.cell(bc_row_idx, 3).value, ws_bc.cell(bc_row_idx, 4).value, ws_bc.cell(bc_row_idx, 5).value = key, contract, group_rate, breakdown, final_buy
                bc_row_idx += 1

    unique_pols = sorted(df_master['POL'].dropna().unique().tolist())
    unique_pods = sorted(df_master['POD'].dropna().unique().tolist())
    unique_places = sorted(df_master['Place'].dropna().unique().tolist())
    if "Search_Lists" in wb.sheetnames: wb.remove(wb["Search_Lists"])
    ws_search = wb.create_sheet("Search_Lists")
    ws_search.sheet_state = 'hidden'
    for c, items in enumerate([unique_pols, unique_pods, unique_places], 1):
        ws_search.cell(1, c).value = ["POL","POD","Place"][c-1]
        for i, v in enumerate(items, 2): ws_search.cell(i, c).value = v

    print("   📊 Writing Pricing Dashboard (V13 Optimized)...")
    ws = wb[SHEET_NAME]
    for m_range in list(ws.merged_cells.ranges): ws.unmerge_cells(str(m_range))
    for r in range(1, max(ws.max_row, 50) + 1):
        for c in range(1, 17):
            cell = ws.cell(r, c)
            cell.value, cell.fill, cell.font, cell.border = None, PatternFill(), Font(), Border()
    ws.auto_filter.ref = None

    C_SEARCH, C_DRY = PatternFill("solid", fgColor="FFF7ED"), PatternFill("solid", fgColor="1E40AF")
    thin = Side(style='thin', color='CBD5E1')
    card_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[1].height = 30
    h_labels = ['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note', 'Source']
    for i, lbl in enumerate(h_labels, 1):
        cell = ws.cell(1, i, value=lbl)
        cell.font, cell.fill, cell.border, cell.alignment = Font(name='Segoe UI', size=9, color="B0B0B0", italic=True), C_SEARCH, card_border, Alignment(horizontal='center', vertical='center')
    for i, ct in enumerate(CONT_NAMES):
        cell = ws.cell(1, COL_FIRST_PRICE + i, value=ct)
        cell.font, cell.fill, cell.border, cell.alignment = Font(name='Segoe UI', size=9, bold=True, color="FFFFFF"), C_DRY, card_border, Alignment(horizontal='center', vertical='center')

    data_font = Font(name='Segoe UI', size=10)
    def _write_date(ws, r, c, v):
        cell = ws.cell(r, c)
        if pd.notna(v) and v != "":
            try:
                cell.value = pd.to_datetime(v)
                cell.number_format = 'DD-MMM'
            except: cell.value = str(v)
        cell.font, cell.border, cell.alignment = data_font, card_border, Alignment(horizontal='center')

    existing_p_map = {}
    if not df_all_charges.empty:
        p_rows = df_all_charges[df_all_charges['Charge_Name'].str.contains('PUC|PSS', case=False, na=False)]
        for _, r in p_rows.iterrows():
            pk = (r['POL'], r['POD'], r.get('Place',''), r['Carrier'], r.get('Commodity',''), r.get('Note',''), r['Container_Type'])
            existing_p_map[pk] = max(existing_p_map.get(pk, 0), float(r['Amount']))

    for r_idx, (_, row) in enumerate(df_master.iterrows(), start=2):
        pol, pod, place, carrier, comm = row['POL'], row['POD'], row['Place'], row['Carrier'], row.get('Commodity', '')
        note_orig = row.get('Note', '')
        is_soc = ("SOC" in str(note_orig).upper())
        is_all_in = row.get('Is_All_In', False)
        ws.cell(r_idx, 1).value, ws.cell(r_idx, 2).value, ws.cell(r_idx, 3).value, ws.cell(r_idx, 4).value, ws.cell(r_idx, 5).value = pol, pod, place, carrier, comm
        _write_date(ws, r_idx, 6, row.get('Eff')); _write_date(ws, r_idx, 7, row.get('Exp'))
        ws.cell(r_idx, 8).value, ws.cell(r_idx, 9).value = note_orig, row.get('Source', '')

        p_truth = {}
        if is_soc and puc_override:
            p_u = str(place).strip().upper()
            # SOC Route Synonym: LAX/LGB <=> Los Angeles
            p_match_keys = [p_u]
            if p_u == "LAX/LGB": p_match_keys.append("LOS ANGELES")
            
            for rn in p_match_keys:
                for kn, rv in puc_override.items():
                    if rn in kn or kn in rn:
                        p_truth = rv
                        break
                if p_truth: break

        for ci, ct in enumerate(CONT_NAMES):
            val = _safe_val(row.get(ct), 0)
            if val != 0:
                f_v = float(val)
                if is_soc:
                    # ONLY subtract existing PUC if the row was an "All-In" row
                    if is_all_in:
                        old_p = existing_p_map.get((pol, pod, place, carrier, comm, note_orig, ct), 0)
                        if old_p > 0: f_v -= old_p
                    # Add Truth PUC
                    f_v += p_truth.get(ct, 0)
                
                if pd.notna(f_v):
                    cell = ws.cell(r_idx, COL_FIRST_PRICE + ci, value=int(f_v))
                    cell.number_format, cell.font, cell.border = '#,##0', data_font, card_border

    end_row = DATA_START_ROW + len(df_master) - 1
    ws.auto_filter.ref = f"A1:P{end_row}"
    ws.freeze_panes = "A2"

    if "Version" in wb.sheetnames: wb.remove(wb["Version"])
    ws_v = wb.create_sheet("Version"); ws_v.sheet_state = 'hidden'
    for ri, (l, v) in enumerate(version_rows, 1): ws_v.cell(ri, 1).value, ws_v.cell(ri, 2).value = l, v

    print("   💾 Saving...")
    wb.save(ERP_FILE)
    print("   🎗️ Restoring Ribbon XML...")
    try:
        sys.path.insert(0, os.path.join(os.path.dirname(BASE_DIR), "ERP", "core"))
        from customui_utils import ensure_customui
        ensure_customui(ERP_FILE, customui_xml_path=os.path.join(os.path.dirname(BASE_DIR), "ERP", "vba", "CustomUI_ERP.xml"))
        print("   ✅ CustomUI14 Ribbon restored")
    except: pass
    print(f"\n✅ DIRECT PARQUET → ERP REFRESH COMPLETE!")

if __name__ == "__main__":
    try:
        res = load_and_process_parquet()
        write_to_erp(*res)
    except Exception as e:
        print(f"❌ Error: {e}")
