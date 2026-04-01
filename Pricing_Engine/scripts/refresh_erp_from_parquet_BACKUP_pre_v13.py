# -*- coding: utf-8 -*-
"""
refresh_erp_from_parquet.py — Direct Parquet → ERP Pipeline
Replaces: create_master_dashboard.py + refresh_erp_data.py (sync_erp.py)
Reads Parquet directly, writes all data into ERP_Master.xlsm.
"""
import os, sys, re, json
import datetime as _dt
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Paths ─────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)  # Pricing_Engine
DATA_DIR = os.path.join(BASE_DIR, "data")
ERP_DIR = os.path.join(BASE_DIR, "..", "ERP", "data")
PARQUET_FILE = os.path.join(DATA_DIR, "Cleaned_Master_History.parquet")
ERP_FILE = os.path.join(ERP_DIR, "ERP_Master.xlsm")
PORT_MAP_FILE = os.path.join(DATA_DIR, "Port_Code_Mapping_Final.xlsx")
PUC_SOC_FILE = os.path.join(DATA_DIR, "PUC_SOC.xlsx")
CONFIG_FILE = os.path.join(BASE_DIR, 'config', 'pipeline_rules.json')

SHEET_NAME = "Pricing Dashboard"
MARKUP_STORE = "Markup_Store"

# ── ERP Layout Constants ──────────────────────────────────────────────
CONT_NAMES = ['20GP', '40GP', '40HQ', '45HQ', '40NOR', '20RF', '40RF']
PRICE_SOURCE = CONT_NAMES
CARRIERS = ["CMA", "COSCO", "EMC", "HPL", "MSC", "ONE", "WHL", "YML"]
CARRIER_DEFAULTS = {"HPL": 75, "WHL": 50}
HEADER_COLOR = "475569"
QUOTE_COLOR = "F8FAFC"
INPUT_COLOR = "EFF6FF"
MARKUP_BG = "F0FDF4"
SEARCH_BG = "FFF7ED"
ACTIVE_BG = "ECFDF5"
DATA_START_ROW = 9
DATA_HEADER_ROW = 8
DATA_TITLE_ROW = 7
HIDDEN_COL_START = 24
SEARCH_POL_COL, SEARCH_POD_COL, SEARCH_PLACE_COL = 1, 2, 3
PUC_COL, PUC_ROW = 4, 2
MARKUP_CARRIER_COL, MARKUP_FIRST_COL, MARKUP_ACTIVE_COL = 9, 10, 17
COL_POL, COL_POD, COL_PLACE, COL_CARRIER = 1, 2, 3, 4
COL_COMMODITY, COL_EFF, COL_EXP, COL_NOTE, COL_SOURCE = 5, 6, 7, 8, 9
COL_FIRST_PRICE = 10

# ══════════════════════════════════════════════════════════════════════
# STEP 1: DATA PROCESSING (from create_master_dashboard.py)
# ══════════════════════════════════════════════════════════════════════

# Import normalization functions from create_master_dashboard
sys.path.insert(0, SCRIPT_DIR)
from create_master_dashboard import (
    normalize_notes, normalize_text_data, load_mapping_files,
    shorten_source_file, apply_one_group_codes, create_puc_lookup_table
)

def _safe_val(val, default=0):
    if isinstance(val, pd.Series):
        return val.iloc[0] if not val.empty else default
    return val if pd.notna(val) else default


def load_and_process_parquet():
    """Load parquet, normalize, pivot to Master wide format + BasicCost."""
    print(f"\n{'='*60}")
    print(f"  DIRECT PARQUET → ERP REFRESH")
    print(f"{'='*60}")

    if not os.path.exists(PARQUET_FILE):
        print(f"❌ Parquet not found: {PARQUET_FILE}")
        sys.exit(1)

    df_all = pd.read_parquet(PARQUET_FILE)
    print(f"[1/6] Parquet loaded: {len(df_all):,} rows")

    # Normalize
    port_map = load_mapping_files()
    df_all = normalize_text_data(df_all, port_map)

    # Re-identify Rate_Type
    df_all.loc[df_all['Source'].str.upper().str.contains("SCFI", na=False), 'Rate_Type'] = 'SCFI'
    df_all.loc[df_all['Source'].str.upper().str.contains("FIX", na=False), 'Rate_Type'] = 'FIX'
    df_all.loc[df_all['Source'].str.upper().str.contains("FAK", na=False), 'Rate_Type'] = 'FAK'
    df_all.loc[df_all['Source'].str.upper().str.contains("OCR", na=False), 'Rate_Type'] = 'OCR'

    # Normalize charges
    mask_scfi_fix = df_all['Rate_Type'].isin(['SCFI', 'FIX', 'OCR'])
    df_all.loc[mask_scfi_fix, 'Charge_Name'] = "Base Ocean Freight"
    mask_fak = df_all['Rate_Type'] == 'FAK'
    mask_base_fak = df_all['Charge_Name'].str.contains('ALL IN|Base|Basic|O/F|Ocean', case=False, na=False)
    df_all.loc[mask_fak & mask_base_fak, 'Charge_Name'] = "Base Ocean Freight"

    df_all = normalize_notes(df_all)
    df_all['Source'] = df_all['Source'].apply(shorten_source_file)

    # ── Master Sheet (Base Ocean Freight, 30-day filter) ──
    print("[2/6] Creating Master pivot...")
    cutoff_date = pd.Timestamp(_dt.date.today()) - pd.Timedelta(days=30)
    df_current = df_all[df_all['Rate_Type'].isin(['FAK', 'SCFI', 'FIX', 'SPECIAL', 'OCR'])].copy()
    df_current['Exp'] = pd.to_datetime(df_current['Exp'], errors='coerce')
    df_current = df_current[df_current['Exp'] >= cutoff_date]

    df_master_raw = df_current[df_current['Charge_Name'] == "Base Ocean Freight"].copy()
    df_master_raw = df_master_raw.sort_values('Amount', ascending=False)
    df_master_raw = df_master_raw.drop_duplicates(
        subset=['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note', 'Source', 'Container_Type'],
        keep='first'
    )

    df_master = df_master_raw.pivot_table(
        index=['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note', 'Source'],
        columns='Container_Type', values='Amount', aggfunc='max'
    ).reset_index()

    if "45'HQ" in df_master.columns:
        df_master = df_master.rename(columns={"45'HQ": "45HQ"})

    desired_order = ['20GP', '40GP', '40HQ', '45HQ', '40NOR', '20RF', '40RF']
    existing_containers = [c for c in desired_order if c in df_master.columns]
    df_master = df_master[['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note', 'Source'] + existing_containers]
    print(f"    → Master: {len(df_master):,} rows")

    # ── BasicCost (full charge breakdown) ──
    print("[3/6] Creating BasicCost breakdown...")
    df_recent = df_current.copy()
    df_recent = apply_one_group_codes(df_recent)

    charge_mapping = {
        'EMF': 'EIC/GFS/BAF/FDI', 'DLF': 'PCS', 'ISPS': 'ISPS/LSF/CMC',
        'BASIC O/F': 'BASIC O/F', 'ALL IN COST': 'BASIC O/F',
        'Base Ocean Freight': 'BASIC O/F', 'COMMISSION': 'HANDLING FEE FOR CARRIER'
    }
    df_recent['Charge_Name_Mapped'] = df_recent['Charge_Name'].replace(charge_mapping)

    df_history_wide = pd.DataFrame()
    if not df_recent.empty:
        df_recent['Charge_Container'] = df_recent['Charge_Name_Mapped'] + '_' + df_recent['Container_Type']
        index_cols = ['Rate_Type', 'Source', 'POL', 'POD', 'Place', 'Carrier', 'Commodity',
                      'Group Rate', 'Eff', 'Exp', 'Contract', 'Note']
        valid_idx = [c for c in index_cols if c in df_recent.columns]
        df_history_wide = df_recent.pivot_table(
            index=valid_idx, columns='Charge_Container', values='Amount', aggfunc='max'
        ).reset_index()
    print(f"    → BasicCost: {len(df_history_wide):,} rows")

    # ── PUC ──
    print("[4/6] Creating PUC lookup...")
    df_puc = create_puc_lookup_table()
    print(f"    → PUC: {len(df_puc):,} rows")

    # ── Reference sheets from FAK ──
    print("[5/6] Extracting FAK reference sheets...")
    REF_SHEETS = {
        'HANDLING FEE CARRIER': 'Handling Fee',
        'ONE GROUP CODE': 'ONE Group Code',
        'WHARFAGE': 'Wharfage', 'GOH': 'GOH', 'HAWAII': 'Hawaii',
    }
    ref_sheet_data = {}
    fak_file = None
    for f in os.listdir(DATA_DIR):
        if f.endswith('.xlsx') and 'FAK' in f.upper() and not f.startswith('~$'):
            fak_file = os.path.join(DATA_DIR, f)
            break
    if fak_file:
        try:
            fak_xls = pd.ExcelFile(fak_file)
            for src_name, dst_name in REF_SHEETS.items():
                if src_name in fak_xls.sheet_names:
                    ref_sheet_data[dst_name] = pd.read_excel(fak_xls, sheet_name=src_name, header=None)
                    print(f"    → {dst_name}: {ref_sheet_data[dst_name].shape[0]} rows")
        except Exception as e:
            print(f"    ⚠️ Error: {e}")

    # ── Version info ──
    EXCLUDE = ['PUC_SOC', 'Port_Code', 'Schedule', 'Master']
    raw_files = sorted([f for f in os.listdir(DATA_DIR)
                        if f.endswith('.xlsx') and not f.startswith('~$')
                        and not any(e in f for e in EXCLUDE)])
    fak_version = 'N/A'
    for rf in raw_files:
        if 'FAK' in rf.upper():
            m = re.search(r'(\d{1,2})\s+([A-Z]{3})\s+NO\.?\s*(\d+)', rf.upper())
            if m:
                fak_version = f"{m.group(1)}{m.group(2)}NO.{m.group(3)}"
            break

    version_rows = [
        ('Ấn bản bảng giá', fak_version),
        ('Tổng số dòng (Master)', len(df_master)),
        ('Ngày normalize', _dt.date.today().strftime('%d-%b-%Y')),
        ('Pipeline', 'Direct Parquet → ERP (no MasterFullPricing.xlsx)'),
    ]
    for rf in raw_files:
        fu = rf.upper()
        label = 'RAW: FAK' if 'FAK' in fu else 'RAW: SCFI' if 'SCFI' in fu else 'RAW: SPECIAL RATE' if 'SPECIAL' in fu or 'FIX' in fu else 'RAW'
        version_rows.append((label, re.sub(r'\.xlsx$', '', rf, flags=re.I)))

    return df_master, df_history_wide, df_puc, ref_sheet_data, version_rows


# ══════════════════════════════════════════════════════════════════════
# STEP 2: WRITE TO ERP (from refresh_erp_data.py)
# ══════════════════════════════════════════════════════════════════════

def _short_source(filename):
    if not filename or pd.isna(filename): return ""
    fname = str(filename).upper()
    if "FAK" in fname: return "FAK"
    elif "SCFI" in fname: return "SCFI"
    elif "FIX" in fname: return "FIX"
    elif "SPECIAL" in fname: return "SPECIAL"
    return str(filename)[:12]

def _short_date(date_val):
    if pd.isna(date_val): return ""
    try:
        dt = pd.to_datetime(date_val) if isinstance(date_val, str) else date_val
        return f"{dt.day}{dt.strftime('%b').upper()}"
    except: return str(date_val)

def read_markup_store(wb):
    store = {}
    if MARKUP_STORE in wb.sheetnames:
        ws = wb[MARKUP_STORE]
        for r in range(2, ws.max_row + 1):
            carrier = ws.cell(r, 1).value
            if carrier:
                vals = [ws.cell(r, c).value or 0 for c in range(2, 2 + len(CONT_NAMES))]
                store[str(carrier).strip()] = vals
    return store

def write_markup_store(wb, store):
    if MARKUP_STORE in wb.sheetnames:
        wb.remove(wb[MARKUP_STORE])
    ws = wb.create_sheet(MARKUP_STORE)
    ws.sheet_state = 'hidden'
    ws.cell(1, 1).value = "Carrier"
    for i, c in enumerate(CONT_NAMES):
        ws.cell(1, 2 + i).value = c
    ws.cell(2, 1).value = "ALL"
    for i, v in enumerate(store.get("ALL", [0]*len(CONT_NAMES))):
        ws.cell(2, 2 + i).value = int(v) if v else 0
    for ci, carrier in enumerate(CARRIERS):
        r = 3 + ci
        ws.cell(r, 1).value = carrier
        vals = store.get(carrier, [CARRIER_DEFAULTS.get(carrier, 0)] * len(CONT_NAMES))
        for i, v in enumerate(vals):
            ws.cell(r, 2 + i).value = int(v) if v else 0


def write_to_erp(df_master, df_history_wide, df_puc, ref_sheet_data, version_rows):
    """Write all data directly into ERP_Master.xlsm."""
    print(f"\n[6/6] Writing to ERP...")

    if not os.path.exists(ERP_FILE):
        print(f"❌ ERP not found: {ERP_FILE}")
        sys.exit(1)

    # Prep master data
    df_master['Source_Short'] = df_master['Source'].apply(_short_source)
    df_master['Eff_Short'] = df_master['Eff'].apply(_short_date)
    df_master['Exp_Short'] = df_master['Exp'].apply(_short_date)

    wb = openpyxl.load_workbook(ERP_FILE, keep_vba=True)
    print(f"   ✓ ERP opened ({len(wb.sheetnames)} sheets)")

    # ── Read existing markups ──
    existing_markups = read_markup_store(wb)
    if existing_markups:
        print(f"   ✓ Found {len(existing_markups)} saved markups")

    # ── Rebuild PUC_Lookup ──
    if not df_puc.empty:
        if "PUC_Lookup" in wb.sheetnames:
            wb.remove(wb["PUC_Lookup"])
        ws_puc = wb.create_sheet("PUC_Lookup")
        ws_puc.sheet_state = 'hidden'
        for ci, h in enumerate(['Place', '20GP', '40GP', '40HQ', '45HQ'], 1):
            ws_puc.cell(1, ci).value = h
        for ri, (_, row) in enumerate(df_puc.iterrows(), 2):
            ws_puc.cell(ri, 1).value = row['Place']
            ws_puc.cell(ri, 2).value = row.get('20GP', 0)
            ws_puc.cell(ri, 3).value = row.get('40GP', 0)
            ws_puc.cell(ri, 4).value = row.get('40HQ', 0)
            ws_puc.cell(ri, 5).value = row.get('45HQ', 0)
        puc_last_row = len(df_puc) + 1
    else:
        puc_last_row = 2

    # ── Build BasicCost_Lookup (from refresh_erp_data.py logic) ──
    print("   📋 Building BasicCost_Lookup...")
    if "BasicCost_Lookup" in wb.sheetnames:
        wb.remove(wb["BasicCost_Lookup"])
    ws_bc = wb.create_sheet("BasicCost_Lookup")
    ws_bc.sheet_state = 'hidden'
    for ci, h in enumerate(['Key', 'Contract', 'GroupRate', 'CostBreakdown', 'TotalCharge'], 1):
        ws_bc.cell(1, ci).value = h

    charge_groups = [
        ("O/F", "BASIC O/F"), ("O/F", "Total Ocean Freight"), ("O/F", "Base Ocean Freight"),
        ("ARB", "ARB/OLF"), ("ISPS", "ISPS/LSF/CMC"),
        ("PSS/PUC", "PSS/PUC"), ("OCS/LSS", "OCS/LSS/EFF/ITC/GFS/ SOC COST HDL FEE"),
        ("PCS/ACS", "PCS/ACS/AGS"), ("GRI", "GRI"), ("EIC/BAF", "EIC/GFS/BAF/FDI"),
        ("WHA/BCO", "WHA/BCO/BCD/CFC/EIC"), ("GARMENT", "GARMENT ADD ON"),
        ("PREMIUM", "PREMIUM ADD ON/HDL FEE US FOR SOC"),
    ]

    # PUC override dict
    puc_override = {}
    if not df_puc.empty:
        puc_cont_map = {'20GP': '20GP', '40GP': '40GP', '40HQ': '40HQ', '45HQ': '45HQ'}
        for _, puc_row in df_puc.iterrows():
            route = str(puc_row.get('Place', '')).strip()
            if route:
                puc_override[route] = {k: float(puc_row.get(v, 0)) if pd.notna(puc_row.get(v)) else 0
                                        for k, v in puc_cont_map.items()}

    bc_row_idx = 2
    seen_keys = set()
    if not df_history_wide.empty:
        df_bc = df_history_wide.copy()
        if 'Exp' in df_bc.columns:
            df_bc = df_bc.sort_values('Exp', ascending=False, na_position='last')
        
        for _, row in df_bc.iterrows():
            pol = str(row.get('POL', '')).strip()
            pod = str(row.get('POD', '')).strip()
            bc_place = str(row.get('Place', '')).strip() if pd.notna(row.get('Place', '')) else ''
            carrier = str(row.get('Carrier', '')).strip()
            contract = row.get('Contract', '')
            group_rate = row.get('Group Rate', '')
            note_val = str(row.get('Note', '')).strip().upper()

            if not pol or not carrier: continue

            for cont in CONT_NAMES:
                key = f"{pol}|{pod}|{bc_place}|{carrier}|{cont}|{note_val}" # Unique by SOC/COC too
                if key in seen_keys: continue
                
                # Check if we have any base price
                has_base = False
                for base_name in ["BASIC O/F", "Total Ocean Freight", "Base Ocean Freight"]:
                    if pd.notna(row.get(f"{base_name}_{cont}")) and row.get(f"{base_name}_{cont}") != 0:
                        has_base = True
                        break
                if not has_base: continue

                seen_keys.add(key)
                
                # Identify components
                components = {}
                for short_name, prefix in charge_groups:
                    val = _safe_val(row.get(f"{prefix}_{cont}"), 0)
                    if val != 0:
                        components[short_name] = max(components.get(short_name, 0), float(val))

                # Special PUC override for SOC
                if note_val == "SOC" and puc_override:
                    for rn, rv in puc_override.items():
                        if rn.upper() in bc_place.upper() or bc_place.upper() in rn.upper():
                            pv = rv.get(cont, 0)
                            if pv != 0: components["PSS/PUC"] = pv
                            break

                # --- 🧠 Intelligent Breakdown Logic ---
                # If we have "Total Ocean Freight" but need to show "O/F"
                # O/F = Total - Sum(all other surcharges found in Parquet)
                total_val = _safe_val(row.get(f"Total Ocean Freight_{cont}", 0))
                
                cost_parts = []
                final_total_buying = 0
                if total_val > 0:
                    # Calculate surcharges (excluding O/F)
                    surcharges_sum = sum(v for k, v in components.items() if k != "O/F")
                    basic_of = total_val - surcharges_sum
                    if basic_of > 0:
                        cost_parts.append(f"O/F ${basic_of:,.0f}")
                    final_total_buying = total_val
                else:
                    # Fallback to standard O/F if no Total is present
                    basic_of = components.get("O/F", 0)
                    if basic_of > 0:
                        cost_parts.append(f"O/F ${basic_of:,.0f}")
                    surcharges_sum = sum(v for k, v in components.items() if k != "O/F")
                    final_total_buying = basic_of + surcharges_sum

                # Add other components
                for k, v in components.items():
                    if k != "O/F" and v != 0:
                        cost_parts.append(f"{k} ${v:,.0f}")

                hdl_col = f"HANDLING FEE FOR CARRIER_{cont}"
                hdl_val = _safe_val(row.get(hdl_col), 0)
                hdl_str = f"\nHDL FEE: ${float(hdl_val):,.0f}" if hdl_val != 0 else ""
                
                contract_str = str(contract) if pd.notna(contract) else ""
                group_str = str(group_rate) if pd.notna(group_rate) and str(group_rate).strip() else "N/A"
                
                breakdown = f"BKG: S/C={contract_str} | Group={group_str}\n"
                if cost_parts: breakdown += f"COST: {' + '.join(cost_parts)}"
                if hdl_str: breakdown += hdl_str

                ws_bc.cell(bc_row_idx, 1).value = key
                ws_bc.cell(bc_row_idx, 2).value = contract_str
                ws_bc.cell(bc_row_idx, 3).value = group_str
                ws_bc.cell(bc_row_idx, 4).value = breakdown
                ws_bc.cell(bc_row_idx, 5).value = final_total_buying
                bc_row_idx += 1
    print(f"   ✓ {bc_row_idx - 2} BasicCost lookup entries")

    # ── Build Search_Lists ──
    unique_pols = sorted(df_master['POL'].dropna().unique().tolist())
    unique_pods = sorted(df_master['POD'].dropna().unique().tolist())
    unique_places = sorted(df_master['Place'].dropna().unique().tolist())
    if "Search_Lists" in wb.sheetnames:
        wb.remove(wb["Search_Lists"])
    ws_search = wb.create_sheet("Search_Lists")
    ws_search.sheet_state = 'hidden'
    ws_search.cell(1, 1).value = "POL"
    ws_search.cell(1, 2).value = "POD"
    ws_search.cell(1, 3).value = "Place"
    for i, v in enumerate(unique_pols, 2): ws_search.cell(i, 1).value = v
    for i, v in enumerate(unique_pods, 2): ws_search.cell(i, 2).value = v
    for i, v in enumerate(unique_places, 2): ws_search.cell(i, 3).value = v
    pol_last = len(unique_pols) + 1
    pod_last = len(unique_pods) + 1
    place_last = len(unique_places) + 1
    print(f"   ✓ Search lists: POL({len(unique_pols)}) POD({len(unique_pods)}) Place({len(unique_places)})")

    # ── Write Reference Sheets (hidden) ──
    print("   📋 Writing reference sheets...")
    HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
    DATA_FONT = Font(name='Arial', size=10)
    for ref_name, ref_df in ref_sheet_data.items():
        if ref_name in wb.sheetnames:
            wb.remove(wb[ref_name])
        ws_ref = wb.create_sheet(ref_name)
        ws_ref.sheet_state = 'hidden'
        for ri, (_, row) in enumerate(ref_df.iterrows(), 1):
            for ci, value in enumerate(row, 1):
                ws_ref.cell(ri, ci, value)
        for cell in ws_ref[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        ws_ref.freeze_panes = 'A2'
        print(f"      → {ref_name} ✓")

    # ── Write Version Sheet (hidden) ──
    if "Version" in wb.sheetnames:
        wb.remove(wb["Version"])
    ws_ver = wb.create_sheet("Version")
    ws_ver.sheet_state = 'hidden'
    ws_ver.column_dimensions['A'].width = 30
    ws_ver.column_dimensions['B'].width = 55
    for ri, (label, value) in enumerate(version_rows, 1):
        ws_ver.cell(ri, 1, label)
        ws_ver.cell(ri, 2, value)

    # ══════════════════════════════════════════════════════════════════
    # PRICING DASHBOARD (exact same layout as refresh_erp_data.py)
    # ══════════════════════════════════════════════════════════════════
    print("   📊 Writing Pricing Dashboard...")
    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ Sheet '{SHEET_NAME}' not found!")
        sys.exit(1)
    ws = wb[SHEET_NAME]

    for merge_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge_range))

    max_clear_row = max(ws.max_row, DATA_START_ROW + 10)
    for r in range(1, max_clear_row + 1):
        for c in range(1, 35):
            cell = ws.cell(r, c)
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font()
            cell.border = Border()

    if ws.tables:
        for t_name in list(ws.tables.keys()):
            del ws.tables[t_name]

    # ── Shared styles ──
    thin = Side(style='thin', color='CBD5E1')
    card_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    input_fill = PatternFill(start_color=INPUT_COLOR, end_color=INPUT_COLOR, fill_type="solid")
    quote_fill = PatternFill(start_color=QUOTE_COLOR, end_color=QUOTE_COLOR, fill_type="solid")
    markup_fill = PatternFill(start_color=MARKUP_BG, end_color=MARKUP_BG, fill_type="solid")
    search_fill = PatternFill(start_color=SEARCH_BG, end_color=SEARCH_BG, fill_type="solid")
    active_fill = PatternFill(start_color=ACTIVE_BG, end_color=ACTIVE_BG, fill_type="solid")
    global_fill = PatternFill(start_color=INPUT_COLOR, end_color=INPUT_COLOR, fill_type="solid")

    # ── ROW 1: Header ──
    ws.row_dimensions[1].height = 32
    ts = datetime.now().strftime('%d%b %H:%M').upper()
    ws['A1'] = ts
    ws['A1'].font = Font(size=9, italic=True, color="CBD5E1")
    ws['A1'].alignment = Alignment(vertical='center')
    ws.merge_cells('A1:B1')
    ws['C1'] = "PRICING ENGINE"
    ws['C1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['C1'].alignment = Alignment(vertical='center')
    ws.cell(1, PUC_COL).value = "PUC"
    ws.cell(1, PUC_COL).font = Font(size=9, bold=True, color="FFFFFF")
    ws.cell(1, PUC_COL).alignment = Alignment(horizontal='center', vertical='center')
    for i, cont in enumerate(['20GP', '40GP', '40HQ', '45HQ']):
        c = ws.cell(1, PUC_COL + 1 + i)
        c.value = cont
        c.font = Font(size=9, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(1, MARKUP_CARRIER_COL).value = "MARKUP"
    ws.cell(1, MARKUP_CARRIER_COL).font = Font(size=10, bold=True, color="FFFFFF")
    ws.cell(1, MARKUP_CARRIER_COL).alignment = Alignment(horizontal='center', vertical='center')
    for i, cont in enumerate(CONT_NAMES):
        c = ws.cell(1, MARKUP_FIRST_COL + i)
        c.value = cont
        c.font = Font(size=9, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(1, MARKUP_ACTIVE_COL).value = "Active"
    ws.cell(1, MARKUP_ACTIVE_COL).font = Font(size=9, bold=True, color="FFFFFF")
    ws.cell(1, MARKUP_ACTIVE_COL).alignment = Alignment(horizontal='center', vertical='center')
    for c in range(1, MARKUP_ACTIVE_COL + 1):
        ws.cell(1, c).fill = header_fill

    # ── ROW 2: Quick Search + PUC + Carrier Markup ──
    ws.row_dimensions[2].height = 28
    for idx, col in enumerate([SEARCH_POL_COL, SEARCH_POD_COL, SEARCH_PLACE_COL]):
        ws.cell(2, col).fill = search_fill
        ws.cell(2, col).font = Font(bold=True, size=10, color="9A3412")
        ws.cell(2, col).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(2, col).border = card_border
    lists = [(SEARCH_POL_COL, f"=Search_Lists!$A$2:$A${pol_last}"),
             (SEARCH_POD_COL, f"=Search_Lists!$B$2:$B${pod_last}"),
             (SEARCH_PLACE_COL, f"=Search_Lists!$C$2:$C${place_last}")]
    for col, formula in lists:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(ws.cell(2, col))

    # PUC route dropdown
    route_cell = ws.cell(PUC_ROW, PUC_COL)
    route_cell.font = Font(bold=True, color="1E40AF", size=10)
    route_cell.fill = input_fill
    route_cell.alignment = Alignment(horizontal='center', vertical='center')
    route_cell.border = card_border
    if not df_puc.empty:
        dv_route = DataValidation(type="list", formula1=f"=PUC_Lookup!$A$2:$A${puc_last_row}", allow_blank=True)
        ws.add_data_validation(dv_route)
        dv_route.add(route_cell)
        puc_places = df_puc['Place'].dropna().unique().tolist()
        route_cell.value = puc_places[0] if puc_places else ""

    # PUC VLOOKUP values
    route_col_let = get_column_letter(PUC_COL)
    puc_cell_refs = []
    for i, col_idx in enumerate([2, 3, 4, 5]):
        cell = ws.cell(PUC_ROW, PUC_COL + 1 + i)
        cell.value = f'=IFERROR(VLOOKUP(${route_col_let}${PUC_ROW},PUC_Lookup!$A:$E,{col_idx},0),0)'
        cell.number_format = '#,##0'
        cell.fill = input_fill
        cell.font = Font(bold=True, size=11, color="1E293B")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = card_border
        puc_cell_refs.append(f'${get_column_letter(PUC_COL + 1 + i)}${PUC_ROW}')
    puc_cell_refs.extend([None, None, None])

    # Carrier dropdown
    carrier_cell = ws.cell(2, MARKUP_CARRIER_COL)
    carrier_cell.value = CARRIERS[0]
    carrier_cell.font = Font(bold=True, size=11, color="166534")
    carrier_cell.fill = markup_fill
    carrier_cell.alignment = Alignment(horizontal='center', vertical='center')
    carrier_cell.border = card_border
    dv_carrier = DataValidation(type="list", formula1=f'"{",".join(CARRIERS)}"')
    ws.add_data_validation(dv_carrier)
    dv_carrier.add(carrier_cell)

    # Carrier markup values
    first_carrier = CARRIERS[0]
    first_vals = existing_markups.get(first_carrier, [CARRIER_DEFAULTS.get(first_carrier, 0)] * len(CONT_NAMES))
    for i in range(len(CONT_NAMES)):
        cell = ws.cell(2, MARKUP_FIRST_COL + i)
        cell.value = first_vals[i] if i < len(first_vals) else 0
        cell.number_format = '#,##0'
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = markup_fill
        cell.border = card_border
        cell.protection = Protection(locked=False)

    # Active summary
    active_cell = ws.cell(2, MARKUP_ACTIVE_COL)
    active_cell.value = f'=SUMPRODUCT((MMULT(IF({MARKUP_STORE}!B3:H{2 + len(CARRIERS)}<>0,1,0),ROW(INDIRECT("1:{len(CONT_NAMES)}"))^0)>0)*1)&" carriers"'
    active_cell.font = Font(size=9, italic=True, color="166534")
    active_cell.alignment = Alignment(horizontal='center', vertical='center')
    active_cell.fill = active_fill
    active_cell.border = card_border

    # ── ROW 3: Quote Gen + Global markup ──
    ws.row_dimensions[3].height = 26
    ws['A3'] = "QUOTE GENERATION"
    ws['A3'].font = Font(size=11, bold=True, color="334155")
    ws['A3'].alignment = Alignment(vertical='center')
    ws['A3'].fill = quote_fill
    ws.merge_cells('A3:E3')
    for c in range(1, 6): ws.cell(3, c).fill = quote_fill
    ws['F3'] = "Generate Quote"
    ws['F3'].font = Font(bold=True, color="FFFFFF", size=10)
    ws['F3'].fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
    ws['F3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['F3'].border = card_border
    ws.merge_cells('F3:H3')
    for c in range(6, 9):
        ws.cell(3, c).fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")

    ws.cell(3, MARKUP_CARRIER_COL).value = "ALL"
    ws.cell(3, MARKUP_CARRIER_COL).font = Font(bold=True, size=10, color="1E40AF")
    ws.cell(3, MARKUP_CARRIER_COL).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(3, MARKUP_CARRIER_COL).fill = global_fill
    ws.cell(3, MARKUP_CARRIER_COL).border = card_border
    all_vals = existing_markups.get("ALL", [0] * len(CONT_NAMES))
    for i in range(len(CONT_NAMES)):
        cell = ws.cell(3, MARKUP_FIRST_COL + i)
        cell.value = all_vals[i] if i < len(all_vals) else 0
        cell.number_format = '#,##0'
        cell.font = Font(bold=True, size=10, color="1E40AF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = global_fill
        cell.border = card_border
        cell.protection = Protection(locked=False)

    # ── ROW 4-5: Customer + Quick preset ──
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 18
    ws['A4'] = "Customer:"
    ws['A4'].font = Font(bold=True, size=10, color="475569")
    ws['A4'].alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells('B4:D4')
    ws.cell(4, 2).border = Border(bottom=thin)
    ws['F4'] = "Quick:"
    ws['F4'].font = Font(bold=True, size=10, color="475569")
    ws['G4'] = "DRY"
    ws['G4'].border = card_border
    ws['G4'].fill = input_fill
    dv_preset = DataValidation(type="list", formula1='"DRY,REEFER,FULL,ALL"')
    ws.add_data_validation(dv_preset)
    dv_preset.add(ws['G4'])
    ws['A5'] = "Filter → Customer → Markup → Generate Quote"
    ws['A5'].font = Font(italic=True, size=8, color="94A3B8")
    ws.merge_cells('A5:H5')

    # ── ROW 6-7: Spacer + Data title ──
    ws.row_dimensions[6].height = 6
    ws.row_dimensions[DATA_TITLE_ROW].height = 28
    ws.cell(DATA_TITLE_ROW, 1).value = f"MASTER PRICING DATA  |  {len(df_master):,} records"
    ws.cell(DATA_TITLE_ROW, 1).font = Font(size=11, bold=True, color="FFFFFF")
    ws.cell(DATA_TITLE_ROW, 1).alignment = Alignment(vertical='center')
    ws.merge_cells(start_row=DATA_TITLE_ROW, start_column=1, end_row=DATA_TITLE_ROW, end_column=16)
    for c in range(1, 17):
        ws.cell(DATA_TITLE_ROW, c).fill = PatternFill(start_color="64748B", end_color="64748B", fill_type="solid")

    # ── ROW 8: Data headers ──
    ws.row_dimensions[DATA_HEADER_ROW].height = 26
    headers = ['POL', 'POD', 'Place', 'Carrier', 'Commodity', 'Eff', 'Exp', 'Note', 'Source',
               '20GP', '40GP', '40HQ', '45HQ', '40NOR', '20RF', '40RF']
    for i, h in enumerate(headers):
        cell = ws.cell(DATA_HEADER_ROW, i + 1)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = card_border

    # ── DATA ROWS WITH FORMULAS ──
    print(f"   Writing {len(df_master):,} rows with formulas...")
    ms_carrier_range = f"{MARKUP_STORE}!$A$3:$A${2 + len(CARRIERS)}"

    for r_idx, (_, row) in enumerate(df_master.iterrows(), start=DATA_START_ROW):
        ws.cell(r_idx, COL_POL).value = row['POL']
        ws.cell(r_idx, COL_POD).value = row['POD']
        ws.cell(r_idx, COL_PLACE).value = row['Place']
        ws.cell(r_idx, COL_CARRIER).value = row['Carrier']
        ws.cell(r_idx, COL_COMMODITY).value = row['Commodity']
        ws.cell(r_idx, COL_EFF).value = row['Eff_Short']
        ws.cell(r_idx, COL_EXP).value = row['Exp_Short']
        ws.cell(r_idx, COL_NOTE).value = row['Note']
        ws.cell(r_idx, COL_SOURCE).value = row['Source_Short']

        for col_idx_src, price_col in enumerate(PRICE_SOURCE):
            base_val = _safe_val(row.get(price_col), 0)
            ws.cell(r_idx, HIDDEN_COL_START + col_idx_src).value = float(base_val) if base_val != 0 else 0

        place_ref = f'$C{r_idx}'
        carrier_ref = f'$D{r_idx}'
        for col_idx, (cont_name, puc_ref) in enumerate(zip(CONT_NAMES, puc_cell_refs)):
            hidden_letter = get_column_letter(HIDDEN_COL_START + col_idx)
            vis_col = COL_FIRST_PRICE + col_idx
            store_data_col = col_idx + 2
            global_col_letter = get_column_letter(MARKUP_FIRST_COL + col_idx)
            global_ref = f"${global_col_letter}$3"
            carrier_cell_ref = f"${global_col_letter}$2"
            ms_data_range = f"{MARKUP_STORE}!${get_column_letter(store_data_col)}$3:${get_column_letter(store_data_col)}${2 + len(CARRIERS)}"
            carrier_lookup = (f'IF({carrier_ref}=$I$2,{carrier_cell_ref},'
                              f'IFERROR(INDEX({ms_data_range},MATCH({carrier_ref},{ms_carrier_range},0)),0))')
            note_ref = f'${get_column_letter(COL_NOTE)}{r_idx}'
            puc_formula = ''
            if puc_ref:
                puc_lookup_col_map = {0: 'B', 1: 'C', 2: 'D', 3: 'E'}
                if col_idx in puc_lookup_col_map:
                    puc_lk_col = puc_lookup_col_map[col_idx]
                    stored_puc = f'IFERROR(SUMPRODUCT((ISNUMBER(SEARCH(PUC_Lookup!$A$2:$A${puc_last_row},{place_ref})))*PUC_Lookup!${puc_lk_col}$2:${puc_lk_col}${puc_last_row}),0)'
                    puc_formula = (f'+IF({note_ref}="SOC",'
                                   f'IF(ISNUMBER(SEARCH(${route_col_let}${PUC_ROW},{place_ref})),{puc_ref},{stored_puc})'
                                   f',0)')

            formula = (f'=IF({hidden_letter}{r_idx}>0,'
                       f'{hidden_letter}{r_idx}+{global_ref}+{carrier_lookup}{puc_formula}'
                       f',"")')
            cell = ws.cell(r_idx, vis_col)
            cell.value = formula
            cell.number_format = '#,##0'

    # Hide helper columns
    for col_idx in range(HIDDEN_COL_START, HIDDEN_COL_START + 7):
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    # Excel Table
    end_row = DATA_START_ROW + len(df_master) - 1
    last_data_col = COL_FIRST_PRICE + len(CONT_NAMES) - 1
    table_ref = f"A{DATA_HEADER_ROW}:{get_column_letter(last_data_col)}{end_row}"
    tab = Table(displayName="MasterPricing", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    ws.add_table(tab)
    ws.auto_filter.ref = table_ref

    # Column widths
    widths = {'A': 10, 'B': 10, 'C': 20, 'D': 12, 'E': 14, 'F': 8, 'G': 8, 'H': 8, 'I': 12}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    for col in range(COL_FIRST_PRICE, COL_FIRST_PRICE + len(CONT_NAMES)):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.column_dimensions[get_column_letter(MARKUP_ACTIVE_COL)].width = 12
    ws.freeze_panes = f"A{DATA_START_ROW}"

    # ── Write Markup Store ──
    if not existing_markups:
        existing_markups["ALL"] = [0] * len(CONT_NAMES)
        for carrier in CARRIERS:
            existing_markups[carrier] = [CARRIER_DEFAULTS.get(carrier, 0)] * len(CONT_NAMES)
    write_markup_store(wb, existing_markups)

    # ── Delete excess rows ──
    if ws.max_row > end_row:
        ws.delete_rows(end_row + 1, ws.max_row - end_row)

    # ── Save ──
    print("   💾 Saving...")
    wb.save(ERP_FILE)
    print(f"\n{'='*60}")
    print(f"✅ DIRECT PARQUET → ERP COMPLETE!")
    print(f"   📊 {len(df_master):,} pricing rows")
    print(f"   📋 BasicCost: {bc_row_idx - 2:,} entries")
    print(f"   🔍 Search: POL({len(unique_pols)}) POD({len(unique_pods)}) Place({len(unique_places)})")
    print(f"   📁 Ref sheets: {', '.join(ref_sheet_data.keys()) or 'none'}")
    print(f"   📁 Hidden: PUC_Lookup, BasicCost_Lookup, Search_Lists, Markup_Store, Version")
    print(f"   ⚠️  Quotes/Jobs sheets: UNTOUCHED")
    print(f"{'='*60}")


if __name__ == "__main__":
    try:
        df_master, df_history_wide, df_puc, ref_sheet_data, version_rows = load_and_process_parquet()
        write_to_erp(df_master, df_history_wide, df_puc, ref_sheet_data, version_rows)
    except PermissionError:
        print("❌ ERROR: Please close ERP_Master.xlsm in Excel first!")
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
