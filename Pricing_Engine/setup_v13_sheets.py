# -*- coding: utf-8 -*-
"""
setup_v13_sheets.py — Add missing columns to Quotes + setup Active Jobs headers
Run once before importing V13 VBA code.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
import os

ERP_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "ERP", "data", "ERP_V13_STAGING.xlsm")

def setup():
    print(f"Opening: {ERP_FILE}")
    wb = openpyxl.load_workbook(ERP_FILE, keep_vba=True)

    # ── Shared styles ──
    hdr_font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='CBD5E1')
    hdr_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ════════════════════════════════════════════════════════════
    # PREP 1 — Add cols 38-42 to Quotes sheet header row
    # ════════════════════════════════════════════════════════════
    ws_q = wb['Quotes']
    new_cols = {
        38: 'StatusDate',
        39: 'Qty',
        40: 'Volume',
        41: 'JobID',
        42: 'ContType',
    }
    for col, name in new_cols.items():
        cell = ws_q.cell(1, col)
        if cell.value is None or cell.value == '':
            cell.value = name
            cell.font = Font(name='Segoe UI', size=10, bold=True)
            cell.alignment = hdr_align
            print(f"  Quotes col {col}: '{name}' added")
        else:
            print(f"  Quotes col {col}: already has '{cell.value}' — skipped")

    # Column widths
    from openpyxl.utils import get_column_letter
    widths = {38: 14, 39: 8, 40: 10, 41: 16, 42: 10}
    for col, w in widths.items():
        ws_q.column_dimensions[get_column_letter(col)].width = w

    print(f"  ✓ Quotes: now {ws_q.max_column} columns")

    # ════════════════════════════════════════════════════════════
    # PREP 2 — Setup Active Jobs sheet headers in row 7
    # ════════════════════════════════════════════════════════════
    ws_j = wb['Active Jobs']

    headers = [
        'Job_ID', 'Quote_ID', 'Customer_ID', 'Customer_Name', 'Customer_Type',
        'Routing', 'Bkg_No', 'Hbl_No', 'ETD', 'ETD_Original', 'ETA',
        'ETA_Alert_Date', 'ATA', 'Carrier', 'Contract_Type', 'Container_Type',
        'Quantity', 'Volume', 'Selling_Rate', 'Buying_Rate', 'Profit',
        'Profit_Margin', 'Status', 'Delay_Count', 'Delay_Log',
        'Door_Delivery', 'Door_Address', 'Door_Status', 'SI_Received',
        'CY_Cutoff', 'Carrier_Com', 'Customer_Com', 'Notes',
        'Created_Date', 'Last_Updated', 'Cost_Breakdown', '\U0001f4e7 Request BKG',
    ]

    # Clear row 1 legacy content (just had "JobID" in A1)
    for c in range(1, 40):
        ws_j.cell(1, c).value = None

    # Write headers in row 7
    for i, h in enumerate(headers, 1):
        cell = ws_j.cell(7, i)
        cell.value = h
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = hdr_border

    # Title row 1: merged
    ws_j.merge_cells('A1:AK1')
    title = ws_j.cell(1, 1)
    title.value = 'ACTIVE JOBS'
    title.font = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
    title.fill = PatternFill('solid', fgColor='294B93')
    title.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    col_widths = {
        'A': 14, 'B': 14, 'C': 12, 'D': 20, 'E': 14,
        'F': 30, 'G': 16, 'H': 16, 'I': 12, 'J': 12, 'K': 12,
        'L': 14, 'M': 12, 'N': 10, 'O': 14, 'P': 14,
        'Q': 10, 'R': 10, 'S': 14, 'T': 14, 'U': 12,
        'V': 14, 'W': 12, 'X': 12, 'Y': 20,
        'Z': 14, 'AA': 20, 'AB': 12, 'AC': 12,
        'AD': 12, 'AE': 16, 'AF': 16, 'AG': 20,
        'AH': 14, 'AI': 14, 'AJ': 30, 'AK': 16,
    }
    for col_letter, w in col_widths.items():
        ws_j.column_dimensions[col_letter].width = w

    # Freeze panes: freeze rows 1-7
    ws_j.freeze_panes = 'A8'

    # Number formats for header guidance (data rows will inherit)
    # S, T, U = currency; V = percentage
    for c in [19, 20, 21]:  # S, T, U
        ws_j.cell(7, c).number_format = '$#,##0'
    ws_j.cell(7, 22).number_format = '0.0%'

    print(f"  ✓ Active Jobs: {len(headers)} headers in row 7, freeze at A8")

    # ── Save ──
    wb.save(ERP_FILE)
    print(f"\n✅ ERP_V13_STAGING.xlsm updated!")
    print(f"   Quotes: cols 38-42 added (StatusDate, Qty, Volume, JobID, ContType)")
    print(f"   Active Jobs: {len(headers)} headers in row 7, data starts row 8")

if __name__ == '__main__':
    setup()
