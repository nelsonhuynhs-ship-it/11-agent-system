# -*- coding: utf-8 -*-
"""
puc_importer.py — PUC/SOC automation pipeline
==============================================================
Imports SOC PUC rates from team file → updates PUC_SOC.xlsx
→ updates ERP PUC_Lookup sheet → updates Parquet PUC corrections.

Usage:
  python puc_importer.py                              # auto-detect newest PUC file
  python puc_importer.py "PUC MAR 2026 (2).xlsx"     # specific file

Container mapping (CRITICAL):
  Source:  20DC, 40HC
  Target:  PUC_20 = 20DC value
           PUC_40 = 40HC value   (40GP uses 40HC rate)
           PUC_40HC = 40HC value (same as PUC_40)
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import os
import glob
from datetime import datetime
import openpyxl

# ── Paths ──────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ENGINE_DIR = os.path.dirname(SCRIPT_DIR)  # Engine_test/
PUC_SOC_FILE = os.path.join(SCRIPT_DIR, "PUC_SOC.xlsx")
PUC_SOC_DATA = os.path.join(SCRIPT_DIR, "data", "PUC_SOC.xlsx")
ERP_FILE = os.path.join(ENGINE_DIR, "ERP", "data", "ERP_Master.xlsm")

# ── Port Name Mapping ─────────────────────────────────────────────
PORT_MAP = {
    "Los Angeles":   "LAX/LGB",
    "Long Beach":    "LAX/LGB",
    "New York":      "NEW YORK, NY",
    "Oakland":       "OAKLAND, CA",
    "Seattle":       "SEATTLE, WA",
    "Portland":      "PORTLAND, OR",
    "Norfolk":       "NORFOLK, VA",
    "Baltimore":     "BALTIMORE, MD",
    "Boston":        "BOSTON, MA",
    "Jacksonville":  "JACKSONVILLE, FL",
    "Wilmington":    "WILMINGTON, NC",
    "Houston":       "HOUSTON, TX",
    "Charleston":    "CHARLESTON, SC",
    "Phoenix":       "PHOENIX, AZ",
    "Savannah":      "SAVANNAH, GA",
    "Miami":         "MIAMI, FL",
    "Chicago":       "CHICAGO, IL",
}

# Reverse map for matching: standardized name → source name
REVERSE_MAP = {}
for src, std in PORT_MAP.items():
    REVERSE_MAP.setdefault(std, []).append(src)


def find_latest_sheet(wb):
    """Find the sheet with the LATEST expiry date."""
    latest_date = None
    latest_ws = None
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=3, values_only=True):
            for cell in row:
                if isinstance(cell, datetime):
                    if latest_date is None or cell > latest_date:
                        latest_date = cell
                        latest_ws = ws
                    break  # Only check first date per row
            if latest_date and latest_ws == ws:
                break  # Found a date in this sheet, enough
    return latest_ws, latest_date


def parse_puc_data(ws):
    """
    Parse PUC data from a worksheet.
    Row 1: (Destination, TYPE, None, Expiration Date)
    Row 2: (None, 20DC, 40HC, None)
    Row 3+: (city_name, 20dc_value, 40hc_value, date)

    Returns dict: {standardized_name: {PUC_20, PUC_40, PUC_40HC}}
    """
    new_rates = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        dest = row[0] if row[0] else None
        val_20 = row[1] if len(row) > 1 else None
        val_40hc = row[2] if len(row) > 2 else None

        if not dest or not isinstance(val_20, (int, float)):
            continue

        dest_str = str(dest).strip()

        # Map port name
        std_name = PORT_MAP.get(dest_str, dest_str.upper())

        # Apply 3-type container mapping rule
        try:
            v20 = float(val_20) if isinstance(val_20, (int, float)) else None
            v40hc = float(val_40hc) if isinstance(val_40hc, (int, float)) else None
        except (ValueError, TypeError):
            continue  # Skip "TBA" or other non-numeric values

        if v20 is None and v40hc is None:
            continue  # Skip rows with no valid data

        new_rates[std_name] = {
            "PUC_20":   v20 or 0,
            "PUC_40":   v40hc or 0,    # 40GP uses 40HC rate
            "PUC_40HC": v40hc or 0,    # same as PUC_40
        }

    return new_rates


def update_puc_soc_file(new_rates, source_filename, expiry):
    """Update PUC_SOC.xlsx with new rates. Handle LA/LB merge."""
    # Use whichever PUC_SOC.xlsx exists
    puc_path = PUC_SOC_FILE if os.path.exists(PUC_SOC_FILE) else PUC_SOC_DATA
    if not os.path.exists(puc_path):
        print(f"[WARN] PUC_SOC.xlsx not found at {puc_path}")
        return 0, 0

    puc_wb = openpyxl.load_workbook(puc_path)
    puc_ws = puc_wb.active

    updated = 0
    added = 0
    rows_to_delete = []
    existing_places = set()

    # Pass 1: Update existing rows and mark duplicates
    for row_idx in range(2, puc_ws.max_row + 1):
        place = puc_ws.cell(row_idx, 1).value
        if not place:
            continue
        std_place = str(place).strip().upper()
        existing_places.add(std_place)

        # Delete individual LA/LB rows (we merge them to LAX/LGB)
        if std_place in ["LONG BEACH, CA", "LOS ANGELES, CA"]:
            rows_to_delete.append(row_idx)
            continue

        # Update if we have new rate
        if std_place in new_rates:
            rates = new_rates[std_place]
            puc_ws.cell(row_idx, 2).value = rates["PUC_20"]
            puc_ws.cell(row_idx, 3).value = rates["PUC_40"]
            puc_ws.cell(row_idx, 4).value = rates["PUC_40HC"]
            updated += 1

    # Delete duplicate rows (reverse order to avoid index shift)
    for row_idx in sorted(rows_to_delete, reverse=True):
        puc_ws.delete_rows(row_idx)

    # Pass 2: Add new destinations not in existing file
    next_row = puc_ws.max_row + 1
    for std_name, rates in new_rates.items():
        if std_name not in existing_places:
            puc_ws.cell(next_row, 1).value = std_name
            puc_ws.cell(next_row, 2).value = rates["PUC_20"]
            puc_ws.cell(next_row, 3).value = rates["PUC_40"]
            puc_ws.cell(next_row, 4).value = rates["PUC_40HC"]
            next_row += 1
            added += 1

    # Update Meta sheet
    if "Meta" not in puc_wb.sheetnames:
        meta_ws = puc_wb.create_sheet("Meta")
    else:
        meta_ws = puc_wb["Meta"]
    meta_ws["A1"] = "Last updated"
    meta_ws["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    meta_ws["A2"] = "Source file"
    meta_ws["B2"] = source_filename
    meta_ws["A3"] = "Valid until"
    meta_ws["B3"] = str(expiry.strftime("%Y-%m-%d") if expiry else "unknown")
    meta_ws["A4"] = "Container mapping"
    meta_ws["B4"] = "PUC_40 = PUC_40HC = source 40HC value"

    puc_wb.save(puc_path)
    print(f"[PUC] PUC_SOC.xlsx updated: {updated} updated, "
          f"{added} added, {len(rows_to_delete)} duplicates removed")

    # Also update the copy in data/ if the main one is at root
    if os.path.exists(PUC_SOC_FILE) and os.path.exists(PUC_SOC_DATA):
        import shutil
        shutil.copy2(PUC_SOC_FILE, PUC_SOC_DATA)
        print(f"[PUC] Synced copy to data/PUC_SOC.xlsx")

    return updated, len(rows_to_delete)


def update_erp_puc_lookup(new_rates):
    """Update the PUC_Lookup sheet in ERP_Master.xlsm."""
    if not os.path.exists(ERP_FILE):
        print(f"[WARN] ERP_Master.xlsm not found at {ERP_FILE}")
        return False

    # Check if Excel has it open
    was_open = False
    xl = None
    try:
        import win32com.client
        xl = win32com.client.GetActiveObject("Excel.Application")
        for wb in xl.Workbooks:
            if "ERP_Master" in wb.Name:
                was_open = True
                wb.Save()
                wb.Close(SaveChanges=True)
                print("[PUC] ERP saved and closed for PUC update")
                break
    except Exception:
        pass

    if was_open:
        import time
        time.sleep(2)

    try:
        erp_wb = openpyxl.load_workbook(ERP_FILE, keep_vba=True)

        if "PUC_Lookup" not in erp_wb.sheetnames:
            print("[WARN] PUC_Lookup sheet not found in ERP — creating")
            ws_puc = erp_wb.create_sheet("PUC_Lookup")
            ws_puc.sheet_state = 'hidden'
            for ci, h in enumerate(['Place', '20GP', '40GP', '40HQ', '45HQ'], 1):
                ws_puc.cell(1, ci).value = h
        else:
            ws_puc = erp_wb["PUC_Lookup"]

        # Update existing rows
        erp_updated = 0
        for row_idx in range(2, ws_puc.max_row + 1):
            place = ws_puc.cell(row_idx, 1).value
            if not place:
                continue
            std_place = str(place).strip().upper()
            if std_place in new_rates:
                rates = new_rates[std_place]
                ws_puc.cell(row_idx, 2).value = rates["PUC_20"]
                ws_puc.cell(row_idx, 3).value = rates["PUC_40"]    # 40GP
                ws_puc.cell(row_idx, 4).value = rates["PUC_40HC"]  # 40HQ
                # 45HQ stays unchanged (col 5)
                erp_updated += 1

        erp_wb.save(ERP_FILE)
        print(f"[PUC] ERP PUC_Lookup updated: {erp_updated} destinations")

        # Reopen if was open
        if was_open:
            try:
                import win32com.client
                if xl is None:
                    xl = win32com.client.Dispatch("Excel.Application")
                xl.Visible = True
                xl.Workbooks.Open(os.path.abspath(ERP_FILE))
                print("[PUC] ERP reopened")
            except Exception as e:
                print(f"[WARN] Could not reopen ERP: {e}")

        return True
    except PermissionError:
        print("[ERROR] ERP file is locked — close Excel and retry")
        return False
    except Exception as e:
        print(f"[ERROR] ERP PUC update failed: {e}")
        return False


def import_puc_file(filepath):
    """
    Full PUC import pipeline:
    1. Find latest sheet in source file
    2. Parse data with port mapping + container mapping
    3. Update PUC_SOC.xlsx
    4. Update ERP PUC_Lookup sheet
    """
    filepath = os.path.abspath(filepath)
    print(f"\n{'='*60}")
    print(f"  PUC IMPORT PIPELINE")
    print(f"{'='*60}")
    print(f"  Source: {os.path.basename(filepath)}")

    if not os.path.exists(filepath):
        print(f"[ERROR] File not found: {filepath}")
        return None

    # 1. Open and find latest sheet
    wb = openpyxl.load_workbook(filepath, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    latest_ws, expiry = find_latest_sheet(wb)
    if not latest_ws:
        print("[ERROR] No valid dates found in any sheet")
        return None
    print(f"  Selected: {latest_ws.title} (valid until {expiry.strftime('%Y-%m-%d')})")

    # 2. Parse data
    new_rates = parse_puc_data(latest_ws)
    wb.close()
    print(f"  Parsed: {len(new_rates)} destinations")

    # Show key rates
    for key_dest in ["LAX/LGB", "NEW YORK, NY", "CHICAGO, IL"]:
        if key_dest in new_rates:
            r = new_rates[key_dest]
            print(f"    {key_dest}: PUC_20=${r['PUC_20']}, PUC_40=${r['PUC_40']}, PUC_40HC=${r['PUC_40HC']}")

    # 3. Update PUC_SOC.xlsx
    print(f"\n[PUC] Updating PUC_SOC.xlsx...")
    updated, removed = update_puc_soc_file(new_rates, os.path.basename(filepath), expiry)

    # 4. Update ERP PUC_Lookup
    print(f"\n[PUC] Updating ERP PUC_Lookup...")
    erp_ok = update_erp_puc_lookup(new_rates)

    result = {
        "source_file": os.path.basename(filepath),
        "source_sheet": latest_ws.title,
        "valid_until": expiry.strftime("%Y-%m-%d") if expiry else "unknown",
        "destinations_parsed": len(new_rates),
        "destinations_updated": updated,
        "duplicates_removed": removed,
        "erp_updated": erp_ok,
        "lax_lgb_puc20": new_rates.get("LAX/LGB", {}).get("PUC_20"),
        "lax_lgb_puc40": new_rates.get("LAX/LGB", {}).get("PUC_40"),
        "lax_lgb_puc40hc": new_rates.get("LAX/LGB", {}).get("PUC_40HC"),
    }

    print(f"\n{'='*60}")
    print(f"  PUC IMPORT COMPLETE")
    print(f"{'='*60}")
    print(f"  Sheet:        {result['source_sheet']}")
    print(f"  Valid until:   {result['valid_until']}")
    print(f"  Destinations:  {result['destinations_parsed']} parsed, {result['destinations_updated']} updated")
    print(f"  Duplicates:    {result['duplicates_removed']} removed")
    print(f"  ERP updated:   {'YES' if result['erp_updated'] else 'NO'}")
    print(f"  LAX/LGB:       PUC_20=${result['lax_lgb_puc20']}, PUC_40=${result['lax_lgb_puc40']}")
    print(f"{'='*60}")

    return result


if __name__ == "__main__":
    if len(sys.argv) > 1:
        result = import_puc_file(sys.argv[1])
    else:
        # Auto-detect newest PUC_*.xlsx in Pricing_Engine/
        search_paths = [
            os.path.join(SCRIPT_DIR, "PUC*.xlsx"),
            os.path.join(SCRIPT_DIR, "PUC *.xlsx"),
        ]
        files = []
        for pattern in search_paths:
            files.extend(glob.glob(pattern))
        # Exclude PUC_SOC.xlsx itself
        files = [f for f in files if "PUC_SOC" not in os.path.basename(f)]

        if not files:
            print("[ERROR] No PUC source file found.")
            print("  Drop file in Pricing_Engine/ and retry, or specify path:")
            print("  python puc_importer.py 'PUC MAR 2026 (2).xlsx'")
            sys.exit(1)

        newest = max(files, key=os.path.getmtime)
        print(f"[AUTO] Using newest file: {os.path.basename(newest)}")
        result = import_puc_file(newest)

    if not result:
        sys.exit(1)
