"""Load weekly capacity signals from team-filled xlsx input."""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from .paths import capacity_input_file
from .schemas import CapacitySignal

log = logging.getLogger(__name__)

EXPECTED_COLUMNS = [
    "week", "carrier", "lane", "dimension", "status",
    "score", "notes", "entered_by", "entered_at",
]


def load_capacity(week: str, xlsx_path: Optional[Path] = None) -> list[CapacitySignal]:
    """Load capacity-{week}.xlsx if it exists, validate, return signals.

    Silently returns [] if the file is missing (expected when team hasn't filled
    it yet). Logs and skips malformed rows.
    """
    xlsx_path = xlsx_path or capacity_input_file(week)
    if not xlsx_path.exists():
        log.info("Capacity input missing for %s: %s", week, xlsx_path)
        return []

    try:
        from openpyxl import load_workbook
    except ImportError:
        log.error("openpyxl not installed; cannot load capacity xlsx")
        return []

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        log.warning("Failed to open %s: %s", xlsx_path, e)
        return []

    ws = wb.active
    if ws is None:
        return []

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        log.warning("Empty xlsx: %s", xlsx_path)
        return []

    header = [str(h).strip().lower() if h is not None else "" for h in header_row]
    col_idx = {name: header.index(name) for name in EXPECTED_COLUMNS if name in header}
    missing = [c for c in EXPECTED_COLUMNS if c not in col_idx]
    if missing:
        log.warning("Capacity xlsx %s missing columns: %s", xlsx_path.name, missing)

    signals: list[CapacitySignal] = []
    for row in rows_iter:
        if not row or all(v is None for v in row):
            continue
        try:
            sig = _row_to_signal(row, col_idx)
        except Exception as e:
            log.warning("Skipping malformed row in %s: %s", xlsx_path.name, e)
            continue
        if sig is not None:
            signals.append(sig)

    wb.close()
    return signals


def _row_to_signal(row: tuple, col_idx: dict[str, int]) -> Optional[CapacitySignal]:
    """Convert an xlsx row tuple into a CapacitySignal."""
    def g(name: str, default=None):
        idx = col_idx.get(name)
        if idx is None or idx >= len(row):
            return default
        v = row[idx]
        return v if v is not None else default

    week = str(g("week", "")).strip()
    carrier = str(g("carrier", "")).strip()
    lane_raw = str(g("lane", "ALL")).strip().upper()
    if lane_raw not in ("WC", "EC", "GULF", "ALL"):
        lane_raw = "ALL"
    dimension = str(g("dimension", "space")).strip().lower()
    if dimension not in ("space", "equipment", "booking_policy"):
        dimension = "space"
    status = str(g("status", "OPEN")).strip().upper()
    if status not in ("OPEN", "TIGHT", "FULL", "ROLLING"):
        status = "OPEN"

    score_raw = g("score", 3)
    try:
        score = int(score_raw)
    except (ValueError, TypeError):
        score = 3

    notes = str(g("notes", "")).strip()
    entered_by = str(g("entered_by", "")).strip()

    entered_at_raw = g("entered_at")
    entered_at: Optional[datetime] = None
    if isinstance(entered_at_raw, datetime):
        entered_at = entered_at_raw
    elif isinstance(entered_at_raw, str) and entered_at_raw:
        try:
            entered_at = datetime.fromisoformat(entered_at_raw)
        except ValueError:
            entered_at = None

    if not week or not carrier:
        return None

    return CapacitySignal(
        week=week,
        carrier=carrier,
        lane=lane_raw,  # type: ignore[arg-type]
        dimension=dimension,  # type: ignore[arg-type]
        status=status,  # type: ignore[arg-type]
        score=max(1, min(5, score)),
        notes=notes,
        entered_by=entered_by,
        entered_at=entered_at,
    )


def average_score(signals: list[CapacitySignal]) -> float:
    """Return average score across all signals, or 0 if empty."""
    if not signals:
        return 0.0
    return sum(s.score for s in signals) / len(signals)
