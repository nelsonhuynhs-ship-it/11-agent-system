"""
markup_engine.py — Sprint 7: Auto-Markup Engine
Reads Markup_Store and PUC_Lookup from ERP_Master.xlsm at startup (cached).
Provides calculate_selling_price() for bot.py to use.

Selling Price Formula (mirrors ERP logic):
  Selling = Base + GlobalMarkup(ALL row) + CarrierMarkup(carrier-specific row) + PUC(if SOC)
"""
import os
import logging
from functools import lru_cache

logger = logging.getLogger(__name__)

# ── Cache ──
_markup_data = None  # dict: { carrier_upper: { cont: markup_value } }
_global_markup = {}  # dict: { cont: global_markup_value }  (row='ALL')
_puc_data = None     # dict: { place_keyword_upper: { cont: puc_value } }
_erp_load_time = None


def _parse_cont_name(col_name: str) -> str | None:
    """Map column header to canonical container code."""
    MAP = {
        '20GP': '20GP', '20DC': '20GP', '20': '20GP',
        '40GP': '40GP', '40DC': '40GP',
        '40HQ': '40HQ', '40HC': '40HQ', '40HG': '40HQ',
        '45HQ': '45HQ', "45'HQ": '45HQ',
        '40NOR': '40NOR',
        '20RF': '20RF', '40RF': '40RF',
    }
    upper = str(col_name).strip().upper()
    return MAP.get(upper)


def load_markup_from_erp(erp_file: str) -> bool:
    """
    Load Markup_Store and PUC_Lookup sheets from ERP_Master.xlsm.
    Returns True on success, False on failure.
    """
    global _markup_data, _global_markup, _puc_data, _erp_load_time
    import openpyxl
    from datetime import datetime

    if not os.path.exists(erp_file):
        logger.warning(f"[Markup] ERP file not found: {erp_file}")
        return False

    try:
        wb = openpyxl.load_workbook(erp_file, keep_vba=True, read_only=True)

        # ── Load Markup_Store ──
        markup_sheet = None
        for s in wb.sheetnames:
            if 'markup' in s.lower():
                markup_sheet = wb[s]
                break

        markup_data = {}
        global_markup = {}

        if markup_sheet:
            rows = list(markup_sheet.iter_rows(values_only=True))
            if rows:
                headers = [str(h).strip() if h else '' for h in rows[0]]
                # headers[0]='Carrier', headers[1+]= container types
                cont_cols = {}
                for ci, h in enumerate(headers[1:], start=1):
                    canonical = _parse_cont_name(h)
                    if canonical:
                        cont_cols[ci] = canonical

                for row in rows[1:]:
                    if not row or row[0] is None:
                        continue
                    carrier = str(row[0]).strip().upper()
                    per_cont = {}
                    for ci, cont in cont_cols.items():
                        try:
                            val = float(row[ci]) if row[ci] is not None else 0.0
                        except (TypeError, ValueError):
                            val = 0.0
                        per_cont[cont] = val

                    if carrier == 'ALL':
                        global_markup = per_cont
                    else:
                        markup_data[carrier] = per_cont

        # ── Load PUC_Lookup ──
        puc_sheet = None
        for s in wb.sheetnames:
            if 'puc' in s.lower():
                puc_sheet = wb[s]
                break

        puc_data = {}
        if puc_sheet:
            rows = list(puc_sheet.iter_rows(values_only=True))
            if rows:
                headers = [str(h).strip() if h else '' for h in rows[0]]
                cont_cols = {}
                for ci, h in enumerate(headers[1:], start=1):
                    canonical = _parse_cont_name(h)
                    if canonical:
                        cont_cols[ci] = canonical

                for row in rows[1:]:
                    if not row or row[0] is None:
                        continue
                    place_key = str(row[0]).strip().upper()
                    per_cont = {}
                    for ci, cont in cont_cols.items():
                        try:
                            val = float(row[ci]) if row[ci] is not None else 0.0
                        except (TypeError, ValueError):
                            val = 0.0
                        per_cont[cont] = val
                    puc_data[place_key] = per_cont

        wb.close()

        _markup_data = markup_data
        _global_markup = global_markup
        _puc_data = puc_data
        _erp_load_time = datetime.now()

        carriers_loaded = list(markup_data.keys())
        puc_places = len(puc_data)
        logger.info(f"[Markup] Loaded: global={global_markup}, carriers={carriers_loaded}, PUC places={puc_places}")
        return True

    except Exception as e:
        logger.error(f"[Markup] Failed to load ERP: {e}")
        return False


def get_carrier_markup(carrier: str, container: str) -> float:
    """Return carrier-specific markup for the given container type."""
    if not _markup_data:
        return 0.0
    carrier_upper = carrier.strip().upper()
    cont = _normalise_cont(container)
    row = _markup_data.get(carrier_upper, {})
    return row.get(cont, 0.0)


def get_global_markup(container: str) -> float:
    """Return global markup (ALL row) for the given container type."""
    if not _global_markup:
        return 0.0
    cont = _normalise_cont(container)
    return _global_markup.get(cont, 0.0)


def get_puc(place: str, container: str) -> float:
    """
    Return PUC (Pick-Up Charge) for a SOC shipment.
    Matches by checking if any PUC_Lookup key appears in place string.
    """
    if not _puc_data:
        return 0.0
    place_upper = str(place).strip().upper()
    cont = _normalise_cont(container)
    for key, per_cont in _puc_data.items():
        if key in place_upper or place_upper in key:
            return per_cont.get(cont, 0.0)
    return 0.0


def _normalise_cont(container: str) -> str:
    """Normalise container string to canonical key."""
    MAP = {
        '20': '20GP', '20GP': '20GP', '20DC': '20GP',
        '40': '40GP', '40GP': '40GP',
        'HQ': '40HQ', '40HQ': '40HQ', '40HC': '40HQ', '40HG': '40HQ',
        '45HQ': '45HQ', "45'HQ": '45HQ',
        '40NOR': '40NOR',
        '20RF': '20RF', '40RF': '40RF', 'RF': '40RF',
    }
    upper = str(container).strip().upper()
    return MAP.get(upper, upper)


def calculate_selling_price(base_price: float, carrier: str, container: str,
                             place: str = '', is_soc: bool = False,
                             adhoc_markup: float = 0.0) -> dict:
    """
    Calculate full selling price from base (net cost).

    Returns dict with:
        selling    — final selling price
        base       — original net cost
        global_mk  — global markup applied
        carrier_mk — carrier-specific markup
        puc        — PUC if SOC, else 0
        breakdown  — human-readable breakdown string
    """
    global_mk = get_global_markup(container)
    carrier_mk = get_carrier_markup(carrier, container)
    puc = get_puc(place, container) if is_soc else 0.0
    selling = base_price + global_mk + carrier_mk + puc + adhoc_markup

    parts = [f"Base ${base_price:,.0f}"]
    if global_mk:
        parts.append(f"G.Markup +${global_mk:,.0f}")
    if carrier_mk:
        parts.append(f"{carrier} +${carrier_mk:,.0f}")
    if puc:
        parts.append(f"PUC +${puc:,.0f}")
    if adhoc_markup:
        label = f"+Custom ${adhoc_markup:,.0f}" if adhoc_markup > 0 else f"-Disc ${abs(adhoc_markup):,.0f}"
        parts.append(label)

    return {
        'selling': selling,
        'base': base_price,
        'global_mk': global_mk,
        'carrier_mk': carrier_mk,
        'puc': puc,
        'adhoc': adhoc_markup,
        'breakdown': ' | '.join(parts),
    }


def is_markup_loaded() -> bool:
    return _markup_data is not None


def markup_summary() -> str:
    """Return a one-line summary of loaded markup data."""
    if not is_markup_loaded():
        return "❌ Markup chưa load"
    carriers = list(_markup_data.keys())
    return (f"✅ Markup: global={_global_markup.get('40HQ', 0)} | "
            f"Carriers={carriers} | PUC={len(_puc_data)} places")
