"""
erp_reader.py — Sprint 8: Read-only ERP Access
Provides CRM, Quote history, and Active Job reads for bot commands and AI context.
All operations are READ-ONLY — never modifies the ERP file.
"""
import os
import logging
from datetime import datetime

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

ERP_FILE = None  # set by init_reader()

# ── Column maps (1-indexed) discovered from ERP inspection ──
QUOTES_HEADER_ROW = 2   # row 2 contains column headers in Quotes sheet
JOBS_HEADER_ROW = 7     # row 7 contains column headers in Active Jobs sheet


def init_reader(erp_file: str):
    global ERP_FILE
    ERP_FILE = erp_file


def _load_sheet_as_df(sheet_keyword: str, header_row: int) -> pd.DataFrame | None:
    """Load a named sheet from ERP into a DataFrame. Thread-safe read-only."""
    if not ERP_FILE or not os.path.exists(ERP_FILE):
        logger.warning("[ERP Reader] ERP file not found")
        return None
    try:
        wb = openpyxl.load_workbook(ERP_FILE, keep_vba=True, read_only=True)
        target = None
        for s in wb.sheetnames:
            if sheet_keyword.lower() in s.lower():
                target = wb[s]
                break
        if target is None:
            wb.close()
            return None

        rows = list(target.iter_rows(values_only=True))
        wb.close()

        if len(rows) < header_row:
            return None

        headers = [str(h).strip() if h else f"Col{i}" for i, h in enumerate(rows[header_row - 1], 1)]
        data_rows = rows[header_row:]
        df = pd.DataFrame(data_rows, columns=headers)
        # Drop rows where all values are None
        df = df.dropna(how='all')
        return df

    except Exception as e:
        logger.error(f"[ERP Reader] Error loading sheet '{sheet_keyword}': {e}")
        return None


# ─────────────────────── QUOTES ───────────────────────

def get_quote_history(customer_name: str = None, status: str = None, limit: int = 10) -> list[dict]:
    """
    Read recent quotes from Quotes sheet.
    Optional filter by customer_name and/or status (PENDING/WIN/LOSS).
    """
    df = _load_sheet_as_df('quot', QUOTES_HEADER_ROW)
    if df is None or df.empty:
        return []

    if customer_name:
        mask = df['Customer_Name'].astype(str).str.upper().str.contains(customer_name.upper(), na=False)
        df = df[mask]

    if status:
        df = df[df['Status'].astype(str).str.upper() == status.upper()]

    # Sort by Quote_Date descending
    if 'Quote_Date' in df.columns:
        df['Quote_Date'] = pd.to_datetime(df['Quote_Date'], errors='coerce')
        df = df.sort_values('Quote_Date', ascending=False)

    records = []
    for _, row in df.head(limit).iterrows():
        try:
            records.append({
                'quote_id':   str(row.get('Quote_ID', '')),
                'date':       pd.to_datetime(row.get('Quote_Date'), errors='coerce'),
                'customer':   str(row.get('Customer_Name', '')),
                'pol':        str(row.get('POL', '')),
                'pod':        str(row.get('POD', '')),
                'place':      str(row.get('Place', '')),
                'carrier':    str(row.get('Carrier', '')),
                'container':  str(row.get('Container_Type', '')),
                'price':      float(row['Price']) if pd.notna(row.get('Price')) else 0,
                'status':     str(row.get('Status', '')),
                'job_id':     str(row.get('Job_ID', '')) if pd.notna(row.get('Job_ID')) else '',
                'exp_date':   pd.to_datetime(row.get('Exp_Date'), errors='coerce'),
            })
        except Exception:
            continue
    return records


def get_quote_stats(customer_name: str) -> dict:
    """Return WIN/LOSS/PENDING counts for a customer."""
    all_quotes = get_quote_history(customer_name, limit=500)
    counter = {'WIN': 0, 'LOSS': 0, 'PENDING': 0, 'total': len(all_quotes)}
    for q in all_quotes:
        s = q['status'].upper()
        if s in counter:
            counter[s] += 1
    if counter['total'] > 0:
        counter['win_rate'] = round(counter['WIN'] / counter['total'] * 100, 1)
    else:
        counter['win_rate'] = 0
    return counter


# ─────────────────────── ACTIVE JOBS ───────────────────────

def get_active_jobs(customer_name: str = None, status_filter: str = None, limit: int = 10) -> list[dict]:
    """
    Read Active Jobs sheet.
    Optional filter by customer_name and/or status.
    """
    df = _load_sheet_as_df('active', JOBS_HEADER_ROW)
    if df is None or df.empty:
        return []

    # Keep only data rows (skip summary rows — they have no Job_ID)
    if 'Job_ID' in df.columns:
        df = df[df['Job_ID'].astype(str).str.match(r'J\d{8,}', na=False)]

    if customer_name:
        mask = df['Customer_Name'].astype(str).str.upper().str.contains(customer_name.upper(), na=False)
        df = df[mask]

    if status_filter:
        status_col = [c for c in df.columns if 'status' in c.lower()]
        if status_col:
            df = df[df[status_col[0]].astype(str).str.upper().str.contains(status_filter.upper(), na=False)]

    records = []
    for _, row in df.head(limit).iterrows():
        try:
            records.append({
                'job_id':      str(row.get('Job_ID', '')),
                'quote_id':    str(row.get('Quote_ID', '')) if pd.notna(row.get('Quote_ID')) else '',
                'customer':    str(row.get('Customer_Name', '')),
                'routing':     str(row.get('Routing', '')),
                'carrier':     str(row.get('Carrier', '')),
                'container':   str(row.get('Container_Type', '')),
                'quantity':    int(row['Quantity']) if pd.notna(row.get('Quantity')) else 1,
                'etd':         pd.to_datetime(row.get('ETD'), errors='coerce'),
                'eta':         pd.to_datetime(row.get('ETA'), errors='coerce'),
                'selling':     float(row['Selling_Rate']) if pd.notna(row.get('Selling_Rate')) else 0,
                'buying':      float(row['Buying_Rate']) if pd.notna(row.get('Buying_Rate')) else 0,
                'bkg_no':      str(row.get('Bkg_No', '')) if pd.notna(row.get('Bkg_No')) else '',
                'status':      _get_job_status(row),
            })
        except Exception:
            continue
    return records


def _get_job_status(row) -> str:
    """Determine job status from available columns."""
    for col in row.index:
        if 'status' in col.lower():
            val = str(row[col])
            if val and val != 'None' and val != 'nan':
                return val
    # Infer from dates
    try:
        eta = pd.to_datetime(row.get('ETA'), errors='coerce')
        ata = pd.to_datetime(row.get('ATA'), errors='coerce')
        if pd.notna(ata):
            return 'Delivered'
        if pd.notna(eta) and eta < pd.Timestamp.now():
            return 'Arrived'
        return 'In_Transit'
    except Exception:
        return 'Unknown'


# ─────────────────────── CRM ───────────────────────

def get_crm_profile(customer_name: str) -> dict | None:
    """Look up a customer in the Customers CRM sheet."""
    df = _load_sheet_as_df('customer', 2)
    if df is None or df.empty:
        return None

    mask = df['Company_Name'].astype(str).str.upper().str.contains(customer_name.upper(), na=False)
    matches = df[mask]
    if matches.empty:
        return None

    row = matches.iloc[0]
    return {
        'id':             str(row.get('Customer_ID', '')),
        'type':           str(row.get('Customer_Type', '')),
        'name':           str(row.get('Company_Name', '')),
        'contact':        str(row.get('Contact_Person', '')) if pd.notna(row.get('Contact_Person')) else '',
        'email':          str(row.get('Email', '')) if pd.notna(row.get('Email')) else '',
        'phone':          str(row.get('Phone', '')) if pd.notna(row.get('Phone')) else '',
        'payment_terms':  str(row.get('Payment_Terms', '')) if pd.notna(row.get('Payment_Terms')) else '',
        'status':         str(row.get('Status', '')) if pd.notna(row.get('Status')) else '',
        'notes':          str(row.get('Notes', '')) if pd.notna(row.get('Notes')) else '',
    }


# ─────────────────────── AI CONTEXT BUILDER ───────────────────────

def build_full_context(customer_name: str) -> str:
    """
    Build a rich AI context string for a customer.
    Merges CRM profile + quote stats + recent quotes + active jobs.
    Injected into Gemini system prompt when customer is mentioned.
    """
    parts = []

    # CRM profile
    crm = get_crm_profile(customer_name)
    if crm:
        parts.append(
            f"CRM: {crm['name']} | Type={crm['type']} | "
            f"Payment={crm['payment_terms']} | Status={crm['status']}"
            + (f" | Notes: {crm['notes']}" if crm['notes'] and crm['notes'] != 'None' else "")
        )

    # Quote stats
    stats = get_quote_stats(customer_name)
    if stats['total'] > 0:
        parts.append(
            f"Quote History: {stats['total']} total | "
            f"WIN={stats['WIN']} | LOSS={stats['LOSS']} | "
            f"PENDING={stats['PENDING']} | Win Rate={stats['win_rate']}%"
        )

    # Recent quotes (last 5)
    recent = get_quote_history(customer_name, limit=5)
    if recent:
        parts.append("Recent Quotes:")
        for q in recent:
            date_str = q['date'].strftime('%d-%b') if pd.notna(q.get('date')) else ''
            parts.append(
                f"  [{q['status']}] {q['quote_id']} {date_str} | "
                f"{q['carrier']} {q['container']} {q['pol']}→{q['place']} ${q['price']:,.0f}"
            )

    # Active jobs
    jobs = get_active_jobs(customer_name, limit=3)
    if jobs:
        parts.append("Active Jobs:")
        for j in jobs:
            etd = j['etd'].strftime('%d-%b') if pd.notna(j.get('etd')) else ''
            eta = j['eta'].strftime('%d-%b') if pd.notna(j.get('eta')) else ''
            parts.append(
                f"  [{j['status']}] {j['job_id']} | "
                f"{j['carrier']} {j['container']}×{j['quantity']} "
                f"{j['routing']} ETD={etd} ETA={eta}"
            )

    if not parts:
        return f"Không có dữ liệu ERP cho khách: {customer_name}"

    return f"=== ERP CONTEXT: {customer_name.upper()} ===\n" + "\n".join(parts)


# ─────────────────────── MONTHLY STATS (Sprint 10) ───────────────────────

def get_monthly_stats(month: str = None) -> dict:
    """
    Aggregate Active Jobs for a given month into dashboard-ready stats.

    Args:
        month: 'YYYY-MM' string — defaults to current month

    Returns dict:
        total_jobs, total_revenue, total_teu, total_customers,
        win_rate, revenue_by_carrier, customer_segments, new_customers
    """
    if month is None:
        month = datetime.now().strftime('%Y-%m')

    try:
        year, mon = int(month[:4]), int(month[5:7])
    except Exception:
        year, mon = datetime.now().year, datetime.now().month

    # ── Load Active Jobs ──
    df = _load_sheet_as_df('active', JOBS_HEADER_ROW)
    stats = {
        'month': month,
        'total_jobs': 0,
        'total_revenue': 0.0,
        'total_teu': 0,
        'total_customers': 0,
        'win_rate': 0.0,
        'revenue_by_carrier': {},
        'customer_segments': {'Direct': 0, 'Coload': 0},
        'new_customers': 0,
    }

    if df is None or df.empty:
        return stats

    # Filter to valid Job_IDs
    if 'Job_ID' in df.columns:
        df = df[df['Job_ID'].astype(str).str.match(r'J\d{8,}', na=False)]

    # Filter by ETD month
    if 'ETD' in df.columns:
        df['ETD'] = pd.to_datetime(df['ETD'], errors='coerce')
        df = df[(df['ETD'].dt.year == year) & (df['ETD'].dt.month == mon)]

    if df.empty:
        return stats

    # Basic counts
    stats['total_jobs'] = len(df)

    if 'Selling_Rate' in df.columns:
        df['Selling_Rate'] = pd.to_numeric(df['Selling_Rate'], errors='coerce').fillna(0)
    if 'Quantity' in df.columns:
        df['Quantity'] = pd.to_numeric(df['Quantity'], errors='coerce').fillna(1)

    stats['total_revenue'] = float(df['Selling_Rate'].sum()) if 'Selling_Rate' in df.columns else 0
    stats['total_teu'] = int(df['Quantity'].sum()) if 'Quantity' in df.columns else stats['total_jobs']

    # Unique customers
    if 'Customer_Name' in df.columns:
        stats['total_customers'] = df['Customer_Name'].dropna().nunique()

    # Revenue by carrier
    if 'Carrier' in df.columns and 'Selling_Rate' in df.columns:
        carrier_rev = (
            df.groupby('Carrier')['Selling_Rate']
            .sum()
            .sort_values(ascending=False)
        )
        stats['revenue_by_carrier'] = {
            str(k): float(v) for k, v in carrier_rev.items() if k and str(k) != 'nan'
        }

    # Customer segments: Direct vs Coload from Customer_Type or Contract_Type
    seg_col = None
    for col in ['Customer_Type', 'Contract_Type', 'Service_Type']:
        if col in df.columns:
            seg_col = col
            break
    if seg_col:
        types = df[seg_col].astype(str).str.upper()
        stats['customer_segments']['Direct'] = int(types.str.contains('DIRECT', na=False).sum())
        stats['customer_segments']['Coload'] = int(
            types.str.contains('COLOAD|CONSOLIDAT|CO-LOAD', na=False, regex=True).sum()
        )
        # Unknown → distribute proportionally or keep as-is
        total_seg = stats['customer_segments']['Direct'] + stats['customer_segments']['Coload']
        if total_seg == 0:
            # Fallback: even split
            half = stats['total_jobs'] // 2
            stats['customer_segments'] = {'Direct': stats['total_jobs'] - half, 'Coload': half}

    # Win rate from Quotes sheet for this month
    try:
        q_df = _load_sheet_as_df('quot', QUOTES_HEADER_ROW)
        if q_df is not None and not q_df.empty:
            if 'Quote_Date' in q_df.columns:
                q_df['Quote_Date'] = pd.to_datetime(q_df['Quote_Date'], errors='coerce')
                q_month = q_df[(q_df['Quote_Date'].dt.year == year) & (q_df['Quote_Date'].dt.month == mon)]
                total_q = len(q_month)
                if total_q > 0:
                    wins = q_month['Status'].astype(str).str.upper().eq('WIN').sum()
                    stats['win_rate'] = round(wins / total_q * 100, 1)
    except Exception:
        pass

    return stats


# ─────────────────────── PIPELINE STATS (Sprint 10b) ───────────────────────

def get_pipeline_stats(month: str = None) -> dict:
    """
    Build sales funnel statistics for a month from Quotes + Active Jobs.

    Returns dict:
        quotes     — All quotes sent this month
        bookings   — Quotes that have a Booking Number (Bkg_No filled)
        wins       — Quotes with status WIN
        shipments  — Active Jobs with ETD in this month
        win_rate   — wins / quotes * 100
    """
    if month is None:
        month = datetime.now().strftime('%Y-%m')

    try:
        year, mon = int(month[:4]), int(month[5:7])
    except Exception:
        year, mon = datetime.now().year, datetime.now().month

    result = {
        'month':     month,
        'quotes':    0,
        'bookings':  0,
        'wins':      0,
        'shipments': 0,
        'win_rate':  0.0,
    }

    # ── From Quotes sheet ──
    try:
        q_df = _load_sheet_as_df('quot', QUOTES_HEADER_ROW)
        if q_df is not None and not q_df.empty:
            if 'Quote_Date' in q_df.columns:
                q_df['Quote_Date'] = pd.to_datetime(q_df['Quote_Date'], errors='coerce')
                q_month = q_df[
                    (q_df['Quote_Date'].dt.year == year) &
                    (q_df['Quote_Date'].dt.month == mon)
                ]
                result['quotes'] = len(q_month)

                # Bookings = quotes with a Booking Number
                bkg_col = next((c for c in q_month.columns
                                if 'bkg' in c.lower() or 'booking' in c.lower()), None)
                if bkg_col:
                    result['bookings'] = int(
                        q_month[bkg_col].apply(
                            lambda x: bool(x and str(x).strip() not in ('', 'None', 'nan'))
                        ).sum()
                    )

                # Wins
                if 'Status' in q_month.columns:
                    result['wins'] = int(
                        q_month['Status'].astype(str).str.upper().eq('WIN').sum()
                    )
    except Exception as e:
        logger.error(f"[ERP Reader] get_pipeline_stats quotes: {e}")

    # ── From Active Jobs ──
    try:
        j_df = _load_sheet_as_df('active', JOBS_HEADER_ROW)
        if j_df is not None and not j_df.empty:
            if 'Job_ID' in j_df.columns:
                j_df = j_df[j_df['Job_ID'].astype(str).str.match(r'J\d{8,}', na=False)]
            if 'ETD' in j_df.columns:
                j_df['ETD'] = pd.to_datetime(j_df['ETD'], errors='coerce')
                j_month = j_df[
                    (j_df['ETD'].dt.year == year) &
                    (j_df['ETD'].dt.month == mon)
                ]
                result['shipments'] = len(j_month)
    except Exception as e:
        logger.error(f"[ERP Reader] get_pipeline_stats jobs: {e}")

    # Win rate
    if result['quotes'] > 0:
        result['win_rate'] = round(result['wins'] / result['quotes'] * 100, 1)

    return result
