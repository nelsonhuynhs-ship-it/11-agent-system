"""
erp_writer.py — Sprint 8: Write Active Jobs & Update Quote Status
Handles Quote → Active Job conversion triggered by /win command.
Follows erp-code-rules.md: only writes to designated sheets, never touches rows 1-8 of Pricing.
"""
import os
import logging
from datetime import datetime

import openpyxl

logger = logging.getLogger(__name__)

ERP_FILE = None  # set by init_writer()

# Active Jobs sheet column map (row 7 = headers, data starts row 8)
# Columns: Job_ID(1), Quote_ID(2), Customer_ID(3), Customer_Name(4), Customer_Type(5),
#          Routing(6), Bkg_No(7), Hbl_No(8), ETD(9), ETD_Original(10), ETA(11),
#          ETA_Alert_Date(12), ATA(13), Carrier(14), Contract_Type(15), Container_Type(16),
#          Quantity(17), Volume(18), Selling_Rate(19), Buying_Rate(20), ...
JOB_COL = {
    'Job_ID': 1, 'Quote_ID': 2, 'Customer_ID': 3, 'Customer_Name': 4,
    'Customer_Type': 5, 'Routing': 6, 'Bkg_No': 7, 'Hbl_No': 8,
    'ETD': 9, 'ETD_Original': 10, 'ETA': 11, 'Carrier': 14,
    'Container_Type': 16, 'Quantity': 17, 'Volume': 18,
    'Selling_Rate': 19, 'Buying_Rate': 20,
}

# Quotes sheet column map (row 2 = headers, data starts row 3)
QUOTE_COL = {
    'Quote_ID': 1, 'Quote_Date': 2, 'Customer_Name': 3, 'POL': 4,
    'POD': 5, 'Place': 6, 'Carrier': 7, 'Contract_No': 8,
    'Rate_Type': 9, 'Container_Type': 10, 'Price': 11,
    'Eff_Date': 12, 'Exp_Date': 13, 'Note': 14,
    'Status': 15, 'Status_Date': 16,
    'Win_Quantity': 19, 'Win_Volume': 20, 'Job_ID': 21,
}


def init_writer(erp_file: str):
    global ERP_FILE
    ERP_FILE = erp_file


def _find_quote_row(ws_quotes, quote_id: str) -> int | None:
    """Find the row number of a quote by Quote_ID in the Quotes sheet."""
    for row_idx, row in enumerate(ws_quotes.iter_rows(min_row=3, values_only=True), start=3):
        if row and str(row[0]).strip() == quote_id.strip():
            return row_idx
    return None


def _find_sheet(wb, keyword: str):
    """Locate sheet by partial name match (safe, no emoji dependency)."""
    for name in wb.sheetnames:
        if keyword.lower() in name.lower():
            return wb[name]
    return None


def _generate_job_id(ws_jobs) -> str:
    """Generate next Job_ID: J + YYYYMMDD + 4-digit seq."""
    today = datetime.now().strftime('%Y%m%d')
    prefix = f"J{today}"
    max_seq = -1
    for row in ws_jobs.iter_rows(min_row=8, max_col=1, values_only=True):
        val = str(row[0]) if row[0] else ''
        if val.startswith(prefix):
            try:
                seq = int(val[len(prefix):])
                max_seq = max(max_seq, seq)
            except ValueError:
                pass
    return f"{prefix}{max_seq + 1:04d}"


def create_active_job(quote_id: str, quantity: int = 1,
                      selling_rate: float = None, buying_rate: float = None) -> dict:
    """
    Convert a PENDING quote into an Active Job.

    Steps:
    1. Read quote data from Quotes sheet by quote_id
    2. Write new row to Active Jobs sheet
    3. Update Quote row: Status=WIN, Job_ID=new_id, Status_Date=today

    Returns dict with 'job_id', 'customer', 'routing', 'carrier', 'container', 'selling', 'error'.
    """
    if not ERP_FILE or not os.path.exists(ERP_FILE):
        return {'error': 'ERP file not found'}

    try:
        wb = openpyxl.load_workbook(ERP_FILE, keep_vba=True)
    except PermissionError:
        return {'error': 'ERP đang mở trong Excel. Đóng file rồi thử lại.'}
    except Exception as e:
        return {'error': f'Không mở được ERP: {str(e)[:80]}'}

    try:
        ws_quotes = _find_sheet(wb, 'quot')
        ws_jobs = _find_sheet(wb, 'active')

        if ws_quotes is None:
            return {'error': 'Không tìm thấy sheet Quotes'}
        if ws_jobs is None:
            return {'error': 'Không tìm thấy sheet Active Jobs'}

        # ── Step 1: Find quote row ──
        quote_row_idx = _find_quote_row(ws_quotes, quote_id)
        if quote_row_idx is None:
            return {'error': f'Không tìm thấy Quote_ID: {quote_id}'}

        # Read quote data
        def qval(col_name):
            col = QUOTE_COL.get(col_name, 0)
            return ws_quotes.cell(quote_row_idx, col).value if col else None

        customer = str(qval('Customer_Name') or '')
        pol = str(qval('POL') or '')
        place = str(qval('Place') or '')
        carrier = str(qval('Carrier') or '')
        container = str(qval('Container_Type') or '')
        base_price = float(qval('Price') or 0)
        routing = f"{pol}→{place}" if place else pol

        if selling_rate is None:
            selling_rate = base_price
        if buying_rate is None:
            buying_rate = base_price  # will be filled from BasicCost_Lookup if available

        # ── Step 2: Write to Active Jobs ──
        job_id = _generate_job_id(ws_jobs)
        now = datetime.now()

        # Find next empty row in Active Jobs (after header row 7)
        next_row = ws_jobs.max_row + 1
        # Verify it's truly empty
        while ws_jobs.cell(next_row, 1).value is not None:
            next_row += 1

        # Write job row
        ws_jobs.cell(next_row, JOB_COL['Job_ID']).value = job_id
        ws_jobs.cell(next_row, JOB_COL['Quote_ID']).value = quote_id
        ws_jobs.cell(next_row, JOB_COL['Customer_Name']).value = customer
        ws_jobs.cell(next_row, JOB_COL['Routing']).value = routing
        ws_jobs.cell(next_row, JOB_COL['Carrier']).value = carrier
        ws_jobs.cell(next_row, JOB_COL['Container_Type']).value = container
        ws_jobs.cell(next_row, JOB_COL['Quantity']).value = quantity
        ws_jobs.cell(next_row, JOB_COL['Volume']).value = quantity  # TEU count = qty
        ws_jobs.cell(next_row, JOB_COL['Selling_Rate']).value = selling_rate
        ws_jobs.cell(next_row, JOB_COL['Buying_Rate']).value = buying_rate
        ws_jobs.cell(next_row, JOB_COL['ETD']).value = now  # placeholder until booking

        # ── Step 3: Update Quote status ──
        ws_quotes.cell(quote_row_idx, QUOTE_COL['Status']).value = 'WIN'
        ws_quotes.cell(quote_row_idx, QUOTE_COL['Status_Date']).value = now
        ws_quotes.cell(quote_row_idx, QUOTE_COL['Win_Quantity']).value = quantity
        ws_quotes.cell(quote_row_idx, QUOTE_COL['Job_ID']).value = job_id

        wb.save(ERP_FILE)
        wb.close()

        logger.info(f"[ERP Writer] Created {job_id} from {quote_id} | {customer} {carrier} {container} ×{quantity}")

        return {
            'job_id': job_id,
            'quote_id': quote_id,
            'customer': customer,
            'routing': routing,
            'carrier': carrier,
            'container': container,
            'quantity': quantity,
            'selling': selling_rate,
            'error': None,
        }

    except Exception as e:
        wb.close()
        logger.error(f"[ERP Writer] Exception: {e}")
        return {'error': str(e)[:120]}


def update_quote_status(quote_id: str, status: str, note: str = '') -> bool:
    """Update a quote's status field directly (e.g., LOSS)."""
    if not ERP_FILE or not os.path.exists(ERP_FILE):
        return False
    try:
        wb = openpyxl.load_workbook(ERP_FILE, keep_vba=True)
        ws_quotes = _find_sheet(wb, 'quot')
        if not ws_quotes:
            wb.close()
            return False
        row_idx = _find_quote_row(ws_quotes, quote_id)
        if not row_idx:
            wb.close()
            return False
        ws_quotes.cell(row_idx, QUOTE_COL['Status']).value = status.upper()
        ws_quotes.cell(row_idx, QUOTE_COL['Status_Date']).value = datetime.now()
        if note:
            ws_quotes.cell(row_idx, QUOTE_COL['Note']).value = note
        wb.save(ERP_FILE)
        wb.close()
        return True
    except Exception as e:
        logger.error(f"[ERP Writer] update_quote_status: {e}")
        return False
