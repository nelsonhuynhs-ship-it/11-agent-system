# -*- coding: utf-8 -*-
"""
N.E.L.S.O.N v2.0 — Nelson's Enterprise Logistics & Sales Ops Network
=====================================================================
Parquet engine + Oracle memory + Sentinel monitoring + Agent architecture

Agents:
  NEXUS    — Orchestrator, routes all tasks
  ENGINE   — Pricing, rate calculations (Parquet)
  LENS     — Analytics, anomaly detection
  SENTINEL — Monitor, heartbeat checks
  ORACLE   — Memory, knowledge base (SQLite)
  NOTIFY   — Alerts, Telegram, email
"""
import logging
import logging.handlers
import sys
import os
import json
import time
from datetime import datetime, date, timedelta

# Fix Windows console encoding
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from telegram import Update
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    filters, ContextTypes
)
from telegram.constants import ChatAction
import pandas as pd
import openpyxl

from config import (
    BOT_TOKEN, ADMIN_CHAT_ID, ADMIN_NAME, GEMINI_API_KEY, GEMINI_MODEL,
    ERP_FILE, LOG_DIR
)
from database import (
    init_db, add_customer_rule, get_customer_rules, delete_customer_rule,
    get_all_customers_with_rules, get_excluded_carriers,
    get_pending_commissions, add_commission, mark_commission_paid,
    save_price_snapshot, get_previous_price
)
from ai_chat import init_gemini, chat_with_ai
from rate_limiter import rate_limiter
from query_parser import parse_rate_query

# ── Sprint 7-10b Module Imports ───────────────────────────────────────────────
from markup_engine import load_markup_from_erp, calculate_selling_price, markup_summary, is_markup_loaded
from customer_profiles import (
    get_profile, enrich_query, format_profile_header,
    list_profile_customers, get_all_lanes
)
from erp_reader import (
    init_reader, get_active_jobs, get_quote_history,
    get_quote_stats, get_crm_profile, build_full_context,
    get_monthly_stats, get_pipeline_stats
)
from erp_writer import init_writer, create_active_job, update_quote_status
from win_loss_analyzer import analyze_by_customer, analyze_by_carrier, analyze_by_route, pending_alerts
from dashboard_builder import build_dashboard
from kpi_store import (
    init_kpi, set_kpi, get_kpi, get_kpi_display, KPI_FIELDS,
    set_leads, get_leads, get_forecast
)
from config import DB_FILE
from query_parser import apply_rate_filters, format_rate_results
# ── Reorg: external formatters ────────────────────────────────────────────────
from quote_formatter import format_quotation
from freetime_formatter import (
    _is_freetime_query, _is_price_query,
    get_freetime_summary, format_freetime_answer,
)
# ── Reorg Phase 4: Parquet query engine ─────────────────────────────────────
from query_engine import (
    load_parquet, load_carrier_rules, query_parquet,
    get_parquet_loaded_time, PARQUET_FILE, CARRIER_RULES_FILE,
)
# ── Bot v6 Agentic Modules (Sprint 12) ───────────────────────────────────────
from rate_expiry_guardian import run_expiry_check, quick_summary as guardian_summary
from customer_intelligence import build_intel_card
from auto_email_booking import handle_booking_request, generate_booking_email
from nl_query_agent import dispatch_nl_query
# ── Bot v6 AI Brain Modules (Sprint 12.5) ──────────────────────────────────
from data_lake import get_lake, init_lake
from etl_sync import run_sync, format_sync_result
from ai_pricing import PricingIntelligence
from ai_sales_intel import SalesIntelligence
from ai_risk_engine import RiskEngine
# ── Bot Menu UI (Sprint 12.5) ───────────────────────────────────────────────
from bot_menu import register_menu_handlers
from intelligence_features import register_intelligence_handlers

# ── N.E.L.S.O.N CTO Agent Integration (Phase 1 Stabilization) ───────────────
# Single polling loop: bot_v5 handles Telegram, routes /task to CTO agent
import threading
_AGENT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.agent', 'agents')
if os.path.isdir(_AGENT_DIR):
    sys.path.insert(0, _AGENT_DIR)
_cto_available = False
try:
    import cto_agent
    _cto_available = True
except Exception as _cto_err:
    logging.getLogger(__name__).info("[Init] CTO Agent not loaded (bot still works): %s", _cto_err)

# ── N.E.L.S.O.N v2.0: ORACLE Memory Layer ───────────────────────────────────
try:
    from memory.oracle import Oracle
    oracle = Oracle()
    _oracle_available = True
except Exception as _oracle_err:
    oracle = None
    _oracle_available = False
    logging.getLogger(__name__).info("[Init] ORACLE not loaded (bot still works): %s", _oracle_err)


# ── Paths ─────────────────────────────────────────────────────────────────────
_THIS_DIR   = os.path.dirname(os.path.abspath(__file__))
_ENGINE_DIR = os.path.join(os.path.dirname(_THIS_DIR), "Pricing_Engine")


# ── Logging ───────────────────────────────────────────────────────────────────
log_file = os.path.join(LOG_DIR, "bot_v5.log")
log_fmt  = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")

_ch = logging.StreamHandler()
_ch.setFormatter(log_fmt)
_ch.setLevel(logging.INFO)

_fh = logging.handlers.RotatingFileHandler(
    log_file, maxBytes=5*1024*1024, backupCount=3, encoding='utf-8'
)
_fh.setFormatter(log_fmt)
_fh.setLevel(logging.INFO)

logging.basicConfig(level=logging.INFO, handlers=[_ch, _fh])
logger = logging.getLogger(__name__)

# ── Global state ──────────────────────────────────────────────────────────────
_last_quote_results: dict = {}            # chat_id → df cached for /savequote




# ══════════════════════════════════════════════════════════════════════════════
# FREETIME HELPER  (logic in freetime_formatter.py)
# ══════════════════════════════════════════════════════════════════════════════

def _get_freetime_summary_local(carrier: str, container: str, pol: str = "HPH") -> str:
    """Thin wrapper — loads rules and delegates to freetime_formatter."""
    rules = load_carrier_rules()
    return get_freetime_summary(carrier, container, rules, pol)





# ══════════════════════════════════════════════════════════════════════════════
# QUOTATION FORMATTER  (logic in quote_formatter.py + freetime_formatter.py)
# ══════════════════════════════════════════════════════════════════════════════
# format_quotation, _smart_note → quote_formatter.py
# _is_freetime_query, get_freetime_summary, format_freetime_answer, _is_price_query → freetime_formatter.py
# carrier advisory notes → carrier_tips.json


# CORE COMMANDS
# ══════════════════════════════════════════════════════════════════════════════

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /start."""
    df = load_parquet()
    count = f"{len(df):,}" if df is not None else "N/A"
    txt = (
        f"Chao *{ADMIN_NAME}*!\n\n"
        f"Em la Nelson Freight Bot v5 (The Logistics Agent).\n"
        f"Du lieu hien tai: *{count}* active rates.\n\n"
        f"*Tra gia (free text):*\n"
        f"  `gia hph lax 40hq` — top 3 options\n"
        f"  `hph chicago soc cont 40` — loc SOC\n\n"
        f"*Lenh:*\n"
        f"  /quote HPH-ATLANTA — tra gia nhanh\n"
        f"  /status — tong quan he thong\n"
        f"  /remember CUSTOMER rule — luu yeu cau KH\n"
        f"  /com — commission pending\n"
        f"  /briefing — bao cao sang\n"
        f"  /help — huong dan day du\n"
    )
    await update.message.reply_text(txt, parse_mode="Markdown")


async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /help."""
    txt = (
        f"*NELSON FREIGHT BOT v4*\n"
        f"\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n\n"
        f"*Tra gia (tu do — khong can /quote):*\n"
        f"  `gia hph lax` — top 3 options\n"
        f"  `hph chicago via tacoma soc` — loc POD + SOC\n"
        f"  `hcm el paso 40hq` — spec container\n\n"
        f"*/quote [query]* — tra gia nhanh\n"
        f"*/status* — tong quan data + AI\n"
        f"*/reload* — refresh Parquet data\n\n"
        f"*Customer Memory:*\n"
        f"  /remember CUSTOMER no ZIM\n"
        f"  /remember CUSTOMER prefer CMA ONE\n"
        f"  /customer NAME — xem profile\n"
        f"  /forget NAME ID — xoa rule\n\n"
        f"*Commission:*\n"
        f"  /com — pending list\n"
        f"  /com add JOB CUS CARRIER CONT QTY AMT\n"
        f"  /com paid ID\n\n"
        f"*ERP:*\n"
        f"  /savequote CUSTOMER 1 2 3\n"
        f"  /quotes [CUSTOMER]\n"
        f"  /wins  /losses\n\n"
        f"*Briefing:* /briefing\n"
    )
    await update.message.reply_text(txt, parse_mode="Markdown")


async def cmd_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /status."""
    df = load_parquet()
    if df is None:
        await update.message.reply_text("Khong load duoc Parquet.")
        return

    carriers = df['Carrier'].nunique()
    carrier_list = ", ".join(sorted(df['Carrier'].dropna().unique()[:8]))
    places  = df['Place'].nunique()
    max_exp = df['Exp'].max().strftime('%d-%b-%Y') if pd.notna(df['Exp'].max()) else "N/A"

    customers = get_all_customers_with_rules()
    pending   = get_pending_commissions()
    com_total = sum(c['total'] for c in pending) if pending else 0
    ai_usage  = rate_limiter.usage_stats()

    txt = (
        f"*SYSTEM STATUS — Bot v5*\n"
        f"\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
        f"Active rates: *{len(df):,}*\n"
        f"Carriers ({carriers}): {carrier_list}...\n"
        f"Places: *{places}*  |  Max Exp: {max_exp}\n\n"
        f"Customers with rules: *{len(customers)}*\n"
        f"Pending com: *{len(pending)}* (${com_total:,.0f})\n\n"
        f"AI: {ai_usage['day_used']}/{ai_usage['day_limit']} today ({ai_usage['model']})\n"
        f"Data loaded: {get_parquet_loaded_time().strftime('%H:%M:%S') if get_parquet_loaded_time() else 'N/A'}\n"
    )
    await update.message.reply_text(txt, parse_mode="Markdown")


async def cmd_reload(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Force reload Parquet."""
    await update.message.reply_text("Dang reload Parquet...")
    df = load_parquet(force=True)
    if df is not None:
        await update.message.reply_text(f"Reload xong: *{len(df):,}* active rates", parse_mode="Markdown")
    else:
        await update.message.reply_text("Khong load duoc Parquet. Kiem tra file.")


# ══════════════════════════════════════════════════════════════════════════════
# /quote COMMAND (wrapper around query engine)
# ══════════════════════════════════════════════════════════════════════════════

async def cmd_quote(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/quote HPH-ATLANTA, /quote HPH LAX SOC cont 40, etc."""
    if not context.args:
        await update.message.reply_text(
            "*Cach dung:*\n"
            "`/quote HPH-ATLANTA`\n"
            "`/quote HCM-CHICAGO VIA TACOMA`\n"
            "`/quote HPH LAX SOC`\n"
            "`/quote HCM-LAX cont 20`",
            parse_mode="Markdown"
        )
        return

    query = " ".join(context.args)
    df = load_parquet()
    if df is None:
        await update.message.reply_text("Khong load duoc du lieu.")
        return

    known_carriers = list(df['Carrier'].dropna().unique())
    parsed = parse_rate_query(query, known_carriers)
    if not parsed.get('pol'):
        parsed['pol'] = 'HPH'

    # Apply customer exclusions
    customers_with_rules = [r['customer'] for r in get_all_customers_with_rules()]
    excluded_note = ""
    for term in list(parsed.get('place_terms', [])):
        if term in customers_with_rules:
            parsed['customer'] = term
            parsed['place_terms'].remove(term)
            excluded = get_excluded_carriers(term)
            if excluded:
                excluded_note = f"\n[!] {term}: loai {', '.join(excluded)}"

    await update.message.chat.send_action(ChatAction.TYPING)
    container = parsed.get('container', '40HQ')
    results   = query_parquet(parsed, top_n=3)

    # Apply customer exclusions
    if parsed.get('customer') and results is not None:
        excluded = get_excluded_carriers(parsed['customer'])
        if excluded:
            results = results[~results['Carrier'].str.upper().isin([e.upper() for e in excluded])]

    chat_id = update.effective_chat.id
    if results is not None and not results.empty:
        _last_quote_results[chat_id] = (results, container, parsed)

    text = format_quotation(results, container, parsed, freetime_fn=_get_freetime_summary_local)
    if excluded_note:
        text += excluded_note

    try:
        await update.message.reply_text(text, parse_mode="Markdown")
    except Exception:
        await update.message.reply_text(text)


# ══════════════════════════════════════════════════════════════════════════════
# CUSTOMER MEMORY (unchanged from v3)
# ══════════════════════════════════════════════════════════════════════════════

async def cmd_remember(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Save customer rule."""
    if not context.args or len(context.args) < 2:
        await update.message.reply_text(
            "*Cach dung:*\n"
            "`/remember PANDA no ZIM`\n"
            "`/remember CREATIVE ship before 25/03`\n"
            "`/remember SIRI prefer CMA ONE`\n"
            "`/remember PANDA note can CO form E`",
            parse_mode="Markdown"
        )
        return
    customer  = context.args[0].upper()
    rest      = " ".join(context.args[1:])
    rest_up   = rest.upper()

    if rest_up.startswith("NO ") or rest_up.startswith("KHONG "):
        carrier = rest.split(maxsplit=1)[1].strip().upper()
        add_customer_rule(customer, "exclude_carrier", carrier)
        await update.message.reply_text(f"Da luu: *{customer}* — Khong di *{carrier}*", parse_mode="Markdown")
    elif "SHIP BEFORE" in rest_up or "XUAT TRUOC" in rest_up:
        date_str = rest.split("before")[-1].strip() if "before" in rest.lower() else rest.split("truoc")[-1].strip()
        add_customer_rule(customer, "ship_before", date_str)
        await update.message.reply_text(f"Da luu: *{customer}* — Ship truoc *{date_str}*", parse_mode="Markdown")
    elif "RECEIVE BEFORE" in rest_up or "NHAN TRUOC" in rest_up:
        date_str = rest.split("before")[-1].strip() if "before" in rest.lower() else rest.split("truoc")[-1].strip()
        add_customer_rule(customer, "receive_before", date_str)
        await update.message.reply_text(f"Da luu: *{customer}* — Nhan truoc *{date_str}*", parse_mode="Markdown")
    elif rest_up.startswith("PREFER ") or rest_up.startswith("UU TIEN "):
        carriers = rest.split(maxsplit=1)[1].strip().upper()
        add_customer_rule(customer, "prefer_carrier", carriers)
        await update.message.reply_text(f"Da luu: *{customer}* — Uu tien *{carriers}*", parse_mode="Markdown")
    elif rest_up.startswith("NOTE ") or rest_up.startswith("GHI CHU "):
        note_val = rest.split(maxsplit=1)[1].strip()
        add_customer_rule(customer, "note", note_val)
        await update.message.reply_text(f"Da luu cho *{customer}*: {note_val}", parse_mode="Markdown")
    else:
        add_customer_rule(customer, "note", rest)
        await update.message.reply_text(f"Da luu cho *{customer}*: {rest}", parse_mode="Markdown")


async def cmd_customer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show customer profile."""
    if not context.args:
        await update.message.reply_text("Dung: `/customer PANDA`", parse_mode="Markdown")
        return
    customer = context.args[0].upper()
    rules = get_customer_rules(customer)
    if not rules:
        await update.message.reply_text(f"*{customer}* — Chua co rules nao.", parse_mode="Markdown")
        return
    labels = {"exclude_carrier": "Khong di", "prefer_carrier": "Uu tien",
               "ship_before": "Ship truoc", "receive_before": "Nhan truoc", "note": "Note"}
    lines = [f"*{customer}* — {len(rules)} rules\n"]
    for r in rules:
        label = labels.get(r['rule_type'], r['rule_type'])
        lines.append(f"  {label}: *{r['rule_value']}* (ID:{r['id']})")
    lines.append(f"\nXoa rule: `/forget {customer} ID`")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def cmd_customers(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """List all customers."""
    customers = get_all_customers_with_rules()
    if not customers:
        await update.message.reply_text("Chua co customer nao co rules.")
        return
    lines = ["*CUSTOMERS* co rules:\n"]
    for c in customers:
        lines.append(f"  - *{c['customer']}* — {c['rule_count']} rules")
    lines.append("\nXem chi tiet: `/customer NAME`")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def cmd_forget(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Delete a customer rule."""
    if not context.args or len(context.args) < 2:
        await update.message.reply_text("Dung: `/forget CUSTOMER ID`", parse_mode="Markdown")
        return
    try:
        rule_id = int(context.args[1])
        delete_customer_rule(rule_id)
        await update.message.reply_text(f"Da xoa rule ID:{rule_id}")
    except ValueError:
        await update.message.reply_text("ID phai la so.")


# ══════════════════════════════════════════════════════════════════════════════
# COMMISSION
# ══════════════════════════════════════════════════════════════════════════════

async def cmd_com(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/com — /com add ... — /com paid ID"""
    args = context.args
    if not args:
        pending = get_pending_commissions()
        if not pending:
            await update.message.reply_text(f"Khong co commission pending nao, {ADMIN_NAME}!")
            return
        total = sum(c['total'] for c in pending)
        lines = [f"*COMMISSION PENDING* — {len(pending)} lo\n"]
        for c in pending:
            lines.append(f"  `ID:{c['id']}` {c['customer']} | {c['carrier']} {c['container']}x{c['quantity']} = *${c['total']:,.0f}* ({c['created_at'][:10]})")
        lines.append(f"\n*Total: ${total:,.0f}*")
        lines.append("Danh dau paid: `/com paid ID`")
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
    elif args[0].lower() == "add" and len(args) >= 7:
        try:
            job_id, customer, carrier, container = args[1], args[2].upper(), args[3].upper(), args[4].upper()
            qty, amount = int(args[5]), float(args[6])
            add_commission(job_id, customer, carrier, container, qty, amount)
            await update.message.reply_text(
                f"Com da tao:\n  {job_id} | {customer} | {carrier} {container}\n  ${amount} x {qty} = *${amount*qty:,.0f}*",
                parse_mode="Markdown"
            )
        except (ValueError, IndexError):
            await update.message.reply_text("Sai format. VD: `/com add J08/03-01 PANDA CMA 40HQ 2 75`", parse_mode="Markdown")
    elif args[0].lower() == "paid" and len(args) >= 2:
        try:
            com_id = int(args[1])
            mark_commission_paid(com_id)
            await update.message.reply_text(f"Com ID:{com_id} da danh dau *PAID*!", parse_mode="Markdown")
        except ValueError:
            await update.message.reply_text("ID phai la so.")
    else:
        await update.message.reply_text(
            "*Commission:*\n`/com` — Xem pending\n`/com add JOB CUS CARRIER CONT QTY AMT`\n`/com paid ID`",
            parse_mode="Markdown"
        )


# ══════════════════════════════════════════════════════════════════════════════
# SPRINT 7: MARKUP ENGINE
# ══════════════════════════════════════════════════════════════════════════════

async def cmd_markup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/markup — Show current markup settings loaded from ERP_Master.xlsm."""
    try:
        if not is_markup_loaded():
            ok = load_markup_from_erp(ERP_FILE)
            if not ok:
                await update.message.reply_text(
                    f"❌ Markup chưa load được.\nKiểm tra ERP file: `{ERP_FILE}`",
                    parse_mode="Markdown"
                )
                return
        summary = markup_summary()
        await update.message.reply_text(
            f"📊 **Markup Engine — Current Settings**\n\n{summary}\n\n"
            f"💡 Selling = Base + Global Markup + Carrier Markup + PUC (SOC)",
            parse_mode="Markdown"
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Lỗi load markup: {str(e)[:150]}")


# ══════════════════════════════════════════════════════════════════════════════
# BRIEFING
# ══════════════════════════════════════════════════════════════════════════════

async def generate_briefing() -> str:
    """Generate morning briefing."""
    df = load_parquet(force=True)
    lines = [f"*MORNING BRIEFING* — {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"]
    if df is not None:
        lines.append(f"Parquet active rates: *{len(df):,}*\n")


    # Sprint 10b: Daily KPI section
    try:
        from datetime import datetime as _dt
        _month = _dt.now().strftime('%Y-%m')
        _stats = get_monthly_stats(_month)
        _targets = get_kpi(_month)
        if _targets:
            _jobs = _stats['total_jobs']
            _wr   = _stats['win_rate']
            _rev  = _stats['total_revenue']
            _t_s  = _targets.get('shipments')
            _t_wr = _targets.get('win_rate')
            _t_r  = _targets.get('revenue')
            kpi_lines = [f"🎯 *Daily KPI — {_dt.now().strftime('%d %b')}*"]
            if _t_s:
                _fc = get_forecast(_jobs, _t_s, _month)
                _b = '█' * int(_jobs/_t_s*10) + '░' * (10 - int(_jobs/_t_s*10))
                kpi_lines.append(f"🚢 Shipments: `{_b}` *{_jobs}/{_t_s:.0f}* "
                                 f"({_jobs/_t_s*100:.0f}%) {_fc['icon']}")
                kpi_lines.append(f"   Forecast EOM: ~{_fc['projected_eom']:.0f} lô | "
                                 f"Còn {_fc['days_remaining']} ngày")
            if _t_wr:
                _wr_icon = '🟢' if _wr >= _t_wr*0.8 else ('🟡' if _wr >= _t_wr*0.5 else '🔴')
                kpi_lines.append(f"🎯 Win Rate: *{_wr:.1f}%* / {_t_wr:.0f}% {_wr_icon}")
            if _t_r:
                kpi_lines.append(f"💰 Revenue: *${_rev:,.0f}* / ${_t_r:,.0f}")
            lines.append("\n".join(kpi_lines) + "\n")
    except Exception as _e:
        pass

    # Price drops from Parquet
    if df is not None:
        try:
            sample = df.sort_values('Amount').drop_duplicates(subset=['Carrier', 'POD', 'Place', 'Container_Type'], keep='first').head(100)
            drops = []
            for _, row in sample.iterrows():
                carrier = str(row.get('Carrier', ''))
                cont    = str(row.get('Container_Type', ''))
                pol_v   = str(row.get('POL', ''))
                pod_v   = str(row.get('POD', ''))
                place_v = str(row.get('Place', ''))
                current = float(row.get('Amount', 0) or 0)
                if current <= 0:
                    continue
                save_price_snapshot(carrier, cont, pol_v, pod_v, place_v, current)
                prev = get_previous_price(carrier, cont, pod_v, place_v)
                if prev and prev['price'] > 0:
                    diff = current - prev['price']
                    if diff < -30:
                        drops.append({'carrier': carrier, 'place': place_v, 'cont': cont,
                                      'old': prev['price'], 'new': current, 'drop': abs(diff)})
            if drops:
                drops.sort(key=lambda x: x['drop'], reverse=True)
                lines.append("*GIA GIAM (co hoi re-quote):*")
                for d in drops[:5]:
                    lines.append(f"  - {d['carrier']} {d['cont']} -> {d['place']}: *-${d['drop']:,.0f}* (${d['old']:,.0f} -> ${d['new']:,.0f})")
                lines.append("")
        except Exception as exc:
            logger.error(f"Briefing price drop error: {exc}")

    # Commission
    pending = get_pending_commissions()
    if pending:
        total = sum(c['total'] for c in pending)
        lines.append(f"*COM PENDING*: {len(pending)} lo (*${total:,.0f}*)")
        for c in pending[:3]:
            lines.append(f"  - {c['customer']} | {c['carrier']} {c['container']}x{c['quantity']} = ${c['total']:,.0f}")
        if len(pending) > 3:
            lines.append(f"  ... va {len(pending)-3} lo nua")
        lines.append("")

    # Deadline rules
    customers = get_all_customers_with_rules()
    deadline_rules = []
    for cust in customers:
        for r in get_customer_rules(cust['customer']):
            if r['rule_type'] in ('ship_before', 'receive_before'):
                deadline_rules.append(f"  - {cust['customer']}: {r['rule_value']}")
    if deadline_rules:
        lines.append("*DEADLINES KHACH HANG:*")
        lines.extend(deadline_rules)

    if len(lines) <= 2:
        lines.append(f"Chua co gi can chu y, {ADMIN_NAME}!")
    return "\n".join(lines)


async def cmd_briefing(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle /briefing."""
    await update.message.reply_text("Dang tao briefing...")
    text = await generate_briefing()
    await update.message.reply_text(text, parse_mode="Markdown")


# ════════════════════════════════════════════
# SPRINT 8: CRM & JOB INTEGRATION
# ════════════════════════════════════════════

async def cmd_win(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/win QUOTE_ID [qty] — Convert quote to Active Job."""
    if not context.args:
        await update.message.reply_text(
            "📝 **Cách dùng:**\n"
            "`/win QUOTE_ID [qty]` — Chuyển quote sang Active Job\n\n"
            "Ví dụ: `/win 10MAR-5 2`",
            parse_mode="Markdown"
        )
        return

    quote_id = context.args[0].upper()
    quantity = 1
    if len(context.args) >= 2:
        try:
            quantity = int(context.args[1])
        except ValueError:
            pass

    await update.message.reply_text(f"⏳ Đang tạo Active Job từ {quote_id}...")

    result = create_active_job(quote_id, quantity)
    if result.get('error'):
        await update.message.reply_text(f"❌ Lỗi: {result['error']}")
        return

    msg = (
        f"✅ **Job tạo thành công!**\n\n"
        f"📋 Job ID: **{result['job_id']}**\n"
        f"📌 Quote: {result['quote_id']}\n"
        f"🏢 Customer: **{result['customer']}**\n"
        f"🗺️ Route: {result['routing']}\n"
        f"🚢 {result['carrier']} {result['container']} × {result['quantity']}\n"
        f"💰 Selling: **${result['selling']:,.0f}**\n\n"
        f"📊 Xem jobs: `/jobs {result['customer']}`"
    )
    await update.message.reply_text(msg, parse_mode="Markdown")
    logger.info(f"WIN job created: {result['job_id']} from {quote_id}")


async def cmd_crm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/crm CUSTOMER — 360° customer view."""
    if not context.args:
        await update.message.reply_text("📝 Dùng: `/crm PANDA`", parse_mode="Markdown")
        return

    customer = ' '.join(context.args).upper()
    await update.message.reply_text(f"🔍 Đang load dữ liệu cho {customer}...")

    lines = [f"🏢 **CRM — {customer}**\n━━━━━━━━━━━━━━━━━━━━"]

    crm = get_crm_profile(customer)
    if crm:
        lines.append(
            f"📋 ID: {crm['id']} | Type: **{crm['type']}**\n"
            f"💳 Payment: {crm['payment_terms']} | Status: {crm['status']}"
        )
        if crm['notes'] and crm['notes'] not in ('None', ''):
            lines.append(f"📝 Notes: _{crm['notes']}_")

    stats = get_quote_stats(customer)
    if stats['total'] > 0:
        lines.append(
            f"\n📊 **Quotes:** {stats['total']} total | "
            f"WIN={stats['WIN']} | LOSS={stats['LOSS']} | "
            f"PENDING={stats['PENDING']} | Win Rate=**{stats['win_rate']}%**"
        )

    recent = get_quote_history(customer, limit=5)
    if recent:
        lines.append("\n📜 **5 Quotes gần nhất:**")
        for q in recent:
            date_str = q['date'].strftime('%d-%b') if hasattr(q.get('date'), 'strftime') else ''
            status_icon = {"WIN": "✅", "LOSS": "❌", "PENDING": "⏳"}.get(q['status'].upper(), "❓")
            lines.append(
                f" {status_icon} `{q['quote_id']} {date_str} | {q['carrier']} {q['container']} "
                f"{q['pol']}→{q['place']} ${q['price']:,.0f}`"
            )

    jobs = get_active_jobs(customer, limit=3)
    if jobs:
        lines.append("\n🚢 **Active Jobs:**")
        for j in jobs:
            etd = j['etd'].strftime('%d-%b') if hasattr(j.get('etd'), 'strftime') else ''
            eta = j['eta'].strftime('%d-%b') if hasattr(j.get('eta'), 'strftime') else ''
            lines.append(
                f" 📦 `{j['job_id']} | {j['carrier']} {j['container']}×{j['quantity']} "
                f"ETD={etd} ETA={eta} [{j['status']}]`"
            )

    if len(lines) <= 2:
        lines.append("\n⚠️ Không tìm thấy dữ liệu ERP cho khách này.")
        lines.append(f"💡 Thêm rules: `/remember {customer} ...`")

    try:
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
    except Exception:
        await update.message.reply_text("\n".join(lines))


async def cmd_jobs(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/jobs [CUSTOMER] — List active jobs."""
    customer = ' '.join(context.args).upper() if context.args else None
    label = f"cho **{customer}**" if customer else "(tất cả)"

    jobs = get_active_jobs(customer_name=customer, limit=10)
    if not jobs:
        await update.message.reply_text(f"📋 Không có Active Jobs {label}.", parse_mode="Markdown")
        return

    lines = [f"🚢 **ACTIVE JOBS** {label} — {len(jobs)} lô\n━━━━━━━━━━━━━━━━━━━━"]
    for j in jobs:
        etd = j['etd'].strftime('%d-%b') if hasattr(j.get('etd'), 'strftime') else '?'
        eta = j['eta'].strftime('%d-%b') if hasattr(j.get('eta'), 'strftime') else '?'
        profit = j['selling'] - j['buying']
        lines.append(
            f"`{j['job_id'][:13]} {j['customer'][:10]}\n"
            f" {j['carrier']:<5} {j['container']:<5}×{j['quantity']} ETD={etd} ETA={eta} P=+${profit:,.0f}`"
        )
    lines.append("\n💡 Chi tiết: `/crm CUSTOMER_NAME`")
    try:
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
    except Exception:
        await update.message.reply_text("\n".join(lines))


async def cmd_history(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/history CUSTOMER — Quote history with win/loss stats."""
    if not context.args:
        await update.message.reply_text("📝 Dùng: `/history PANDA`", parse_mode="Markdown")
        return

    customer = ' '.join(context.args).upper()
    quotes = get_quote_history(customer, limit=15)
    stats = get_quote_stats(customer)

    if not quotes:
        await update.message.reply_text(f"📋 Không có quotes cho **{customer}**.", parse_mode="Markdown")
        return

    lines = [
        f"📜 **HISTORY — {customer}**",
        f"Win Rate: **{stats['win_rate']}%** | WIN={stats['WIN']} LOSS={stats['LOSS']} PENDING={stats['PENDING']}",
        "━━━━━━━━━━━━━━━━━━━━"
    ]
    status_icons = {"WIN": "✅", "LOSS": "❌", "PENDING": "⏳", "AUTO-LOST": "🔴"}
    for q in quotes:
        date_str = q['date'].strftime('%d-%b') if hasattr(q.get('date'), 'strftime') else ''
        icon = status_icons.get(q['status'].upper(), "❓")
        job_tag = f"→{q['job_id']}" if q['job_id'] else ""
        lines.append(
            f"{icon} `{q['quote_id']:<12} {date_str:<6} | "
            f"{q['carrier']:<5} {q['container']:<5} ${q['price']:>7,.0f} "
            f"{q['pol']}→{q['place'][:12]}{job_tag}`"
        )
    lines.append("\n💡 `/win QUOTE_ID` — Chuyển PENDING → Active Job")
    try:
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
    except Exception:
        await update.message.reply_text("\n".join(lines))



# ════════════════════════════════════════════
# SPRINT 10: VISUAL DASHBOARD & KPI
# ════════════════════════════════════════════

async def cmd_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/report [YYYY-MM] — Generate monthly dashboard image."""
    from datetime import datetime as dt
    month = context.args[0] if context.args else dt.now().strftime('%Y-%m')

    # Validate format
    try:
        dt.strptime(month, '%Y-%m')
    except ValueError:
        await update.message.reply_text("❌ Format: `/report 2026-03`", parse_mode="Markdown")
        return

    await update.message.reply_text(f"📊 Đang tạo dashboard {month}...")

    try:
        stats = get_monthly_stats(month)
        kpi_targets = get_kpi(month)
        img_buf = build_dashboard(stats, kpi_targets, month)

        caption = (
            f"📊 **Freight Report — {month}**\n"
            f"🚢 {stats['total_jobs']} lô | 📦 {stats['total_teu']} TEU | "
            f"💰 ${stats['total_revenue']:,.0f} | 🎯 Win {stats['win_rate']}%"
        )
        await update.message.reply_photo(photo=img_buf, caption=caption, parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"❌ Lỗi tạo report: {str(e)[:150]}")
        logger.error(f"[Report] {e}")


async def cmd_setkpi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/setkpi FIELD VALUE [YYYY-MM] — Set KPI target."""
    valid_fields = ', '.join(f"`{f}`" for f in KPI_FIELDS)
    if len(context.args) < 2:
        await update.message.reply_text(
            f"🎯 **Set KPI Target**\n\n"
            f"Cú pháp: `/setkpi FIELD VALUE`\n\n"
            f"Fields: {valid_fields}\n\n"
            f"Ví dụ:\n"
            f"`/setkpi shipments 60`\n"
            f"`/setkpi revenue 200000`\n"
            f"`/setkpi win_rate 40`\n"
            f"`/setkpi new_customers 3`",
            parse_mode="Markdown"
        )
        return

    field = context.args[0].lower()
    try:
        target = float(context.args[1])
    except ValueError:
        await update.message.reply_text("❌ Value phải là số. Ví dụ: `/setkpi shipments 60`", parse_mode="Markdown")
        return

    month = context.args[2] if len(context.args) >= 3 else None
    ok = set_kpi(field, target, month)

    if not ok:
        await update.message.reply_text(
            f"❌ Field không hợp lệ: `{field}`\nCác field hợp lệ: {valid_fields}",
            parse_mode="Markdown"
        )
        return

    from kpi_store import KPI_FIELDS as KF
    label, unit = KF.get(field, (field, ''))
    fmt = f"${target:,.0f}" if field == 'revenue' else f"{target:,.0f} {unit}"
    mon_display = month or __import__('datetime').datetime.now().strftime('%Y-%m')
    await update.message.reply_text(
        f"✅ **KPI đã set!**\n{label}: **{fmt}** (tháng {mon_display})\n\n"
        f"Xem tiến độ: `/kpi`",
        parse_mode="Markdown"
    )


async def cmd_kpi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/kpi [YYYY-MM] — Show KPI progress vs actuals."""
    from datetime import datetime as dt
    month = context.args[0] if context.args else dt.now().strftime('%Y-%m')

    targets = get_kpi(month)
    stats = get_monthly_stats(month)

    if not targets:
        no_kpi = get_kpi_display(month)
        await update.message.reply_text(no_kpi, parse_mode="Markdown")
        return

    lines = [f"🎯 **KPI Progress — {month}**", "━━━━━━━━━━━━━━━━━━━━"]

    actuals = {
        'shipments':     stats['total_jobs'],
        'revenue':       stats['total_revenue'],
        'win_rate':      stats['win_rate'],
        'new_customers': stats.get('new_customers', 0),
    }
    labels = {
        'shipments':     ('🚢 Shipments', 'lô'),
        'revenue':       ('💰 Revenue', 'USD'),
        'win_rate':      ('🎯 Win Rate', '%'),
        'new_customers': ('🏢 New KH', 'KH'),
    }

    for field, (label, unit) in labels.items():
        target = targets.get(field)
        actual = actuals.get(field, 0)
        if target is None:
            continue

        pct = min(actual / target * 100, 100) if target > 0 else 0
        bar_filled = int(pct / 10)
        bar = '█' * bar_filled + '░' * (10 - bar_filled)
        status = '✅' if pct >= 100 else ('🟡' if pct >= 70 else '🔴')

        fmt_a = f"${actual:,.0f}" if field == 'revenue' else f"{actual:,.0f}"
        fmt_t = f"${target:,.0f}" if field == 'revenue' else f"{target:,.0f}"
        lines.append(
            f"{status} **{label}**\n"
            f"   `{bar}` {pct:.0f}%\n"
            f"   {fmt_a} / {fmt_t} {unit}"
        )

    lines.append("\n💡 Cập nhật target: `/setkpi FIELD VALUE`")
    lines.append("📊 Dashboard ảnh: `/report`")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")




# ════════════════════════════════════════════
# SPRINT 10b: KPI INTELLIGENCE
# ════════════════════════════════════════════

async def cmd_forecast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/forecast [YYYY-MM] — Project end-of-month KPI based on current pace."""
    from datetime import datetime as dt
    month = context.args[0] if context.args else dt.now().strftime('%Y-%m')

    try:
        dt.strptime(month, '%Y-%m')
    except ValueError:
        await update.message.reply_text("❌ Format: `/forecast 2026-03`", parse_mode="Markdown")
        return

    targets = get_kpi(month)
    stats = get_monthly_stats(month)
    actual_jobs = stats['total_jobs']
    actual_rev = stats['total_revenue']

    if not targets:
        await update.message.reply_text(
            f"⚠️ Chưa có KPI target cho {month}.\nSet trước: `/setkpi shipments 60`",
            parse_mode="Markdown"
        )
        return

    lines = [f"📊 **KPI Forecast — {month}**", "━━━━━━━━━━━━━━━━━━━━"]

    # Shipments forecast
    target_s = targets.get('shipments')
    if target_s:
        fc = get_forecast(actual_jobs, target_s, month)
        lines.append(
            f"🚢 **Shipments**\n"
            f"   Hiện tại: **{actual_jobs:,}** lô (ngày {fc['days_elapsed']}/{fc['days_total']})\n"
            f"   Pace avg: {fc['daily_avg']:.1f} lô/ngày\n"
            f"   Dự báo EOM: ~**{fc['projected_eom']:.0f}** / {target_s:.0f} lô\n"
            f"   {fc['icon']} Achievement: **{fc['pct_of_target']:.0f}%**"
        )

    # Revenue forecast
    target_r = targets.get('revenue')
    if target_r:
        fc_r = get_forecast(actual_rev, target_r, month)
        lines.append(
            f"\n💰 **Revenue**\n"
            f"   Hiện tại: **${actual_rev:,.0f}** / ${target_r:,.0f}\n"
            f"   Dự báo EOM: ~**${fc_r['projected_eom']:,.0f}**\n"
            f"   {fc_r['icon']} Achievement: **{fc_r['pct_of_target']:.0f}%**"
        )

    # Win rate (no projection — use actual)
    target_wr = targets.get('win_rate')
    if target_wr:
        actual_wr = stats['win_rate']
        icon = '🟢' if actual_wr >= target_wr * 0.8 else ('🟡' if actual_wr >= target_wr * 0.5 else '🔴')
        lines.append(
            f"\n🎯 **Win Rate**\n"
            f"   Hiện tại: **{actual_wr:.1f}%** / {target_wr:.0f}%\n"
            f"   {icon} {'On track' if icon == '🟢' else 'At risk' if icon == '🟡' else 'Critical'}"
        )

    if target_s and fc['days_remaining'] > 0:
        remaining = max(target_s - actual_jobs, 0)
        need_per_day = remaining / fc['days_remaining'] if fc['days_remaining'] > 0 else 0
        lines.append(
            f"\n⚡ **Action needed:**\n"
            f"   Còn {fc['days_remaining']} ngày | Cần thêm {remaining:.0f} lô\n"
            f"   → Cần close **{need_per_day:.1f} lô/ngày** để đạt target"
        )

    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def cmd_pipeline(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/pipeline [YYYY-MM] — Show sales funnel: Leads → Quotes → Bookings → Shipments."""
    from datetime import datetime as dt
    month = context.args[0] if context.args else dt.now().strftime('%Y-%m')

    try:
        dt.strptime(month, '%Y-%m')
    except ValueError:
        await update.message.reply_text("❌ Format: `/pipeline 2026-03`", parse_mode="Markdown")
        return

    await update.message.reply_text(f"🔄 Đang tải pipeline {month}...")

    pipe = get_pipeline_stats(month)
    leads = get_leads(month)

    lines = [f"📊 **Pipeline — {month}**", "━━━━━━━━━━━━━━━━━━━━"]

    # Funnel bars (scale to 12 chars width)
    top = max(leads, pipe['quotes'], pipe['bookings'], pipe['shipments'], 1)

    def bar(n):
        filled = round(n / top * 12) if top > 0 else 0
        return '█' * filled + '░' * (12 - filled)

    lines.append(f"👥 Leads:      {leads:>3}  `{bar(leads)}`")
    lines.append(f"📋 Quotes:     {pipe['quotes']:>3}  `{bar(pipe['quotes'])}`")
    lines.append(f"✅ Bookings:   {pipe['bookings']:>3}  `{bar(pipe['bookings'])}`")
    lines.append(f"🚢 Shipments:  {pipe['shipments']:>3}  `{bar(pipe['shipments'])}`")
    lines.append("")

    # Conversion rates
    if pipe['quotes'] > 0 and leads > 0:
        qr = pipe['quotes'] / leads * 100
        lines.append(f"📈 Lead→Quote: **{qr:.0f}%**")
    if pipe['bookings'] > 0 and pipe['quotes'] > 0:
        br = pipe['bookings'] / pipe['quotes'] * 100
        lines.append(f"📈 Quote→Booking: **{br:.0f}%**")
    if pipe['shipments'] > 0 and pipe['quotes'] > 0:
        wr = pipe['shipments'] / pipe['quotes'] * 100
        lines.append(f"🎯 Win rate: **{wr:.0f}%**")

    lines.append("\n💡 Update leads: `/setleads 48`")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def cmd_setleads(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/setleads N [YYYY-MM] — Set manual lead count for pipeline tracking."""
    if not context.args:
        await update.message.reply_text(
            "📝 Dùng: `/setleads 48` hoặc `/setleads 48 2026-03`",
            parse_mode="Markdown"
        )
        return
    try:
        count = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ N phải là số nguyên. Ví dụ: `/setleads 48`", parse_mode="Markdown")
        return

    month = context.args[1] if len(context.args) >= 2 else None
    ok = set_leads(count, month)
    mon_display = month or __import__('datetime').datetime.now().strftime('%Y-%m')
    if ok:
        await update.message.reply_text(
            f"✅ **Leads set:** {count} leads cho {mon_display}\n"
            f"Xem funnel: `/pipeline`",
            parse_mode="Markdown"
        )
    else:
        await update.message.reply_text("❌ Lỗi lưu leads.")





async def cmd_analyze(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sprint 9: /analyze — Win/Loss AI Deep-Dive.

    Usage:
      /analyze PANDA             → Customer analysis
      /analyze carrier CMA       → Carrier performance
      /analyze route HPH DENVER  → Lane market intel
      /analyze pending           → Overdue follow-ups
    """
    if not context.args:
        await update.message.reply_text(
            "🤖 **AI Win/Loss Analysis**\n\n"
            "`/analyze PANDA` — Customer deep-dive\n"
            "`/analyze carrier CMA` — Carrier performance\n"
            "`/analyze route HPH DENVER` — Lane intel\n"
            "`/analyze pending` — Quotes quá hạn follow-up",
            parse_mode="Markdown"
        )
        return

    args = context.args
    mode = args[0].lower()

    await update.message.reply_text("🤖 Đang phân tích... (AI đang xử lý)")

    try:
        if mode == 'pending':
            result = await pending_alerts()

        elif mode == 'carrier' and len(args) >= 2:
            carrier = args[1].upper()
            result = await analyze_by_carrier(carrier)

        elif mode == 'route' and len(args) >= 3:
            pol = args[1].upper()
            place = ' '.join(args[2:]).upper()
            result = await analyze_by_route(pol, place)

        else:
            # Default: customer analysis (first arg = customer name)
            customer = ' '.join(args).upper()
            result = await analyze_by_customer(customer)

        # Telegram message limit is 4096 chars
        if len(result) > 4000:
            result = result[:3990] + "\n... _(xem chi tiết trong ERP)_"

        try:
            await update.message.reply_text(result, parse_mode="Markdown")
        except Exception:
            await update.message.reply_text(result)

    except Exception as e:
        await update.message.reply_text(f"❌ Lỗi phân tích: {str(e)[:150]}")
        logger.error(f"[Analyze] Error: {e}")



# ════════════════════════════════════════════
# SPRINT 10: VISUAL DASHBOARD & KPI
# ════════════════════════════════════════════

async def send_morning_briefing(context: ContextTypes.DEFAULT_TYPE):
    """Scheduled 7:30 briefing."""
    try:
        text = await generate_briefing()
        await context.bot.send_message(chat_id=ADMIN_CHAT_ID, text=text, parse_mode="Markdown")
        logger.info("Morning briefing sent!")
    except Exception as exc:
        logger.error(f"Briefing error: {exc}")


# ══════════════════════════════════════════════════════════════════════════════
# ERP INTEGRATION (/savequote /quotes /wins /losses)
# ══════════════════════════════════════════════════════════════════════════════

async def cmd_savequote(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/savequote CUSTOMER 1 2 3 — Save selected options to ERP Quotes sheet."""
    if not context.args or len(context.args) < 2:
        await update.message.reply_text(
            "Dung: `/savequote CUSTOMER 1 2 3`\nChay `/quote` truoc de xem gia.",
            parse_mode="Markdown"
        )
        return
    chat_id = update.effective_chat.id
    if chat_id not in _last_quote_results:
        await update.message.reply_text("Chua co ket qua /quote. Chay /quote truoc.")
        return

    results_df, container, parsed = _last_quote_results[chat_id]
    customer = context.args[0].upper()
    try:
        indices = [int(x) for x in context.args[1:]]
    except ValueError:
        await update.message.reply_text("So thu tu phai la so. VD: `/savequote PANDA 1 2`")
        return

    if not os.path.exists(ERP_FILE):
        await update.message.reply_text("Khong tim thay ERP_Master.xlsm")
        return
    try:
        wb = openpyxl.load_workbook(ERP_FILE, keep_vba=True)
    except Exception as exc:
        if 'locked' in str(exc).lower() or 'permission' in str(exc).lower():
            await update.message.reply_text("ERP_Master.xlsm dang mo trong Excel. Dong file roi thu lai.")
        else:
            await update.message.reply_text(f"Loi mo ERP: {str(exc)[:100]}")
        return

    qs_name = next((s for s in wb.sheetnames if 'quot' in s.lower()), None)
    if not qs_name:
        await update.message.reply_text("Khong tim thay sheet Quotes trong ERP")
        return

    ws       = wb[qs_name]
    next_row = ws.max_row + 1
    now      = datetime.now()
    quote_id = f"{now.strftime('%d%b').upper()}-{next_row - 2}"
    saved    = 0
    rows_list = list(results_df.iterrows())

    for idx in indices:
        if idx < 1 or idx > len(rows_list):
            continue
        _, row = rows_list[idx - 1]
        ws.cell(next_row, 1, quote_id)
        ws.cell(next_row, 2, now)
        ws.cell(next_row, 3, customer)
        ws.cell(next_row, 4, str(row.get('POL', '')))
        ws.cell(next_row, 5, str(row.get('POD', '')))
        ws.cell(next_row, 6, str(row.get('Place', '')))
        ws.cell(next_row, 7, str(row.get('Carrier', '')))
        ws.cell(next_row, 9, str(row.get('Commodity', '')))
        ws.cell(next_row, 10, container)
        try:
            ws.cell(next_row, 11, float(row.get('Amount', 0)))
        except Exception:
            ws.cell(next_row, 11, row.get('Amount', 0))
        try:
            ws.cell(next_row, 12, pd.to_datetime(row.get('Eff', ''), errors='coerce'))
            ws.cell(next_row, 13, pd.to_datetime(row.get('Exp', ''), errors='coerce'))
        except Exception:
            pass
        ws.cell(next_row, 14, str(row.get('Note', '')))
        ws.cell(next_row, 15, 'PENDING')
        ws.cell(next_row, 16, now)
        next_row += 1
        saved += 1

    wb.save(ERP_FILE)
    wb.close()
    await update.message.reply_text(
        f"Da luu *{saved}* quotes cho *{customer}*\n"
        f"Quote ID: *{quote_id}*\n"
        f"Xem: `/quotes {customer}`",
        parse_mode="Markdown"
    )


async def cmd_quotes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/quotes [CUSTOMER] — List recent quotes."""
    if not os.path.exists(ERP_FILE):
        await update.message.reply_text("Khong tim thay ERP_Master.xlsm")
        return
    try:
        wb = openpyxl.load_workbook(ERP_FILE, read_only=True, keep_vba=True)
    except Exception:
        await update.message.reply_text("Khong doc duoc ERP file")
        return
    qs_name = next((s for s in wb.sheetnames if 'quot' in s.lower()), None)
    if not qs_name:
        wb.close()
        await update.message.reply_text("Khong tim thay sheet Quotes")
        return

    ws              = wb[qs_name]
    customer_filter = context.args[0].upper() if context.args else None
    quotes = []
    for r in range(3, ws.max_row + 1):
        qid = ws.cell(r, 1).value
        if not qid:
            continue
        cust = str(ws.cell(r, 3).value or '')
        if customer_filter and customer_filter not in cust.upper():
            continue
        quotes.append({
            'id': qid, 'customer': cust,
            'carrier': ws.cell(r, 7).value,
            'container': ws.cell(r, 10).value,
            'price': ws.cell(r, 11).value,
            'place': ws.cell(r, 6).value,
            'status': ws.cell(r, 15).value or 'PENDING',
        })
    wb.close()

    if not quotes:
        await update.message.reply_text(f"Khong co quotes" + (f" cho {customer_filter}" if customer_filter else ""))
        return

    icons = {'WIN': 'WIN', 'LOST': 'LOST', 'PENDING': '...'}
    title = f"*QUOTES* ({len(quotes)} total)"
    if customer_filter:
        title += f" — {customer_filter}"
    lines = [title + "\n"]
    for q in reversed(quotes[-10:]):
        icon = icons.get(str(q['status']).upper(), '...')
        try:
            price_str = f"${int(float(q['price'])):,}"
        except Exception:
            price_str = str(q['price'])
        lines.append(f"[{icon}] `{q['id']}` {q['customer']} | {q['carrier']} {q['container']} {price_str}")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def cmd_wins(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/wins — WIN quotes."""
    if not os.path.exists(ERP_FILE):
        await update.message.reply_text("Khong tim thay ERP_Master.xlsm")
        return
    try:
        wb = openpyxl.load_workbook(ERP_FILE, read_only=True, keep_vba=True)
    except Exception:
        await update.message.reply_text("Khong doc duoc ERP file")
        return
    qs_name = next((s for s in wb.sheetnames if 'quot' in s.lower()), None)
    if not qs_name:
        wb.close()
        await update.message.reply_text("Khong tim thay sheet Quotes")
        return
    ws = wb[qs_name]
    wins = []
    for r in range(3, ws.max_row + 1):
        if str(ws.cell(r, 15).value or '').upper() == 'WIN':
            wins.append({
                'id': ws.cell(r, 1).value,
                'customer': ws.cell(r, 3).value,
                'carrier': ws.cell(r, 7).value,
                'container': ws.cell(r, 10).value,
                'price': ws.cell(r, 11).value,
                'qty': ws.cell(r, 19).value or 1,
            })
    wb.close()
    if not wins:
        await update.message.reply_text("Chua co quote WIN nao.")
        return
    total_rev = sum(float(w['price'] or 0) * int(w['qty'] or 1) for w in wins)
    lines = [f"*WIN QUOTES* — {len(wins)} quotes\n"]
    for w in wins[-10:]:
        try:
            price_str = f"${int(float(w['price'])):,}"
        except Exception:
            price_str = str(w['price'])
        lines.append(f"  [WIN] `{w['id']}` {w['customer']} | {w['carrier']} {w['container']}x{w['qty']} {price_str}")
    lines.append(f"\n*Total Revenue: ${total_rev:,.0f}*")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def cmd_losses(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/losses — Analyze LOST quotes."""
    if not os.path.exists(ERP_FILE):
        await update.message.reply_text("Khong tim thay ERP_Master.xlsm")
        return
    try:
        wb = openpyxl.load_workbook(ERP_FILE, read_only=True, keep_vba=True)
    except Exception:
        await update.message.reply_text("Khong doc duoc ERP file")
        return
    qs_name = next((s for s in wb.sheetnames if 'quot' in s.lower()), None)
    if not qs_name:
        wb.close()
        await update.message.reply_text("Khong tim thay sheet Quotes")
        return
    ws = wb[qs_name]
    losses = []
    for r in range(3, ws.max_row + 1):
        if str(ws.cell(r, 15).value or '').upper() == 'LOST':
            losses.append({'customer': ws.cell(r, 3).value, 'carrier': ws.cell(r, 7).value})
    wb.close()
    if not losses:
        await update.message.reply_text("Chua co quote LOST nao.")
        return
    from collections import Counter
    by_cust    = Counter(l['customer'] for l in losses)
    by_carrier = Counter(l['carrier'] for l in losses)
    lines = [f"*LOST ANALYSIS* — {len(losses)} quotes\n"]
    lines.append("*By Customer:*")
    for cust, count in by_cust.most_common(5):
        lines.append(f"  {cust}: {count}")
    lines.append("\n*By Carrier:*")
    for carrier, count in by_carrier.most_common(5):
        lines.append(f"  {carrier}: {count}")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


# ══════════════════════════════════════════════════════════════════════════════
# FREE TEXT HANDLER
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# /task COMMAND — N.E.L.S.O.N CTO Agent
# ══════════════════════════════════════════════════════════════════════════════

async def cmd_task(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/task DESCRIPTION — Route task to N.E.L.S.O.N CTO Agent."""
    if not _cto_available:
        await update.message.reply_text(
            "❌ CTO Agent chưa sẵn sàng.\n"
            "Kiểm tra `.agent/agents/` folder."
        )
        return
    if not context.args:
        await update.message.reply_text(
            "🧠 **N.E.L.S.O.N CTO Agent**\n"
            "Dùng: `/task rebuild ERP staging`\n"
            "Hoặc: `/task check CRM sheet`\n\n"
            "Agent sẽ: Plan → Backup → Build → Review → Report",
            parse_mode="Markdown"
        )
        return
    task_desc = " ".join(context.args)
    await update.message.reply_text(f"🧠 NEXUS: Analyzing task: {task_desc[:60]}...")
    # Run CTO agent in background thread (it's synchronous)
    def _run_cto():
        try:
            cto_agent.handle_command(f"/task {task_desc}")
        except Exception as e:
            logger.error(f"[CTO] Error: {e}")
    threading.Thread(target=_run_cto, daemon=True).start()


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Route free text:
      - Detected as price query → ENGINE (Parquet)
      - CTO Agent commands → NEXUS
      - Everything else → Gemini AI fallback
    ORACLE: saves all messages for context memory.
    """
    text = update.message.text.strip()
    if not text:
        return

    # ORACLE: remember user message
    user_id = str(update.effective_user.id)
    if _oracle_available:
        try:
            username = update.effective_user.first_name or update.effective_user.username or "?"
            oracle.remember(user_id, "user", text)
            oracle.upsert_profile(user_id, username=username)
        except Exception:
            pass

    await update.message.chat.send_action(ChatAction.TYPING)

    # Route 0: CTO Agent slash-like commands in free text
    text_lower = text.lower()
    if _cto_available and text_lower.startswith(("/task ", "/status", "/pause", "/rollback", "/approve", "/reject", "/health", "/backlog", "/test")):
        # These are CTO agent commands — route to N.E.L.S.O.N
        def _run_cto():
            try:
                cto_agent.handle_command(text)
            except Exception as e:
                logger.error(f"[CTO] Error: {e}")
        threading.Thread(target=_run_cto, daemon=True).start()
        return

    # Route 1: freetime-only query (DEM/DET/Power charge)
    if _is_freetime_query(text):
        rules = load_carrier_rules()
        response = format_freetime_answer(text, rules)
        try:
            await update.message.reply_text(response, parse_mode="Markdown")
        except Exception:
            await update.message.reply_text(response)
        # ORACLE: save so follow-up "còn ONE?" knows context
        if _oracle_available:
            try:
                oracle.remember(user_id, "assistant", response[:2000], agent="ENGINE")
            except Exception:
                pass
        return

    # Route 2: price query
    if _is_price_query(text):
        df = load_parquet()
        if df is None:
            await update.message.reply_text("Khong load duoc du lieu Parquet.")
            return
        known_carriers = list(df['Carrier'].dropna().unique())
        parsed = parse_rate_query(text, known_carriers)
        if not parsed.get('pol'):
            parsed['pol'] = 'HPH'

        container = parsed.get('container', '40HQ')
        results   = query_parquet(parsed, top_n=3)

        chat_id = update.effective_chat.id
        if results is not None and not results.empty:
            _last_quote_results[chat_id] = (results, container, parsed)

        response = format_quotation(results, container, parsed)
        try:
            await update.message.reply_text(response, parse_mode="Markdown")
        except Exception:
            await update.message.reply_text(response)
        # ORACLE: save so follow-up questions know the price context
        if _oracle_available:
            try:
                oracle.remember(user_id, "assistant", response[:2000], agent="ENGINE")
            except Exception:
                pass
        return

    # Fallback: Gemini AI
    if not GEMINI_API_KEY:
        await update.message.reply_text(
            f"AI chat chua kich hoat. Dung /help de xem cac lenh co san."
        )
        return

    can_go, limit_msg = rate_limiter.can_request()
    if not can_go:
        await update.message.reply_text(limit_msg)
        return

    df       = load_parquet()
    # ORACLE: inject context for AI
    ai_context = ""
    if _oracle_available:
        try:
            ai_context = oracle.build_context(user_id)
        except Exception:
            pass
    response = await chat_with_ai(text, pricing_df=df, oracle_context=ai_context)
    if response:
        rate_limiter.record_request()
        # ORACLE: remember assistant reply
        if _oracle_available:
            try:
                oracle.remember(user_id, "assistant", response[:2000], agent="NEXUS")
            except Exception:
                pass
        if len(response) > 4000:
            for i in range(0, len(response), 4000):
                await update.message.reply_text(response[i:i+4000])
        else:
            try:
                await update.message.reply_text(response, parse_mode="Markdown")
            except Exception:
                await update.message.reply_text(response)
    else:
        await update.message.reply_text("AI khong phan hoi. Thu lai sau.")


async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Log and notify errors."""
    logger.error(f"Error: {context.error}")
    if update and update.effective_message:
        await update.effective_message.reply_text(f"Loi: {str(context.error)[:200]}")



# ══════════════════════════════════════════════════════════════════════════════
# BOT v6 AGENTIC COMMANDS (Sprint 12)
# ══════════════════════════════════════════════════════════════════════════════

async def cmd_guardian(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/guardian — Proactive Rate Expiry Guardian (Bot v6 Feature #1)."""
    await update.message.reply_text("🔍 Đang scan rates sắp hết hạn...")
    df = load_parquet()
    if df is None:
        await update.message.reply_text("❌ Không load được Parquet data.")
        return
    jobs = get_active_jobs(limit=50)
    sent = await run_expiry_check(
        update.message.bot,
        update.effective_chat.id,
        df, jobs
    )
    if not sent:
        summary = guardian_summary(df)
        await update.message.reply_text(
            f"✅ Tất cả clear!\n{summary}\n\nKhông có rate nào hết hạn trong 14 ngày tới."
        )


async def cmd_intel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/intel CUSTOMER — Customer 360° Intelligence (Bot v6 Feature #6)."""
    if not context.args:
        await update.message.reply_text(
            "📊 **Customer 360° Intelligence**\n"
            "Dùng: `/intel HML` hoặc `/intel SIRI`\n\n"
            "Hiển thị: Profile · Win rate · Active jobs · Rate opportunities · Negotiation playbook",
            parse_mode="Markdown"
        )
        return

    customer = " ".join(context.args).upper()
    await update.message.reply_text(f"🔍 Đang phân tích {customer}...")

    try:
        df      = load_parquet()
        crm     = get_crm_profile(customer)
        quotes  = get_quote_history(customer, limit=50)
        jobs    = get_active_jobs(customer_name=customer, limit=10)

        # Get static profile if available
        from customer_profiles import get_profile
        static_profile = get_profile(customer)

        card = build_intel_card(
            customer_name=customer,
            crm_profile=crm,
            quote_history=quotes,
            active_jobs=jobs,
            parquet_df=df,
            static_profile=static_profile,
        )
        try:
            await update.message.reply_text(card, parse_mode="Markdown")
        except Exception:
            await update.message.reply_text(card)
    except Exception as e:
        await update.message.reply_text(f"❌ Lỗi: {str(e)[:200]}")
        logger.error(f"[cmd_intel] {e}")


async def cmd_book(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/book QUOTE_ID — Auto-Email Booking Agent (Bot v6 Feature #2)."""
    if not context.args:
        await update.message.reply_text(
            "📧 **Auto Booking Email**\n"
            "Dùng: `/book QUOTE_ID`\n\n"
            "Bot sẽ:\n"
            "1. Lấy thông tin quote từ ERP\n"
            "2. Generate email booking template cho carrier\n"
            "3. Mở Outlook draft (nếu có) hoặc gửi nội dung email qua Telegram",
            parse_mode="Markdown"
        )
        return

    quote_id = context.args[0].upper()
    await update.message.reply_text(f"📝 Đang tạo booking email cho {quote_id}...")

    try:
        # Find quote details from ERP
        all_quotes = get_quote_history(limit=200)
        quote = next(
            (q for q in all_quotes if quote_id.upper() in str(q.get('quote_id', '')).upper()),
            None
        )
        if not quote:
            await update.message.reply_text(
                f"❌ Không tìm thấy quote `{quote_id}` trong ERP.\n"
                f"Dùng `/quotes` để xem danh sách quotes.",
                parse_mode="Markdown"
            )
            return

        # Build job dict for email generation
        job_dict = {
            'job_id':    quote.get('quote_id', quote_id),
            'quote_id':  quote.get('quote_id', quote_id),
            'customer':  quote.get('customer', 'N/A'),
            'carrier':   quote.get('carrier', 'N/A'),
            'container': quote.get('container', '40HQ'),
            'quantity':  1,
            'pol':       quote.get('pol', 'HPH'),
            'pod':       quote.get('pod', 'USTIW'),
            'place':     quote.get('place', 'N/A'),
            'routing':   f"{quote.get('pol','HPH')} → {quote.get('place','N/A')}",
            'commodity': 'General Cargo',
            'weight':    'Per PKG',
            'etd':       quote.get('exp_date'),
            'service_type': 'COC',
        }

        await handle_booking_request(update.message.bot, update.effective_chat.id, job_dict)

    except Exception as e:
        await update.message.reply_text(f"❌ Lỗi tạo booking email: {str(e)[:200]}")
        logger.error(f"[cmd_book] {e}")


async def cmd_ask(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/ask QUESTION — Natural Language Database Query (Bot v6 Feature #10)."""
    if not context.args:
        await update.message.reply_text(
            "💬 **Natural Language Query**\n"
            "Hỏi bất kỳ câu hỏi nào về data:\n\n"
            "Ví dụ:\n"
            "`/ask Carrier nào margin cao nhất?`\n"
            "`/ask Khách nào chưa order 30 ngày?`\n"
            "`/ask Giá tuyến Denver đang thế nào?`\n"
            "`/ask CMA vs ONE win rate ai cao hơn?`\n"
            "`/ask Route nào hay đi nhất?`",
            parse_mode="Markdown"
        )
        return

    question = " ".join(context.args)
    await update.message.reply_text(f"🤔 Đang phân tích: \"{question}\"...")

    try:
        df         = load_parquet()
        all_jobs   = get_active_jobs(limit=100)
        all_quotes = get_quote_history(limit=500)

        answer = dispatch_nl_query(
            question=question,
            parquet_df=df,
            all_jobs=all_jobs,
            all_quotes=all_quotes,
        )
        try:
            await update.message.reply_text(answer, parse_mode="Markdown")
        except Exception:
            await update.message.reply_text(answer)

    except Exception as e:
        await update.message.reply_text(f"❌ Lỗi query: {str(e)[:200]}")
        logger.error(f"[cmd_ask] {e}")


# ══════════════════════════════════════════════════════════════════════════════
# BOT v6 AI BRAIN COMMANDS (Sprint 12.5)
# ══════════════════════════════════════════════════════════════════════════════

async def cmd_sync(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/sync — Force ETL sync ERP → DataLake."""
    await update.message.reply_text("🔄 Đang sync ERP → DataLake...")
    df = load_parquet(force=True)
    result = run_sync(df, ERP_FILE)
    await update.message.reply_text(format_sync_result(result))


async def cmd_predict(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/predict POL PLACE CUSTOMER — AI đề xuất giá tối ưu."""
    if not context.args or len(context.args) < 2:
        await update.message.reply_text(
            "🧠 **Pricing Intelligence**\n"
            "Dùng: `/predict HPH Denver HML`\n\n"
            "AI sẽ phân tích market + win history → đề xuất giá tối ưu",
            parse_mode="Markdown"
        )
        return
    pol      = context.args[0].upper()
    place    = context.args[1].title()
    customer = context.args[2].upper() if len(context.args) >= 3 else "GENERAL"
    await update.message.reply_text(f"🧠 Đang phân tích {pol}→{place} cho {customer}...")
    try:
        lake = get_lake()
        pi   = PricingIntelligence(lake)
        result = pi.suggest(pol=pol, place=place, customer=customer)
        text   = pi.format_suggestion(result)
        await update.message.reply_text(text)
    except Exception as e:
        await update.message.reply_text(f"❌ Lỗi: {str(e)[:200]}")
        logger.error(f"[cmd_predict] {e}")


async def cmd_whywon(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/whywon QUOTE_ID — AI giải thích tại sao thắng/thua."""
    if not context.args:
        await update.message.reply_text(
            "🔍 **Win/Loss Analysis**\n"
            "Dùng: `/whywon 10MAR-5`\n\n"
            "AI phân tích tại sao thắng/thua quote dựa trên market data + history",
            parse_mode="Markdown"
        )
        return
    quote_id = context.args[0].upper()
    await update.message.reply_text(f"🔍 Đang phân tích {quote_id}...")
    try:
        all_quotes = get_quote_history(limit=500)
        quote = next(
            (q for q in all_quotes if quote_id in str(q.get('quote_id', '')).upper()),
            None
        )
        if not quote:
            await update.message.reply_text(f"❌ Không tìm thấy quote `{quote_id}`", parse_mode="Markdown")
            return
        lake = get_lake()
        si   = SalesIntelligence(lake)
        text = si.explain_quote(quote)
        await update.message.reply_text(text)
    except Exception as e:
        await update.message.reply_text(f"❌ Lỗi: {str(e)[:200]}")
        logger.error(f"[cmd_whywon] {e}")


async def cmd_reachout(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/reachout — Danh sách khách hàng nguy cơ churn."""
    await update.message.reply_text("📣 Đang scan customer churn risk...")
    try:
        lake = get_lake()
        si   = SalesIntelligence(lake)
        from customer_profiles import list_profile_customers
        codes = list_profile_customers() or []
        # Also add customers from quote history
        all_quotes = get_quote_history(limit=500)
        for q in all_quotes:
            c = str(q.get('customer', '')).upper()
            if c and c not in codes:
                codes.append(c)
        churn_list = si.detect_churn(codes)
        text = si.format_reachout_list(churn_list)
        await update.message.reply_text(text)
    except Exception as e:
        await update.message.reply_text(f"❌ Lỗi: {str(e)[:200]}")
        logger.error(f"[cmd_reachout] {e}")


async def cmd_risk(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/risk CUSTOMER — Multi-dimension risk assessment."""
    if not context.args:
        await update.message.reply_text(
            "⚡ **Risk Assessment**\n"
            "Dùng: `/risk HML` hoặc `/risk SIRI`\n\n"
            "4 chiều: Weight · Rate Expiry · Space · Payment",
            parse_mode="Markdown"
        )
        return
    customer = " ".join(context.args).upper()
    await update.message.reply_text(f"⚡ Đang assess risk cho {customer}...")
    try:
        df = load_parquet()
        lake = get_lake()
        re = RiskEngine(lake, df)
        assessment = re.assess_customer(customer)
        text = re.format_risk_card(assessment)
        await update.message.reply_text(text)
    except Exception as e:
        await update.message.reply_text(f"❌ Lỗi: {str(e)[:200]}")
        logger.error(f"[cmd_risk] {e}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# N.E.L.S.O.N v2.0 COMMANDS (/memory /profile /clearhistory)
# ══════════════════════════════════════════════════════════════════════════════

async def cmd_memory(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/memory — Show ORACLE memory stats."""
    if not _oracle_available:
        await update.message.reply_text("❌ ORACLE not available.")
        return
    s = oracle.stats()
    text = (
        "🧠 **ORACLE Memory Stats**\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"💬 Messages: {s['total_messages']}\n"
        f"👥 Users: {s['unique_users']}\n"
        f"📇 Profiles: {s['profiles']}\n"
        f"📋 Pending tasks: {s['pending_tasks']}\n"
        f"📦 Total tasks: {s['total_tasks']}\n"
        f"💾 DB size: {s['db_size_kb']} KB"
    )
    await update.message.reply_text(text, parse_mode="Markdown")


async def cmd_profile(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/profile — Show your ORACLE profile."""
    if not _oracle_available:
        await update.message.reply_text("❌ ORACLE not available.")
        return
    user_id = str(update.effective_user.id)
    p = oracle.get_profile(user_id)
    if not p:
        await update.message.reply_text("📇 No profile yet. Send a message to start tracking.")
        return
    text = (
        "📇 **Your Profile**\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"👤 Username: {p.get('username', '?')}\n"
        f"🏷️ Segment: {p.get('segment', 'unknown')}\n"
        f"⚡ Risk: {p.get('risk_level', 'normal')}\n"
        f"📊 Deals: {p.get('deal_count', 0)}\n"
        f"🛣️ Top route: {p.get('top_route', 'N/A')}\n"
        f"🕐 Last seen: {p.get('last_seen', 'N/A')}"
    )
    await update.message.reply_text(text, parse_mode="Markdown")


async def cmd_clearhistory(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/clearhistory — Clear your ORACLE conversation memory."""
    if not _oracle_available:
        await update.message.reply_text("❌ ORACLE not available.")
        return
    user_id = str(update.effective_user.id)
    deleted = oracle.forget(user_id)
    await update.message.reply_text(
        f"🗑️ Cleared {deleted} messages from memory."
    )


def main():
    logger.info("=" * 60)
    logger.info("  N.E.L.S.O.N v2.0 — Enterprise Logistics & Sales Ops")
    logger.info("  NEXUS · ENGINE · LENS · SENTINEL · ORACLE · NOTIFY")
    logger.info("=" * 60)

    init_db()

    # ── Sprint 7-10b: Init ERP reader/writer + KPI + Markup ───────────────────
    try:
        init_reader(ERP_FILE)
        logger.info(f"[Init] ERP Reader: {ERP_FILE}")
    except Exception as e:
        logger.warning(f"[Init] ERP Reader failed: {e}")

    try:
        init_writer(ERP_FILE)
        logger.info("[Init] ERP Writer ready")
    except Exception as e:
        logger.warning(f"[Init] ERP Writer failed: {e}")

    try:
        init_kpi(DB_FILE)
        logger.info("[Init] KPI Store ready")
    except Exception as e:
        logger.warning(f"[Init] KPI Store failed: {e}")

    try:
        ok = load_markup_from_erp(ERP_FILE)
        logger.info(f"[Init] Markup Engine: {'OK' if ok else 'no data'}")
    except Exception as e:
        logger.warning(f"[Init] Markup failed: {e}")

    # Pre-load Parquet
    df = load_parquet()
    if df is not None:
        logger.info(f"[Init] ENGINE: Parquet loaded ({len(df)} rows)")
    else:
        logger.warning("Parquet not found — check PARQUET_FILE path")

    # ── N.E.L.S.O.N v2.0: ORACLE init ────────────────────────────────────────
    if _oracle_available:
        logger.info(f"[Init] ORACLE: {oracle.stats()}")
    else:
        logger.warning("[Init] ORACLE: not available")

    # Pre-load carrier rules
    load_carrier_rules()

    # AI Brain: Init DataLake (Sprint 12.5)
    try:
        sync_result = run_sync(df, ERP_FILE)
        if sync_result['ok']:
            logger.info(f"[Init] AI Brain DataLake ready: {sync_result['rates']} rates, {sync_result['quotes']} quotes")
        else:
            logger.warning(f"[Init] AI Brain partial: {sync_result.get('error','')}")
    except Exception as e:
        logger.warning(f"[Init] AI Brain failed (bot still works): {e}")

    # Init Gemini AI
    if GEMINI_API_KEY:
        if init_gemini():
            logger.info("[Init] Gemini AI ready (%s)", GEMINI_MODEL)
        else:
            logger.warning("[Init] Gemini AI init failed")
    else:
        logger.info("[Init] AI chat disabled (no GEMINI_API_KEY)")

    # ── N.E.L.S.O.N CTO Agent: start specialist threads ──────────────────────
    if _cto_available:
        try:
            cto_agent.start_specialists()
            logger.info("[Init] N.E.L.S.O.N CTO Agent specialists started")
        except Exception as e:
            logger.warning(f"[Init] CTO Agent specialists failed (bot still works): {e}")

    app = Application.builder().token(BOT_TOKEN).build()

    # Commands
    app.add_handler(CommandHandler("start",      cmd_start))
    app.add_handler(CommandHandler("help",       cmd_help))
    app.add_handler(CommandHandler("status",     cmd_status))
    app.add_handler(CommandHandler("reload",     cmd_reload))
    app.add_handler(CommandHandler("quote",      cmd_quote))
    app.add_handler(CommandHandler("remember",   cmd_remember))
    app.add_handler(CommandHandler("customer",   cmd_customer))
    app.add_handler(CommandHandler("customers",  cmd_customers))
    app.add_handler(CommandHandler("forget",     cmd_forget))
    app.add_handler(CommandHandler("com",        cmd_com))
    app.add_handler(CommandHandler("briefing",   cmd_briefing))
    app.add_handler(CommandHandler("savequote",  cmd_savequote))
    app.add_handler(CommandHandler("quotes",     cmd_quotes))
    app.add_handler(CommandHandler("wins",       cmd_wins))
    app.add_handler(CommandHandler("losses",     cmd_losses))

    # Sprint 7
    app.add_handler(CommandHandler("markup",     cmd_markup))
    # Sprint 8
    app.add_handler(CommandHandler("crm",        cmd_crm))
    app.add_handler(CommandHandler("jobs",       cmd_jobs))
    app.add_handler(CommandHandler("history",    cmd_history))
    app.add_handler(CommandHandler("win",        cmd_win))
    # Sprint 9
    app.add_handler(CommandHandler("analyze",    cmd_analyze))
    # Sprint 10
    app.add_handler(CommandHandler("report",     cmd_report))
    app.add_handler(CommandHandler("setkpi",     cmd_setkpi))
    app.add_handler(CommandHandler("kpi",        cmd_kpi))
    # Sprint 10b
    app.add_handler(CommandHandler("forecast",   cmd_forecast))
    app.add_handler(CommandHandler("pipeline",   cmd_pipeline))
    app.add_handler(CommandHandler("setleads",   cmd_setleads))
    # Sprint 12 — Bot v6 Agentic
    app.add_handler(CommandHandler("guardian",   cmd_guardian))
    app.add_handler(CommandHandler("intel",      cmd_intel))
    app.add_handler(CommandHandler("book",       cmd_book))
    app.add_handler(CommandHandler("ask",        cmd_ask))
    # Sprint 12.5 — AI Brain
    app.add_handler(CommandHandler("sync",       cmd_sync))
    app.add_handler(CommandHandler("predict",    cmd_predict))
    app.add_handler(CommandHandler("whywon",     cmd_whywon))
    app.add_handler(CommandHandler("reachout",   cmd_reachout))
    app.add_handler(CommandHandler("risk",       cmd_risk))

    # N.E.L.S.O.N CTO Agent
    if _cto_available:
        app.add_handler(CommandHandler("task",       cmd_task))

    # N.E.L.S.O.N v2.0 — ORACLE commands
    app.add_handler(CommandHandler("memory",       cmd_memory))
    app.add_handler(CommandHandler("profile",      cmd_profile))
    app.add_handler(CommandHandler("clearhistory", cmd_clearhistory))

    # Sprint 12.5b — Email Intelligence Features (new analytics commands)
    register_intelligence_handlers(app, skip_commands=["intel", "risk"])

    # Sprint 12.5 — Menu UI
    register_menu_handlers(app)

    # Free text
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    app.add_error_handler(error_handler)

    # Morning briefing 7:30 VN time
    jq = app.job_queue
    if jq:
        from datetime import time as dtime
        jq.run_daily(send_morning_briefing, time=dtime(hour=7, minute=30), name="morning_briefing")
        logger.info("Morning briefing scheduled at 07:30 daily")
        # v6: Rate Expiry Guardian — daily 06:00
        async def _scheduled_guardian(ctx):
            df = load_parquet()
            jobs = get_active_jobs(limit=50)
            await run_expiry_check(ctx.bot, ADMIN_CHAT_ID, df, jobs)
        jq.run_daily(_scheduled_guardian, time=dtime(hour=6, minute=0), name="rate_expiry_guardian")
        logger.info("Rate Expiry Guardian scheduled at 06:00 daily")
        # v6 AI Brain: ETL Sync — daily 05:30
        async def _scheduled_etl_sync(ctx):
            df = load_parquet(force=True)
            sync_result = run_sync(df, ERP_FILE)
            if sync_result['ok']:
                await ctx.bot.send_message(ADMIN_CHAT_ID, format_sync_result(sync_result))
        jq.run_daily(_scheduled_etl_sync, time=dtime(hour=5, minute=30), name="etl_sync")
        logger.info("ETL Sync scheduled at 05:30 daily")
        # N.E.L.S.O.N v2.0: SENTINEL heartbeat — 08:00 daily
        async def _scheduled_sentinel(ctx):
            try:
                from agents.sentinel import Sentinel
                sentinel = Sentinel()
                sentinel.morning_briefing()
            except Exception as e:
                logger.error(f"[SENTINEL] Briefing failed: {e}")
        jq.run_daily(_scheduled_sentinel, time=dtime(hour=8, minute=0), name="sentinel_briefing")
        logger.info("SENTINEL briefing scheduled at 08:00 daily")

        # N.E.L.S.O.N v2.0: Rate prediction refresh — 07:45 daily (before SENTINEL)
        async def _scheduled_rate_prediction(ctx):
            try:
                from intelligence.rate_predictor import RatePredictor
                predictor = RatePredictor()
                forecasts = predictor.top_routes_forecast()
                # Cache in Oracle for morning briefing
                if _oracle_available and forecasts:
                    import json as _json
                    oracle.upsert_profile(
                        "system_cache",
                        notes=_json.dumps(forecasts, ensure_ascii=False)[:500]
                    )
                logger.info(f"[RatePredictor] {len(forecasts)} routes forecasted")
            except Exception as e:
                logger.warning(f"[RatePredictor] Refresh failed: {e}")
        jq.run_daily(_scheduled_rate_prediction, time=dtime(hour=7, minute=45), name="rate_prediction_daily")
        logger.info("Rate prediction scheduled at 07:45 daily")

        # N.E.L.S.O.N v2.0: Email Intelligence scan — 21:00 daily (after hours)
        async def _scheduled_email_intel(ctx):
            try:
                from intelligence.email_intel import EmailIntel
                intel = EmailIntel()
                results = intel.scan_email_db(days=1, limit=50)
                if results:
                    logger.info(f"[EmailIntel] Processed {len(results)} emails")
                    # Push summary to admin
                    signals_summary = []
                    for r in results[:5]:
                        s = r.get("signals", {})
                        signals_summary.append(
                            f"  {r['user_id']}: {s.get('sentiment','?')} "
                            f"| {s.get('intent','?')} "
                            f"| urgency {s.get('urgency_score','?')}"
                        )
                    if signals_summary:
                        msg = (
                            f"📧 <b>Email Intel Scan — {len(results)} emails</b>\n"
                            + "\n".join(signals_summary)
                        )
                        await ctx.bot.send_message(ADMIN_CHAT_ID, msg, parse_mode="HTML")
            except Exception as e:
                logger.warning(f"[EmailIntel] Scan failed: {e}")
        jq.run_daily(_scheduled_email_intel, time=dtime(hour=21, minute=0), name="email_intel_daily")
        logger.info("Email Intel scan scheduled at 21:00 daily")

    logger.info("Log: %s", log_file)
    logger.info("Bot v5 started. Listening...")
    logger.info("Bot v5 ready! Press Ctrl+C to stop.")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    max_retries, retry_delay = 5, 10
    for attempt in range(max_retries):
        try:
            main()
            break
        except KeyboardInterrupt:
            logger.info("Bot stopped by user.")
            break
        except Exception as exc:
            logger.error("Bot crashed: %s", exc, exc_info=True)
            if attempt < max_retries - 1:
                wait = retry_delay * (attempt + 1)
                logger.info("Restarting in %ds... (attempt %d/%d)", wait, attempt + 2, max_retries)
                time.sleep(wait)
            else:
                logger.error("Max retries (%d) reached. Exiting.", max_retries)
